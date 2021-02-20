from builtins import Exception
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtCore import Qt
import pymysql
from PyQt5.QtWidgets import QDialog
import threading
import time
import sys
import globalValues
import datetime
from PyQt5.QtCore import QDir, Qt, QUrl, pyqtSignal, QPoint, QRect, QObject
from PyQt5.QtMultimedia import QMediaContent, QMediaPlayer
from PyQt5.QtMultimediaWidgets import QVideoWidget
from PyQt5.QtWidgets import (QApplication, QFileDialog, QHBoxLayout, QLabel,
        QPushButton, QSizePolicy, QSlider, QStyle, QVBoxLayout, QWidget)
from PyQt5.QtWidgets import QMainWindow,QWidget, QPushButton, QAction
from PyQt5.QtGui import QIcon, QPainter, QImage, QPixmap
from PyQt5.QtMultimedia import QMediaContent, QMediaPlayer, QVideoFrame, QAbstractVideoSurface, QAbstractVideoBuffer, QVideoSurfaceFormat
from panelMesBox import Ui_mes_box
import os
import openpyxl

# import win32api
# import win32print

from datetime import date
from fpdf import FPDF
import glob

from base64 import b64decode, b64encode
from Crypto.PublicKey import RSA
from Crypto.Cipher import PKCS1_OAEP
# import win32com.client

globalValues.colorForm = 1

app = QtWidgets.QApplication(sys.argv)

class VideoFrameGrabber(QAbstractVideoSurface):
    frameAvailable = pyqtSignal(QImage)

    def __init__(self, widget: QWidget, parent: QObject):
        super().__init__(parent)

        self.widget = widget

    def supportedPixelFormats(self, handleType):
        return [QVideoFrame.Format_ARGB32, QVideoFrame.Format_ARGB32_Premultiplied,
                QVideoFrame.Format_RGB32, QVideoFrame.Format_RGB24, QVideoFrame.Format_RGB565,
                QVideoFrame.Format_RGB555, QVideoFrame.Format_ARGB8565_Premultiplied,
                QVideoFrame.Format_BGRA32, QVideoFrame.Format_BGRA32_Premultiplied, QVideoFrame.Format_BGR32,
                QVideoFrame.Format_BGR24, QVideoFrame.Format_BGR565, QVideoFrame.Format_BGR555,
                QVideoFrame.Format_BGRA5658_Premultiplied, QVideoFrame.Format_AYUV444,
                QVideoFrame.Format_AYUV444_Premultiplied, QVideoFrame.Format_YUV444,
                QVideoFrame.Format_YUV420P, QVideoFrame.Format_YV12, QVideoFrame.Format_UYVY,
                QVideoFrame.Format_YUYV, QVideoFrame.Format_NV12, QVideoFrame.Format_NV21,
                QVideoFrame.Format_IMC1, QVideoFrame.Format_IMC2, QVideoFrame.Format_IMC3,
                QVideoFrame.Format_IMC4, QVideoFrame.Format_Y8, QVideoFrame.Format_Y16,
                QVideoFrame.Format_Jpeg, QVideoFrame.Format_CameraRaw, QVideoFrame.Format_AdobeDng]

    def isFormatSupported(self, format):
        imageFormat = QVideoFrame.imageFormatFromPixelFormat(format.pixelFormat())
        size = format.frameSize()

        return imageFormat != QImage.Format_Invalid and not size.isEmpty() and \
               format.handleType() == QAbstractVideoBuffer.NoHandle

    def start(self, format: QVideoSurfaceFormat):
        imageFormat = QVideoFrame.imageFormatFromPixelFormat(format.pixelFormat())
        size = format.frameSize()

        if imageFormat != QImage.Format_Invalid and not size.isEmpty():
            self.imageFormat = imageFormat
            self.imageSize = size
            self.sourceRect = format.viewport()

            super().start(format)

            self.widget.updateGeometry()
            self.updateVideoRect()

            return True
        else:
            return False

    def stop(self):
        self.currentFrame = QVideoFrame()
        self.targetRect = QRect()

        super().stop()

        self.widget.update()

    def present(self, frame):
        if frame.isValid():
            cloneFrame = QVideoFrame(frame)
            cloneFrame.map(QAbstractVideoBuffer.ReadOnly)
            image = QImage(cloneFrame.bits(), cloneFrame.width(), cloneFrame.height(),
                           QVideoFrame.imageFormatFromPixelFormat(cloneFrame.pixelFormat()))
            self.frameAvailable.emit(image)  # this is very important
            cloneFrame.unmap()

        if self.surfaceFormat().pixelFormat() != frame.pixelFormat() or \
                self.surfaceFormat().frameSize() != frame.size():
            self.setError(QAbstractVideoSurface.IncorrectFormatError)
            self.stop()

            return False
        else:
            self.currentFrame = frame

            self.widget.repaint(self.targetRect)

            return True

    def updateVideoRect(self):
        size = self.surfaceFormat().sizeHint()
        size.scale(self.widget.size().boundedTo(size), Qt.KeepAspectRatio)

        self.targetRect = QRect(QPoint(0, 0), size)
        self.targetRect.moveCenter(self.widget.rect().center())

    def paint(self, painter):
        if self.currentFrame.map(QAbstractVideoBuffer.ReadOnly):
            oldTransform = self.painter.transform()

        if self.surfaceFormat().scanLineDirection() == QVideoSurfaceFormat.BottomToTop:
            self.painter.scale(1, -1)
            self.painter.translate(0, -self.widget.height())

        image = QImage(self.currentFrame.bits(), self.currentFrame.width(), self.currentFrame.height(),
                       self.currentFrame.bytesPerLine(), self.imageFormat)

        self.painter.drawImage(self.targetRect, image, self.sourceRect)

        self.painter.setTransform(oldTransform)

        self.currentFrame.unmap()

class Slider(QtWidgets.QSlider):

    def mousePressEvent(self, event):
        super(Slider, self).mousePressEvent(event)
        if event.button() == QtCore.Qt.LeftButton:
            val = self.pixelPosToRangeValue(event.pos())
            self.setValue(val)

    def pixelPosToRangeValue(self, pos):
        opt = QtWidgets.QStyleOptionSlider()
        self.initStyleOption(opt)
        gr = self.style().subControlRect(QtWidgets.QStyle.CC_Slider, opt, QtWidgets.QStyle.SC_SliderGroove, self)
        sr = self.style().subControlRect(QtWidgets.QStyle.CC_Slider, opt, QtWidgets.QStyle.SC_SliderHandle, self)

        if self.orientation() == QtCore.Qt.Horizontal:
            sliderLength = sr.width()
            sliderMin = gr.x()
            sliderMax = gr.right() - sliderLength + 1
        else:
            sliderLength = sr.height()
            sliderMin = gr.y()
            sliderMax = gr.bottom() - sliderLength + 1;
        pr = pos - sr.center() + sr.topLeft()
        p = pr.x() if self.orientation() == QtCore.Qt.Horizontal else pr.y()
        return QtWidgets.QStyle.sliderValueFromPosition(self.minimum(), self.maximum(), p - sliderMin,
                                               sliderMax - sliderMin, opt.upsideDown)

class Ui_PanelJournalTS(QDialog):

    firstCallChangeScroll = True

    con = pymysql.connections

    lstLight = [[]]
    lstDark = [[]]
    lengthLight = 0
    lengthDark = 0
    delta = 40
    num = 0

    indexStrOld = -1

    curImageKppIn = QImage()
    curImageKppOut = QImage()
    curImageWghtEmpty = QImage()
    curImageWghtLoad = QImage()

    pathFileVideo = ""

    pathFileCur = ''

    lstChl = ['КппВъезд', 'КппВыезд', 'ВесыВъезд', 'ВесыВыезд']

    isSearchGRZ = False
    isSearchOrder = False
    isSearchCompany = False
    isSearchDate = False

    dataSearch = ''

    def __init__(self):
        super().__init__()
        self.runUiContextTable()

    def runUiContextTable(self):
        self.setObjectName("Dialog")
        self.setFixedSize(1920, 1080)
        self.setWindowFlags(QtCore.Qt.CustomizeWindowHint | QtCore.Qt.FramelessWindowHint)

        self.lblBack = QtWidgets.QLabel(self)
        self.lblBack.setGeometry(QtCore.QRect(0, 0, 1921, 1081))
        self.lblBack.setText("")
        self.lblBack.setObjectName("lblBack")
        self.frmTableTS = QtWidgets.QFrame(self)
        self.frmTableTS.setGeometry(QtCore.QRect(0, 31, 1920, 490))
        self.frmTableTS.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmTableTS.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmTableTS.setObjectName("frmTableTS")
        self.tblTS = QtWidgets.QTableWidget(self.frmTableTS)
        self.tblTS.setGeometry(QtCore.QRect(20, 79, 1878, 400))
        self.tblTS.setObjectName("tblTS")
        self.tblTS.setRowCount(0)

        self.tblTS.setColumnCount(17)

        if (globalValues.colorForm == 1):
            self.tblTS.setStyleSheet("QTableWidget {background-color: rgb(235,235,235);\n"
                                     "border: 1px solid rgb(150,150,150);\n"
                                     "gridline-color: rgb(89,89,89);\n"
                                     "border-bottom-left-radius: 5px;\n"
                                     "border-bottom-right-radius: 0px;\n"
                                     "color:black;}\n"
                                     "QLineEdit {background-color: white;}\n"
                                     "QHeaderView::section {\n"
                                     "gridline-color: rgb(89,89,89);\n"
                                     "background-color: rgb(142,187,208);\n"
                                     "color: black;};")

        elif (globalValues.colorForm == 0):
            self.tblTS.setStyleSheet("QTableWidget {background-color: rgb(42,42,42);\n"
                                     "border: 1px solid rgb(63,63,63);\n"
                                     "border-radius:5px;\n"
                                     "gridline-color: rgb(89,89,89);\n"
                                     "border-bottom-right-radius: 0px;\n"
                                     "color:white;}\n"
                                     "QLineEdit {background-color: white;}\n"
                                     "QHeaderView::section {\n"
                                     "gridline-color: rgb(89,89,89);\n"
                                     "background-color: rgb(50,75,115);};")

        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(8, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(9, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(10, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(11, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(12, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QtGui.QColor(84, 122, 181))
        brush = QtGui.QBrush(QtGui.QColor(255, 255, 255))
        brush.setStyle(QtCore.Qt.SolidPattern)
        item.setForeground(brush)
        self.tblTS.setHorizontalHeaderItem(13, item)
        self.tblTS.horizontalHeader().setDefaultSectionSize(134)
        self.tblTS.horizontalHeader().setMinimumSectionSize(39)
        self.lblMainBoxCams = QtWidgets.QLabel(self.frmTableTS)
        self.lblMainBoxCams.setEnabled(False)
        self.lblMainBoxCams.setGeometry(QtCore.QRect(10, 10, 1900, 480))
        self.lblMainBoxCams.setText("")
        self.lblMainBoxCams.setObjectName("lblMainBoxCams")
        self.label_10 = QtWidgets.QLabel(self.frmTableTS)
        self.label_10.setGeometry(QtCore.QRect(690, 60, 536, 20))
        self.label_10.setObjectName("label_10")
        self.label_8 = QtWidgets.QLabel(self.frmTableTS)
        self.label_8.setGeometry(QtCore.QRect(1824, 60, 74, 20))
        self.label_8.setText("")
        self.label_8.setObjectName("label_8")
        self.label_9 = QtWidgets.QLabel(self.frmTableTS)
        self.label_9.setGeometry(QtCore.QRect(20, 60, 47, 20))
        self.label_9.setText("")
        self.label_9.setObjectName("label_9")
        self.label_7 = QtWidgets.QLabel(self.frmTableTS)
        self.label_7.setGeometry(QtCore.QRect(1762, 60, 135, 20))
        self.label_7.setText("")
        self.label_7.setObjectName("label_7")
        self.label_6 = QtWidgets.QLabel(self.frmTableTS)
        self.label_6.setGeometry(QtCore.QRect(21, 60, 669, 20))
        self.label_6.setObjectName("label_6")
        self.label_11 = QtWidgets.QLabel(self.frmTableTS)
        self.label_11.setGeometry(QtCore.QRect(1226, 60, 134, 20))
        self.label_11.setObjectName("label_11")
        self.label_12 = QtWidgets.QLabel(self.frmTableTS)
        self.label_12.setGeometry(QtCore.QRect(1360, 60, 402, 20))
        self.label_12.setObjectName("label_12")
        self.btnFilterRefresh = QtWidgets.QPushButton(self.frmTableTS)
        self.btnFilterRefresh.setGeometry(QtCore.QRect(1610, 20, 31, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnFilterRefresh.setFont(font)
        self.btnFilterRefresh.setText("")
        self.btnFilterRefresh.setIconSize(QtCore.QSize(25, 25))
        self.btnFilterRefresh.setObjectName("btnFilterRefresh")
        self.frmSrchDateRng = QtWidgets.QFrame(self.frmTableTS)
        self.frmSrchDateRng.setGeometry(QtCore.QRect(1000, 20, 431, 31))
        self.frmSrchDateRng.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmSrchDateRng.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmSrchDateRng.setObjectName("frmSrchDateRng")
        self.label_DateRng_4 = QtWidgets.QLabel(self.frmSrchDateRng)
        self.label_DateRng_4.setGeometry(QtCore.QRect(11, 5, 61, 21))
        self.label_DateRng_4.setObjectName("label_DateRng_4")
        self.label_dateAt_4 = QtWidgets.QLabel(self.frmSrchDateRng)
        self.label_dateAt_4.setGeometry(QtCore.QRect(77, 5, 24, 21))
        self.label_dateAt_4.setObjectName("label_dateAt_4")
        self.DTEditRngAt = QtWidgets.QDateTimeEdit(self.frmSrchDateRng)
        self.DTEditRngAt.setGeometry(QtCore.QRect(101, 0, 147, 31))
        self.DTEditRngAt.setObjectName("DTEditRngAt_3")
        self.DTEditRngTo = QtWidgets.QDateTimeEdit(self.frmSrchDateRng)
        self.DTEditRngTo.setGeometry(QtCore.QRect(281, 0, 147, 31))
        self.DTEditRngTo.setObjectName("DTEditRngTo_4")
        self.label_DateTo_4 = QtWidgets.QLabel(self.frmSrchDateRng)
        self.label_DateTo_4.setGeometry(QtCore.QRect(257, 5, 21, 21))
        self.label_DateTo_4.setObjectName("label_DateTo_4")
        self.btnFilter = QtWidgets.QPushButton(self.frmTableTS)
        self.btnFilter.setGeometry(QtCore.QRect(1440, 20, 161, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnFilter.setFont(font)
        self.btnFilter.setObjectName("btnFilter")
        self.label_filtertype = QtWidgets.QLabel(self.frmTableTS)
        self.label_filtertype.setGeometry(QtCore.QRect(800, 20, 111, 31))
        self.label_filtertype.setObjectName("label_filtertype")
        self.comboFilter = QtWidgets.QComboBox(self.frmTableTS)
        self.comboFilter.setGeometry(QtCore.QRect(920, 20, 81, 31))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(11)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.comboFilter.setFont(font)
        self.comboFilter.setObjectName("comboFilter")
        self.comboFilter.addItem("")
        self.comboFilter.addItem("")
        self.comboFilter.addItem("")
        self.label_iconfilter = QtWidgets.QLabel(self.frmTableTS)
        self.label_iconfilter.setGeometry(QtCore.QRect(770, 22, 25, 25))
        self.label_iconfilter.setText("")
        self.label_iconfilter.setObjectName("label_iconfilter")
        self.line = QtWidgets.QFrame(self.frmTableTS)
        self.line.setGeometry(QtCore.QRect(750, 20, 2, 31))
        self.line.setFrameShape(QtWidgets.QFrame.VLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.frmSrchDateFix = QtWidgets.QFrame(self.frmTableTS)
        self.frmSrchDateFix.setGeometry(QtCore.QRect(1000, 20, 431, 31))
        self.frmSrchDateFix.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmSrchDateFix.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmSrchDateFix.setObjectName("frmSrchDateFix")
        self.label_DateFix = QtWidgets.QLabel(self.frmSrchDateFix)
        self.label_DateFix.setGeometry(QtCore.QRect(10, 5, 51, 21))
        self.label_DateFix.setObjectName("label_DateFix")
        self.DTEitFix = QtWidgets.QDateTimeEdit(self.frmSrchDateFix)
        self.DTEitFix.setGeometry(QtCore.QRect(70, 0, 147, 31))
        self.DTEitFix.setObjectName("DTEitFix")
        self.btnSearch = QtWidgets.QPushButton(self.frmTableTS)
        self.btnSearch.setGeometry(QtCore.QRect(600, 20, 102, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnSearch.setFont(font)
        icon = QtGui.QIcon()
        self.btnSearch.setIcon(icon)
        self.btnSearch.setIconSize(QtCore.QSize(25, 25))
        self.btnSearch.setObjectName("btnSearch")
        self.leSearchTblTS = QtWidgets.QLineEdit(self.frmTableTS)
        self.leSearchTblTS.setGeometry(QtCore.QRect(20, 20, 501, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.leSearchTblTS.setFont(font)
        self.leSearchTblTS.setObjectName("leSearchTblTS")
        self.comboSearch = QtWidgets.QComboBox(self.frmTableTS)
        self.comboSearch.setGeometry(QtCore.QRect(530, 20, 61, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.comboSearch.setFont(font)
        self.comboSearch.setObjectName("comboSearch")
        self.comboSearch.addItem("")
        self.comboSearch.addItem("")
        self.comboSearch.addItem("")
        self.comboSearch.addItem("")
        self.comboSearch.addItem("")
        self.btnSrchRefresh = QtWidgets.QPushButton(self.frmTableTS)
        self.btnSrchRefresh.setGeometry(QtCore.QRect(710, 20, 31, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnSrchRefresh.setFont(font)
        self.btnSrchRefresh.setText("")
        icon1 = QtGui.QIcon()
        self.btnSrchRefresh.setIcon(icon1)
        self.btnSrchRefresh.setIconSize(QtCore.QSize(25, 25))
        self.btnSrchRefresh.setObjectName("btnSrchRefresh")
        self.line_2 = QtWidgets.QFrame(self.frmTableTS)
        self.line_2.setGeometry(QtCore.QRect(760, 20, 2, 31))
        self.line_2.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")
        self.line_3 = QtWidgets.QFrame(self.frmTableTS)
        self.line_3.setGeometry(QtCore.QRect(1650, 20, 2, 31))
        self.line_3.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_3.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_3.setObjectName("line_3")
        self.line_4 = QtWidgets.QFrame(self.frmTableTS)
        self.line_4.setGeometry(QtCore.QRect(1660, 20, 2, 31))
        self.line_4.setFrameShape(QtWidgets.QFrame.VLine)
        self.line_4.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_4.setObjectName("line_4")
        self.btnFilterPrint = QtWidgets.QPushButton(self.frmTableTS)
        self.btnFilterPrint.setGeometry(QtCore.QRect(1805, 20, 91, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnFilterPrint.setFont(font)
        self.btnFilterPrint.setIconSize(QtCore.QSize(25, 25))
        self.btnFilterPrint.setObjectName("btnFilterPrint")
        self.btnFilterSave = QtWidgets.QPushButton(self.frmTableTS)
        self.btnFilterSave.setGeometry(QtCore.QRect(1673, 20, 121, 31))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnFilterSave.setFont(font)
        self.btnFilterSave.setIconSize(QtCore.QSize(25, 25))
        self.btnFilterSave.setObjectName("btnFilterSave")
        self.verticalScrollBar = QtWidgets.QScrollBar(self.frmTableTS)
        self.verticalScrollBar.setGeometry(QtCore.QRect(1897, 103, 10, 376))
        self.verticalScrollBar.setOrientation(QtCore.Qt.Vertical)
        self.verticalScrollBar.setObjectName("verticalScrollBar")
        self.lblMainBoxCams.raise_()
        self.tblTS.raise_()
        self.label_10.raise_()
        self.label_8.raise_()
        self.label_9.raise_()
        self.label_7.raise_()
        self.label_6.raise_()
        self.label_11.raise_()
        self.label_12.raise_()
        self.btnFilterRefresh.raise_()
        self.btnFilter.raise_()
        self.label_filtertype.raise_()
        self.comboFilter.raise_()
        self.label_iconfilter.raise_()
        self.line.raise_()
        self.btnSearch.raise_()
        self.leSearchTblTS.raise_()
        self.comboSearch.raise_()
        self.btnSrchRefresh.raise_()
        self.line_2.raise_()
        self.line_3.raise_()
        self.line_4.raise_()
        self.btnFilterPrint.raise_()
        self.btnFilterSave.raise_()
        self.verticalScrollBar.raise_()
        self.frmSrchDateFix.raise_()
        self.frmSrchDateRng.raise_()
        self.frmINFO = QtWidgets.QFrame(self)
        self.frmINFO.setGeometry(QtCore.QRect(0, 520, 1921, 561))
        self.frmINFO.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmINFO.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmINFO.setObjectName("frmINFO")
        self.TabFrameVideo = QtWidgets.QTabWidget(self.frmINFO)
        self.TabFrameVideo.setGeometry(QtCore.QRect(730, 20, 1171, 521))
        self.TabFrameVideo.setObjectName("TabFrameVideo")
        self.tabKPP = QtWidgets.QWidget()
        self.tabKPP.setObjectName("tabKPP")
        self.label_51 = QtWidgets.QLabel(self.tabKPP)
        self.label_51.setGeometry(QtCore.QRect(0, 0, 1171, 498))
        self.label_51.setText("")
        self.label_51.setObjectName("label_51")
        self.le_VideoBack_2 = QtWidgets.QLineEdit(self.tabKPP)
        self.le_VideoBack_2.setEnabled(False)
        self.le_VideoBack_2.setGeometry(QtCore.QRect(650, 15, 470, 470))
        self.le_VideoBack_2.setObjectName("le_VideoBack_2")
        self.frmMediaPanelKppIn = QtWidgets.QFrame(self.tabKPP)
        self.frmMediaPanelKppIn.setGeometry(QtCore.QRect(650, 460, 470, 25))
        self.frmMediaPanelKppIn.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmMediaPanelKppIn.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmMediaPanelKppIn.setObjectName("frmMediaPanelKppIn")
        self.btnPlayKppIn = QtWidgets.QPushButton(self.frmMediaPanelKppIn)
        self.btnPlayKppIn.setGeometry(QtCore.QRect(50, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnPlayKppIn.setFont(font)
        self.btnPlayKppIn.setText("")
        self.btnPlayKppIn.setIconSize(QtCore.QSize(25, 25))
        self.btnPlayKppIn.setObjectName("btnPlayKppIn")
        self.SldrKppIn = Slider(self.frmMediaPanelKppIn)
        self.SldrKppIn.setGeometry(QtCore.QRect(156, 6, 300, 13))
        self.SldrKppIn.setOrientation(QtCore.Qt.Horizontal)
        self.SldrKppIn.setObjectName("SldrKppIn")
        self.btnStopKppIn = QtWidgets.QPushButton(self.frmMediaPanelKppIn)
        self.btnStopKppIn.setGeometry(QtCore.QRect(0, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnStopKppIn.setFont(font)
        self.btnStopKppIn.setText("")
        self.btnStopKppIn.setIconSize(QtCore.QSize(25, 25))
        self.btnStopKppIn.setObjectName("btnStopKppIn")
        self.btnStepbackKppIn = QtWidgets.QPushButton(self.frmMediaPanelKppIn)
        self.btnStepbackKppIn.setGeometry(QtCore.QRect(25, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnStepbackKppIn.setFont(font)
        self.btnStepbackKppIn.setText("")
        self.btnStepbackKppIn.setIconSize(QtCore.QSize(25, 25))
        self.btnStepbackKppIn.setObjectName("btnStepbackKppIn")
        self.btnStepforwardKppIn = QtWidgets.QPushButton(self.frmMediaPanelKppIn)
        self.btnStepforwardKppIn.setGeometry(QtCore.QRect(75, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnStepforwardKppIn.setFont(font)
        self.btnStepforwardKppIn.setText("")
        self.btnStepforwardKppIn.setIconSize(QtCore.QSize(25, 25))
        self.btnStepforwardKppIn.setObjectName("btnStepforwardKppIn")
        self.cbKppIn = QtWidgets.QComboBox(self.frmMediaPanelKppIn)
        self.cbKppIn.setGeometry(QtCore.QRect(100, 0, 45, 25))
        self.cbKppIn.setObjectName("cbKppIn")
        self.cbKppIn.addItem("")
        self.cbKppIn.addItem("")
        self.cbKppIn.addItem("")
        self.cbKppIn.addItem("")
        self.le_VideoKppIn = QVideoWidget(self.tabKPP)
        self.le_VideoKppIn.setGeometry(QtCore.QRect(652, 48, 466, 411))
        self.le_VideoKppIn.setObjectName("le_VideoKppIn")
        self.frmMediaPanelKppOut = QtWidgets.QFrame(self.tabKPP)
        self.frmMediaPanelKppOut.setGeometry(QtCore.QRect(50, 460, 470, 25))
        self.frmMediaPanelKppOut.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmMediaPanelKppOut.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmMediaPanelKppOut.setObjectName("frmMediaPanelKppOut")
        self.btnPlayKppOut = QtWidgets.QPushButton(self.frmMediaPanelKppOut)
        self.btnPlayKppOut.setGeometry(QtCore.QRect(50, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnPlayKppOut.setFont(font)
        self.btnPlayKppOut.setText("")
        self.btnPlayKppOut.setIconSize(QtCore.QSize(25, 25))
        self.btnPlayKppOut.setObjectName("btnPlayKppOut")
        self.SldrKppOut = Slider(self.frmMediaPanelKppOut)
        self.SldrKppOut.setGeometry(QtCore.QRect(156, 6, 300, 13))
        self.SldrKppOut.setOrientation(QtCore.Qt.Horizontal)
        self.SldrKppOut.setObjectName("SldrKppOut")
        self.btnStopKppOut = QtWidgets.QPushButton(self.frmMediaPanelKppOut)
        self.btnStopKppOut.setGeometry(QtCore.QRect(0, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnStopKppOut.setFont(font)
        self.btnStopKppOut.setText("")
        self.btnStopKppOut.setIconSize(QtCore.QSize(25, 25))
        self.btnStopKppOut.setObjectName("btnStopKppOut")
        self.btnStepbackKppOut = QtWidgets.QPushButton(self.frmMediaPanelKppOut)
        self.btnStepbackKppOut.setGeometry(QtCore.QRect(25, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnStepbackKppOut.setFont(font)
        self.btnStepbackKppOut.setText("")
        self.btnStepbackKppOut.setIconSize(QtCore.QSize(25, 25))
        self.btnStepbackKppOut.setObjectName("btnStepbackKppOut")
        self.btnStepforwardKppOut = QtWidgets.QPushButton(self.frmMediaPanelKppOut)
        self.btnStepforwardKppOut.setGeometry(QtCore.QRect(75, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnStepforwardKppOut.setFont(font)
        self.btnStepforwardKppOut.setText("")
        self.btnStepforwardKppOut.setIconSize(QtCore.QSize(25, 25))
        self.btnStepforwardKppOut.setObjectName("btnStepforwardKppOut")
        self.cbKppOut = QtWidgets.QComboBox(self.frmMediaPanelKppOut)
        self.cbKppOut.setGeometry(QtCore.QRect(100, 0, 45, 25))
        self.cbKppOut.setObjectName("cbKppOut")
        self.cbKppOut.addItem("")
        self.cbKppOut.addItem("")
        self.cbKppOut.addItem("")
        self.cbKppOut.addItem("")
        self.le_VideoKppOut = QVideoWidget(self.tabKPP)
        self.le_VideoKppOut.setGeometry(QtCore.QRect(52, 48, 466, 411))
        self.le_VideoKppOut.setObjectName("le_VideoKppOut")
        self.le_VideoBack_4 = QtWidgets.QLineEdit(self.tabKPP)
        self.le_VideoBack_4.setEnabled(False)
        self.le_VideoBack_4.setGeometry(QtCore.QRect(50, 15, 470, 470))
        self.le_VideoBack_4.setObjectName("le_VideoBack_4")
        self.label_15 = QtWidgets.QLabel(self.tabKPP)
        self.label_15.setGeometry(QtCore.QRect(660, 20, 161, 21))
        self.label_15.setObjectName("label_15")
        self.label_16 = QtWidgets.QLabel(self.tabKPP)
        self.label_16.setGeometry(QtCore.QRect(60, 20, 161, 21))
        self.label_16.setObjectName("label_16")
        self.label_51.raise_()
        self.le_VideoBack_4.raise_()
        self.le_VideoBack_2.raise_()
        self.frmMediaPanelKppIn.raise_()
        self.le_VideoKppIn.raise_()
        self.frmMediaPanelKppOut.raise_()
        self.le_VideoKppOut.raise_()
        self.label_15.raise_()
        self.label_16.raise_()
        self.TabFrameVideo.addTab(self.tabKPP, "")
        self.tabWeight = QtWidgets.QWidget()
        self.tabWeight.setObjectName("tabWeight")
        self.le_VideoBack = QtWidgets.QLineEdit(self.tabWeight)
        self.le_VideoBack.setEnabled(False)
        self.le_VideoBack.setGeometry(QtCore.QRect(650, 15, 470, 470))
        self.le_VideoBack.setObjectName("le_VideoBack")
        self.label_50 = QtWidgets.QLabel(self.tabWeight)
        self.label_50.setGeometry(QtCore.QRect(0, 0, 1171, 498))
        self.label_50.setText("")
        self.label_50.setObjectName("label_50")
        self.frmMediaPanelWeightIn = QtWidgets.QFrame(self.tabWeight)
        self.frmMediaPanelWeightIn.setGeometry(QtCore.QRect(650, 460, 470, 25))
        self.frmMediaPanelWeightIn.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmMediaPanelWeightIn.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmMediaPanelWeightIn.setObjectName("frmMediaPanelWeightIn")
        self.btnPlayWeightIn = QtWidgets.QPushButton(self.frmMediaPanelWeightIn)
        self.btnPlayWeightIn.setGeometry(QtCore.QRect(50, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnPlayWeightIn.setFont(font)
        self.btnPlayWeightIn.setText("")
        self.btnPlayWeightIn.setIconSize(QtCore.QSize(25, 25))
        self.btnPlayWeightIn.setObjectName("btnPlayWeightIn")
        self.SldrWeightIn = Slider(self.frmMediaPanelWeightIn)
        self.SldrWeightIn.setGeometry(QtCore.QRect(156, 6, 300, 13))
        self.SldrWeightIn.setOrientation(QtCore.Qt.Horizontal)
        self.SldrWeightIn.setObjectName("SldrWeightIn")
        self.btnStopWeightIn = QtWidgets.QPushButton(self.frmMediaPanelWeightIn)
        self.btnStopWeightIn.setGeometry(QtCore.QRect(0, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnStopWeightIn.setFont(font)
        self.btnStopWeightIn.setText("")
        self.btnStopWeightIn.setIconSize(QtCore.QSize(25, 25))
        self.btnStopWeightIn.setObjectName("btnStopWeightIn")
        self.btnStepbackWeightIn = QtWidgets.QPushButton(self.frmMediaPanelWeightIn)
        self.btnStepbackWeightIn.setGeometry(QtCore.QRect(25, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnStepbackWeightIn.setFont(font)
        self.btnStepbackWeightIn.setText("")
        self.btnStepbackWeightIn.setIconSize(QtCore.QSize(25, 25))
        self.btnStepbackWeightIn.setObjectName("btnStepbackWeightIn")
        self.btnStepforwardWeightIn = QtWidgets.QPushButton(self.frmMediaPanelWeightIn)
        self.btnStepforwardWeightIn.setGeometry(QtCore.QRect(75, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnStepforwardWeightIn.setFont(font)
        self.btnStepforwardWeightIn.setText("")
        self.btnStepforwardWeightIn.setIconSize(QtCore.QSize(25, 25))
        self.btnStepforwardWeightIn.setObjectName("btnStepforwardWeightIn")
        self.cbWeightIn = QtWidgets.QComboBox(self.frmMediaPanelWeightIn)
        self.cbWeightIn.setGeometry(QtCore.QRect(100, 0, 45, 25))
        self.cbWeightIn.setObjectName("cbWeightIn")
        self.cbWeightIn.addItem("")
        self.cbWeightIn.addItem("")
        self.cbWeightIn.addItem("")
        self.cbWeightIn.addItem("")
        self.btnPlayWeightIn.raise_()
        self.SldrWeightIn.raise_()
        self.btnStopWeightIn.raise_()
        self.btnStepforwardWeightIn.raise_()
        self.btnStepbackWeightIn.raise_()
        self.cbWeightIn.raise_()
        self.le_VideoWeightIn = QVideoWidget(self.tabWeight)
        self.le_VideoWeightIn.setGeometry(QtCore.QRect(652, 48, 466, 411))
        self.le_VideoWeightIn.setObjectName("le_VideoWeightIn")
        self.le_VideoBack_3 = QtWidgets.QLineEdit(self.tabWeight)
        self.le_VideoBack_3.setEnabled(False)
        self.le_VideoBack_3.setGeometry(QtCore.QRect(50, 15, 470, 470))
        self.le_VideoBack_3.setObjectName("le_VideoBack_3")
        self.frmMediaPanelWeightOut = QtWidgets.QFrame(self.tabWeight)
        self.frmMediaPanelWeightOut.setGeometry(QtCore.QRect(50, 460, 470, 25))
        self.frmMediaPanelWeightOut.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmMediaPanelWeightOut.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmMediaPanelWeightOut.setObjectName("frmMediaPanelWeightOut")
        self.btnPlayWeightOut = QtWidgets.QPushButton(self.frmMediaPanelWeightOut)
        self.btnPlayWeightOut.setGeometry(QtCore.QRect(50, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnPlayWeightOut.setFont(font)
        self.btnPlayWeightOut.setText("")
        self.btnPlayWeightOut.setIconSize(QtCore.QSize(25, 25))
        self.btnPlayWeightOut.setObjectName("btnPlayWeightOut")
        self.SldrWeightOut = Slider(self.frmMediaPanelWeightOut)
        self.SldrWeightOut.setGeometry(QtCore.QRect(156, 6, 300, 13))
        self.SldrWeightOut.setOrientation(QtCore.Qt.Horizontal)
        self.SldrWeightOut.setObjectName("SldrWeightOut")
        self.btnStopWeightOut = QtWidgets.QPushButton(self.frmMediaPanelWeightOut)
        self.btnStopWeightOut.setGeometry(QtCore.QRect(0, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnStopWeightOut.setFont(font)
        self.btnStopWeightOut.setText("")
        self.btnStopWeightOut.setIconSize(QtCore.QSize(25, 25))
        self.btnStopWeightOut.setObjectName("btnStopWeightOut")
        self.btnStepbackWeightOut = QtWidgets.QPushButton(self.frmMediaPanelWeightOut)
        self.btnStepbackWeightOut.setGeometry(QtCore.QRect(25, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnStepbackWeightOut.setFont(font)
        self.btnStepbackWeightOut.setText("")
        self.btnStepbackWeightOut.setIconSize(QtCore.QSize(25, 25))
        self.btnStepbackWeightOut.setObjectName("btnStepbackWeightOut")
        self.btnStepforwardWeightOut = QtWidgets.QPushButton(self.frmMediaPanelWeightOut)
        self.btnStepforwardWeightOut.setGeometry(QtCore.QRect(75, 0, 25, 25))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnStepforwardWeightOut.setFont(font)
        self.btnStepforwardWeightOut.setText("")
        self.btnStepforwardWeightOut.setIconSize(QtCore.QSize(25, 25))
        self.btnStepforwardWeightOut.setObjectName("btnStepforwardWeightOut")
        self.cbWeightOut = QtWidgets.QComboBox(self.frmMediaPanelWeightOut)
        self.cbWeightOut.setGeometry(QtCore.QRect(100, 0, 45, 25))
        self.cbWeightOut.setObjectName("cbWeightOut")
        self.cbWeightOut.addItem("")
        self.cbWeightOut.addItem("")
        self.cbWeightOut.addItem("")
        self.cbWeightOut.addItem("")
        self.le_VideoWeightOut = QVideoWidget(self.tabWeight)
        self.le_VideoWeightOut.setGeometry(QtCore.QRect(52, 48, 466, 411))
        self.le_VideoWeightOut.setObjectName("le_VideoWeightOut")
        self.label_14 = QtWidgets.QLabel(self.tabWeight)
        self.label_14.setGeometry(QtCore.QRect(660, 20, 161, 21))
        self.label_14.setObjectName("label_14")
        self.label_13 = QtWidgets.QLabel(self.tabWeight)
        self.label_13.setGeometry(QtCore.QRect(60, 20, 161, 21))
        self.label_13.setObjectName("label_13")
        self.label_50.raise_()
        self.le_VideoBack.raise_()
        self.frmMediaPanelWeightIn.raise_()
        self.le_VideoWeightIn.raise_()
        self.le_VideoBack_3.raise_()
        self.frmMediaPanelWeightOut.raise_()
        self.le_VideoWeightOut.raise_()
        self.label_14.raise_()
        self.label_13.raise_()
        self.TabFrameVideo.addTab(self.tabWeight, "")
        self.label_56 = QtWidgets.QLabel(self.frmINFO)
        self.label_56.setGeometry(QtCore.QRect(40, 290, 141, 21))
        self.label_56.setObjectName("label_56")
        self.label_57 = QtWidgets.QLabel(self.frmINFO)
        self.label_57.setGeometry(QtCore.QRect(225, 23, 291, 25))
        self.label_57.setObjectName("label_57")
        self.label_59 = QtWidgets.QLabel(self.frmINFO)
        self.label_59.setGeometry(QtCore.QRect(40, 110, 81, 21))
        self.label_59.setObjectName("label_59")
        self.label_60 = QtWidgets.QLabel(self.frmINFO)
        self.label_60.setGeometry(QtCore.QRect(40, 170, 151, 21))
        self.label_60.setObjectName("label_60")
        self.label_61 = QtWidgets.QLabel(self.frmINFO)
        self.label_61.setGeometry(QtCore.QRect(40, 350, 161, 21))
        self.label_61.setObjectName("label_61")
        self.label_64 = QtWidgets.QLabel(self.frmINFO)
        self.label_64.setGeometry(QtCore.QRect(40, 260, 151, 21))
        self.label_64.setObjectName("label_64")
        self.label_65 = QtWidgets.QLabel(self.frmINFO)
        self.label_65.setGeometry(QtCore.QRect(40, 230, 121, 21))
        self.label_65.setObjectName("label_65")
        self.label_67 = QtWidgets.QLabel(self.frmINFO)
        self.label_67.setGeometry(QtCore.QRect(197, 23, 28, 28))
        self.label_67.setText("")
        self.label_67.setObjectName("label_67")
        self.label_68 = QtWidgets.QLabel(self.frmINFO)
        self.label_68.setGeometry(QtCore.QRect(40, 200, 81, 21))
        self.label_68.setObjectName("label_68")
        self.label_70 = QtWidgets.QLabel(self.frmINFO)
        self.label_70.setGeometry(QtCore.QRect(40, 470, 241, 21))
        self.label_70.setObjectName("label_70")
        self.label_71 = QtWidgets.QLabel(self.frmINFO)
        self.label_71.setGeometry(QtCore.QRect(40, 80, 151, 21))
        self.label_71.setObjectName("label_71")
        self.label_73 = QtWidgets.QLabel(self.frmINFO)
        self.label_73.setGeometry(QtCore.QRect(40, 500, 251, 21))
        self.label_73.setObjectName("label_73")
        self.label_80 = QtWidgets.QLabel(self.frmINFO)
        self.label_80.setGeometry(QtCore.QRect(40, 320, 161, 21))
        self.label_80.setObjectName("label_80")
        self.label_82 = QtWidgets.QLabel(self.frmINFO)
        self.label_82.setGeometry(QtCore.QRect(40, 410, 231, 21))
        self.label_82.setObjectName("label_82")
        self.label_84 = QtWidgets.QLabel(self.frmINFO)
        self.label_84.setGeometry(QtCore.QRect(40, 440, 241, 21))
        self.label_84.setObjectName("label_84")
        self.label_86 = QtWidgets.QLabel(self.frmINFO)
        self.label_86.setGeometry(QtCore.QRect(430, 110, 78, 21))
        self.label_86.setObjectName("label_86")
        self.label_front3_3 = QtWidgets.QLabel(self.frmINFO)
        self.label_front3_3.setGeometry(QtCore.QRect(20, 20, 701, 521))
        self.label_front3_3.setText("")
        self.label_front3_3.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.label_front3_3.setObjectName("label_front3_3")
        self.line_5 = QtWidgets.QFrame(self.frmINFO)
        self.line_5.setGeometry(QtCore.QRect(225, 50, 285, 3))
        self.line_5.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_5.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_5.setObjectName("line_5")
        self.line_6 = QtWidgets.QFrame(self.frmINFO)
        self.line_6.setGeometry(QtCore.QRect(40, 150, 670, 3))
        self.line_6.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_6.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_6.setObjectName("line_6")
        self.line_8 = QtWidgets.QFrame(self.frmINFO)
        self.line_8.setGeometry(QtCore.QRect(40, 390, 670, 3))
        self.line_8.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_8.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_8.setObjectName("line_8")
        self.lblMainBoxCams_2 = QtWidgets.QLabel(self.frmINFO)
        self.lblMainBoxCams_2.setEnabled(False)
        self.lblMainBoxCams_2.setGeometry(QtCore.QRect(10, 10, 1900, 541))
        self.lblMainBoxCams_2.setText("")
        self.lblMainBoxCams_2.setObjectName("lblMainBoxCams_2")
        self.le_Organisation = QtWidgets.QLineEdit(self.frmINFO)
        self.le_Organisation.setGeometry(QtCore.QRect(200, 80, 511, 21))
        self.le_Organisation.setObjectName("le_Organisation")
        self.le_GRZNumber = QtWidgets.QLineEdit(self.frmINFO)
        self.le_GRZNumber.setGeometry(QtCore.QRect(200, 110, 221, 21))
        self.le_GRZNumber.setObjectName("le_GRZNumber")
        self.le_CarModel = QtWidgets.QLineEdit(self.frmINFO)
        self.le_CarModel.setGeometry(QtCore.QRect(510, 110, 201, 21))
        self.le_CarModel.setObjectName("le_CarModel")
        self.le_ZakazNumber = QtWidgets.QLineEdit(self.frmINFO)
        self.le_ZakazNumber.setGeometry(QtCore.QRect(200, 170, 221, 21))
        self.le_ZakazNumber.setObjectName("le_ZakazNumber")
        self.le_ZakazDate = QtWidgets.QLineEdit(self.frmINFO)
        self.le_ZakazDate.setGeometry(QtCore.QRect(200, 200, 221, 21))
        self.le_ZakazDate.setObjectName("le_ZakazDate")
        self.le_ZakazState = QtWidgets.QLineEdit(self.frmINFO)
        self.le_ZakazState.setGeometry(QtCore.QRect(200, 230, 221, 21))
        self.le_ZakazState.setObjectName("le_ZakazState")
        self.le_ComeToObject = QtWidgets.QLineEdit(self.frmINFO)
        self.le_ComeToObject.setGeometry(QtCore.QRect(200, 260, 221, 21))
        self.le_ComeToObject.setObjectName("le_ComeToObject")
        self.le_GoneFromObject = QtWidgets.QLineEdit(self.frmINFO)
        self.le_GoneFromObject.setGeometry(QtCore.QRect(200, 290, 221, 21))
        self.le_GoneFromObject.setObjectName("le_GoneFromObject")
        self.le_ComeToPoligon = QtWidgets.QLineEdit(self.frmINFO)
        self.le_ComeToPoligon.setGeometry(QtCore.QRect(200, 320, 221, 21))
        self.le_ComeToPoligon.setObjectName("le_ComeToPoligon")
        self.le_DriveOverTime = QtWidgets.QLineEdit(self.frmINFO)
        self.le_DriveOverTime.setGeometry(QtCore.QRect(200, 350, 221, 21))
        self.le_DriveOverTime.setObjectName("le_DriveOverTime")
        self.le_WeightObject = QtWidgets.QLineEdit(self.frmINFO)
        self.le_WeightObject.setGeometry(QtCore.QRect(290, 410, 131, 21))
        self.le_WeightObject.setObjectName("le_WeightObject")
        self.le_VolumeObject = QtWidgets.QLineEdit(self.frmINFO)
        self.le_VolumeObject.setGeometry(QtCore.QRect(290, 440, 131, 21))
        self.le_VolumeObject.setObjectName("le_VolumeObject")
        self.le_WeightPoligon = QtWidgets.QLineEdit(self.frmINFO)
        self.le_WeightPoligon.setGeometry(QtCore.QRect(290, 470, 131, 21))
        self.le_WeightPoligon.setObjectName("le_WeightPoligon")
        self.le_VolumePoligon = QtWidgets.QLineEdit(self.frmINFO)
        self.le_VolumePoligon.setGeometry(QtCore.QRect(290, 500, 131, 21))
        self.le_VolumePoligon.setObjectName("le_VolumePoligon")
        self.lblMainBoxCams_2.raise_()
        self.label_front3_3.raise_()
        self.TabFrameVideo.raise_()
        self.label_56.raise_()
        self.label_57.raise_()
        self.label_59.raise_()
        self.label_60.raise_()
        self.label_61.raise_()
        self.label_64.raise_()
        self.label_65.raise_()
        self.label_67.raise_()
        self.label_68.raise_()
        self.label_70.raise_()
        self.label_71.raise_()
        self.label_73.raise_()
        self.label_80.raise_()
        self.label_82.raise_()
        self.label_84.raise_()
        self.label_86.raise_()
        self.line_5.raise_()
        self.line_6.raise_()
        self.line_8.raise_()
        self.le_Organisation.raise_()
        self.le_GRZNumber.raise_()
        self.le_CarModel.raise_()
        self.le_ZakazNumber.raise_()
        self.le_ZakazDate.raise_()
        self.le_ZakazState.raise_()
        self.le_ComeToObject.raise_()
        self.le_GoneFromObject.raise_()
        self.le_ComeToPoligon.raise_()
        self.le_DriveOverTime.raise_()
        self.le_WeightObject.raise_()
        self.le_VolumeObject.raise_()
        self.le_WeightPoligon.raise_()
        self.le_VolumePoligon.raise_()
        self.frmHeadExit = QtWidgets.QFrame(self)
        self.frmHeadExit.setGeometry(QtCore.QRect(0, 0, 1920, 31))
        self.frmHeadExit.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmHeadExit.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmHeadExit.setObjectName("frmHeadExit")
        self.lblBarClose = QtWidgets.QLabel(self.frmHeadExit)
        self.lblBarClose.setGeometry(QtCore.QRect(25, 0, 1896, 30))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setUnderline(True)
        font.setWeight(50)
        self.lblBarClose.setFont(font)
        self.lblBarClose.setObjectName("lblBarClose")
        self.lblIconBarClose = QtWidgets.QLabel(self.frmHeadExit)
        self.lblIconBarClose.setGeometry(QtCore.QRect(2, 0, 20, 30))
        self.lblIconBarClose.setText("")
        self.lblIconBarClose.setObjectName("lblIconBarClose")
        self.btnCloseMainForm = QtWidgets.QPushButton(self.frmHeadExit)
        self.btnCloseMainForm.setGeometry(QtCore.QRect(1890, 2, 26, 26))
        self.btnCloseMainForm.setText("")
        self.btnCloseMainForm.setObjectName("btnCloseMainForm")
        self.label = QtWidgets.QLabel(self.frmHeadExit)
        self.label.setGeometry(QtCore.QRect(0, 0, 50, 30))
        self.label.setText("")
        self.label.setObjectName("label")
        self.label.raise_()
        self.lblBarClose.raise_()
        self.lblIconBarClose.raise_()
        self.btnCloseMainForm.raise_()
        self.lblBack.raise_()
        self.frmINFO.raise_()
        self.frmTableTS.raise_()
        self.frmHeadExit.raise_()

        self.retranslateUi()
        QtCore.QMetaObject.connectSlotsByName(self)
        # self.runThJournal()
        self.firstCall()

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("Dialog", "Dialog"))
        item = self.tblTS.horizontalHeaderItem(0)
        item.setText(_translate("Dialog", "ГРЗ ТС"))
        item = self.tblTS.horizontalHeaderItem(1)
        item.setText(_translate("Dialog", "#З/Н"))
        item = self.tblTS.horizontalHeaderItem(2)
        item.setText(_translate("Dialog", "Организация"))
        item = self.tblTS.horizontalHeaderItem(3)
        item.setText(_translate("Dialog", "Марка ТС"))
        item = self.tblTS.horizontalHeaderItem(4)
        item.setText(_translate("Dialog", "Дата"))
        item = self.tblTS.horizontalHeaderItem(5)
        item.setText(_translate("Dialog", "В.въезда"))
        item = self.tblTS.horizontalHeaderItem(6)
        item.setText(_translate("Dialog", "В.выезда"))
        item = self.tblTS.horizontalHeaderItem(7)
        item.setText(_translate("Dialog", "М.груза,т"))
        item = self.tblTS.horizontalHeaderItem(8)
        item.setText(_translate("Dialog", "О.груза,м3"))
        item = self.tblTS.horizontalHeaderItem(9)
        item.setText(_translate("Dialog", "В. в пути"))
        item = self.tblTS.horizontalHeaderItem(10)
        item.setText(_translate("Dialog", "В.Въезда"))
        item = self.tblTS.horizontalHeaderItem(11)
        item.setText(_translate("Dialog", "М.груза,т"))
        item = self.tblTS.horizontalHeaderItem(12)
        item.setText(_translate("Dialog", "О.груза,м3"))
        item = self.tblTS.horizontalHeaderItem(13)
        item.setText(_translate("Dialog", "Состояние З/Н"))
        self.label_10.setText(_translate("Dialog",
                                         "                                                                    СТРОИТЕЛЬНЫЙ ОБЬЕКТ"))
        self.label_6.setText(_translate("Dialog",
                                        "                                                                                             ОБЩАЯ ИНФОРМАЦИЯ"))
        self.label_11.setText(_translate("Dialog", "             МАРШРУТ"))
        self.label_12.setText(
            _translate("Dialog", "                                                           ПОЛИГОН"))
        self.label_DateRng_4.setText(_translate("Dialog", "Период:"))
        self.label_dateAt_4.setText(_translate("Dialog", "От"))
        self.label_DateTo_4.setText(_translate("Dialog", "До"))
        self.btnFilter.setText(_translate("Dialog", " Применить фильтр"))
        self.label_filtertype.setText(_translate("Dialog", "Тип фильтра:"))
        self.comboFilter.setItemText(0, _translate("Dialog", "Нет"))
        self.comboFilter.setItemText(1, _translate("Dialog", "Дата"))
        self.comboFilter.setItemText(2, _translate("Dialog", "Период"))
        self.label_iconfilter.setToolTip(_translate("Dialog", "Фильтр для поиска"))
        self.label_DateFix.setText(_translate("Dialog", "Дата :"))
        self.btnSearch.setText(_translate("Dialog", "      Поиск"))
        # self.leSearchTblTS.setText(_translate("Dialog", "Поиск"))
        self.comboSearch.setItemText(0, _translate("Dialog", "Нет"))
        self.comboSearch.setItemText(1, _translate("Dialog", "ГРЗ"))
        self.comboSearch.setItemText(2, _translate("Dialog", "З/Н"))
        self.comboSearch.setItemText(3, _translate("Dialog", "Орг."))
        self.comboSearch.setItemText(4, _translate("Dialog", "Дата"))
        self.btnFilterPrint.setText(_translate("Dialog", "     Печать"))
        self.btnFilterSave.setText(_translate("Dialog", "     Сохранить"))
        self.cbKppIn.setItemText(0, _translate("Dialog", "0.5x"))
        self.cbKppIn.setItemText(1, _translate("Dialog", "1x"))
        self.cbKppIn.setItemText(2, _translate("Dialog", "1.5x"))
        self.cbKppIn.setItemText(3, _translate("Dialog", "2x"))
        self.cbKppOut.setItemText(0, _translate("Dialog", "0.5x"))
        self.cbKppOut.setItemText(1, _translate("Dialog", "1x"))
        self.cbKppOut.setItemText(2, _translate("Dialog", "1.5x"))
        self.cbKppOut.setItemText(3, _translate("Dialog", "2x"))
        self.label_15.setText(_translate("Dialog", "НАПРАВЛЕНИЕ: ВЫЕЗД"))
        self.label_16.setText(_translate("Dialog", "НАПРАВЛЕНИЕ: ВЪЕЗД"))
        self.TabFrameVideo.setTabText(self.TabFrameVideo.indexOf(self.tabKPP), _translate("Dialog", " КПП"))
        self.cbWeightIn.setItemText(0, _translate("Dialog", "0.5x"))
        self.cbWeightIn.setItemText(1, _translate("Dialog", "1x"))
        self.cbWeightIn.setItemText(2, _translate("Dialog", "1.5x"))
        self.cbWeightIn.setItemText(3, _translate("Dialog", "2x"))
        self.cbWeightOut.setItemText(0, _translate("Dialog", "0.5x"))
        self.cbWeightOut.setItemText(1, _translate("Dialog", "1x"))
        self.cbWeightOut.setItemText(2, _translate("Dialog", "1.5x"))
        self.cbWeightOut.setItemText(3, _translate("Dialog", "2x"))
        self.label_14.setText(_translate("Dialog", "НАПРАВЛЕНИЕ: ВЫЕЗД"))
        self.label_13.setText(_translate("Dialog", "НАПРАВЛЕНИЕ: ВЪЕЗД"))
        self.TabFrameVideo.setTabText(self.TabFrameVideo.indexOf(self.tabWeight), _translate("Dialog", " ВЕСЫ"))
        self.label_56.setText(_translate("Dialog", "Выехал с обьекта:"))
        self.label_57.setText(_translate("Dialog", "Информация о транспортном средстве"))
        self.label_59.setText(_translate("Dialog", "№ ГРЗ ТС:"))
        self.label_60.setText(_translate("Dialog", "Номер З/Н:"))
        self.label_61.setText(_translate("Dialog", "Время в пути:"))
        self.label_64.setText(_translate("Dialog", "Приехал на обьект:"))
        self.label_65.setText(_translate("Dialog", "Состояние З/Н:"))
        self.label_68.setText(_translate("Dialog", "Дата З/Н:"))
        self.label_70.setText(_translate("Dialog", "Масса грунта в ТС на полигоне:"))
        self.label_71.setText(_translate("Dialog", "Организация:"))
        self.label_73.setText(_translate("Dialog", "Обьем грунта в ТС на полигоне:"))
        self.label_80.setText(_translate("Dialog", "Прибыл на полигон:"))
        self.label_82.setText(_translate("Dialog", "Масса грунта в ТС на объекте:"))
        self.label_84.setText(_translate("Dialog", "Обьем грунта в ТС на объекте:"))
        self.label_86.setText(_translate("Dialog", "Марка ТС:"))
        self.lblBarClose.setText(_translate("Dialog", "АСМК-ГРУНТ"))

    def firstCall(self):

        self.cbKppIn.setCurrentIndex(1)
        self.cbKppOut.setCurrentIndex(1)
        self.cbWeightIn.setCurrentIndex(1)
        self.cbWeightOut.setCurrentIndex(1)

        self.mediaPlayerKppIn = QMediaPlayer(None, QMediaPlayer.VideoSurface)
        # self.mediaGrabberKppIn = QMediaPlayer(None, QMediaPlayer.VideoSurface)
        self.mediaPlayerKppIn.setVideoOutput(self.le_VideoKppIn)

        # self.grabberKppIn = VideoFrameGrabber(self.le_VideoGRZTS, self)
        # self.mediaGrabberKppIn.setVideoOutput(self.grabberKppIn)
        # self.grabberKppIn.frameAvailable.connect(self.process_frame_kppin)

        self.mediaPlayerKppOut = QMediaPlayer(None, QMediaPlayer.VideoSurface)
        # self.mediaGrabberKppOut = QMediaPlayer(None, QMediaPlayer.VideoSurface)
        self.mediaPlayerKppOut.setVideoOutput(self.le_VideoKppOut)

        # self.grabberKppOut = VideoFrameGrabber(self.le_VideoGRUNT, self)
        # self.mediaGrabberKppOut.setVideoOutput(self.grabberKppOut)
        # self.grabberKppOut.frameAvailable.connect(self.process_frame_kppout)

        self.mediaPlayerWeightIn = QMediaPlayer(None, QMediaPlayer.VideoSurface)
        # self.mediaGrabberWghtEmpty = QMediaPlayer(None, QMediaPlayer.VideoSurface)
        self.mediaPlayerWeightIn.setVideoOutput(self.le_VideoWeightIn)
        #
        # self.grabberWghtEmpty = VideoFrameGrabber(self.le_VideoWghtEmpty, self)
        # self.mediaGrabberWghtEmpty.setVideoOutput(self.grabberWghtEmpty)
        # self.grabberWghtEmpty.frameAvailable.connect(self.process_frame_wght_empty)
        #
        self.mediaPlayerWeightOut = QMediaPlayer(None, QMediaPlayer.VideoSurface)
        # self.mediaGrabberWghtLoad = QMediaPlayer(None, QMediaPlayer.VideoSurface)
        self.mediaPlayerWeightOut.setVideoOutput(self.le_VideoWeightOut)
        #
        # self.grabberWghtLoad = VideoFrameGrabber(self.le_VideoWghtFull, self)
        # self.mediaGrabberWghtLoad.setVideoOutput(self.grabberWghtLoad)
        # self.grabberWghtLoad.frameAvailable.connect(self.process_frame_wght_load)

        self.leSearchTblTS.setPlaceholderText('Поиск')
        self.DTEitFix.setFont(QtGui.QFont("MS Shell Dlg 2", 11))
        self.DTEditRngTo.setFont(QtGui.QFont("MS Shell Dlg 2", 11))
        self.DTEditRngAt.setFont(QtGui.QFont("MS Shell Dlg 2", 11))

        self.lstLight = [[self.lblBack, "background-color: rgb(66,66,66);"],
                         [self.frmTableTS, "background-color: rgb(242,242,242);"],
                         [self.tblTS, "QTableWidget {background-color: rgb(235,235,235);\n"
                                      "border: 1px solid rgb(150,150,150);\n"
                                      "gridline-color: rgb(89,89,89);\n"
                                      "border-bottom-left-radius: 5px;\n"
                                      "border-bottom-right-radius: 0px;\n"
                                      "color:black;}\n"
                                      "QLineEdit {background-color: white;}\n"
                                      "QHeaderView::section {\n"
                                      "gridline-color: rgb(89,89,89);\n"
                                      "background-color: rgb(142,187,208);\n"
                                      "color: black;};\n"
                                      ""],
                         [self.lblMainBoxCams, "background-color: rgb(242,242,242);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(205,205,205);"],
                         [self.label_10, "background-color: rgb(142,187,208);\n"
                                         "border: 1px solid rgb(150,150,150);\n"
                                         "border-left-color: rgb(0,0,0);"],
                         [self.label_8, "background-color: rgb(142,187,208);\n"
                                        "border-radius: 5px;\n"
                                        "border: 1px solid rgb(150,150,150);\n"
                                        "border-bottom-left-radius: 0px;\n"
                                        "border-bottom-right-radius: 0px;\n"
                                        "border-top-left-radius: 0px;\n"
                                        "border-right-color: rgb(150,150,150);\n"
                                        "border-left-color: rgb(0,0,0);"],
                         [self.label_9, "background-color: rgb(142,187,208);\n"
                                        "border-radius: 5px;\n"
                                        "border: 1px solid rgb(150,150,150);\n"
                                        "border-bottom-left-radius: 0px;\n"
                                        "border-bottom-right-radius: 0px;\n"
                                        "border-top-right-radius: 0px;\n"
                                        "border-left-color: rgb(150,150,150);"],
                         [self.label_7, "background-color: rgb(142,187,208);\n"
                                        "border-radius: 5px;\n"
                                        "border: 1px solid rgb(150,150,150);\n"
                                        "border-bottom-left-radius: 0px;\n"
                                        "border-bottom-right-radius: 0px;\n"
                                        "border-top-left-radius: 0px;\n"
                                        "border-right-color: rgb(0,0,0);\n"
                                        "border-left-color: rgb(0,0,0);"],
                         [self.label_6, "background-color: rgb(142,187,208);\n"
                                        "border-radius: 5px;\n"
                                        "border: 1px solid rgb(150,150,150);\n"
                                        "border-bottom-left-radius: 0px;\n"
                                        "border-bottom-right-radius: 0px;\n"
                                        "border-top-right-radius: 0px;\n"
                                        "border-right-color: rgb(110,110,110);\n"
                                        "border-left-color: rgb(205,205,205);"],
                         [self.label_11, "background-color: rgb(142,187,208);\n"
                                         "border: 1px solid rgb(150,150,150);\n"
                                         "border-left-color: rgb(0,0,0);"],
                         [self.label_12, "background-color: rgb(142,187,208);\n"
                                         "border: 1px solid rgb(150,150,150);\n"
                                         "border-left-color: rgb(0,0,0);"],
                         [self.btnFilterRefresh, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                 "color: rgb(0,0,0);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(135,135,135);\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconrefreshgrey.png);}\n"
                                                                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                                                              "color: rgb(255, 255, 255);\n"
                                                                                              "border-radius:3px;\n"
                                                                                              "border:1px solid rgb(135,135,135);\n"
                                                                                              "image: url(" + globalValues.pathStyleImgs + "iconrefreshwhite.png);}\n"
                                                                                                                                           "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                                                                                                           "color: rgb(255, 255, 255);\n"
                                                                                                                                           "border-radius:3px;\n"
                                                                                                                                           "border:1px solid rgb(135,135,135);\n"
                                                                                                                                           "image: url(" + globalValues.pathStyleImgs + "iconrefreshwhite.png);};"],
                         [self.frmSrchDateRng, "background-color: rgb(242,242,242);"],
                         [self.label_DateRng_4, "background-color: rgb(242,242,242);\n"
                                                "color: rgb(0,0,0);\n"
                                                "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_dateAt_4, "background-color: rgb(242,242,242);\n"
                                               "color: rgb(0,0,0);\n"
                                               "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.DTEditRngAt, "background-color: rgb(227,227,227);\n"
                                            "color: black;\n"
                                            "border-radius: 3px;\n"
                                            "border: 1px solid rgb(135,135,135);\n"
                                            "font: 11pt \"MS Shell Dlg 2\";\n"
                                            "border: 1px solid rgb(63,63,63)"],
                         [self.DTEditRngTo, "background-color: rgb(227,227,227);\n"
                                            "color: black;\n"
                                            "border-radius: 3px;\n"
                                            "border: 1px solid rgb(135,135,135);\n"
                                            "font: 11pt \"MS Shell Dlg 2\";\n"
                                            "border: 1px solid rgb(63,63,63)"],
                         [self.label_DateTo_4, "background-color: rgb(242,242,242);\n"
                                               "color: rgb(0,0,0);\n"
                                               "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.btnFilter, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                          "color: rgb(0,0,0);\n"
                                          "border-radius:3px;\n"
                                          "border:1px solid rgb(135,135,135);}\n"
                                          "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                          "color: rgb(255, 255, 255);\n"
                                          "border-radius:3px;\n"
                                          "border:1px solid rgb(135,135,135);}\n"
                                          "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                          "color: rgb(255, 255, 255);\n"
                                          "border-radius:3px;\n"
                                          "border:1px solid rgb(135,135,135);};"],
                         [self.label_filtertype, "background-color: rgb(242,242,242);\n"
                                                 "color: rgb(0,0,0);\n"
                                                 "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.comboFilter, "background-color: rgb(227,227,227);\n"
                                            "color: black;\n"
                                            "border-radius: 3px;\n"
                                            "border: 1px solid rgb(135,135,135);"],
                         [self.label_iconfilter, "background-color: rgb(242,242,242);\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconfilter2.png);"],
                         [self.line, "background-color: rgb(242,242,242);"],
                         [self.frmSrchDateFix, "background-color: rgb(242,242,242);"],
                         [self.label_DateFix, "background-color: rgb(242,242,242);\n"
                                              "color: rgb(0,0,0);\n"
                                              "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.DTEitFix, "background-color: rgb(227,227,227);\n"
                                         "color: black;\n"
                                         "border-radius: 3px;\n"
                                         "border: 1px solid rgb(135,135,135);\n"
                                         "font: 11pt \"MS Shell Dlg 2\";\n"
                                         "border: 1px solid rgb(63,63,63)"],
                         [self.btnSearch, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                          "color: rgb(0,0,0);\n"
                                          "border-radius:3px;\n"
                                          "border:1px solid rgb(135,135,135);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "iconsearch8.png);}\n"
                                                                                       "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                                                       "color: rgb(255, 255, 255);\n"
                                                                                       "border-radius:3px;\n"
                                                                                       "border:1px solid rgb(135,135,135);\n"
                                                                                       "image: url(" + globalValues.pathStyleImgs + "iconsearchwhite.png);}\n"
                                                                                                                                    "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                                                                                                    "color: rgb(255, 255, 255);\n"
                                                                                                                                    "border-radius:3px;\n"
                                                                                                                                    "border:1px solid rgb(135,135,135);\n"
                                                                                                                                    "image: url(" + globalValues.pathStyleImgs + "iconsearchwhite.png);};"],
                         [self.leSearchTblTS, "background-color: rgb(235,235,235);\n"
                                              "color: rgb(0,0,0);\n"
                                              "border-radius: 3px;\n"
                                              "border: 1px solid rgb(150,150,150);"],
                         [self.comboSearch, "background-color: rgb(227,227,227);\n"
                                            "color: black;\n"
                                            "border-radius: 3px;\n"
                                            "border: 1px solid rgb(135,135,135);"],
                         [self.btnSrchRefresh, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                               "color: rgb(0,0,0);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconrefreshgrey.png);}\n"
                                                                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                                                            "color: rgb(255, 255, 255);\n"
                                                                                            "border-radius:3px;\n"
                                                                                            "border:1px solid rgb(135,135,135);\n"
                                                                                            "image: url(" + globalValues.pathStyleImgs + "iconrefreshwhite.png);}\n"
                                                                                                                                         "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                                                                                                         "color: rgb(255, 255, 255);\n"
                                                                                                                                         "border-radius:3px;\n"
                                                                                                                                         "border:1px solid rgb(135,135,135);\n"
                                                                                                                                         "image: url(" + globalValues.pathStyleImgs + "iconrefreshwhite.png);};"],
                         [self.line_2, "background-color: rgb(242,242,242);"],
                         [self.line_3, "background-color: rgb(242,242,242);"],
                         [self.line_4, "background-color: rgb(242,242,242);"],
                         [self.btnFilterPrint, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                               "color: rgb(0,0,0);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconprint8grey.png);}\n"
                                                                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                                                            "color: rgb(255, 255, 255);\n"
                                                                                            "border-radius:3px;\n"
                                                                                            "border:1px solid rgb(135,135,135);\n"
                                                                                            "image: url(" + globalValues.pathStyleImgs + "iconprint8.png);}\n"
                                                                                                                                         "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                                                                                                         "color: rgb(255, 255, 255);\n"
                                                                                                                                         "border-radius:3px;\n"
                                                                                                                                         "border:1px solid rgb(135,135,135);\n"
                                                                                                                                         "image: url(" + globalValues.pathStyleImgs + "iconprint8.png);};"],
                         [self.btnFilterSave, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                              "color: rgb(0,0,0);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconsaveGREY.png);}\n"
                                                                                           "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                                                           "color: rgb(255, 255, 255);\n"
                                                                                           "border-radius:3px;\n"
                                                                                           "border:1px solid rgb(135,135,135);\n"
                                                                                           "image: url(" + globalValues.pathStyleImgs + "iconsaveWHITE.png);}\n"
                                                                                                                                        "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                                                                                                        "color: rgb(255, 255, 255);\n"
                                                                                                                                        "border-radius:3px;\n"
                                                                                                                                        "border:1px solid rgb(135,135,135);\n"
                                                                                                                                        "image: url(" + globalValues.pathStyleImgs + "iconsaveWHITE.png);};"],
                         [self.verticalScrollBar, "QScrollBar:vertical {\n"
                                                  "border: 1px solid rgb(89,89,89);\n"
                                                  " background: rgb(255,255,255);\n"
                                                  "width:10px;\n"
                                                  "margin: 0px 0px 0px 0px;\n"
                                                  "}\n"
                                                  "QScrollBar::handle:vertical {\n"
                                                  "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0 rgb(142,187,208), stop: 0.5 rgb(142,187,208), stop:1 rgb(142,187,208));\n"
                                                  "min-height: 0px;\n"
                                                  "}\n"
                                                  "QScrollBar::add-line:vertical {\n"
                                                  " background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop: 0 rgb(142,187,208), stop: 0.5 rgb(142,187,208),  stop:1 rgb(142,187,208));\n"
                                                  "height: 0px;\n"
                                                  "subcontrol-position: bottom;\n"
                                                  "subcontrol-origin: margin;\n"
                                                  "}\n"
                                                  "QScrollBar::sub-line:vertical {\n"
                                                  "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0  rgb(142,187,208), stop: 0.5 rgb(142,187,208),  stop:1 rgb(142,187,208));\n"
                                                  "height: 0 px;\n"
                                                  "subcontrol-position: top;\n"
                                                  "subcontrol-origin: margin;\n"
                                                  "}"],
                         [self.frmINFO, "background-color: rgb(242,242,242);"],
                         [self.TabFrameVideo, "QTabWidget::pane {border:none;\n"
                                              "border-radius: 5px}\n"
                                              "QTabBar::tab {background-color: rgb(142,187,208);\n"
                                              "border: 1px solid rgb(150,150,150);\n"
                                              "border-left-color: rgb(66,66,66);\n"
                                              "padding: 5px;\n"
                                              "padding-bottom: 4px;\n"
                                              "padding-top: 4px;\n"
                                              "padding-right:14px;\n"
                                              "padding-left: 7px;\n"
                                              "font: 8pt \"MS Shell Dlg 2\";\n"
                                              "color:black;}\n"
                                              "QTabBar::tab:selected {background-color: rgb(172,221,243);\n"
                                              "margin-bottom: -1px;\n"
                                              " color:black;\n"
                                              "border: 1px solid rgb(150,150,150);\n"
                                              "border-left-color: rgb(0,0,0);}\n"
                                              "QTabBar::tab:hover {background-color: rgb(160,208,230);}\n"
                                              "QTabBar::tab:selected:hover {background-color: rgb(172,221,243);}"],
                         [self.tabKPP, ""],
                         [self.le_VideoBack_2, "background-color: rgb(235,235,235);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(150,150,150);"],
                         [self.label_50, "background-color: rgb(235,235,235);\n"
                                         "border: 2px solid rgb(150,150,150);\n"
                                         "border-radius: 5px;\n"
                                         "border-top-left-radius: 0px;\n"
                                         "border-bottom-left-radius: 5px;\n"
                                         "border-bottom-right-radius: 5px;"],
                         [self.le_VideoKppOut, "background-color: rgb(235,235,235);\n"
                                               "border: none;\n"
                                               "image: url(" + globalValues.pathStyleImgs + "camDefault460x262grey.png);"],
                         [self.frmMediaPanelKppIn, "background-color: rgb(227,227,227);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 2px solid rgb(150,150,150);\n"
                                                   "border-top: 1px solid rgb(150,150,150);\n"
                                                   "border-top-left-radius: 0px;\n"
                                                   "border-top-right-radius: 0px;"],
                         [self.btnPlayKppIn, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:0px;\n"
                                             "border:2px solid rgb(150,150,150);\n"
                                             "border-left:1px solid rgb(150,150,150);\n"
                                             "border-right:1px solid rgb(150,150,150);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                                                          "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                          "color: rgb(255, 255, 255);\n"
                                                                                          "border-radius:0px;\n"
                                                                                          "border:2px solid rgb(150,150,150);\n"
                                                                                          "border-left:1px solid rgb(150,150,150);\n"
                                                                                          "border-right:1px solid rgb(150,150,150);\n"
                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                                                                                                       "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                       "color: rgb(255, 255, 255);\n"
                                                                                                                                       "border-radius:0px;\n"
                                                                                                                                       "border:2px solid rgb(150,150,150);\n"
                                                                                                                                       "border-left:1px solid rgb(150,150,150);\n"
                                                                                                                                       "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                       "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}"],
                         [self.SldrKppIn, "QSlider {border-radius: none;\n"
                                          "border: none;}\n"
                                          "QSlider::groove:horizontal {\n"
                                          "    background-color: rgb(252,252,252);\n"
                                          "    border: 1px solid rgb(64,64,64);\n"
                                          "    height: 10px;\n"
                                          "    margin: 0px;\n"
                                          "    }\n"
                                          "QSlider::handle:horizontal {\n"
                                          "    background-color: rgb(64,64,64);\n"
                                          "    border: 1px solid rgb(64,64,64);\n"
                                          "    height: 10px;\n"
                                          "    width:  10px;\n"
                                          "    margin: -15px 0px;}"],
                         [self.btnStopKppIn, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:0px;\n"
                                             "border:2px solid rgb(150,150,150);\n"
                                             "border-left:2px solid rgb(150,150,150);\n"
                                             "border-right:1px solid rgb(150,150,150);\n"
                                             "border-bottom-left-radius: 5px;\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);}\n"
                                                                                          "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                          "color: rgb(255, 255, 255);\n"
                                                                                          "border-radius:0px;\n"
                                                                                          "border:2px solid rgb(150,150,150);\n"
                                                                                          "border-left:2px solid rgb(150,150,150);\n"
                                                                                          "border-right:1px solid rgb(150,150,150);\n"
                                                                                          "border-bottom-left-radius: 5px;\n"
                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);}\n"
                                                                                                                                       "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                       "color: rgb(255, 255, 255);\n"
                                                                                                                                       "border-radius:0px;\n"
                                                                                                                                       "border:2px solid rgb(150,150,150);\n"
                                                                                                                                       "border-left:2px solid rgb(150,150,150);\n"
                                                                                                                                       "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                       "border-bottom-left-radius: 5px;\n"
                                                                                                                                       "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);};"],
                         [self.btnStepbackKppIn, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:0px;\n"
                                                 "border:2px solid rgb(150,150,150);\n"
                                                 "border-left:1px solid rgb(150,150,150);\n"
                                                 "border-right:1px solid rgb(150,150,150);\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}\n"
                                                                                              "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                              "color: rgb(255, 255, 255);\n"
                                                                                              "border-radius:0px;\n"
                                                                                              "border:2px solid rgb(150,150,150);\n"
                                                                                              "border-left:1px solid rgb(150,150,150);\n"
                                                                                              "border-right:1px solid rgb(150,150,150);\n"
                                                                                              "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}\n"
                                                                                                                                           "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                           "color: rgb(255, 255, 255);\n"
                                                                                                                                           "border-radius:0px;\n"
                                                                                                                                           "border:2px solid rgb(150,150,150);\n"
                                                                                                                                           "border-left:1px solid rgb(150,150,150);\n"
                                                                                                                                           "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                           "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}"],
                         [self.btnStepforwardKppIn, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}\n"
                                                                                                 "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                                 "color: rgb(255, 255, 255);\n"
                                                                                                 "border-radius:0px;\n"
                                                                                                 "border:2px solid rgb(150,150,150);\n"
                                                                                                 "border-left:1px solid rgb(150,150,150);\n"
                                                                                                 "border-right:1px solid rgb(150,150,150);\n"
                                                                                                 "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}\n"
                                                                                                                                              "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                              "color: rgb(255, 255, 255);\n"
                                                                                                                                              "border-radius:0px;\n"
                                                                                                                                              "border:2px solid rgb(150,150,150);\n"
                                                                                                                                              "border-left:1px solid rgb(150,150,150);\n"
                                                                                                                                              "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                              "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}"],
                         [self.cbKppIn, "QComboBox {\n"
                                        "background: white;\n"
                                        "border: 2px solid rgb(150,150,150);\n"
                                        "border-right: 1px solid rgb(150,150,150);\n"
                                        "border-bottom: 2px solid rgb(150,150,150);\n"
                                        "border-radius: 0px;\n"
                                        "}\n"
                                        "QComboBox:editable {\n"
                                        "background-color: rgb(142,187,208);\n"
                                        "color: black;\n"
                                        "border-radius: 0px;\n"
                                        "}"],
                         [self.le_VideoKppIn, "background-color: rgb(235,235,235);\n"
                                              "border: none;\n"
                                              "image: url(" + globalValues.pathStyleImgs + "camDefault460x262grey.png);"],
                         [self.frmMediaPanelKppOut, "background-color: rgb(227,227,227);\n"
                                                    "border-radius: 5px;\n"
                                                    "border: 2px solid rgb(150,150,150);\n"
                                                    "border-top: 1px solid rgb(150,150,150);\n"
                                                    "border-top-left-radius: 0px;\n"
                                                    "border-top-right-radius: 0px;"],
                         [self.btnPlayKppOut, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:0px;\n"
                                              "border:2px solid rgb(150,150,150);\n"
                                              "border-left:1px solid rgb(150,150,150);\n"
                                              "border-right:1px solid rgb(150,150,150);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                                                           "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                           "color: rgb(255, 255, 255);\n"
                                                                                           "border-radius:0px;\n"
                                                                                           "border:2px solid rgb(150,150,150);\n"
                                                                                           "border-left:1px solid rgb(150,150,150);\n"
                                                                                           "border-right:1px solid rgb(150,150,150);\n"
                                                                                           "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                                                                                                        "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                        "color: rgb(255, 255, 255);\n"
                                                                                                                                        "border-radius:0px;\n"
                                                                                                                                        "border:2px solid rgb(150,150,150);\n"
                                                                                                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                                                                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}"],
                         [self.SldrKppOut, "QSlider {border-radius: none;\n"
                                           "border: none;}\n"
                                           "QSlider::groove:horizontal {\n"
                                           "    background-color: rgb(252,252,252);\n"
                                           "    border: 1px solid rgb(64,64,64);\n"
                                           "    height: 10px;\n"
                                           "    margin: 0px;\n"
                                           "    }\n"
                                           "QSlider::handle:horizontal {\n"
                                           "    background-color: rgb(64,64,64);\n"
                                           "    border: 1px solid rgb(64,64,64);\n"
                                           "    height: 10px;\n"
                                           "    width:  10px;\n"
                                           "    margin: -15px 0px;}"],
                         [self.btnStopKppOut, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:0px;\n"
                                              "border:2px solid rgb(150,150,150);\n"
                                              "border-left:2px solid rgb(150,150,150);\n"
                                              "border-right:1px solid rgb(150,150,150);\n"
                                              "border-bottom-left-radius: 5px;\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);}\n"
                                                                                           "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                           "color: rgb(255, 255, 255);\n"
                                                                                           "border-radius:0px;\n"
                                                                                           "border:2px solid rgb(150,150,150);\n"
                                                                                           "border-left:2px solid rgb(150,150,150);\n"
                                                                                           "border-right:1px solid rgb(150,150,150);\n"
                                                                                           "border-bottom-left-radius: 5px;\n"
                                                                                           "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);}\n"
                                                                                                                                        "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                        "color: rgb(255, 255, 255);\n"
                                                                                                                                        "border-radius:0px;\n"
                                                                                                                                        "border:2px solid rgb(150,150,150);\n"
                                                                                                                                        "border-left:2px solid rgb(150,150,150);\n"
                                                                                                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                        "border-bottom-left-radius: 5px;\n"
                                                                                                                                        "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);};"],
                         [self.btnStepbackKppOut, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:0px;\n"
                                                  "border:2px solid rgb(150,150,150);\n"
                                                  "border-left:1px solid rgb(150,150,150);\n"
                                                  "border-right:1px solid rgb(150,150,150);\n"
                                                  "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}\n"
                                                                                               "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                               "color: rgb(255, 255, 255);\n"
                                                                                               "border-radius:0px;\n"
                                                                                               "border:2px solid rgb(150,150,150);\n"
                                                                                               "border-left:1px solid rgb(150,150,150);\n"
                                                                                               "border-right:1px solid rgb(150,150,150);\n"
                                                                                               "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}\n"
                                                                                                                                            "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                            "color: rgb(255, 255, 255);\n"
                                                                                                                                            "border-radius:0px;\n"
                                                                                                                                            "border:2px solid rgb(150,150,150);\n"
                                                                                                                                            "border-left:1px solid rgb(150,150,150);\n"
                                                                                                                                            "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                            "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}"],
                         [self.btnStepforwardKppOut, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:0px;\n"
                                                     "border:2px solid rgb(150,150,150);\n"
                                                     "border-left:1px solid rgb(150,150,150);\n"
                                                     "border-right:1px solid rgb(150,150,150);\n"
                                                     "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}\n"
                                                                                                  "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                                  "color: rgb(255, 255, 255);\n"
                                                                                                  "border-radius:0px;\n"
                                                                                                  "border:2px solid rgb(150,150,150);\n"
                                                                                                  "border-left:1px solid rgb(150,150,150);\n"
                                                                                                  "border-right:1px solid rgb(150,150,150);\n"
                                                                                                  "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}\n"
                                                                                                                                               "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                               "color: rgb(255, 255, 255);\n"
                                                                                                                                               "border-radius:0px;\n"
                                                                                                                                               "border:2px solid rgb(150,150,150);\n"
                                                                                                                                               "border-left:1px solid rgb(150,150,150);\n"
                                                                                                                                               "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                               "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}"],
                         [self.cbKppOut, "QComboBox {\n"
                                         "background: white;\n"
                                         "border: 2px solid rgb(150,150,150);\n"
                                         "border-right: 1px solid rgb(150,150,150);\n"
                                         "border-bottom: 2px solid rgb(150,150,150);\n"
                                         "border-radius: 0px;\n"
                                         "}\n"
                                         "QComboBox:editable {\n"
                                         "background-color: rgb(142,187,208);\n"
                                         "color: black;\n"
                                         "border-radius: 0px;\n"
                                         "}"],
                         [self.le_VideoBack_3, "background-color: rgb(235,235,235);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(150,150,150);"],
                         [self.label_16, "background-color: rgb(235,235,235);\n"
                                         "font: 11pt \"MS Shell Dlg 2\";\n"
                                         "border: none;"],
                         [self.label_15, "background-color: rgb(235,235,235);\n"
                                         "font: 11pt \"MS Shell Dlg 2\";\n"
                                         "border: none;"],
                         [self.label_51, "background-color: rgb(235,235,235);\n"
                                         "border: 2px solid rgb(150,150,150);\n"
                                         "border-radius: 5px;\n"
                                         "border-top-left-radius: 0px;\n"
                                         "border-bottom-left-radius: 5px;\n"
                                         "border-bottom-right-radius: 5px;"],
                         [self.le_VideoBack, "background-color: rgb(235,235,235);\n"
                                             "border-radius: 5px;\n"
                                             "border: 2px solid rgb(150,150,150);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "camDefault460x460grey.png);"],
                         [self.le_VideoWeightOut, "background-color: rgb(235,235,235);\n"
                                                  "border: none;\n"
                                                  "image: url(" + globalValues.pathStyleImgs + "camDefault460x262grey.png);"],
                         [self.frmMediaPanelWeightIn, "background-color: rgb(227,227,227);\n"
                                                      "border-radius: 5px;\n"
                                                      "border: 2px solid rgb(150,150,150);\n"
                                                      "border-top: 1px solid rgb(150,150,150);\n"
                                                      "border-top-left-radius: 0px;\n"
                                                      "border-top-right-radius: 0px;"],
                         [self.btnPlayWeightIn, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:0px;\n"
                                                "border:2px solid rgb(150,150,150);\n"
                                                "border-left:1px solid rgb(150,150,150);\n"
                                                "border-right:1px solid rgb(150,150,150);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                                                             "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                             "color: rgb(255, 255, 255);\n"
                                                                                             "border-radius:0px;\n"
                                                                                             "border:2px solid rgb(150,150,150);\n"
                                                                                             "border-left:1px solid rgb(150,150,150);\n"
                                                                                             "border-right:1px solid rgb(150,150,150);\n"
                                                                                             "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                                                                                                          "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                          "color: rgb(255, 255, 255);\n"
                                                                                                                                          "border-radius:0px;\n"
                                                                                                                                          "border:2px solid rgb(150,150,150);\n"
                                                                                                                                          "border-left:1px solid rgb(150,150,150);\n"
                                                                                                                                          "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}"],
                         [self.SldrWeightIn, "QSlider {border-radius: none;\n"
                                             "border: none;}\n"
                                             "QSlider::groove:horizontal {\n"
                                             "    background-color: rgb(252,252,252);\n"
                                             "    border: 1px solid rgb(64,64,64);\n"
                                             "    height: 10px;\n"
                                             "    margin: 0px;\n"
                                             "    }\n"
                                             "QSlider::handle:horizontal {\n"
                                             "    background-color: rgb(64,64,64);\n"
                                             "    border: 1px solid rgb(64,64,64);\n"
                                             "    height: 10px;\n"
                                             "    width:  10px;\n"
                                             "    margin: -15px 0px;}"],
                         [self.btnStopWeightIn, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:0px;\n"
                                                "border:2px solid rgb(150,150,150);\n"
                                                "border-left:2px solid rgb(150,150,150);\n"
                                                "border-right:1px solid rgb(150,150,150);\n"
                                                "border-bottom-left-radius: 5px;\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);}\n"
                                                                                             "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                             "color: rgb(255, 255, 255);\n"
                                                                                             "border-radius:0px;\n"
                                                                                             "border:2px solid rgb(150,150,150);\n"
                                                                                             "border-left:2px solid rgb(150,150,150);\n"
                                                                                             "border-right:1px solid rgb(150,150,150);\n"
                                                                                             "border-bottom-left-radius: 5px;\n"
                                                                                             "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);}\n"
                                                                                                                                          "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                          "color: rgb(255, 255, 255);\n"
                                                                                                                                          "border-radius:0px;\n"
                                                                                                                                          "border:2px solid rgb(150,150,150);\n"
                                                                                                                                          "border-left:2px solid rgb(150,150,150);\n"
                                                                                                                                          "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                          "border-bottom-left-radius: 5px;\n"
                                                                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);};"],
                         [self.btnStepbackWeightIn, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}\n"
                                                                                                 "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                                 "color: rgb(255, 255, 255);\n"
                                                                                                 "border-radius:0px;\n"
                                                                                                 "border:2px solid rgb(150,150,150);\n"
                                                                                                 "border-left:1px solid rgb(150,150,150);\n"
                                                                                                 "border-right:1px solid rgb(150,150,150);\n"
                                                                                                 "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}\n"
                                                                                                                                              "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                              "color: rgb(255, 255, 255);\n"
                                                                                                                                              "border-radius:0px;\n"
                                                                                                                                              "border:2px solid rgb(150,150,150);\n"
                                                                                                                                              "border-left:1px solid rgb(150,150,150);\n"
                                                                                                                                              "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                              "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}"],
                         [self.btnStepforwardWeightIn, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                       "color: rgb(255, 255, 255);\n"
                                                       "border-radius:0px;\n"
                                                       "border:2px solid rgb(150,150,150);\n"
                                                       "border-left:1px solid rgb(150,150,150);\n"
                                                       "border-right:1px solid rgb(150,150,150);\n"
                                                       "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}\n"
                                                                                                    "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                                    "color: rgb(255, 255, 255);\n"
                                                                                                    "border-radius:0px;\n"
                                                                                                    "border:2px solid rgb(150,150,150);\n"
                                                                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                                                                    "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}\n"
                                                                                                                                                 "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                                 "color: rgb(255, 255, 255);\n"
                                                                                                                                                 "border-radius:0px;\n"
                                                                                                                                                 "border:2px solid rgb(150,150,150);\n"
                                                                                                                                                 "border-left:1px solid rgb(150,150,150);\n"
                                                                                                                                                 "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                                 "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}"],
                         [self.cbWeightIn, "QComboBox {\n"
                                           "background: white;\n"
                                           "border: 2px solid rgb(150,150,150);\n"
                                           "border-right: 1px solid rgb(150,150,150);\n"
                                           "border-bottom: 2px solid rgb(150,150,150);\n"
                                           "border-radius: 0px;\n"
                                           "}\n"
                                           "QComboBox:editable {\n"
                                           "background-color: rgb(142,187,208);\n"
                                           "color: black;\n"
                                           "border-radius: 0px;\n"
                                           "}"],
                         [self.le_VideoWeightIn, "background-color: rgb(235,235,235);\n"
                                                 "border: none;\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "camDefault460x262grey.png);"],
                         [self.le_VideoBack_4, "background-color: rgb(235,235,235);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(150,150,150);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "camDefault460x460grey.png);"],
                         [self.frmMediaPanelWeightOut, "background-color: rgb(227,227,227);\n"
                                                       "border-radius: 5px;\n"
                                                       "border: 2px solid rgb(150,150,150);\n"
                                                       "border-top: 1px solid rgb(150,150,150);\n"
                                                       "border-top-left-radius: 0px;\n"
                                                       "border-top-right-radius: 0px;"],
                         [self.btnPlayWeightOut, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:0px;\n"
                                                 "border:2px solid rgb(150,150,150);\n"
                                                 "border-left:1px solid rgb(150,150,150);\n"
                                                 "border-right:1px solid rgb(150,150,150);\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                                                              "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                              "color: rgb(255, 255, 255);\n"
                                                                                              "border-radius:0px;\n"
                                                                                              "border:2px solid rgb(150,150,150);\n"
                                                                                              "border-left:1px solid rgb(150,150,150);\n"
                                                                                              "border-right:1px solid rgb(150,150,150);\n"
                                                                                              "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                                                                                                           "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                           "color: rgb(255, 255, 255);\n"
                                                                                                                                           "border-radius:0px;\n"
                                                                                                                                           "border:2px solid rgb(150,150,150);\n"
                                                                                                                                           "border-left:1px solid rgb(150,150,150);\n"
                                                                                                                                           "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                           "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}"],
                         [self.SldrWeightOut, "QSlider {border-radius: none;\n"
                                              "border: none;}\n"
                                              "QSlider::groove:horizontal {\n"
                                              "    background-color: rgb(252,252,252);\n"
                                              "    border: 1px solid rgb(64,64,64);\n"
                                              "    height: 10px;\n"
                                              "    margin: 0px;\n"
                                              "    }\n"
                                              "QSlider::handle:horizontal {\n"
                                              "    background-color: rgb(64,64,64);\n"
                                              "    border: 1px solid rgb(64,64,64);\n"
                                              "    height: 10px;\n"
                                              "    width:  10px;\n"
                                              "    margin: -15px 0px;}"],
                         [self.btnStopWeightOut, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:0px;\n"
                                                 "border:2px solid rgb(150,150,150);\n"
                                                 "border-left:2px solid rgb(150,150,150);\n"
                                                 "border-right:1px solid rgb(150,150,150);\n"
                                                 "border-bottom-left-radius: 5px;\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);}\n"
                                                                                              "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                              "color: rgb(255, 255, 255);\n"
                                                                                              "border-radius:0px;\n"
                                                                                              "border:2px solid rgb(150,150,150);\n"
                                                                                              "border-left:2px solid rgb(150,150,150);\n"
                                                                                              "border-right:1px solid rgb(150,150,150);\n"
                                                                                              "border-bottom-left-radius: 5px;\n"
                                                                                              "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);}\n"
                                                                                                                                           "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                           "color: rgb(255, 255, 255);\n"
                                                                                                                                           "border-radius:0px;\n"
                                                                                                                                           "border:2px solid rgb(150,150,150);\n"
                                                                                                                                           "border-left:2px solid rgb(150,150,150);\n"
                                                                                                                                           "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                           "border-bottom-left-radius: 5px;\n"
                                                                                                                                           "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);};"],
                         [self.btnStepbackWeightOut, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:0px;\n"
                                                     "border:2px solid rgb(150,150,150);\n"
                                                     "border-left:1px solid rgb(150,150,150);\n"
                                                     "border-right:1px solid rgb(150,150,150);\n"
                                                     "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}\n"
                                                                                                  "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                                  "color: rgb(255, 255, 255);\n"
                                                                                                  "border-radius:0px;\n"
                                                                                                  "border:2px solid rgb(150,150,150);\n"
                                                                                                  "border-left:1px solid rgb(150,150,150);\n"
                                                                                                  "border-right:1px solid rgb(150,150,150);\n"
                                                                                                  "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}\n"
                                                                                                                                               "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                               "color: rgb(255, 255, 255);\n"
                                                                                                                                               "border-radius:0px;\n"
                                                                                                                                               "border:2px solid rgb(150,150,150);\n"
                                                                                                                                               "border-left:1px solid rgb(150,150,150);\n"
                                                                                                                                               "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                               "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}"],
                         [self.btnStepforwardWeightOut, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}\n"
                                                                                                     "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                                                                     "color: rgb(255, 255, 255);\n"
                                                                                                     "border-radius:0px;\n"
                                                                                                     "border:2px solid rgb(150,150,150);\n"
                                                                                                     "border-left:1px solid rgb(150,150,150);\n"
                                                                                                     "border-right:1px solid rgb(150,150,150);\n"
                                                                                                     "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}\n"
                                                                                                                                                  "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                                                                                                                  "color: rgb(255, 255, 255);\n"
                                                                                                                                                  "border-radius:0px;\n"
                                                                                                                                                  "border:2px solid rgb(150,150,150);\n"
                                                                                                                                                  "border-left:1px solid rgb(150,150,150);\n"
                                                                                                                                                  "border-right:1px solid rgb(150,150,150);\n"
                                                                                                                                                  "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}"],
                         [self.cbWeightOut, "QComboBox {\n"
                                            "background: white;\n"
                                            "border: 2px solid rgb(150,150,150);\n"
                                            "border-right: 1px solid rgb(150,150,150);\n"
                                            "border-bottom: 2px solid rgb(150,150,150);\n"
                                            "border-radius: 0px;\n"
                                            "}\n"
                                            "QComboBox:editable {\n"
                                            "background-color: rgb(142,187,208);\n"
                                            "color: black;\n"
                                            "border-radius: 0px;\n"
                                            "}"],
                         [self.label_13, "background-color: rgb(235,235,235);\n"
                                         "font: 11pt \"MS Shell Dlg 2\";\n"
                                         "border: none;"],
                         [self.label_14, "background-color: rgb(235,235,235);\n"
                                         "font: 11pt \"MS Shell Dlg 2\";\n"
                                         "border: none;"],
                         [self.label_56, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_57, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_59, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_60, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_61, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_64, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_65, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_67, "background-color: rgb(235,235,235);\n"
                                         "color: white;\n"
                                         "image: url(" + globalValues.pathStyleImgs + "iconinfo2.png);"],
                         [self.label_68, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_70, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_71, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_73, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_80, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_82, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_84, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_86, "background-color: rgb(235,235,235);\n"
                                         "color: rgb(0,0,0);\n"
                                         "font: 12pt \"MS Shell Dlg 2\";"],
                         [self.label_front3_3, "background-color: rgb(235,235,235);\n"
                                               "color: white;\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(150,150,150);"],
                         [self.line_5, "background-color: rgb(242,242,242);"],
                         [self.line_6, "background-color: rgb(242,242,242);"],
                         [self.line_8, "background-color: rgb(242,242,242);"],
                         [self.lblMainBoxCams_2, "background-color: rgb(242,242,242);\n"
                                                 "border-radius: 5px;\n"
                                                 "border: 2px solid rgb(205,205,205);"],
                         [self.le_GoneFromObject, "background-color: rgb(255,255,255);\n"
                                                  "color: rgb(0,0,0);\n"
                                                  "border-radius:3px;\n"
                                                  "border: 1px solid rgb(150,150,150);"],
                         [self.le_WeightObject, "background-color: rgb(255,255,255);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border-radius:3px;\n"
                                                "border: 1px solid rgb(150,150,150);"],
                         [self.le_CarModel, "background-color: rgb(255,255,255);\n"
                                            "color: rgb(0,0,0);\n"
                                            "border-radius:3px;\n"
                                            "border: 1px solid rgb(150,150,150);"],
                         [self.le_VolumeObject, "background-color: rgb(255,255,255);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border-radius:3px;\n"
                                                "border: 1px solid rgb(150,150,150);"],
                         [self.le_VolumePoligon, "background-color: rgb(255,255,255);\n"
                                                 "color: rgb(0,0,0);\n"
                                                 "border-radius:3px;\n"
                                                 "border: 1px solid rgb(150,150,150);"],
                         [self.le_ComeToPoligon, "background-color: rgb(255,255,255);\n"
                                                 "color: rgb(0,0,0);\n"
                                                 "border-radius:3px;\n"
                                                 "border: 1px solid rgb(150,150,150);"],
                         [self.le_Organisation, "background-color: rgb(255,255,255);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border-radius:3px;\n"
                                                "border: 1px solid rgb(150,150,150);"],
                         [self.le_ComeToObject, "background-color: rgb(255,255,255);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border-radius:3px;\n"
                                                "border: 1px solid rgb(150,150,150);"],
                         [self.le_WeightPoligon, "background-color: rgb(255,255,255);\n"
                                                 "color: rgb(0,0,0);\n"
                                                 "border-radius:3px;\n"
                                                 "border: 1px solid rgb(150,150,150);"],
                         [self.le_ZakazNumber, "background-color: rgb(255,255,255);\n"
                                               "color: rgb(0,0,0);\n"
                                               "border-radius:3px;\n"
                                               "border: 1px solid rgb(150,150,150);"],
                         [self.le_DriveOverTime, "background-color: rgb(255,255,255);\n"
                                                 "color: rgb(0,0,0);\n"
                                                 "border-radius:3px;\n"
                                                 "border: 1px solid rgb(150,150,150);"],
                         [self.le_ZakazState, "background-color: rgb(255,255,255);\n"
                                              "color: rgb(0,0,0);\n"
                                              "border-radius:3px;\n"
                                              "border: 1px solid rgb(150,150,150);"],
                         [self.le_GRZNumber, "background-color: rgb(255,255,255);\n"
                                             "color: rgb(0,0,0);\n"
                                             "border-radius:3px;\n"
                                             "border: 1px solid rgb(150,150,150);"],
                         [self.le_ZakazDate, "background-color: rgb(255,255,255);\n"
                                             "color: rgb(0,0,0);\n"
                                             "border-radius:3px;\n"
                                             "border: 1px solid rgb(150,150,150);"],
                         [self.frmHeadExit, "background-color: rgb(205,205,205);"],
                         [self.lblBarClose, "background-color: rgb(255,255,255);\n"
                                            "color: rgb(0,0,0);\n"
                                            ""],
                         [self.lblIconBarClose, "background-color: rgb(255,255,255);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "sinaps1717.png);"],
                         [self.btnCloseMainForm, "QPushButton:!hover{ background-color: rgb(255,255,255);\n"
                                                 "border-radius: 5px;\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconextbl1919.png);}\n"
                                                                                              "QPushButton:hover { background-color: rgb(84,122,181);\n"
                                                                                              "border-radius: 3px;\n"
                                                                                              "image: url(" + globalValues.pathStyleImgs + "iconext1919.png);}\n"
                                                                                                                                           "QPushButton:hover:pressed { background-color: rgb(50,75,115);\n"
                                                                                                                                           "border-radius: 3px;\n"
                                                                                                                                           "image: url(" + globalValues.pathStyleImgs + "iconext1919.png);};\n"
                                                                                                                                                                                        ""],
                         [self.label, "background-color: rgb(255,255,255);"]]

        self.lstDark = [[self.lblBack, "background-color: rgb(66,66,66);"],
                        [self.frmTableTS, "background-color: rgb(66,66,66);\n"
                                          "color: white;"],
                        [self.tblTS, "QTableWidget {background-color: rgb(42,42,42);\n"
                                     "border: 1px solid rgb(63,63,63);\n"
                                     "border-radius:5px;\n"
                                     "gridline-color: rgb(89,89,89);\n"
                                     "border-bottom-right-radius: 0px;\n"
                                     "color:white;}\n"
                                     "QLineEdit {background-color: white;}\n"
                                     "QHeaderView::section {\n"
                                     "gridline-color: rgb(89,89,89);\n"
                                     "background-color: rgb(50,75,115);};"],
                        [self.lblMainBoxCams, "background-color: rgb(62,62,62);\n"
                                              "border-radius: 5px;"],
                        [self.label_10, "background-color: rgb(50,75,115);\n"
                                        "color: white;\n"
                                        "border: 1px solid rgb(63,63,63);\n"
                                        "border-left-color: rgb(0,0,0);\n"
                                        "border-right-color: rgb(30,55,95);"],
                        [self.label_8, "background-color: rgb(50,75,115);\n"
                                       "border-radius: 5px;\n"
                                       "border: 1px solid rgb(63,63,63);\n"
                                       "border-bottom-left-radius: 0px;\n"
                                       "border-bottom-right-radius: 0px;\n"
                                       "border-top-left-radius: 0px;\n"
                                       "border-right-color: rgb(62,62,62);\n"
                                       "border-left-color: rgb(0,0,0);"],
                        [self.label_9, "background-color: rgb(62,62,62);\n"
                                       "border-radius: 0px;\n"
                                       "border: 0px solid rgb(62,62,62);\n"
                                       "border-bottom-left-radius: 0px;\n"
                                       "border-bottom-right-radius: 0px;\n"
                                       "border-top-right-radius: 0px;\n"
                                       "border-left-color: rgb(62,62,62);"],
                        [self.label_7, "background-color: rgb(50,75,115);\n"
                                       "border-radius: 5px;\n"
                                       "border: 1px solid rgb(63,63,63);\n"
                                       "border-bottom-left-radius: 0px;\n"
                                       "border-bottom-right-radius: 0px;\n"
                                       "border-top-left-radius: 0px;\n"
                                       "border-right-color: rgb(0,0,0);\n"
                                       "border-left-color: rgb(0,0,0);"],
                        [self.label_6, "background-color: rgb(50,75,115);\n"
                                       "color: white;\n"
                                       "border-radius: 5px;\n"
                                       "border: 1px solid rgb(63,63,63);\n"
                                       "border-bottom-left-radius: 0px;\n"
                                       "border-bottom-right-radius: 0px;\n"
                                       "border-top-right-radius: 0px;\n"
                                       "border-left-color: rgb(70,95,135);\n"
                                       "border-right-color: rgb(30,55,95);"],
                        [self.label_11, "background-color: rgb(50,75,115);\n"
                                        "color: white;\n"
                                        "border: 1px solid rgb(63,63,63);\n"
                                        "border-left-color: rgb(0,0,0);\n"
                                        "border-right-color: rgb(30,55,95);"],
                        [self.label_12, "background-color: rgb(50,75,115);\n"
                                        "color: white;\n"
                                        "border: 1px solid rgb(63,63,63);\n"
                                        "border-left-color: rgb(0,0,0);\n"
                                        "border-right-color: rgb(30,55,95);"],
                        [self.btnFilterRefresh, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(63,63,63);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconrefresh.png);}\n"
                                                                                             "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                                                             "color: rgb(255, 255, 255);\n"
                                                                                             "border-radius:3px;\n"
                                                                                             "border:1px solid rgb(63,63,63);\n"
                                                                                             "image: url(" + globalValues.pathStyleImgs + "iconrefresh.png);}\n"
                                                                                                                                          "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                                                                                                          "color: rgb(255, 255, 255);\n"
                                                                                                                                          "border-radius:3px;\n"
                                                                                                                                          "border:1px solid rgb(63,63,63);\n"
                                                                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconrefresh.png);};"],
                        [self.frmSrchDateRng, "background-color: rgb(62,62,62);"],
                        [self.label_DateRng_4, "background-color: rgb(62,62,62);\n"
                                               "color: rgb(255,255,255);\n"
                                               "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_dateAt_4, "background-color: rgb(62,62,62);\n"
                                              "color: rgb(255,255,255);\n"
                                              "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.DTEditRngAt, "background-color: rgb(89,89,89);\n"
                                           "color: rgb(255,255,255);\n"
                                           "font: 11pt \"MS Shell Dlg 2\";\n"
                                           "border-radius: 3px;\n"
                                           "border: 1px solid rgb(63,63,63)"],
                        [self.DTEditRngTo, "background-color: rgb(89,89,89);\n"
                                           "color: rgb(255,255,255);\n"
                                           "font: 11pt \"MS Shell Dlg 2\";\n"
                                           "border-radius: 3px;\n"
                                           "border: 1px solid rgb(63,63,63)"],
                        [self.label_DateTo_4, "background-color: rgb(62,62,62);\n"
                                              "color: rgb(255,255,255);\n"
                                              "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.btnFilter, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                         "color: rgb(255, 255, 255);\n"
                                         "border-radius:3px;\n"
                                         "border:1px solid rgb(63,63,63);}\n"
                                         "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                         "color: rgb(255, 255, 255);\n"
                                         "border-radius:3px;\n"
                                         "border:1px solid rgb(63,63,63);}\n"
                                         "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                         "color: rgb(255, 255, 255);\n"
                                         "border-radius:3px;\n"
                                         "border:1px solid rgb(63,63,63);};"],
                        [self.label_filtertype, "background-color: rgb(62,62,62);\n"
                                                "color: rgb(255,255,255);\n"
                                                "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.comboFilter, "background-color: rgb(89,89,89);\n"
                                           "font: 11pt \"MS Shell Dlg 2\";\n"
                                           "color: rgb(255,255,255);\n"
                                           "border-radius: 3px;\n"
                                           "border: 1px solid rgb(63,63,63)"],
                        [self.label_iconfilter, "background-color: rgb(62,62,62);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconfilter2.png);"],
                        [self.line, "background-color: rgb(89,89,89);"],
                        [self.frmSrchDateFix, "background-color: rgb(62,62,62);"],
                        [self.label_DateFix, "background-color: rgb(62,62,62);\n"
                                             "color: rgb(255,255,255);\n"
                                             "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.DTEitFix, "background-color: rgb(89,89,89);\n"
                                        "color: rgb(255,255,255);\n"
                                        "font: 11pt \"MS Shell Dlg 2\";\n"
                                        "border-radius: 3px;\n"
                                        "border: 1px solid rgb(63,63,63)"],
                        [self.btnSearch, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                         "color: rgb(255, 255, 255);\n"
                                         "border-radius:3px;\n"
                                         "border:1px solid rgb(63,63,63);\n"
                                         "image: url(" + globalValues.pathStyleImgs + "iconsearchwhite.png);}\n"
                                                                                      "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                                                      "color: rgb(255, 255, 255);\n"
                                                                                      "border-radius:3px;\n"
                                                                                      "border:1px solid rgb(63,63,63);\n"
                                                                                      "image: url(" + globalValues.pathStyleImgs + "iconsearchwhite.png);}\n"
                                                                                                                                   "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                                                                                                   "color: rgb(255, 255, 255);\n"
                                                                                                                                   "border-radius:3px;\n"
                                                                                                                                   "border:1px solid rgb(63,63,63);\n"
                                                                                                                                   "image: url(" + globalValues.pathStyleImgs + "iconsearchwhite.png);};"],
                        [self.leSearchTblTS, "background-color: rgb(42, 42, 42);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius: 3px;"],
                        [self.comboSearch, "background-color: rgb(89,89,89);\n"
                                           "color: white;\n"
                                           "border-radius: 3px;\n"
                                           "border: 1px solid rgb(63,63,63)"],
                        [self.btnSrchRefresh, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconrefreshwhite.png);}\n"
                                                                                           "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                                                           "color: rgb(255, 255, 255);\n"
                                                                                           "border-radius:3px;\n"
                                                                                           "border:1px solid rgb(63,63,63);\n"
                                                                                           "image: url(" + globalValues.pathStyleImgs + "iconrefreshwhite.png);}\n"
                                                                                                                                        "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                                                                                                        "color: rgb(255, 255, 255);\n"
                                                                                                                                        "border-radius:3px;\n"
                                                                                                                                        "border:1px solid rgb(63,63,63);\n"
                                                                                                                                        "image: url(" + globalValues.pathStyleImgs + "iconrefreshwhite.png);};"],
                        [self.line_2, "background-color: rgb(89,89,89);"],
                        [self.line_3, "background-color: rgb(89,89,89);"],
                        [self.line_4, "background-color: rgb(89,89,89);"],
                        [self.btnFilterPrint, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconprint8.png);}\n"
                                                                                           "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                                                           "color: rgb(255, 255, 255);\n"
                                                                                           "border-radius:3px;\n"
                                                                                           "border:1px solid rgb(63,63,63);\n"
                                                                                           "image: url(" + globalValues.pathStyleImgs + "iconprint8.png);}\n"
                                                                                                                                        "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                                                                                                        "color: rgb(255, 255, 255);\n"
                                                                                                                                        "border-radius:3px;\n"
                                                                                                                                        "border:1px solid rgb(63,63,63);\n"
                                                                                                                                        "image: url(" + globalValues.pathStyleImgs + "iconprint8.png);};"],
                        [self.btnFilterSave, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                             "color: rgb(255,255,255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconsaveWHITE.png);}\n"
                                                                                          "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                                                          "color: rgb(255, 255, 255);\n"
                                                                                          "border-radius:3px;\n"
                                                                                          "border:1px solid rgb(135,135,135);\n"
                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconsaveWHITE.png);}\n"
                                                                                                                                       "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                                                                                                       "color: rgb(255, 255, 255);\n"
                                                                                                                                       "border-radius:3px;\n"
                                                                                                                                       "border:1px solid rgb(135,135,135);\n"
                                                                                                                                       "image: url(" + globalValues.pathStyleImgs + "iconsaveWHITE.png);};"],
                        [self.verticalScrollBar, "QScrollBar:vertical {\n"
                                                 "border: 1px solid rgb(63,63,63);\n"
                                                 " background: rgb(63,63,63);\n"
                                                 "width:10px;\n"
                                                 "margin: 0px 0px 0px 0px;\n"
                                                 "}\n"
                                                 "QScrollBar::add-page:vertical {background: rgb(89,89,89);}\n"
                                                 "QScrollBar::sub-page:vertical {background: rgb(89,89,89);}\n"
                                                 "QScrollBar::handle:vertical {\n"
                                                 "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0 rgb(50,75,115), stop: 0.5 rgb(50,75,115), stop:1 rgb(50,75,115));\n"
                                                 "min-height: 0px;\n"
                                                 "}\n"
                                                 "QScrollBar::add-line:vertical {\n"
                                                 " background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop: 0 rgb(50,75,115), stop: 0.5 rgb(50,75,115),  stop:1 rgb(50,75,115));\n"
                                                 "height: 0px;\n"
                                                 "subcontrol-position: bottom;\n"
                                                 "subcontrol-origin: margin;\n"
                                                 "}\n"
                                                 "QScrollBar::sub-line:vertical {\n"
                                                 "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0  rgb(50,75,115), stop: 0.5 rgb(50,75,115),  stop:1 rgb(50,75,115));\n"
                                                 "height: 0 px;\n"
                                                 "subcontrol-position: top;\n"
                                                 "subcontrol-origin: margin;\n"
                                                 "}"],
                        [self.frmINFO, "background-color: rgb(66,66,66);\n"
                                       "color: white;"],
                        [self.TabFrameVideo, "QTabWidget::pane {border:none;\n"
                                             "border-radius: 5px}\n"
                                             "QTabBar::tab {background: rgb(50,75,115);\n"
                                             "border: 1px solid rgb(30,55,95);\n"
                                             "border-left-color: rgb(0,0,0);\n"
                                             "padding: 5px;\n"
                                             "padding-bottom: 4px;\n"
                                             "padding-top: 4px;\n"
                                             "padding-right:14px;\n"
                                             "padding-left: 7px;\n"
                                             "font: 8pt \"MS Shell Dlg 2\";\n"
                                             "color:white;}\n"
                                             "QTabBar::tab:selected {background: rgb(84,122,181);\n"
                                             "margin-bottom: -1px;\n"
                                             "color:white;\n"
                                             "border: 0px solid(75,75,75);}\n"
                                             "QTabBar::tab:hover {background-color: rgb(70,95,135);}\n"
                                             "QTabBar::tab:selected:hover {background-color: rgb(84,122,181);}"],
                        [self.label_51, "background-color: rgb(75,75,75);\n"
                                        "border-top-left-radius: 0px;\n"
                                        "border-bottom-left-radius: 10px;\n"
                                        "border-bottom-right-radius: 10px;\n"
                                        "border-top-right-radius: 10px;"],
                        [self.le_VideoBack_2, "background-color: rgb(0,0,0);\n"
                                              "border-radius: 5px;\n"
                                              "border: 2px solid rgb(42,42,42);"],
                        [self.frmMediaPanelKppIn, "background-color: rgb(89,89,89);\n"
                                                  "border-radius: 5px;\n"
                                                  "border: 2px solid rgb(42,42,42);\n"
                                                  "border-top: 1px solid rgb(42,42,42);\n"
                                                  "border-top-left-radius: 0px;\n"
                                                  "border-top-right-radius: 0px;"],
                        [self.btnPlayKppIn, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:0px;\n"
                                            "border:2px solid rgb(42,42,42);\n"
                                            "border-left:1px solid rgb(42,42,42);\n"
                                            "border-right:1px solid rgb(42,42,42);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                                                         "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                         "color: rgb(255, 255, 255);\n"
                                                                                         "border-radius:0px;\n"
                                                                                         "border:2px solid rgb(42,42,42);\n"
                                                                                         "border-left:1px solid rgb(42,42,42);\n"
                                                                                         "border-right:1px solid rgb(42,42,42);\n"
                                                                                         "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                                                                                                      "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                      "color: rgb(255, 255, 255);\n"
                                                                                                                                      "border-radius:0px;\n"
                                                                                                                                      "border:2px solid rgb(42,42,42);\n"
                                                                                                                                      "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                      "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                      "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}"],
                        [self.SldrKppIn, "QSlider {border-radius: none;\n"
                                         "border: none;}\n"
                                         "QSlider::groove:horizontal {\n"
                                         "    background-color: rgb(252,252,252);\n"
                                         "    border: 1px solid rgb(64,64,64);\n"
                                         "    height: 10px;\n"
                                         "    margin: 0px;\n"
                                         "    }\n"
                                         "QSlider::handle:horizontal {\n"
                                         "    background-color: rgb(64,64,64);\n"
                                         "    border: 1px solid rgb(64,64,64);\n"
                                         "    height: 10px;\n"
                                         "    width:  10px;\n"
                                         "    margin: -15px 0px;}"],
                        [self.btnStopKppIn, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:0px;\n"
                                            "border:2px solid rgb(42,42,42);\n"
                                            "border-left:2px solid rgb(42,42,42);\n"
                                            "border-right:1px solid rgb(42,42,42);\n"
                                            "border-bottom-left-radius: 5px;\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);}\n"
                                                                                         "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                         "color: rgb(255, 255, 255);\n"
                                                                                         "border-radius:0px;\n"
                                                                                         "border:2px solid rgb(42,42,42);\n"
                                                                                         "border-left:2px solid rgb(42,42,42);\n"
                                                                                         "border-right:1px solid rgb(42,42,42);\n"
                                                                                         "border-bottom-left-radius: 5px;\n"
                                                                                         "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);}\n"
                                                                                                                                      "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                      "color: rgb(255, 255, 255);\n"
                                                                                                                                      "border-radius:0px;\n"
                                                                                                                                      "border:2px solid rgb(42,42,42);\n"
                                                                                                                                      "border-left:2px solid rgb(42,42,42);\n"
                                                                                                                                      "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                      "border-bottom-left-radius: 5px;\n"
                                                                                                                                      "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);};"],
                        [self.btnStepbackKppIn, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:0px;\n"
                                                "border:2px solid rgb(42,42,42);\n"
                                                "border-left:1px solid rgb(42,42,42);\n"
                                                "border-right:1px solid rgb(42,42,42);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}\n"
                                                                                             "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                             "color: rgb(255, 255, 255);\n"
                                                                                             "border-radius:0px;\n"
                                                                                             "border:2px solid rgb(42,42,42);\n"
                                                                                             "border-left:1px solid rgb(42,42,42);\n"
                                                                                             "border-right:1px solid rgb(42,42,42);\n"
                                                                                             "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}\n"
                                                                                                                                          "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                          "color: rgb(255, 255, 255);\n"
                                                                                                                                          "border-radius:0px;\n"
                                                                                                                                          "border:2px solid rgb(42,42,42);\n"
                                                                                                                                          "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                          "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}"],
                        [self.btnStepforwardKppIn, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                   "color: rgb(255, 255, 255);\n"
                                                   "border-radius:0px;\n"
                                                   "border:2px solid rgb(42,42,42);\n"
                                                   "border-left:1px solid rgb(42,42,42);\n"
                                                   "border-right:1px solid rgb(42,42,42);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}\n"
                                                                                                "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                                "color: rgb(255, 255, 255);\n"
                                                                                                "border-radius:0px;\n"
                                                                                                "border:2px solid rgb(42,42,42);\n"
                                                                                                "border-left:1px solid rgb(42,42,42);\n"
                                                                                                "border-right:1px solid rgb(42,42,42);\n"
                                                                                                "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}\n"
                                                                                                                                             "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                             "color: rgb(255, 255, 255);\n"
                                                                                                                                             "border-radius:0px;\n"
                                                                                                                                             "border:2px solid rgb(42,42,42);\n"
                                                                                                                                             "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                             "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                             "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}"],
                        [self.cbKppIn, "QComboBox {\n"
                                       "background: white;\n"
                                       "border: 2px solid rgb(42,42,42);\n"
                                       "border-right: 1px solid rgb(42,42,42);\n"
                                       "border-bottom: 2px solid rgb(42,42,42);\n"
                                       "border-radius: 0px;\n"
                                       "color: white;\n"
                                       "}\n"
                                       "QComboBox:editable {\n"
                                       "background-color: rgb(50,75,115);\n"
                                       "color: black;\n"
                                       "border-radius: 0px;\n"
                                       "}"],
                        [self.le_VideoKppOut, "background-color: rgb(0,0,0);\n"
                                              "border: none;\n"
                                              "image: url(" + globalValues.pathStyleImgs + "camDefault395x262.png);"],
                        [self.frmMediaPanelKppOut, "background-color: rgb(89,89,89);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 2px solid rgb(42,42,42);\n"
                                                   "border-top: 1px solid rgb(42,42,42);\n"
                                                   "border-top-left-radius: 0px;\n"
                                                   "border-top-right-radius: 0px;"],
                        [self.btnPlayKppOut, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:0px;\n"
                                             "border:2px solid rgb(42,42,42);\n"
                                             "border-left:1px solid rgb(42,42,42);\n"
                                             "border-right:1px solid rgb(42,42,42);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                                                          "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                          "color: rgb(255, 255, 255);\n"
                                                                                          "border-radius:0px;\n"
                                                                                          "border:2px solid rgb(42,42,42);\n"
                                                                                          "border-left:1px solid rgb(42,42,42);\n"
                                                                                          "border-right:1px solid rgb(42,42,42);\n"
                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                                                                                                       "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                       "color: rgb(255, 255, 255);\n"
                                                                                                                                       "border-radius:0px;\n"
                                                                                                                                       "border:2px solid rgb(42,42,42);\n"
                                                                                                                                       "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                       "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                       "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}"],
                        [self.SldrKppOut, "QSlider {border-radius: none;\n"
                                          "border: none;}\n"
                                          "QSlider::groove:horizontal {\n"
                                          "    background-color: rgb(252,252,252);\n"
                                          "    border: 1px solid rgb(64,64,64);\n"
                                          "    height: 10px;\n"
                                          "    margin: 0px;\n"
                                          "    }\n"
                                          "QSlider::handle:horizontal {\n"
                                          "    background-color: rgb(64,64,64);\n"
                                          "    border: 1px solid rgb(64,64,64);\n"
                                          "    height: 10px;\n"
                                          "    width:  10px;\n"
                                          "    margin: -15px 0px;}"],
                        [self.btnStopKppOut, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:0px;\n"
                                             "border:2px solid rgb(42,42,42);\n"
                                             "border-left:1px solid rgb(42,42,42);\n"
                                             "border-right:1px solid rgb(42,42,42);\n"
                                             "border-bottom-left-radius: 5px;\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);}\n"
                                                                                          "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                          "color: rgb(255, 255, 255);\n"
                                                                                          "border-radius:0px;\n"
                                                                                          "border:2px solid rgb(42,42,42);\n"
                                                                                          "border-left:1px solid rgb(42,42,42);\n"
                                                                                          "border-right:1px solid rgb(42,42,42);\n"
                                                                                          "border-bottom-left-radius: 5px;\n"
                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);}\n"
                                                                                                                                       "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                       "color: rgb(255, 255, 255);\n"
                                                                                                                                       "border-radius:0px;\n"
                                                                                                                                       "border:2px solid rgb(42,42,42);\n"
                                                                                                                                       "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                       "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                       "border-bottom-left-radius: 5px;\n"
                                                                                                                                       "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);};"],
                        [self.btnStepbackKppOut, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:0px;\n"
                                                 "border:2px solid rgb(42,42,42);\n"
                                                 "border-left:1px solid rgb(42,42,42);\n"
                                                 "border-right:1px solid rgb(42,42,42);\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}\n"
                                                                                              "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                              "color: rgb(255, 255, 255);\n"
                                                                                              "border-radius:0px;\n"
                                                                                              "border:2px solid rgb(42,42,42);\n"
                                                                                              "border-left:1px solid rgb(42,42,42);\n"
                                                                                              "border-right:1px solid rgb(42,42,42);\n"
                                                                                              "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}\n"
                                                                                                                                           "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                           "color: rgb(255, 255, 255);\n"
                                                                                                                                           "border-radius:0px;\n"
                                                                                                                                           "border:2px solid rgb(42,42,42);\n"
                                                                                                                                           "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                           "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                           "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}"],
                        [self.btnStepforwardKppOut, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}\n"
                                                                                                 "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                                 "color: rgb(255, 255, 255);\n"
                                                                                                 "border-radius:0px;\n"
                                                                                                 "border:2px solid rgb(42,42,42);\n"
                                                                                                 "border-left:1px solid rgb(42,42,42);\n"
                                                                                                 "border-right:1px solid rgb(42,42,42);\n"
                                                                                                 "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}\n"
                                                                                                                                              "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                              "color: rgb(255, 255, 255);\n"
                                                                                                                                              "border-radius:0px;\n"
                                                                                                                                              "border:2px solid rgb(42,42,42);\n"
                                                                                                                                              "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                              "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                              "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}"],
                        [self.cbKppOut, "QComboBox {\n"
                                        "background: white;\n"
                                        "border: 2px solid rgb(42,42,42);\n"
                                        "border-right: 1px solid rgb(42,42,42);\n"
                                        "border-bottom: 2px solid rgb(42,42,42);\n"
                                        "border-radius: 0px;\n"
                                        "}\n"
                                        "QComboBox:editable {\n"
                                        "background-color: rgb(50,75,115);\n"
                                        "color: black;\n"
                                        "border-radius: 0px;\n"
                                        "}"],
                        [self.le_VideoKppIn, "background-color: rgb(0,0,0);\n"
                                             "border: none;\n"
                                             "image: url(" + globalValues.pathStyleImgs + "camDefault395x262.png);"],
                        [self.le_VideoBack_4, "background-color: rgb(0,0,0);\n"
                                              "border-radius: 5px;\n"
                                              "border: 2px solid rgb(42,42,42);"],
                        [self.label_15, "background-color: rgb(0,0,0);\n"
                                        "color: white;\n"
                                        "font: 11pt \"MS Shell Dlg 2\";\n"
                                        "border: none;"],
                        [self.label_16, "background-color: rgb(0,0,0);\n"
                                        "color: white;\n"
                                        "font: 11pt \"MS Shell Dlg 2\";\n"
                                        "border: none;"],
                        [self.tabWeight, ""],
                        [self.le_VideoBack, "background-color: rgb(0,0,0);\n"
                                            "border-radius: 5px;\n"
                                            "border: 2px solid rgb(42,42,42);"],
                        [self.label_50, "background-color: rgb(75,75,75);\n"
                                        "border-top-left-radius: 0px;\n"
                                        "border-bottom-left-radius: 5px;\n"
                                        "border-bottom-right-radius: 5px;\n"
                                        "border-top-right-radius: 5px;"],
                        [self.frmMediaPanelWeightIn, "background-color: rgb(89,89,89);\n"
                                                     "border-radius: 5px;\n"
                                                     "border: 2px solid rgb(42,42,42);\n"
                                                     "border-top: 1px solid rgb(42,42,42);\n"
                                                     "border-top-left-radius: 0px;\n"
                                                     "border-top-right-radius: 0px;"],
                        [self.btnPlayWeightIn, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:0px;\n"
                                               "border:2px solid rgb(42,42,42);\n"
                                               "border-left:1px solid rgb(42,42,42);\n"
                                               "border-right:1px solid rgb(42,42,42);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                                                            "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                            "color: rgb(255, 255, 255);\n"
                                                                                            "border-radius:0px;\n"
                                                                                            "border:2px solid rgb(42,42,42);\n"
                                                                                            "border-left:1px solid rgb(42,42,42);\n"
                                                                                            "border-right:1px solid rgb(42,42,42);\n"
                                                                                            "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                                                                                                         "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                         "color: rgb(255, 255, 255);\n"
                                                                                                                                         "border-radius:0px;\n"
                                                                                                                                         "border:2px solid rgb(42,42,42);\n"
                                                                                                                                         "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                         "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                         "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}"],
                        [self.SldrWeightIn, "QSlider {border-radius: none;\n"
                                            "border: none;}\n"
                                            "QSlider::groove:horizontal {\n"
                                            "    background-color: rgb(252,252,252);\n"
                                            "    border: 1px solid rgb(64,64,64);\n"
                                            "    height: 10px;\n"
                                            "    margin: 0px;\n"
                                            "    }\n"
                                            "QSlider::handle:horizontal {\n"
                                            "    background-color: rgb(64,64,64);\n"
                                            "    border: 1px solid rgb(64,64,64);\n"
                                            "    height: 10px;\n"
                                            "    width:  10px;\n"
                                            "    margin: -15px 0px;}"],
                        [self.btnStopWeightIn, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:0px;\n"
                                               "border:2px solid rgb(42,42,42);\n"
                                               "border-left:1px solid rgb(42,42,42);\n"
                                               "border-right:1px solid rgb(42,42,42);\n"
                                               "border-bottom-left-radius: 5px;\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);}\n"
                                                                                            "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                            "color: rgb(255, 255, 255);\n"
                                                                                            "border-radius:0px;\n"
                                                                                            "border:2px solid rgb(42,42,42);\n"
                                                                                            "border-left:1px solid rgb(42,42,42);\n"
                                                                                            "border-right:1px solid rgb(42,42,42);\n"
                                                                                            "border-bottom-left-radius: 5px;\n"
                                                                                            "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);}\n"
                                                                                                                                         "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                         "color: rgb(255, 255, 255);\n"
                                                                                                                                         "border-radius:0px;\n"
                                                                                                                                         "border:2px solid rgb(42,42,42);\n"
                                                                                                                                         "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                         "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                         "border-bottom-left-radius: 5px;\n"
                                                                                                                                         "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);};"],
                        [self.btnStepbackWeightIn, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                   "color: rgb(255, 255, 255);\n"
                                                   "border-radius:0px;\n"
                                                   "border:2px solid rgb(42,42,42);\n"
                                                   "border-left:1px solid rgb(42,42,42);\n"
                                                   "border-right:1px solid rgb(42,42,42);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}\n"
                                                                                                "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                                "color: rgb(255, 255, 255);\n"
                                                                                                "border-radius:0px;\n"
                                                                                                "border:2px solid rgb(42,42,42);\n"
                                                                                                "border-left:1px solid rgb(42,42,42);\n"
                                                                                                "border-right:1px solid rgb(42,42,42);\n"
                                                                                                "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}\n"
                                                                                                                                             "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                             "color: rgb(255, 255, 255);\n"
                                                                                                                                             "border-radius:0px;\n"
                                                                                                                                             "border:2px solid rgb(42,42,42);\n"
                                                                                                                                             "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                             "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                             "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}"],
                        [self.btnStepforwardWeightIn, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                      "color: rgb(255, 255, 255);\n"
                                                      "border-radius:0px;\n"
                                                      "border:2px solid rgb(42,42,42);\n"
                                                      "border-left:1px solid rgb(42,42,42);\n"
                                                      "border-right:1px solid rgb(42,42,42);\n"
                                                      "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}\n"
                                                                                                   "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                                   "color: rgb(255, 255, 255);\n"
                                                                                                   "border-radius:0px;\n"
                                                                                                   "border:2px solid rgb(42,42,42);\n"
                                                                                                   "border-left:1px solid rgb(42,42,42);\n"
                                                                                                   "border-right:1px solid rgb(42,42,42);\n"
                                                                                                   "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}\n"
                                                                                                                                                "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                                "color: rgb(255, 255, 255);\n"
                                                                                                                                                "border-radius:0px;\n"
                                                                                                                                                "border:2px solid rgb(42,42,42);\n"
                                                                                                                                                "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                                "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                                "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}"],
                        [self.cbWeightIn, "QComboBox {\n"
                                          "background: white;\n"
                                          "border: 2px solid rgb(42,42,42);\n"
                                          "border-right: 1px solid rgb(42,42,42);\n"
                                          "border-bottom: 2px solid rgb(42,42,42);\n"
                                          "border-radius: 0px;\n"
                                          "}\n"
                                          "QComboBox:editable {\n"
                                          "background-color: rgb(50,75,115);\n"
                                          "color: black;\n"
                                          "border-radius: 0px;\n"
                                          "}"],
                        [self.le_VideoWeightOut, "background-color: rgb(0,0,0);\n"
                                                 "border: none;\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "camDefault395x262.png);"],
                        [self.le_VideoBack_3, "background-color: rgb(0,0,0);\n"
                                              "border-radius: 5px;\n"
                                              "border: 2px solid rgb(42,42,42);"],
                        [self.frmMediaPanelWeightOut, "background-color: rgb(89,89,89);\n"
                                                      "border-radius: 5px;\n"
                                                      "border: 2px solid rgb(42,42,42);\n"
                                                      "border-top: 1px solid rgb(42,42,42);\n"
                                                      "border-top-left-radius: 0px;\n"
                                                      "border-top-right-radius: 0px;"],
                        [self.btnPlayWeightOut, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:0px;\n"
                                                "border:2px solid rgb(42,42,42);\n"
                                                "border-left:1px solid rgb(42,42,42);\n"
                                                "border-right:1px solid rgb(42,42,42);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                                                             "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                             "color: rgb(255, 255, 255);\n"
                                                                                             "border-radius:0px;\n"
                                                                                             "border:2px solid rgb(42,42,42);\n"
                                                                                             "border-left:1px solid rgb(42,42,42);\n"
                                                                                             "border-right:1px solid rgb(42,42,42);\n"
                                                                                             "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                                                                                                          "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                          "color: rgb(255, 255, 255);\n"
                                                                                                                                          "border-radius:0px;\n"
                                                                                                                                          "border:2px solid rgb(42,42,42);\n"
                                                                                                                                          "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                          "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}"],
                        [self.SldrWeightOut, "QSlider {border-radius: none;\n"
                                             "border: none;}\n"
                                             "QSlider::groove:horizontal {\n"
                                             "    background-color: rgb(252,252,252);\n"
                                             "    border: 1px solid rgb(64,64,64);\n"
                                             "    height: 10px;\n"
                                             "    margin: 0px;\n"
                                             "    }\n"
                                             "QSlider::handle:horizontal {\n"
                                             "    background-color: rgb(64,64,64);\n"
                                             "    border: 1px solid rgb(64,64,64);\n"
                                             "    height: 10px;\n"
                                             "    width:  10px;\n"
                                             "    margin: -15px 0px;}"],
                        [self.btnStopWeightOut, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:0px;\n"
                                                "border:2px solid rgb(42,42,42);\n"
                                                "border-left:1px solid rgb(42,42,42);\n"
                                                "border-right:1px solid rgb(42,42,42);\n"
                                                "border-bottom-left-radius: 5px;\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);}\n"
                                                                                             "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                             "color: rgb(255, 255, 255);\n"
                                                                                             "border-radius:0px;\n"
                                                                                             "border:2px solid rgb(42,42,42);\n"
                                                                                             "border-left:1px solid rgb(42,42,42);\n"
                                                                                             "border-right:1px solid rgb(42,42,42);\n"
                                                                                             "border-bottom-left-radius: 5px;\n"
                                                                                             "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);}\n"
                                                                                                                                          "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                          "color: rgb(255, 255, 255);\n"
                                                                                                                                          "border-radius:0px;\n"
                                                                                                                                          "border:2px solid rgb(42,42,42);\n"
                                                                                                                                          "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                          "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                          "border-bottom-left-radius: 5px;\n"
                                                                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);};"],
                        [self.btnStepbackWeightOut, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}\n"
                                                                                                 "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                                 "color: rgb(255, 255, 255);\n"
                                                                                                 "border-radius:0px;\n"
                                                                                                 "border:2px solid rgb(42,42,42);\n"
                                                                                                 "border-left:1px solid rgb(42,42,42);\n"
                                                                                                 "border-right:1px solid rgb(42,42,42);\n"
                                                                                                 "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}\n"
                                                                                                                                              "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                              "color: rgb(255, 255, 255);\n"
                                                                                                                                              "border-radius:0px;\n"
                                                                                                                                              "border:2px solid rgb(42,42,42);\n"
                                                                                                                                              "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                              "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                              "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}"],
                        [self.btnStepforwardWeightOut, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                       "color: rgb(255, 255, 255);\n"
                                                       "border-radius:0px;\n"
                                                       "border:2px solid rgb(42,42,42);\n"
                                                       "border-left:1px solid rgb(42,42,42);\n"
                                                       "border-right:1px solid rgb(42,42,42);\n"
                                                       "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}\n"
                                                                                                    "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                                                                    "color: rgb(255, 255, 255);\n"
                                                                                                    "border-radius:0px;\n"
                                                                                                    "border:2px solid rgb(42,42,42);\n"
                                                                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                                                                    "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}\n"
                                                                                                                                                 "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                                                                                                                 "color: rgb(255, 255, 255);\n"
                                                                                                                                                 "border-radius:0px;\n"
                                                                                                                                                 "border:2px solid rgb(42,42,42);\n"
                                                                                                                                                 "border-left:1px solid rgb(42,42,42);\n"
                                                                                                                                                 "border-right:1px solid rgb(42,42,42);\n"
                                                                                                                                                 "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}"],
                        [self.cbWeightOut, "QComboBox {\n"
                                           "background: white;\n"
                                           "border: 2px solid rgb(42,42,42);\n"
                                           "border-right: 1px solid rgb(42,42,42);\n"
                                           "border-bottom: 2px solid rgb(42,42,42);\n"
                                           "border-radius: 0px;\n"
                                           "}\n"
                                           "QComboBox:editable {\n"
                                           "background-color: rgb(50,75,115);\n"
                                           "color: black;\n"
                                           "border-radius: 0px;\n"
                                           "}"],
                        [self.le_VideoWeightIn, "background-color: rgb(0,0,0);\n"
                                                "border: none;\n"
                                                "image: url(" + globalValues.pathStyleImgs + "camDefault395x262.png);"],
                        [self.label_14, "background-color: rgb(0,0,0);\n"
                                        "color: white;\n"
                                        "font: 11pt \"MS Shell Dlg 2\";\n"
                                        "border: none;"],
                        [self.label_13, "background-color: rgb(0,0,0);\n"
                                        "color: white;\n"
                                        "font: 11pt \"MS Shell Dlg 2\";\n"
                                        "border: none;"],
                        [self.label_56, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_57, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_59, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_60, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_61, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_64, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_65, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_67, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "image: url(" + globalValues.pathStyleImgs + "iconinfo2.png);"],
                        [self.label_68, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_70, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_71, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_73, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_80, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_82, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_84, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_86, "background-color: rgb(75,75,75);\n"
                                        "color: white;\n"
                                        "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.label_front3_3, "background-color: rgb(75,75,75);\n"
                                              "border-radius: 5px;"],
                        [self.line_5, "background-color: rgb(75,75,75);"],
                        [self.line_6, "background-color: rgb(75,75,75);"],
                        [self.line_8, "background-color: rgb(75,75,75);"],
                        [self.lblMainBoxCams_2, "background-color: rgb(62,62,62);\n"
                                                "border-radius: 5px;"],
                        [self.le_Organisation, "background-color: rgb(89,89,89);\n"
                                               "color: white;\n"
                                               "border-radius: 3px;\n"
                                               "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_GRZNumber, "background-color: rgb(89,89,89);\n"
                                            "color: white;\n"
                                            "border-radius: 3px;\n"
                                            "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_CarModel, "background-color: rgb(89,89,89);\n"
                                           "color: white;\n"
                                           "border-radius: 3px;\n"
                                           "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_ZakazNumber, "background-color: rgb(89,89,89);\n"
                                              "color: white;\n"
                                              "border-radius: 3px;\n"
                                              "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_ZakazDate, "background-color: rgb(89,89,89);\n"
                                            "color: white;\n"
                                            "border-radius: 3px;\n"
                                            "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_ZakazState, "background-color: rgb(89,89,89);\n"
                                             "color: white;\n"
                                             "border-radius: 3px;\n"
                                             "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_ComeToObject, "background-color: rgb(89,89,89);\n"
                                               "color: white;\n"
                                               "border-radius: 3px;\n"
                                               "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_GoneFromObject, "background-color: rgb(89,89,89);\n"
                                                 "color: white;\n"
                                                 "border-radius: 3px;\n"
                                                 "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_ComeToPoligon, "background-color: rgb(89,89,89);\n"
                                                "color: white;\n"
                                                "border-radius: 3px;\n"
                                                "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_DriveOverTime, "background-color: rgb(89,89,89);\n"
                                                "color: white;\n"
                                                "border-radius: 3px;\n"
                                                "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_WeightObject, "background-color: rgb(89,89,89);\n"
                                               "color: white;\n"
                                               "border-radius: 3px;\n"
                                               "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_VolumeObject, "background-color: rgb(89,89,89);\n"
                                               "color: white;\n"
                                               "border-radius: 3px;\n"
                                               "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_WeightPoligon, "background-color: rgb(89,89,89);\n"
                                                "color: white;\n"
                                                "border-radius: 3px;\n"
                                                "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.le_VolumePoligon, "background-color: rgb(89,89,89);\n"
                                                "color: white;\n"
                                                "border-radius: 3px;\n"
                                                "font: 12pt \"MS Shell Dlg 2\";"],
                        [self.lblBarClose, "background-color: rgb(255,255,255);\n"
                                           "color: rgb(0,0,0);\n"
                                           ""],
                        [self.lblIconBarClose, "background-color: rgb(255,255,255);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "sinaps1717.png);"],
                        [self.btnCloseMainForm, "QPushButton:!hover{ background-color: rgb(255,255,255);\n"
                                                "border-radius: 5px;\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconextbl1919.png);}\n"
                                                                                             "QPushButton:hover { background-color: rgb(84,122,181);\n"
                                                                                             "border-radius: 3px;\n"
                                                                                             "image: url(" + globalValues.pathStyleImgs + "iconext1919.png);}\n"
                                                                                                                                          "QPushButton:hover:pressed { background-color: rgb(50,75,115);\n"
                                                                                                                                          "border-radius: 3px;\n"
                                                                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconext1919.png);};\n"
                                                                                                                                                                                       ""],
                        [self.label, "background-color: rgb(255,255,255);"]]

        self.lengthLight = len(self.lstLight)
        self.lengthDark = len(self.lstDark)

        self.changeCOLORMainPanelGrunt()

        myQHeaderView = self.tblTS.horizontalHeader()
        for i in range(14):
            myQHeaderView.setSectionResizeMode(i, QtWidgets.QHeaderView.Fixed)
        myQHeaderView.setStretchLastSection(False)

        dateToday = datetime.date.today().strftime('%d.%m.%Y')
        str_date_today = str(dateToday)
        cur_time = datetime.datetime.time(datetime.datetime.now())
        str_cur_time = str(cur_time)
        len_cur_time = len(str_cur_time)
        str_cur_time = str_cur_time[0: (len_cur_time - 10)]
        date_time_at = str_date_today + " " + '00:00'
        date_time_to = str_date_today + " " + str_cur_time
        dateCur = str_date_today
        datetimeTo = QtCore.QDateTime.fromString(date_time_to, 'd.M.yyyy hh:mm')
        datetimeAt = QtCore.QDateTime.fromString(date_time_at, 'd.M.yyyy hh:mm')
        date = QtCore.QDateTime.fromString(dateCur, 'd.M.yyyy')
        self.DTEditRngTo.setDateTime(datetimeTo)
        self.DTEditRngAt.setDateTime(datetimeAt)
        self.DTEitFix.setDateTime(date)
        self.btnCloseMainForm.clicked.connect(self.closeForm)
        self.btnSrchRefresh.clicked.connect(self.updateTable)
        self.btnFilterRefresh.clicked.connect(self.updateTable)
        self.comboFilter.currentIndexChanged.connect(self.changeStCb)
        self.btnSearch.clicked.connect(self.searchEls)
        self.btnFilter.clicked.connect(self.searchFilterEls)
        self.tblTS.itemClicked.connect(self.pasteInfoTS)

        self.btnPlayKppIn.clicked.connect(self.playVideoKppIn)
        # self.btnPauseGRZTS.clicked.connect(self.pauseVideoKppIn)
        self.btnStopKppIn.clicked.connect(self.stopVideoKppIn)

        self.btnPlayKppOut.clicked.connect(self.playVideoKppOut)
        # self.btnPauseGRUNT.clicked.connect(self.pauseVideoKppOut)
        self.btnStopKppOut.clicked.connect(self.stopVideoKppOut)

        self.btnPlayWeightIn.clicked.connect(self.playVideoWeightIn)
        # self.btnPauseGRUNT.clicked.connect(self.pauseVideoKppOut)
        self.btnStopWeightIn.clicked.connect(self.stopVideoWeightIn)

        self.btnPlayWeightOut.clicked.connect(self.playVideoWeightOut)
        # self.btnPauseGRUNT.clicked.connect(self.pauseVideoKppOut)
        self.btnStopWeightOut.clicked.connect(self.stopVideoWeightOut)

        # self.btnPlayWghtEmpty.clicked.connect(self.playVideoWghtEmpty)
        # self.btnPauseWghtEmpty.clicked.connect(self.pauseVideoWghtEmpty)
        # self.btnStopWghtEmpte.clicked.connect(self.stopVideoWghtEmpty)
        #
        # self.btnPlauWghtFull.clicked.connect(self.playVideoWghtLoad)
        # self.btnPauseWghtFull.clicked.connect(self.pauseVideoWghtLoad)
        # self.btnStopWghtFull.clicked.connect(self.stopVideoWghtLoad)

        if (globalValues.find_zakaz == False):

            self.connectToMySql()
            self.updateTable()
            self.changeStCb()

        self.tblTS.setColumnWidth(14, 0)
        self.tblTS.setColumnWidth(15, 0)
        self.tblTS.setColumnWidth(16, 0)



        # self.setMediaFrm('E:/project.mp4', self.mediaPlayerWeightEmpty)
        # self.setMediaFrm('E:/project.mp4', self.mediaGrabberWghtEmpty)
        # self.setMediaFrm('E:/project.mp4', self.mediaPlayerWeightLoad)
        # self.setMediaFrm('E:/project.mp4', self.mediaGrabberWghtLoad)

        #KppIn

        #KppIn
        self.SldrKppIn.sliderMoved.connect(self.setPositionKppIn)
        self.SldrKppIn.sliderPressed.connect(self.pressSldrKppIn)
        self.mediaPlayerKppIn.positionChanged.connect(self.positionChangedKppIn)
        self.mediaPlayerKppIn.durationChanged.connect( self.durationChangedKppIn)
        self.btnStepbackKppIn.clicked.connect(self.jumpBackKppIn)
        self.btnStepforwardKppIn.clicked.connect(self.jumpForwardKppIn)
        self.cbKppIn.currentIndexChanged.connect(self.playBackRateKppIn)
        self.mediaPlayerKppIn.error.connect(self.handleError)

        #KppOut
        self.SldrKppOut.sliderMoved.connect(self.setPositionKppOut)
        self.SldrKppOut.sliderPressed.connect(self.pressSldrKppOut)
        self.mediaPlayerKppOut.positionChanged.connect(self.positionChangedKppOut)
        self.mediaPlayerKppOut.durationChanged.connect(self.durationChangedKppOut)
        self.btnStepbackKppOut.clicked.connect(self.jumpBackKppOut)
        self.btnStepforwardKppOut.clicked.connect(self.jumpForwardKppOut)
        self.cbKppOut.currentIndexChanged.connect(self.playBackRateKppOut)
        self.mediaPlayerKppOut.error.connect(self.handleError)

        #WeightIn
        self.SldrWeightIn.sliderMoved.connect(self.setPositionWeightIn)
        self.SldrWeightIn.sliderPressed.connect(self.pressSldrWeightIn)
        self.mediaPlayerWeightIn.positionChanged.connect(self.positionChangedWeightIn)
        self.mediaPlayerWeightIn.durationChanged.connect(self.durationChangedWeightIn)
        self.btnStepbackWeightIn.clicked.connect(self.jumpBackWeightIn)
        self.btnStepforwardWeightIn.clicked.connect(self.jumpForwardWeightIn)
        self.cbWeightIn.currentIndexChanged.connect(self.playBackRateWeightIn)
        self.mediaPlayerWeightIn.error.connect(self.handleError)

        #WeightOut
        self.SldrWeightOut.sliderMoved.connect(self.setPositionWeightOut)
        self.SldrWeightOut.sliderPressed.connect(self.pressSldrWeightOut)
        self.mediaPlayerWeightOut.positionChanged.connect(self.positionChangedWeightOut)
        self.mediaPlayerWeightOut.durationChanged.connect(self.durationChangedWeightOut)
        self.btnStepbackWeightOut.clicked.connect(self.jumpBackWeightOut)
        self.btnStepforwardWeightOut.clicked.connect(self.jumpForwardWeightOut)
        self.cbWeightOut.currentIndexChanged.connect(self.playBackRateWeightOut)
        self.mediaPlayerWeightOut.error.connect(self.handleError)

        # self.SldrWghtEmpty.sliderMoved.connect(self.setPositionWghtEmpty)
        # self.mediaPlayerWeightEmpty.positionChanged.connect(self.positionChangedWghtEmpty)
        # self.mediaPlayerWeightEmpty.durationChanged.connect(self.durationChangedWghtEmpty)
        # self.mediaPlayerWeightEmpty.error.connect(self.handleError)
        #
        # self.SldrWghtFull.sliderMoved.connect(self.setPositionWghtLoad)
        # self.mediaPlayerWeightLoad.positionChanged.connect(self.positionChangedWghtLoad)
        # self.mediaPlayerWeightLoad.durationChanged.connect(self.durationChangedWghtLoad)
        # self.mediaPlayerWeightLoad.error.connect(self.handleError)

        # self.btnMakeFrameGRZTS.clicked.connect(self.showImgFrmKppIn)
        # self.btnMakeFrameGRUNT.clicked.connect(self.showImgFrmKppOut)
        # self.btnMakeFrameWghtEmpty.clicked.connect(self.showImgFrmWghtEmpty)
        # self.btnMakeFrameWghtFull.clicked.connect(self.showImgFrmWghtLoad)

        self.tblTS.verticalHeader().hide()
        self.tblTS.horizontalScrollBar().hide()
        self.tblTS.verticalScrollBar().hide()
        self.verticalScrollBar.valueChanged.connect(self.sync_func)
        self.btnFilterSave.clicked.connect(self.saveReport)
        self.btnFilterPrint.clicked.connect(self.printData)
        self.btnFilterPrint.setEnabled(False)
        self.btnFilterSave.setEnabled(False)

        self.tblTS.setSelectionBehavior(QtWidgets.QTableView.SelectRows)



        # checkMp4 = 'mp4' in globalValues.pathFileVideoKppIn
        # if (os.path.exists(globalValues.pathFileVideoKppIn) and checkMp4):
        # self.setMediaFrm('', self.mediaPlayerKppIn)
        #     # self.setMediaFrm(globalValues.pathFileVideoKppIn, self.mediaGrabberKppIn)
        #
        # checkMp4 = 'mp4' in globalValues.pathFileVideoKppOut
        # if (os.path.exists(globalValues.pathFileVideoKppOut) and checkMp4):
        # self.setMediaFrm('', self.mediaPlayerKppOut)
        #
        # checkMp4 = 'mp4' in globalValues.pathFileVideoWeightIn
        # if (os.path.exists(globalValues.pathFileVideoWeightIn) and checkMp4):
        # self.setMediaFrm('', self.mediaPlayerWeightIn)
        #
        # checkMp4 = 'mp4' in globalValues.pathFileVideoWeightOut
        # if (os.path.exists(globalValues.pathFileVideoWeightOut) and checkMp4):
        # self.setMediaFrm('', self.mediaPlayerWeightOut)
        #     # self.setMediaFrm(globalValues.pathFileVideoKppOut, self.mediaGrabberKppOut)

    def sync_func(self):
        self.tblTS.verticalScrollBar().setValue(self.verticalScrollBar.value())

    def changeColor(self, object, str):
        object.setStyleSheet(str)

    def changeCOLORMainPanelGrunt(self):

            delta = int((0.65/self.lengthDark)*1000)

            if (globalValues.colorForm == 1):

                    for i in range(self.lengthLight):
                        startTime = round(time.time() * 1000)
                        while True:
                                    obj = self.lstLight[i][0]
                                    style = self.lstLight[i][1]
                                    self.changeColor(obj, style)
                                    # print('changeLight!')
                                    if (abs(round(time.time()*1000) - startTime) > delta):
                                            break

            elif (globalValues.colorForm == 0):

                    for i in range(self.lengthDark):
                        startTime = round(time.time() * 1000)
                        while True:
                                    # print(delta)
                                    # print(abs(round(time.time()*1000) - startTime))
                                    obj = self.lstDark[i][0]
                                    style = self.lstDark[i][1]
                                    self.changeColor(obj, style)
                                    # print('changeDark!')
                                    if (abs(round(time.time()*1000) - startTime) > delta):
                                            break

    def thChangeScroll(self, check):
        try:

            num_delta = 200
            if (self.firstCallChangeScroll):
                self.firstCallChangeScroll = False
                num_delta = 350

            # print('Delta: ' + str(num_delta))

            start_time = round(time.time() * 100)
            while True:
                numRows = self.tblTS.verticalScrollBar().maximum()
                delta = round(abs(time.time() * 100) - start_time)

                # print(num_delta)

                if numRows != 0:
                    print('changeBarScroll!!!' + str(numRows))
                    time.sleep(0.2)
                    self.verticalScrollBar.setMaximum(self.tblTS.verticalScrollBar().maximum())
                    break
                # print('checkingScroll!!!: ' + str(numRows))

                if delta > num_delta:
                    if (check):
                        # print('changeScrollTimer')
                        self.verticalScrollBar.setMaximum(self.tblTS.verticalScrollBar().maximum())
                    break

                time.sleep(0.1)

                if (globalValues.stopAll):
                    break
        except Exception as ex:
            globalValues.writeLogData('Поток обработки изменения скролла', str(ex))

    def setMediaFrm(self, pathFile, obMedia):
        # fileName, _ = QFileDialog.getOpenFileName(self, "Open Movie", QDir.homePath())

        if (pathFile != ''):
            obMedia.setMedia(QMediaContent(QUrl.fromLocalFile(pathFile)))
            # self.btnPlayKPPOUT.setEnabled(True)
            print('checkingpath: ' + pathFile)

    def process_frame_kppin(self, image):
        self.curImageKppIn = image

    def showImgFrmKppIn(self):
        curImage = self.curImageKppIn.scaled(470, 470, QtCore.Qt.KeepAspectRatio)
        # self.curImage.save('E:/2.jpg')
        self.lbl_frameGRZTS.setPixmap(QPixmap.fromImage(curImage))

    def process_frame_kppout(self, image):
        self.curImageKppOut = image

    def showImgFrmKppOut(self):
        curImage = self.curImageKppOut.scaled(470, 470, QtCore.Qt.KeepAspectRatio)
        self.lbl_frameGRUNT.setPixmap(QPixmap.fromImage(curImage))

    # def process_frame_wght_empty(self, image):
    #     self.curImageWghtEmpty = image
    #
    # def showImgFrmWghtEmpty(self):
    #     curImage = self.curImageWghtEmpty.scaled(470, 470, QtCore.Qt.KeepAspectRatio)
    #     # self.curImage.save('E:/2.jpg')
    #     self.le_frmWghtEmpty.setPixmap(QPixmap.fromImage(curImage))
    #
    # def process_frame_wght_load(self, image):
    #     self.curImageWghtLoad = image
    #
    # def showImgFrmWghtLoad(self):
    #     curImage = self.curImageWghtLoad.scaled(470, 470, QtCore.Qt.KeepAspectRatio)
    #     # self.curImage.save('E:/2.jpg')
    #     self.le_frmWghtFull.setPixmap(QPixmap.fromImage(curImage))


 #Window0

    #KppIn

    def playVideoKppIn(self):
        # print('StartingPlay!!!')

        checkMp4 = 'mp4' in globalValues.pathFileVideoKppIn
        if (os.path.exists(globalValues.pathFileVideoKppIn) and checkMp4):

            try:
                self.playBackRateKppIn()
                if self.mediaPlayerKppIn.state() != QMediaPlayer.PlayingState:
                    self.mediaPlayerKppIn.play()
                    # self.mediaGrabberKppIn.play()

                    if (globalValues.colorForm == 1):
                        self.btnPlayKppIn.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.PNG);}\n"
                                                        "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.PNG);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.PNG);}")
                    else:
                        self.btnPlayKppIn.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.PNG);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.PNG);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.PNG);}")
                else:
                    self.pauseVideoKppIn()

                    if (globalValues.colorForm == 1):
                        self.btnPlayKppIn.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}")
                    else:
                        self.btnPlayKppIn.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}")
            except Exception as ex:
                globalValues.writeLogData('Функция воспроизведения видеофайла', str(ex))

    def jumpBackKppIn(self):

        try:
            delta = int (self.SldrKppIn.maximum()/10)
            pos = 0
            if (self.SldrKppIn.sliderPosition() - delta <= 0):
                pos = 0
            else:
                pos = int(self.SldrKppIn.sliderPosition() - delta)
            self.SldrKppIn.setValue(pos)
            self.mediaPlayerKppIn.setPosition(pos)
            # self.mediaGrabberKppIn.setPosition(pos)

        except Exception as ex:
            globalValues.writeLogData('Функция промотки назад видеофайла', str(ex))

    def jumpForwardKppIn(self):
        try:
            maxValSldr  = self.SldrKppIn.maximum()
            delta = int( maxValSldr / 10)
            pos = 0
            if (self.SldrKppIn.sliderPosition() + delta >= maxValSldr):
                pos = maxValSldr
            else:
                pos = int(self.SldrKppIn.sliderPosition() + delta)
            self.SldrKppIn.setValue(pos)
            self.mediaPlayerKppIn.setPosition(pos)
            # self.mediaGrabberKppIn.setPosition(pos)

        except Exception as ex:
            globalValues.writeLogData('Функция промотки вперёд видеофайла', str(ex))

    def playBackRateKppIn(self):
        try:
            dataRate = self.cbKppIn.currentText()
            dataRate = dataRate.replace('x', '')
            dataRate = float(dataRate)
            # print(dataRate)
            self.mediaPlayerKppIn.setPlaybackRate(dataRate)
            # self.mediaGrabberKppIn.setPlaybackRate(dataRate)
        except Exception as ex:
            globalValues.writeLogData('Функция смены скорости воспроизведения видеофайла', str(ex))

    def pauseVideoKppIn(self):
        if self.mediaPlayerKppIn.state() != QMediaPlayer.PausedState:
            self.mediaPlayerKppIn.pause()
            # self.mediaGrabberKppIn.pause()

    def stopVideoKppIn(self):
        try:
            if self.mediaPlayerKppIn.state() != QMediaPlayer.StoppedState:
                self.mediaPlayerKppIn.stop()
                # self.mediaGrabberKppIn.stop()
                if (globalValues.colorForm == 1):
                    self.btnPlayKppIn.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}")
                else:
                    self.btnPlayKppIn.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}")
        except Exception as ex:
            globalValues.writeLogData('Функция остановки видеофайла', str(ex))

    def positionChangedKppIn(self, position):
        try:
            self.SldrKppIn.setValue(position)

            if (position >= self.SldrKppIn.maximum()):
                self.mediaPlayerKppIn.setPosition(0)
                # self.mediaGrabberKppIn.setPosition(0)
                self.mediaPlayerKppIn.stop()
                # self.mediaGrabberKppIn.stop()

                if (globalValues.colorForm == 1):
                    self.btnPlayKppIn.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}")
                else:
                    self.btnPlayKppIn.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}")

        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def durationChangedKppIn(self, duration):
        try:
            self.SldrKppIn.setRange(0, duration)
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def setPositionKppIn(self, position):
        # print('Position: ' + str(position))
        try:
            self.mediaPlayerKppIn.setPosition(position)
            # self.mediaGrabberKppIn.setPosition(position)
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def pressSldrKppIn(self):
        try:
            # print('Clicked!')
            self.setPositionKppIn(self.SldrKppIn.sliderPosition())
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

# KppOut

    def playVideoKppOut(self):

        checkMp4 = 'mp4' in globalValues.pathFileVideoKppOut
        if (os.path.exists(globalValues.pathFileVideoKppOut) and checkMp4):

            try:
                self.playBackRateKppOut()
                if self.mediaPlayerKppOut.state() != QMediaPlayer.PlayingState:
                    self.mediaPlayerKppOut.play()
                    # self.mediaGrabberKppOut.play()

                    if (globalValues.colorForm == 1):
                        self.btnPlayKppOut.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.PNG);}\n"
                                                        "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.PNG);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.PNG);}")
                    else:
                        self.btnPlayKppOut.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.PNG);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.PNG);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.PNG);}")
                else:
                    self.pauseVideoKppOut()

                    if (globalValues.colorForm == 1):
                        self.btnPlayKppOut.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}")
                    else:
                        self.btnPlayKppOut.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}")
            except Exception as ex:
                globalValues.writeLogData('Функция воспроизведения видеофайла', str(ex))

    def pauseVideoKppOut(self):
        if self.mediaPlayerKppOut.state() != QMediaPlayer.PausedState:
            self.mediaPlayerKppOut.pause()
            # self.mediaGrabberKppOut.pause()

    def stopVideoKppOut(self):
        try:
            if self.mediaPlayerKppOut.state() != QMediaPlayer.StoppedState:
                self.mediaPlayerKppOut.stop()
                # self.mediaGrabberKppOut.stop()

                if (globalValues.colorForm == 1):
                    self.btnPlayKppOut.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}")
                else:
                    self.btnPlayKppOut.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}")

        except Exception as ex:
            globalValues.writeLogData('Функция остановки воспроизведения видеофайла', str(ex))

    def positionChangedKppOut(self, position):
        try:
            self.SldrKppOut.setValue(position)

            if (position >= self.SldrKppOut.maximum()):
                self.mediaPlayerKppOut.setPosition(0)
                # self.mediaGrabberKppOut.setPosition(0)
                self.mediaPlayerKppOut.stop()
                # self.mediaGrabberKppOut.stop()

                if (globalValues.colorForm == 1):
                    self.btnPlayKppOut.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}")
                else:
                    self.btnPlayKppOut.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}")

        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def durationChangedKppOut(self, duration):
        try:
            self.SldrKppOut.setRange(0, duration)
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def setPositionKppOut(self, position):
        try:
            self.mediaPlayerKppOut.setPosition(position)
            # self.mediaGrabberKppOut.setPosition(position)
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def pressSldrKppOut(self):
        try:
            # print('Ckick2')
            self.setPositionKppOut(self.SldrKppOut.sliderPosition())
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def jumpBackKppOut(self):

        try:
            delta = int (self.SldrKppOut.maximum()/10)
            pos = 0
            if (self.SldrKppOut.sliderPosition() - delta <= 0):
                pos = 0
            else:
                pos = int(self.SldrKppOut.sliderPosition() - delta)
            self.SldrKppOut.setValue(pos)
            self.mediaPlayerKppOut.setPosition(pos)
            # self.mediaGrabberKppOut.setPosition(pos)

        except Exception as ex:
            globalValues.writeLogData('Функция промотки назад видеофайла', str(ex))

    def jumpForwardKppOut(self):
        try:
            maxValSldr  = self.SldrKppOut.maximum()
            delta = int( maxValSldr / 10)
            pos = 0
            if (self.SldrKppOut.sliderPosition() + delta >= maxValSldr):
                pos = maxValSldr
            else:
                pos = int(self.SldrKppOut.sliderPosition() + delta)
            self.SldrKppOut.setValue(pos)
            self.mediaPlayerKppOut.setPosition(pos)
            # self.mediaGrabberKppOut.setPosition(pos)

        except Exception as ex:
            globalValues.writeLogData('Функция промотки вперёд видеофайла', str(ex))

    def playBackRateKppOut(self):
        try:
            dataRate = self.cbKppOut.currentText()
            dataRate = dataRate.replace('x', '')
            dataRate = float(dataRate)
            # print(dataRate)
            self.mediaPlayerKppOut.setPlaybackRate(dataRate)
            # self.mediaGrabberKppOut.setPlaybackRate(dataRate)
        except Exception as ex:
            globalValues.writeLogData('Функция смены скорости воспроизведения видеофайла', str(ex))

#WeightIn

    def playVideoWeightIn(self):

        checkMp4 = 'mp4' in globalValues.pathFileVideoWeightIn
        if (os.path.exists(globalValues.pathFileVideoWeightIn) and checkMp4):

            try:
                self.playBackRateWeightIn()
                if self.mediaPlayerWeightIn.state() != QMediaPlayer.PlayingState:
                    self.mediaPlayerWeightIn.play()
                    # self.mediaGrabberKppOut.play()

                    if (globalValues.colorForm == 1):
                        self.btnPlayWeightIn.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.PNG);}\n"
                                                        "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.PNG);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.PNG);}")
                    else:
                        self.btnPlayWeightIn.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.PNG);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.PNG);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.PNG);}")
                else:
                    self.pauseVideoWeightIn()

                    if (globalValues.colorForm == 1):
                        self.btnPlayWeightIn.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}")
                    else:
                        self.btnPlayWeightIn.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}")
            except Exception as ex:
                globalValues.writeLogData('Функция воспроизведения видеофайла', str(ex))

    def pauseVideoWeightIn(self):
        if self.mediaPlayerWeightIn.state() != QMediaPlayer.PausedState:
            self.mediaPlayerWeightIn.pause()
            # self.mediaGrabberKppOut.pause()

    def stopVideoWeightIn(self):
        try:
            if self.mediaPlayerWeightIn.state() != QMediaPlayer.StoppedState:
                self.mediaPlayerWeightIn.stop()
                # self.mediaGrabberKppOut.stop()

                if (globalValues.colorForm == 1):
                    self.btnPlayWeightIn.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}")
                else:
                    self.btnPlayWeightIn.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}")

        except Exception as ex:
            globalValues.writeLogData('Функция остановки воспроизведения видеофайла', str(ex))

    def positionChangedWeightIn(self, position):
        try:
            self.SldrWeightIn.setValue(position)

            if (position >= self.SldrWeightIn.maximum()):
                self.mediaPlayerWeightIn.setPosition(0)
                # self.mediaGrabberKppOut.setPosition(0)
                self.mediaPlayerWeightIn.stop()
                # self.mediaGrabberKppOut.stop()

                if (globalValues.colorForm == 1):
                    self.btnPlayWeightIn.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}")
                else:
                    self.btnPlayWeightIn.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}")

        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def durationChangedWeightIn(self, duration):
        try:
            self.SldrWeightIn.setRange(0, duration)
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def setPositionWeightIn(self, position):
        try:
            self.mediaPlayerWeightIn.setPosition(position)
            # self.mediaGrabberKppOut.setPosition(position)
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def pressSldrWeightIn(self):
        try:
            self.setPositionWeightIn(self.SldrWeightIn.sliderPosition())
            print('checking!!!')
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def jumpBackWeightIn(self):

        try:
            delta = int (self.SldrWeightIn.maximum()/10)
            pos = 0
            if (self.SldrWeightIn.sliderPosition() - delta <= 0):
                pos = 0
            else:
                pos = int(self.SldrWeightIn.sliderPosition() - delta)
            self.SldrWeightIn.setValue(pos)
            self.mediaPlayerWeightIn.setPosition(pos)
            # self.mediaGrabberKppOut.setPosition(pos)

        except Exception as ex:
            globalValues.writeLogData('Функция промотки назад видеофайла', str(ex))

    def jumpForwardWeightIn(self):
        try:
            maxValSldr  = self.SldrWeightIn.maximum()
            delta = int( maxValSldr / 10)
            pos = 0
            if (self.SldrWeightIn.sliderPosition() + delta >= maxValSldr):
                pos = maxValSldr
            else:
                pos = int(self.SldrWeightIn.sliderPosition() + delta)
            self.SldrWeightIn.setValue(pos)
            self.mediaPlayerWeightIn.setPosition(pos)
            # self.mediaGrabberKppOut.setPosition(pos)

        except Exception as ex:
            globalValues.writeLogData('Функция промотки вперёд видеофайла', str(ex))

    def playBackRateWeightIn(self):
        try:
            dataRate = self.cbWeightIn.currentText()
            dataRate = dataRate.replace('x', '')
            dataRate = float(dataRate)
            # print(dataRate)
            self.mediaPlayerWeightIn.setPlaybackRate(dataRate)
            # self.mediaGrabberKppOut.setPlaybackRate(dataRate)
        except Exception as ex:
            globalValues.writeLogData('Функция смены скорости воспроизведения видеофайла', str(ex))

#WeightOut

    def playVideoWeightOut(self):

        checkMp4 = 'mp4' in globalValues.pathFileVideoWeightOut
        if (os.path.exists(globalValues.pathFileVideoWeightOut) and checkMp4):

            try:
                self.playBackRateWeightOut()
                if self.mediaPlayerWeightOut.state() != QMediaPlayer.PlayingState:
                    self.mediaPlayerWeightOut.play()
                    # self.mediaGrabberKppOut.play()

                    if (globalValues.colorForm == 1):
                        self.btnPlayWeightOut.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.PNG);}\n"
                                                        "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.PNG);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.PNG);}")
                    else:
                        self.btnPlayWeightOut.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.PNG);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.PNG);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.PNG);}")
                else:
                    self.pauseVideoWeightOut()

                    if (globalValues.colorForm == 1):
                        self.btnPlayWeightOut.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-left:1px solid rgb(150,150,150);\n"
                                                        "border-right:1px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}")
                    else:
                        self.btnPlayWeightOut.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}")
            except Exception as ex:
                globalValues.writeLogData('Функция воспроизведения видеофайла', str(ex))

    def pauseVideoWeightOut(self):
        if self.mediaPlayerWeightOut.state() != QMediaPlayer.PausedState:
            self.mediaPlayerWeightOut.pause()
            # self.mediaGrabberKppOut.pause()

    def stopVideoWeightOut(self):
        try:
            if self.mediaPlayerWeightOut.state() != QMediaPlayer.StoppedState:
                self.mediaPlayerWeightOut.stop()
                # self.mediaGrabberKppOut.stop()

                if (globalValues.colorForm == 1):
                    self.btnPlayWeightOut.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}")
                else:
                    self.btnPlayWeightOut.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}")

        except Exception as ex:
            globalValues.writeLogData('Функция остановки воспроизведения видеофайла', str(ex))

    def positionChangedWeightOut(self, position):
        try:
            self.SldrWeightOut.setValue(position)

            if (position >= self.SldrWeightOut.maximum()):
                self.mediaPlayerWeightOut.setPosition(0)
                # self.mediaGrabberKppOut.setPosition(0)
                self.mediaPlayerWeightOut.stop()
                # self.mediaGrabberKppOut.stop()

                if (globalValues.colorForm == 1):
                    self.btnPlayWeightOut.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-left:1px solid rgb(150,150,150);\n"
                                                    "border-right:1px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}")
                else:
                    self.btnPlayWeightOut.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}")

        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def durationChangedWeightOut(self, duration):
        try:
            self.SldrWeightOut.setRange(0, duration)
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def setPositionWeightOut(self, position):
        try:
            self.mediaPlayerWeightOut.setPosition(position)
            # self.mediaGrabberKppOut.setPosition(position)
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def pressSldrWeightOut(self):
        try:
            self.setPositionWeightOut(self.SldrWeightOut.sliderPosition())
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def jumpBackWeightOut(self):

        try:
            delta = int (self.SldrWeightOut.maximum()/10)
            pos = 0
            if (self.SldrWeightOut.sliderPosition() - delta <= 0):
                pos = 0
            else:
                pos = int(self.SldrWeightOut.sliderPosition() - delta)
            self.SldrWeightOut.setValue(pos)
            self.mediaPlayerWeightOut.setPosition(pos)
            # self.mediaGrabberKppOut.setPosition(pos)

        except Exception as ex:
            globalValues.writeLogData('Функция промотки назад видеофайла', str(ex))

    def jumpForwardWeightOut(self):
        try:
            maxValSldr  = self.SldrWeightOut.maximum()
            delta = int( maxValSldr / 10)
            pos = 0
            if (self.SldrWeightOut.sliderPosition() + delta >= maxValSldr):
                pos = maxValSldr
            else:
                pos = int(self.SldrWeightOut.sliderPosition() + delta)
            self.SldrWeightOut.setValue(pos)
            self.mediaPlayerWeightOut.setPosition(pos)
            # self.mediaGrabberKppOut.setPosition(pos)

        except Exception as ex:
            globalValues.writeLogData('Функция промотки вперёд видеофайла', str(ex))

    def playBackRateWeightOut(self):
        try:
            dataRate = self.cbWeightOut.currentText()
            dataRate = dataRate.replace('x', '')
            dataRate = float(dataRate)
            # print(dataRate)
            self.mediaPlayerWeightOut.setPlaybackRate(dataRate)
            # self.mediaGrabberKppOut.setPlaybackRate(dataRate)
        except Exception as ex:
            globalValues.writeLogData('Функция смены скорости воспроизведения видеофайла', str(ex))


#
# # Window2
#
#     def playVideoWghtEmpty(self):
#         print('StartingPlay!!!')
#         if self.mediaPlayerWeightEmpty.state() != QMediaPlayer.PlayingState:
#             self.mediaPlayerWeightEmpty.play()
#             self.mediaGrabberWghtEmpty.play()
#
#     def pauseVideoWghtEmpty(self):
#         if self.mediaPlayerWeightEmpty.state() != QMediaPlayer.PausedState:
#             self.mediaPlayerWeightEmpty.pause()
#             self.mediaGrabberWghtEmpty.pause()
#
#     def stopVideoWghtEmpty(self):
#         if self.mediaPlayerWeightEmpty.state() != QMediaPlayer.StoppedState:
#             self.mediaPlayerWeightEmpty.stop()
#             self.mediaGrabberWghtEmpty.stop()
#
#     def positionChangedWghtEmpty(self, position):
#         self.SldrWghtEmpty.setValue(position)
#
#     def durationChangedWghtEmpty(self, duration):
#         self.SldrWghtEmpty.setRange(0, duration)
#
#     def setPositionWghtEmpty(self, position):
#         print(position)
#         self.mediaPlayerWeightEmpty.setPosition(position)
#         self.mediaGrabberWghtEmpty.setPosition(position)
#
# # Window3
#
#     def playVideoWghtLoad(self):
#         print('StartingPlay!!!')
#         if self.mediaPlayerWeightLoad.state() != QMediaPlayer.PlayingState:
#             self.mediaPlayerWeightLoad.play()
#             self.mediaGrabberWghtLoad.play()
#
#     def pauseVideoWghtLoad(self):
#         if self.mediaPlayerWeightLoad.state() != QMediaPlayer.PausedState:
#             self.mediaPlayerWeightLoad.pause()
#             self.mediaGrabberWghtLoad.pause()
#
#     def stopVideoWghtLoad(self):
#         if self.mediaPlayerWeightLoad.state() != QMediaPlayer.StoppedState:
#             self.mediaPlayerWeightLoad.stop()
#             self.mediaGrabberWghtLoad.stop()
#
#     def positionChangedWghtLoad(self, position):
#         self.SldrWghtFull.setValue(position)
#
#     def durationChangedWghtLoad(self, duration):
#         self.SldrWghtFull.setRange(0, duration)
#
#     def setPositionWghtLoad(self, position):
#         print(position)
#         self.mediaPlayerWeightLoad.setPosition(position)
#         self.mediaGrabberWghtLoad.setPosition(position)

    def handleError(self):
        try:
            globalValues.writeLogData("Ошибка воспроизведения медиаплеера ", self.mediaPlayerKppIn.errorString())
        except Exception as ex:
            globalValues.writeLogData('Функция обработки ошибки воспроизведения видеофайла', str(ex))

    def updateTable(self):

        try:

            if globalValues.debug:
                self.con = pymysql.connect(host='localhost',
                                      port=3306,
                                      user='sergey',
                                      passwd='34ubitav',
                                      db=globalValues.dbMySqlName)
            else:
                if (self.checkMySql):
                    self.connectToMySql()

            cur = self.con.cursor()

            try:

                countRowsOld = 0

                start_time = round(time.time())

                with self.con:
                # if True:

                    sqlID = 'SELECT id FROM ' +globalValues.tblsDB[1]+ ' order by id desc limit 1'

                    cur.execute(sqlID)

                    rows_id = cur.fetchall()

                    num_end_order = rows_id[0][0]

                    sqlCom = ''

                    rowsMain = []

                    rowsOrder = []

                    rowsPolygon = []

                    if (num_end_order > globalValues.valNumShowOeder):

                        print(num_end_order)
                        num_start_order = str(num_end_order - globalValues.valNumShowOeder)

                        sqlCom = 'SELECT * FROM ' + globalValues.tblsDB[1] + ' LIMIT ' + num_start_order + ', ' + str(num_end_order) + ''

                        cur.execute(sqlCom)

                        rowsMain = cur.fetchall()

                        sqlCom = 'SELECT * FROM ' + globalValues.tblsDB[2] + ' LIMIT ' + num_start_order + ', ' + str(num_end_order) + ''

                        cur.execute(sqlCom)

                        rowsOrder = cur.fetchall()

                        sqlCom = 'SELECT * FROM ' + globalValues.tblsDB[5] + ' LIMIT ' + num_start_order + ', ' + str(num_end_order) + ''

                        cur.execute(sqlCom)

                        rowsPolygon = cur.fetchall()

                        # rowsOrder = reversed(rowsOrder)

                        print('LengthDataFromBD: ', len(rowsMain), ', ', len(rowsOrder), ', ', len(rowsPolygon))

                    else:

                        cur.execute("SELECT * FROM " + globalValues.tblsDB[1])

                        rowsMain = cur.fetchall()

                        cur.execute("SELECT * FROM " + globalValues.tblsDB[2])

                        rowsOrder = cur.fetchall()

                        # rowsOrder = reversed(rowsOrder)

                        cur.execute("SELECT * FROM " + globalValues.tblsDB[5])

                        rowsPolygon = cur.fetchall()


                    countRows = 0

                    for row in rowsMain:
                        countRows += 1

                    if (countRows != 0 and countRows != countRowsOld):
                        self.tblTS.setRowCount(0)

                        valWdg = 0


                        # print('CheckingData!')
                        # print(str(rowsOrder[0][4]))
                        # print(str(rowsOrder[1][4]))
                        # print(str(rowsOrder[2][4]))
                        # print(str(rowsOrder[3][4]))
                        # print(str(rowsOrder[0][5]))
                        # print(str(rowsOrder[1][5]))
                        # print(str(rowsOrder[2][5]))
                        # print(str(rowsOrder[3][5]))

                        for rowMain in rowsMain:

                            try:

                                print('CheckWDG!!')
                                print(valWdg)

                                valColor = globalValues.colorForm

                                self.tblTS.insertRow(0)
                                item = QtWidgets.QTableWidgetItem(str(rowMain[2]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 0, item)

                                item = QtWidgets.QTableWidgetItem(str(rowMain[1]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 1, item)

                                item = QtWidgets.QTableWidgetItem(str(rowsOrder[valWdg][4]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 2, item)

                                item = QtWidgets.QTableWidgetItem(str(rowsOrder[valWdg][5]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 3, item)

                                item = QtWidgets.QTableWidgetItem(str(rowsOrder[valWdg][3]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 4, item)


                                if globalValues.polygon:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[10]))
                                else:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[12]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 5, item)

                                if globalValues.polygon:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[11]))
                                else:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[13]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 6, item)

                                if globalValues.polygon:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[5]))
                                else:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[14]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 7, item)

                                if globalValues.polygon:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[6]))
                                else:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[15]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 8, item)

                                if globalValues.polygon:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[12]))
                                else:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[10]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 9, item)

                                if globalValues.polygon:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[12]))
                                else:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[10]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 10, item)

                                if globalValues.polygon:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[14]))
                                else:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[5]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 11, item)

                                if globalValues.polygon:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[15]))
                                else:
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[6]))
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 12, item)

                                # item = QtWidgets.QTableWidgetItem(str(rowsPolygon[valWdg][2]))
                                # item.setTextAlignment(Qt.AlignCenter)
                                # if (globalValues.colorForm == 1):
                                #     item.setForeground(QtGui.QBrush(Qt.black))
                                # else:
                                #     item.setForeground(QtGui.QBrush(Qt.white))
                                # item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                # self.tblTS.setItem(0, 13, item)


                                strStOrder = str(rowMain[7])

                                item = QtWidgets.QTableWidgetItem(strStOrder)
                                item.setTextAlignment(Qt.AlignCenter)
                                if (globalValues.colorForm == 1):
                                    item.setForeground(QtGui.QBrush(Qt.black))
                                else:
                                    item.setForeground(QtGui.QBrush(Qt.white))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)

                                if (strStOrder == 'не выполняется'):
                                    if (valColor == 0):
                                        item.setBackground(QtGui.QColor(149, 26, 26))
                                        item.setForeground(QtGui.QColor(255, 255, 255))
                                    else:
                                        item.setBackground(QtGui.QColor(197, 104, 104))
                                        item.setForeground(QtGui.QColor(0, 0, 0))
                                elif (strStOrder == 'выполняется'):
                                    if (valColor == 0):
                                        item.setBackground(QtGui.QColor(149, 26, 26))
                                        item.setForeground(QtGui.QColor(255, 255, 255))
                                    else:
                                        item.setBackground(QtGui.QColor(197, 104, 104))
                                        item.setForeground(QtGui.QColor(0, 0, 0))
                                elif (strStOrder == 'перемещение'):
                                    if (valColor == 0):
                                        item.setBackground(QtGui.QColor(6, 81, 16))
                                        item.setForeground(QtGui.QColor(255, 255, 255))
                                    else:
                                        item.setBackground(QtGui.QColor(84, 181, 100))
                                        item.setForeground(QtGui.QColor(0, 0, 0))
                                elif (strStOrder == 'выполнен'):
                                    if (valColor == 0):
                                        item.setBackground(QtGui.QColor(6, 81, 16))
                                        item.setForeground(QtGui.QColor(255, 255, 255))
                                    else:
                                        item.setBackground(QtGui.QColor(84, 181, 100))
                                        item.setForeground(QtGui.QColor(0, 0, 0))

                                self.tblTS.setItem(0, 13, item)

                                item = QtWidgets.QTableWidgetItem(str(rowsMain[valWdg][0]))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 14, item)

                                item = QtWidgets.QTableWidgetItem(str(rowsOrder[valWdg][0]))
                                item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                self.tblTS.setItem(0, 15, item)

                                # item = QtWidgets.QTableWidgetItem(str(rowsPolygon[valWdg][0]))
                                # item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                # self.tblTS.setItem(0, 16, item)

                                valWdg += 1

                            except Exception as ex:
                                self.tblTS.removeRow(0)
                                globalValues.writeLogData('Not data in mysqlBD', str(ex))

                        countRowsOld = countRows

                    th_scroll = threading.Thread(target=self.thChangeScroll, args=(True,))
                    th_scroll.start()

            except Exception as ex:
                globalValues.writeLogData('Поток обновления панели журнала ТС', str(ex))

            cur.close()
            self.con.close()
        except Exception as ex:
            globalValues.writeLogData('Поток обновления панели журнала ТС', str(ex))
            uiMes = Ui_mes_box()
            uiMes.lblStrInfo.setText('Отсутствует соединение с БД!')
            uiMes.btnOK.hide()
            if (globalValues.colorForm == 1):
                uiMes.label_iconques.setStyleSheet("background-color: rgb(235,235,235);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
            else:
                uiMes.label_iconques.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
            uiMes.btnCancel.setText('Продолжить')
            uiMes.exec_()
            return

    def closeForm(self):
        self.close()

    def connectToMySql(self):
        i = 0
        while True:
            try:
                    self.con = pymysql.connect(host=globalValues.my_sql_localhost,
                                                   port=globalValues.my_sql_port,
                                                   user=globalValues.my_sql_name,
                                                   passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)
                    if (self.con.open):
                        break

            except Exception as ex:
                globalValues.writeLogData('Функция подключения к БД панель журнала ТС', str(ex))

            i += 1

            if (i > 2):
                break

    def checkMySql(self):
        try:
                if (self.con.open):
                        return False
                else:
                        return True
        except Exception as ex:
                globalValues.writeLogData('Функция проверки подключения к бд MySql панель журнала ТС', str(ex))

    def changeStCb(self):
        try:
            dataCb = self.comboFilter.currentIndex()
            if (dataCb == 2):
                self.frmSrchDateFix.hide()
                self.frmSrchDateRng.show()
            elif (dataCb == 1):
                self.frmSrchDateRng.hide()
                self.frmSrchDateFix.show()
            elif (dataCb == 0):
                self.frmSrchDateRng.hide()
                self.frmSrchDateFix.hide()

        except Exception as ex:
            globalValues.writeLogData('Функция смены фильра поиска панели журнала', str(ex))

    def searchEls(self):
        try:
            print('StartingSearch!!!')
            dataCb = self.comboSearch.currentIndex()

            if (dataCb == 1):

                # self.updateTable()

                self.dataSearch = self.leSearchTblTS.text()

                try:
                            # countItems = self.tblTS.rowCount()
                            # if (countItems == ''):
                            #     countItems = 0
                            #
                            # countItems = int(countItems)
                            #
                            # for i in range(countItems - 1, -1,  -1):
                            #     dataItem = str(self.tblTS.item(i, 0).text())
                            #     if (dataItem == ''):
                            #         dataItem = ' '
                            #     is_data_in_db = dataSearch in dataItem
                            #     if (is_data_in_db == False):
                            #         self.tblTS.removeRow(i)

                            self.isSearchGRZ = True
                            self.searchDataInBD()


                except Exception as ex:
                    globalValues.writeLogData('Поиск по ГРЗ в таблице журнала ТС', str(ex))

            if (dataCb == 2):

                # self.updateTable()

                self.dataSearch = self.leSearchTblTS.text()

                try:

                    # countItems = self.tblTS.rowCount()
                    #
                    # if (countItems == ''):
                    #     countItems = 0
                    #
                    # countItems = int(countItems)
                    #
                    # for i in range(countItems - 1, -1, -1):
                    #
                    #     dataItem = str(self.tblTS.item(i, 1).text())
                    #
                    #     if (dataItem == ''):
                    #         dataItem = ' '
                    #
                    #     is_data_in_db = dataSearch in dataItem
                    #
                    #     if (is_data_in_db == False):
                    #         self.tblTS.removeRow(i)

                    self.isSearchOrder = True
                    self.searchDataInBD()

                except Exception as ex:

                    globalValues.writeLogData('Поиск по З/Н в таблице журнала ТС', str(ex))

            if (dataCb == 3):

                # self.updateTable()

                self.dataSearch = self.leSearchTblTS.text()

                try:

                            # countItems = self.tblTS.rowCount()
                            # if (countItems == ''):
                            #     countItems = 0
                            #
                            # countItems = int(countItems)
                            #
                            # for i in range(countItems - 1, -1,  -1):
                            #     dataItem = str(self.tblTS.item(i, 2).text())
                            #     if (dataItem == ''):
                            #         dataItem = ' '
                            #     is_data_in_db = dataSearch in dataItem
                            #     if (is_data_in_db == False):
                            #         self.tblTS.removeRow(i)

                        self.isSearchCompany = True
                        self.searchDataInBD()

                except Exception as ex:
                    globalValues.writeLogData('Поиск по организации в таблице журнала ТС', str(ex))

            if (dataCb == 4):

                # self.updateTable()

                self.dataSearch = self.leSearchTblTS.text()

                try:

                            # countItems = self.tblTS.rowCount()
                            # if (countItems == ''):
                            #     countItems = 0
                            #
                            # countItems = int(countItems)
                            #
                            # for i in range(countItems - 1, -1,  -1):
                            #     dataItem = str(self.tblTS.item(i, 4).text())
                            #     if (dataItem == ''):
                            #         dataItem = ' '
                            #     is_data_in_db = dataSearch in dataItem
                            #     if (is_data_in_db == False):
                            #         self.tblTS.removeRow(i)

                        self.isSearchDate = True
                        self.searchDataInBD()

                except Exception as ex:
                    globalValues.writeLogData('Поиск по Дате в таблице журнала ТС', str(ex))

            elif (dataCb == 0):
                self.updateTable()
        except Exception as ex:
            globalValues.writeLogData('Функция поиска панель журнала ТС', str(ex))

    def searchDataInBD(self):
        try:
            for i in range(3):
                try:
                    if (self.checkMySql):
                        self.connectToMySql()

                    cur = self.con.cursor()

                    with self.con:
                    # if True:

                        rowsMain = []

                        rowsOrder = []

                        rowsPolygon = []

                        if (self.isSearchGRZ):

                            self.isSearchGRZ = False

                            sqlQ = "SELECT * FROM " + globalValues.tblsDB[1] + " WHERE number_grz LIKE '%" + self.dataSearch + "%'"

                            cur.execute(sqlQ)

                            rowsMain = cur.fetchall()

                            sqlQ = "SELECT * FROM " + globalValues.tblsDB[2] + " WHERE car_grz LIKE '%" + self.dataSearch + "%'"

                            cur.execute(sqlQ)

                            rowsOrder = cur.fetchall()

                            # rowsOrder = reversed(rowsOrder)

                            # sqlQ = "SELECT * FROM " + globalValues.tblsDB[5] + " WHERE car_grz LIKE '" + self.dataSearch + "%'"
                            #
                            # cur.execute(sqlQ)
                            #
                            # rowsPolygon = cur.fetchall()

                        if (self.isSearchOrder):

                            self.isSearchOrder = False

                            sqlQ = "SELECT * FROM " + globalValues.tblsDB[1] + " WHERE number_order LIKE " + self.dataSearch

                            cur.execute(sqlQ)

                            rowsMain = cur.fetchall()

                            sqlQ = "SELECT * FROM " + globalValues.tblsDB[2] + " WHERE number_order LIKE " + self.dataSearch

                            cur.execute(sqlQ)

                            rowsOrder = cur.fetchall()

                        if (self.isSearchDate):

                            self.isSearchDate = False

                            sqlQ = "SELECT * FROM " + globalValues.tblsDB[1] + " WHERE date LIKE '" + self.dataSearch + "%'"

                            cur.execute(sqlQ)

                            rowsMain = cur.fetchall()

                            sqlQ = "SELECT * FROM " + globalValues.tblsDB[2] + " WHERE date_work LIKE '" + self.dataSearch + "%'"

                            cur.execute(sqlQ)

                            rowsOrder = cur.fetchall()

                        if (self.isSearchCompany):

                            self.isSearchCompany = False

                            sqlQ = "SELECT * FROM " + globalValues.tblsDB[1] + " WHERE name_company LIKE '%" + self.dataSearch + "%'"

                            cur.execute(sqlQ)

                            rowsMain = cur.fetchall()

                            sqlQ = "SELECT * FROM " + globalValues.tblsDB[2] + " WHERE name_company LIKE '%" + self.dataSearch + "%'"

                            cur.execute(sqlQ)

                            rowsOrder = cur.fetchall()

                        countRows = 0

                        for row in rowsMain:
                            countRows += 1

                        if (countRows != 0):
                            self.tblTS.setRowCount(0)

                            valWdg = 0

                            # print('CheckingData!')
                            # print(str(rowsOrder[0][4]))
                            # print(str(rowsOrder[1][4]))
                            # print(str(rowsOrder[2][4]))
                            # print(str(rowsOrder[3][4]))
                            # print(str(rowsOrder[0][5]))
                            # print(str(rowsOrder[1][5]))
                            # print(str(rowsOrder[2][5]))
                            # print(str(rowsOrder[3][5]))



                            for rowMain in rowsMain:

                                try:

                                    print('CheckWDG!!')
                                    print(valWdg)

                                    valColor = globalValues.colorForm

                                    self.tblTS.insertRow(0)
                                    item = QtWidgets.QTableWidgetItem(str(rowMain[2]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 0, item)

                                    item = QtWidgets.QTableWidgetItem(str(rowMain[1]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 1, item)

                                    item = QtWidgets.QTableWidgetItem(str(rowsOrder[valWdg][4]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 2, item)

                                    item = QtWidgets.QTableWidgetItem(str(rowsOrder[valWdg][5]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 3, item)

                                    item = QtWidgets.QTableWidgetItem(str(rowsOrder[valWdg][3]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 4, item)

                                    if globalValues.polygon:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[10]))
                                    else:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[12]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 5, item)

                                    if globalValues.polygon:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[11]))
                                    else:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[13]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 6, item)

                                    if globalValues.polygon:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[5]))
                                    else:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[14]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 7, item)

                                    if globalValues.polygon:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[6]))
                                    else:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[15]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 8, item)

                                    if globalValues.polygon:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[12]))
                                    else:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[10]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 9, item)

                                    if globalValues.polygon:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[12]))
                                    else:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[10]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 10, item)

                                    if globalValues.polygon:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[14]))
                                    else:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[5]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 11, item)

                                    if globalValues.polygon:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[15]))
                                    else:
                                        item = QtWidgets.QTableWidgetItem(str(rowMain[6]))
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 12, item)

                                    # item = QtWidgets.QTableWidgetItem(str(rowsPolygon[valWdg][2]))
                                    # item.setTextAlignment(Qt.AlignCenter)
                                    # if (globalValues.colorForm == 1):
                                    #     item.setForeground(QtGui.QBrush(Qt.black))
                                    # else:
                                    #     item.setForeground(QtGui.QBrush(Qt.white))
                                    # item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    # self.tblTS.setItem(0, 13, item)

                                    strStOrder = str(rowMain[7])

                                    item = QtWidgets.QTableWidgetItem(strStOrder)
                                    item.setTextAlignment(Qt.AlignCenter)
                                    if (globalValues.colorForm == 1):
                                        item.setForeground(QtGui.QBrush(Qt.black))
                                    else:
                                        item.setForeground(QtGui.QBrush(Qt.white))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)

                                    if (strStOrder == 'не выполняется'):
                                        if (valColor == 0):
                                            item.setBackground(QtGui.QColor(149, 26, 26))
                                            item.setForeground(QtGui.QColor(255, 255, 255))
                                        else:
                                            item.setBackground(QtGui.QColor(197, 104, 104))
                                            item.setForeground(QtGui.QColor(0, 0, 0))
                                    elif (strStOrder == 'выполняется'):
                                        if (valColor == 0):
                                            item.setBackground(QtGui.QColor(149, 26, 26))
                                            item.setForeground(QtGui.QColor(255, 255, 255))
                                        else:
                                            item.setBackground(QtGui.QColor(197, 104, 104))
                                            item.setForeground(QtGui.QColor(0, 0, 0))
                                    elif (strStOrder == 'перемещение'):
                                        if (valColor == 0):
                                            item.setBackground(QtGui.QColor(6, 81, 16))
                                            item.setForeground(QtGui.QColor(255, 255, 255))
                                        else:
                                            item.setBackground(QtGui.QColor(84, 181, 100))
                                            item.setForeground(QtGui.QColor(0, 0, 0))
                                    elif (strStOrder == 'выполнен'):
                                        if (valColor == 0):
                                            item.setBackground(QtGui.QColor(6, 81, 16))
                                            item.setForeground(QtGui.QColor(255, 255, 255))
                                        else:
                                            item.setBackground(QtGui.QColor(84, 181, 100))
                                            item.setForeground(QtGui.QColor(0, 0, 0))

                                    self.tblTS.setItem(0, 13, item)

                                    item = QtWidgets.QTableWidgetItem(str(rowsMain[valWdg][0]))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 14, item)

                                    item = QtWidgets.QTableWidgetItem(str(rowsOrder[valWdg][0]))
                                    item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    self.tblTS.setItem(0, 15, item)

                                    # item = QtWidgets.QTableWidgetItem(str(rowsPolygon[valWdg][0]))
                                    # item.setFlags(QtCore.Qt.ItemIsSelectable | QtCore.Qt.ItemIsEnabled)
                                    # self.tblTS.setItem(0, 16, item)

                                    valWdg += 1

                                except Exception as ex:
                                    self.tblTS.removeRow(0)
                                    globalValues.writeLogData('Not data in mysqlBD', str(ex))

                            countRowsOld = countRows
                        else:
                            self.tblTS.setRowCount(0)

                        th_scroll = threading.Thread(target=self.thChangeScroll, args=(True,))
                        th_scroll.start()

                    break
                except Exception as ex:
                    globalValues.writeLogData('Проверка подключения к БД', str(ex))

                    time.sleep(0.3)

        except Exception as ex:
            globalValues.writeLogData('Функция поиска данны в БД', str(ex))

    def searchFilterEls(self):
        try:
            print('startingUpdateFilterTbl!!!')

            dataCb = self.comboFilter.currentIndex()

            if (dataCb == 1):

                self.updateTable()


                dataDate = self.DTEitFix.text()
                dataDate = dataDate[0 : 10]

                try:
                            countItems = self.tblTS.rowCount()
                            if (countItems == ''):
                                countItems = 0

                            countItems = int(countItems)

                            for i in range(countItems - 1, -1,  -1):
                                dataItem = str(self.tblTS.item(i, 4).text())
                                if (dataItem == ''):
                                    dataItem = ' '
                                is_data_in_db = dataDate in dataItem
                                if (is_data_in_db == False):
                                    self.tblTS.removeRow(i)


                except Exception as ex:
                    globalValues.writeLogData('Поиск по фильтру дата в таблице журнала ТС', str(ex))

            elif (dataCb == 2):
                self.updateTable()
                dataDateAt = self.DTEditRngAt.text()
                lenDateAt = len(dataDateAt)
                dataTimeAt = dataDateAt[11 : lenDateAt]
                if (len(dataTimeAt) == 4):
                    dataTimeAt = '0' + dataTimeAt
                dataDateAt = dataDateAt[0:10]


                dataDateTo = self.DTEditRngTo.text()
                lenDateTo = len(dataDateTo)
                dataTimeTo = dataDateTo[11: lenDateTo]
                if (len(dataTimeTo) == 4):
                    dataTimeTo = '0' + dataTimeTo
                dataDateTo = dataDateTo[0:10]

                startTimeVal = self.valueTimeToIntSearch(dataTimeAt)
                endTimeVal = self.valueTimeToIntSearch(dataTimeTo)



                try:
                    countItems = self.tblTS.rowCount()

                    if (countItems == ''):
                        countItems = 0

                    countItems = int(countItems)


                    for i in range(countItems - 1, -1, -1):

                        dataItem = str(self.tblTS.item(i, 4).text())
                        dataTimeItem = str(self.tblTS.item(i, 5).text())
                        dataTimeItem = dataTimeItem.replace('.', ':')
                        if (dataItem == ''):
                            dataItem = '0'

                        print(dataItem)

                        if self.checkDataTime(dataItem, dataDateAt, dataDateTo, startTimeVal, endTimeVal, dataTimeItem):
                            self.tblTS.removeRow(i)


                except Exception as ex:
                    print('Error!')
                    globalValues.writeLogData('Поиск по фильтру период даты в таблице журнала ТС', str(ex))

            elif (dataCb == 0):
                self.updateTable()
        except Exception as ex:
            globalValues.writeLogData('Функция поиска по фильтру журнал ТС', str(ex))

    def pasteInfoTS(self):

        indexStr = self.tblTS.currentRow()

        if (indexStr != self.indexStrOld):

            globalValues.pathFileVideoKppIn = ''
            globalValues.pathFileVideoKppOut = ''
            globalValues.pathFileVideoWeightIn = ''
            globalValues.pathFileVideoWeightOut = ''

            self.btnFilterSave.setEnabled(True)

            try:
                if globalValues.debug:
                    self.con = pymysql.connect(host='localhost',
                                               port=3306,
                                               user='sergey',
                                               passwd='34ubitav',
                                               db=globalValues.dbMySqlName)
                else:
                    if (self.checkMySql):
                        self.connectToMySql()

                cur = self.con.cursor()

                try:

                    start_time = round(time.time())

                    # with self.con:
                    #
                    #     id = int(self.tblTS.item(indexStr, 15).text())
                    #
                    #     strid = str(id)
                    #
                    #     sqlQ = "SELECT name_operator, driver_document FROM " + globalValues.tblsDB[2] + " WHERE id = " + strid
                    #
                    #     cur.execute(sqlQ)
                    #
                    #     rowsOrder = cur.fetchone()

                        # self.le_DriverFIO.setText(str(rowsOrder[0]))
                        # self.le_DrivePassportNumber.setText(str(rowsOrder[1]))

                except Exception as ex:
                    globalValues.writeLogData('Функция заполнения данных о ТС', str(ex))


                self.le_Organisation.setText(self.tblTS.item(indexStr, 2).text())
                self.le_GRZNumber.setText(self.tblTS.item(indexStr, 0).text())
                self.le_CarModel.setText(self.tblTS.item(indexStr, 3).text())
                self.le_ZakazNumber.setText(self.tblTS.item(indexStr, 1).text())
                self.le_ZakazDate.setText(self.tblTS.item(indexStr, 4).text())
                self.le_ZakazState.setText(self.tblTS.item(indexStr, 13).text())
                self.le_ComeToObject.setText(self.tblTS.item(indexStr, 5).text())
                self.le_GoneFromObject.setText(self.tblTS.item(indexStr, 6).text())
                self.le_ComeToPoligon.setText(self.tblTS.item(indexStr, 10).text())
                self.le_DriveOverTime.setText(self.tblTS.item(indexStr, 9).text())
                self.le_WeightObject.setText(self.tblTS.item(indexStr, 7).text())
                self.le_VolumeObject.setText(self.tblTS.item(indexStr, 8).text())
                self.le_WeightPoligon.setText(self.tblTS.item(indexStr, 11).text())
                self.le_VolumePoligon.setText(self.tblTS.item(indexStr, 12).text())

                numOrder = self.le_ZakazNumber.text()
                dateCur = self.le_ZakazDate.text()

                for i in range(4):
                    pathFldrVideo= globalValues.curDisk + ':/ACMK/ВидеоСобытий/' + self.lstChl[i] + '/' + dateCur
                    self.checkFolderLongPath(pathFldrVideo)
                    lstDataRes = self.checkAndPathVideoForReport(numOrder, pathFldrVideo)
                    if (lstDataRes[0]):
                        pathFileVideo = lstDataRes[1]
                        print('CheckingPath: ' + pathFileVideo)
                        checkMp4 = 'mp4' in pathFileVideo
                        pathFileVideo = pathFileVideo.replace('\\', '/')
                        print(pathFileVideo)
                        if (os.path.exists(pathFileVideo) and checkMp4):
                            if (i == 0):
                                globalValues.pathFileVideoKppOut = pathFileVideo
                                self.setMediaFrm(pathFileVideo, self.mediaPlayerKppOut)
                            elif (i == 1):
                                globalValues.pathFileVideoKppIn = pathFileVideo
                                self.setMediaFrm(pathFileVideo, self.mediaPlayerKppIn)
                            elif (i == 2):
                                globalValues.pathFileVideoWeightOut = pathFileVideo
                                self.setMediaFrm(pathFileVideo, self.mediaPlayerWeightOut)
                            elif (i == 3):
                                globalValues.pathFileVideoWeightIn = pathFileVideo
                                self.setMediaFrm(pathFileVideo, self.mediaPlayerWeightIn)
                    else:
                        pathFldrVideo = globalValues.curDisk + ':/ACMK/ВидеоСобытий/' + self.lstChl[i]
                        pathFoldersLst = os.listdir(pathFldrVideo)
                        pathLstDir = []
                        for el in pathFoldersLst:
                            if (os.path.isfile(el) == False):
                                pathLstDir.append(el)
                        pathLstDir.sort(key=self.sortByLength)
                        i_0 = 0
                        checkExistNextDate = False
                        dateNext = ''
                        for el in pathLstDir:
                            checkDate = dateCur in el
                            if (checkDate == True and ((i_0 + 1) != len(pathLstDir))):
                                checkExistNextDate = True
                                dateNext = pathLstDir[i_0 + 1]
                                break
                            i_0 += 1
                        # print(dateNext)
                        # print(pathLstDir)
                        if (checkExistNextDate):
                            pathFldrVideoNext = globalValues.curDisk + ':/ACMK/ВидеоСобытий/' + self.lstChl[i] + '/' + dateNext

                            print('CheckingDate!!!' + str(pathFldrVideoNext))

                            lstDataRes = self.checkAndPathVideoForReport(numOrder, pathFldrVideoNext)
                            if (lstDataRes[0]):
                                pathFileVideo = lstDataRes[1]
                                print('CheckingPath: ' + pathFileVideo)
                                checkMp4 = 'mp4' in pathFileVideo
                                pathFileVideo = pathFileVideo.replace('\\', '/')
                                print(pathFileVideo)
                                if (os.path.exists(pathFileVideo) and checkMp4):
                                    if (i == 0):
                                        globalValues.pathFileVideoKppOut = pathFileVideo
                                        self.setMediaFrm(pathFileVideo, self.mediaPlayerKppOut)
                                    elif (i == 1):
                                        globalValues.pathFileVideoKppIn = pathFileVideo
                                        self.setMediaFrm(pathFileVideo, self.mediaPlayerKppIn)
                                    elif (i == 2):
                                        globalValues.pathFileVideoWeightOut = pathFileVideo
                                        self.setMediaFrm(pathFileVideo, self.mediaPlayerWeightOut)
                                    elif (i == 3):
                                        globalValues.pathFileVideoWeightIn = pathFileVideo
                                        self.setMediaFrm(pathFileVideo, self.mediaPlayerWeightIn)

                    print('NumI: ' + str(i))

                strZN = self.le_ZakazNumber.text()

                pathFiles = glob.glob(globalValues.pathReports + '/*.pdf')
                print(pathFiles)
                if (len(pathFiles) != 0):
                    dataSearchOrder = strZN
                    dataSearchFormat = '.pdf'

                    for el in pathFiles:
                        print(el)
                        is_order = dataSearchOrder in el
                        is_fmt = dataSearchFormat in el
                        if (is_order and is_fmt):
                            self.btnFilterPrint.setEnabled(True)
                            break

                    if (is_order == False or is_fmt == False):
                        self.btnFilterPrint.setEnabled(False)
                else:
                    self.btnFilterPrint.setEnabled(False)

                cur.close()

            except Exception as ex:
                globalValues.writeLogData('Функция заполнения данных о ТС в журнале ТС', str(ex))

        self.indexStrOld = indexStr

        self.stopVideoKppIn()
        self.stopVideoKppOut()
        self.stopVideoWeightIn()
        self.stopVideoWeightOut()

    def saveReport(self):
        try:
            file_path = globalValues.pathReports + '/'
            num_rep = self.le_ZakazNumber.text()
            name_comp = self.le_Organisation.text()
            number_grz = self.le_GRZNumber.text()
            model_ts = self.le_CarModel.text()
            number_order = self.le_ZakazNumber.text()
            date_order = self.le_ZakazDate.text()
            st_order = self.le_ZakazState.text()
            date_in_object = self.le_ComeToObject.text()
            time_in_object = self.le_ComeToObject.text()
            date_out_object = self.le_GoneFromObject.text()
            time_out_object = self.le_GoneFromObject.text()
            date_in_poly = self.le_ComeToPoligon.text()
            time_in_poly = self.le_ComeToPoligon.text()
            time_in_road = self.le_DriveOverTime.text()
            weight_obj = self.le_WeightObject.text()
            weight_poly = self.le_WeightPoligon.text()
            vol_obj = self.le_VolumeObject.text()
            vol_poly = self.le_VolumePoligon.text()
            pathImgWeight = ''
            pathImgScan = ''

            rep = self.checkAndPathImgsForReport(number_order, globalValues.pathWeightImg)
            if rep[0]:
                pathImgWeight = rep[1]
                pathImgScan = rep[2]

            self.createReport(file_path, number_order, name_comp, number_grz, model_ts, number_order, date_order, date_order, st_order, date_in_object, time_in_object, date_out_object, time_out_object, date_in_poly, time_in_poly, time_in_road, weight_obj, vol_obj, weight_poly, vol_poly, pathImgWeight, pathImgScan)

            strDataToTS = 'Выполнено сохранение З/Н #' + str(number_order)
            globalValues.writeEventToDBJournalMain('Журнал З/Н ТС', strDataToTS)

        except Exception as ex:
            globalValues.writeLogData('Функция сохранения З/Н', str(ex))

    def createReport(self, file_PATH, RepNum, OrgName, GRZ, CarModel, TalonNum, TalonDate, ZNDate, ZNState, DateObjCome, TimeObjCome, DateObjGone, TimeObjGone, DatePolyCome, TimePolyCome, TimeInRoad, WeightObj, VolObj, WeightPoly, VolPoly, pathImgGRZ, pathImgGRUNT):
        try:
            # Формируем локальные переменные из переданных в функцию для дальнейшей вставки в пдф
            strRepNum = "Отчет #" + str(RepNum)
            strOrgName = "Организация:   " + str(OrgName)
            # strGRZNum = "Номер ГРЗ ТС:  " + str(GRZ[0]) + str(GRZ[1]) + str(GRZ[2]) + str(GRZ[3]) + str(GRZ[4]) + str(GRZ[5]) + str(GRZ[6]) + " " + str(GRZ[7]) + str(GRZ[8]) + str(GRZ[9])
            strGRZNum = "Номер ГРЗ ТС:  " + str(GRZ[0 : 6]) + " " + str(GRZ[6 : len(GRZ)])
            strCarModel = "Модель ТС:       " + str(CarModel)
            strTalonNum = "Номер талона:             " + str(TalonNum)
            strTalonDate = "Дата выдачи талона:  " + str(TalonDate)
            strZNState = "Состояние З/Н:            " + str(ZNState)
            strZNDate = "Дата закрытия З/Н:     " + str(ZNDate)
            strTimeObjCome = "Приехал на обьект:     " + str(DateObjCome) + " / " + str(TimeObjCome)
            strTimeObjGone = "Выехал с обьекта:       " + str(DateObjGone) + " / " + str(TimeObjGone)
            strTimePolyCome = "Приехал на полигон:   " + str(DatePolyCome) + " / " + str(TimePolyCome)
            strTimeInRoad = "Время в пути:               " + str(TimeInRoad)
            strWeightObj = "Масса грунта на обьекте:     " + str(WeightObj)
            strVolObj = "Объем грунта на обьекте:    " + str(VolObj)
            strWeightPoly = "Масса грунта на полигоне:   " + str(WeightPoly)
            strVolPoly = "Объем грунта на полигоне:  " + str(VolPoly)
            strFrameGRZ = "Фотофиксация ГРЗ ТС"
            strFrameGRUNT = "Фотофиксация ГРУНТ"
            strRepDate = "Дата создания отчета: " + str(date.today().day) + "." + str(date.today().month) + "." + str(date.today().year) + "г."

            # Заполнение страницы
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font('DejaVu', '', 'E:/grunt64venv/Lib/site-packages/fpdf/font/DejaVuSans.ttf', uni=True)
            pdf.set_font('DejaVu', '', 14)
            pdf.cell(200, 10, txt=strRepNum, ln=1, align="C")
            pdf.line(86, 18, 134, 18)
            pdf.set_font('DejaVu', '', 10)
            pdf.cell(70, 2, txt="", ln=1, align="L")
            pdf.cell(70, 5, txt=strOrgName, ln=1, align="L")
            pdf.cell(70, 5, txt=strGRZNum, ln=1, align="L")
            pdf.cell(70, 5, txt=strCarModel, ln=1, align="L")
            pdf.cell(70, 2, txt="", ln=1, align="L")
            pdf.line(11, 38, 198, 38)
            pdf.cell(70, 5, txt=strTalonNum, ln=1, align="L")
            pdf.cell(70, 5, txt=strTalonDate, ln=1, align="L")
            pdf.cell(70, 5, txt=strZNState, ln=1, align="L")
            pdf.cell(70, 5, txt=strZNDate, ln=1, align="L")
            pdf.cell(70, 5, txt=strTimeObjCome, ln=1, align="L")
            pdf.cell(70, 5, txt=strTimeObjGone, ln=1, align="L")
            pdf.cell(70, 5, txt=strTimePolyCome, ln=1, align="L")
            pdf.cell(70, 5, txt=strTimeInRoad, ln=1, align="L")
            pdf.cell(70, 2, txt="", ln=1, align="L")
            pdf.line(11, 80, 198, 80)
            pdf.cell(70, 5, txt=strWeightObj, ln=1, align="L")
            pdf.cell(70, 5, txt=strVolObj, ln=1, align="L")
            pdf.cell(70, 5, txt=strWeightPoly, ln=1, align="L")
            pdf.cell(70, 5, txt=strVolPoly, ln=1, align="L")
            pdf.line(11, 102, 198, 102)
            if (os.path.exists(pathImgGRZ)):
                pdf.image(pathImgGRZ, x=45, y=110, w=130)
            if (os.path.exists(pathImgGRUNT)):
                pdf.image(pathImgGRUNT, x=75, y=191, w=70)
            pdf.cell(200, 2, txt="", ln=1, align="C")
            pdf.cell(200, 8, txt=strFrameGRZ, ln=1, align="C")
            pdf.cell(200, 73, txt="", ln=1, align="C")
            pdf.cell(200, 8, txt=strFrameGRUNT, ln=1, align="C")
            pdf.cell(200, 83, txt="", ln=1, align="C")
            pdf.cell(190, 1, txt=strRepDate, ln=1, align="R")

            # Делаем имя файла и сохраняем
            filename = str(file_PATH) + str(RepNum) + "_" + str(date.today()) + ".pdf"
            pdf.output(filename)
            self.pathFileCur = filename
            # self.pFile_PARTH = file_PATH
            # self.pFile_NAME = str(RepNum) + "_" + str(date.today()) + ".pdf"
            time.sleep(0.3)
            if (os.path.exists(filename)):
                # print(filename)
                os.system(filename)
            self.btnFilterPrint.setEnabled(True)

            return True
        except Exception as ex:
            globalValues.writeLogData('Функция создания отчёта в пдф', str(ex))
            uiMes = Ui_mes_box()
            uiMes.lblStrInfo.setText('Произошла ошибка при создании отчёта!')
            uiMes.btnOK.hide()
            if (globalValues.colorForm == 1):
                uiMes.label_iconques.setStyleSheet("background-color: rgb(235,235,235);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
            else:
                uiMes.label_iconques.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
            uiMes.btnCancel.setText('Продолжить')
            uiMes.exec_()
            return False

    def printData(self):
        self.printReport(self.pathFileCur)

    def printReport(self, pathFile):
        try:
            # PPrinter = ping3.ping(DefaultPrinter_IP)
            # if (str(PPrinter) != 'None'):
                  #формируем полное имя файла (путь каталога+имя, оба параметра получаем из функции создания отчета через глобал параметры)
                #скрипт печати
            check_file = 'pdf' in pathFile
            if (os.path.exists(pathFile) and check_file):
                # win32api.ShellExecute(
                #     0,
                #     "print",
                #     pathFile,
                #     '/d:"%s"' % win32print.GetDefaultPrinter(),
                #     ".",
                #     0
                # )
                PPrinter = ""

                os.startfile(pathFile, "print")

                strData = str(pathFile)
                strDataToTS = 'Выполнена печать З/Н(' + strData + ')'
                globalValues.writeEventToDBJournalMain('Журнал З/Н ТС', strDataToTS)

                print("PrinterCheckOK!!!")
            # else:
            #     print("PrinterCheck FALSE!!!")
                return True
        except Exception as ex:
            globalValues.writeLogData('Функция печати файла отчёта', str(ex))
            globalValues.writeLogData('Функция создания отчёта в пдф', str(ex))
            uiMes = Ui_mes_box()
            uiMes.lblStrInfo.setText('Произошла ошибка при печати отчёта!')
            uiMes.btnOK.hide()
            if (globalValues.colorForm == 1):
                uiMes.label_iconques.setStyleSheet("background-color: rgb(235,235,235);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
            else:
                uiMes.label_iconques.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
            uiMes.btnCancel.setText('Продолжить')
            uiMes.exec_()
            return False

    def checkFolderLongPath(self, pathFolder):
        try:
            i = 0
            listPath = []
            for element in pathFolder:
                if (element == '/' and i != 0):
                    listPath.append(pathFolder[0:i])
                i += 1
            # listPath = pathFolder.split('/')
            # listPath.pop(0)
            listPath.append(pathFolder)
            for elPath in listPath:
                if (os.path.exists(elPath) == False):
                    try:
                        os.mkdir(elPath)
                    except Exception as ex:
                        globalValues.writeLogData('Функция проверки и создания папки хранилища', str(ex))

            if (len(listPath) > 0):
                return True
            else:
                return False
        except Exception as ex:
            globalValues.writeLogData('Функция проверки ссылки на папку хранилища', str(ex))
            return False

    def sortByLength(self, inputStr):
        return len(inputStr)

    def checkAndPathVideoForReport(self, numOrder, pathFldrVideo):
        try:
            lengthPath = len(pathFldrVideo)
            pathFile = ''
            pathFileVideo = ''

            for name in glob.glob(pathFldrVideo + '/*.mp4'):
                pathFile = name
                name = name[lengthPath + 1:len(name)]
                checkOrder = str(numOrder) in name

                if (checkOrder):
                    pathFileVideo = pathFile
                    return True, pathFileVideo

            return False, pathFileVideo
        except Exception as ex:
            globalValues.writeLogData('Функция проверки видео для отчёта', str(ex))

    def checkAndPathImgsForReport(self, numOrder, pathImg):
        try:
            lengthPath = len(pathImg)
            pathFile = ''
            isWeight = False
            isScan = False
            pathImgWeight = ''
            pathImgScan = ''

            for name in glob.glob(pathImg + '/*.jpg'):
                pathFile = name
                name = name[lengthPath:len(name)]
                checkOrder = str(numOrder) in name
                checkWeight = 'weight' in name
                checkScan = 'scan' in name
                if (checkOrder and checkWeight):
                    pathImgWeight = pathFile
                    isWeight = True
                if (checkOrder and checkScan):
                    pathImgScan = pathFile
                    isScan = True
                if (isScan and isWeight):
                    return True, pathImgWeight, pathImgScan

            if (isWeight or isScan):
                return True, pathImgWeight, pathImgScan
            else:
                return False, pathImgWeight, pathImgScan
        except Exception as ex:
            globalValues.writeLogData('Функция проверки картинок для отчёта', str(ex))

    def checkDataTime(self, timeEl, timeAt, timeTo, startTimeVal, endTimeVal, dataTimeItem):
        try:
            valDateAt = int(timeAt[0:2])
            valMonthAt = int(timeAt[3:5])
            valYearAt = int(timeAt[6:10])

            valDateTo = int(timeTo[0:2])
            valMonthTo = int(timeTo[3:5])
            valYearTo = int(timeTo[6:10])

            valDateEl = int(timeEl[0:2])
            valMonthEl = int(timeEl[3:5])
            valYearEl = int(timeEl[6:10])

            if (valMonthAt == valMonthTo and valYearAt == valYearTo):
                print('changeDay')
                if (valDateAt <= valDateEl <= valDateTo):
                    timeElTbl = self.valueTimeToIntTbl(dataTimeItem)

                    if (valDateEl == valDateAt):
                        if (timeElTbl < startTimeVal):
                            return True
                    if (valDateEl == valDateTo):
                        if (timeElTbl > endTimeVal):
                            return True

                    return False
                return True
            elif (valMonthAt != valMonthTo and valYearAt == valYearTo):
                print('changeMonth')
                if ((valDateAt <= valDateEl and valMonthEl == valMonthAt) or (valDateEl <= valDateTo and valMonthEl == valMonthTo) or (valMonthAt < valMonthEl < valMonthTo)):
                    timeElTbl = self.valueTimeToIntTbl(dataTimeItem)

                    if (valDateEl == valDateAt and valMonthEl == valMonthAt):
                        if (timeElTbl < startTimeVal):
                            return True
                            # self.listItem.append(i)
                    if (valDateEl == valDateTo and valMonthEl == valMonthTo):
                        if (timeElTbl > endTimeVal):
                            return True

                    return False
                return True
            elif (valYearAt != valYearTo):
                print('changeYear')
                timeElTbl = self.valueTimeToIntTbl(dataTimeItem)

                if ((valYearEl == valYearAt and valMonthEl > valMonthAt) or (
                        valYearEl == valYearTo and valMonthEl < valMonthTo) or (valYearAt < valYearEl < valYearTo) or (
                        valYearEl == valYearAt and valMonthEl == valMonthAt and valDateEl >= valDateAt) or (
                        valYearEl == valYearTo and valMonthEl == valMonthTo and valDateEl <= valDateTo)):

                    if (valDateEl == valDateAt and valMonthEl == valMonthAt and valYearEl == valYearAt):
                        if (timeElTbl < startTimeVal):
                            return True
                    if (valDateEl == valDateTo and valMonthEl == valMonthTo and valYearEl == valYearTo):
                        if (timeElTbl > endTimeVal):
                            return True

                    return False

                return True
            else:
                print('NotCase!')
                return False

        except Exception as ex:
            globalValues.writeLogData('Функция проверки записи датаэдита', str(ex))

    def valueTimeToIntTbl(self, strTime):
        strTime = strTime
        strTime = strTime.replace(':', '')
        return(int(strTime))

    def valueTimeToIntSearch(self, strTime):
        strTime = strTime + '00'
        strTime = strTime.replace(':', '')
        return(int(strTime))

    def decryptData(self, listEls):
        try:
            private_key = """-----BEGIN RSA PRIVATE KEY-----
                MIIEowIBAAKCAQEA0c45dZkOTBkAi/eP4i6Grc1GVPVw1a+dni+ciOqACFISHEls
                PHGUIBIpzAxVUt7Kgc3QPZ9amMewFTz+whoe7HjAf+wtC+ISBDw8s7CQxknbeUAX
                NYeShbtP1l8Mz6sHJmSk97qFMdyICOALIKY6ky0rYxG7AsqCbsVBe2MKGqKKGuho
                c4Ei35vNBVPn3KLNmSqjfKk4QwNmttwP6gEqldZwT6pjz2wgA42CMQYkVEQ4w4Bw
                a0Go/C1AsrwLd7q3tgqM+tAEctICAjNAdTKMhFWMOdUn60yaEFcZPo7EE0dUw5ZY
                Up91fqQBDoxt2D3p0iWmTEKZPgx9Tbtu+Z6/JwIDAQABAoIBACe6f1Lvaq+qRFo8
                xLg1yzb6GglYeMdd++DKbz/V9+ybbeaBWMeRUlVIWzXSWA3bNkmiKX6hwEwR9Bvx
                cuRageSRcRJILLeFVZgLuArmsmN59N9e7YYrZ+l+8L1NPmXMowv4HuzyGuq4MeJM
                Wo8SKyFXelHGN71tj4lePOoadP1Z1eS9MvScAfs0Noek71CiBpdWG9dmPksRSNOl
                y37W3MypfdD/9na+sXoIxCZGoLudY/KheSBZ++m/pZZave/g4xrHFfKRR7xtGNI7
                CKCJNaijqnJK62EEKwv4SWZIarfRBjhTDcbXfz+Kz1PT6NSvzoG9/7YpXFkhCRio
                mEZmNFECgYEA1D5gCtvgz5nXKcMVBRkPDgU4Kv9EPXZwiut0zW6XNZpaF6neXjXm
                Rx633aHRWPzhZLX0jrgkNlCDgmxL9HiVu/erGql9bpwraZi4J7nft9voifKnaXva
                EsOLEGE3k5oVX3xkgkgfqRLEGFppSIqJkdv8CgX+oqJzq8euje/B2lUCgYEA/Q8s
                fhDHQn9GxJjwEKaWR/IU0EDzfOWVit7IBnCjb3v+1RwrAbT55hTFLW3FoILzcX1I
                Tr7kQpyScly1ucjALBAEQOYciEVeFLvAGhoiuJCiahTmQI+4QGrjHD5hhjNkvqrN
                3kDZh986aLYNBLK61mx6Ml1Dyas0b3ZBVrlxZ4sCgYAeTY2O31fYrCFRQB4nLS2+
                FbawROPsVpW47+csUYbbS19jk4hBMTbgnp0n0qu+JdTUeToiil35N0OfgnDRxcmz
                HahbVSmoejmkiP56BYrQiGBKGdAXOmynUy3ut8Kkm1JD4NHE3CFRFXHT/Eyd49HC
                doMktzhk5gbX1tmwQDQQRQKBgDDIrhEXdvJQyvm3agArvSjdeDm1a7sWH0AINpNX
                P4qMYtH+fiP0GYDLXD+nu8N3uyqTtk7H6gUVXf4B9V59Xt6fr9I7CiETDlH858mg
                ZDUkXMsKgGDN0/1HHcUiGXbfjXpcPxerdMQGuqHZBqVzNyWDAAOZiynjgVZDe9EW
                KtCFAoGBAM87MQ8T3pORPACj2nUojbjypT0RskqU0GHmqKFkoo4zqY0IPM/dN8Ac
                DB3iPC4i90SWvJyqLd1MV3xm7+uQ9l/p3Jvo7/ClVlPRKZYHbPieBZABMgc+VQ1k
                xNEm2eKuM7oat7/vTmGsCFuMwd9+4Z7HrfGdoS6ilJRvX4jAzimw
                -----END RSA PRIVATE KEY-----"""
            ListData = []


            for el in listEls:
                # print(dataElement)
                    rsa_key = RSA.importKey(private_key)
                    rsa_key = PKCS1_OAEP.new(rsa_key)
                    raw_cipher_data = b64decode(el)
                    decrypted = rsa_key.decrypt(raw_cipher_data)
                    strDecrypted = decrypted.decode('utf8')
                    ListData.append(strDecrypted)

            return ListData

        except Exception as ex:
            globalValues.writeLogData('Функция декодирования данных проверки устройств в файле', str(ex))

    def encrypt(self, strEncrypted, numberRow, numberCol):
        try:

            public_key = """-----BEGIN PUBLIC KEY-----
        MIIBIjANBgkqhkiG9w0BAQEFAAOCAQ8AMIIBCgKCAQEA0c45dZkOTBkAi/eP4i6G
        rc1GVPVw1a+dni+ciOqACFISHElsPHGUIBIpzAxVUt7Kgc3QPZ9amMewFTz+whoe
        7HjAf+wtC+ISBDw8s7CQxknbeUAXNYeShbtP1l8Mz6sHJmSk97qFMdyICOALIKY6
        ky0rYxG7AsqCbsVBe2MKGqKKGuhoc4Ei35vNBVPn3KLNmSqjfKk4QwNmttwP6gEq
        ldZwT6pjz2wgA42CMQYkVEQ4w4Bwa0Go/C1AsrwLd7q3tgqM+tAEctICAjNAdTKM
        hFWMOdUn60yaEFcZPo7EE0dUw5ZYUp91fqQBDoxt2D3p0iWmTEKZPgx9Tbtu+Z6/
        JwIDAQAB
        -----END PUBLIC KEY-----"""
            rsa_key = RSA.importKey(public_key)
            rsa_key = PKCS1_OAEP.new(rsa_key)
            encrypted = rsa_key.encrypt(strEncrypted.encode('utf8'))
            encrypted_text = b64encode(encrypted)

            pathFile = str(os.getenv('APPDATA')) + r'\Sinaps\cryptoSSA.xlsx'
            # excel = win32com.client.Dispatch('Excel.Application')
            # workbook = excel.Workbooks.open(pathFile, True, False, None, '34ubitav')
            # sheet = workbook.Worksheets('Лист1')
            # excel.Visible = False
            # excel.DisplayAlerts = False

            pathXls = globalValues.pathFileXls
            print(pathXls)

            book = openpyxl.load_workbook(pathXls)
            sheets = book.sheetnames
            print('qwer')
            for sheet in sheets:
                print(str(sheet))
                print('qwerty!!!')
                if (str(sheet) == 'Лист1'):
                    sheet_main = book[sheet]
                    encrypted_text = str(encrypted_text)
                    length = len(encrypted_text)
                    encrypted_text = encrypted_text[2: length - 1]
                    sheet_main.cell(numberRow, numberCol).value = str(encrypted_text)

            # workbook.Close(True, pathFile)
            # excel.Quit

        except Exception as ex:
            globalValues.writeLogData('Функция кодирования данных и сохранения в запароленный файл', str(ex))

    def th_find_zakaz(self):

        print('')

        try:

            firstCall = True
            findOldData = False
            checkExit = False
            lstDataBD = []
            lstFindData = []

            strGRZ = 'C747PC7999'
            strGRZ_1 = 'C747PC799'
            num_order_ = 100002
            arr_ = []
            arr_.append('search')
            arr_.append(strGRZ)
            lstFindData.append(arr_)
            arr_ = []
            arr_.append('search')
            arr_.append(strGRZ_1)
            lstFindData.append(arr_)

            # lstDataToLoad = []
            arr = []
            arr.append('load')
            arr.append(strGRZ_1)
            arr.append(num_order_)
            arr.append('27190')
            arr.append('18.01')
            arr.append('15:00:06')
            arr.append('16:48:36')
            lstFindData.append(arr)
            print('First: ', lstFindData)
            lstFindData.pop()
            print('Sec: ', lstFindData)



            print(lstFindData)
            # print(lstFindData[len(lstFindData)-1])
            # lstFindData.pop()
            # print(lstFindData)
            start_time = round(time.time()*100)
            print('start time: ', start_time)
            delta_search = 6000

            while True:

                print(len(lstFindData))

                if (abs(round(time.time()*100) - start_time) > delta_search and len(lstFindData) > 0):
                    print(abs(round(time.time()) - start_time))
                    print(start_time)
                    print(round(time.time()))
                    start_time = round(time.time())
                    findOldData = True
                    print('checkingTrueSearch!')

                if (firstCall):

                    firstCall = False
                    i = 5
                    # pathFile = globalValues.pathDefFldr + r'\cryptoSSA.xlsx'
                    # excel = win32com.client.Dispatch('Excel.Application')
                    # workbook = excel.Workbooks.open(pathFile, True, False, None, '34ubitav')
                    # sheet = workbook.Worksheets('Лист1')
                    # excel.Visible = False
                    # excel.DisplayAlerts = False

                    pathXls = globalValues.pathFileXls
                    print(pathXls)

                    book = openpyxl.load_workbook(pathXls)
                    sheets = book.sheetnames
                    print('qwer')
                    for sheet in sheets:
                        print(str(sheet))
                        print('qwerty!!!')
                        if (str(sheet) == 'Лист1'):
                            sheet_main = book[sheet]

                            dataElement = sheet_main.cell(i, 1).value

                            while dataElement != None:
                                arrBD = []
                                for j in range(1, 6, 1):
                                    arrBD.append(sheet_main.cell(i, j).value)
                                lstDataBD.append(arrBD)
                                i += 1
                                dataElement = sheet_main.cell(i, 1).value
                                print(dataElement)

                            print(lstDataBD)
                            print(len(lstDataBD))

                nameTbl = globalValues.tblsDB[1]

                # print(nameTbl)

                if (globalValues.find_zakaz or findOldData):

                    findOldData = False
                    lstSearchBD = lstDataBD
                    count_data = 0
                    valueCell = ''


                    lstCheckBD = []
                    lstNotConBD = []

                    start_time = round(time.time()*100)
                    checkFindOrder = False
                    numCheckBD = 0

                    while True:

                        print('serhio!')

                        lstCheckBD = []
                        lstNotConBD = []

                        for elBD in lstSearchBD:
                            print(lstSearchBD)
                            curPort = int(elBD[1])
                            lstLogPas = []
                            lstLogPas.append(elBD[2])
                            lstLogPas.append(elBD[3])
                            lstLogPas = self.decryptData(lstLogPas)
                            # print(lstLogPas)

                            strAbout = 'База данных: ' + str(elBD[0]) + ', ' + str(elBD[4])
                            checkBD = False

                            try:

                                print('checking!')

                                con = pymysql.connect(host= elBD[0], port= curPort, user= lstLogPas[0], passwd= lstLogPas[1], db= elBD[4])

                                checkBD = con.open
                                lstCheckBD.append(checkBD)

                                print('cheecking!!!')

                                cur = con.cursor()

                                lstSearchData = []

                                print(globalValues.find_zakaz)

                                if (globalValues.find_zakaz):
                                    lstSearchData.append(lstFindData[len(lstFindData)-1])
                                    globalValues.find_zakaz = False
                                    # lstFindData.pop()
                                else:
                                    lstSearchData = lstFindData

                                print(lstSearchData)

                                k = 0

                                for el in lstSearchData:

                                    str_grz = el[1]

                                    if (el[0] == 'load'):
                                        cur.execute("SELECT * FROM " + str(nameTbl) + " WHERE number_grz='" + str_grz + "' and number_order=" + el[2] + " and state_order='перемещение'")
                                    else:
                                        cur.execute("SELECT * FROM " + str(nameTbl) + " WHERE number_grz='" + str_grz + "' and state_order='перемещение'")

                                    row = cur.fetchall()

                                    print(len(row))
                                    print('Sergio!!!')
                                    print(row)

                                    if (len(row) > 0):
                                        checkFindOrder = True
                                        lstFindData.pop(k)

                                    if (checkFindOrder and str(el[0]) == 'load'):
                                        # globalValues.checkLoadData = False

                                        cur.execute("UPDATE " + str(nameTbl) + " SET "
                                           "weight_bd='" + str(el[3]) + "', volume_bd='" + str(el[4]) + "'"
                                            ", time_entry_bd='" + str(el[5]) + "', time_check_out_bd='" + str(el[6]) + "'"
                                             " WHERE number_grz='" + str_grz + "' and number_order=" + el[2] + " and state_order='перемещение'")
                                        con.commit()

                                    elif (checkFindOrder):

                                            lstAllData = []
                                            for el in row:
                                                lstAllData.append(el[1])
                                                lstAllData.append(el[5])
                                                lstAllData.append(el[6])
                                                lstAllData.append(el[8])
                                                lstAllData.append(el[9])
                                                lstAllData.append(el[10])
                                                lstAllData.append(el[11])
                                                break

                                            lstData = str(lstAllData[1]).split(' ')
                                            lstAllData[1] = lstData[0]
                                            lstData = str(lstAllData[2]).split(' ')
                                            lstAllData[2] = lstData[0]

                                            try:
                                                if globalValues.debug:
                                                    self.con = pymysql.connect(host='localhost',
                                                                               port=3306,
                                                                               user='sergey',
                                                                               passwd='34ubitav',
                                                                               db=globalValues.dbMySqlName)
                                                else:
                                                    if (self.checkMySql):
                                                        self.connectToMySql()

                                                cur_my = self.con.cursor()

                                                print(lstAllData[1])

                                                # lstAllData[0] = 100003

                                                cur_my.execute("UPDATE " + str(nameTbl) + " SET number_order=" + str(lstAllData[0]) + ", "
                                                    "weight_bd='" + str(lstAllData[1]) + "', volume_bd='" + str(lstAllData[2]) + "'"
                                                    ", time_entry_bd='" + str(lstAllData[5]) + "', time_check_out_bd='" + str(lstAllData[6]) + "'"
                                                    " WHERE number_grz='" + str_grz + "' and state_order='выполняется'")

                                                self.con.commit()

                                                print(self.con.open)





                                                cur_my.close()

                                                self.con.close()

                                                print('qwertyui')

                                            except Exception as ex:
                                                globalValues.writeLogData('Запись данных с объекта в БД', str(ex))

                                            print(lstAllData)

                                    k += 1

                                cur.close()

                                con.close()

                                if (checkFindOrder):
                                        break

                            except Exception as ex:
                                lstCheckBD.append(checkBD)
                                globalValues.writeLogData(strAbout, str(ex))
                                lstNotConBD.append(elBD)


                        if (checkFindOrder):
                                print('OrderFind!!!')
                                break

                        numCheckBD += 1

                        print(numCheckBD)

                        if (numCheckBD >= 3):
                            # arr = []
                            # arr.append(strGRZ)
                            # arr.append(num_order)
                            # lstFindData.append(arr)
                            print('notDatainBD')
                            break

                        lstSearchBD = lstNotConBD

                        time.sleep(1)

                        print('checkingBD!!!')

                            # con = pymysql.connect(host='localhost', port=3306, user='sergey', passwd='34ubitav', db=globalValues.dbMySqlName)

                    print(lstCheckBD)

                # print(lstFindData)

                time.sleep(0.05)

                # break



        except Exception as ex:
            globalValues.writeLogData('Поток поиска З/Н', str(ex))

    def read_data_from_xls(self, row, col):
        try:

            # pathFile = str(os.getenv('APPDATA')) + r'\Sinaps\cryptoSSA.xlsx'
            # excel = win32com.client.Dispatch('Excel.Application')
            # workbook = excel.Workbooks.open(pathFile, True, False, None, '34ubitav')
            # sheet = workbook.Worksheets('Лист1')
            # excel.Visible = False
            # excel.DisplayAlerts = False

            pathXls = globalValues.pathFileXls
            print(pathXls)

            book = openpyxl.load_workbook(pathXls)
            sheets = book.sheetnames
            print('qwer')
            for sheet in sheets:
                print(str(sheet))
                print('qwerty!!!')
                if (str(sheet) == 'Лист1'):
                    sheet_main = book[sheet]

                    # for i in range(lengtnListEls):
                    dataElement = sheet_main.cell(row, col).value

                    print(dataElement)

            # workbook.Close(False)
            # excel.Quit

        except Exception as ex:
            globalValues.writeLogData('Считывание данных из файла', str(ex))

if __name__ == "__main__":

    globalValues.find_zakaz = True

    ui = Ui_PanelJournalTS()

    # ui.show()

    ui.th_find_zakaz()

    # sys.exit(app.exec_())
