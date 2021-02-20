from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QDialog, QApplication
from PyQt5.QtGui import QPixmap, QScreen
# import pyautoguiogui
from base64 import b64decode, b64encode
from Crypto.PublicKey import RSA
from Crypto.Cipher import PKCS1_OAEP
# import win32com.client
import os
import subprocess
import globalValues
import time
import openpyxl
# from MainPanelGrunt import app
import sys
from PyQt5.QtGui import QMovie

globalValues.colorForm = 1



class Ui_panel_autologin(QDialog):

    isChangeLogin = False

    change_login = False
    check_login = False
    check_password = False
    check_good_autologin = False

    lstLight = [[]]
    lstDark = [[]]
    lengthLight = 0
    lengthDark = 0
    app = QApplication(sys.argv)

    def __init__(self):
        pathFolder = str(os.getenv('APPDATA')) + r'\Sinaps'
        self.checkFolderPath(pathFolder, True)
        # self.screenImgDisplay()
        super().__init__()
        self.run_ui()

    def run_ui(self):
        self.setObjectName("panel_autologin")
        self.setFixedSize(1920, 1080)
        self.setWindowFlags(QtCore.Qt.CustomizeWindowHint | QtCore.Qt.FramelessWindowHint)
        self.setAutoFillBackground(False)
        self.leLogin = QtWidgets.QLineEdit(self)
        self.leLogin.setGeometry(QtCore.QRect(828, 560, 271, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.leLogin.setFont(font)
        self.leLogin.setObjectName("leLogin")
        self.lePassword = QtWidgets.QLineEdit(self)
        self.lePassword.setGeometry(QtCore.QRect(828, 597, 271, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.lePassword.setFont(font)
        self.lePassword.setObjectName("lePassword")
        self.lbl = QtWidgets.QLabel(self)
        self.lbl.setGeometry(QtCore.QRect(873, 520, 181, 23))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.lbl.setFont(font)
        self.lbl.setObjectName("lbl")
        self.btnInput = QtWidgets.QPushButton(self)
        self.btnInput.setGeometry(QtCore.QRect(913, 650, 101, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.btnInput.setFont(font)
        self.btnInput.setObjectName("butInput")
        self.label_33 = QtWidgets.QLabel(self)
        self.label_33.setGeometry(QtCore.QRect(758, 378, 404, 22))
        self.label_33.setObjectName("label_33")
        self.label_34 = QtWidgets.QLabel(self)
        self.label_34.setGeometry(QtCore.QRect(762, 381, 16, 16))
        self.label_34.setText("")
        self.label_34.setObjectName("label_34")
        self.btnCloseMainForm = QtWidgets.QPushButton(self)
        self.btnCloseMainForm.setGeometry(QtCore.QRect(1140, 378, 22, 22))
        self.btnCloseMainForm.setText("")
        self.btnCloseMainForm.setObjectName("btnCloseMainForm")
        self.label = QtWidgets.QLabel(self)
        self.label.setGeometry(QtCore.QRect(758, 399, 404, 302))
        self.label.setText("")
        self.label.setObjectName("label")
        self.label_setscreen = QtWidgets.QLabel(self)
        self.label_setscreen.setGeometry(QtCore.QRect(0, 0, 1920, 1080))
        self.label_setscreen.setObjectName("label_setscreen")
        self.label_gradient = QtWidgets.QLabel(self)
        self.label_gradient.setGeometry(QtCore.QRect(0, 0, 1920, 1080))
        self.label_gradient.setObjectName("label_gradient")
        # self.label_4 = QtWidgets.QLabel(self)
        # self.label_4.setGeometry(QtCore.QRect(0, 600, 47, 13))
        # self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self)
        self.label_5.setGeometry(QtCore.QRect(1161, 381, 31, 318))
        self.label_5.setObjectName("label_5")
        self.label_6 = QtWidgets.QLabel(self)
        self.label_6.setGeometry(QtCore.QRect(727, 381, 31, 318))
        self.label_6.setObjectName("label_6")
        self.label_7 = QtWidgets.QLabel(self)
        self.label_7.setGeometry(QtCore.QRect(758, 350, 403, 31))
        self.label_7.setObjectName("label_7")
        self.label_8 = QtWidgets.QLabel(self)
        self.label_8.setGeometry(QtCore.QRect(1161, 320, 61, 61))
        self.label_8.setText("")
        self.label_8.setObjectName("label_8")
        self.label_10 = QtWidgets.QLabel(self)
        self.label_10.setGeometry(QtCore.QRect(1161, 699, 61, 61))
        self.label_10.setText("")
        self.label_10.setObjectName("label_10")
        self.label_11 = QtWidgets.QLabel(self)
        self.label_11.setGeometry(QtCore.QRect(697, 320, 61, 61))
        self.label_11.setText("")
        self.label_11.setObjectName("label_11")
        self.label_12 = QtWidgets.QLabel(self)
        self.label_12.setGeometry(QtCore.QRect(758, 699, 403, 31))
        self.label_12.setObjectName("label_12")
        self.label_13 = QtWidgets.QLabel(self)
        self.label_13.setGeometry(QtCore.QRect(696, 699, 62, 61))
        self.label_13.setText("")
        self.label_13.setObjectName("label_13")
        self.lbl_2 = QtWidgets.QLabel(self)
        self.lbl_2.setGeometry(QtCore.QRect(770, 410, 161, 23))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(18)
        font.setBold(True)
        font.setWeight(75)
        self.lbl_2.setFont(font)
        self.lbl_2.setObjectName("lbl_2")
        self.lbl_3 = QtWidgets.QLabel(self)
        self.lbl_3.setGeometry(QtCore.QRect(770, 435, 301, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(9)
        font.setBold(False)
        font.setWeight(50)
        self.lbl_3.setFont(font)
        self.lbl_3.setAlignment(QtCore.Qt.AlignLeading|QtCore.Qt.AlignLeft|QtCore.Qt.AlignTop)
        self.lbl_3.setObjectName("lbl_3")
        self.label_setscreen.raise_()
        self.label_gradient.raise_()
        # self.label_4.raise_()
        self.label_11.raise_()
        self.label_6.raise_()
        self.label_12.raise_()
        self.label_7.raise_()
        self.label_8.raise_()
        self.label_10.raise_()
        self.label_5.raise_()
        self.label.raise_()
        self.label_33.raise_()
        self.label_13.raise_()
        self.btnCloseMainForm.raise_()
        self.label_34.raise_()
        self.lePassword.raise_()
        self.lbl.raise_()
        self.leLogin.raise_()
        self.btnInput.raise_()
        self.lbl_2.raise_()
        self.lbl_3.raise_()

        self.retranslateUi()
        QtCore.QMetaObject.connectSlotsByName(self)
        # self.runThJournal()
        self.firstStart()

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("panel_autologin", "Панель авторизации"))
        # self.leLogin.setText(_translate("panel_autologin", "Логин"))
        # self.lePassword.setText(_translate("panel_autologin", "Пароль"))
        self.lbl.setText(_translate("panel_autologin", "АВТОРИЗАЦИЯ"))
        self.btnInput.setText(_translate("panel_autologin", "ВХОД"))
        self.label_33.setText(_translate("panel_autologin", "         Вход в систему"))
        # self.label_setscreen.setText(_translate("panel_autologin", "TextLabel"))
        # self.label_gradient.setText(_translate("panel_autologin", "TextLabel"))
        # self.label_4.setText(_translate("panel_autologin", "TextLabel"))
        self.lbl_2.setText(_translate("panel_autologin", "АСМК-ГРУНТ"))
        self.lbl_3.setText(_translate("panel_autologin", "Автоматизированная система мониторинга\n"
                                                         "и контроля транспортировки отходов строительства"))

    # def screenImgDisplay(self):
    #
    #     pathImg = globalValues.curDisk + '/Sinaps/screenShot.png'
    #
    #     try:
    #         # if(os.path.exists(pathImg)):
    #         #     os.remove(pathImg)
    #         # print('Start')
    #         # global app
    #         # QScreen.grabWindow(app.primaryScreen(), QApplication.desktop().winId()).save(pathImg, 'png')
    #         print('End')
    #         # screen = pyautogui.screenshot(pathImg)
    #         # subprocess.call(['attrib', '+H', pathImg])
    #     except Exception as ex:
    #         print(ex)
    #
    #         globalValues.writeLogData('Функция создания скриншота рабочего стола', str(ex))

    def firstStart(self):

        pathImg = globalValues.curDisk + '/Sinaps/screenShot.png'
        # print(pathImg)
        # pathImg = pathImg.replace('\\', '/')

        self.lstLight = [[self, "background-color: rgb(0,0,0);"],
                         [self.leLogin, "background-color: rgb(255,255,255);\n"
                                        "color: rgb(0,0,0);\n"
                                        "border-radius:3px;\n"
                                        "border: 1px solid rgb(150,150,150);"],
                         [self.lePassword, "background-color: rgb(255,255,255);\n"
                                           "color: rgb(0,0,0);\n"
                                           "border-radius:3px;\n"
                                           "border: 1px solid rgb(150,150,150);"],
                         [self.lbl, "background-color: rgb(235,235,235);\n"
                                    "color: black;"],
                         [self.btnInput, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                         "color: rgb(0,0,0);\n"
                                         "border-radius:3px;\n"
                                         "border:1px solid rgb(135,135,135);}\n"
                                         "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                         "color: rgb(255, 255, 255);\n"
                                         "border-radius:3px;\n"
                                         "border:1px solid rgb(63,63,63);}\n"
                                         "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                         "color: rgb(255, 255, 255);\n"
                                         "border-radius:3px;\n"
                                         "border:1px solid rgb(63,63,63);};"],
                         [self.label_33, "background-color: rgb(255,255,255);\n"
                                         "border: 1px solid rgb(255,255,255);\n"
                                         "border-bottom-color: rgb(205,205,205);\n"
                                         "border-top-color: rgb(150,150,150);\n"
                                         "border-left-color: rgb(150,150,150);\n"
                                         "border-right-color: rgb(150,150,150);"],
                         [self.label_34, "background-color: rgb(255,255,255);\n"
                                         "image: url(" + globalValues.pathStyleImgs + "sinaps1717.png);"],
                         [self.btnCloseMainForm, "QPushButton:!hover{ background-color: rgb(255,255,255);\n"
                                                 "border: 1px solid rgb(205,205,205);\n"
                                                 "border-top-color: rgb(150, 150, 150);\n"
                                                 "border-left-color: 1px rgb(255,255,255);\n"
                                                 "border-right-color: rgb(150,150,150);\n"
                                                 "border-radius: 0px;\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconextblack.png);}\n"
                                                                                              "QPushButton:hover { background-color: rgb(84,122,181);\n"
                                                                                              "border: 1px solid rgb(205,205,205);\n"
                                                                                              "border-top-color: rgb(150, 150, 150);\n"
                                                                                              "border-left-color: 1px rgb(255,255,255);\n"
                                                                                              "border-right-color: rgb(150,150,150);\n"
                                                                                              "border-radius: 0px;\n"
                                                                                              "image: url(" + globalValues.pathStyleImgs + "iconextwhite.png);}\n"
                                                                                                                                           "QPushButton:hover:pressed { background-color: rgb(50,75,115);\n"
                                                                                                                                           "border: 1px solid rgb(205,205,205);\n"
                                                                                                                                           "border-top-color: rgb(150, 150, 150);\n"
                                                                                                                                           "border-left-color: 1px rgb(255,255,255);\n"
                                                                                                                                           "border-right-color: rgb(150,150,150);\n"
                                                                                                                                           "border-radius: 0px;\n"
                                                                                                                                           "image: url(" + globalValues.pathStyleImgs + "iconextwhite.png);}"],
                         [self.label, "background-color: rgb(242,242,242);\n"
                                      "border: 1px solid rgb(242,242,242);\n"
                                      "border-bottom-color: rgb(150,150,150);\n"
                                      "border-left-color: rgb(150,150,150);\n"
                                      "border-right-color: rgb(150,150,150);"],
                         [self.label_setscreen, "image: url(" + pathImg + ");"],
                         [self.label_gradient,
                          "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:1, stop:0 rgba(205,205,205, 110), stop:1 rgba(205,205,205, 110));"],
                         [self.label_5,
                          "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:1, y2:1, stop:0.146453 rgba(242,242,242, 255), stop:1 rgba(0, 0, 0, 0));"],
                         [self.label_6,
                          "background-color: qlineargradient(spread:pad, x1:1, y1:0.511, x2:0, y2:0.5, stop:0.146453 rgba(242,242,242, 255), stop:1 rgba(0, 0, 0, 0))"],
                         [self.label_7,
                          "background-color: qlineargradient(spread:pad, x1:1, y1:1, x2:1, y2:0, stop:0.146453 rgba(242,242,242, 255), stop:1 rgba(0, 0, 0, 0));"],
                         [self.label_8,
                          "background-color: qradialgradient(spread:pad, cx:0, cy:1, radius:0.5, fx:0, fy:1, stop:0.146453 rgba(242,242,242, 255), stop:1 rgba(0, 0, 0, 0));"],
                         [self.label_10,
                          "background-color: qradialgradient(spread:pad, cx:0, cy:0, radius:0.5, fx:0, fy:0, stop:0.146453 rgba(242,242,242, 255), stop:1 rgba(0, 0, 0, 0));"],
                         [self.label_11,
                          "background-color: qradialgradient(spread:pad, cx:1, cy:1, radius:0.5, fx:1, fy:1, stop:0.146453 rgba(242,242,242, 255), stop:1 rgba(0, 0, 0, 0));"],
                         [self.label_12,
                          "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:1, stop:0.146453 rgba(242,242,242, 255), stop:1 rgba(0, 0, 0, 0));"],
                         [self.label_13,
                          "background-color: qradialgradient(spread:pad, cx:1, cy:0, radius:0.5, fx:1, fy:0, stop:0.146453 rgba(242,242,242, 255), stop:1 rgba(0, 0, 0, 0));"],
                         [self.lbl_2, "background-color: rgb(242,242,242);\n"
                                      "color: rgb(0,0,0);\n"
                                      "border-radius: 5px;"],
                         [self.lbl_3, "background-color: rgb(242,242,242);\n"
                                      "color: rgb(0,0,0);\n"
                                      "border-radius: 5px;"]]
        self.lstDark = [[self, "background-color: rgb(0,0,0);"],
                        [self.leLogin, "background-color: rgb(42, 42, 42);\n"
                                       "color: rgb(255, 255, 255);\n"
                                       "border-radius: 3px;"],
                        [self.lePassword, "background-color: rgb(42, 42, 42);\n"
                                          "color: rgb(255, 255, 255);\n"
                                          "border-radius: 3px;"],
                        [self.lbl, "background-color: rgb(62,62,62);\n"
                                   "color: rgb(255,255,255);\n"
                                   "border-radius: 5px;"],
                        [self.btnInput, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                        "color: rgb(255, 255, 255);\n"
                                        "border-radius:3px;\n"
                                        "border:1px solid rgb(63,63,63);}\n"
                                        "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                        "color: rgb(255, 255, 255);\n"
                                        "border-radius:3px;\n"
                                        "border:1px solid rgb(63,63,63);}\n"
                                        "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                        "color: rgb(255, 255, 255);\n"
                                        "border-radius:3px;\n"
                                        "border:1px solid rgb(63,63,63);};"],
                        [self.label_33, "background-color: rgb(220,220,220);"],
                        [self.label_34, "background-color: rgb(220,220,220);\n"
                                        "image: url(" + globalValues.pathStyleImgs + "sinaps1717.png);"],
                        [self.btnCloseMainForm, "QPushButton:!hover{ background-color: rgb(220,220,220);\n"
                                                "border: 0px solid rgb(220,220,220);\n"
                                                "border-radius: 0px;\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconextblack.png);}\n"
                                                                                             "QPushButton:hover { background-color: rgb(84,122,181);\n"
                                                                                             "border: 0px solid rgb(84,122,181);\n"
                                                                                             "border-radius: 0px;\n"
                                                                                             "image: url(" + globalValues.pathStyleImgs + "iconextwhite.png);}\n"
                                                                                                                                          "QPushButton:hover:pressed { background-color: rgb(50,75,115);\n"
                                                                                                                                          "border: 0px solid rgb(50,75,115);\n"
                                                                                                                                          "border-radius: 0px;\n"
                                                                                                                                          "image: url(" + globalValues.pathStyleImgs + "iconextwhite.png);}"],
                        [self.label, "background-color: rgb(62,62,62);"],
                        [self.label_setscreen, "image: url(" + pathImg + ");"],
                        [self.label_gradient,
                         "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:1, stop:0 rgba(0, 0, 0, 211), stop:1 rgba(0, 0, 0, 210));"],
                        [self.label_5,
                         "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:1, y2:1, stop:0.146453 rgba(0, 0, 0, 255), stop:1 rgba(0, 0, 0, 0));"],
                        [self.label_6,
                         "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:0, y2:0, stop:0.146453 rgba(0, 0, 0, 255), stop:1 rgba(0, 0, 0, 0));"],
                        [self.label_7,
                         "background-color: qlineargradient(spread:pad, x1:1, y1:1, x2:1, y2:0, stop:0.146453 rgba(0, 0, 0, 255), stop:1 rgba(0, 0, 0, 0));"],
                        [self.label_8,
                         "background-color: qradialgradient(spread:pad, cx:0, cy:1, radius:0.5, fx:0, fy:1, stop:0.272727 rgba(0, 0, 0, 255), stop:1 rgba(0, 0, 0, 0));"],
                        [self.label_10,
                         "background-color: qradialgradient(spread:pad, cx:0, cy:0, radius:0.5, fx:0, fy:0, stop:0.146453 rgba(0, 0, 0, 255), stop:1 rgba(0, 0, 0, 0));"],
                        [self.label_11,
                         "background-color: qradialgradient(spread:pad, cx:1, cy:1, radius:0.5, fx:1, fy:1, stop:0.146453 rgba(0, 0, 0, 255), stop:1 rgba(0, 0, 0, 0));"],
                        [self.label_12,
                         "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:0, y2:1, stop:0.146453 rgba(0, 0, 0, 255), stop:1 rgba(0, 0, 0, 0));"],
                        [self.label_13,
                         "background-color: qradialgradient(spread:pad, cx:1, cy:0, radius:0.5, fx:1, fy:0, stop:0.146453 rgba(0, 0, 0, 255), stop:1 rgba(0, 0, 0, 0));"],
                        [self.lbl_2, "background-color: rgb(62,62,62);\n"
                                     "color: rgb(255,255,255);\n"
                                     "border-radius: 5px;"],
                        [self.lbl_3, "background-color: rgb(62,62,62);\n"
                                     "color: rgb(255,255,255);\n"
                                     "border-radius: 5px;"]]

        self.lengthLight = len(self.lstLight)
        self.lengthDark = len(self.lstDark)

        self.changeCOLORMainPanelGrunt()

        self.btnInput.clicked.connect(self.checkAccount)
        self.leLogin.mousePressEvent = self.clearTextLogin
        self.lePassword.mousePressEvent = self.clearTextPassword
        self.btnInput.setObjectName("butInput")
        self.leLogin.setPlaceholderText('Введите логин')
        self.lePassword.setPlaceholderText('Введите пароль')
        self.lePassword.setEchoMode(QtWidgets.QLineEdit.Password)
        self.btnCloseMainForm.clicked.connect(self.closeForm)
        # self.firstDecryptedDataInXls()
        pathFolder = str(os.getenv('APPDATA')) + r'\Sinaps'
        self.checkFolderPath(pathFolder, True)

    def changeColor(self, object, str):
            object.setStyleSheet(str)

    def changeCOLORMainPanelGrunt(self):

            delta = int((1/self.lengthDark)*1000)


            if (globalValues.colorForm == 1):

                    for i in range(self.lengthLight):
                        startTime = round(time.time() * 1000)
                        while True:
                                    obj = self.lstLight[i][0]
                                    style = self.lstLight[i][1]
                                    self.changeColor(obj, style)
                                    # print('changeLight!')
                                    if (abs(round(time.time()*1000) - startTime) > delta):
                                            break

            elif (globalValues.colorForm == 0):

                    for i in range(self.lengthDark):
                        startTime = round(time.time() * 1000)
                        while True:
                                    # print(delta)
                                    # print(abs(round(time.time()*1000) - startTime))
                                    obj = self.lstDark[i][0]
                                    style = self.lstDark[i][1]
                                    self.changeColor(obj, style)
                                    # print('changeDark!')
                                    if (abs(round(time.time()*1000) - startTime) > delta):
                                            break

    def decrypt(self, message):
        private_key =  """-----BEGIN RSA PRIVATE KEY-----
    MIIEowIBAAKCAQEA0c45dZkOTBkAi/eP4i6Grc1GVPVw1a+dni+ciOqACFISHEls
    PHGUIBIpzAxVUt7Kgc3QPZ9amMewFTz+whoe7HjAf+wtC+ISBDw8s7CQxknbeUAX
    NYeShbtP1l8Mz6sHJmSk97qFMdyICOALIKY6ky0rYxG7AsqCbsVBe2MKGqKKGuho
    c4Ei35vNBVPn3KLNmSqjfKk4QwNmttwP6gEqldZwT6pjz2wgA42CMQYkVEQ4w4Bw
    a0Go/C1AsrwLd7q3tgqM+tAEctICAjNAdTKMhFWMOdUn60yaEFcZPo7EE0dUw5ZY
    Up91fqQBDoxt2D3p0iWmTEKZPgx9Tbtu+Z6/JwIDAQABAoIBACe6f1Lvaq+qRFo8
    xLg1yzb6GglYeMdd++DKbz/V9+ybbeaBWMeRUlVIWzXSWA3bNkmiKX6hwEwR9Bvx
    cuRageSRcRJILLeFVZgLuArmsmN59N9e7YYrZ+l+8L1NPmXMowv4HuzyGuq4MeJM
    Wo8SKyFXelHGN71tj4lePOoadP1Z1eS9MvScAfs0Noek71CiBpdWG9dmPksRSNOl
    y37W3MypfdD/9na+sXoIxCZGoLudY/KheSBZ++m/pZZave/g4xrHFfKRR7xtGNI7
    CKCJNaijqnJK62EEKwv4SWZIarfRBjhTDcbXfz+Kz1PT6NSvzoG9/7YpXFkhCRio
    mEZmNFECgYEA1D5gCtvgz5nXKcMVBRkPDgU4Kv9EPXZwiut0zW6XNZpaF6neXjXm
    Rx633aHRWPzhZLX0jrgkNlCDgmxL9HiVu/erGql9bpwraZi4J7nft9voifKnaXva
    EsOLEGE3k5oVX3xkgkgfqRLEGFppSIqJkdv8CgX+oqJzq8euje/B2lUCgYEA/Q8s
    fhDHQn9GxJjwEKaWR/IU0EDzfOWVit7IBnCjb3v+1RwrAbT55hTFLW3FoILzcX1I
    Tr7kQpyScly1ucjALBAEQOYciEVeFLvAGhoiuJCiahTmQI+4QGrjHD5hhjNkvqrN
    3kDZh986aLYNBLK61mx6Ml1Dyas0b3ZBVrlxZ4sCgYAeTY2O31fYrCFRQB4nLS2+
    FbawROPsVpW47+csUYbbS19jk4hBMTbgnp0n0qu+JdTUeToiil35N0OfgnDRxcmz
    HahbVSmoejmkiP56BYrQiGBKGdAXOmynUy3ut8Kkm1JD4NHE3CFRFXHT/Eyd49HC
    doMktzhk5gbX1tmwQDQQRQKBgDDIrhEXdvJQyvm3agArvSjdeDm1a7sWH0AINpNX
    P4qMYtH+fiP0GYDLXD+nu8N3uyqTtk7H6gUVXf4B9V59Xt6fr9I7CiETDlH858mg
    ZDUkXMsKgGDN0/1HHcUiGXbfjXpcPxerdMQGuqHZBqVzNyWDAAOZiynjgVZDe9EW
    KtCFAoGBAM87MQ8T3pORPACj2nUojbjypT0RskqU0GHmqKFkoo4zqY0IPM/dN8Ac
    DB3iPC4i90SWvJyqLd1MV3xm7+uQ9l/p3Jvo7/ClVlPRKZYHbPieBZABMgc+VQ1k
    xNEm2eKuM7oat7/vTmGsCFuMwd9+4Z7HrfGdoS6ilJRvX4jAzimw
    -----END RSA PRIVATE KEY-----"""

        rsa_key = RSA.importKey(private_key)
        rsa_key = PKCS1_OAEP.new(rsa_key)
        raw_cipher_data = b64decode(message)
        decrypted = rsa_key.decrypt(raw_cipher_data)
        return decrypted.decode('utf8')

    def checkAccount(self):

            print('1234567')
            global check_login
            global check_password
            self.check_login = False
            self.check_password = False

            # pathXls = globalValues.pathFileXls

            # subprocess.call(['attrib', '+H', pathXls])

            try:
                # xlApp = win32com.client.Dispatch("Excel.Application")
                # filename,password = pathXls, '34ubitav'
                # xlwb = xlApp.Workbooks.Open(filename, False, True, None, password)
                # xlApp.Visible = False
                # sheet = xlwb.Worksheets('Лист1')

                pathXls = globalValues.pathFileXls
                print(pathXls)

                book = openpyxl.load_workbook(pathXls)
                sheets = book.sheetnames
                print('qwer')
                for sheet in sheets:
                    print(str(sheet))
                    print('qwerty!!!')
                    if (str(sheet) == 'Лист1'):
                        sheet_main = book[sheet]


                        # valLoginCode = 'zidWCmCnRKrL8IJkdhDJri/tSFj3rCyQ/KX7qN8vTbS0c/reiy4rPQ7Ppg/zpI25XM35gLQc7d8aYPclLEapcx7AeG9xtV3O9yWF4pkMaWx7iuT7XQNezYka7pGPnOchf56spATKZJKlEFhAdyExUeQYVwSwEyNqdFvNYiZ+Z4XONBOsOv7VmSxYllhdqrt3ZItoMj47qJ1iE1rylsz+Ccjxm470pfvdl0pVldnCR4mv6s3Q6USBI8eL/Y7L/uumzul+xpqi3EAAwgIwD8QOjtbNFA+oiH97ieBvhfyysXM155gwc9YCrOAAhKZ3xzOpAy2G+VxUtW7gqP4AvT+9Cw=='
                    # print(len(valLoginCode))
                    # print(valLoginCode)
                        for i_0 in range(100):
                            valLoginCode = sheet_main.cell(i_0+1,1).value
                            # print(len(valLoginCode))
                            # print(valLoginCode)
                            if (valLoginCode == None):
                                break
                            # print('check')
                            login_decode = self.decrypt(valLoginCode)
                            print('Log: ' + str(login_decode))
                            if (self.leLogin.text() == str(login_decode)):
                                self.check_login = True
                                valPass = sheet_main.cell(i_0+1,2).value
                                # print(valPass)
                                if (valPass == None):
                                    break
                                pass_decode = self.decrypt(valPass)
                                # print(pass_decode)
                                print('Pwd: ' + str(pass_decode))
                                if (self.lePassword.text() == str(pass_decode)):
                                    self.check_password = True
                                    # print('goodChecking!!!')
                                    break

                # xlwb.Close(False)
                #
                # xlApp.Quit

            except Exception as ex:
                globalValues.writeLogData('Функция проверки логина', str(ex))

            # self.leLogin, "background-color: rgb(255,255,255);\n"
            # "color: rgb(0,0,0);\n"
            # "border-radius:3px;\n"
            # "border: 1px solid rgb(150,150,150);"],
            # [self.lePassword, "background-color: rgb(255,255,255);\n"
            # "color: rgb(0,0,0);\n"
            # "border-radius:3px;\n"
            # "border: 1px solid rgb(150,150,150);"],

            #Dark

            # self.leLogin, "background-color: rgb(42, 42, 42);\n"
            #               "color: rgb(255, 255, 255);\n"
            #               "border-radius: 3px;"],
            # [self.lePassword, "background-color: rgb(42, 42, 42);\n"
            #                   "color: rgb(255, 255, 255);\n"
            #                   "border-radius: 3px;"],


            if (self.check_login == False):
                if (globalValues.colorForm == 1):
                    self.leLogin.setStyleSheet("background-color: rgb(255,255,255);\n"
            "color: rgb(255,0,0);\n"
            "border-radius:3px;\n"
            "border: 1px solid rgb(150,150,150);")
                else:
                    self.leLogin.setStyleSheet("background-color: rgb(42, 42, 42);\n"
                          "color: rgb(255, 0, 0);\n"
                          "border-radius: 3px;")
                self.leLogin.setText('Неверный логин')

            else:
                if (globalValues.colorForm == 1):
                    self.leLogin.setStyleSheet("background-color: rgb(255,255,255);\n"
                                               "color: rgb(0,255,0);\n"
                                               "border-radius:3px;\n"
                                               "border: 1px solid rgb(150,150,150);")
                else:
                    self.leLogin.setStyleSheet("background-color: rgb(42, 42, 42);\n"
                                               "color: rgb(0, 255, 0);\n"
                                               "border-radius: 3px;")

            if (self.check_password == False):
                if (globalValues.colorForm == 1):
                    self.lePassword.setStyleSheet("background-color: rgb(255,255,255);\n"
                                               "color: rgb(255, 0, 0);\n"
                                               "border-radius:3px;\n"
                                               "border: 1px solid rgb(150,150,150);")
                else:
                    self.lePassword.setStyleSheet("background-color: rgb(42, 42, 42);\n"
                                               "color: rgb(255, 0, 0);\n"
                                               "border-radius: 3px;")
                self.lePassword.setEchoMode(QtWidgets.QLineEdit.Normal)
                self.lePassword.setText('Неверный пароль')

            if (self.check_password and self.check_login):
                self.check_good_autologin = True
                self.hide()
                if (self.isChangeLogin == False):

                    # print(globalValues.curUserName)
                    # print('qwewqeqweqwewqe')
                    listEls = [[2, 3], [2, 4], [2, 5], [2, 6], [2, 7], [2, 8], [2, 9], [2, 10], [2, 11], [2, 12], [2, 13], [2, 14], [2, 15], [2, 16], [2, 17], [2, 18]]
                    listElsLogin = [[1, 3], [1, 4], [1, 5], [1, 6], [1, 7], [1, 8], [1, 9], [1, 10], [1, 11], [1, 12], [1, 13], [1, 14], [1, 15], [1, 16], [1, 17], [1, 18]]
                    # print('qqqqqqqqqqqqqqqqqqqqwewewe')
                    globalValues.statusGoodLoginDev = self.decryptDataFromXlsList(listEls)
                    globalValues.loginDataList = self.decryptDataFromXlsListDataLogin(listElsLogin)

                    # print('///////////////////////qwsaqwsaq')
                    print('DataStatus: ', globalValues.statusGoodLoginDev)
                    print('DataLogin: ', globalValues.loginDataList)
                    globalValues.my_sql_name = globalValues.loginDataList[0]
                    globalValues.my_sql_password = globalValues.loginDataList[1]


                else:
                    self.isChangeLogin = False

                globalValues.curUserName = self.leLogin.text()
                # print(globalValues.curUserName)
                self.close()

    def clearTextLogin(self, event):

        # print('ClearLogin!!!')

        if (globalValues.colorForm == 1):
            self.leLogin.setStyleSheet("background-color: rgb(255,255,255);\n"
                                          "color: rgb(0,0,0);\n"
                                          "border-radius:3px;\n"
                                          "border: 1px solid rgb(150,150,150);")
        else:
            self.leLogin.setStyleSheet("background-color: rgb(42, 42, 42);\n"
                                          "color: rgb(255, 255, 255);\n"
                                          "border-radius: 3px;")
        if(self.leLogin.text() == 'Неверный логин'):
            self.leLogin.clear()

    def clearTextPassword(self, event):
        self.lePassword.setEchoMode(QtWidgets.QLineEdit.Password)
        if (globalValues.colorForm == 1):
            self.lePassword.setStyleSheet("background-color: rgb(255,255,255);\n"
                                          "color: rgb(0, 0, 0);\n"
                                          "border-radius:3px;\n"
                                          "border: 1px solid rgb(150,150,150);")
        else:
            self.lePassword.setStyleSheet("background-color: rgb(42, 42, 42);\n"
                                          "color: rgb(255, 255, 255);\n"
                                          "border-radius: 3px;")
        if(self.lePassword.text() == 'Неверный пароль'):
            self.lePassword.clear()

    def closeEvent(self, event):
        event.accept()

    def closeForm(self):
        print('CloseWindow!')
        # self.closeForm()
        self.close()

    def decryptDataFromXlsList(self, listEls):
        try:
            private_key = """-----BEGIN RSA PRIVATE KEY-----
                MIIEowIBAAKCAQEA0c45dZkOTBkAi/eP4i6Grc1GVPVw1a+dni+ciOqACFISHEls
                PHGUIBIpzAxVUt7Kgc3QPZ9amMewFTz+whoe7HjAf+wtC+ISBDw8s7CQxknbeUAX
                NYeShbtP1l8Mz6sHJmSk97qFMdyICOALIKY6ky0rYxG7AsqCbsVBe2MKGqKKGuho
                c4Ei35vNBVPn3KLNmSqjfKk4QwNmttwP6gEqldZwT6pjz2wgA42CMQYkVEQ4w4Bw
                a0Go/C1AsrwLd7q3tgqM+tAEctICAjNAdTKMhFWMOdUn60yaEFcZPo7EE0dUw5ZY
                Up91fqQBDoxt2D3p0iWmTEKZPgx9Tbtu+Z6/JwIDAQABAoIBACe6f1Lvaq+qRFo8
                xLg1yzb6GglYeMdd++DKbz/V9+ybbeaBWMeRUlVIWzXSWA3bNkmiKX6hwEwR9Bvx
                cuRageSRcRJILLeFVZgLuArmsmN59N9e7YYrZ+l+8L1NPmXMowv4HuzyGuq4MeJM
                Wo8SKyFXelHGN71tj4lePOoadP1Z1eS9MvScAfs0Noek71CiBpdWG9dmPksRSNOl
                y37W3MypfdD/9na+sXoIxCZGoLudY/KheSBZ++m/pZZave/g4xrHFfKRR7xtGNI7
                CKCJNaijqnJK62EEKwv4SWZIarfRBjhTDcbXfz+Kz1PT6NSvzoG9/7YpXFkhCRio
                mEZmNFECgYEA1D5gCtvgz5nXKcMVBRkPDgU4Kv9EPXZwiut0zW6XNZpaF6neXjXm
                Rx633aHRWPzhZLX0jrgkNlCDgmxL9HiVu/erGql9bpwraZi4J7nft9voifKnaXva
                EsOLEGE3k5oVX3xkgkgfqRLEGFppSIqJkdv8CgX+oqJzq8euje/B2lUCgYEA/Q8s
                fhDHQn9GxJjwEKaWR/IU0EDzfOWVit7IBnCjb3v+1RwrAbT55hTFLW3FoILzcX1I
                Tr7kQpyScly1ucjALBAEQOYciEVeFLvAGhoiuJCiahTmQI+4QGrjHD5hhjNkvqrN
                3kDZh986aLYNBLK61mx6Ml1Dyas0b3ZBVrlxZ4sCgYAeTY2O31fYrCFRQB4nLS2+
                FbawROPsVpW47+csUYbbS19jk4hBMTbgnp0n0qu+JdTUeToiil35N0OfgnDRxcmz
                HahbVSmoejmkiP56BYrQiGBKGdAXOmynUy3ut8Kkm1JD4NHE3CFRFXHT/Eyd49HC
                doMktzhk5gbX1tmwQDQQRQKBgDDIrhEXdvJQyvm3agArvSjdeDm1a7sWH0AINpNX
                P4qMYtH+fiP0GYDLXD+nu8N3uyqTtk7H6gUVXf4B9V59Xt6fr9I7CiETDlH858mg
                ZDUkXMsKgGDN0/1HHcUiGXbfjXpcPxerdMQGuqHZBqVzNyWDAAOZiynjgVZDe9EW
                KtCFAoGBAM87MQ8T3pORPACj2nUojbjypT0RskqU0GHmqKFkoo4zqY0IPM/dN8Ac
                DB3iPC4i90SWvJyqLd1MV3xm7+uQ9l/p3Jvo7/ClVlPRKZYHbPieBZABMgc+VQ1k
                xNEm2eKuM7oat7/vTmGsCFuMwd9+4Z7HrfGdoS6ilJRvX4jAzimw
                -----END RSA PRIVATE KEY-----"""
            lengtnListEls = len(listEls)
            ListData = []

            # pathFile = str(os.getenv('APPDATA')) + r'\Sinaps\cryptoSSA.xlsx'
            # excel = win32com.client.Dispatch('Excel.Application')
            # workbook = excel.Workbooks.open(pathFile, True, False, None, '34ubitav')
            # sheet = workbook.Worksheets('Лист1')
            # excel.Visible = False
            # excel.DisplayAlerts = False

            pathXls = globalValues.pathFileXls
            print(pathXls)

            book = openpyxl.load_workbook(pathXls)
            sheets = book.sheetnames
            print('qwer')
            for sheet in sheets:
                print(str(sheet))
                print('qwerty!!!')
                if (str(sheet) == 'Лист1'):
                    sheet_main = book[sheet]

                    for i in range(lengtnListEls):
                        dataElement = sheet_main.cell(listEls[i][0], listEls[i][1]).value
                        # print(dataElement)
                        if (str(dataElement) != 'None' and str(dataElement) != ''):
                            rsa_key = RSA.importKey(private_key)
                            rsa_key = PKCS1_OAEP.new(rsa_key)
                            raw_cipher_data = b64decode(dataElement)
                            decrypted = rsa_key.decrypt(raw_cipher_data)
                            strDecrypted = decrypted.decode('utf8')
                            print('CheckingData: ', strDecrypted, listEls[i][0], listEls[i][1])
                            if (strDecrypted == 'TrueDBSSA'):
                                ListData.append(True)
                            else:
                                ListData.append(False)
                        else:
                            ListData.append(False)
            # workbook.Close(False)
            # excel.Quit

            return ListData

            # rsa_key = RSA.importKey(private_key)
            # rsa_key = PKCS1_OAEP.new(rsa_key)
            # raw_cipher_data = b64decode(dataElement)
            # decrypted = rsa_key.decrypt(raw_cipher_data)
            # return decrypted.decode('utf8')

        except Exception as ex:
            globalValues.writeLogData('Функция декодирования данных проверки устройств в файле', str(ex))

    def decryptDataFromXlsListDataLogin(self, listEls):
        try:
            private_key = """-----BEGIN RSA PRIVATE KEY-----
                MIIEowIBAAKCAQEA0c45dZkOTBkAi/eP4i6Grc1GVPVw1a+dni+ciOqACFISHEls
                PHGUIBIpzAxVUt7Kgc3QPZ9amMewFTz+whoe7HjAf+wtC+ISBDw8s7CQxknbeUAX
                NYeShbtP1l8Mz6sHJmSk97qFMdyICOALIKY6ky0rYxG7AsqCbsVBe2MKGqKKGuho
                c4Ei35vNBVPn3KLNmSqjfKk4QwNmttwP6gEqldZwT6pjz2wgA42CMQYkVEQ4w4Bw
                a0Go/C1AsrwLd7q3tgqM+tAEctICAjNAdTKMhFWMOdUn60yaEFcZPo7EE0dUw5ZY
                Up91fqQBDoxt2D3p0iWmTEKZPgx9Tbtu+Z6/JwIDAQABAoIBACe6f1Lvaq+qRFo8
                xLg1yzb6GglYeMdd++DKbz/V9+ybbeaBWMeRUlVIWzXSWA3bNkmiKX6hwEwR9Bvx
                cuRageSRcRJILLeFVZgLuArmsmN59N9e7YYrZ+l+8L1NPmXMowv4HuzyGuq4MeJM
                Wo8SKyFXelHGN71tj4lePOoadP1Z1eS9MvScAfs0Noek71CiBpdWG9dmPksRSNOl
                y37W3MypfdD/9na+sXoIxCZGoLudY/KheSBZ++m/pZZave/g4xrHFfKRR7xtGNI7
                CKCJNaijqnJK62EEKwv4SWZIarfRBjhTDcbXfz+Kz1PT6NSvzoG9/7YpXFkhCRio
                mEZmNFECgYEA1D5gCtvgz5nXKcMVBRkPDgU4Kv9EPXZwiut0zW6XNZpaF6neXjXm
                Rx633aHRWPzhZLX0jrgkNlCDgmxL9HiVu/erGql9bpwraZi4J7nft9voifKnaXva
                EsOLEGE3k5oVX3xkgkgfqRLEGFppSIqJkdv8CgX+oqJzq8euje/B2lUCgYEA/Q8s
                fhDHQn9GxJjwEKaWR/IU0EDzfOWVit7IBnCjb3v+1RwrAbT55hTFLW3FoILzcX1I
                Tr7kQpyScly1ucjALBAEQOYciEVeFLvAGhoiuJCiahTmQI+4QGrjHD5hhjNkvqrN
                3kDZh986aLYNBLK61mx6Ml1Dyas0b3ZBVrlxZ4sCgYAeTY2O31fYrCFRQB4nLS2+
                FbawROPsVpW47+csUYbbS19jk4hBMTbgnp0n0qu+JdTUeToiil35N0OfgnDRxcmz
                HahbVSmoejmkiP56BYrQiGBKGdAXOmynUy3ut8Kkm1JD4NHE3CFRFXHT/Eyd49HC
                doMktzhk5gbX1tmwQDQQRQKBgDDIrhEXdvJQyvm3agArvSjdeDm1a7sWH0AINpNX
                P4qMYtH+fiP0GYDLXD+nu8N3uyqTtk7H6gUVXf4B9V59Xt6fr9I7CiETDlH858mg
                ZDUkXMsKgGDN0/1HHcUiGXbfjXpcPxerdMQGuqHZBqVzNyWDAAOZiynjgVZDe9EW
                KtCFAoGBAM87MQ8T3pORPACj2nUojbjypT0RskqU0GHmqKFkoo4zqY0IPM/dN8Ac
                DB3iPC4i90SWvJyqLd1MV3xm7+uQ9l/p3Jvo7/ClVlPRKZYHbPieBZABMgc+VQ1k
                xNEm2eKuM7oat7/vTmGsCFuMwd9+4Z7HrfGdoS6ilJRvX4jAzimw
                -----END RSA PRIVATE KEY-----"""
            lengtnListEls = len(listEls)
            ListData = []

            # pathFile = str(os.getenv('APPDATA')) + r'\Sinaps\cryptoSSA.xlsx'
            # excel = win32com.client.Dispatch('Excel.Application')
            # workbook = excel.Workbooks.open(pathFile, True, False, None, '34ubitav')
            # sheet = workbook.Worksheets('Лист1')
            # excel.Visible = False
            # excel.DisplayAlerts = False

            pathXls = globalValues.pathFileXls
            print(pathXls)

            book = openpyxl.load_workbook(pathXls)
            sheets = book.sheetnames
            print('qwer')
            for sheet in sheets:
                print(str(sheet))
                print('qwerty!!!')
                if (str(sheet) == 'Лист1'):
                    sheet_main = book[sheet]

                    for i in range(lengtnListEls):
                        dataElement = sheet_main.cell(listEls[i][0], listEls[i][1]).value
                        # print(dataElement)
                        if (str(dataElement) != 'None' and str(dataElement) != ''):
                            rsa_key = RSA.importKey(private_key)
                            rsa_key = PKCS1_OAEP.new(rsa_key)
                            raw_cipher_data = b64decode(dataElement)
                            decrypted = rsa_key.decrypt(raw_cipher_data)
                            textDecrypted = decrypted.decode('utf8')
                            ListData.append(str(textDecrypted))
                        else:
                            ListData.append(str(dataElement))
            # workbook.Close(False)
            # excel.Quit

            return ListData

            # rsa_key = RSA.importKey(private_key)
            # rsa_key = PKCS1_OAEP.new(rsa_key)
            # raw_cipher_data = b64decode(dataElement)
            # decrypted = rsa_key.decrypt(raw_cipher_data)
            # return decrypted.decode('utf8')

        except Exception as ex:
            globalValues.writeLogData('Функция декодирования данных логина и пароля в файле', str(ex))

    def checkFolderPath(self, pathFolder, isHide):
            try:

                    if (os.path.exists(pathFolder) == False):
                        try:
                            os.mkdir(pathFolder)
                            # globalValsSBV.curPathFolderToWrite = pathFolder
                            # if (isHide):
                            #     subprocess.call(['attrib', '+H', pathFolder])
                            return False
                        except Exception as ex:
                            globalValues.writeLogData('Функция проверки и создания папки хранилища', str(ex))
                    else:
                        return True

            except Exception as ex:
                globalValues.writeLogData('Функция проверки ссылки на текущую папку записи', str(ex))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    ui = Ui_panel_autologin()
    ui.show()
    sys.exit(app.exec_())
