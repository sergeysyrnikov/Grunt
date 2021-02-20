from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtGui import QImage, QPixmap, QGuiApplication, QMovie
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QDialog, QLabel, QMessageBox, QWidget, QTableWidgetItem, QAbstractItemView, QDesktopWidget, QHeaderView, QMenu
from threading import Thread
import globalValues
import serial
from time import sleep
import os
import threading
from queue import Queue
import pymysql
import psycopg2
import time
from time import sleep
import serial.tools.list_ports
from setNetworkCam import Ui_NetworkCam
import cv2
from PyQt5.QtGui import QImage
import socket
import datetime
# import ping3
import requests
from base64 import b64decode, b64encode
from Crypto.PublicKey import RSA
from Crypto.Cipher import PKCS1_OAEP
# import win32com.client
# import pythoncom as pt
import subprocess
from PanelArchiveSettings import Ui_Storage
import writeCams as wrtCams
from panAboutSystMenu import Ui_systemMenu
from getDataFull import MeasureVol
import mysql.connector
import openpyxl
import socket

pathImgGray = globalValues.pathImage + 'icongreykrug3131.png'
pathImgRed = globalValues.pathImage + 'iconredkrug3131.png'
pathImgGreen = globalValues.pathImage + 'icongreenkrug3131.png'

pathImgRedCom = globalValues.pathImage + 'iconredkrug4141.png'
pathImgGreenCom = globalValues.pathImage + 'icongreenkrug4141.png'



q = Queue()
Debug = False

# globalValues.colorForm = 1

# class QMovieLabel(QtGui.QLabel):
#     def __init__(self, fileName, parent=None):
#         super(QMovieLabel, self).__init__(parent)
#         m = QtGui.QMovie(fileName)
#         self.setMovie(m)
#         m.start()
#
#     def setMovie(self, movie):
#         super(QMovieLabel, self).setMovie(movie)
#         s=movie.currentImage().size()
#         self._movieWidth = s.width()
#         self._movieHeight = s.height()


globalValues.colorForm = 1

class Ui_SettingsDevice(QDialog):

    con = pymysql.connections

    conn_pg = psycopg2._psycopg.connection

    listRtsp = ['', '', '', '', '', '', '', '', '', '', '', '']
    is_good_cam_0 = False
    is_good_cam_1 = False
    is_good_cam_2 = False
    is_good_cam_3 = False
    if globalValues.debugCamScan:
        is_good_cam_4 = True
        is_good_cam_5 = True

    else:
        is_good_cam_4 = False
        is_good_cam_5 = False

    pathDefCamImg = globalValues.pathStyleImgs + 'camDefault350x350.PNG'
    pathDefCamImgGrey = globalValues.pathStyleImgs + 'camDefault350x350grey.PNG'
    pathDefNotCamScan = globalValues.pathStyleImgs + 'camDefault300x300.PNG'
    pathCamNotCon = globalValues.pathStyleImgs + 'iconcamred.png'
    pathCamCon = globalValues.pathStyleImgs + 'iconcamgreen.png'

    rtsp_cam_out_gard = 'rtsp://192.168.0.88:554/user=admin&password=admin&channel=1&stream=0.sdp'

    closePanel = QtCore.pyqtSignal()

    error_weight_signal = QtCore.pyqtSignal()
    error_traffic_signal = QtCore.pyqtSignal()
    is_good_con_weight = QtCore.pyqtSignal()
    is_good_con_traffic = QtCore.pyqtSignal()

    is_error_weight = False
    is_error_traffic = False

    is_good_start_weight = False
    is_good_start_traffic = False

    checkCamAfterAdd = False
    indexAddCam = 0

    numTab = 0

    listIDSetCam = []
    listNameChannel = []
    listArtTblCam = ['Имя камеры: ', 'IP-адрес: ', 'Маска: ', 'Шлюз: ']
    listChannel = ['Камера въезд КПП', 'Камера выезд КПП', 'Камера въезд Весы', 'Камера выезд Весы', 'Сканер ВК1', 'Сканер ВК2']

    indexCamClick = 99

    num_first_call_mysql = False
    num_first_call_pgsql = False

    checkCamNow = True

    goodLoginDBMySQL = False
    goodLoginDBPgSQL = False
    goodLoginCam = [True, True, True, True, True, True]

    passwdLoginCams = [['',''], ['',''], ['',''], ['',''], ['',''], ['','']]

    lstLight = [[]]
    lstDark = [[]]
    lengthLight = 0
    lengthDark = 0

    checkFirstStart = True

    checkStartDB = False

    refreshTree = False

    def __init__(self):
        super().__init__()
        self.runSetUi()

    def runSetUi(self):
        self.setObjectName("Dialog")
        self.resize(632, 386)
        self.setToolTip("")
        self.advices = QtWidgets.QTabWidget(self)
        self.advices.setGeometry(QtCore.QRect(-4, 0, 641, 391))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.advices.setFont(font)
        self.advices.setToolTipDuration(-1)
        self.advices.setObjectName("advices")
        self.cams = QtWidgets.QWidget()
        self.cams.setObjectName("cams")
        self.lblVideoCam = QtWidgets.QLabel(self.cams)
        self.lblVideoCam.setGeometry(QtCore.QRect(225, 10, 345, 345))
        self.lblVideoCam.setText("")
        self.lblVideoCam.setObjectName("lblVideoCam")
        self.lblPoolCams = QtWidgets.QLabel(self.cams)
        self.lblPoolCams.setGeometry(QtCore.QRect(-2, 10, 211, 33))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.lblPoolCams.setFont(font)
        self.lblPoolCams.setScaledContents(False)
        self.lblPoolCams.setObjectName("lblPoolCams")
        self.label_13 = QtWidgets.QLabel(self.cams)
        self.label_13.setGeometry(QtCore.QRect(6, 12, 25, 25))
        self.label_13.setText("")
        self.label_13.setObjectName("label_13")
        self.line_2 = QtWidgets.QFrame(self.cams)
        self.line_2.setGeometry(QtCore.QRect(10, 42, 190, 3))
        self.line_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_2.setObjectName("line_2")
        self.treeWidget = QtWidgets.QTreeWidget(self.cams)
        self.treeWidget.setGeometry(QtCore.QRect(-2, 15, 211, 341))
        self.treeWidget.setObjectName("treeWidget")
        brush = QtGui.QBrush(QtGui.QColor(0, 0, 0))
        brush.setStyle(QtCore.Qt.SolidPattern)
        self.treeWidget.headerItem().setForeground(0, brush)
        self.label_3 = QtWidgets.QLabel(self.cams)
        self.label_3.setGeometry(QtCore.QRect(205, 32, 3, 16))
        self.label_3.setText("")
        self.label_3.setObjectName("label_3")
        self.lblContCam = QtWidgets.QLabel(self.cams)
        self.lblContCam.setGeometry(QtCore.QRect(585, 10, 61, 191))
        self.lblContCam.setText("")
        self.lblContCam.setObjectName("lblContCam")
        self.btnArchive = QtWidgets.QPushButton(self.cams)
        self.btnArchive.setGeometry(QtCore.QRect(595, 20, 30, 30))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnArchive.setFont(font)
        self.btnArchive.setContextMenuPolicy(QtCore.Qt.DefaultContextMenu)
        self.btnArchive.setToolTipDuration(-1)
        self.btnArchive.setText("")
        self.btnArchive.setObjectName("btnArchive")
        self.btnDelCam = QtWidgets.QPushButton(self.cams)
        self.btnDelCam.setGeometry(QtCore.QRect(595, 90, 30, 30))
        self.btnDelCam.setText("")
        self.btnDelCam.setObjectName("btnDelCam")
        self.btnAddCam = QtWidgets.QPushButton(self.cams)
        self.btnAddCam.setGeometry(QtCore.QRect(595, 55, 30, 30))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnAddCam.setFont(font)
        self.btnAddCam.setContextMenuPolicy(QtCore.Qt.DefaultContextMenu)
        self.btnAddCam.setToolTipDuration(-1)
        self.btnAddCam.setText("")
        self.btnAddCam.setObjectName("btnAddCam")
        self.btnSetCam = QtWidgets.QPushButton(self.cams)
        self.btnSetCam.setGeometry(QtCore.QRect(595, 125, 30, 30))
        self.btnSetCam.setText("")
        self.btnSetCam.setObjectName("btnSetCam")
        self.btnInfo = QtWidgets.QPushButton(self.cams)
        self.btnInfo.setGeometry(QtCore.QRect(595, 160, 30, 30))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnInfo.setFont(font)
        self.btnInfo.setContextMenuPolicy(QtCore.Qt.DefaultContextMenu)
        self.btnInfo.setToolTipDuration(-1)
        self.btnInfo.setText("")
        self.btnInfo.setObjectName("btnInfo")
        self.treeWidget.raise_()
        self.lblVideoCam.raise_()
        self.lblPoolCams.raise_()
        self.label_13.raise_()
        self.line_2.raise_()
        self.label_3.raise_()
        self.lblContCam.raise_()
        self.btnArchive.raise_()
        self.btnDelCam.raise_()
        self.btnAddCam.raise_()
        self.btnSetCam.raise_()
        self.btnInfo.raise_()
        self.advices.addTab(self.cams, "")
        self.weight = QtWidgets.QWidget()
        self.weight.setObjectName("weight")
        self.lcdWeight = QtWidgets.QLCDNumber(self.weight)
        self.lcdWeight.setGeometry(QtCore.QRect(123, 220, 391, 111))
        self.lcdWeight.setObjectName("lcdWeight")
        self.lblSemaLeftInUp = QtWidgets.QLabel(self.weight)
        self.lblSemaLeftInUp.setEnabled(False)
        self.lblSemaLeftInUp.setGeometry(QtCore.QRect(80, 228, 31, 31))
        self.lblSemaLeftInUp.setText("")
        self.lblSemaLeftInUp.setObjectName("lblSemaLeftInUp")
        self.lblSemaLeftOutUp = QtWidgets.QLabel(self.weight)
        self.lblSemaLeftOutUp.setEnabled(False)
        self.lblSemaLeftOutUp.setGeometry(QtCore.QRect(33, 228, 31, 31))
        self.lblSemaLeftOutUp.setText("")
        self.lblSemaLeftOutUp.setObjectName("lblSemaLeftOutUp")
        self.lblSemaLeftOutDown = QtWidgets.QLabel(self.weight)
        self.lblSemaLeftOutDown.setEnabled(False)
        self.lblSemaLeftOutDown.setGeometry(QtCore.QRect(33, 292, 31, 31))
        self.lblSemaLeftOutDown.setText("")
        self.lblSemaLeftOutDown.setObjectName("lblSemaLeftOutDown")
        self.lblSemaLeftInDown = QtWidgets.QLabel(self.weight)
        self.lblSemaLeftInDown.setEnabled(False)
        self.lblSemaLeftInDown.setGeometry(QtCore.QRect(80, 292, 31, 31))
        self.lblSemaLeftInDown.setText("")
        self.lblSemaLeftInDown.setObjectName("lblSemaLeftInDown")
        self.lblSemaRightInDown = QtWidgets.QLabel(self.weight)
        self.lblSemaRightInDown.setEnabled(False)
        self.lblSemaRightInDown.setGeometry(QtCore.QRect(525, 292, 31, 31))
        self.lblSemaRightInDown.setText("")
        self.lblSemaRightInDown.setObjectName("lblSemaRightInDown")
        self.lblSemaRightInUp = QtWidgets.QLabel(self.weight)
        self.lblSemaRightInUp.setEnabled(False)
        self.lblSemaRightInUp.setGeometry(QtCore.QRect(525, 228, 31, 31))
        self.lblSemaRightInUp.setText("")
        self.lblSemaRightInUp.setObjectName("lblSemaRightInUp")
        self.lblSemaRightOutDown = QtWidgets.QLabel(self.weight)
        self.lblSemaRightOutDown.setEnabled(False)
        self.lblSemaRightOutDown.setGeometry(QtCore.QRect(572, 292, 31, 31))
        self.lblSemaRightOutDown.setText("")
        self.lblSemaRightOutDown.setObjectName("lblSemaRightOutDown")
        self.lblSemaRightOutUp = QtWidgets.QLabel(self.weight)
        self.lblSemaRightOutUp.setEnabled(False)
        self.lblSemaRightOutUp.setGeometry(QtCore.QRect(572, 228, 31, 31))
        self.lblSemaRightOutUp.setText("")
        self.lblSemaRightOutUp.setObjectName("lblSemaRightOutUp")
        self.lblStWeightImg = QtWidgets.QLabel(self.weight)
        self.lblStWeightImg.setEnabled(False)
        self.lblStWeightImg.setGeometry(QtCore.QRect(300, 60, 41, 41))
        self.lblStWeightImg.setText("")
        self.lblStWeightImg.setObjectName("lblStWeightImg")
        self.lblStTraffImg = QtWidgets.QLabel(self.weight)
        self.lblStTraffImg.setEnabled(False)
        self.lblStTraffImg.setGeometry(QtCore.QRect(300, 110, 41, 41))
        self.lblStTraffImg.setText("")
        self.lblStTraffImg.setObjectName("lblStTraffImg")
        self.cbWeight = QtWidgets.QComboBox(self.weight)
        self.cbWeight.setGeometry(QtCore.QRect(220, 70, 69, 22))
        self.cbWeight.setObjectName("cbWeight")
        self.cbTraff = QtWidgets.QComboBox(self.weight)
        self.cbTraff.setGeometry(QtCore.QRect(220, 120, 69, 22))
        self.cbTraff.setObjectName("cbTraff")
        self.lblContDataWeight = QtWidgets.QLabel(self.weight)
        self.lblContDataWeight.setGeometry(QtCore.QRect(20, 174, 596, 171))
        self.lblContDataWeight.setText("")
        self.lblContDataWeight.setObjectName("lblContDataWeight")
        self.btnOpenPorts = QtWidgets.QPushButton(self.weight)
        self.btnOpenPorts.setGeometry(QtCore.QRect(360, 90, 191, 30))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.btnOpenPorts.setFont(font)
        self.btnOpenPorts.setObjectName("btnOpenPorts")
        self.lblContComs = QtWidgets.QLabel(self.weight)
        self.lblContComs.setGeometry(QtCore.QRect(20, 15, 596, 150))
        self.lblContComs.setText("")
        self.lblContComs.setObjectName("lblContComs")
        self.lblNameWeight = QtWidgets.QLabel(self.weight)
        self.lblNameWeight.setGeometry(QtCore.QRect(100, 72, 91, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lblNameWeight.setFont(font)
        self.lblNameWeight.setObjectName("lblNameWeight")
        self.lblNameTraff = QtWidgets.QLabel(self.weight)
        self.lblNameTraff.setGeometry(QtCore.QRect(100, 120, 101, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lblNameTraff.setFont(font)
        self.lblNameTraff.setObjectName("lblNameTraff")
        self.lblIconCom = QtWidgets.QLabel(self.weight)
        self.lblIconCom.setGeometry(QtCore.QRect(30, 20, 23, 23))
        self.lblIconCom.setText("")
        self.lblIconCom.setObjectName("lblIconCom")
        self.lblIconDataWeightTr = QtWidgets.QLabel(self.weight)
        self.lblIconDataWeightTr.setGeometry(QtCore.QRect(30, 178, 23, 23))
        self.lblIconDataWeightTr.setText("")
        self.lblIconDataWeightTr.setObjectName("lblIconDataWeightTr")
        self.label_135 = QtWidgets.QLabel(self.weight)
        self.label_135.setGeometry(QtCore.QRect(12, 10, 611, 341))
        self.label_135.setText("")
        self.label_135.setObjectName("label_135")
        self.label_53 = QtWidgets.QLabel(self.weight)
        self.label_53.setGeometry(QtCore.QRect(28, 220, 41, 111))
        self.label_53.setText("")
        self.label_53.setObjectName("label_53")
        self.label_54 = QtWidgets.QLabel(self.weight)
        self.label_54.setGeometry(QtCore.QRect(75, 220, 41, 111))
        self.label_54.setText("")
        self.label_54.setObjectName("label_54")
        self.label_55 = QtWidgets.QLabel(self.weight)
        self.label_55.setGeometry(QtCore.QRect(520, 220, 41, 111))
        self.label_55.setText("")
        self.label_55.setObjectName("label_55")
        self.label_56 = QtWidgets.QLabel(self.weight)
        self.label_56.setGeometry(QtCore.QRect(567, 220, 41, 111))
        self.label_56.setText("")
        self.label_56.setObjectName("label_56")
        self.line_7 = QtWidgets.QFrame(self.weight)
        self.line_7.setGeometry(QtCore.QRect(210, 40, 176, 10))
        self.line_7.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_7.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_7.setObjectName("line_7")
        self.lblNameWeight_2 = QtWidgets.QLabel(self.weight)
        self.lblNameWeight_2.setGeometry(QtCore.QRect(210, 20, 181, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lblNameWeight_2.setFont(font)
        self.lblNameWeight_2.setObjectName("lblNameWeight_2")
        self.line_8 = QtWidgets.QFrame(self.weight)
        self.line_8.setGeometry(QtCore.QRect(190, 200, 228, 10))
        self.line_8.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_8.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_8.setObjectName("line_8")
        self.lblNameWeight_3 = QtWidgets.QLabel(self.weight)
        self.lblNameWeight_3.setGeometry(QtCore.QRect(190, 180, 231, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.lblNameWeight_3.setFont(font)
        self.lblNameWeight_3.setObjectName("lblNameWeight_3")
        self.label_135.raise_()
        self.lblContDataWeight.raise_()
        self.label_55.raise_()
        self.label_56.raise_()
        self.label_54.raise_()
        self.label_53.raise_()
        self.lblContComs.raise_()
        self.lcdWeight.raise_()
        self.lblSemaLeftInUp.raise_()
        self.lblSemaLeftOutUp.raise_()
        self.lblSemaLeftOutDown.raise_()
        self.lblSemaLeftInDown.raise_()
        self.lblSemaRightInDown.raise_()
        self.lblSemaRightInUp.raise_()
        self.lblSemaRightOutDown.raise_()
        self.lblSemaRightOutUp.raise_()
        self.lblStWeightImg.raise_()
        self.lblStTraffImg.raise_()
        self.cbWeight.raise_()
        self.cbTraff.raise_()
        self.btnOpenPorts.raise_()
        self.lblNameWeight.raise_()
        self.lblNameTraff.raise_()
        self.lblIconCom.raise_()
        self.lblIconDataWeightTr.raise_()
        self.line_7.raise_()
        self.lblNameWeight_2.raise_()
        self.line_8.raise_()
        self.lblNameWeight_3.raise_()
        self.advices.addTab(self.weight, "")
        self.shlag = QtWidgets.QWidget()
        self.shlag.setObjectName("shlag")
        self.label = QtWidgets.QLabel(self.shlag)
        self.label.setEnabled(False)
        self.label.setGeometry(QtCore.QRect(32, 63, 52, 15))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.shlag)
        self.label_2.setEnabled(False)
        self.label_2.setGeometry(QtCore.QRect(32, 97, 91, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.label_17 = QtWidgets.QLabel(self.shlag)
        self.label_17.setEnabled(False)
        self.label_17.setGeometry(QtCore.QRect(32, 132, 61, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_17.setFont(font)
        self.label_17.setObjectName("label_17")
        self.label_39 = QtWidgets.QLabel(self.shlag)
        self.label_39.setGeometry(QtCore.QRect(10, 10, 251, 341))
        self.label_39.setText("")
        self.label_39.setObjectName("label_39")
        self.leIpGard = QtWidgets.QLineEdit(self.shlag)
        self.leIpGard.setGeometry(QtCore.QRect(127, 60, 111, 20))
        self.leIpGard.setText("")
        self.leIpGard.setObjectName("leIpGard")
        self.leMaskGard = QtWidgets.QLineEdit(self.shlag)
        self.leMaskGard.setGeometry(QtCore.QRect(127, 95, 111, 20))
        self.leMaskGard.setText("")
        self.leMaskGard.setObjectName("leMaskGard")
        self.leGatewayGard = QtWidgets.QLineEdit(self.shlag)
        self.leGatewayGard.setGeometry(QtCore.QRect(127, 130, 111, 20))
        self.leGatewayGard.setText("")
        self.leGatewayGard.setObjectName("leGatewayGard")
        self.label_40 = QtWidgets.QLabel(self.shlag)
        self.label_40.setGeometry(QtCore.QRect(20, 20, 231, 181))
        self.label_40.setText("")
        self.label_40.setObjectName("label_40")
        self.label_18 = QtWidgets.QLabel(self.shlag)
        self.label_18.setGeometry(QtCore.QRect(70, 22, 145, 16))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_18.setFont(font)
        self.label_18.setObjectName("label_18")
        self.line_3 = QtWidgets.QFrame(self.shlag)
        self.line_3.setGeometry(QtCore.QRect(70, 40, 141, 16))
        self.line_3.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_3.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_3.setObjectName("line_3")
        self.label_41 = QtWidgets.QLabel(self.shlag)
        self.label_41.setGeometry(QtCore.QRect(20, 210, 231, 131))
        self.label_41.setText("")
        self.label_41.setObjectName("label_41")
        self.label_19 = QtWidgets.QLabel(self.shlag)
        self.label_19.setGeometry(QtCore.QRect(50, 212, 171, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_19.setFont(font)
        self.label_19.setObjectName("label_19")
        self.line_4 = QtWidgets.QFrame(self.shlag)
        self.line_4.setGeometry(QtCore.QRect(50, 232, 168, 15))
        self.line_4.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_4.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_4.setObjectName("line_4")
        self.label_42 = QtWidgets.QLabel(self.shlag)
        self.label_42.setGeometry(QtCore.QRect(270, 10, 351, 341))
        self.label_42.setText("")
        self.label_42.setObjectName("label_42")
        self.btnSaveSetGard = QtWidgets.QPushButton(self.shlag)
        self.btnSaveSetGard.setGeometry(QtCore.QRect(30, 165, 91, 25))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btnSaveSetGard.setFont(font)
        self.btnSaveSetGard.setObjectName("btnSaveSetGard")
        self.label_43 = QtWidgets.QLabel(self.shlag)
        self.label_43.setGeometry(QtCore.QRect(300, 16, 160, 30))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(5)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.label_43.sizePolicy().hasHeightForWidth())
        self.label_43.setSizePolicy(sizePolicy)
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_43.setFont(font)
        self.label_43.setObjectName("label_43")
        self.lblGardVideoCam = QtWidgets.QLabel(self.shlag)
        self.lblGardVideoCam.setGeometry(QtCore.QRect(295, 40, 300, 300))
        self.lblGardVideoCam.setText("")
        self.lblGardVideoCam.setObjectName("lblGardVideoCam")
        self.btnCheckConGard = QtWidgets.QPushButton(self.shlag)
        self.btnCheckConGard.setGeometry(QtCore.QRect(146, 165, 91, 25))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btnCheckConGard.setFont(font)
        self.btnCheckConGard.setObjectName("btnCheckConGard")
        self.label_45 = QtWidgets.QLabel(self.shlag)
        self.label_45.setGeometry(QtCore.QRect(30, 260, 91, 61))
        self.label_45.setText("")
        self.label_45.setObjectName("label_45")
        self.label_46 = QtWidgets.QLabel(self.shlag)
        self.label_46.setGeometry(QtCore.QRect(147, 260, 91, 61))
        self.label_46.setText("")
        self.label_46.setObjectName("label_46")
        self.label_20 = QtWidgets.QLabel(self.shlag)
        self.label_20.setGeometry(QtCore.QRect(53, 262, 47, 15))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_20.setFont(font)
        self.label_20.setObjectName("label_20")
        self.label_21 = QtWidgets.QLabel(self.shlag)
        self.label_21.setGeometry(QtCore.QRect(170, 262, 47, 15))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_21.setFont(font)
        self.label_21.setObjectName("label_21")
        self.lblImgGardClose = QtWidgets.QLabel(self.shlag)
        self.lblImgGardClose.setEnabled(False)
        self.lblImgGardClose.setGeometry(QtCore.QRect(178, 285, 31, 31))
        self.lblImgGardClose.setText("")
        self.lblImgGardClose.setObjectName("lblImgGardClose")
        self.lblImgGardOpen = QtWidgets.QLabel(self.shlag)
        self.lblImgGardOpen.setEnabled(False)
        self.lblImgGardOpen.setGeometry(QtCore.QRect(60, 285, 31, 31))
        self.lblImgGardOpen.setText("")
        self.lblImgGardOpen.setObjectName("lblImgGardOpen")
        self.label_42.raise_()
        self.label_39.raise_()
        self.label_40.raise_()
        self.label.raise_()
        self.label_2.raise_()
        self.label_17.raise_()
        self.leIpGard.raise_()
        self.leMaskGard.raise_()
        self.leGatewayGard.raise_()
        self.label_18.raise_()
        self.line_3.raise_()
        self.label_41.raise_()
        self.label_19.raise_()
        self.line_4.raise_()
        self.btnSaveSetGard.raise_()
        self.label_43.raise_()
        self.lblGardVideoCam.raise_()
        self.btnCheckConGard.raise_()
        self.label_45.raise_()
        self.label_46.raise_()
        self.label_20.raise_()
        self.label_21.raise_()
        self.lblImgGardClose.raise_()
        self.lblImgGardOpen.raise_()
        self.advices.addTab(self.shlag, "")
        self.tab_2 = QtWidgets.QWidget()
        self.tab_2.setObjectName("tab_2")
        self.label_69 = QtWidgets.QLabel(self.tab_2)
        self.label_69.setEnabled(False)
        self.label_69.setGeometry(QtCore.QRect(60, 92, 81, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_69.setFont(font)
        self.label_69.setObjectName("label_69")
        self.label_72 = QtWidgets.QLabel(self.tab_2)
        self.label_72.setEnabled(False)
        self.label_72.setGeometry(QtCore.QRect(60, 132, 51, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_72.setFont(font)
        self.label_72.setObjectName("label_72")
        self.label_77 = QtWidgets.QLabel(self.tab_2)
        self.label_77.setEnabled(False)
        self.label_77.setGeometry(QtCore.QRect(60, 172, 51, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_77.setFont(font)
        self.label_77.setObjectName("label_77")
        self.label_78 = QtWidgets.QLabel(self.tab_2)
        self.label_78.setEnabled(False)
        self.label_78.setGeometry(QtCore.QRect(60, 212, 51, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_78.setFont(font)
        self.label_78.setObjectName("label_78")
        self.label_49 = QtWidgets.QLabel(self.tab_2)
        self.label_49.setGeometry(QtCore.QRect(17, 15, 298, 331))
        self.label_49.setText("")
        self.label_49.setObjectName("label_49")
        self.label_50 = QtWidgets.QLabel(self.tab_2)
        self.label_50.setGeometry(QtCore.QRect(325, 15, 300, 331))
        self.label_50.setText("")
        self.label_50.setObjectName("label_50")
        self.btnCheckDBMy = QtWidgets.QPushButton(self.tab_2)
        self.btnCheckDBMy.setGeometry(QtCore.QRect(90, 270, 161, 25))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btnCheckDBMy.setFont(font)
        self.btnCheckDBMy.setObjectName("btnCheckDBMy")
        self.label_51 = QtWidgets.QLabel(self.tab_2)
        self.label_51.setGeometry(QtCore.QRect(100, 22, 145, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_51.setFont(font)
        self.label_51.setObjectName("label_51")
        self.line_5 = QtWidgets.QFrame(self.tab_2)
        self.line_5.setGeometry(QtCore.QRect(100, 44, 146, 10))
        self.line_5.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_5.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_5.setObjectName("line_5")
        self.line_6 = QtWidgets.QFrame(self.tab_2)
        self.line_6.setGeometry(QtCore.QRect(380, 42, 180, 10))
        self.line_6.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_6.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_6.setObjectName("line_6")
        self.label_52 = QtWidgets.QLabel(self.tab_2)
        self.label_52.setGeometry(QtCore.QRect(380, 20, 181, 20))
        font = QtGui.QFont()
        font.setPointSize(12)
        self.label_52.setFont(font)
        self.label_52.setObjectName("label_52")
        self.leDBIpMy = QtWidgets.QLineEdit(self.tab_2)
        self.leDBIpMy.setGeometry(QtCore.QRect(180, 90, 101, 20))
        self.leDBIpMy.setText("")
        self.leDBIpMy.setObjectName("leDBIpMy")
        self.leDBPortMy = QtWidgets.QLineEdit(self.tab_2)
        self.leDBPortMy.setGeometry(QtCore.QRect(180, 130, 101, 20))
        self.leDBPortMy.setText("")
        self.leDBPortMy.setObjectName("leDBPortMy")
        self.leDBPassMy = QtWidgets.QLineEdit(self.tab_2)
        self.leDBPassMy.setGeometry(QtCore.QRect(180, 210, 101, 20))
        self.leDBPassMy.setText("")
        self.leDBPassMy.setObjectName("leDBPassMy")
        self.leDBLoginMy = QtWidgets.QLineEdit(self.tab_2)
        self.leDBLoginMy.setGeometry(QtCore.QRect(180, 170, 101, 20))
        self.leDBLoginMy.setText("")
        self.leDBLoginMy.setObjectName("leDBLoginMy")
        self.leDBPortPg = QtWidgets.QLineEdit(self.tab_2)
        self.leDBPortPg.setGeometry(QtCore.QRect(480, 130, 101, 20))
        self.leDBPortPg.setText("")
        self.leDBPortPg.setObjectName("leDBPortPg")
        self.label_129 = QtWidgets.QLabel(self.tab_2)
        self.label_129.setEnabled(False)
        self.label_129.setGeometry(QtCore.QRect(360, 172, 51, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_129.setFont(font)
        self.label_129.setObjectName("label_129")
        self.label_130 = QtWidgets.QLabel(self.tab_2)
        self.label_130.setEnabled(False)
        self.label_130.setGeometry(QtCore.QRect(360, 92, 81, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_130.setFont(font)
        self.label_130.setObjectName("label_130")
        self.leDBIpPg = QtWidgets.QLineEdit(self.tab_2)
        self.leDBIpPg.setGeometry(QtCore.QRect(480, 90, 101, 20))
        self.leDBIpPg.setText("")
        self.leDBIpPg.setObjectName("leDBIpPg")
        self.label_131 = QtWidgets.QLabel(self.tab_2)
        self.label_131.setEnabled(False)
        self.label_131.setGeometry(QtCore.QRect(360, 212, 51, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_131.setFont(font)
        self.label_131.setObjectName("label_131")
        self.btnCheckDBPg = QtWidgets.QPushButton(self.tab_2)
        self.btnCheckDBPg.setGeometry(QtCore.QRect(390, 270, 161, 25))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btnCheckDBPg.setFont(font)
        self.btnCheckDBPg.setObjectName("btnCheckDBPg")
        self.label_132 = QtWidgets.QLabel(self.tab_2)
        self.label_132.setEnabled(False)
        self.label_132.setGeometry(QtCore.QRect(360, 132, 51, 16))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.label_132.setFont(font)
        self.label_132.setObjectName("label_132")
        self.leDBLoginPg = QtWidgets.QLineEdit(self.tab_2)
        self.leDBLoginPg.setGeometry(QtCore.QRect(480, 170, 101, 20))
        self.leDBLoginPg.setText("")
        self.leDBLoginPg.setObjectName("leDBLoginPg")
        self.leDBPassPg = QtWidgets.QLineEdit(self.tab_2)
        self.leDBPassPg.setGeometry(QtCore.QRect(480, 210, 101, 20))
        self.leDBPassPg.setText("")
        self.leDBPassPg.setObjectName("leDBPassPg")
        self.label_134 = QtWidgets.QLabel(self.tab_2)
        self.label_134.setGeometry(QtCore.QRect(10, 10, 620, 343))
        self.label_134.setText("")
        self.label_134.setObjectName("label_134")
        self.label_134.raise_()
        self.label_49.raise_()
        self.label_50.raise_()
        self.label_69.raise_()
        self.label_72.raise_()
        self.label_77.raise_()
        self.label_78.raise_()
        self.btnCheckDBMy.raise_()
        self.label_51.raise_()
        self.line_5.raise_()
        self.line_6.raise_()
        self.label_52.raise_()
        self.leDBIpMy.raise_()
        self.leDBPortMy.raise_()
        self.leDBPassMy.raise_()
        self.leDBLoginMy.raise_()
        self.leDBPortPg.raise_()
        self.label_129.raise_()
        self.label_130.raise_()
        self.leDBIpPg.raise_()
        self.label_131.raise_()
        self.btnCheckDBPg.raise_()
        self.label_132.raise_()
        self.leDBLoginPg.raise_()
        self.leDBPassPg.raise_()
        self.advices.addTab(self.tab_2, "")
        self.Scaner = QtWidgets.QWidget()
        self.Scaner.setObjectName("Scaner")
        self.lblVideoScan = QtWidgets.QLabel(self.Scaner)
        self.lblVideoScan.setGeometry(QtCore.QRect(170, 30, 300, 300))
        self.lblVideoScan.setText("")
        self.lblVideoScan.setObjectName("lblVideoScan")
        self.btnConnScan = QtWidgets.QPushButton(self.Scaner)
        self.btnConnScan.setGeometry(QtCore.QRect(20, 30, 91, 25))
        font = QtGui.QFont()
        font.setPointSize(10)
        self.btnConnScan.setFont(font)
        self.btnConnScan.setObjectName("btnConnScan")
        self.advices.addTab(self.Scaner, "")
        self.tab = QtWidgets.QWidget()
        self.tab.setObjectName("tab")
        self.lblCamOut = QtWidgets.QLabel(self.tab)
        self.lblCamOut.setEnabled(False)
        self.lblCamOut.setGeometry(QtCore.QRect(230, 40, 41, 41))
        self.lblCamOut.setText("")
        self.lblCamOut.setObjectName("lblCamOut")
        self.lblCamIn = QtWidgets.QLabel(self.tab)
        self.lblCamIn.setEnabled(False)
        self.lblCamIn.setGeometry(QtCore.QRect(230, 100, 41, 41))
        self.lblCamIn.setText("")
        self.lblCamIn.setObjectName("lblCamIn")
        self.lblCamWeight1 = QtWidgets.QLabel(self.tab)
        self.lblCamWeight1.setEnabled(False)
        self.lblCamWeight1.setGeometry(QtCore.QRect(230, 160, 41, 41))
        self.lblCamWeight1.setText("")
        self.lblCamWeight1.setObjectName("lblCamWeight1")
        self.lblCamWeight2 = QtWidgets.QLabel(self.tab)
        self.lblCamWeight2.setEnabled(False)
        self.lblCamWeight2.setGeometry(QtCore.QRect(230, 220, 41, 41))
        self.lblCamWeight2.setText("")
        self.lblCamWeight2.setObjectName("lblCamWeight2")

        self.lblMovieWhite = QtWidgets.QLabel(self)
        self.lblMovieWhite.setGeometry(QtCore.QRect(291, 168, 50, 50))
        self.movieWh = QMovie(globalValues.pathStyleImgs + "animkight3.gif")
        size = QtCore.QSize(50, 50)
        self.movieWh.setScaledSize(size)
        self.lblMovieWhite.setMovie(self.movieWh)
        self.movieWh.start()

        self.label_26 = QtWidgets.QLabel(self.tab)
        self.label_26.setEnabled(False)
        self.label_26.setGeometry(QtCore.QRect(50, 50, 161, 21))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label_26.setFont(font)
        self.label_26.setObjectName("label_26")
        self.label_27 = QtWidgets.QLabel(self.tab)
        self.label_27.setEnabled(False)
        self.label_27.setGeometry(QtCore.QRect(50, 110, 161, 21))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label_27.setFont(font)
        self.label_27.setObjectName("label_27")
        self.label_28 = QtWidgets.QLabel(self.tab)
        self.label_28.setEnabled(False)
        self.label_28.setGeometry(QtCore.QRect(50, 170, 165, 21))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label_28.setFont(font)
        self.label_28.setObjectName("label_28")
        self.label_29 = QtWidgets.QLabel(self.tab)
        self.label_29.setEnabled(False)
        self.label_29.setGeometry(QtCore.QRect(50, 230, 165, 21))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label_29.setFont(font)
        self.label_29.setObjectName("label_29")
        self.lblStDBMy = QtWidgets.QLabel(self.tab)
        self.lblStDBMy.setEnabled(False)
        self.lblStDBMy.setGeometry(QtCore.QRect(540, 185, 41, 41))
        self.lblStDBMy.setText("")
        self.lblStDBMy.setObjectName("lblStDBMy")
        self.label_60 = QtWidgets.QLabel(self.tab)
        self.label_60.setEnabled(False)
        self.label_60.setGeometry(QtCore.QRect(50, 290, 161, 21))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label_60.setFont(font)
        self.label_60.setObjectName("label_60")
        self.lblStWeightAndSema = QtWidgets.QLabel(self.tab)
        self.lblStWeightAndSema.setEnabled(False)
        self.lblStWeightAndSema.setGeometry(QtCore.QRect(540, 125, 41, 41))
        self.lblStWeightAndSema.setText("")
        self.lblStWeightAndSema.setObjectName("lblStWeightAndSema")
        self.label_62 = QtWidgets.QLabel(self.tab)
        self.label_62.setEnabled(False)
        self.label_62.setGeometry(QtCore.QRect(360, 135, 161, 21))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label_62.setFont(font)
        self.label_62.setObjectName("label_62")
        self.label_63 = QtWidgets.QLabel(self.tab)
        self.label_63.setEnabled(False)
        self.label_63.setGeometry(QtCore.QRect(360, 195, 161, 21))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label_63.setFont(font)
        self.label_63.setObjectName("label_63")
        self.lblStGard = QtWidgets.QLabel(self.tab)
        self.lblStGard.setEnabled(False)
        self.lblStGard.setGeometry(QtCore.QRect(230, 280, 41, 41))
        self.lblStGard.setText("")
        self.lblStGard.setObjectName("lblStGard")
        self.lblStDBPg = QtWidgets.QLabel(self.tab)
        self.lblStDBPg.setEnabled(False)
        self.lblStDBPg.setGeometry(QtCore.QRect(540, 245, 41, 41))
        self.lblStDBPg.setText("")
        self.lblStDBPg.setObjectName("lblStDBPg")
        self.label_66 = QtWidgets.QLabel(self.tab)
        self.label_66.setEnabled(False)
        self.label_66.setGeometry(QtCore.QRect(360, 255, 161, 21))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label_66.setFont(font)
        self.label_66.setObjectName("label_66")
        self.label_206 = QtWidgets.QLabel(self.tab)
        self.label_206.setGeometry(QtCore.QRect(15, 15, 298, 331))
        self.label_206.setText("")
        self.label_206.setObjectName("label_206")
        self.label_207 = QtWidgets.QLabel(self.tab)
        self.label_207.setGeometry(QtCore.QRect(10, 10, 620, 343))
        self.label_207.setText("")
        self.label_207.setObjectName("label_207")
        self.label_208 = QtWidgets.QLabel(self.tab)
        self.label_208.setGeometry(QtCore.QRect(325, 15, 300, 331))
        self.label_208.setText("")
        self.label_208.setObjectName("label_208")
        self.line_19 = QtWidgets.QFrame(self.tab)
        self.line_19.setGeometry(QtCore.QRect(50, 82, 221, 16))
        self.line_19.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_19.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_19.setObjectName("line_19")
        self.line_20 = QtWidgets.QFrame(self.tab)
        self.line_20.setGeometry(QtCore.QRect(50, 142, 221, 16))
        self.line_20.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_20.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_20.setObjectName("line_20")
        self.line_21 = QtWidgets.QFrame(self.tab)
        self.line_21.setGeometry(QtCore.QRect(50, 202, 221, 16))
        self.line_21.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_21.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_21.setObjectName("line_21")
        self.line_23 = QtWidgets.QFrame(self.tab)
        self.line_23.setGeometry(QtCore.QRect(360, 167, 221, 16))
        self.line_23.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_23.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_23.setObjectName("line_23")
        self.line_24 = QtWidgets.QFrame(self.tab)
        self.line_24.setGeometry(QtCore.QRect(360, 227, 221, 16))
        self.line_24.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_24.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_24.setObjectName("line_24")
        self.line_25 = QtWidgets.QFrame(self.tab)
        self.line_25.setGeometry(QtCore.QRect(50, 262, 221, 16))
        self.line_25.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_25.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_25.setObjectName("line_25")
        self.label_31 = QtWidgets.QLabel(self.tab)
        self.label_31.setEnabled(False)
        self.label_31.setGeometry(QtCore.QRect(360, 75, 161, 21))
        font = QtGui.QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setWeight(50)
        self.label_31.setFont(font)
        self.label_31.setObjectName("label_31")
        self.lblScan = QtWidgets.QLabel(self.tab)
        self.lblScan.setEnabled(False)
        self.lblScan.setGeometry(QtCore.QRect(540, 65, 41, 41))
        self.lblScan.setText("")
        self.lblScan.setObjectName("lblScan")
        self.line_26 = QtWidgets.QFrame(self.tab)
        self.line_26.setGeometry(QtCore.QRect(360, 107, 221, 16))
        self.line_26.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_26.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_26.setObjectName("line_26")
        self.label_207.raise_()
        self.label_208.raise_()
        self.label_206.raise_()
        self.lblCamOut.raise_()
        self.lblCamIn.raise_()
        self.lblCamWeight1.raise_()
        self.lblCamWeight2.raise_()
        self.label_26.raise_()
        self.label_27.raise_()
        self.label_28.raise_()
        self.label_29.raise_()
        self.lblStDBMy.raise_()
        self.label_60.raise_()
        self.lblStWeightAndSema.raise_()
        self.label_62.raise_()
        self.label_63.raise_()
        self.lblStGard.raise_()
        self.lblStDBPg.raise_()
        self.label_66.raise_()
        self.line_19.raise_()
        self.line_20.raise_()
        self.line_21.raise_()
        self.line_23.raise_()
        self.line_24.raise_()
        self.line_25.raise_()
        self.label_31.raise_()
        self.lblScan.raise_()
        self.line_26.raise_()
        self.advices.addTab(self.tab, "")
        self.lblBack = QtWidgets.QLabel(self)
        self.lblBack.setGeometry(QtCore.QRect(0, 0, 632, 386))
        self.lblBack.setText("")
        self.lblBack.setObjectName("lblBack")
        self.lblBack.raise_()
        self.advices.raise_()

        self.lblMovieWhite.raise_()


        self.retranslateUi()
        QtCore.QMetaObject.connectSlotsByName(self)
        self.firstSettings()

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("Dialog", "Панель системного меню"))

        self.lblPoolCams.setText(_translate("Dialog", "         Список камер"))
        self.treeWidget.headerItem().setText(0, _translate("Dialog", "Cams"))
        self.btnArchive.setToolTip(_translate("Dialog", "Архив"))
        self.btnDelCam.setToolTip(_translate("Dialog", "Удалить"))
        self.btnAddCam.setToolTip(_translate("Dialog", "Добавить"))
        self.btnSetCam.setToolTip(_translate("Dialog", "Параметры"))
        self.btnInfo.setToolTip(_translate("Dialog", "Справка"))
        self.advices.setTabText(self.advices.indexOf(self.cams), _translate("Dialog", "Камеры"))
        self.btnOpenPorts.setText(_translate("Dialog", "Установить соединение"))
        self.lblNameWeight.setText(_translate("Dialog", "Весы :"))
        self.lblNameTraff.setText(_translate("Dialog", "Светофоры : "))
        self.lblIconCom.setToolTip(_translate("Dialog", "Настройки COM-портов"))
        self.lblIconDataWeightTr.setToolTip(_translate("Dialog", "Весы и светофоры"))
        self.lblNameWeight_2.setText(_translate("Dialog", "Настройки COM-портов"))
        self.lblNameWeight_3.setText(_translate("Dialog", "Показания весового комплекса"))
        self.advices.setTabText(self.advices.indexOf(self.weight), _translate("Dialog", "Весы"))
        self.label.setText(_translate("Dialog", "ip адрес:"))
        self.label_2.setText(_translate("Dialog", "маска подсети:"))
        self.label_17.setText(_translate("Dialog", "шлюз:"))
        self.label_18.setText(_translate("Dialog", "Сетевые настройки"))
        self.label_19.setText(_translate("Dialog", "Положение шлагбаума"))
        self.btnSaveSetGard.setToolTip(_translate("Dialog", "сохранить настройки"))
        self.btnSaveSetGard.setText(_translate("Dialog", "Сохранить"))
        self.label_43.setText(_translate("Dialog", " ВИДЕОПОТОК С КАМЕРЫ"))
        self.btnCheckConGard.setToolTip(_translate("Dialog", "проверить соединение"))
        self.btnCheckConGard.setText(_translate("Dialog", "Проверить"))
        self.label_20.setText(_translate("Dialog", "Открыт"))
        self.label_21.setText(_translate("Dialog", "Закрыт"))
        self.advices.setTabText(self.advices.indexOf(self.shlag), _translate("Dialog", "Шлагбаум"))
        self.label_69.setText(_translate("Dialog", "ip адрес БД :"))
        self.label_72.setText(_translate("Dialog", "Порт :"))
        self.label_77.setText(_translate("Dialog", "Логин :"))
        self.label_78.setText(_translate("Dialog", "Пароль :"))
        self.btnCheckDBMy.setToolTip(_translate("Dialog", "сохранить настройки"))
        self.btnCheckDBMy.setText(_translate("Dialog", "Проверить соединение"))
        self.label_51.setText(_translate("Dialog", "База данных MySQL"))
        self.label_52.setText(_translate("Dialog", "База данных PostgreSQL"))
        self.label_129.setText(_translate("Dialog", "Логин :"))
        self.label_130.setText(_translate("Dialog", "ip адрес БД :"))
        self.label_131.setText(_translate("Dialog", "Пароль :"))
        self.btnCheckDBPg.setToolTip(_translate("Dialog", "сохранить настройки"))
        self.btnCheckDBPg.setText(_translate("Dialog", "Проверить соединение"))
        self.label_132.setText(_translate("Dialog", "Порт :"))
        self.advices.setTabText(self.advices.indexOf(self.tab_2), _translate("Dialog", "База данных"))
        self.btnConnScan.setToolTip(_translate("Dialog", "сохранить настройки"))
        self.btnConnScan.setText(_translate("Dialog", "Включить"))
        self.advices.setTabText(self.advices.indexOf(self.Scaner), _translate("Dialog", "Сканер"))
        self.label_26.setText(_translate("Dialog", "КАМЕРА ВЪЕЗД"))
        self.label_27.setText(_translate("Dialog", "КАМЕРА ВЫЕЗД"))
        self.label_28.setText(_translate("Dialog", "КАМЕРА ВЕСЫ ВЪЕЗД"))
        self.label_29.setText(_translate("Dialog", "КАМЕРА ВЕСЫ ВЫЕЗД"))
        self.label_60.setText(_translate("Dialog", "ШЛАГБАУМ"))
        self.label_62.setText(_translate("Dialog", "ВЕСЫ И СВЕТОФОРЫ"))
        self.label_63.setText(_translate("Dialog", "БД MySQL"))
        self.label_66.setText(_translate("Dialog", "БД PostgreSQL"))
        self.label_31.setText(_translate("Dialog", "СКАНЕР"))
        self.advices.setTabText(self.advices.indexOf(self.tab), _translate("Dialog", "Статус"))

    def firstSettings(self):

        self.lblSemaLeftInUp.setStyleSheet("QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
         "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        self.lblSemaLeftOutUp.setStyleSheet("QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
         "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        self.lblSemaLeftOutDown.setStyleSheet("QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
         "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        self.lblSemaLeftInDown.setStyleSheet("QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
         "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        self.lblSemaRightInDown.setStyleSheet("QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
         "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        self.lblSemaRightInUp.setStyleSheet("QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
         "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        self.lblSemaRightOutDown.setStyleSheet("QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
         "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        self.lblSemaRightOutUp.setStyleSheet("QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
         "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        self.lblStWeightImg.setStyleSheet("QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
         "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        self.lblStTraffImg.setStyleSheet("QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
         "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")

        self.lblImgGardClose.setStyleSheet("QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
         "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        self.lblImgGardOpen.setStyleSheet("QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
         "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")

        self.lblCamOut.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                         "image: url(" + globalValues.pathStyleImgs + "icongreykrug4141.png);")
        self.lblCamIn.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                        "image: url(" + globalValues.pathStyleImgs + "icongreykrug4141.png);")
        self.lblCamWeight1.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "image: url(" + globalValues.pathStyleImgs + "icongreykrug4141.png);")
        self.lblCamWeight2.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "image: url(" + globalValues.pathStyleImgs + "icongreykrug4141.png);")

        # self.lblCamScan1.setStyleSheet(
        #     "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
        #     "image: url(" + globalValues.pathStyleImgs + "icongreykrug4141.png);")
        # self.lblCamScan2.setStyleSheet(
        #     "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
        #     "image: url(" + globalValues.pathStyleImgs + "icongreykrug4141.png);")

        self.lblStGard.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                         "image: url(" + globalValues.pathStyleImgs + "icongreykrug4141.png);")
        self.lblStDBPg.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                         "image: url(" + globalValues.pathStyleImgs + "icongreykrug4141.png);")
        self.lblStWeightAndSema.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                         "image: url(" + globalValues.pathStyleImgs + "icongreykrug4141.png);")
        self.lblStDBMy.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                         "image: url(" + globalValues.pathStyleImgs + "icongreykrug4141.png);")
        self.lblScan.setStyleSheet(
            "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
            "image: url(" + globalValues.pathStyleImgs + "icongreykrug4141.png);")


        if globalValues.debugCamScan:
            self.lblCamScan1.setStyleSheet(
                'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                'image: url(' + pathImgGreenCom + ');};')
            self.lblCamScan2.setStyleSheet(
                'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                'image: url(' + pathImgGreenCom + ');};')

        self.lstLight = [[self.advices, "background-color: rgb(252,252,252);"],
                         [self.lblVideoCam, "background-color: rgb(235,235,235);\n"
                                            "border-radius: 10px;\n"
                                            "border: 2px solid rgb(150,150,150);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "camDefault350x350grey.png);"],
                         [self.lblPoolCams, "background-color: rgb(235,235,235);\n"
                                            "border: 1px solid rgb(150,150,150);\n"
                                            "color: rgb(0,0,0);\n"
                                            "border-radius: 5px;\n"
                                            "border-bottom-color: rgb(235,235,235);\n"
                                            "border-left-color: rgb(235,235,235);"],
                         [self.label_13, "background-color: rgb(235,235,235);\n"
                                         "image: url(" + globalValues.pathStyleImgs + "iconworkpanel1.png);"],
                         [self.line_2, "background-color: rgb(235,235,235);"],
                         [self.treeWidget, "QTreeWidget {background-color: rgb(235,235,235);\n"
                                           "border: 1px solid rgb(150,150,150);\n"
                                           "border-radius: 5px;\n"
                                           "color: rgb(0,0,0)}\n"
                                           "QHeaderView::section {\n"
                                           "background-color: rgb(235,235,235);\n"
                                           "color: rgb(89,89,89);\n"
                                           "border: 1px solid rgb(150,150,150);};\n"
                                           "QTreeView::branch:open:has-children:!has-sublings;\n"
                                           "QTreeView::branch:open:has-children:has-sublings {};\n"
                                           ""],
                         [self.label_3, "background-color: rgb(235,235,235);"],
                         [self.lblContCam, "background-color: rgb(235,235,235);\n"
                                           "border-radius: 10px;\n"
                                           "border: 1px solid rgb(150,150,150);"],
                         [self.btnArchive, "QPushButton {background-color: rgb(242,242,242); \n"
                                           "color: rgb(255,255,255);\n"
                                           "border-radius:3px;\n"
                                           "border:1px solid rgb(63,63,63);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "iconarchset2.png);}\n"
                                                                                        "QPushButton:hover {background-color: rgb(242,242,242); \n"
                                                                                        "color: rgb(255,255,255);\n"
                                                                                        "border-radius:3px;\n"
                                                                                        "border:1px solid rgb(84,122,181);\n"
                                                                                        "image: url(" + globalValues.pathStyleImgs + "iconarchset2.png);}\n"
                                                                                                                                     "QPushButton:hover:pressed {background-color: rgb(164,164,164); \n"
                                                                                                                                     "color: rgb(255,255,255);\n"
                                                                                                                                     "border-radius:3px;\n"
                                                                                                                                     "border:1px solid rgb(63,63,63);\n"
                                                                                                                                     "image: url(" + globalValues.pathStyleImgs + "iconarchset2.png);};"],
                         [self.btnDelCam,
                          "QPushButton {background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(223, 0, 0, 255), stop:1 rgba(255, 100, 100, 255));\n"
                          "color: rgb(255,255,255);\n"
                          "border-radius:3px;\n"
                          "border:1px solid rgb(63,63,63);\n"
                          "image: url(" + globalValues.pathStyleImgs + "iconminuscam.png);}\n"
                                                                       "QPushButton:hover {background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(223, 0, 0, 255), stop:1 rgba(255, 100, 100, 255));\n"
                                                                       "color: rgb(255,255,255);\n"
                                                                       "border-radius:3px;\n"
                                                                       "border:1px solid rgb(177, 30, 30);\n"
                                                                       "image: url(" + globalValues.pathStyleImgs + "iconminuscam.png);}\n"
                                                                                                                    "QPushButton:hover:pressed {background-color: rgb(149,26,26);\n"
                                                                                                                    "color: rgb(255,255,255);\n"
                                                                                                                    "border-radius:3px;\n"
                                                                                                                    "border:1px solid rgb(63,63,63);\n"
                                                                                                                    "image: url(" + globalValues.pathStyleImgs + "iconminuscam.png);};"],
                         [self.btnAddCam,
                          "QPushButton {background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(9, 131, 25, 255), stop:1 rgba(9, 185, 32, 255));\n"
                          "color: rgb(255,255,255);\n"
                          "border-radius:3px;\n"
                          "border:1px solid rgb(63,63,63);\n"
                          "image: url(" + globalValues.pathStyleImgs + "iconpluscam3.png);}\n"
                                                                       "QPushButton::hover {background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(9, 131, 25, 255), stop:1 rgba(9, 185, 32, 255));\n"
                                                                       "color: rgb(255,255,255);\n"
                                                                       "border-radius:3px;\n"
                                                                       "border:1px solid rgb(7, 106, 20);\n"
                                                                       "image: url(" + globalValues.pathStyleImgs + "iconpluscam3.png);}\n"
                                                                                                                    "QPushButton:hover:pressed {background-color: rgb(6,81,16);\n"
                                                                                                                    "color: rgb(255, 255, 255);\n"
                                                                                                                    "border:1px solid rgb(66,66,66);\n"
                                                                                                                    "image: url(" + globalValues.pathStyleImgs + "iconpluscam3.png);};"],
                         [self.btnSetCam,
                          "QPushButton {background-color:  qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(89, 89, 89, 255), stop:1 rgba(143, 141, 141, 255));\n"
                          "color: rgb(255,255,255);\n"
                          "border-radius:3px;\n"
                          "border:1px solid rgb(63,63,63);\n"
                          "image: url(" + globalValues.pathStyleImgs + "iconprintset10.png);}\n"
                                                                       "QPushButton:hover {background-color:  qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(89, 89, 89, 255), stop:1 rgba(143, 141, 141, 255));\n"
                                                                       "color: rgb(255,255,255);\n"
                                                                       "border-radius:3px;\n"
                                                                       "border:1px solid rgb(121, 121, 121);\n"
                                                                       "image: url(" + globalValues.pathStyleImgs + "iconprintset10.png);}\n"
                                                                                                                    "QPushButton:hover:pressed {background-color: rgb(85,85,85);\n"
                                                                                                                    "color: rgb(255, 255, 255);\n"
                                                                                                                    "border:1px solid rgb(66,66,66);\n"
                                                                                                                    "image: url(" + globalValues.pathStyleImgs + "iconprintset10.png);};"],
                         [self.btnInfo, "QPushButton {background-color: rgb(52,152,219);\n"
                                        "color: rgb(255,255,255);\n"
                                        "border-radius:3px;\n"
                                        "border:1px solid rgb(63,63,63);\n"
                                        "image: url(" + globalValues.pathStyleImgs + "iconqstn7.png);}\n"
                                                                                     "QPushButton:hover {background-color: rgb(52,152,219);\n"
                                                                                     "color: rgb(255,255,255);\n"
                                                                                     "border-radius:3px;\n"
                                                                                     "border:1px solid rgb(57, 170, 244);\n"
                                                                                     "image: url(" + globalValues.pathStyleImgs + "iconqstn7.png);}\n"
                                                                                                                                  "QPushButton:hover:pressed {background-color: rgb(40,116,167);\n"
                                                                                                                                  "color: rgb(255, 255, 255);\n"
                                                                                                                                  "border:1px solid rgb(66,66,66);\n"
                                                                                                                                  "image: url(" + globalValues.pathStyleImgs + "iconqstn8.png);};"],
                         [self.lcdWeight, "background-color: rgb(227,227,227);\n"
                                          "color: rgb(0,0,0);\n"
                                          "border-radius:3px;\n"
                                          "border:1px solid rgb(135,135,135);"],
                         [self.cbWeight, "background-color: rgb(227,227,227);\n"
                                         "color: black;\n"
                                         "border-radius: 3px;\n"
                                         "border: 1px solid rgb(135,135,135);"],
                         [self.cbTraff, "background-color: rgb(227,227,227);\n"
                                        "color: black;\n"
                                        "border-radius: 3px;\n"
                                        "border: 1px solid rgb(135,135,135);"],
                         [self.lblContDataWeight, "background-color: rgb(235,235,235);\n"
                                                  "border-radius: 10px;\n"
                                                  "border: 1px solid rgb(150,150,150);"],
                         [self.btnOpenPorts, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                             "color: rgb(0,0,0);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(135,135,135);}\n"
                                             "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);}\n"
                                             "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);};"],
                         [self.lblContComs, "background-color: rgb(235,235,235);\n"
                                            "border-radius: 10px;\n"
                                            "border: 1px solid rgb(150,150,150);"],
                         [self.lblNameWeight, "background-color: rgb(235,235,235);\n"
                                              "color: black;"],
                         [self.lblNameTraff, "background-color: rgb(235,235,235);\n"
                                             "color: black;"],
                         [self.lblIconCom, "background-color: rgb(235,235,235);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "iconcomport2323.png);"],
                         [self.lblIconDataWeightTr, "background-color: rgb(235,235,235);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconweightnewgrey2323.png);"],
                         [self.label_135, "background-color: rgb(242,242,242);\n"
                                          "border-radius: 10px;\n"
                                          "border: 1px solid rgb(205,205,205);"],
                         [self.label_53, "background-color: rgb(227,227,227);\n"
                                         "border-radius: 20px;\n"
                                         "border: 2px solid rgb(150,150,150);"],
                         [self.label_54, "background-color: rgb(227,227,227);\n"
                                         "border-radius: 20px;\n"
                                         "border: 2px solid rgb(150,150,150);"],
                         [self.label_55, "background-color: rgb(227,227,227);\n"
                                         "border-radius: 20px;\n"
                                         "border: 2px solid rgb(150,150,150);"],
                         [self.label_56, "background-color: rgb(227,227,227);\n"
                                         "border-radius: 20px;\n"
                                         "border: 2px solid rgb(150,150,150);"],
                         [self.line_7, "background-color: rgb(235,235,235);\n"
                                       "color: black;"],
                         [self.lblNameWeight_2, "background-color: rgb(235,235,235);\n"
                                                "color: black;"],
                         [self.line_8, "background-color: rgb(235,235,235);\n"
                                       "color: black;"],
                         [self.lblNameWeight_3, "background-color: rgb(235,235,235);\n"
                                                "color: black;"],
                         [self.label, "background-color: rgb(235,235,235);\n"
                                      "color: black;"],
                         [self.label_2, "background-color: rgb(235,235,235);\n"
                                        "color: black;"],
                         [self.label_17, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_39, "background-color: rgb(242,242,242);\n"
                                         "border-radius: 10px;\n"
                                         "border: 1px solid rgb(205,205,205);"],
                         [self.leIpGard, "background-color: rgb(255,255,255);\n"
                                         "color: rgb(0,0,0);\n"
                                         "border-radius:3px;\n"
                                         "border: 1px solid rgb(150,150,150);"],
                         [self.leMaskGard, "background-color: rgb(255,255,255);\n"
                                           "color: rgb(0,0,0);\n"
                                           "border-radius:3px;\n"
                                           "border: 1px solid rgb(150,150,150);"],
                         [self.leGatewayGard, "background-color: rgb(255,255,255);\n"
                                              "color: rgb(0,0,0);\n"
                                              "border-radius:3px;\n"
                                              "border: 1px solid rgb(150,150,150);"],
                         [self.label_40, "background-color: rgb(235,235,235);\n"
                                         "border-radius: 10px;\n"
                                         "border: 1px solid rgb(150,150,150);"],
                         [self.label_18, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.line_3, "background-color: rgb(235,235,235);\n"
                                       "color: black;"],
                         [self.label_41, "background-color: rgb(235,235,235);\n"
                                         "border-radius: 10px;\n"
                                         "border: 1px solid rgb(150,150,150);"],
                         [self.label_19, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.line_4, "background-color: rgb(235,235,235);\n"
                                       "color: black;"],
                         [self.label_42, "background-color: rgb(242,242,242);\n"
                                         "border-radius: 10px;\n"
                                         "border: 1px solid rgb(205,205,205);"],
                         [self.btnSaveSetGard, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                               "color: rgb(0,0,0);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(135,135,135);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(63,63,63);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(63,63,63);};"],
                         [self.label_43,
                          "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(252,252,252, 255), stop:1 rgba(150,150,150, 255));\n"
                          "color: rgb(0,0,0);\n"
                          "border-radius: 8px;\n"
                          "border: 2px solid rgb(150,150,150);"],
                         [self.lblGardVideoCam, "background-color: rgb(235,235,235);\n"
                                                "border-radius: 10px;\n"
                                                "border: 2px solid rgb(150,150,150);"],
                         [self.btnCheckConGard, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(135,135,135);}\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(63,63,63);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(63,63,63);};"],
                         [self.label_45, "background-color: rgb(227,227,227);\n"
                                         "border-radius: 10px;\n"
                                         "border:1px solid rgb(135,135,135);"],
                         [self.label_46, "background-color: rgb(227,227,227);\n"
                                         "border-radius: 10px;\n"
                                         "border:1px solid rgb(135,135,135);"],
                         [self.label_20, "background-color: rgb(227,227,227);\n"
                                         "color: black;"],
                         [self.label_21, "background-color: rgb(227,227,227);\n"
                                         "color: black;"],
                         [self.label_69, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_72, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_77, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_78, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_49, "background-color: rgb(235,235,235);\n"
                                         "border-radius: 10px;\n"
                                         "border: 1px solid rgb(150,150,150);"],
                         [self.label_50, "background-color: rgb(235,235,235);\n"
                                         "border-radius: 10px;\n"
                                         "border: 1px solid rgb(150,150,150);"],
                         [self.btnCheckDBMy, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                             "color: rgb(0,0,0);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(135,135,135);}\n"
                                             "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);}\n"
                                             "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);};"],
                         [self.label_51, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.line_5, "background-color: rgb(235,235,235);\n"
                                       "color: black;"],
                         [self.line_6, "background-color: rgb(235,235,235);\n"
                                       "color: black;"],
                         [self.label_52, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.leDBIpMy, "background-color: rgb(255,255,255);\n"
                                         "color: rgb(0,0,0);\n"
                                         "border-radius:3px;\n"
                                         "border: 1px solid rgb(150,150,150);"],
                         [self.leDBPortMy, "background-color: rgb(255,255,255);\n"
                                           "color: rgb(0,0,0);\n"
                                           "border-radius:3px;\n"
                                           "border: 1px solid rgb(150,150,150);"],
                         [self.leDBPassMy, "background-color: rgb(255,255,255);\n"
                                           "color: rgb(0,0,0);\n"
                                           "border-radius:3px;\n"
                                           "border: 1px solid rgb(150,150,150);"],
                         [self.leDBLoginMy, "background-color: rgb(255,255,255);\n"
                                            "color: rgb(0,0,0);\n"
                                            "border-radius:3px;\n"
                                            "border: 1px solid rgb(150,150,150);"],
                         [self.leDBPortPg, "background-color: rgb(255,255,255);\n"
                                           "color: rgb(0,0,0);\n"
                                           "border-radius:3px;\n"
                                           "border: 1px solid rgb(150,150,150);"],
                         [self.label_129, "background-color: rgb(235,235,235);\n"
                                          "color: black;"],
                         [self.label_130, "background-color: rgb(235,235,235);\n"
                                          "color: black;"],
                         [self.leDBIpPg, "background-color: rgb(255,255,255);\n"
                                         "color: rgb(0,0,0);\n"
                                         "border-radius:3px;\n"
                                         "border: 1px solid rgb(150,150,150);"],
                         [self.label_131, "background-color: rgb(235,235,235);\n"
                                          "color: black;"],
                         [self.btnCheckDBPg, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                             "color: rgb(0,0,0);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(135,135,135);}\n"
                                             "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);}\n"
                                             "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);};"],
                         [self.label_132, "background-color: rgb(235,235,235);\n"
                                          "color: black;"],
                         [self.leDBLoginPg, "background-color: rgb(255,255,255);\n"
                                            "color: rgb(0,0,0);\n"
                                            "border-radius:3px;\n"
                                            "border: 1px solid rgb(150,150,150);"],
                         [self.leDBPassPg, "background-color: rgb(255,255,255);\n"
                                           "color: rgb(0,0,0);\n"
                                           "border-radius:3px;\n"
                                           "border: 1px solid rgb(150,150,150);"],
                         [self.label_134, "background-color: rgb(242,242,242);\n"
                                          "border-radius: 10px;\n"
                                          "border: 1px solid rgb(205,205,205);"],
                         [self.lblVideoScan, "background-color: rgb(235,235,235);\n"
                                             "border-radius: 10px;\n"
                                             "border: 2px solid rgb(150,150,150);"],
                         [self.btnConnScan, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                            "color: rgb(0,0,0);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(135,135,135);}\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);}\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);};"],
                         [self.label_26, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_27, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_28, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_29, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_60, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_62, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_63, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_66, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.label_206, "background-color: rgb(235,235,235);\n"
                                          "border-radius: 10px;\n"
                                          "border: 1px solid rgb(150,150,150);"],
                         [self.label_207, "background-color: rgb(242,242,242);\n"
                                          "border-radius: 10px;\n"
                                          "border: 1px solid rgb(205,205,205);"],
                         [self.label_208, "background-color: rgb(235,235,235);\n"
                                          "border-radius: 10px;\n"
                                          "border: 1px solid rgb(150,150,150);"],
                         [self.line_19, "background-color: rgb(235,235,235);\n"
                                        "color: black;"],
                         [self.line_20, "background-color: rgb(235,235,235);\n"
                                        "color: black;"],
                         [self.line_21, "background-color: rgb(235,235,235);\n"
                                        "color: black;"],
                         [self.line_23, "background-color: rgb(235,235,235);\n"
                                        "color: black;"],
                         [self.line_24, "background-color: rgb(235,235,235);\n"
                                        "color: black;"],
                         [self.line_25, "background-color: rgb(235,235,235);\n"
                                        "color: black;"],
                         [self.label_31, "background-color: rgb(235,235,235);\n"
                                         "color: black;"],
                         [self.line_26, "background-color: rgb(235,235,235);\n"
                                        "color: black;"],
                         [self.lblBack, "background-color: rgb(242,242,242);"]]
        self.lstDark = [[self.advices, "background-color: rgb(66,66,66);"],
                        [self.btnDelCam,
                         "QPushButton {background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(149, 26, 26, 255), stop:1 rgba(225, 38, 38, 255));\n"
                         "color: rgb(255,255,255);\n"
                         "border-radius:3px;\n"
                         "border:1px solid rgb(63,63,63);\n"
                         "image: url(" + globalValues.pathStyleImgs + "iconminuscam.png);}\n"
                                                                      "QPushButton:hover {background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(149, 26, 26, 255), stop:1 rgba(225, 38, 38, 255));\n"
                                                                      "color: rgb(255,255,255);\n"
                                                                      "border-radius:3px;\n"
                                                                      "border:1px solid rgb(177, 30, 30);\n"
                                                                      "image: url(" + globalValues.pathStyleImgs + "iconminuscam.png);}\n"
                                                                                                                   "QPushButton:hover:pressed {background-color: rgb(149,26,26);\n"
                                                                                                                   "color: rgb(255,255,255);\n"
                                                                                                                   "border-radius:3px;\n"
                                                                                                                   "border:1px solid rgb(63,63,63);\n"
                                                                                                                   "image: url(" + globalValues.pathStyleImgs + "iconminuscam.png);};"],
                        [self.btnSetCam,
                         "QPushButton {background-color:  qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(89, 89, 89, 255), stop:1 rgba(143, 141, 141, 255));\n"
                         "color: rgb(255,255,255);\n"
                         "border-radius:3px;\n"
                         "border:1px solid rgb(63,63,63);\n"
                         "image: url(" + globalValues.pathStyleImgs + "iconprintset10.png);}\n"
                                                                      "QPushButton:hover {background-color:  qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(89, 89, 89, 255), stop:1 rgba(143, 141, 141, 255));\n"
                                                                      "color: rgb(255,255,255);\n"
                                                                      "border-radius:3px;\n"
                                                                      "border:1px solid rgb(121, 121, 121);\n"
                                                                      "image: url(" + globalValues.pathStyleImgs + "iconprintset10.png);}\n"
                                                                                                                   "QPushButton:hover:pressed {background-color: rgb(85,85,85);\n"
                                                                                                                   "color: rgb(255, 255, 255);\n"
                                                                                                                   "border:1px solid rgb(66,66,66);\n"
                                                                                                                   "image: url(" + globalValues.pathStyleImgs + "iconprintset10.png);};"],
                        [self.btnAddCam,
                         "QPushButton {background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(6, 81, 16, 255), stop:1 rgba(7, 142, 25, 255));\n"
                         "color: rgb(255,255,255);\n"
                         "border-radius:3px;\n"
                         "border:1px solid rgb(63,63,63);\n"
                         "image: url(" + globalValues.pathStyleImgs + "iconpluscam3.png);}\n"
                                                                      "QPushButton::hover {background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(6, 81, 16, 255), stop:1 rgba(7, 142, 25, 255));\n"
                                                                      "color: rgb(255,255,255);\n"
                                                                      "border-radius:3px;\n"
                                                                      "border:1px solid rgb(7, 106, 20);\n"
                                                                      "image: url(" + globalValues.pathStyleImgs + "iconpluscam3.png);}\n"
                                                                                                                   "QPushButton:hover:pressed {background-color: rgb(6,81,16);\n"
                                                                                                                   "color: rgb(255, 255, 255);\n"
                                                                                                                   "border:1px solid rgb(66,66,66);\n"
                                                                                                                   "image: url(" + globalValues.pathStyleImgs + "iconpluscam3.png);};"],
                        [self.lblVideoCam, "background-color: rgb(0,0,0);\n"
                                           "border-radius: 10px;\n"
                                           "border: 2px solid rgb(42,42,42);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "camDefault350x350.png);"],
                        [self.lblContCam, "background-color: rgb(75,75,75);\n"
                                          "border-radius: 10px;"],
                        [self.lblPoolCams, "background-color: rgb(89,89,89);\n"
                                           "color: white;\n"
                                           "border-radius: 5px;\n"
                                           "border: 1px solid rgb(66,66,66); \n"
                                           "border-bottom-color: rgb(89,89,89);"],
                        [self.label_13, "background-color: rgb(89,89,89);\n"
                                        "image: url(" + globalValues.pathStyleImgs + "iconworkpanel1.png);"],
                        [self.line_2, "background-color: rgb(89,89,89);"],
                        [self.treeWidget, "QTreeWidget {background-color: rgb(89,89,89);\n"
                                          "border: 1px solid rgb(66,66,66);\n"
                                          "border-radius: 5px;\n"
                                          "color: rgb(255,255,255)}\n"
                                          "QHeaderView::section {\n"
                                          "background-color: rgb(89,89,89);\n"
                                          "color: rgb(89,89,89);\n"
                                          "border: 1px solid rgb(89,89,89);};\n"
                                          "QTreeView::branch:open:has-children:!has-sublings;\n"
                                          "QTreeView::branch:open:has-children:has-sublings {};\n"
                                          ""],
                        [self.label_3, "background-color: rgb(89,89,89);"],
                        [self.btnInfo, "QPushButton {background-color: rgb(52,152,219);\n"
                                       "color: rgb(255,255,255);\n"
                                       "border-radius:3px;\n"
                                       "border:1px solid rgb(63,63,63);\n"
                                       "image: url(" + globalValues.pathStyleImgs + "iconqstn7.png);}\n"
                                                                                    "QPushButton:hover {background-color: rgb(52,152,219);\n"
                                                                                    "color: rgb(255,255,255);\n"
                                                                                    "border-radius:3px;\n"
                                                                                    "border:1px solid rgb(57, 170, 244);\n"
                                                                                    "image: url(" + globalValues.pathStyleImgs + "iconqstn7.png);}\n"
                                                                                                                                 "QPushButton:hover:pressed {background-color: rgb(40,116,167);\n"
                                                                                                                                 "color: rgb(255, 255, 255);\n"
                                                                                                                                 "border:1px solid rgb(66,66,66);\n"
                                                                                                                                 "image: url(" + globalValues.pathStyleImgs + "iconqstn8.png);};"],
                        [self.btnArchive, "QPushButton {background-color: rgb(227,227,227); \n"
                                          "color: rgb(255,255,255);\n"
                                          "border-radius:3px;\n"
                                          "border:1px solid rgb(63,63,63);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "iconarchset2.png);}\n"
                                                                                       "QPushButton:hover {background-color: rgb(227,227,227); \n"
                                                                                       "color: rgb(255,255,255);\n"
                                                                                       "border-radius:3px;\n"
                                                                                       "border:1px solid rgb(84,122,181);\n"
                                                                                       "image: url(" + globalValues.pathStyleImgs + "iconarchset2.png);}\n"
                                                                                                                                    "QPushButton:hover:pressed {background-color: rgb(164,164,164); \n"
                                                                                                                                    "color: rgb(255,255,255);\n"
                                                                                                                                    "border-radius:3px;\n"
                                                                                                                                    "border:1px solid rgb(63,63,63);\n"
                                                                                                                                    "image: url(" + globalValues.pathStyleImgs + "iconarchset2.png);};"],
                        [self.lcdWeight, "background-color: rgb(89,89,89);\n"
                                         " border-radius: 3px;\n"
                                         " border: 1px solid rgb(63,63,63);"],
                        [self.cbWeight, "background-color: rgb(89,89,89);\n"
                                        "color: white;\n"
                                        "border-radius: 3px;\n"
                                        "border: 1px solid rgb(63,63,63);"],
                        [self.cbTraff, "background-color: rgb(89,89,89);\n"
                                       "color: white;\n"
                                       "border-radius: 3px;\n"
                                       "border: 1px solid rgb(63,63,63);"],
                        [self.lblContDataWeight, "background-color: rgb(75,75,75);\n"
                                                 "border-radius: 10px;"],
                        [self.btnOpenPorts, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);}\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);}\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);};"],
                        [self.lblContComs, "background-color: rgb(75,75,75);\n"
                                           "border-radius: 10px;"],
                        [self.lblNameWeight, "background-color: rgb(75,75,75);\n"
                                             "color: white;"],
                        [self.lblNameTraff, "background-color: rgb(75,75,75);\n"
                                            "color: white;"],
                        [self.lblIconCom, "background-color: rgb(75,75,75);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "iconcomport2323.png);"],
                        [self.lblIconDataWeightTr, "background-color: rgb(75,75,75);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "iconweightnew2323.png);"],
                        [self.label_135, "background-color: rgb(62,62,62);\n"
                                         "border-radius: 10px;"],
                        [self.label_53, "background-color: rgb(89,89,89);\n"
                                        "border-radius: 20px;"],
                        [self.label_54, "background-color: rgb(89,89,89);\n"
                                        "border-radius: 20px;"],
                        [self.label_55, "background-color: rgb(89,89,89);\n"
                                        "border-radius: 20px;"],
                        [self.label_56, "background-color: rgb(89,89,89);\n"
                                        "border-radius: 20px;"],
                        [self.line_7, "background-color: rgb(75,75,75);"],
                        [self.lblNameWeight_2, "background-color: rgb(75,75,75);\n"
                                               "color: white;"],
                        [self.line_8, "background-color: rgb(75,75,75);"],
                        [self.lblNameWeight_3, "background-color: rgb(75,75,75);\n"
                                               "color: white;"],
                        [self.label, "background-color: rgb(75,75,75);\n"
                                     "color: white;"],
                        [self.label_2, "background-color: rgb(75,75,75);\n"
                                       "color: white;"],
                        [self.label_17, "background-color: rgb(75,75,75);\n"
                                        "color: white;"],
                        [self.label_39, "background-color: rgb(62,62,62);\n"
                                        "border-radius: 10px;"],
                        [self.leIpGard, "background-color: white;"],
                        [self.leMaskGard, "background-color: white;"],
                        [self.leGatewayGard, "background-color: white;"],
                        [self.label_40, "background-color: rgb(75,75,75);\n"
                                        "border-radius: 10px;"],
                        [self.label_18, "background-color: rgb(75,75,75);\n"
                                        "color: white;"],
                        [self.line_3, "background-color: rgb(75,75,75);"],
                        [self.label_41, "background-color: rgb(75,75,75);\n"
                                        "border-radius: 10px;"],
                        [self.label_19, "background-color: rgb(75,75,75);\n"
                                        "color: white;"],
                        [self.line_4, "background-color: rgb(75,75,75);"],
                        [self.label_42, "background-color: rgb(62,62,62);\n"
                                        "border-radius: 10px;"],
                        [self.btnSaveSetGard, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);};"],
                        [self.label_43, "\n"
                                        "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 255), stop:1 rgba(62, 62, 62, 255));\n"
                                        "color: rgb(255,255,255);\n"
                                        "border-radius: 8px;\n"
                                        "border: 2px solid rgb(43,43,44);"],
                        [self.lblGardVideoCam, "background-color: rgb(0,0,0);\n"
                                               "border-radius: 10px;\n"
                                               "border: 4px solid rgb(42,42,42);\n"
                                               ""],
                        [self.btnCheckConGard, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(63,63,63);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(63,63,63);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(63,63,63);};"],
                        [self.label_45, "background-color: rgb(89,89,89);\n"
                                        "border-radius: 10px;"],
                        [self.label_46, "background-color: rgb(89,89,89);\n"
                                        "border-radius: 10px;"],
                        [self.label_20, "background-color: rgb(89,89,89);\n"
                                        "color: white;"],
                        [self.label_21, "background-color: rgb(89,89,89);\n"
                                        "color: white;"],
                        [self.label_49, "background-color: rgb(75,75,75);\n"
                                        "border-radius: 10px;"],
                        [self.label_50, "background-color: rgb(75,75,75);\n"
                                        "border-radius: 10px;"],
                        [self.btnCheckDBMy, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);}\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);}\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);};"],
                        [self.label_51, "background-color: rgb(75,75,75);\n"
                                        "color: white;"],
                        [self.line_5, "background-color: rgb(75,75,75);\n"
                                      "color: white;"],
                        [self.line_6, "background-color: rgb(75,75,75);\n"
                                      "color: white;"],
                        [self.label_52, "background-color: rgb(75,75,75);\n"
                                        "color: white;"],
                        [self.leDBIpMy, "background-color: white;\n"
                                        "border-radius: 3px;"],
                        [self.leDBPortPg, "background-color: white;\n"
                                          "border-radius: 3px;"],
                        [self.label_129, "background-color: rgb(75,75,75);\n"
                                         "color: white;"],
                        [self.label_130, "background-color: rgb(75,75,75);\n"
                                         "color: white;"],
                        [self.leDBIpPg, "background-color: white;\n"
                                        "border-radius: 3px;"],
                        [self.label_131, "background-color: rgb(75,75,75);\n"
                                         "color: white;"],
                        [self.btnCheckDBPg, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);}\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);}\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);};"],
                        [self.label_132, "background-color: rgb(75,75,75);\n"
                                         "color: white;"],
                        [self.leDBLoginPg, "background-color: white;\n"
                                           "border-radius: 3px;"],
                        [self.leDBPassPg, "background-color: white;\n"
                                          "border-radius: 3px;"],
                        [self.label_134, "background-color: rgb(62,62,62);\n"
                                         "border-radius: 10px;"],
                        [self.leDBPortMy, "background-color: white;\n"
                                          "border-radius: 3px;"],
                        [self.label_77, "background-color: rgb(75,75,75);\n"
                                        "color: white;"],
                        [self.label_72, "background-color: rgb(75,75,75);\n"
                                        "color: white;"],
                        [self.leDBLoginMy, "background-color: white;\n"
                                           "border-radius: 3px;"],
                        [self.label_78, "background-color: rgb(75,75,75);\n"
                                        "color: white;"],
                        [self.leDBPassMy, "background-color: white;\n"
                                          "border-radius: 3px;"],
                        [self.label_69, "background-color: rgb(75,75,75);\n"
                                        "color: white;"],
                        [self.lblVideoScan, "background-color: rgb(0,0,0);\n"
                                            "border-radius: 10px;\n"
                                            "border: 4px solid rgb(42,42,42);\n"
                                            ""],
                        [self.btnConnScan, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                           "color: rgb(255, 255, 255);\n"
                                           "border-radius:3px;\n"
                                           "border:1px solid rgb(63,63,63);}\n"
                                           "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                           "color: rgb(255, 255, 255);\n"
                                           "border-radius:3px;\n"
                                           "border:1px solid rgb(63,63,63);}\n"
                                           "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                           "color: rgb(255, 255, 255);\n"
                                           "border-radius:3px;\n"
                                           "border:1px solid rgb(63,63,63);};"],
                        [self.label_26, "color: rgb(255,255,255);\n"
                                        "background-color: rgb(75,75,75);\n"
                                        "font: 75 12pt \"MS Shell Dlg 2\";\n"
                                        ""],
                        [self.label_27, "color: rgb(255,255,255);\n"
                                        "background-color: rgb(75,75,75);\n"
                                        "font: 75 12pt \"MS Shell Dlg 2\";\n"
                                        ""],
                        [self.label_28, "color: rgb(255,255,255);\n"
                                        "background-color: rgb(75,75,75);\n"
                                        "font: 75 12pt \"MS Shell Dlg 2\";\n"
                                        ""],
                        [self.label_29, "color: rgb(255,255,255);\n"
                                        "background-color: rgb(75,75,75);\n"
                                        "font: 75 12pt \"MS Shell Dlg 2\";\n"
                                        ""],
                        [self.label_60, "color: rgb(255,255,255);\n"
                                        "background-color: rgb(75,75,75);\n"
                                        "font: 75 12pt \"MS Shell Dlg 2\";\n"
                                        ""],
                        [self.label_62, "color: rgb(255,255,255);\n"
                                        "background-color: rgb(75,75,75);\n"
                                        "font: 75 12pt \"MS Shell Dlg 2\";\n"
                                        ""],
                        [self.label_63, "color: rgb(255,255,255);\n"
                                        "background-color: rgb(75,75,75);\n"
                                        "font: 75 12pt \"MS Shell Dlg 2\";\n"
                                        ""],
                        [self.label_66, "color: rgb(255,255,255);\n"
                                        "background-color: rgb(75,75,75);\n"
                                        "font: 75 12pt \"MS Shell Dlg 2\";\n"
                                        ""],
                        [self.label_206, "background-color: rgb(75,75,75);\n"
                                         "border-radius: 10px;"],
                        [self.label_207, "background-color: rgb(62,62,62);\n"
                                         "border-radius: 10px;"],
                        [self.label_208, "background-color: rgb(75,75,75);\n"
                                         "border-radius: 10px;"],
                        [self.line_19, "background-color: rgb(75,75,75);\n"
                                       "color: white;"],
                        [self.line_20, "background-color: rgb(75,75,75);\n"
                                       "color: white;"],
                        [self.line_21, "background-color: rgb(75,75,75);\n"
                                       "color: white;"],
                        [self.line_23, "background-color: rgb(75,75,75);\n"
                                       "color: white;"],
                        [self.line_24, "background-color: rgb(75,75,75);\n"
                                       "color: white;"],
                        [self.line_25, "background-color: rgb(75,75,75);\n"
                                       "color: white;"],
                        [self.line_26, "background-color: rgb(75,75,75);\n"
                                       "color: white;"],
                        [self.label_31, "color: rgb(255,255,255);\n"
                                        "background-color: rgb(75,75,75);\n"
                                        "font: 75 12pt \"MS Shell Dlg 2\";\n"
                                        ""],
                        [self.lblBack, "background-color: rgb(66,66,66);"]]

        self.lengthLight = len(self.lstLight)
        self.lengthDark = len(self.lstDark)

        # self.changeCOLORMainPanelGrunt()
        # self.lblVideoCam.setPixmap(QtGui.QPixmap(self.pathDefCamImg))

        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(True)
        font.setItalic(False)
        # font.setWeight(9)
        itemInKPP = QtWidgets.QTreeWidgetItem(self.treeWidget)
        itemInKPP.setText(0, 'Камера въезд КПП')
        itemInKPP.setFont(0,font)
        itemInKPP.setIcon(0, QtGui.QIcon(self.pathCamNotCon))
        self.treeWidget.addTopLevelItem(itemInKPP)
        itemOutKPP = QtWidgets.QTreeWidgetItem(self.treeWidget)
        itemOutKPP.setText(0, 'Камера выезд КПП')
        itemOutKPP.setFont(0,font)
        itemOutKPP.setIcon(0, QtGui.QIcon(self.pathCamNotCon))
        self.treeWidget.addTopLevelItem(itemOutKPP)
        itemInWeight = QtWidgets.QTreeWidgetItem(self.treeWidget)
        itemInWeight.setText(0, 'Камера въезд Весы')
        itemInWeight.setFont(0, font)
        itemInWeight.setIcon(0, QtGui.QIcon(self.pathCamNotCon))
        self.treeWidget.addTopLevelItem(itemInWeight)
        itemOutWeight = QtWidgets.QTreeWidgetItem(self.treeWidget)
        itemOutWeight.setText(0, 'Камера выезд Весы')
        itemOutWeight.setFont(0, font)
        itemOutWeight.setIcon(0, QtGui.QIcon(self.pathCamNotCon))
        self.treeWidget.addTopLevelItem(itemOutWeight)

        # itemVolVideo0 = QtWidgets.QTreeWidgetItem(self.treeWidget)
        # itemVolVideo0.setText(0, 'Сканер ВК1')
        # itemVolVideo0.setFont(0, font)
        # itemVolVideo0.setIcon(0, QtGui.QIcon(self.pathCamNotCon))
        # self.treeWidget.addTopLevelItem(itemVolVideo0)
        #
        # itemVolVideo1 = QtWidgets.QTreeWidgetItem(self.treeWidget)
        # itemVolVideo1.setText(0, 'Сканер ВК2')
        # itemVolVideo1.setFont(0, font)
        # itemVolVideo1.setIcon(0, QtGui.QIcon(self.pathCamNotCon))
        # self.treeWidget.addTopLevelItem(itemVolVideo1)

        self.advices.setCurrentIndex(0)

        # self.advices.setCurrentIndex(0)
        # self.cams.hide()

        # self.check()
        self.btnOpenPorts.clicked.connect(self.openPorts)
        self.btnCheckDBMy.clicked.connect(self.checkConDB)
        self.btnCheckDBPg.clicked.connect(self.checkConDBpg)
        self.btnAddCam.clicked.connect(self.openPanelAddNetworkSet)
        self.btnDelCam.clicked.connect(self.deleteCamSet)
        self.treeWidget.doubleClicked.connect(self.editDataInCam)
        self.treeWidget.clicked.connect(self.openCamVideo)
        self.btnSetCam.clicked.connect(self.editDataInCam)
        self.btnArchive.clicked.connect(self.openPanelStorage)
        self.btnConnScan.clicked.connect(self.startMeasuringVol)


        self.leIpGard.setPlaceholderText('192.168.0.100')
        self.leMaskGard.setPlaceholderText('255.255.255.0')
        self.leGatewayGard.setPlaceholderText('192.168.0.1')

        list = serial.tools.list_ports.comports()

        print('ComPorts: ', list)



        connected = []
        for element in list:
            dataEl = str(element.device)
            dataLst = dataEl.split('S')
            dataEl = 'Com ' + str(dataLst[1])
            connected.append(dataEl)
            print('ComPortsName: ', dataEl)


        num_it = 0

        self.cbWeight.clear()
        self.cbTraff.clear()

        self.cbWeight.addItems(connected)
        self.cbTraff.addItems(connected)

        # pathData = str(os.getenv('APPDATA'))
        pathData = globalValues.pathDefaultData
        pathFolder = pathData + '/Sinaps'
        self.checkFolderPath(pathFolder, True)

        pathDataCom = pathData + '/Sinaps/dataComPorts.txt'
        pathDataCams = pathData + '/Sinaps/dataCams.txt'
        pathDataDBmySql = pathData + '/Sinaps/dataDBmySql.txt'
        pathDataDBpgSql = pathData + '/Sinaps/dataDBpgSql.txt'

        if (os.path.exists(pathDataCom) == False):
            file_com = open(pathDataCom, 'w')
            file_com.write('')
            file_com.close()
        else:
            try:
                file_com = open(pathDataCom, 'r')
                data_coms = file_com.read()
                file_com.close()

                if(data_coms != ''):
                        list_coms = data_coms.split(' ')
                        data_com_weight = list_coms[0]
                        data_com_traffic = list_coms[1]
                        num_items_weight = self.cbWeight.count()

                        for i in range(num_items_weight):
                            str_from_cb = self.cbWeight.itemText(i)
                            if ( data_com_weight == str_from_cb):
                                self.cbWeight.setCurrentIndex(i)
                                break

                        num_items_traffic = self.cbTraff.count()

                        for i in range(num_items_traffic):
                            str_from_cb_tr = self.cbTraff.itemText(i)
                            if ( data_com_traffic == str_from_cb_tr):
                                self.cbTraff.setCurrentIndex(i)
                                break

            except Exception as ex:
                globalValues.writeLogData('Проверка и считывание данных о портах из файла', str(ex))

        if (os.path.exists(pathDataCams) == False):
            file_cam = open(pathDataCams, 'w')
            file_cam.write('')
            file_cam.close()
        else:
            try:
                file_cam = open(pathDataCams, 'r')
                data_cams = file_cam.read()
                file_cam.close()

                if (data_cams != ''):
                    dataLst = data_cams.split(' ')
                    checkAddCam = dataLst[2]
                    if (checkAddCam == '1'):
                        self.btnAddCam.setEnabled(False)
                    else:
                        self.btnAddCam.setEnabled(True)

            except Exception as ex:
                globalValues.writeLogData('Проверка и считывание данных о видеокамере из файла', str(ex))

        if (os.path.exists(pathDataDBmySql) == False):
            fileDBmy = open(pathDataDBmySql, 'w')
            strDataToDBmySql = 'localhost 3306'
            fileDBmy.write(strDataToDBmySql)
            fileDBmy.close()

            # subprocess.call(['attrib', '+H', pathDataDBmySql])

            self.leDBIpMy.setText('localhost')
            self.leDBPortMy.setText('3306')
        else:
            try:
                    fileDBmy = open(pathDataDBmySql, 'r')
                    strDataMy = fileDBmy.read()
                    listDataMy = strDataMy.split(' ')
                    self.leDBIpMy.setText(listDataMy[0])
                    self.leDBPortMy.setText(listDataMy[1])
                    fileDBmy.close()
            except Exception as ex:
                    globalValues.writeLogData('Чтение данных из файла mySql', str(ex))

        if (os.path.exists(pathDataDBpgSql) == False):
            fileDBpg = open(pathDataDBpgSql, 'w')
            strDataToDBpgSql = 'localhost 5432'
            fileDBpg.write(strDataToDBpgSql)
            fileDBpg.close()

            # subprocess.call(['attrib', '+H', pathDataDBpgSql])

            self.leDBIpPg.setText('localhost')
            self.leDBPortPg.setText('5432')
        else:
            try:
                    fileDBpg = open(pathDataDBpgSql, 'r')
                    strDataPg = fileDBpg.read()
                    listDataPg = strDataPg.split(' ')
                    self.leDBIpPg.setText(listDataPg[0])
                    self.leDBPortPg.setText(listDataPg[1])
                    fileDBpg.close()
            except Exception as ex:
                    globalValues.writeLogData('Чтение данных из файла pgSql', str(ex))

        self.leDBPassMy.setEchoMode(QtWidgets.QLineEdit.Password)
        self.leDBPassPg.setEchoMode(QtWidgets.QLineEdit.Password)

        # self.leDBIpPg.setText('localhost')
        # self.leDBLoginPg.setText('postgres')
        # self.leDBPassPg.setText('sinaps281082')
        # self.leDBNamePg.setText('trassir3')
        # self.leDBIpMy.setText('localhost')
        # self.leDBLoginMy.setText('sergey')
        # self.leDBPassMy.setText('34ubitav')
        # self.leDBNameMy.setText('cars')
        # self.leDBPortMy.setText('3306')
        # self.leDBPortPg.setText('5432')
        # globalValues.my_sql_db = self.leDBNameMy.text()

        globalValues.my_sql_localhost = self.leDBIpMy.text()
        globalValues.my_sql_name = self.leDBLoginMy.text()
        globalValues.my_sql_password = self.leDBPassMy.text()
        globalValues.my_sql_port = int(self.leDBPortMy.text())
        # globalValues.my_pg_db = self.leDBNamePg.text()
        globalValues.my_pg_localhost = self.leDBIpPg.text()
        globalValues.my_pg_name = self.leDBLoginPg.text()
        globalValues.my_pg_password = self.leDBPassPg.text()
        globalValues.my_pg_port = int(self.leDBPortPg.text())

        self.readStrgStAndPath()

        self.btnInfo.clicked.connect(self.openPanAboutSystem)
        if self.checkLogin():
            self.checkArchive()

    def changeColor(self, object, str):
            object.setStyleSheet(str)

    def changeCOLORMainPanelGrunt(self, timeChangeColor):

        delta = int((timeChangeColor / self.lengthDark) * 1000)

        if globalValues.checkConDbMySql:
            if (globalValues.colorForm == 1):
                self.btnCheckDBMy.setStyleSheet("QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(0,255,0);}\n"
                                                "\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(0,255,0);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(0,255,0);};")
            else:

                self.btnCheckDBMy.setStyleSheet("QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(0,255,0);}\n"
                                                "\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(0,255,0);}\n"
                                                "\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(0,255,0);};")
        else:
            if (globalValues.colorForm == 1):
                self.btnCheckDBMy.setStyleSheet("QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);};")
            else:
                self.btnCheckDBMy.setStyleSheet("QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);};")

        if globalValues.checkConDbPgSql:
            if (globalValues.colorForm == 1):
                self.btnCheckDBPg.setStyleSheet("QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(0,255,0);}\n"
                                                "\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(0,255,0);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(0,255,0);};")
            else:
                self.btnCheckDBPg.setStyleSheet("QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(0,255,0);}\n"
                                                "\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(0,255,0);}\n"
                                                "\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(0,255,0);};")
        else:
            if (globalValues.colorForm == 1):
                self.btnCheckDBPg.setStyleSheet("QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);};")
            else:
                self.btnCheckDBPg.setStyleSheet("QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);};")

        if (globalValues.colorForm == 1):

            for i in range(self.lengthLight):
                startTime = round(time.time() * 1000)
                while True:
                    obj = self.lstLight[i][0]
                    style = self.lstLight[i][1]
                    self.changeColor(obj, style)
                    # print('changeLight!')
                    if (abs(round(time.time() * 1000) - startTime) > delta):
                        break

        elif (globalValues.colorForm == 0):

            for i in range(self.lengthDark):
                startTime = round(time.time() * 1000)
                while True:
                    # print(delta)
                    # print(abs(round(time.time()*1000) - startTime))
                    obj = self.lstDark[i][0]
                    style = self.lstDark[i][1]
                    self.changeColor(obj, style)
                    # print('changeDark!')
                    if (abs(round(time.time() * 1000) - startTime) > delta):
                        break

    def readStrgStAndPath(self):
        try:
            pathData = str(os.getenv('APPDATA')) + r'\Sinaps\dataStrg.txt'
            if (os.path.exists(pathData)):
                file_strg = open(pathData, 'r')
                data = file_strg.read()
                file_strg.close()
                if (data != ''):
                    dataLst = data.split(' ')
                    globalValues.pathStrg = dataLst[0]
                    if (dataLst[1] == '1'):
                        globalValues.stateStorage = True
                    else:
                        globalValues.stateStorage = False
                    if (dataLst[2] != ''):
                        globalValues.numberFolderDateSave = int(dataLst[2])
            else:
                file_strg = open(pathData, 'w')
                file_strg.close()

        except Exception as ex:
            globalValues.writeLogData('Функция считывания состояния архива', str(ex))

    def openPanelAddNetworkSet(self):
        pathData = str(os.getenv('APPDATA'))
        pathDataCams = pathData + r'\Sinaps\dataCams.txt'
        checkAddCam = 0
        uiNetCam = Ui_NetworkCam()

        # numCountUseItem = len(self.listNameChannel)

        try:
            file_cam = open(pathDataCams, 'r')
            data_cams = file_cam.read()
            file_cam.close()

            if (data_cams != ''):
                dataLst = data_cams.split(' ')
                uiNetCam.leNewCamName.setText(dataLst[0])
                uiNetCam.leNewCamIP.setText(dataLst[1])

        except Exception as ex:
            globalValues.writeLogData('Проверка и считывание данных о видеокамере из файла', str(ex))

        for strUseNameChl in self.listNameChannel:
             numItems = uiNetCam.comboNewCamChlName.count()
             for i in range(numItems):
                if(strUseNameChl == uiNetCam.comboNewCamChlName.itemText(i)):
                        uiNetCam.comboNewCamChlName.removeItem(i)
                        break

        uiNetCam.exec_()

        if (uiNetCam.checkSaveBtn == True):
            try:

                if (self.checkMySql):
                        self.con = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                                   user=globalValues.my_sql_name,
                                                   passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)

                with self.con:
                        print('GoodSQL')

                        cur = self.con.cursor()
                        query = ("INSERT INTO " + globalValues.tblsDB[3] + " (name, ip, channel) VALUES ( %s, %s, %s )")
                        # ( % s, \'B126PE777\', 5135, 9, \'выполняется\')'

                        passwdCam = str(uiNetCam.leNewCamPass.text())
                        loginCam = str(uiNetCam.leNewCamLogin.text())
                        nameChannel = str(uiNetCam.comboNewCamChlName.currentText())
                        print('GoodSQL')
                        # hash = hashlib.sha512(passwordCam.encode())
                        # data = str(hash.hexdigest())
                        # data = passwordCam
                        k = 0
                        z = 0
                        for row in self.listChannel:
                            if (row == nameChannel):
                                self.passwdLoginCams[z][0] = loginCam
                                self.passwdLoginCams[z][1] = passwdCam
                                globalValues.encryptDecryptNum[k+2] = True
                            k += 1
                            z += 1
                        print('GoodSQL')

                        cur.execute(query, (uiNetCam.leNewCamName.text(), uiNetCam.leNewCamIP.text(), nameChannel))

                        self.con.commit()
                        print('DBGood')
                        sleep(0.2)
                        strDataToJurnal = 'Добавлены сетевые настройки камеры для канала: ' + uiNetCam.comboNewCamChlName.currentText()
                        globalValues.writeEventToDBJournalMain('Камеры', strDataToJurnal)
                        self.checkCamAfterAdd = True
                        sleep(0.2)
                        i_0 = 0
                        for listIt in self.listChannel:
                           if (listIt == str(uiNetCam.comboNewCamChlName.currentText())):
                                   print('checkInt!' + str(i_0))
                                   self.indexAddCam = i_0
                                   break
                           i_0 += 1
                        self.openCamVideo()
                        cur.close()

                if (uiNetCam.comboNewCamChlName.count() == 1):
                    self.btnAddCam.setEnabled(False)
                    checkAddCam = 1

            except  Exception as ex:
                print(str(ex))
                print('ExitError')
                globalValues.writeLogData('Функция добавления новых настроек камеры', str(ex))

        try:
            file_cam = open(pathDataCams, 'w')
            dataToWrt = uiNetCam.leNewCamName.text() + ' ' + uiNetCam.leNewCamIP.text() + ' ' + str(checkAddCam)
            file_cam.write(dataToWrt)
            file_cam.close()
        except Exception as ex:
            globalValues.writeLogData('Проверка и считывание данных о видеокамере из файла', str(ex))

    def switchDisabled(self, number):
            print('CheckingNum: ', self.numTab)
            if number == 1:
                            self.advices.setCurrentIndex(0)
                            self.advices.setTabEnabled(0, True)
                            self.advices.setTabEnabled(1, False)
                            self.advices.setTabEnabled(2, False)
                            self.advices.setTabEnabled(3, False)
                            self.advices.setTabEnabled(4, False)
                            self.advices.setTabEnabled(5, False)
            elif number == 2:
                            self.advices.setCurrentIndex(1)
                            self.advices.setTabEnabled(0, False)
                            self.advices.setTabEnabled(1, True)
                            self.advices.setTabEnabled(2, False)
                            self.advices.setTabEnabled(3, False)
                            self.advices.setTabEnabled(4, False)
                            self.advices.setTabEnabled(5, False)
            elif number == 3:
                            self.advices.setCurrentIndex(2)
                            self.advices.setTabEnabled(0, False)
                            self.advices.setTabEnabled(1, False)
                            self.advices.setTabEnabled(2, True)
                            self.advices.setTabEnabled(3, False)
                            self.advices.setTabEnabled(4, False)
                            self.advices.setTabEnabled(5, False)
            elif number == 4:
                            self.advices.setCurrentIndex(3)
                            self.advices.setTabEnabled(0, False)
                            self.advices.setTabEnabled(1, False)
                            self.advices.setTabEnabled(2, False)
                            self.advices.setTabEnabled(3, True)
                            self.advices.setTabEnabled(4, False)
                            self.advices.setTabEnabled(5, False)
            elif number == 5:
                            self.advices.setCurrentIndex(4)
                            self.advices.setTabEnabled(0, False)
                            self.advices.setTabEnabled(1, False)
                            self.advices.setTabEnabled(2, False)
                            self.advices.setTabEnabled(3, False)
                            self.advices.setTabEnabled(4, True)
                            self.advices.setTabEnabled(5, False)
                            if (globalValues.stateScaner):
                                print('CheckingCam!!!')
                                globalValues.stopVideoScaner = False
                                th_video_scan = threading.Thread(target=self.showVideoScaner)
                                th_video_scan.start()
            elif number == 6:
                            self.advices.setCurrentIndex(5)
                            self.advices.setTabEnabled(0, False)
                            self.advices.setTabEnabled(1, False)
                            self.advices.setTabEnabled(2, False)
                            self.advices.setTabEnabled(3, False)
                            self.advices.setTabEnabled(4, False)
                            self.advices.setTabEnabled(5, True)

    def startCheckConMySql(self):
            self.thStartEncryptedAndDecryptedData()
            if (self.checkStartDB == False):
                th_check_mysql = threading.Thread(target=self.checkConDB)
                th_check_mysql.start()
            self.checkStartDB == False

    def startCheckConPgSql(self):
            th_check_pgsql = threading.Thread(target=self.checkConDBpg)
            th_check_pgsql.start()

    def runGif(self):
        self.lblMovie = QtWidgets.QLabel(self)
        self.lblMovie.setGeometry(QtCore.QRect(100, 100, 50, 50))
        self.lblMovie.raise_()
        movie = QMovie(globalValues.pathStyleImgs + "animdark3.gif")
        size = QtCore.QSize(50, 50)
        movie.setScaledSize(size)
        self.lblMovie.setMovie(movie)
        movie.start()

    def checkConDB(self):

        pathData = str(os.getenv('APPDATA'))
        pathDataDBmySql = pathData + r'\Sinaps\dataDBmySql.txt'
        dataToWrite = self.leDBIpMy.text() + ' ' + self.leDBPortMy.text()
        self.writeDataToTxtFile(dataToWrite, pathDataDBmySql)

        # try:
        #         if (self.num_first_call_mysql != 0):
        #                 if (self.cur.connection):
        #                     self.cur.close()
        # except Exception as ex:
        #         print('ErrorDbMy!')
        #         globalValues.writeLogData('Функция проверки состояния содинения с бд MySQL', str(ex))

        # pt.CoInitialize()

        if (globalValues.statusGoodLoginDev[0] and globalValues.statusGoodLoginDev[1] and self.leDBPassMy.text() == '' and self.leDBLoginMy.text() == ''):
            globalValues.my_sql_name = globalValues.loginDataList[0]
            globalValues.my_sql_password = globalValues.loginDataList[1]
            print('checkAutoLoginMySql!!!')
        else:
            globalValues.my_sql_name = self.leDBLoginMy.text()
            globalValues.my_sql_password = self.leDBPassMy.text()

        # globalValues.my_sql_db = self.leDBNameMy.text()
        globalValues.my_sql_localhost = self.leDBIpMy.text()
        globalValues.my_sql_port = int(self.leDBPortMy.text())



        try:
            # print(globalValues.my_sql_password)
            # print(globalValues.my_sql_name)

            # GRANT < privileges > ON database. * TO 'user' @ 'localhost'IDENTIFIED BY 'password';

            print('CheckingPass: ', globalValues.my_sql_password)
            print('CheckingName: ', globalValues.my_sql_name)

            self.con = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                  user=globalValues.my_sql_name,
                                  passwd=globalValues.my_sql_password)
            if (self.con.open):
                if (self.goodLoginDBMySQL == False):
                        globalValues.encryptDecryptNum[0] = True
                self.goodLoginDBMySQL = True
                print('GoodCheckDbMySql!')
                globalValues.checkConDbMySql = True
                if (globalValues.colorForm == 1):
                    self.btnCheckDBMy.setStyleSheet("QPushButton:!hover {background-color: rgb(227,227,227);\n"
                        "color: rgb(0,0,0);\n"
                        "border-radius:3px;\n"
                        "border:1px solid rgb(0,255,0);}\n"
                        "\n"
                        "QPushButton:hover {background-color: rgb(84,122,181);\n"
                        "color: rgb(255, 255, 255);\n"
                        "border-radius:3px;\n"
                        "border:1px solid rgb(0,255,0);}\n"
                        "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                        "color: rgb(255, 255, 255);\n"
                        "border-radius:3px;\n"
                        "border:1px solid rgb(0,255,0);};")
                else:

                    self.btnCheckDBMy.setStyleSheet("QPushButton:!hover {background-color: rgb(89,89,89);\n"
                        "color: rgb(255, 255, 255);\n"
                        "border-radius:3px;\n"
                        "border:1px solid rgb(0,255,0);}\n"
                        "\n"
                        "QPushButton:hover {background-color: rgb(84,122,181);\n"
                        "color: rgb(255, 255, 255);\n"
                        "border-radius:3px;\n"
                        "border:1px solid rgb(0,255,0);}\n"
                        "\n"
                        "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                        "color: rgb(255, 255, 255);\n"
                        "border-radius:3px;\n"
                        "border:1px solid rgb(0,255,0);};")

                # self.cur.close()
                self.lblStDBMy.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
            'image: url(' + pathImgGreenCom + ');};')
                self.num_first_call_mysql += 1


                createDBMySqlWithTbls(globalValues.my_sql_localhost, globalValues.my_sql_port, globalValues.my_sql_name, globalValues.my_sql_password, globalValues.dbMySqlName, globalValues.tblsDB)

                self.con.close()

                self.con = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                  user=globalValues.my_sql_name,
                                  passwd=globalValues.my_sql_password,  db=globalValues.dbMySqlName)

                self.leDBLoginMy.setText('')
                self.leDBPassMy.setText('')

                globalValues.writeEventToDBJournalMain('Базы данных', 'Проверка соединения с бд MySQL выполнена успешно.')

        except Exception as ex:

            self.goodLoginDBMySQL = False

            globalValues.boolCheckInsertPassword[0] = True

            # globalValues.encryptDecryptNum[0] = True

            globalValues.writeLogData('Функция проверки содинения с бд MySQL', str(ex))

            # globalValues.writeEventToDBJournalMain('Базы данных', 'Не удалось установить соединение с бд MySQL')

            print("Connection Error MySQL")

            self.lblMovieWhite.hide()
            globalValues.checkConDbMySql = False
            if (globalValues.colorForm == 1):
                self.btnCheckDBMy.setStyleSheet("QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);};")
            else:
                self.btnCheckDBMy.setStyleSheet("QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);};")

            self.lblStDBMy.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
            'image: url(' + pathImgRedCom + ');};')

            self.num_first_call_mysql = 0
            # con = None

        if (self.num_first_call_mysql == 1):
            # sleep(1)

            self.check()
            self.openPorts()
            sleep(0.85)
            self.checkAllCams()
            globalValues.writeEventToDBJournalMain('Базы данных', 'Установлено успешное соединение с бд MySQL.')

    def thStartEncryptedAndDecryptedData(self):
         th_encrypted_decrypted = threading.Thread(target=self.thEncryptedAndDecryptedData)
         th_encrypted_decrypted.start()

    def thEncryptedAndDecryptedData(self):

           try:
               # pt.CoInitialize()

               def encrypt(strEncrypted, numberRow, numberCol):
                   try:
                       print('DataIn: ', strEncrypted, numberRow, numberCol)
                       public_key = """-----BEGIN PUBLIC KEY-----
                   MIIBIjANBgkqhkiG9w0BAQEFAAOCAQ8AMIIBCgKCAQEA0c45dZkOTBkAi/eP4i6G
                   rc1GVPVw1a+dni+ciOqACFISHElsPHGUIBIpzAxVUt7Kgc3QPZ9amMewFTz+whoe
                   7HjAf+wtC+ISBDw8s7CQxknbeUAXNYeShbtP1l8Mz6sHJmSk97qFMdyICOALIKY6
                   ky0rYxG7AsqCbsVBe2MKGqKKGuhoc4Ei35vNBVPn3KLNmSqjfKk4QwNmttwP6gEq
                   ldZwT6pjz2wgA42CMQYkVEQ4w4Bwa0Go/C1AsrwLd7q3tgqM+tAEctICAjNAdTKM
                   hFWMOdUn60yaEFcZPo7EE0dUw5ZYUp91fqQBDoxt2D3p0iWmTEKZPgx9Tbtu+Z6/
                   JwIDAQAB
                   -----END PUBLIC KEY-----"""
                       rsa_key = RSA.importKey(public_key)
                       rsa_key = PKCS1_OAEP.new(rsa_key)
                       encrypted = rsa_key.encrypt(strEncrypted.encode('utf8'))
                       encrypted_text = b64encode(encrypted)

                       # pathFile = str(os.getenv('APPDATA')) + r'\Sinaps\cryptoSSA.xlsx'
                       # excel = win32com.client.Dispatch('Excel.Application')
                       # workbook = excel.Workbooks.open(pathFile, True, False, None, '34ubitav')
                       # sheet = workbook.Worksheets('Лист1')
                       # excel.Visible = False
                       # excel.DisplayAlerts = False

                       pathXls = globalValues.pathFileXls
                       print(pathXls)

                       book = openpyxl.load_workbook(pathXls)
                       sheets = book.sheetnames
                       print('qwer')
                       for sheet in sheets:
                           print(str(sheet))
                           print('qwerty!!!')
                           if (str(sheet) == 'Лист1'):
                               sheet_main = book[sheet]
                               encrypted_text = str(encrypted_text)
                               length = len(encrypted_text)
                               encrypted_text = encrypted_text[2: length - 1]
                               sheet_main.cell(numberRow, numberCol).value = str(encrypted_text)
                               print(encrypted_text)

                       book.save(pathXls)
                       book.close()

                       # workbook.Close(True, pathFile)
                       # excel.Quit

                   except Exception as ex:
                       globalValues.writeLogData('Функция кодирования данных и сохранения в запароленный файл', str(ex))

               while True:
                   if(globalValues.stopAll):
                       break

                   if (globalValues.stopEncryptedDecrypt):
                       globalValues.stopEncryptedDecrypt = False
                       break

                   # sleep(1)
                   # print('Checking!!!!!!!!!!!!!')

                   #MySQL
                   if (globalValues.encryptDecryptNum[0]):
                           print('TrueDB')
                           encrypt(globalValues.my_sql_name, 1, 3)
                           encrypt('TrueDBSSA', 2, 3)
                           encrypt(globalValues.my_sql_password, 1, 4)
                           encrypt('TrueDBSSA', 2, 4)
                           globalValues.statusGoodLoginDev[0] = True
                           globalValues.statusGoodLoginDev[1] = True
                           globalValues.encryptDecryptNum[0] = False

                   if (globalValues.boolCheckInsertPassword[0]):
                           print('Sergik', globalValues.statusGoodLoginDev[0], globalValues.statusGoodLoginDev[1])
                           if (globalValues.statusGoodLoginDev[0] or globalValues.statusGoodLoginDev[1]):
                                   print('FalseDBmySql!!!')
                                   globalValues.statusGoodLoginDev[0] = False
                                   globalValues.statusGoodLoginDev[1] = False
                                   print('ChangeFalseElsMySql')
                                   encrypt('FalseDBSSA', 2, 3)
                                   encrypt('FalseDBSSA', 2, 4)
                           globalValues.boolCheckInsertPassword[0] = False

                   #PG

                   if (globalValues.encryptDecryptNum[1]):
                           print('11111111111111111111')
                           encrypt(globalValues.my_pg_name, 1, 5)
                           encrypt('TrueDBSSA', 2, 5)
                           encrypt(globalValues.my_pg_password, 1, 6)
                           encrypt('TrueDBSSA', 2, 6)
                           globalValues.statusGoodLoginDev[2] = True
                           globalValues.statusGoodLoginDev[3] = True
                           globalValues.encryptDecryptNum[1] = False
                   # print('qwerty')
                   if (globalValues.boolCheckInsertPassword[1]):
                           print('22222222222222222')
                           if (globalValues.statusGoodLoginDev[2] or globalValues.statusGoodLoginDev[3]):
                                   encrypt('FalseDBSSA', 2, 5)
                                   encrypt('FalseDBSSA', 2, 6)
                                   globalValues.statusGoodLoginDev[2] = False
                                   globalValues.statusGoodLoginDev[3] = False
                                   print('ChangeFalseElsPgSql')
                           globalValues.boolCheckInsertPassword[1] = False

                   #Cam0

                   if (globalValues.encryptDecryptNum[2]):
                        print('EncryptedDataCam0')
                        encrypt(self.passwdLoginCams[0][0], 1, 7)
                        encrypt(self.passwdLoginCams[0][1], 1, 8)
                        globalValues.encryptDecryptNum[2] = False

                   if (globalValues.boolCheckInsertPassword[2]):
                           print('FalseCheckCam0')
                           if (globalValues.statusGoodLoginDev[4] or globalValues.statusGoodLoginDev[5]):
                                   encrypt('FalseDBSSA', 2, 7)
                                   encrypt('FalseDBSSA', 2, 8)
                                   globalValues.statusGoodLoginDev[4] = False
                                   globalValues.statusGoodLoginDev[5] = False
                                   print('ChangeFalseCam0')
                           globalValues.boolCheckInsertPassword[2] = False

                   if (globalValues.boolCheckInsertTrueCamPassword[0]):
                       print('TrueCheckCam0')
                       encrypt('TrueDBSSA', 2, 7)
                       encrypt('TrueDBSSA', 2, 8)
                       globalValues.statusGoodLoginDev[4] = True
                       globalValues.statusGoodLoginDev[5] = True
                       globalValues.boolCheckInsertTrueCamPassword[0] = False

                   #Cam1

                   if (globalValues.encryptDecryptNum[3]):
                        print('EncryptedDataCam1')
                        encrypt(self.passwdLoginCams[1][0], 1, 9)
                        encrypt('TrueDBSSA', 2, 9)
                        encrypt(self.passwdLoginCams[1][1], 1, 10)
                        encrypt('TrueDBSSA', 2, 10)
                        globalValues.encryptDecryptNum[3] = False
                   if (globalValues.boolCheckInsertPassword[3]):
                           print('FalseCheckCam1')
                           if (globalValues.statusGoodLoginDev[6] or globalValues.statusGoodLoginDev[7]):
                                   encrypt('FalseDBSSA', 2, 9)
                                   encrypt('FalseDBSSA', 2, 10)
                                   globalValues.statusGoodLoginDev[6] = False
                                   globalValues.statusGoodLoginDev[7] = False
                                   print('ChangeFalseCam1')
                           globalValues.boolCheckInsertPassword[3] = False
                   if (globalValues.boolCheckInsertTrueCamPassword[1]):
                       encrypt('TrueDBSSA', 2, 9)
                       encrypt('TrueDBSSA', 2, 10)
                       globalValues.statusGoodLoginDev[6] = True
                       globalValues.statusGoodLoginDev[7] = True
                       globalValues.boolCheckInsertTrueCamPassword[1] = False

                   # Cam2

                   if (globalValues.encryptDecryptNum[4]):
                           print('EncryptedDataCam2')
                           encrypt(self.passwdLoginCams[2][0], 1, 11)
                           encrypt('TrueDBSSA', 2, 11)
                           encrypt(self.passwdLoginCams[2][1], 1, 12)
                           encrypt('TrueDBSSA', 2, 12)
                           globalValues.encryptDecryptNum[4] = False
                   if (globalValues.boolCheckInsertPassword[4]):
                           print('FalseCheckCam2')
                           if (globalValues.statusGoodLoginDev[8] or globalValues.statusGoodLoginDev[9]):
                                   encrypt('FalseDBSSA', 2, 11)
                                   encrypt('FalseDBSSA', 2, 12)
                                   globalValues.statusGoodLoginDev[8] = False
                                   globalValues.statusGoodLoginDev[9] = False
                                   print('ChangeFalseCam2')
                           globalValues.boolCheckInsertPassword[4] = False
                   if (globalValues.boolCheckInsertTrueCamPassword[2]):
                       encrypt('TrueDBSSA', 2, 11)
                       encrypt('TrueDBSSA', 2, 12)
                       globalValues.statusGoodLoginDev[8] = True
                       globalValues.statusGoodLoginDev[9] = True
                       globalValues.boolCheckInsertTrueCamPassword[2] = False

                   # Cam3

                   if (globalValues.encryptDecryptNum[5]):
                           print('EncryptedDataCam3')
                           encrypt(self.passwdLoginCams[3][0], 1, 13)
                           encrypt('TrueDBSSA', 2, 13)
                           encrypt(self.passwdLoginCams[3][1], 1, 14)
                           encrypt('TrueDBSSA', 2, 14)
                           globalValues.encryptDecryptNum[5] = False
                   if (globalValues.boolCheckInsertPassword[5]):
                           print('FalseCheckCam3')
                           if (globalValues.statusGoodLoginDev[10] or globalValues.statusGoodLoginDev[11]):
                                   encrypt('FalseDBSSA', 2, 13)
                                   encrypt('FalseDBSSA', 2, 14)
                                   globalValues.statusGoodLoginDev[10] = False
                                   globalValues.statusGoodLoginDev[11] = False
                                   print('ChangeFalseCam3')
                           globalValues.boolCheckInsertPassword[5] = False
                   if (globalValues.boolCheckInsertTrueCamPassword[3]):
                       encrypt('TrueDBSSA', 2, 13)
                       encrypt('TrueDBSSA', 2, 14)
                       globalValues.statusGoodLoginDev[10] = True
                       globalValues.statusGoodLoginDev[11] = True
                       globalValues.boolCheckInsertTrueCamPassword[3] = False

                   #Cam4

                   if (globalValues.encryptDecryptNum[6]):
                           print('EncryptedDataCam4')
                           encrypt(self.passwdLoginCams[4][0], 1, 15)
                           encrypt('TrueDBSSA', 2, 15)
                           encrypt(self.passwdLoginCams[4][1], 1, 16)
                           encrypt('TrueDBSSA', 2, 16)
                           globalValues.encryptDecryptNum[6] = False
                   if (globalValues.boolCheckInsertPassword[6]):
                           print('FalseCheckCam4')
                           if (globalValues.statusGoodLoginDev[12] or globalValues.statusGoodLoginDev[13]):
                                   encrypt('FalseDBSSA', 2, 15)
                                   encrypt('FalseDBSSA', 2, 16)
                                   globalValues.statusGoodLoginDev[12] = False
                                   globalValues.statusGoodLoginDev[13] = False
                                   print('ChangeFalseCam4')
                           globalValues.boolCheckInsertPassword[6] = False
                   if (globalValues.boolCheckInsertTrueCamPassword[4]):
                       encrypt('TrueDBSSA', 2, 15)
                       encrypt('TrueDBSSA', 2, 16)
                       globalValues.statusGoodLoginDev[12] = True
                       globalValues.statusGoodLoginDev[13] = True
                       globalValues.boolCheckInsertTrueCamPassword[4] = False

                   #Cam5

                   if (globalValues.encryptDecryptNum[7]):
                           print('EncryptedDataCam5')
                           encrypt(self.passwdLoginCams[5][0], 1, 17)
                           encrypt('TrueDBSSA', 2, 17)
                           encrypt(self.passwdLoginCams[5][1], 1, 18)
                           encrypt('TrueDBSSA', 2, 18)
                           globalValues.encryptDecryptNum[7] = False
                   if (globalValues.boolCheckInsertPassword[7]):
                           print('FalseCheckCam3')
                           if (globalValues.statusGoodLoginDev[14] or globalValues.statusGoodLoginDev[15]):
                                   encrypt('FalseDBSSA', 2, 17)
                                   encrypt('FalseDBSSA', 2, 18)
                                   globalValues.statusGoodLoginDev[14] = False
                                   globalValues.statusGoodLoginDev[15] = False
                                   print('ChangeFalseCam5')
                           globalValues.boolCheckInsertPassword[7] = False
                   if (globalValues.boolCheckInsertTrueCamPassword[5]):
                       encrypt('TrueDBSSA', 2, 17)
                       encrypt('TrueDBSSA', 2, 18)
                       print('checkTrueCam5!')
                       globalValues.statusGoodLoginDev[14] = True
                       globalValues.statusGoodLoginDev[15] = True
                       globalValues.boolCheckInsertTrueCamPassword[5] = False

                   # print('checkSt!!!')

                   sleep(0.1)

           except Exception as ex:
               globalValues.writeLogData('Поток кодирования и декодирования данных', str(ex))

           # pt.CoUninitialize()

           print('EndingEncrypto!!!')

    def checkConDBpg(self):

        pathData = str(os.getenv('APPDATA'))
        pathDataDBmySql = pathData + r'\Sinaps\dataDBpgSql.txt'
        dataToWrite = self.leDBIpPg.text() + ' ' + self.leDBPortPg.text()
        self.writeDataToTxtFile(dataToWrite, pathDataDBmySql)

        if (globalValues.statusGoodLoginDev[2] and globalValues.statusGoodLoginDev[3] and self.leDBPassPg.text() == '' and self.leDBLoginPg.text() == ''):
            globalValues.my_pg_name = globalValues.loginDataList[2]
            globalValues.my_pg_password = globalValues.loginDataList[3]
            print('checkAutoLoginPg!!!')
        else:
            # globalValues.my_pg_db = self.leDBNamePg.text()
            # globalValues.my_pg_db = ''
            globalValues.my_pg_localhost = self.leDBIpPg.text()
            globalValues.my_pg_name = self.leDBLoginPg.text()
            globalValues.my_pg_password = self.leDBPassPg.text()
            globalValues.my_pg_port = int(self.leDBPortPg.text())


        try:
                if (self.num_first_call_pgsql != 0):
                        if (self.conn_pg is not None):
                            self.conn_pg.close()
        except Exception as ex:
                print('ErrorDbPg!')
                globalValues.writeLogData('Функция проверки состояния соединения с бд PG', str(ex))

        try:
            self.conn_pg = psycopg2.connect(dbname=globalValues.my_pg_db, user=globalValues.my_pg_name,
                                       password=globalValues.my_pg_password, host=globalValues.my_pg_localhost,
                                       port=globalValues.my_pg_port)

            # cur_pg = self.conn_pg.cursor()

            with self.conn_pg:
                if (self.conn_pg is not None):
                    if (self.goodLoginDBPgSQL == False):
                        globalValues.encryptDecryptNum[1] = True
                    self.goodLoginDBPgSQL = True
                    globalValues.checkConDbPgSql = True
                    if (globalValues.colorForm == 1):
                        self.btnCheckDBPg.setStyleSheet("QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                        "color: rgb(0,0,0);\n"
                                                        "border-radius:3px;\n"
                                                        "border:1px solid rgb(0,255,0);}\n"
                                                        "\n"
                                                        "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:3px;\n"
                                                        "border:1px solid rgb(0,255,0);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:3px;\n"
                                                        "border:1px solid rgb(0,255,0);};")
                    else:
                        self.btnCheckDBPg.setStyleSheet("QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:3px;\n"
                                                        "border:1px solid rgb(0,255,0);}\n"
                                                        "\n"
                                                        "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:3px;\n"
                                                        "border:1px solid rgb(0,255,0);}\n"
                                                        "\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:3px;\n"
                                                        "border:1px solid rgb(0,255,0);};")

                    self.num_first_call_pgsql += 1
                    globalValues.writeEventToDBJournalMain('Базы данных', 'Проверка соединения с бд PgSQL выполнена успешно.')

            self.lblStDBPg.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
            'image: url(' + pathImgGreenCom+ ');};')

            if (self.num_first_call_pgsql == 1):
                # print('qqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqqq')
                globalValues.writeEventToDBJournalMain('Базы данных', 'Установлено успешное соединение с бд PgSQL.')


        except Exception as ex:

            print("Connection Error PgSQL")

            self.goodLoginDBPgSQL = False
            globalValues.boolCheckInsertPassword[1] = True
            # globalValues.encryptDecryptNum[1] = True

            self.num_first_call_pgsql = 0

            globalValues.checkConDbPgSql = False

            if (globalValues.colorForm == 1):
                self.btnCheckDBPg.setStyleSheet("QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);};")
            else:
                self.btnCheckDBPg.setStyleSheet("QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);}\n"
                                                "\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(255,0,0);};")

            self.lblStDBPg.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
            'image: url(' + pathImgRedCom + ');};')

            globalValues.writeLogData('Функция проверки содинения с бд PG', str(ex))

            globalValues.writeEventToDBJournalMain('Базы данных', 'Не удалось установить соединение с бд PgSQL')

    def check(self):
        try:
                th_refresh_sys_menu = Thread(target=self.thread_my)
                th_refresh_sys_menu.start()

        except Exception as ex:
                globalValues.writeLogData('Функция обновления компонентов системного меню ', str(ex))

    def openPanAboutSystem(self):
        try:
            uiPanAbout = Ui_systemMenu()
            uiPanAbout.exec_()
        except Exception as ex:
            globalValues.writeLogData('Функция справки системного меню', str(ex))

    def openPorts(self):
        try:
                pathData = globalValues.curDisk + '/Sinaps/dataComPorts.txt'
                file_com = open(pathData, 'w')
                dataAll = self.cbWeight.currentText() + ' ' + self.cbTraff.currentText()
                file_com.write(dataAll)
                file_com.close()

                # self.cbTraff.setCurrentIndex(2)

                portWeight = self.cbWeight.currentText()
                portTraffic = self.cbTraff.currentText()

                globalValues.outComTraf = False
                globalValues.outComWeight = False

                #Запуск потока кодирования данных

                # z = threading.Thread(target=self.thEncryptedAndDecryptedData)
                # z.start()

                if (globalValues.debugPortWeight == False):
                    if (self.is_good_start_weight == False):
                        globalValues.stopComPortWeight = True
                        time.sleep(0.3)
                        globalValues.stopComPortWeight = False
                        x = threading.Thread(target=self.openPortWeightThread, args=(q, portWeight, ))
                        x.start()

                else:
                    self.lblStWeightAndSema.setStyleSheet(
                        'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGreenCom + ');};')
                    self.lblStWeightImg.setStyleSheet(
                        'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGreen + ');};')

                    currentIndexTab = self.advices.currentIndex()
                    self.advices.setCurrentIndex(0)
                    self.advices.setCurrentIndex(currentIndexTab)

                    self.is_good_con_weight.emit()

                if (globalValues.debugPorts == False):
                    if(self.is_good_start_traffic == False):

                        globalValues.stopComPortTraffic = True
                        time.sleep(0.3)
                        globalValues.stopComPortTraffic = False

                        y = threading.Thread(target=self.openPortTrafficThread, args=(q, portTraffic, ))
                        y.start()
                else:
                    # print('startingComs!')
                    #
                    #
                    self.lblStWeightAndSema.setStyleSheet(
                            'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGreenCom + ');};')
                    self.lblStTraffImg.setStyleSheet(
                            'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGreen + ');};')

                    currentIndexTab = self.advices.currentIndex()
                    self.advices.setCurrentIndex(0)
                    self.advices.setCurrentIndex(currentIndexTab)

                    self.is_good_con_traffic.emit()



        except Exception as ex:
                print(ex)

    def openPortWeightThread(self, output_queue, portWeight):
        try:
            iter = 0
            time_start = 0
            check_start = True
            checkMeasure = False
            while True:
                if (check_start):
                    try:

                        try:
                            con = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                                  user=globalValues.my_sql_name,
                                                  passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)

                            # cur = con.cursor()

                        except Exception as ex:
                            globalValues.writeLogData('Подключение к БД MySql в потоке обработки данных с УЦИ весов НВТ-9',
                                                      str(ex))

                        port_weight = str(portWeight)

                        lstPorts = port_weight.split(' ')
                        port_weight = '/dev/ttyS' + lstPorts[1]
                        print(port_weight)

                        # print()

                        ser_weight = serial.Serial(
                            port=port_weight, \
                            baudrate=9600, \
                            parity=serial.PARITY_NONE, \
                            stopbits=serial.STOPBITS_ONE, \
                            bytesize=serial.EIGHTBITS, \
                            timeout=0)

                        print('Start Com Port!!!')

                        time_get_data = round(time.time())
                        line = []

                        self.is_good_start_weight = False

                        Debug = globalValues.debugPorts

                        strValueWeight = ''
                        intValueWeightCur = 0
                        intValueWeightOld = ''
                        intValueWeight = 0
                        countIter = 0
                        countIterDefaultBack = 0
                        strNumberGRZCur = ''
                        checkWriteValue = True

                        time_start = round(time.time())
                        check_ready_file = True

                        check_start = True

                        check_change_sema = False

                        is_change_sema_default = False

                        check_measure = True
                        time_send_value = round(time.time())
                        time_check = round(time.time())
                        wdg_i = 0


                        while True:
                            if (globalValues.stopAll):
                                print('EndComWeight!')
                                break

                            if (abs(round(time.time()) - time_check) > 2):
                                time_check = round(time.time())

                            for c in ser_weight.read():

                                h = hex(c)
                                line.append(h)
                                if h == '0xa':
                                    countLine = len(line)
                                    # print(countLine)
                                    if (line[0] == '0x77') and (countLine == 14):
                                        time_get_data = round(time.time())
                                        strValueWeight = ''
                                        # print(line)
                                        # print('goodData!!!')
                                        valWdg = 0
                                        for element in line:
                                            if (valWdg > 2 and valWdg < 10):
                                                strElement = str(element)
                                                strElement = strElement[3]
                                                # print('checking strElement: ' + strElement)
                                                strValueWeight = strValueWeight + strElement
                                                # print('checking strValueData: ' + strValueWeight)
                                                if (valWdg == 9):
                                                    for dataElement in strValueWeight:
                                                        if (dataElement == '0'):
                                                            strValueWeight = strValueWeight[1::]
                                                        else:
                                                            break
                                                    if (line[2] == '0x2d' and strValueWeight != ''):
                                                        strValueWeight = ('-' + strValueWeight)

                                            valWdg += 1
                                        if (strValueWeight == ''):
                                            strValueWeight = '0'
                                        intValueWeight = int(strValueWeight)
                                        intValueWeightCur = intValueWeight
                                        if (intValueWeightCur > globalValues.valWeightMaxMeas and intValueWeightCur == intValueWeightOld):
                                            countIter += 1
                                        else:
                                            if (intValueWeightCur != intValueWeightOld and intValueWeightCur > globalValues.valWeightMaxMeas):
                                                if (countIter > 0):
                                                    countIter -= 1
                                            else:
                                                countIter = 0
                                                checkWriteValue = True

                                        if (intValueWeightCur == intValueWeightOld and intValueWeightCur <= globalValues.valWeightToWriteVal):
                                            countIterDefaultBack += 1
                                        else:
                                            if (intValueWeightCur != intValueWeightOld and intValueWeightCur <= globalValues.valWeightToWriteVal):
                                                if (countIterDefaultBack > 0):
                                                    countIterDefaultBack -= 1
                                            else:
                                                countIterDefaultBack = 0

                                        intValueWeightOld = intValueWeightCur
                                        # print('Value weight: ' + str(intValueWeight))

                                    line = []
                                    break

                            global valueWeight

                            globalValues.value = intValueWeightCur

                            if (intValueWeightCur > globalValues.valWeightMaxMeas and check_measure and countIter > 20):
                                is_change_sema_default = True
                                countIterDefaultBack = 0
                                output_queue.put('measure')
                                print('StartingMeasure!!!')
                                check_measure = False

                            if (countIterDefaultBack > 500):
                                countIterDefaultBack = 0
                                globalValues.startDeltaWeight = intValueWeightCur
                                print('WritingDelta' + str(intValueWeightCur))
                                if (is_change_sema_default):
                                    is_change_sema_default = False
                                    output_queue.put('default')
                                    globalValues

                            if (countIter > 400 and checkWriteValue):
                                # output_queue.put('ready_weight')
                                check_change_sema = True
                                countIter = 0
                                checkWriteValue = False
                                intValueWithDelta = int(intValueWeight - globalValues.startDeltaWeight)
                                strValueWithDelta = str(intValueWithDelta)

                                print('WritingDelta' + str(strValueWithDelta))

                                globalValues.str_ready_weight = strValueWithDelta

                                output_queue.put('weight_ready')
                                time.sleep(0.001)

                            if (Debug == False):
                                if (abs(round(time.time()) - time_get_data) >= 5):
                                    self.btnOpenPorts.show()
                                    self.cbWeight.setEnabled(True)
                                    globalValues.writeEventToDBJournalMain('Весы', 'Не удалось установить соединение с терминалом весов НВТ-9')
                                    ser_weight.flush()
                                    ser_weight.close()
                                    print('NotConnectWeight!')
                                    # self.setFixedSize(360, 200)
                                    # self.btnOpenPorts.show()
                                    self.lblStWeightImg.setStyleSheet(
                                        'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                        'image: url(' + pathImgRed + ');};')

                                    self.lblStWeightAndSema.setStyleSheet(
                                        'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                        'image: url(' + pathImgRedCom + ');};')

                                    currentIndexTab = self.advices.currentIndex()
                                    self.advices.setCurrentIndex(0)
                                    self.advices.setCurrentIndex(currentIndexTab)

                                    self.is_good_start_weight = False
                                    if (iter == 0):
                                        self.error_weight_signal.emit()
                                    break

                            if (abs(round(time.time()) - time_start) >= 5 and check_start):
                                check_start = False

                                globalValues.writeEventToDBJournalMain('Весы', 'Успешно установлено соединение с УЦИ весов')

                                self.cbWeight.setEnabled(False)

                                self.is_good_start_weight = True

                                if (self.is_good_start_traffic):
                                    self.btnOpenPorts.hide()
                                    self.lblStWeightAndSema.setStyleSheet(
                                        'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                        'image: url(' + pathImgGreenCom + ');};')

                                # print(str(abs(round(time.time()) - time_start)))
                                self.lblStWeightImg.setStyleSheet(
                                    'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                    'image: url(' + pathImgGreen + ');};')

                                currentIndexTab = self.advices.currentIndex()
                                self.advices.setCurrentIndex(0)
                                self.advices.setCurrentIndex(currentIndexTab)

                                self.is_good_con_weight.emit()

                            if (intValueWeightCur < globalValues.valWeightToWriteVal and check_change_sema):

                                bool_check_write_data = False
                                bool_check_write_car_weight_empty = False

                                if (con.open == False):
                                    try:
                                        con = pymysql.connect(host=globalValues.my_sql_localhost,
                                                              port=globalValues.my_sql_port,
                                                              user=globalValues.my_sql_name,
                                                              passwd=globalValues.my_sql_password,
                                                              db=globalValues.dbMySqlName)

                                        # cur = con.cursor()

                                    except Exception as ex:
                                        globalValues.writeLogData(
                                            'Подключение к БД MySql в потоке обработки данных с УЦИ весов НВТ-9', str(ex))

                                try:
                                    # with con:
                                    with con:
                                        cur = con.cursor()

                                        cur.execute("SELECT * FROM " + globalValues.tblsDB[1])

                                        rows = cur.fetchall()

                                        valWdg = 0

                                        strNumberGRZCur = globalValues.str_grz_weight
                                        strWeigthReady = globalValues.str_ready_weight

                                        strGRZFromDB = ''

                                        # f = open('E:/check.txt', 'w')
                                        # f.write(strNumberGRZCur + ' lala ')
                                        # f.close()
                                        print('checkingBD')

                                        for row in reversed(rows):
                                            bool_check_el_db_weight_empty = 'кг' in str(row[3])
                                            bool_check_el_db_weight_load = 'кг' in str(row[4])
                                            bool_check_el_db_st_zakaz = 'перемещение' in str(row[7])



                                            if ((str(row[2]) == strNumberGRZCur) and (bool_check_el_db_st_zakaz == False)):
                                                print('checking WriteeeeData!')

                                                globalValues.is_read_depth = True

                                                if (bool_check_el_db_weight_empty == False):
                                                    strGRZFromDB = str(row[2])

                                                    valWeightReady = int(strWeigthReady)
                                                    valWeightEmpty = 0
                                                    i = 0
                                                    for el in globalValues.listPoolTS:
                                                        el = str(el).upper()
                                                        if (el == strGRZFromDB.upper()):
                                                            valWeightEmpty = int(globalValues.listWeightEmptyTS[i])
                                                            i += 1
                                                    if globalValues.debugTestSystem:
                                                        valWeightEmpty = 100
                                                    if (abs(valWeightReady - valWeightEmpty) > globalValues.deltaWeight):
                                                        checkMeasure = True
                                                        query = ("UPDATE " + globalValues.tblsDB[1] + " SET weight_empty = (%s) WHERE id= (%s)")
                                                        strWeightEmptyTS = str(valWeightEmpty)
                                                        cur.execute(query, ((strWeightEmptyTS + ' кг.'), str(row[0])))
                                                        con.commit()
                                                        bool_check_el_db_weight_empty = True

                                                if (bool_check_el_db_weight_empty == False):

                                                    strGRZFromDB = str(row[2])
                                                    bool_check_write_data = True
                                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET weight_empty = (%s) WHERE id= (%s)")
                                                    cur.execute(query, ((strWeigthReady + ' кг.'), str(row[0])))
                                                    con.commit()
                                                    globalValues.curTalon = str(row[1])
                                                    bool_check_write_car_weight_empty = True
                                                    globalValues.refreshTblMain = True
                                                    break

                                                    # globalValues.str_grz_weight = ''

                                                else:
                                                    print('checking WriteeeeData222222222!')
                                                    # if (str(row[9]) == 'не измерена'):
                                                    if (bool_check_el_db_weight_load == False):
                                                        strGRZFromDB = str(row[2])
                                                        globalValues.curTalon = str(row[1])
                                                        bool_check_write_data = True

                                                        query = ("UPDATE " + globalValues.tblsDB[1] + " SET weight_load = (%s) WHERE id= (%s)")
                                                        cur.execute(query, ((strWeigthReady + ' кг.'), str(row[0])))
                                                        con.commit()
                                                        if checkMeasure:
                                                            checkMeasure = False
                                                            strValueWeightEmpty = strWeightEmptyTS
                                                        else:
                                                            strValueWeightEmpty = str(row[3])

                                                            strValueWeightEmpty = strValueWeightEmpty[0:len(strValueWeightEmpty) - 4]

                                                        print(strValueWeightEmpty)

                                                        value_delta_weight = abs(int(strWeigthReady) - int(strValueWeightEmpty))

                                                        val_volume = value_delta_weight/1510
                                                        val_volume = round(val_volume, 2)
                                                        strVolume = str(val_volume) + ' м3'

                                                        print(strVolume)

                                                        query = ("UPDATE " + globalValues.tblsDB[1] + " SET weight = (%s) WHERE id= (%s)")
                                                        cur.execute(query, ((str(value_delta_weight) + ' кг.'), str(row[0])))
                                                        con.commit()

                                                        query = ("UPDATE " + globalValues.tblsDB[1] + " SET time_check_out = (%s) WHERE id= (%s)")
                                                        str_time_out = self.strCurTime()
                                                        cur.execute(query, (str_time_out, str(row[0])))
                                                        con.commit()

                                                        query = ("UPDATE " + globalValues.tblsDB[1] + " SET volume = (%s) WHERE id= (%s)")
                                                        cur.execute(query, (strVolume, str(row[0])))
                                                        con.commit()

                                                        str_time_in = str(row[10])
                                                        val_time_job = 0


                                                        str_time_job = self.createTimingInJob(str_time_in, str_time_out)

                                                        query = ("UPDATE " + globalValues.tblsDB[1] + " SET time = (%s) WHERE id= (%s)")
                                                        cur.execute(query, (str_time_job, str(row[0])))
                                                        con.commit()

                                                        query = ("UPDATE " + globalValues.tblsDB[1] + " SET state_order = (%s) WHERE id= (%s)")
                                                        cur.execute(query, ('перемещение', str(row[0])))
                                                        con.commit()



                                                        globalValues.refreshTblMain = True
                                                        break

                                            valWdg += 1

                                        if (bool_check_write_data):
                                            curListRtsp = []
                                            curListIp = []
                                            curListRtsp.append(globalValues.listRtsp[2])
                                            curListIp.append(globalValues.listIp[2])
                                            curListRtsp.append(globalValues.listRtsp[3])
                                            curListIp.append(globalValues.listIp[3])

                                            name_img_weight = 'weight'

                                            print('CheckRtsp!')
                                            print(curListRtsp)



                                            curLstRtspScan = []
                                            curLstIpScan = []
                                            curLstRtspScan.append(globalValues.listRtsp[4])
                                            curLstIpScan.append(globalValues.listIp[4])
                                            curLstRtspScan.append(globalValues.listRtsp[5])
                                            curLstIpScan.append(globalValues.listIp[5])

                                            name_img_scan = 'scan'

                                            print(curLstRtspScan)

                                            th_create_img_weight = threading.Thread(target=createImgsWeight, args=(0, curListRtsp, curListIp, globalValues.pathWeightImg, globalValues.curTalon, name_img_weight, ))
                                            th_create_img_weight.start()

                                            th_create_img_scan = threading.Thread(target= createImgsWeight, args = (0, curLstRtspScan, curLstIpScan, globalValues.pathScanImgs, globalValues.curTalon, name_img_scan, ))
                                            th_create_img_scan.start()

                                            if (bool_check_write_car_weight_empty):
                                                dataInJournal = 'Выполнена запись массы ненагруженного ТС ГРЗ: ' + strGRZFromDB
                                            else:
                                                dataInJournal = 'Выполнена запись массы нагруженного ТС ГРЗ: ' + strGRZFromDB
                                            globalValues.writeEventToDBJournalMain('Весы', dataInJournal)
                                        else:
                                            dataInJournal = 'Не выполнена запись массы TC ГРЗ: ' + strNumberGRZCur
                                            globalValues.writeEventToDBJournalMain('Весы', dataInJournal)

                                        cur.close()
                                except Exception as ex:
                                    globalValues.writeLogData('Функция записи данных массы Весы', str(ex))
                                globalValues.str_grz_weight = ''
                                globalValues.str_ready_weight = ''
                                output_queue.put('default')
                                check_measure = True
                                check_change_sema = False

                        print('ExitThreadWeight!')

                    except Exception as ex:
                        globalValues.writeLogData('Поток обработки данных с весов', str(ex))
                        ser_weight.close()
                        self.btnOpenPorts.show()
                        self.cbWeight.setEnabled(True)
                        self.lblStWeightImg.setStyleSheet(
                            'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgRed + ');};')

                        self.lblStWeightAndSema.setStyleSheet(
                            'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgRedCom + ');};')

                        currentIndexTab = self.advices.currentIndex()
                        self.advices.setCurrentIndex(0)
                        self.advices.setCurrentIndex(currentIndexTab)

                        self.is_good_start_weight = False
                        time.sleep(1)

                    time_start = round(time.time()*100)
                    check_start = False

                if (globalValues.stopAll or globalValues.stopComPortWeight):
                    print('endWeight!!!')
                    break

                if (abs(round(time.time()*100 - time_start) > 500)):
                    iter += 1
                    check_start = True
                    if (iter >= 5):
                        break

                time.sleep(0.1)

            globalValues.outComWeight = True
            strDataInJournal = 'Проблемы соединения. Требуется перезапуск!'
            globalValues.writeEventToDBJournalMain('Весы', strDataInJournal)

        except Exception as ex:
            globalValues.writeLogData('Поток обработки весов', str(ex))
            self.btnOpenPorts.show()
            self.cbWeight.setEnabled(True)
            self.lblStWeightImg.setStyleSheet(
                'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                'image: url(' + pathImgRed + ');};')

            self.lblStWeightAndSema.setStyleSheet(
                'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                'image: url(' + pathImgRedCom + ');};')

            currentIndexTab = self.advices.currentIndex()
            self.advices.setCurrentIndex(0)
            self.advices.setCurrentIndex(currentIndexTab)
            self.is_good_start_weight = False

    def openPortTrafficThread(self, input_queue, portTraffic):
        try:
            iter = 0
            time_start = 0
            check_start = True

            while True:
                if (check_start):
                    try:
                            port_traffic = str(portTraffic)

                            lstPorts = port_traffic.split(' ')
                            port_traffic = '/dev/ttyS' + lstPorts[1]
                            print(port_traffic)

                            ser_traffic = serial.Serial(
                                port=port_traffic, \
                                baudrate=9600, \
                                parity=serial.PARITY_NONE, \
                                stopbits=serial.STOPBITS_ONE, \
                                bytesize=serial.EIGHTBITS, \
                                timeout=0)

                            check_default = True
                            check_default_traffic_left = False
                            check_default_traffic_right = False

                            self.is_good_start_traffic = False

                            Debug = globalValues.debugPorts

                            globalValues.check_sema_1_in = 0
                            globalValues.check_sema_1_out = 0
                            globalValues.check_sema_2_in = 0
                            globalValues.check_sema_2_out = 0

                            values_default_checking = bytearray([0x40, 0x31, 0x30, 0x41, 0x31, 0x0D])
                            values_default_data_green_default_in = bytearray([0x40, 0x31, 0x30, 0x30, 0x33, 0x30, 0x34, 0x0D])
                            values_default_traffic_light_green_out = bytearray([0x40, 0x31, 0x30, 0x30, 0x34, 0x30, 0x35, 0x0D])
                            values_default_traffic_light_all_red = bytearray([0x40, 0x31, 0x30, 0x30, 0x38, 0x30, 0x39, 0x0D])

                            values = values_default_data_green_default_in

                            millisStart = int(round(time.time() * 1000))
                            millis = int(round(time.time() * 1000))

                            value_iter = 0

                            line = []

                            time_finish = 0
                            check_finish = False

                            lineNew = []

                            time_get_data_traffic = round(time.time())
                            time_start = round(time.time())
                            check_start = True
                            data_queue = ''

                            while True:

                                if (globalValues.stopAll):
                                        print('EndComTraffic!')
                                        break

                                if ((int(round(time.time() * 1000)) - millis) > 50):
                                    if ((value_iter % 2) == 1):
                                        millis = int(round(time.time() * 1000))
                                        ser_traffic.write(values)
                                    else:
                                        millis = int(round(time.time() * 1000))
                                        ser_traffic.write(values_default_checking)

                                for c in ser_traffic.read():
                                    h = hex(c)
                                    line.append(h)
                                    lineNew.append(h)
                                    countLine = len(line)
                                    # may be h = 0xd or 0x0D
                                    if h == '0xd':
                                        countLine = len(line)
                                        # may be '0x3E'
                                        # and (countLine == 8)
                                        if (line[0] == '0x3e'):
                                            time_get_data_traffic = round(time.time())

                                        line = []

                                #Debug
                                if (Debug == False):
                                        if (abs(round(time.time()) - time_get_data_traffic) >= 7):
                                            self.cbTraff.setEnabled(True)
                                            self.btnOpenPorts.show()

                                            globalValues.writeEventToDBJournalMain('Весы', 'Не удалось установить соединение с контроллером светофоров')
                                            ser_traffic.flush()
                                            ser_traffic.close()
                                            # self.setFixedSize(360, 200)
                                            # self.btnOpenPorts.show()
                                            # self.lblStTraffImg.setPixmap(QtGui.QPixmap(pathImgRedCom))
                                            self.lblStTraffImg.setStyleSheet('QLabel:!hover { image: url(' + pathImgRed + '); background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));};')
                                            self.lblStWeightAndSema.setStyleSheet(
                                                'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                                'image: url(' + pathImgRedCom + ');};')

                                            currentIndexTab = self.advices.currentIndex()
                                            self.advices.setCurrentIndex(0)
                                            self.advices.setCurrentIndex(currentIndexTab)
                                            if (iter == 0):
                                                self.error_traffic_signal.emit()
                                            self.is_good_start_traffic = False
                                            print('NotConnectTraffic!')
                                            break


                                if (abs(round(time.time()) - time_start) >= 6 and check_start):
                                    check_start = False
                                    self.cbTraff.setEnabled(False)
                                    globalValues.writeEventToDBJournalMain('Весы', 'Успешно выполнено соединение с контроллером светофоров')
                                    self.is_good_start_traffic = True

                                    if (self.is_good_start_weight):
                                        self.btnOpenPorts.hide()
                                        print('goodWorkingWeight!')
                                        self.lblStWeightAndSema.setStyleSheet(
                                            'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                            'image: url(' + pathImgGreenCom + ');};')
                                    self.lblStTraffImg.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                'image: url(' + pathImgGreen + ');};')

                                    currentIndexTab = self.advices.currentIndex()
                                    self.advices.setCurrentIndex(0)
                                    self.advices.setCurrentIndex(currentIndexTab)

                                    self.is_good_con_traffic.emit()

                                value_iter += 1

                                if (self.is_good_start_traffic):

                                    if (values == values_default_traffic_light_green_out):
                                        globalValues.check_sema_1_in = 2
                                        globalValues.check_sema_1_out = 1
                                        globalValues.check_sema_2_in = 2
                                        globalValues.check_sema_2_out = 1
                                    else:
                                        if (values == values_default_traffic_light_all_red):
                                            globalValues.check_sema_1_in = 1
                                            globalValues.check_sema_1_out = 1
                                            globalValues.check_sema_2_in = 1
                                            globalValues.check_sema_2_out = 1
                                        else:
                                            if (values == values_default_data_green_default_in):
                                                globalValues.check_sema_1_in = 1
                                                globalValues.check_sema_1_out = 2
                                                globalValues.check_sema_2_in = 1
                                                globalValues.check_sema_2_out = 2
                                            else:
                                                globalValues.check_sema_1_in = 0
                                                globalValues.check_sema_1_out = 0
                                                globalValues.check_sema_2_in = 0
                                                globalValues.check_sema_2_out = 0

                                trafWeightReady = False
                                trafMeasure = False
                                trafDefault = False

                                if (globalValues.startHandSt):
                                    if (globalValues.trafWeightReady):
                                        globalValues.trafWeightReady = False
                                        values = values_default_traffic_light_green_out
                                        print('changeSemaGreenIn')
                                        globalValues.writeEventToDBJournalMain('Светофоры', 'Переключены светофоры в режим выезд ТС')

                                    if (globalValues.trafMeasure):
                                        globalValues.trafMeasure = False
                                        values = values_default_traffic_light_all_red
                                        print('changeSemaAllRed')
                                        globalValues.writeEventToDBJournalMain('Светофоры','Переключены светофоры в режим запрет движения ТС')

                                    if (globalValues.trafDefault):
                                        globalValues.trafDefault = False
                                        values = values_default_data_green_default_in
                                        print('changeSemaGreenOut')
                                        globalValues.writeEventToDBJournalMain('Светофоры','Переключены светофоры в режим въезд ТС')

                                else:

                                        if (q.empty() == False):
                                            data_queue = input_queue.get()


                                        if (str(data_queue) == 'weight_ready'):
                                            values = values_default_traffic_light_green_out
                                            data_queue = ''
                                            globalValues.writeEventToDBJournalMain('Светофоры','Переключены светофоры в режим выезд ТС')

                                        if (str(data_queue) == 'measure'):
                                            values = values_default_traffic_light_all_red
                                            globalValues.writeEventToDBJournalMain('Светофоры','Переключены светофоры в режим запрет движения ТС')
                                            data_queue = ''

                                        if (str(data_queue) == 'default'):
                                            values = values_default_data_green_default_in
                                            data_queue = ''
                                            globalValues.writeEventToDBJournalMain('Светофоры','Переключены светофоры в режим въезд ТС')
                                            print('checking Default!!!')

                            print('ExitThreadTraffic!')

                    except Exception as ex:
                        globalValues.writeLogData('Поток обработки данных со светофоров', str(ex))
                        ser_traffic.close()
                        self.btnOpenPorts.show()
                        self.cbTraff.setEnabled(True)
                        self.lblStTraffImg.setStyleSheet(
                            'QLabel:!hover { image: url(' + pathImgRed + '); background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));};')
                        self.lblStWeightAndSema.setStyleSheet(
                            'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgRedCom + ');};')

                        currentIndexTab = self.advices.currentIndex()
                        self.advices.setCurrentIndex(0)
                        self.advices.setCurrentIndex(currentIndexTab)
                        self.is_good_start_traffic = False
                        time.sleep(1)

                    time_start = round(time.time() * 100)
                    check_start = False

                if (globalValues.stopAll or globalValues.stopComPortTraffic):
                    print('endTraffic!!!')
                    break

                if (abs(round(time.time()*100 - time_start) > 500)):
                    iter += 1
                    check_start = True
                    if (iter >= 5):
                        break

                time.sleep(0.1)

            globalValues.outComTraf = True
            strDataInJournal = 'Проблемы соединения. Требуется перезапуск!'
            globalValues.writeEventToDBJournalMain('Светофоры', strDataInJournal)

        except Exception as ex:
            globalValues.writeLogData('Поток обработки данных светофоров', str(ex))

            self.btnOpenPorts.show()
            self.cbTraff.setEnabled(True)
            self.lblStTraffImg.setStyleSheet(
                'QLabel:!hover { image: url(' + pathImgRed + '); background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));};')
            self.lblStWeightAndSema.setStyleSheet(
                'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                'image: url(' + pathImgRedCom + ');};')

            currentIndexTab = self.advices.currentIndex()
            self.advices.setCurrentIndex(0)
            self.advices.setCurrentIndex(currentIndexTab)
            self.is_good_start_traffic = False

    def openPanelStorage(self):
        try:
            uiStor = Ui_Storage()
            uiStor.exec_()
            if self.checkLogin():
                self.checkArchive()
        except Exception as ex:
            globalValues.writeLogData('Функция запуска панели настройки архива', str(ex))

    # def get_data_from_text(self, strData, delimiter):
    #     strSearch = ''
    #     val_iter = 0
    #     j_delim = 0
    #     val_index_left = 0
    #     val_index_right = 0
    #     for element in strData:
    #         lengthDelimiter = len(delimiter)
    #         if (lengthDelimiter != 0):
    #             if (element == delimiter[0]):
    #                 val_index_left = val_iter
    #                 j_delim = 0
    #                 for element_del in delimiter:
    #                     if strData[val_iter + j_delim] != element_del:
    #                         break
    #                     print(strData[val_iter + j_delim])
    #                     if ((j_delim+1) == lengthDelimiter):
    #                         strSearch = strData[0:val_index_left]
    #                         lengthData = len(strData)
    #                         strData = strData[(val_index_left+lengthDelimiter):(lengthData)]
    #                         return strSearch, strData
    #                     j_delim += 1
    #
    #         val_iter += 1

    def thread_my(self):
        # imgRed = QtGui.QPixmap('E:/img/Green.png')
        # imgGreen = QtGui.QPixmap('E:/12032020Grunt/mainProjectGrunt/Green.png')

        try:

            value_old_0 = 0
            value_old_1 = 0
            value_old_2 = 0
            value_old_3 = 0

            countOld = 0

            start_time = round(time.time())
            start_time_refresh = round(time.time())
            start_time_lcd_loop = round(time.time())
            start_time_gif = round(time.time())
            checkRefresh = False
            start_time_check = round(time.time())
            checkStart = False

            checkStartRS = True

            try:
                    con = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                               user=globalValues.my_sql_name,
                                               passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)

            except Exception as ex:
                    globalValues.writeLogData('Поток обновления системного меню, подключение к бд MySql', str(ex))
                    # if (globalValues.colorForm == 1):
                    self.lblMovieWhite.hide()
                    # else:
                    #     self.lblMovie.hide()
                    self.advices.setEnabled(True)
                    self.checkFirstStart = False

            # delta_time = 15

            check_call = True

            while True:
                if(globalValues.stopAll):
                        break

                if(globalValues.stopThreadSysMenu):
                    globalValues.stopThreadSysMenu = False
                    break

                # if (abs(round(time.time()) - start_time) > 1):
                #     start_time = round(time.time())
                #     currentIndexTab = self.advices.currentIndex()
                #     self.advices.setCurrentIndex(0)
                #     self.advices.setCurrentIndex(currentIndexTab)

                if ( abs(round(time.time() - start_time_lcd_loop)) > 1):
                        start_time_lcd_loop = round(time.time())
                        valueCur = globalValues.value
                        self.lcdWeight.display(str(valueCur))

                        if (globalValues.check_sema_1_out != value_old_0):
                            if (globalValues.check_sema_1_out == 1):
                                self.lblSemaLeftOutUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgRed + ');};')
                                time.sleep(0.02)
                                self.lblSemaLeftOutDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGray + ');};')
                                time.sleep(0.02)
                            if (globalValues.check_sema_1_out == 2):
                                self.lblSemaLeftOutDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGreen + ');};')
                                time.sleep(0.02)
                                self.lblSemaLeftOutUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGray + ');};')
                                time.sleep(0.02)
                            currentIndexTab = self.advices.currentIndex()
                            self.advices.setCurrentIndex(0)
                            self.advices.setCurrentIndex(currentIndexTab)

                        if (globalValues.check_sema_1_in != value_old_1):
                            if (globalValues.check_sema_1_in == 1):
                                self.lblSemaLeftInUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgRed + ');};')
                                time.sleep(0.02)
                                self.lblSemaLeftInDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGray + ');};')
                                time.sleep(0.02)
                            if (globalValues.check_sema_1_in == 2):
                                self.lblSemaLeftInDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGreen + ');};')
                                time.sleep(0.02)
                                self.lblSemaLeftInUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGray + ');};')
                                time.sleep(0.02)
                            currentIndexTab = self.advices.currentIndex()
                            self.advices.setCurrentIndex(0)
                            self.advices.setCurrentIndex(currentIndexTab)

                        if (globalValues.check_sema_2_in != value_old_2):
                            if (globalValues.check_sema_2_in == 1):
                                self.lblSemaRightInUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgRed + ');};')
                                time.sleep(0.02)
                                self.lblSemaRightInDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGray + ');};')
                                time.sleep(0.02)
                            if (globalValues.check_sema_2_in == 2):
                                self.lblSemaRightInDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGreen + ');};')
                                time.sleep(0.02)
                                self.lblSemaRightInUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGray + ');};')
                                time.sleep(0.02)
                            currentIndexTab = self.advices.currentIndex()
                            self.advices.setCurrentIndex(0)
                            self.advices.setCurrentIndex(currentIndexTab)

                        if (globalValues.check_sema_2_out != value_old_3):
                            if (globalValues.check_sema_2_out == 1):
                                self.lblSemaRightOutUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgRed + ');};')
                                time.sleep(0.02)
                                self.lblSemaRightOutDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGray + ');};')
                                time.sleep(0.02)
                            if (globalValues.check_sema_2_out == 2):
                                self.lblSemaRightOutDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGreen + ');};')
                                time.sleep(0.02)
                                self.lblSemaRightOutUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                        'image: url(' + pathImgGray + ');};')
                                time.sleep(0.02)
                            currentIndexTab = self.advices.currentIndex()
                            self.advices.setCurrentIndex(0)
                            self.advices.setCurrentIndex(currentIndexTab)

                        value_old_2 = globalValues.check_sema_2_in
                        value_old_3 = globalValues.check_sema_2_out
                        # print('wwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwwww' + str(globalValues.check_sema_1_out))

                        value_old_0 = globalValues.check_sema_1_out
                        value_old_1 = globalValues.check_sema_1_in

                        if (globalValues.stopSysPanel == True):
                            globalValues.stopSysPanel = False
                            break

                if (checkStart == False):
                    if (abs(round(time.time()) - start_time_check) > 5):
                        checkStart = True

                if(checkStart or self.refreshTree):
                        if (abs(round(time.time()) - start_time_refresh) > 0.3 or self.checkCamAfterAdd or self.refreshTree):
                            self.refreshTree = False
                            # self.checkCamAfterAdd = False
                            checkRefresh = True
                            start_time_refresh = round(time.time())
                            try:
                                    if(con.open == False):
                                                    con = pymysql.connect(host=globalValues.my_sql_localhost,
                                                                          port=globalValues.my_sql_port,
                                                                          user=globalValues.my_sql_name,
                                                                          passwd=globalValues.my_sql_password,
                                                                          db=globalValues.dbMySqlName)

                                    # with con:
                                    if True:
                                        numCams = 4
                                        # numCams = 6
                                        cur = con.cursor()
                                        cur.execute("SELECT * FROM " + globalValues.tblsDB[3])
                                        con.commit()
                                        rows = cur.fetchall()
                                        countNew = 0
                                        for row in rows:
                                                countNew += 1

                                        if (countNew != countOld or checkRefresh):

                                            self.listIDSetCam.clear()
                                            self.listNameChannel.clear()

                                            for z_0 in range(numCams):
                                                    self.treeWidget.topLevelItem(z_0).takeChildren().clear()

                                            for row in rows:
                                                self.listIDSetCam.append(str(row[0]))
                                                self.listNameChannel.append(str(row[3]))
                                                for i_0 in range(numCams):
                                                    checkBreak = False
                                                    dataTopItemTree = self.treeWidget.topLevelItem(i_0).text(0)
                                                    if(str(row[3]).lower() == dataTopItemTree.lower()):
                                                        curItem = self.treeWidget.topLevelItem(i_0)

                                                        for j_0 in range(2):
                                                            data = str(self.listArtTblCam[j_0]) + str(row[j_0 + 1])
                                                            QtWidgets.QTreeWidgetItem(curItem, [data])
                                                            checkBreak = True
                                                    if(checkBreak):
                                                        break

                                            # print(self.listIDSetCam)

                                            countOld = countNew

                                        cur.close()

                            except Exception as ex:
                                    globalValues.writeLogData('Поток обновления системного меню, проверка записей настроек камер в бд MySql', str(ex))

                if (self.checkFirstStart):

                    # if (abs(round(time.time()) - start_time_gif) > globalValues.timeLoadSys and checkStartRS):
                    #
                    #     print('starting Cam Measuring!')
                    #
                    #     checkStartRS = False
                    #
                    #     th_measuring = threading.Thread(target=self.measuringVol)
                    #     th_measuring.start()

                    if (abs(round(time.time()) - start_time_gif) > globalValues.timeLoadSys + globalValues.delta_time):

                        print('CheckingFinish!')

                        self.lblMovieWhite.hide()

                        # self.lblMovie.hide()
                        self.advices.setEnabled(True)
                        self.checkFirstStart = False
                        if (globalValues.debugCamScan):
                            curItem4 = self.treeWidget.topLevelItem(4)
                            curItem4.setIcon(0, QtGui.QIcon(globalValues.pathStyleImgs + 'iconcamgreen.png'))
                            curItem5 = self.treeWidget.topLevelItem(5)
                            curItem5.setIcon(0, QtGui.QIcon(globalValues.pathStyleImgs + 'iconcamgreen.png'))
                            self.goodLoginCam[4] = True
                            self.goodLoginCam[5] = True
                        self.autoStartArchive()
                        if self.checkLogin():
                            self.checkArchive()

                        if globalValues.debugPorts:
                            print('CheckingFinish!')
                            if self.goodLoginDBMySQL or self.goodLoginDBPgSQL:
                                print('CheckingFinish!')
                                if self.is_good_start_weight or self.is_good_start_traffic:
                                    print('CheckingFinish!')
                                    globalValues.stopSysPanel = True
                                    globalValues.stopCamInSet = True
                                    self.hide()
                            print('checkingExit!!!')
                            print(self.goodLoginDBPgSQL)
                            print(self.goodLoginDBMySQL)
                            print(self.is_good_start_traffic)
                            print(self.is_good_start_weight)

                        else:
                            if self.goodLoginDBMySQL and self.goodLoginDBPgSQL:
                                if self.is_good_start_weight and self.is_good_start_traffic:
                                    globalValues.stopSysPanel = True
                                    globalValues.stopCamInSet = True
                                    self.hide()

                        if globalValues.debugPortWeight:
                            globalValues.stopSysPanel = True
                            globalValues.stopCamInSet = True
                            self.hide()

                if (globalValues.stateScaner and self.advices.currentIndex() == 4 and check_call):
                    print('CheckingCam!!!')
                    check_call = False
                    globalValues.stopVideoScaner = False
                    th_video_scan = threading.Thread(target= self.showVideoScaner)
                    th_video_scan.start()
                elif (globalValues.stateScaner == False or self.advices.currentIndex() != 4):
                    globalValues.stopVideoScaner = True
                    check_call = True

                print(self.advices.currentIndex())

                sleep(0.1)

            con.close()

        except Exception as ex:
            globalValues.writeLogData('Поток обработки данных системного меню', str(ex))

    def showVideoScaner(self):
        try:

            self.lblVideoScan.setPixmap(QtGui.QPixmap(self.pathDefNotCamScan))

            globalValues.stopVideoScaner = False

            cap = cv2.VideoCapture(2)

            cap.set(cv2.CAP_PROP_BUFFERSIZE, 2)

            # cap.set()
            frame_rate = 16
            width_default = 300
            koef_ratio = 640 / width_default
            height_default = int(480 / koef_ratio)
            prev = 0
            # width_full = 1670
            # height_full = 940
            # print(str(koef_ratio))

            print('StartingVideo!')

            print(cap.isOpened())
            zx = 0
            while (globalValues.stopAll == False and globalValues.stopVideoScaner == False):

                ret, frame = cap.read()
                # print(ret)
                if ret:

                    time_elapsed = time.time() - prev

                    if time_elapsed > 1. / frame_rate:
                        zx += 1
                        prev = time.time()

                        frame = cv2.resize(frame, (width_default, height_default))
                        rgbImage = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                        h, w, ch = rgbImage.shape
                        bytesPerLine = ch * w
                        imgQt = QImage(rgbImage.data, w, h, bytesPerLine, QImage.Format_RGB888)
                        pixCam = QtGui.QPixmap(imgQt)
                        # cv2.imwrite('E:/' + str(zx) + '.jpeg', frame)
                        self.lblVideoScan.setPixmap(pixCam)

                else:
                    self.lblVideoScan.setPixmap(QtGui.QPixmap(self.pathDefNotCamScan))
                    break



            self.lblVideoScan.setPixmap(QtGui.QPixmap(self.pathDefNotCamScan))
            # cap.destroyAllWindows()
            cap.release()
            self.lblVideoScan.updateGeometry()
            self.lblVideoScan.update()

            print('EndingVideo!')

        except Exception as ex:
            globalValues.writeLogData('Поток возведения видео со сканера', str(ex))

    def deleteCamSet(self):
            # self.movie.stop()
            try:

                indexItem = int(self.treeWidget.currentIndex().row())
                print('checkingDeleteItem!' + str(indexItem))
                curItem = self.treeWidget.topLevelItem(indexItem)
                globalValues.stopCamInSet = True
                self.listRtsp[indexItem] = ''
                globalValues.rtspCams[indexItem] = ''
                globalValues.ipsCams[indexItem] = ''

                if (self.checkMySql):
                        self.con = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                                   user=globalValues.my_sql_name,
                                                   passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)

                with self.con:
                    cur = self.con.cursor()
                    query = ("DELETE FROM " + globalValues.tblsDB[3] + " where channel = (%s)")
                    strNameChannel = self.treeWidget.topLevelItem(indexItem).text(0)

                    if (strNameChannel != ''):
                            cur.execute(query, (strNameChannel))
                            self.con.commit()
                            globalValues.lstGoodCams[indexItem] = False
                            strDataToJurnal = 'Удалены сетевые настройки камеры для канала: ' + strNameChannel
                            globalValues.writeEventToDBJournalMain('Камеры', strDataToJurnal)

                    cur.close()

                curItem.setIcon(0, QtGui.QIcon(globalValues.pathStyleImgs + 'iconcamred.png'))
                self.btnAddCam.setEnabled(True)
                if (self.goodLoginCam[indexItem]):
                        globalValues.boolCheckInsertPassword[indexItem + 2] = True
                self.goodLoginCam[indexItem] = False
                self.insertStatusRed(indexItem, pathImgRedCom)



            except Exception as ex:
                print(str(ex))
                globalValues.writeLogData('Функция удаления сетевых настроек камеры', str(ex))

    def openCamVideo(self):
        # print('123213123123213')
        # self.lblMovieWhite.show()
        # print('123')
        try:
            if (self.checkCamAfterAdd):
                print('InAfterAddCam')
                self.checkCamAfterAdd = False
                curIndex = self.indexAddCam
                curItem = self.treeWidget.topLevelItem(curIndex).text(0)
                print(curItem)
                curItemEl = self.treeWidget.topLevelItem(curIndex)
                globalValues.isNextConCam = True
                time.sleep(1)
            else:

                curIndex = int(self.treeWidget.currentIndex().row())
                curItem = self.treeWidget.currentItem().text(0)
                curItemEl = self.treeWidget.topLevelItem(curIndex)
            print('CurIndex: ' + str(curIndex))

            checkTop = False


            for strNameChlUse in self.listChannel:
                    if(strNameChlUse == curItem):
                        checkTop = True
                        break

            if ((curIndex != self.indexCamClick and checkTop) or globalValues.isNextConCam):
                globalValues.isNextConCam = False
                if (self.indexCamClick != 100):
                    print('checkingClickCod')
                    globalValues.stopCamInSet = False
                    rtsp_link = ''
                    rtsp_link_main = ''
                    ipDev = ''
                    checkIn = False
                    try:
                            print('In1!!!')

                            if (self.checkMySql):
                                    self.con = pymysql.connect(host=globalValues.my_sql_localhost,
                                                               port=globalValues.my_sql_port,
                                                               user=globalValues.my_sql_name,
                                                               passwd=globalValues.my_sql_password,
                                                               db=globalValues.dbMySqlName)

                            with self.con:
                                    print('In2!!!')
                                    print('CheckingInConopenCamVideo!')
                                    cur = self.con.cursor()
                                    cur.execute("SELECT * FROM " + globalValues.tblsDB[3])
                                    print('In3!!!')
                                    rows = cur.fetchall()
                                    print('In4!!!')
                                    # print(curItem)

                                    for row in reversed(rows):
                                            print(row)
                                            if (curItem == row[3]):
                                                 i = curIndex
                                                 # j = 0
                                                 #
                                                 # print('CurIndex: ' + str(i))

                                                 # if (i == 0):
                                                 #     j = i
                                                 # elif (i == 1):
                                                 #     j = i + 1
                                                 # elif (i == 2):
                                                 #     j = i + 2
                                                 # elif (i == 3):
                                                 #     j = i + 3
                                                 # else:
                                                 #     j = i + 3

                                                 j = 2*i


                                                 dataLogin = ''
                                                 dataPassWd = ''

                                                 if (self.passwdLoginCams[i][0] == '' and self.passwdLoginCams[i][1] == ''):
                                                         dataLogin = str(globalValues.loginDataList[j + 4])
                                                         dataPassWd = str(globalValues.loginDataList[j + 5])
                                                 else:
                                                         dataLogin = self.passwdLoginCams[i][0]
                                                         dataPassWd = self.passwdLoginCams[i][1]
                                                         print('checking: ' + str(dataLogin) + str(dataPassWd))

                                                 if (globalValues.oneCheckCamEdit):
                                                     globalValues.oneCheckCamEdit = False
                                                     if (globalValues.debugCam):

                                                         rtsp_link = 'rtsp://' + globalValues.ipCam + ':554/user=' + dataLogin + '&password=' + dataPassWd + '&channel=1&stream=0.sdp'
                                                     else:
                                                         rtsp_link = globalValues.rtspMainLink[curIndex]
                                                         rtsp_link_main = globalValues.rtspMainLink[curIndex+6]
                                                         print('ChekingRtsp: ')
                                                         print(globalValues.ipCam)
                                                         rtsp_link = rtsp_link.replace('strIP', globalValues.ipCam).replace('login', dataLogin).replace('pwd', dataPassWd)
                                                         rtsp_link_main = rtsp_link_main.replace('strIP', globalValues.ipCam).replace('login', dataLogin).replace('pwd', dataPassWd)
                                                         print(rtsp_link)
                                                         print(rtsp_link_main)
                                                         # rtsp_link = 'rtsp://' + dataLogin + ':' + dataPassWd + '@' + globalValues.ipCam + ':554/cam/realmonitor?channel=1&subtype=1'
                                                     ipDev = globalValues.ipCam
                                                 else:
                                                     if (globalValues.debugCam):
                                                         rtsp_link = 'rtsp://' + str(row[2]) + ':554/user=' + dataLogin + '&password=' + dataPassWd + '&channel=1&stream=0.sdp'
                                                     else:
                                                         rtsp_link = globalValues.rtspMainLink[curIndex]
                                                         rtsp_link_main = globalValues.rtspMainLink[curIndex + 6]
                                                         print('ChekingRtsp: ')
                                                         print(str(row[2]))
                                                         rtsp_link = rtsp_link.replace('strIP', str(row[2])).replace('login', dataLogin).replace('pwd', dataPassWd)
                                                         rtsp_link_main = rtsp_link_main.replace('strIP',globalValues.ipCam).replace('login', dataLogin).replace('pwd', dataPassWd)
                                                         print(rtsp_link)
                                                         print(rtsp_link_main)
                                                     ipDev = str(row[2])
                                                 print(rtsp_link)
                                                 checkIn = True
                                                 break
                                    cur.close()
                    except Exception as ex:
                           print(str(ex))
                    print(str(ipDev) + 'iiiiiiiiiiiiippppppppppppppp' + str(curIndex))
                    print(checkIn)
                    if (checkIn):
                        self.threadingCams(curIndex, rtsp_link, rtsp_link_main, ipDev, False)
                    else:
                        print("NotData!!! InDBrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrrr" + str(curIndex))
                        if globalValues.debugCamScan:
                            curItemEl.setIcon(0, QtGui.QIcon(globalValues.pathStyleImgs + 'iconcamgreen.png'))
                        else:
                            curItemEl.setIcon(0, QtGui.QIcon(globalValues.pathStyleImgs + 'iconcamred.png'))
                        self.lblMovieWhite.hide()
                        if (self.goodLoginCam[curIndex]):
                                globalValues.boolCheckInsertPassword[curIndex + 2] = True
                                print('WriteFalse!!!')
                        self.goodLoginCam[curIndex] = False
                        self.insertStatusRed(curIndex, pathImgRedCom)
                        globalValues.stopCamInSet = True
                self.indexCamClick = curIndex

            if (curIndex == 0):
                x = "Имя" in curItem
                if (x):
                     self.indexCamClick = 100

            print('checkStCam: ' + str(globalValues.lstGoodCams))

        except Exception as ex:
            globalValues.writeLogData('Функция открытия камеры ', str(ex))

    def autoStartArchive(self):
        try:
            pathFolder = ''

            try:
                pathData = globalValues.pathDefFldr + '/dataStrg.txt'
                file_strg = open(pathData, 'r')
                data = file_strg.read()
                file_strg.close()
                if (data != ''):
                    dataLst = data.split(' ')
                    pathFolder = dataLst[0]
                    if (int(dataLst[1]) == 0):
                        globalValues.stateStorage = False
                    elif(int(dataLst[1]) == 1):
                        globalValues.stateStorage = True
                    if (dataLst[2] != ''):
                        globalValues.numberFolderDateSave = int(dataLst[2])

            except Exception as ex:
                globalValues.writeLogData('Функция записи состояния архива в файл', str(ex))

            if globalValues.stateStorage:
                listRtsp = []
                listIp = []
                listChnl = []
                listNum = []
                pathDefFolder = globalValues.diskForTimeFiles
                delta = globalValues.deltaWriting
                name_channel = ''
                print('checkingRtsp: ' + str(globalValues.listRtsp))
                for i in range(4):
                    if globalValues.lstGoodCams[i]:
                        if (globalValues.listRtsp[i] != ''):
                            listRtsp.append(globalValues.listRtsp[i])
                            listIp.append(globalValues.listIp[i])
                            if (i == 2):
                                name_channel = 'ВесыВъезд'
                            elif (i == 3):
                                name_channel = 'ВесыВыезд'
                            elif (i == 1):
                                name_channel = 'КппВыезд'
                            elif (i == 0):
                                name_channel = 'КппВъезд'
                            else:
                                name_channel = 'КаналЗаписи'
                            listChnl.append(name_channel)
                            listNum.append(i)

                if (len(listIp) > 0 and pathFolder != ''):
                    print('CheckingData!!!!')
                    print(listRtsp)
                    print(listIp)
                    print(listChnl)
                    print(listNum)
                    globalValues.stopStrg = False

                    # listRtsp[2] = 'rtsp://admin:qwe123456@10.2.165.101:554/Streaming/Channels/102'
                    # listRtsp[3] = 'rtsp://admin:qwe123456@10.2.165.102:554/Streaming/Channels/102'

                    th_start_writing_cams = threading.Thread(target=wrtCams.startingWritingCams, args=(pathFolder, pathDefFolder, delta, listRtsp, listIp, listChnl, listNum,))
                    th_start_writing_cams.start()
                    globalValues.stateStorage = True


                    try:
                        pathData = str(os.getenv('APPDATA')) + r'\Sinaps\dataStrg.txt'
                        file_strg = open(pathData, 'w')
                        dataAll = str(globalValues.pathStrg) + ' 1 ' + str(globalValues.numberFolderDateSave)
                        file_strg.write(dataAll)
                        file_strg.close()
                    except Exception as ex:
                        globalValues.writeLogData('Функция записи состояния архива в файл', str(ex))

                else:
                    try:
                        pathData = str(os.getenv('APPDATA')) + r'\Sinaps\dataStrg.txt'
                        file_strg = open(pathData, 'w')
                        dataAll = str(globalValues.pathStrg) + ' 0 ' + str(globalValues.numberFolderDateSave)
                        file_strg.write(dataAll)
                        file_strg.close()
                        globalValues.stateStorage = False
                    except Exception as ex:
                        globalValues.writeLogData('Функция записи состояния архива в файл', str(ex))


        except Exception as ex:
            globalValues.writeLogData('Функция автозапуска архива', str(ex))

    def closeEvent(self, event):
        # print('closePanel!!!')
        globalValues.stopSysPanel = True
        globalValues.stopCamInSet = True
        globalValues.stopVideoScaner = True
        # self.closePanel.emit()
        # print('closePanel!!!')
        event.accept()
        self.hide()

    def threadingCams(self, numCam, rtsp_link, rtsp_link_main, ip, firstCheck):
        # if (globalValues.colorForm == 1):

        # else:
        #     self.lblMovie.show()
        self.advices.setEnabled(False)
        try:
            if (firstCheck == False):
                globalValues.stopCamInSet = True
                sleep(0.25)
            globalValues.stopCamInSet = False

            print('startCam!!!')
            self.lblMovieWhite.show()

            self.listRtsp[numCam] = rtsp_link
            self.listRtsp[numCam+6] = rtsp_link_main
            globalValues.rtspCams[numCam] = rtsp_link
            globalValues.ipsCams[numCam] = ip
            th_out_gard = threading.Thread(target=self.setVideoCam, args=(rtsp_link, numCam, ip, firstCheck, ))
            th_out_gard.start()
            sleep(0.1)

        except Exception as ex:
            strDataInLog = 'Функция запуска потока работы камеры ' + str(ip)
            globalValues.writeLogData(strDataInLog, str(ex))

    def insertStatusRed(self, numCam, path):
            if (numCam == 0):
                    self.is_good_cam_0 = False
                    globalValues.camsSt[0] = False
                    self.lblCamOut.setStyleSheet(
                            'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + path + ');};')
            elif (numCam == 1):
                    self.is_good_cam_1 = False
                    globalValues.camsSt[1] = False
                    self.lblCamIn.setStyleSheet(
                            'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + path + ');};')
            elif (numCam == 2):
                    self.is_good_cam_2 = False
                    globalValues.camsSt[2] = False
                    self.lblCamWeight1.setStyleSheet(
                            'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + path + ');};')
            elif (numCam == 3):
                    self.is_good_cam_3 = False
                    globalValues.camsSt[3] = False
                    self.lblCamWeight2.setStyleSheet(
                            'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + path + ');};')
            # elif (numCam == 4):
            #         self.is_good_cam_4 = False
            #         globalValues.camsSt[4] = False
            #         self.lblCamScan1.setStyleSheet(
            #                 'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
            #                 'image: url(' + path + ');};')
            # elif (numCam == 5):
            #         self.is_good_cam_5 = False
            #         globalValues.camsSt[5] = False
            #         self.lblCamScan2.setStyleSheet(
            #                 'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
            #                 'image: url(' + path + ');};')

    def setVideoCam(self, rtsp_link, numCam, ip, firstCheck):

        try:
                # globalValues.lstIpCams[numCam] = ''

                globalValues.lstGoodCams[numCam] = False

                # if (globalValues.debugCamScan):
                #     curItem4 = self.treeWidget.topLevelItem(4)
                #     curItem4.setIcon(0, QtGui.QIcon(globalValues.pathStyleImgs + 'iconcamgreen.png'))
                #     curItem5 = self.treeWidget.topLevelItem(5)
                #     curItem5.setIcon(0, QtGui.QIcon(globalValues.pathStyleImgs + 'iconcamgreen.png'))
                #     self.goodLoginCam[4] = True
                #     self.goodLoginCam[5] = True

                isGoodIP = False
                if (len(ip) >= 7):
                        isGoodIP = self.checkCamInNetwork(ip)

                if (isGoodIP):
                        # print('goodCheckingTCPCam!')
                        print('InCamCheckFrame' + str(ip))
                        frame_rate = 20
                        prev = 0
                        curItem = self.treeWidget.topLevelItem(numCam)
                        # rtsp_my = self.rtsp_link

                        # os.environ["OPENCV_FFMPEG_CAPTURE_OPTIONS"] = "rtsp_transport;udp"

                        start_time = round(time.time())
                        num_not_con = 0
                        is_good_con = False

                        self.updateGeometry()

                        while True:
                            if (globalValues.stopThreadCams):
                                print('StopCams')
                                break

                            if (globalValues.stopCamInSet):
                                print('StopCamsSet')
                                break

                            print('IndexEl: ' + str(numCam))

                            print('startOpening' + str(ip))
                            start_time_open = round(time.time())
                            print(rtsp_link)
                            # cap = cv2.VideoCapture(rtsp_link)
                            cap = cv2.VideoCapture(rtsp_link, cv2.CAP_FFMPEG)

                            cap.set(cv2.CAP_PROP_BUFFERSIZE, 2)

                            # if (cap.isOpened()):
                            #     print('Good!!!')

                            # while True:
                            #     ret, frame = cap.read()
                            #     print(ret)
                            #     if ret == True:
                            #         # if True:
                            #         #     img = cv2.imread('/home/sergey/dog.jpg')
                            #         cv2.imshow('serhio', frame)
                            #     # cv2.waitKey()
                            #     # cv2.destroyAllWindows()
                            #     cv2.waitKey(5)
                            #     #
                            #     # if ret == False:
                            #     #     break

                            print('EndOpen:' + str(abs(round(time.time()) - start_time_open)))

                            if(cap.isOpened()):
                                    if (self.goodLoginCam[numCam] == False or firstCheck):
                                        print('checking numCam: ' + str(numCam))
                                        globalValues.boolCheckInsertTrueCamPassword[numCam] = True

                                    # globalValues.lstIpCams[numCam] = str(ip)

                                    print('NumCam: ' + str(numCam))

                                    globalValues.listRtsp[numCam] = rtsp_link
                                    globalValues.listIp[numCam] = ip

                                    self.goodLoginCam[numCam] = True

                                    globalValues.lstGoodCams[numCam] = True

                                    curItem.setIcon(0, QtGui.QIcon(globalValues.pathStyleImgs + 'iconcamgreen.png'))
                                    # self.treeWidget.setEnabled(True)
                                    if (numCam == 0):
                                            self.is_good_cam_0 = True
                                            globalValues.camsSt[0] = True
                                            self.lblCamOut.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                                    'image: url(' + pathImgGreenCom + ');};')
                                            globalValues.listRtsp[6] = rtsp_link

                                    elif (numCam == 1):
                                            self.is_good_cam_1 = True
                                            globalValues.camsSt[1] = True
                                            self.lblCamIn.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                                    'image: url(' + pathImgGreenCom + ');};')
                                            globalValues.listRtsp[7] = rtsp_link

                                    elif (numCam == 2):
                                            self.is_good_cam_2 = True
                                            globalValues.camsSt[2] = True
                                            self.lblCamWeight1.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                                    'image: url(' + pathImgGreenCom + ');};')
                                            globalValues.listRtsp[8] = rtsp_link

                                    elif (numCam == 3):
                                            self.is_good_cam_3 = True
                                            globalValues.camsSt[3] = True
                                            self.lblCamWeight2.setStyleSheet(
                                                    'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                                    'image: url(' + pathImgGreenCom + ');};')
                                            globalValues.listRtsp[9] = rtsp_link

                                    # elif (numCam == 4):
                                    #     self.is_good_cam_4 = True
                                    #     globalValues.camsSt[4] = True
                                    #     self.lblCamScan1.setStyleSheet(
                                    #         'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                    #         'image: url(' + pathImgGreenCom + ');};')
                                    #     globalValues.listRtsp[10] = rtsp_link
                                    #
                                    # elif (numCam == 5):
                                    #     self.is_good_cam_5 = True
                                    #     globalValues.camsSt[5] = True
                                    #     self.lblCamScan2.setStyleSheet(
                                    #         'QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                                    #         'image: url(' + pathImgGreenCom + ');};')
                                    #     globalValues.listRtsp[11] = rtsp_link

                                    # globalValues.refreshTblSetCam = True
                                    print('CamIsOpened!')
                                    num_not_con = 0
                                    is_good_con = True
                                    if (self.checkFirstStart == False):
                                        # if (globalValues.colorForm == 1):
                                        #     self.lblMovieWhite.hide()
                                        #     self.lblMovie.hide()
                                        # else:
                                        self.lblMovieWhite.hide()
                                            # self.lblMovie.hide()
                                        self.advices.setEnabled(True)

                                    if(firstCheck):
                                            strDataToJurnal = 'Выполнена успешная проверка камеры для канала: ' + str(curItem.text(0))
                                            globalValues.writeEventToDBJournalMain('Камеры', strDataToJurnal)
                                            break
                                    else:
                                            strDataToJurnal = 'Выполнено успешное открытие камеры для канала: ' + str(curItem.text(0))
                                            globalValues.writeEventToDBJournalMain('Камеры', strDataToJurnal)

                            print('CheckIn!!!123')
                            # cap.set(cv2.CAP_PROP_BUFFERSIZE, 2)
                            # cap.set()
                            width_default = 345
                            koef_ratio = 1980/width_default
                            height_default = int(1020/koef_ratio)
                            width_full = 1670
                            height_full = 940
                            # print(str(koef_ratio))

                            while (globalValues.stopAll == False):
                                # print('SerhioCheck!!!')
                                # if (globalValues.stopAll):
                                #     break
                                if (globalValues.stopThreadCams):
                                    break

                                if (globalValues.stopCamInSet):
                                    break

                                # print('qwertyhg')

                                        # print(numCam)

                                ret, frame = cap.read()

                                print(ret)
                                # print(cap.isOpened())

                                # print(globalValues.stopCamInSet)

                                # print(numCam)
                                time_elapsed = time.time() - prev

                                # cv2.waitKey(5)

                                # print('CamYessss!')

                                if time_elapsed > 1. / frame_rate and ret == True:
                                    prev = time.time()
                                    print('12345')
                                    # if (ret == False):
                                    #     is_good_con = False
                                    #     print('1234567')
                                    #     break

                                        # https://stackoverflow.com/a/55468544/6622587

                                    # p = convertToQtFormat.scaled(800, 600, Qt.KeepAspectRatio)
                                    # pixCam = pixCam.scaled(400, 226, Qt.KeepAspectRatio)
                                    # p = convertToQtFormat
                                        # print('In')
                                    frame = cv2.resize(frame, (width_default, height_default))
                                    rgbImage = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                                    h, w, ch = rgbImage.shape
                                    bytesPerLine = ch * w
                                    imgQt = QImage(rgbImage.data, w, h, bytesPerLine,
                                                       QImage.Format_RGB888)
                                    pixCam = QtGui.QPixmap(imgQt)
                                    self.lblVideoCam.setPixmap(pixCam)
                                    print('InsertData', numCam)
                            if (globalValues.colorForm == 1):
                                self.lblVideoCam.setPixmap(QtGui.QPixmap(self.pathDefCamImgGrey))
                            else:
                                self.lblVideoCam.setPixmap(QtGui.QPixmap(self.pathDefCamImg))

                            cap.release()

                            # self.check_not_connected = True
                            if (globalValues.stopAll):
                                break

                                # print('checkErrorCam!!!')
                            num_not_con += 1
                            print(num_not_con)
                            if(2 < abs(round(time.time()) - start_time) < 6 and cap.isOpened() == False):
                                    print('111111111')
                                    if (self.goodLoginCam[numCam]):
                                            globalValues.boolCheckInsertPassword[numCam + 2] = True
                                    self.goodLoginCam[numCam] = False
                                    if (self.checkFirstStart == False):
                                        # if (globalValues.colorForm == 1):
                                        self.lblMovieWhite.hide()
                                        #     self.lblMovie.hide()
                                        # else:
                                        #     self.lblMovieWhite.hide()
                                        #     self.lblMovie.hide()
                                        self.advices.setEnabled(True)
                                    break


                        if (globalValues.stopThreadCams == False):
                                if (is_good_con == False):
                                    print('llllllllllllllllllllllllllllllllllllll')
                                    curItem.setIcon(0, QtGui.QIcon(globalValues.pathStyleImgs + 'iconcamred.png'))
                                    self.insertStatusRed(numCam, pathImgRedCom)
                                    if (self.goodLoginCam[numCam]):
                                            globalValues.boolCheckInsertPassword[numCam +2] = True
                                    self.goodLoginCam[numCam] = False

                                    # self.treeWidget.setEnabled(True)

                        print('ExitCamInSet: ' + str(numCam))
                else:
                        curItem = self.treeWidget.topLevelItem(numCam)
                        if globalValues.debugCamScan:
                            curItem.setIcon(0, QtGui.QIcon(globalValues.pathStyleImgs + 'iconcamgreencam.png'))
                        else:
                            curItem.setIcon(0, QtGui.QIcon(globalValues.pathStyleImgs + 'iconcamred.png'))
                        self.insertStatusRed(numCam, pathImgRedCom)
                        if (self.goodLoginCam[numCam]):
                                globalValues.boolCheckInsertPassword[numCam + 2] = True
                        self.goodLoginCam[numCam] = False
                        # self.treeWidget.setEnabled(True)
                        print('ExitCam: ' + str(numCam))
                        strDataToJurnal = 'Не удалось открыть камеру для канала: ' + str(curItem.text(0))
                        globalValues.writeEventToDBJournalMain('Камеры', strDataToJurnal)
                        if (self.checkFirstStart == False):
                            # if (globalValues.colorForm == 1):
                            self.lblMovieWhite.hide()
                            #     self.lblMovie.hide()
                            # else:
                            #     self.lblMovieWhite.hide()
                            #     self.lblMovie.hide()
                            self.advices.setEnabled(True)

                self.updateGeometry()
                NumIn = self.advices.currentIndex()
                self.advices.setCurrentIndex(NumIn + 1)
                self.advices.setCurrentIndex(NumIn)
                print('checkStCam: ' + str(globalValues.lstGoodCams))

        except Exception as ex:
            globalValues.writeLogData( 'Функция открытия камеры в системных настройках' ,str(ex))
            self.lblMovieWhite.hide()
            self.advices.setEnabled(True)

    def editDataInCam(self):
            # print('EditCam')
            checkEdit = False
            try:

                    indexItem = int(self.treeWidget.currentIndex().row())
                    strItem = self.treeWidget.topLevelItem(indexItem).text(0)
                    self.listRtsp[indexItem] = ''
                    globalValues.rtspCams[indexItem] = ''
                    globalValues.ipsCams[indexItem] = ''

                    if (self.checkMySql):
                            self.con = pymysql.connect(host=globalValues.my_sql_localhost,
                                                       port=globalValues.my_sql_port,
                                                       user=globalValues.my_sql_name,
                                                       passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)

                    with self.con:
                            cur = self.con.cursor()
                            cur.execute("SELECT * FROM " + globalValues.tblsDB[3])

                            rows = cur.fetchall()

                            valId = 0
                            checkInData = False
                            for row in reversed(rows):
                                    if (strItem == row[3]):
                                            checkInData = True
                                            uiNetCam = Ui_NetworkCam()
                                            uiNetCam.leNewCamName.setText(str(row[1]))
                                            uiNetCam.leNewCamIP.setText(str(row[2]))
                                            # uiNetCam.leNewCamMask.setText(str(row[3]))
                                            # uiNetCam.leNewCamGate.setText(str(row[4]))
                                            # uiNetCam.leNewCamLogin.setText(str(row[5]))
                                            # uiNetCam.leNewCamPass.setText('admin')
                                            # uiNetCam.comboNewCamChlName.clear()
                                            listNameCB = []
                                            for strUseNameChl in self.listNameChannel:
                                                    numItems = uiNetCam.comboNewCamChlName.count()

                                                    for i in range(numItems):
                                                            if (strUseNameChl == uiNetCam.comboNewCamChlName.itemText(i)):
                                                                    uiNetCam.comboNewCamChlName.removeItem(i)
                                                                    break
                                            countCB = uiNetCam.comboNewCamChlName.count()
                                            listNameCB.append(str(row[3]))
                                            for i in range(countCB):
                                                listNameCB.append(uiNetCam.comboNewCamChlName.itemText(i))
                                            uiNetCam.comboNewCamChlName.clear()
                                            for dataCB in listNameCB:
                                                 uiNetCam.comboNewCamChlName.addItem(dataCB)

                                            uiNetCam.labelNewCam.setGeometry(108, 15, 191, 25)
                                            uiNetCam.labelNewCam.setText('Редактирование камеры')

                                            uiNetCam.btnAddNewCam.setText('Сохранить')

                                            uiNetCam.exec_()



                                            if (uiNetCam.checkSaveBtn == True):
                                                    try:
                                                            # self.con = pymysql.connect(
                                                            #         host=globalValues.my_sql_localhost,
                                                            #         port=globalValues.my_sql_port,
                                                            #         user=globalValues.my_sql_name,
                                                            #         passwd=globalValues.my_sql_password,
                                                            #         db=globalValues.dbMySqlName)

                                                            # if (self.checkMySql):
                                                            #     self.con = pymysql.connect(
                                                            #         host=globalValues.my_sql_localhost,
                                                            #         port=globalValues.my_sql_port,
                                                            #         user=globalValues.my_sql_name,
                                                            #         passwd=globalValues.my_sql_password,
                                                            #         db=globalValues.dbMySqlName)
                                                            #
                                                            # with self.con:
                                                            if True:
                                                                    # cur = self.con.cursor()
                                                                    query = ("UPDATE " + globalValues.tblsDB[3] + " SET name = (%s), ip = (%s), channel = (%s) WHERE channel = (%s)")
                                                                    # ( % s, \'B126PE777\', 5135, 9, \'выполняется\')'

                                                                    loginCam = str(uiNetCam.leNewCamLogin.text())

                                                                    passwdCam = str(uiNetCam.leNewCamPass.text())
                                                                    nameChannel = str(uiNetCam.comboNewCamChlName.currentText())

                                                                    # hash = hashlib.sha512(passwordCam.encode())
                                                                    # data = str(hash.hexdigest())

                                                                    globalValues.ipCam = uiNetCam.leNewCamIP.text()

                                                                    cur.execute(query, (uiNetCam.leNewCamName.text(),uiNetCam.leNewCamIP.text(), nameChannel, strItem))

                                                                    self.con.commit()
                                                                    cur.close()
                                                                    if (uiNetCam.comboNewCamChlName.currentText() != strItem):
                                                                            curItem = self.treeWidget.topLevelItem(indexItem)
                                                                            globalValues.lstGoodCams[indexItem] = False
                                                                            curItem.setIcon(0, QtGui.QIcon(globalValues.pathStyleImgs + 'iconcamred.png'))
                                                                            self.insertStatusRed(indexItem, pathImgRedCom)
                                                                            if (self.goodLoginCam[indexItem]):
                                                                                    globalValues.boolCheckInsertPassword[indexItem + 2] = True
                                                                            self.goodLoginCam[indexItem] = False
                                                                    k = 0
                                                                    z = 0

                                                                    for row in self.listChannel:
                                                                        if (row == nameChannel):
                                                                            self.passwdLoginCams[z][0] = loginCam
                                                                            self.passwdLoginCams[z][1] = passwdCam
                                                                            globalValues.encryptDecryptNum[k + 2] = True
                                                                        k += 1
                                                                        z += 1

                                                                    strDataToJurnal = 'Выполнено редактирование сетевых настроек камеры для канала: ' + uiNetCam.comboNewCamChlName.currentText()
                                                                    globalValues.writeEventToDBJournalMain('Камеры', strDataToJurnal)

                                                            globalValues.stopCamInSet = True
                                                            globalValues.isNextConCam = True
                                                            self.refreshTree = True
                                                            globalValues.oneCheckCamEdit = True
                                                            sleep(0.25)
                                                            self.checkCamAfterAdd = True
                                                            i_0 = 0
                                                            for listIt in self.listChannel:
                                                                    if (listIt == str(uiNetCam.comboNewCamChlName.currentText())):
                                                                            print('checkInt!' + str(i_0))
                                                                            self.indexAddCam = i_0
                                                                            # time.sleep(1)
                                                                            break
                                                                    i_0 += 1
                                                            checkEdit = True
                                                    except  Exception as ex:
                                                            print(str(ex))
                                                            globalValues.writeLogData('Функция редактирования сетевых настроек камеры', str(ex))
                                            else:
                                                cur.close()
                                            break

                                    valId += 1
                            if (checkInData == False):
                                print('NotDataInDB')

                            if checkEdit:
                                # sleep(15)
                                self.openCamVideo()
                                sleep(0.1)
                                globalValues.refreshTblSetCam = True

            except Exception as ex:
                    globalValues.writeLogData('Функция редатирования сетевых настроек камеры', str(ex))

    def checkCamInNetwork(self, strIps):

            # serversocket.setsockopt(SOL_SOCKET, SO_REUSEADDR, 1)

            # os.system('sudo chmod u+s `which ping')
            # time.sleep(0.15)
            # os.system('sudo chmod ug+s')

            checkGoodIp = False
            # strIp = '192.168.0.88'
            strIpHTTP = 'http://' + strIps
            start_time = round(time.time())
            print('Ping IP: ' + strIps)
            try:
                dataPing = ping(strIps)
            except Exception as ex:
                globalValues.writeLogData('Общий сбой сети', str(ex))
                return checkGoodIp
            print('qweqweqwe' + str(dataPing))

            if (str(dataPing) == 'False'):
                    checkGoodIp = False
                    print('NotCam' + strIpHTTP)
                    return checkGoodIp

            else:
                    try:
                            data = requests.head(strIpHTTP, verify=False, timeout=4)
                            checkGoodIp = True
                            print('Good opening: ' + strIps)
                            return checkGoodIp
                            # print(abs(round(time.time()) - start_time))
                    except Exception as ex:
                            checkGoodIp = False
                            print('NotCam' + strIpHTTP)

                            return checkGoodIp
                            # print(abs(round(time.time()) - start_time))

    def checkAllCams(self):
            try:
                    if (self.checkMySql):
                            self.con = pymysql.connect(host=globalValues.my_sql_localhost,
                                                       port=globalValues.my_sql_port,
                                                       user=globalValues.my_sql_name,
                                                       passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)
                            print('ConnectionBDInCheckCam!!!')
                    # with self.con:
                    if True:
                            cur = self.con.cursor()
                            cur.execute("SELECT * FROM " + globalValues.tblsDB[3])
                            rows = cur.fetchall()
                            cur.close()
                            # globalValues.lstIpCams = []
                    for i in range(4):
                    # for i in range(6):
                            try:
                                    curItem = self.treeWidget.topLevelItem(i).text(0)
                                    curIndex = i
                                    rtsp_link_cam = ''
                                    rtsp_link_main = ''
                                    ipDev = ''

                                    for row in reversed(rows):
                                        if (curItem == row[3]):
                                            # j = 0
                                            # if (i == 0):
                                            #     j = i
                                            # elif (i == 1):
                                            #     j = i + 1
                                            # elif (i == 2):
                                            #     j = i + 2
                                            # elif (i == 3):
                                            #     j = i + 3
                                            # else:
                                            #     j = i + 3
                                            # else:

                                            j = 2*i

                                            dataLogin = str(globalValues.loginDataList[j+4])
                                            dataPassWd = str(globalValues.loginDataList[j+5])
                                            # globalValues.lstIpCams[i] = str(row[2])
                                            if (globalValues.debugCam):
                                                rtsp_link_cam = 'rtsp://' + str(row[2]) + ':554/user=' + dataLogin + '&password=' + dataPassWd + '&channel=1&stream=0.sdp'
                                            else:
                                                rtsp_link = globalValues.rtspMainLink[curIndex]
                                                rtsp_link_main = globalValues.rtspMainLink[curIndex+6]
                                                print('ChekingRtsp: ')
                                                print(str(row[2]))
                                                rtsp_link_cam = rtsp_link.replace('strIP', str(row[2])).replace('login', dataLogin).replace('pwd', dataPassWd)
                                                rtsp_link_main = rtsp_link_main.replace('strIP', str(row[2])).replace('login', dataLogin).replace('pwd', dataPassWd)
                                                print(rtsp_link_cam)
                                                print(rtsp_link_main)
                                                # rtsp_link_cam = 'rtsp://' + dataLogin + ':' + dataPassWd + '@' + str(row[2]) + ':554/cam/realmonitor?channel=1&subtype=1'

                                            ipDev = str(row[2])
                                            checkIn = True
                                            break
                                    # print('qqqqqqqqqqqqqqq' + str(rtsp_link_cam))
                                    print('IPCam: ' + ipDev)
                                    self.threadingCams(curIndex, rtsp_link_cam, rtsp_link_main, ipDev, True)
                            except Exception as ex:
                                    globalValues.writeLogData('Функция проверки камер ', str(ex))

            except Exception as ex:
                    globalValues.writeLogData('Функция запуска проверки камер видеонаблюдения', str(ex))

    def checkMySql(self):
        try:
                if (self.con.open):
                        return False
                else:
                        return True
        except Exception as ex:
                globalValues.writeLogData('Функция проверки подключения к бд MySql', str(ex))

    def checkFolderPath(self, pathFolder, isHide):
            try:

                    if (os.path.exists(pathFolder) == False):
                        try:
                            os.mkdir(pathFolder)
                            # globalValsSBV.curPathFolderToWrite = pathFolder
                            # if (isHide):
                            #     subprocess.call(['attrib', '+H', pathFolder])
                            return False
                        except Exception as ex:
                            globalValues.writeLogData('Функция проверки и создания папки хранилища', str(ex))
                    else:
                        return True

            except Exception as ex:
                globalValues.writeLogData('Функция проверки ссылки на текущую папку записи', str(ex))

    def writeDataToTxtFile(self, data, path_doc):
        try:
                file = open(path_doc, 'w')
                file.write(data)
                file.close()
        except Exception as ex:
                globalValues.writeLogData('Функция сохранения данных в текстовый файл', str(ex))

    def currentPosMovie(self, widthMovie, heightMovie):
        x = int((self.width() - widthMovie) / 2) + self.pos().x()
        y = int((self.height() - heightMovie) / 2) + self.pos().y()
        return x,y

    def strCurTime(self):
        str_cur_time = str(datetime.datetime.time(datetime.datetime.now()))
        len_cur_time = len(str_cur_time)
        str_cur_time = str_cur_time[0: (len_cur_time - 7)]
        str_cur_time = str_cur_time.replace(':', '.')
        return str_cur_time

    def createTimingInJob(self, str_time_in, str_time_out):

        try:

            strValTimeIn = str_time_in.replace('.', '')
            strValTimeOut = str_time_out.replace('.', '')

            num_in_sec = int(strValTimeIn[len(strValTimeIn) - 2: len(strValTimeIn)])
            num_out_sec = int(strValTimeOut[len(strValTimeOut) - 2: len(strValTimeOut)])

            num_in_min = int(strValTimeIn[len(strValTimeIn) - 4: len(strValTimeIn) - 2])
            num_out_min = int(strValTimeOut[len(strValTimeIn) - 4: len(strValTimeIn) - 2])

            num_in_hour = int(strValTimeIn[len(strValTimeIn) - 6: len(strValTimeIn) - 4])
            num_out_hour = int(strValTimeOut[len(strValTimeIn) - 6: len(strValTimeIn) - 4])

            # print(num_in_sec)
            # print(num_out_sec)
            # print(num_in_hour)
            # print(num_out_hour)
            # print(num_in_min)
            # print(num_out_min)

            val_job_sec = 0
            val_job_min = 0
            val_job_hour = 0

            sub_min = 0
            sub_hour = 0

            if (num_out_sec >= num_in_sec):
                val_job_sec = abs(num_out_sec - num_in_sec)
            else:
                val_job_sec = abs((60 - num_in_sec) + num_out_sec)
                sub_min = 1

            # print(val_job_sec)

            if (num_out_min > num_in_min or (num_out_min == num_in_min and num_out_sec >= num_in_sec)):
                val_job_min = abs(num_out_min - num_in_min)
            else:
                val_job_min = abs((60 - num_in_min) + num_out_min)
                sub_hour = 1

            val_job_min -= sub_min

            # print(val_job_min)

            if (num_out_hour > num_in_hour or (num_out_hour == num_in_hour and num_out_min > num_in_min) or (
                    num_out_hour == num_in_hour and num_out_min == num_in_min and num_out_sec >= num_in_sec)):
                val_job_hour = abs(num_out_hour - num_in_hour)
            else:
                val_job_hour = abs((24 - num_in_hour) + num_out_hour)

            val_job_hour -= sub_hour

            # print(val_job_hour)

            str_sec = str(val_job_sec)
            str_min = str(val_job_min)
            str_hour = str(val_job_hour)

            if (len(str_sec) < 2):
                str_sec = '0' + str_sec

            if (len(str_min) < 2):
                str_min = '0' + str_min

            if (len(str_hour) < 2):
                str_hour = '0' + str_hour

            str_time_job = str_hour + '.' + str_min + '.' + str_sec

            return str_time_job

        except Exception as ex:
            globalValues.writeLogData('Формирование времени работ', str(ex))

    def checkArchive(self):
        try:
            if globalValues.stateStorage:
                strDataTip = 'Для доступа остановите архив.'
                self.btnAddCam.setEnabled(False)
                self.btnAddCam.setToolTip(strDataTip)
                self.btnDelCam.setEnabled(False)
                self.btnDelCam.setToolTip(strDataTip)
                self.btnSetCam.setEnabled(False)
                self.btnSetCam.setToolTip(strDataTip)
            else:
                self.btnAddCam.setEnabled(True)
                self.btnAddCam.setToolTip('')
                self.btnDelCam.setEnabled(True)
                self.btnDelCam.setToolTip('')
                self.btnSetCam.setEnabled(True)
                self.btnSetCam.setToolTip('')

        except Exception as ex:
            globalValues.writeLogData('Функция проверки архива', str(ex))

    def checkLogin(self):
        try:
            if (globalValues.curUserName == 'operator'):
                self.btnAddCam.setEnabled(False)
                self.btnDelCam.setEnabled(False)
                self.btnSetCam.setEnabled(False)
                # self.btnArchive.setEnabled(False)
                self.btnOpenPorts.setEnabled(False)
                self.btnCheckDBPg.setEnabled(False)
                self.btnCheckDBMy.setEnabled(False)
                return False
            else:
                self.btnAddCam.setEnabled(True)
                self.btnDelCam.setEnabled(True)
                self.btnSetCam.setEnabled(True)
                # self.btnArchive.setEnabled(True)
                self.btnOpenPorts.setEnabled(True)
                self.btnCheckDBPg.setEnabled(True)
                self.btnCheckDBMy.setEnabled(True)
                return True

        except Exception as ex:
            globalValues.writeLogData('Функция проверки логина, панель системных настроек', str(ex))

    def startMeasuringVol(self):
        try:
            self.btnConnScan.setEnabled(False)
            globalValues.stopVideoScaner = True
            time.sleep(0.25)
            th_measuring = threading.Thread(target=self.measuringVol)
            th_measuring.start()
            th_video_scan = threading.Thread(target=self.showVideoScaner)
            th_video_scan.start()
        except Exception as ex:
            globalValues.writeLogData('Функция запуска измерения объёма', str(ex))

    def measuringVol(self):
        try:
            obMes = MeasureVol()
            obMes.runCam(self.lblScan, self.btnConnScan)
        except Exception as ex:
            globalValues.writeLogData('Функция измерения объёма', str(ex))

def ping(host):
        """
        Returns True if host (str) responds to a ping request.
        Remember that a host may not respond to a ping (ICMP) request even if the host name is valid.
        """

        # Option for the number of packets as a function of
        param = '-c'

        # Building the command. Ex: "ping -c 1 google.com"
        command = ['ping', param, '1', host]

        return subprocess.call(command) == 0

def checkFolderLongPath(pathFolder):
        try:
            i = 0
            listPath = []
            for element in pathFolder:
                if (element == '/' and i != 0):
                    listPath.append(pathFolder[0:i])
                i += 1
            # listPath = str(pathFolder).split('/')
            # listPath.pop(0)
            listPath.append(pathFolder)

            # print(listPath)

            for elPath in listPath:
                if (os.path.exists(elPath) == False):
                    try:
                        os.mkdir(elPath)
                    except Exception as ex:
                        globalValues.writeLogData('Функция проверки и создания папки хранилища', str(ex))

        except Exception as ex:
            globalValues.writeLogData('Функция проверки ссылки на папку хранилища', str(ex))

def checkCamInNetwork(strIps):
        checkGoodIp = False
        # strIp = '192.168.0.88'
        strIpHTTP = 'http://' + strIps
        # print(strIpHTTP)
        start_time = round(time.time())
        try:
            dataPing = ping(strIps)
        except Exception as ex:
            return checkGoodIp
        if (str(dataPing) == 'False'):
            checkGoodIp = False
            return checkGoodIp
            # print(abs(round(time.time()) - start_time))
        else:
            try:
                # print('goodping')
                data = requests.head(strIpHTTP, verify=False, timeout=4)

                checkGoodIp = True
                # print('Good opening: ' + strIps)
                return checkGoodIp
                # print(abs(round(time.time()) - start_time))
            except Exception as ex:
                checkGoodIp = False
                return checkGoodIp
                # print(abs(round(time.time()) - start_time))

def createImgsWeight(numChanel, listRtsp, listIP, pathSaveFldr, numOrder, nameImg):
        try:
            pathImg = pathSaveFldr + '/' + str(numOrder) + '_' + nameImg

            print('Sergio: ' + str(pathImg))

            checkJob = False

            # if (os.path.exists(pathImg) == False):

            if True:
                checkFolderLongPath(pathSaveFldr)
                checkJob = False
                numCheckIP = 0
                numCv2 = 0
                while True:
                    str_Ip = listIP[numChanel]
                    rtsp_link = listRtsp[numChanel]
                    if (checkCamInNetwork(str_Ip)):

                        # cap = cv2.VideoCapture(rtsp_link, cv2.CAP_FFMPEG)
                        cap = cv2.VideoCapture(rtsp_link)
                        cap.set(cv2.CAP_PROP_BUFFERSIZE, 3)

                        if (cap.isOpened()):
                            while (True):
                                if (globalValues.stopAll):
                                    break
                                ret, frame = cap.read()
                                if ret:
                                    cv2.imwrite(pathImg, frame)
                                    checkJob = True
                                    break
                            # cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)

                        else:
                            numCv2 += 1
                        cap.release()
                    else:
                        if (numChanel == 0):
                            numChanel = 1
                        elif (numChanel == 1):
                            numChanel = 0
                        str_Ip = listIP[numChanel]
                        rtsp_link = listRtsp[numChanel]

                        if (checkCamInNetwork(str_Ip)):
                            # cap = cv2.VideoCapture(rtsp_link, cv2.CAP_FFMPEG)
                            cap = cv2.VideoCapture(rtsp_link)
                            cap.set(cv2.CAP_PROP_BUFFERSIZE, 3)

                            if (cap.isOpened()):
                                while (True):
                                    if (globalValues.stopAll):
                                        break
                                    ret, frame = cap.read()
                                    if ret:
                                        pathImg = pathSaveFldr + '/' + str(numOrder) + '_weight.jpg'
                                        cv2.imwrite(pathImg, frame)
                                        checkJob = True
                                        break
                            else:
                                numCv2 += 1
                            cap.release()
                        else:
                            numCheckIP += 1

                    if (checkJob == False and (numCv2 > 1 or numCheckIP > 2)):
                        break
                    if (checkJob):
                        break

            if checkJob:
                strDataInJournal = 'Выполнено сохранение кадра З/Н #' + str(numOrder)
                globalValues.writeEventToDBJournalMain('Весы', strDataInJournal)

        except Exception as ex:
            globalValues.writeLogData('Функция создания скрина события, весы', str(ex))

class switch(QDialog):
    def __init__(self, value):
        self.value = value  # значение, которое будем искать
        self.fall = False   # для пустых case блоков

    def __iter__(self):     # для использования в цикле for
        """ Возвращает один раз метод match и завершается """
        yield self.match
        raise StopIteration

    def match(self, *args):
        """ Указывает, нужно ли заходить в тестовый вариант """
        if self.fall or not args:
            # пустой список аргументов означает последний блок case
            # fall означает, что ранее сработало условие и нужно заходить
            #   в каждый case до первого break
            return True
        elif self.value in args:
            self.fall = True
            return True
        return False

def decryptDataFromXls(numberRow, numberCol):
        try:
                private_key = """-----BEGIN RSA PRIVATE KEY-----
            MIIEowIBAAKCAQEA0c45dZkOTBkAi/eP4i6Grc1GVPVw1a+dni+ciOqACFISHEls
            PHGUIBIpzAxVUt7Kgc3QPZ9amMewFTz+whoe7HjAf+wtC+ISBDw8s7CQxknbeUAX
            NYeShbtP1l8Mz6sHJmSk97qFMdyICOALIKY6ky0rYxG7AsqCbsVBe2MKGqKKGuho
            c4Ei35vNBVPn3KLNmSqjfKk4QwNmttwP6gEqldZwT6pjz2wgA42CMQYkVEQ4w4Bw
            a0Go/C1AsrwLd7q3tgqM+tAEctICAjNAdTKMhFWMOdUn60yaEFcZPo7EE0dUw5ZY
            Up91fqQBDoxt2D3p0iWmTEKZPgx9Tbtu+Z6/JwIDAQABAoIBACe6f1Lvaq+qRFo8
            xLg1yzb6GglYeMdd++DKbz/V9+ybbeaBWMeRUlVIWzXSWA3bNkmiKX6hwEwR9Bvx
            cuRageSRcRJILLeFVZgLuArmsmN59N9e7YYrZ+l+8L1NPmXMowv4HuzyGuq4MeJM
            Wo8SKyFXelHGN71tj4lePOoadP1Z1eS9MvScAfs0Noek71CiBpdWG9dmPksRSNOl
            y37W3MypfdD/9na+sXoIxCZGoLudY/KheSBZ++m/pZZave/g4xrHFfKRR7xtGNI7
            CKCJNaijqnJK62EEKwv4SWZIarfRBjhTDcbXfz+Kz1PT6NSvzoG9/7YpXFkhCRio
            mEZmNFECgYEA1D5gCtvgz5nXKcMVBRkPDgU4Kv9EPXZwiut0zW6XNZpaF6neXjXm
            Rx633aHRWPzhZLX0jrgkNlCDgmxL9HiVu/erGql9bpwraZi4J7nft9voifKnaXva
            EsOLEGE3k5oVX3xkgkgfqRLEGFppSIqJkdv8CgX+oqJzq8euje/B2lUCgYEA/Q8s
            fhDHQn9GxJjwEKaWR/IU0EDzfOWVit7IBnCjb3v+1RwrAbT55hTFLW3FoILzcX1I
            Tr7kQpyScly1ucjALBAEQOYciEVeFLvAGhoiuJCiahTmQI+4QGrjHD5hhjNkvqrN
            3kDZh986aLYNBLK61mx6Ml1Dyas0b3ZBVrlxZ4sCgYAeTY2O31fYrCFRQB4nLS2+
            FbawROPsVpW47+csUYbbS19jk4hBMTbgnp0n0qu+JdTUeToiil35N0OfgnDRxcmz
            HahbVSmoejmkiP56BYrQiGBKGdAXOmynUy3ut8Kkm1JD4NHE3CFRFXHT/Eyd49HC
            doMktzhk5gbX1tmwQDQQRQKBgDDIrhEXdvJQyvm3agArvSjdeDm1a7sWH0AINpNX
            P4qMYtH+fiP0GYDLXD+nu8N3uyqTtk7H6gUVXf4B9V59Xt6fr9I7CiETDlH858mg
            ZDUkXMsKgGDN0/1HHcUiGXbfjXpcPxerdMQGuqHZBqVzNyWDAAOZiynjgVZDe9EW
            KtCFAoGBAM87MQ8T3pORPACj2nUojbjypT0RskqU0GHmqKFkoo4zqY0IPM/dN8Ac
            DB3iPC4i90SWvJyqLd1MV3xm7+uQ9l/p3Jvo7/ClVlPRKZYHbPieBZABMgc+VQ1k
            xNEm2eKuM7oat7/vTmGsCFuMwd9+4Z7HrfGdoS6ilJRvX4jAzimw
            -----END RSA PRIVATE KEY-----"""

                pathFile = str(os.getenv('APPDATA')) + r'\Sinaps\cryptoSSA.xlsx'
                # excel = win32com.client.Dispatch('Excel.Application')
                # workbook = excel.Workbooks.open(pathFile, True, False, None, '34ubitav')
                # sheet = workbook.Worksheets('Лист1')
                # excel.Visible = False
                # excel.DisplayAlerts = False

                pathXls = globalValues.pathFileXls
                print(pathXls)

                book = openpyxl.load_workbook(pathXls)
                sheets = book.sheetnames
                print('qwer')
                for sheet in sheets:
                    print(str(sheet))
                    print('qwerty!!!')
                    if (str(sheet) == 'Лист1'):
                        sheet_main = book[sheet]
                        dataElement = str(sheet_main.cell(numberRow, numberCol).value)

                # workbook.Close(False)
                # excel.Quit
                # print('WorkingWithXlsx!')
                dataElement = 'None'
                if (dataElement == 'None'):
                        print('ReturnFalseNotEls')
                        return 'FalseDBSSA'
                else:
                        print('decryptDataInEl')
                        print(dataElement)
                        rsa_key = RSA.importKey(private_key)
                        rsa_key = PKCS1_OAEP.new(rsa_key)
                        raw_cipher_data = b64decode(dataElement)
                        decrypted = rsa_key.decrypt(raw_cipher_data)
                        return decrypted.decode('utf8')

        except Exception as ex:
                globalValues.writeLogData('Функция декодирования данных в файле', str(ex))

def createDBMySqlWithTbls(databaseServerIP, dbPort, databaseUserName, databaseUserPassword, databaseName, tblList):

    isDBInTO = False

    try:

        con = pymysql.connect(host=databaseServerIP, port= dbPort, user=databaseUserName, password=databaseUserPassword)

        cur = con.cursor()
        sqlShowDB = "SHOW DATABASES"
        cur.execute(sqlShowDB)
        dbList = cur.fetchall()
        dbName = databaseName.lower()
        for nameDb in dbList:
            isExistDB = dbName in str(nameDb)
            if (isExistDB):
                isDBInTO = True
                break
        if (isDBInTO == False):
            print('CheckingStartCreateBD!!!')
            sqlCrDB = "CREATE DATABASE " + dbName
            cur.execute(sqlCrDB)
            cur.close()
            con.close()
            print('CheckingCreateBD!!!')
            time.sleep(0.1)
            con = pymysql.connect(host=databaseServerIP, port= dbPort, user=databaseUserName, password=databaseUserPassword, db= dbName)
            cur = con.cursor()

            #CreateMainDataMeasuringTbl

            print('Create Tbl 1!')

            sqlCrTbl = "CREATE TABLE " + tblList[1] + " (ID int NOT NULL AUTO_INCREMENT PRIMARY KEY," \
                       " number_order int DEFAULT '0', number_grz VARCHAR(15) DEFAULT 'нет инф-ии', weight_empty VARCHAR(20) DEFAULT 'не измерена'," \
                       " weight_load VARCHAR (20) DEFAULT 'не измерена', weight VARCHAR (20) DEFAULT 'не измерена', volume VARCHAR (20) DEFAULT 'не измерен'," \
                       "state_order VARCHAR (20) DEFAULT 'выполняется', date VARCHAR (15) DEFAULT '00.00.00', time VARCHAR (15) DEFAULT '00:00:00', time_entry VARCHAR (15) DEFAULT '00:00:00'," \
                       "time_check_out VARCHAR (15) DEFAULT '00:00:00', time_entry_bd VARCHAR (15) DEFAULT '00:00:00', time_check_out_bd VARCHAR (15) DEFAULT '00:00:00'," \
                       "weight_bd VARCHAR (20) DEFAULT 'не измерена', volume_bd VARCHAR (20) DEFAULT 'не измерен', name_company VARCHAR (80) DEFAULT 'инф-ия отсутствует'," \
                       " reg_date TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP) DEFAULT CHARSET=utf8 ROW_FORMAT=COMPACT;"

            cur.execute(sqlCrTbl)
            con.commit()

            # CreateAllDataOrder

            print('Create Tbl 2!')

            sqlCrTbl = "CREATE TABLE " + tblList[2] + " (ID int NOT NULL AUTO_INCREMENT PRIMARY KEY," \
                       " car_name VARCHAR(40) DEFAULT 'инф-ия отсутствует', number_order VARCHAR(40) DEFAULT 'инф-ия отсутствует', date_work VARCHAR(20) DEFAULT 'инф-ия отсутствует'," \
                       " name_company VARCHAR (80) DEFAULT 'инф-ия отсутствует', car_model VARCHAR (40) DEFAULT 'инф-ия отсутствует', car_grz VARCHAR (40) DEFAULT 'инф-ия отсутствует'," \
                       " car_driver VARCHAR (40) DEFAULT 'инф-ия отсутствует', driver_document VARCHAR (80) DEFAULT 'инф-ия отсутствует', task_work VARCHAR (500) DEFAULT 'инф-ия отсутствует'," \
                       " time_come_in VARCHAR (40) DEFAULT '00:00:00', time_come_out VARCHAR (40) DEFAULT '00:00:00', name_operator VARCHAR (50) DEFAULT 'инф-ия отсутствует'," \
                       " name_object VARCHAR (20) DEFAULT 'Площадка №', num_talon int DEFAULT 0, reg_date TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP) DEFAULT CHARSET=utf8 ROW_FORMAT=COMPACT;"

            cur.execute(sqlCrTbl)
            con.commit()

            # CreateJournal

            sqlCrTbl = "CREATE TABLE " + tblList[0] + " (ID int NOT NULL AUTO_INCREMENT PRIMARY KEY," \
                       " date VARCHAR(10) DEFAULT 'нет инф-ии', time VARCHAR(30) DEFAULT 'нет инф-ии', object VARCHAR(30) DEFAULT 'нет инф-ии'," \
                       " data VARCHAR (100) DEFAULT 'нет инф-ии'," \
                       " reg_date TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP) DEFAULT CHARSET=utf8 ROW_FORMAT=COMPACT;"

            cur.execute(sqlCrTbl)
            con.commit()

            # CreateSetTblCam

            sqlCrTbl = "CREATE TABLE " + tblList[3] + " (ID int NOT NULL AUTO_INCREMENT PRIMARY KEY," \
                       " name VARCHAR(30) DEFAULT '', ip VARCHAR(30) DEFAULT '', channel VARCHAR(30) DEFAULT ''," \
                       " mask VARCHAR (30) DEFAULT '', gateway VARCHAR (30) DEFAULT ''," \
                       " reg_date TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP) DEFAULT CHARSET=utf8 ROW_FORMAT=COMPACT;"

            cur.execute(sqlCrTbl)
            con.commit()

            # CreateTrackingTbl

            sqlCrTbl = "CREATE TABLE " + tblList[4] + " (ID int NOT NULL AUTO_INCREMENT PRIMARY KEY," \
                       " url VARCHAR(800) DEFAULT ''," \
                       " num_talon int DEFAULT 0, reg_date TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP) DEFAULT CHARSET=utf8 ROW_FORMAT=COMPACT;"

            cur.execute(sqlCrTbl)
            con.commit()

            sqlCrTbl = "CREATE TABLE " + tblList[5] + " (ID int NOT NULL AUTO_INCREMENT PRIMARY KEY," \
                                                      " weight VARCHAR (20) DEFAULT 'не измерена', volume VARCHAR (20) DEFAULT 'не измерен'," \
                                                      "time_entry VARCHAR (15) DEFAULT '00:00:00', time_out VARCHAR (15) DEFAULT '00:00:00', time_travel VARCHAR (15) DEFAULT '00:00:00'," \
                                                      " num_talon int DEFAULT 0, car_grz VARCHAR (15) DEFAULT 'нет инф-ии', name_company VARCHAR (80) DEFAULT 'инф-ия отсутствует'," \
                                                      "date_work VARCHAR(20) DEFAULT 'инф-ия отсутствует'," \
                                                      " reg_date TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP) DEFAULT CHARSET=utf8 ROW_FORMAT=COMPACT;"
            cur.execute(sqlCrTbl)
            con.commit()

            ###Create Tbl TS Pools

            sqlCrTbl = "CREATE TABLE " + tblList[6] + " (ID int NOT NULL AUTO_INCREMENT PRIMARY KEY," \
                                                      " company VARCHAR (50) DEFAULT 'нет инф-ии', name_ts VARCHAR (20) DEFAULT 'нет инф-ии'," \
                                                      "model_ts VARCHAR (30) DEFAULT 'нет инф-ии', reg_number VARCHAR (20) DEFAULT 'нет инф-ии'," \
                                                      "manufact_year int DEFAULT 0, eco_class_fuel int DEFAULT 0," \
                                                      "power_engine_ls int DEFAULT 0, power_engine_kwt int DEFAULT 0," \
                                                      "volume_engine int DEFAULT 0, weight_empty int DEFAULT 0," \
                                                      "max_weight int DEFAULT 0, carrying_capacity int DEFAULT 0," \
                                                      "volume_body int DEFAULT 0, wheel_formula VARCHAR (20) DEFAULT 'нет инф-ии'," \
                                                      " reg_date TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP) DEFAULT CHARSET=utf8 ROW_FORMAT=COMPACT;"

            cur.execute(sqlCrTbl)
            con.commit()

            # # CreateSettingsDB
            #
            # sqlCrTbl = "CREATE TABLE " + tblList[3] + " (ID int NOT NULL AUTO_INCREMENT PRIMARY KEY," \
            #            " host VARCHAR(30) DEFAULT '', port int NOT NULL DEFAULT 3306," \
            #            " reg_date TIMESTAMP NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP)"
            #
            # cur.execute(sqlCrTbl)
            # con.commit()


            cur.close()
            con.close()

            globalValues.writeEventToDBJournalMain('Базы данных', 'База данных MySql успешно создана')

        else:
            # print('DBExist!')
            con.close()

    except Exception as ex:

        globalValues.writeLogData('Функция создания и проверки бд MySQL', str(ex))

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    uiSetMain = Ui_SettingsDevice()
    uiSetMain.show()
    sys.exit(app.exec_())
