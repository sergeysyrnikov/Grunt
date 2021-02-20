import openpyxl
import time
import globalValues
import datetime

pathFile = 'xlsTiming/timing.xlsx'

lst_data_new = []

def dataFromXls(pathFile):

    try:
        book = openpyxl.load_workbook(pathFile)

        sheets = book.sheetnames

        dataListIP = []

        rtsp_link_list = []

        lst_time_data = []

        firstCall = True

        lst_data = []

        for sheet in sheets:
            print(str(sheet))
            if (str(sheet) == 'Лист1'):
                sheet_main = book[sheet]
                i = 1
                while True:
                    # print(sheet_main.cell(i, 1).value)
                    if (sheet_main.cell(i, 1).value != None):
                        date = sheet_main.cell(i, 1).value
                        time_in = str(sheet_main.cell(i, 3).value)
                        time_out = str(sheet_main.cell(i, 4).value)
                        lst_all_data = [date, time_in, time_out]
                        if (sheet_main.cell(i, 1).value == 1):
                            if (firstCall == False):
                                lst_data.append(lst_time_data)
                            else:
                                firstCall = False
                            lst_time_data = []

                        print(sheet_main.cell(i, 1).value)
                        # valueIP = sheet_main.cell(i, 1).value
                        lst_time_data.append(lst_all_data)
                    else:
                        lst_data.append(lst_time_data)
                        break
                    i += 1

        print(lst_data)
        print(len(lst_data))

        return lst_data
    except Exception as ex:
        globalValues.writeLogData('Функция считывания данных их файла эксель', str(ex))

lst_data_new = dataFromXls(pathFile)




def time_cur_in_out(lst_data_timing):
    try:
        dateToday = datetime.date.today().strftime('%d.%m.%Y')
        str_date_today = str(dateToday)

        lstDate = str_date_today.split('.')
        #
        # print(lstDate[0])
        # print(lstDate[1])

        cur_time = datetime.datetime.time(datetime.datetime.now())
        str_cur_time = str(cur_time)

        lstTiming = str_cur_time.split(':')
        hour = lstTiming[0]
        min = lstTiming[1]
        # print(hour)
        # print(min)

        str_cur_time = hour + min

        num_month = int(lstDate[1])
        num_day = int(lstDate[0]) - 1
        # print(lst_data_new[num_month][num_day])

        str_time_in = str(lst_data_timing[num_month][num_day][1])
        str_time_in = str_time_in[0:len(str_time_in) - 3]
        str_time_in = str_time_in.replace(':', '')
        val_time_in = int(str_time_in)
        str_time_out = str(lst_data_timing[num_month][num_day][2])
        str_time_out = str_time_out[0:len(str_time_out) - 3]
        str_time_out = str_time_out.replace(':', '')
        val_time_out = int(str_time_out)

        val_cur_time = int(str_cur_time)

        return val_cur_time, val_time_in, val_time_out

        print(val_cur_time)
        print(val_time_in)
        print(val_time_out)
    except Exception as ex:
        globalValues.writeLogData('Функция создания текущего времени, времни восхода, времени заката', str(ex))


res_def = time_cur_in_out(lst_data_new)

val_cur_time = res_def[0]
val_time_in = res_def[1]
val_time_out = res_def[2]



firstCallDay = True
firstCallNight = True

while True:

    if (val_time_in <= val_cur_time <= val_time_out):
        print('checkDay!!!')

        if firstCallDay:
            print('WorkMyCodOnvifDay')
            firstCallDay = False
            firstCallNight = True
    else:
        print('checkingNight')

        if firstCallNight:
            print('WorkMyCodOnvifNight')
            firstCallDay = True
            firstCallNight = False

    time.sleep(30)


