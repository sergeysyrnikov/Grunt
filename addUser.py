from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QDialog
import globalValues
import time
import sys
from panelMesBox import Ui_mes_box
import os
from base64 import b64decode, b64encode
from Crypto.PublicKey import RSA
from Crypto.Cipher import PKCS1_OAEP
# import win32com.client
import openpyxl

app = QtWidgets.QApplication(sys.argv)

globalValues.colorForm = 0

class Ui_addUser(QDialog):

    lstLight = [[]]
    lstDark = [[]]
    lengthLight = 0
    lengthDark = 0
    delta = 40
    num = 0

    def __init__(self):
        super().__init__()
        self.runUi()

    def runUi(self):
        self.setObjectName("Dialog")
        self.setFixedSize(290, 128)
        self.leLogin = QtWidgets.QLineEdit(self)
        self.leLogin.setGeometry(QtCore.QRect(10, 10, 271, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.leLogin.setFont(font)
        self.leLogin.setObjectName("leLogin")
        self.butInput = QtWidgets.QPushButton(self)
        self.butInput.setGeometry(QtCore.QRect(95, 90, 101, 26))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(10)
        font.setBold(False)
        font.setWeight(50)
        self.butInput.setFont(font)
        self.butInput.setObjectName("butInput")
        self.lePassword = QtWidgets.QLineEdit(self)
        self.lePassword.setGeometry(QtCore.QRect(10, 47, 271, 31))
        font = QtGui.QFont()
        font.setFamily("Arial")
        font.setPointSize(12)
        self.lePassword.setFont(font)
        self.lePassword.setObjectName("lePassword")
        self.label = QtWidgets.QLabel(self)
        self.label.setGeometry(QtCore.QRect(0, 0, 291, 131))
        self.label.setText("")
        self.label.setObjectName("label")
        self.label.raise_()
        self.lePassword.raise_()
        self.butInput.raise_()
        self.leLogin.raise_()

        self.retranslateUi()
        QtCore.QMetaObject.connectSlotsByName(self)
        # self.runThJournal()
        self.firstCall()

    def retranslateUi(self):
        _translate = QtCore.QCoreApplication.translate
        self.setWindowTitle(_translate("Dialog", "Добавление пользователя"))
        # self.leLogin.setText(_translate("Dialog", "Логин"))
        self.butInput.setText(_translate("Dialog", "Добавить"))
        # self.lePassword.setText(_translate("Dialog", "Пароль"))

    def firstCall(self):

        self.leLogin.setPlaceholderText('Логин')
        self.lePassword.setPlaceholderText('Пароль')

        self.lstLight = [[self.leLogin,"background-color: rgb(255,255,255);\n"
"color: rgb(0,0,0);\n"
"border-radius:3px;\n"
"border: 1px solid rgb(150,150,150);"],
        [self.butInput,"QPushButton:!hover {background-color: rgb(227,227,227);\n"
"color: rgb(0,0,0);\n"
"border-radius:3px;\n"
"border:1px solid rgb(135,135,135);}\n"
"QPushButton:hover {background-color: rgb(84,122,181);\n"
"color: rgb(255, 255, 255);\n"
"border-radius:3px;\n"
"border:1px solid rgb(63,63,63);}\n"
"QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
"color: rgb(255, 255, 255);\n"
"border-radius:3px;\n"
"border:1px solid rgb(63,63,63);};"],
        [self.lePassword,"background-color: rgb(255,255,255);\n"
"color: rgb(0,0,0);\n"
"border-radius:3px;\n"
"border: 1px solid rgb(150,150,150);"],
        [self.label,"background-color: rgb(242,242,242);"]]
        self.lstDark = [[self.leLogin,"background-color: rgb(42, 42, 42);\n"
"color: rgb(255, 255, 255);\n"
"border-radius: 3px;"],
        [self.butInput,"QPushButton:!hover {background-color: rgb(89,89,89);\n"
"color: rgb(255, 255, 255);\n"
"border-radius:3px;\n"
"border:1px solid rgb(63,63,63);}\n"
"QPushButton:hover {background-color: rgb(84,122,181);\n"
"color: rgb(255, 255, 255);\n"
"border-radius:3px;\n"
"border:1px solid rgb(63,63,63);}\n"
"QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
"color: rgb(255, 255, 255);\n"
"border-radius:3px;\n"
"border:1px solid rgb(63,63,63);};"],
        [self.lePassword,"background-color: rgb(42, 42, 42);\n"
"color: rgb(255, 255, 255);\n"
"border-radius: 3px;"],
        [self.label,"background-color: rgb(62,62,62);"]]

        self.lePassword.setEchoMode(QtWidgets.QLineEdit.Password)
        self.lengthLight = len(self.lstLight)
        self.lengthDark = len(self.lstDark)
        self.changeCOLORMainPanelGrunt(0.1)

        self.butInput.clicked.connect(self.createNewUser)

    def changeColor(self, object, str):
            object.setStyleSheet(str)

    def changeCOLORMainPanelGrunt(self, timeChangeColor):

        delta = int((timeChangeColor / self.lengthDark) * 1000)

        if (globalValues.colorForm == 1):

            for i in range(self.lengthLight):
                startTime = round(time.time() * 1000)
                while True:
                    obj = self.lstLight[i][0]
                    style = self.lstLight[i][1]
                    self.changeColor(obj, style)
                    # print('changeLight!')
                    if (abs(round(time.time() * 1000) - startTime) > delta):
                        break

        elif (globalValues.colorForm == 0):

            for i in range(self.lengthDark):
                startTime = round(time.time() * 1000)
                while True:
                    # print(delta)
                    # print(abs(round(time.time()*1000) - startTime))
                    obj = self.lstDark[i][0]
                    style = self.lstDark[i][1]
                    self.changeColor(obj, style)
                    # print('changeDark!')
                    if (abs(round(time.time() * 1000) - startTime) > delta):
                        break

    def createNewUser(self):
        try:
            if (self.leLogin.text() != '' and self.lePassword.text() != ''):
                print('Good!')
                pathFile = globalValues.pathDefFldr + '\cryptoSSA.xlsx'
                strLogin = self.leLogin.text()
                strPwd = self.lePassword.text()

                self.encryptUser(strLogin, strPwd, pathFile)

                self.close()
            else:
                uiMes = Ui_mes_box()
                uiMes.lblStrInfo.setText('Некорректная запись данных!')
                uiMes.btnOK.hide()
                if (globalValues.colorForm == 1):
                    uiMes.label_iconques.setStyleSheet("background-color: rgb(235,235,235);\n"
                                                       "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
                else:
                    uiMes.label_iconques.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                       "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
                uiMes.btnCancel.setText('Продолжить')
                uiMes.exec_()
        except Exception as ex:
            globalValues.writeLogData('Функция добавления нового пользователя в систему', str(ex))


    def encryptUser(self, strEncryptLogin, strEncryptPwd, pathFile):
        try:
            if (os.path.exists(pathFile)):
                public_key = """-----BEGIN PUBLIC KEY-----
                MIIBIjANBgkqhkiG9w0BAQEFAAOCAQ8AMIIBCgKCAQEA0c45dZkOTBkAi/eP4i6G
                rc1GVPVw1a+dni+ciOqACFISHElsPHGUIBIpzAxVUt7Kgc3QPZ9amMewFTz+whoe
                7HjAf+wtC+ISBDw8s7CQxknbeUAXNYeShbtP1l8Mz6sHJmSk97qFMdyICOALIKY6
                ky0rYxG7AsqCbsVBe2MKGqKKGuhoc4Ei35vNBVPn3KLNmSqjfKk4QwNmttwP6gEq
                ldZwT6pjz2wgA42CMQYkVEQ4w4Bwa0Go/C1AsrwLd7q3tgqM+tAEctICAjNAdTKM
                hFWMOdUn60yaEFcZPo7EE0dUw5ZYUp91fqQBDoxt2D3p0iWmTEKZPgx9Tbtu+Z6/
                JwIDAQAB
                -----END PUBLIC KEY-----"""
                rsa_key = RSA.importKey(public_key)
                rsa_key = PKCS1_OAEP.new(rsa_key)
                encryptedLogin = rsa_key.encrypt(strEncryptLogin.encode('utf8'))
                encrypted_text_login = b64encode(encryptedLogin)

                encryptedPwd = rsa_key.encrypt(strEncryptPwd.encode('utf8'))
                encrypted_text_pwd = b64encode(encryptedPwd)

            #     excel = win32com.client.Dispatch('Excel.Application')
            #     workbook = excel.Workbooks.open(pathFile, True, False, None, '34ubitav')
            #     sheet = workbook.Worksheets('Лист1')
            #     excel.Visible = False
            #     excel.DisplayAlerts = False

                pathXls = globalValues.pathFileXls
                print(pathXls)

                book = openpyxl.load_workbook(pathXls)
                sheets = book.sheetnames
                print('qwer')
                for sheet in sheets:
                    print(str(sheet))
                    print('qwerty!!!')
                    if (str(sheet) == 'Лист1'):
                        sheet_main = book[sheet]

                        encrypted_text_login = str(encrypted_text_login)
                        lenLogin = len(encrypted_text_login)
                        encrypted_text_login = encrypted_text_login[2: lenLogin - 1]

                        encrypted_text_pwd = str(encrypted_text_pwd)
                        lenPwd = len(encrypted_text_pwd)
                        encrypted_text_pwd = encrypted_text_pwd[2: lenPwd - 1]

                        numEndRow = 0
                        for i in range(1, 1000):
                            print(i)
                            value = sheet_main.cell(i, 1).value
                            print(value)
                            if (value == None):
                                numEndRow = i
                                break
                        # print(numEndRow)
                        # print(encrypted_text_login)
                        # print(encrypted_text_pwd)
                        sheet_main.cell(numEndRow, 1).value = str(encrypted_text_login)
                        sheet_main.cell(numEndRow, 2).value = str(encrypted_text_pwd)

                book.save(pathXls)
                book.close()

                # workbook.Close(True, pathFile)
                # excel.Quit
            else:
                uiMes = Ui_mes_box()
                uiMes.lblStrInfo.setText('Отсутствует файл для записи!')
                uiMes.btnOK.hide()
                if (globalValues.colorForm == 1):
                    uiMes.label_iconques.setStyleSheet("background-color: rgb(235,235,235);\n"
                                                       "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
                else:
                    uiMes.label_iconques.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                       "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
                uiMes.btnCancel.setText('Продолжить')
                uiMes.exec_()
        except Exception as ex:
            globalValues.writeLogData('Функция кодирования данных и сохранения в запароленный файл', str(ex))

if __name__ == "__main__":
    uiPan = Ui_addUser()
    uiPan.show()
    sys.exit(app.exec_())
