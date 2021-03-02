import cv2
import os
import gc
# from cffi.model import global_cache
from onvif import ONVIFCamera
from PyQt5.QtWidgets import QDialog, QMessageBox, QApplication, QSlider, QStyleOptionSlider, QStyle, QFrame, QWidget
from PyQt5.QtCore import QLocale
from PyQt5.QtGui import QImage, QPixmap, QScreen, QFont
from PyQt5.QtMultimedia import QMediaContent, QMediaPlayer
from PyQt5.QtCore import Qt, QUrl, QRect
from PyQt5.QtMultimediaWidgets import QVideoWidget
from panelMapUrl import Ui_DialogMap
from panelLogin import Ui_panel_autologin
from systemsettings import Ui_SettingsDevice
from panelZakaz import PanelZakaz
import pymysql
import threading
import webbrowser
from time import sleep
import psycopg2
import globalValues
from contextTab import Ui_ContextTable
import time
import datetime
import sys
from aboutSystem import Ui_about_system_menu
from aboutSystemMenu import Ui_AboutSysMenu
from panelJournal import Ui_Journal
from panelMode import Ui_select_st_working
from panelMesBox import Ui_mes_box
from PanelJournalTS import Ui_PanelJournalTS
from PanelTSBase import Ui_TS_Base
from PanelAddZN import Ui_AddZakaz
from PanelReportZN import Ui_panel_report
import glob
from openpyxl import load_workbook
from PanelTalon import Ui_Talon
from functools import partial
import requests
from panRtsp import Ui_panRtspCams
import subprocess
# import memory_profiler
# from pympler import muppy, summary
# import ping3
# import openpyxl
# from traits.tests.check_timing import global_value

# Back up the reference to the exceptionhook
sys._excepthook = sys.excepthook

def my_exception_hook(exctype, value, traceback):
    # Print the error and traceback
    print(exctype, value, traceback)
    # Call the normal Exception hook after
    sys._excepthook(exctype, value, traceback)
    sys.exit(1)

# Set the exception hook to our wrapping function
sys.excepthook = my_exception_hook


timeout = 2
check = True

listID = []
listIDZAKAZData = []

is_error_con_weight = False
is_error_con_traffic = False

# rtsp_cam_out_gard = 'rtsp://admin:admin@10.2.165.103:554/cam/realmonitor?channel=1&subtype=0'
# rtsp_cam_in_gard = 'rtsp://admin:admin@10.2.165.104:554/cam/realmonitor?channel=1&subtype=0'
# rtsp_cam_in_weight = 'rtsp://admin:qwe123456@10.2.165.101:554/cam/realmonitor?channel=1&subtype=0'
# rtsp_cam_out_weight = 'rtsp://admin:qwe123456@10.2.165.102:554/cam/realmonitor?channel=1&subtype=0'

pathData = globalValues.pathDefFldr + '/dataComPorts.txt'

pathImgGray = globalValues.pathImage + 'icongreykrug3131.png'
pathImgRed = globalValues.pathImage + 'iconredkrug3131.png'
pathImgGreen = globalValues.pathImage + 'icongreenkrug3131.png'

pathImgRedCom = globalValues.pathImage + 'iconredkrug4141.png'
pathImgGreenCom = globalValues.pathImage + 'icongreenkrug4141.png'

app = QtWidgets.QApplication(sys.argv)

pixCamNotCon = QPixmap(globalValues.pathStyleImgs + 'camDefault400x400.png')




pixCamNotConGrey = QPixmap(globalValues.pathStyleImgs + 'camDefault400x400grey.png')
pixCamNotConFull = QPixmap(globalValues.pathStyleImgs + 'camDefaultFull.PNG')

valueWeight = 12

# I18N_QT_PATH = 'E:/venv/Lib/site-packages/PyQt5/Qt/translations/'

listIp = []
listRtsp = []

firstCallChangeScroll = True

# globalValues.colorForm = 1


class Slider(QtWidgets.QSlider):

    value = 0

    def mousePressEvent(self, event):
        super(Slider, self).mousePressEvent(event)
        if event.button() == Qt.LeftButton:
            val = self.pixelPosToRangeValue(event.pos())
            self.setValue(val)

    def pixelPosToRangeValue(self, pos):
        opt = QtWidgets.QStyleOptionSlider()
        self.initStyleOption(opt)
        gr = self.style().subControlRect(QtWidgets.QStyle.CC_Slider, opt, QtWidgets.QStyle.SC_SliderGroove, self)
        sr = self.style().subControlRect(QtWidgets.QStyle.CC_Slider, opt, QtWidgets.QStyle.SC_SliderHandle, self)

        if self.orientation() == Qt.Horizontal:
            sliderLength = sr.width()
            sliderMin = gr.x()
            sliderMax = gr.right() - sliderLength + 1
        else:
            sliderLength = sr.height()
            sliderMin = gr.y()
            sliderMax = gr.bottom() - sliderLength + 1;
        pr = pos - sr.center() + sr.topLeft()
        p = pr.x() if self.orientation() == Qt.Horizontal else pr.y()
        return QtWidgets.QStyle.sliderValueFromPosition(self.minimum(), self.maximum(), p - sliderMin,
                                               sliderMax - sliderMin, opt.upsideDown)

    # def wheelEvent(self, QWheelEvent):
    #     print('cgewjgjhewgfhjg123')

class Ui_MainFormGrunt(QDialog):

    con = pymysql.connections
    firstCallInThPG = True
    check_full_cam0 = False
    check_full_cam1 = False
    check_full_cam2 = False
    check_full_cam3 = False
    is_discon_cam0 = True
    num__wdg_call_cam0 = 0
    num__wdg_call_cam1 = 0
    num__wdg_call_cam2 = 0
    num__wdg_call_cam3 = 0
    check_error_weight = False
    check_error_traffic = False
    check_good_con_weight = False
    check_good_con_traffic = False
    uiDev = Ui_SettingsDevice()
    # uiContext = Ui_ContextTable()
    is_first_in_openDev = True
    is_set_context_show = False

    is_first_con_menu = True
    item_cur_row = 0
    item_cur_col = 0
    num_search_cat = 0
    list_rows_tbl = []
    countCols = 0
    countRows = 0
    xContext = 0
    yContext = 0
    is_open_pan_context = True
    currentModeStWorking = 0

    rtsp_cam_out_gard = ''
    rtsp_cam_in_gard = ''
    rtsp_cam_in_weight = ''
    rtsp_cam_out_weight = ''

    defCon = 99

    checkLoginEvent = False

    listChannel = ['Камера въезд КПП', 'Камера выезд КПП', 'Камера въезд Весы', 'Камера выезд Весы']

    numCurChannel = 99

    lstLight = [[]]
    lstDark = [[]]

    numColor = 0
    lengthLight = 0
    lengthDark = 0

    refreshOneTblMain = False
    
    refreshTblInSearch = False

    checkStCams = [False, False, False, False, False, False]

    checkHideHandSt = False

    checkFirstCallStWork = True

    exitLogin = False

    firstCheckListStTs = True

    pathFolderVideoStrg = ''

    numTestVideoCr = 0

    pathFileErr = globalValues.pathDefFldr + '/checkErrGrunt.txt'

    pathDefNotCamScan = globalValues.pathStyleImgs + 'camDefault300x300.PNG'

    pos = 0

    # pathFileVideo = "E:/Storage/04.05.2020/192.168.1.88/1.mp4"

    # checkHandSt = False


    #PG

    def __init__(self):
        super().__init__()
        self.runUi()

    def runUi(self):
        self.setObjectName("MainWindow")
        self.setEnabled(True)
        self.setFixedSize(1920, 1080)
        self.setWindowFlags Qt.CustomizeWindowHint | Qt.FramelessWindowHint)
        font = QFont()
        font.setKerning(True)
        self.setFont(font)
        self.setAutoFillBackground(False)
        self.setStyleSheet("")
        self.mainFormPr = QtWidgets.QWidget(self)
        self.mainFormPr.setObjectName("mainFormPr")
        self.frmCam = QtWidgets.QFrame(self.mainFormPr)
        self.frmCam.setGeometry QRect(0, 132, 855, 948))
        self.frmCam.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmCam.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmCam.setObjectName("frmCam")
        self.lblTitleCams = QtWidgets.QLabel(self.frmCam)
        self.lblTitleCams.setGeometry QRect(9, 0, 205, 40))
        font = QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lblTitleCams.setFont(font)
        self.lblTitleCams.setText("")
        self.lblTitleCams.setAlignment Qt.AlignLeading Qt.AlignLeft Qt.AlignTop)
        self.lblTitleCams.setObjectName("lblTitleCams")
        self.lblCatCams = QtWidgets.QLabel(self.frmCam)
        self.lblCatCams.setGeometry QRect(47, 5, 160, 20))
        font = QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lblCatCams.setFont(font)
        self.lblCatCams.setObjectName("lblCatCams")
        self.lblIconCams = QtWidgets.QLabel(self.frmCam)
        self.lblIconCams.setGeometry QRect(19, 5, 23, 23))
        self.lblIconCams.setText("")
        self.lblIconCams.setObjectName("lblIconCams")
        self.lblMainBoxCams = QtWidgets.QLabel(self.frmCam)
        # self.lblMainBoxCams.setEnabled(False)
        self.lblMainBoxCams.setGeometry QRect(9, 30, 841, 901))
        self.lblMainBoxCams.setText("")
        self.lblMainBoxCams.setObjectName("lblMainBoxCams")
        self.lblCam1 = QtWidgets.QLabel(self.frmCam)
        self.lblCam1.setGeometry QRect(444, 35, 160, 30))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(5)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lblCam1.sizePolicy().hasHeightForWidth())
        self.lblCam1.setSizePolicy(sizePolicy)
        font = QFont()
        font.setPointSize(10)
        self.lblCam1.setFont(font)
        self.lblCam1.setObjectName("lblCam1")
        self.lblCam2 = QtWidgets.QLabel(self.frmCam)
        self.lblCam2.setGeometry QRect(25, 486, 160, 30))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(5)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lblCam2.sizePolicy().hasHeightForWidth())
        self.lblCam2.setSizePolicy(sizePolicy)
        font = QFont()
        font.setPointSize(10)
        self.lblCam2.setFont(font)
        self.lblCam2.setObjectName("lblCam2")
        self.lblCam4 = QtWidgets.QLabel(self.frmCam)
        self.lblCam4.setGeometry QRect(444, 486, 160, 30))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(5)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lblCam4.sizePolicy().hasHeightForWidth())
        self.lblCam4.setSizePolicy(sizePolicy)
        font = QFont()
        font.setPointSize(10)
        self.lblCam4.setFont(font)
        self.lblCam4.setObjectName("lblCam4")
        self.lblCam0 = QtWidgets.QLabel(self.frmCam)
        self.lblCam0.setGeometry QRect(25, 35, 160, 30))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(5)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lblCam0.sizePolicy().hasHeightForWidth())
        self.lblCam0.setSizePolicy(sizePolicy)
        font = QFont()
        font.setPointSize(10)
        self.lblCam0.setFont(font)
        self.lblCam0.setObjectName("lblCam0")
        self.leCam0 = QtWidgets.QLabel(self.frmCam)
        # self.leCam0.setEnabled(False)
        self.leCam0.setGeometry QRect(20, 59, 400, 400))
        self.leCam0.setText("")
        self.leCam0.setObjectName("leCam0")
        self.leCam1 = QtWidgets.QLabel(self.frmCam)
        # self.leCam1.setEnabled(False)
        self.leCam1.setGeometry QRect(439, 59, 400, 400))
        font = QFont()
        font.setPointSize(14)
        font.setStrikeOut(False)
        self.leCam1.setFont(font)
        self.leCam1.setText("")
        self.leCam1.setObjectName("leCam1")
        self.leCam2 = QtWidgets.QLabel(self.frmCam)
        # self.leCam2.setEnabled(False)
        self.leCam2.setGeometry QRect(20, 510, 400, 400))
        self.leCam2.setText("")
        self.leCam2.setObjectName("leCam2")
        self.leCam3 = QtWidgets.QLabel(self.frmCam)
        # self.leCam3.setEnabled(False)
        self.leCam3.setGeometry QRect(439, 510, 400, 400))
        self.leCam3.setText("")
        self.leCam3.setObjectName("leCam3")

        self.btnCamSet0 = QtWidgets.QPushButton(self.frmCam)
        self.btnCamSet0.setGeometry QRect(355, 425, 25, 25))
        self.btnCamSet0.setText("")
        self.btnCamSet0.setObjectName("btnCamSet0")
        self.btnCamSet2 = QtWidgets.QPushButton(self.frmCam)
        self.btnCamSet2.setGeometry QRect(357, 875, 25, 25))
        self.btnCamSet2.setText("")
        self.btnCamSet2.setObjectName("btnCamSet2")
        self.btnCamSet1 = QtWidgets.QPushButton(self.frmCam)
        self.btnCamSet1.setGeometry QRect(776, 425, 25, 25))
        self.btnCamSet1.setText("")
        self.btnCamSet1.setObjectName("btnCamSet1")
        self.btnCamSet3 = QtWidgets.QPushButton(self.frmCam)
        self.btnCamSet3.setGeometry QRect(776, 875, 25, 25))
        self.btnCamSet3.setText("")
        self.btnCamSet3.setObjectName("btnCamSet3")

        self.btnZoomOut0 = QtWidgets.QPushButton(self.frmCam)
        self.btnZoomOut0.setGeometry QRect(325, 425, 25, 25))
        self.btnZoomOut0.setText("")
        self.btnZoomOut0.setObjectName("btnZoomOut0")
        self.btnZoomIn0 = QtWidgets.QPushButton(self.frmCam)
        self.btnZoomIn0.setGeometry QRect(295, 425, 25, 25))
        self.btnZoomIn0.setText("")
        self.btnZoomIn0.setObjectName("btnZoomIn0")
        self.btnZoomOut1 = QtWidgets.QPushButton(self.frmCam)
        self.btnZoomOut1.setGeometry QRect(745, 425, 25, 25))
        self.btnZoomOut1.setText("")
        self.btnZoomOut1.setObjectName("btnZoomOut1")
        self.btnZoomIn1 = QtWidgets.QPushButton(self.frmCam)
        self.btnZoomIn1.setGeometry QRect(715, 425, 25, 25))
        self.btnZoomIn1.setText("")
        self.btnZoomIn1.setObjectName("btnZoomIn1")
        self.btnZoomOut3 = QtWidgets.QPushButton(self.frmCam)
        self.btnZoomOut3.setGeometry QRect(745, 875, 25, 25))
        self.btnZoomOut3.setText("")
        self.btnZoomOut3.setObjectName("btnZoomOut3")
        self.btnZoomIn3 = QtWidgets.QPushButton(self.frmCam)
        self.btnZoomIn3.setGeometry QRect(715, 875, 25, 25))
        self.btnZoomIn3.setText("")
        self.btnZoomIn3.setObjectName("btnZoomIn3")
        self.btnZoomOut2 = QtWidgets.QPushButton(self.frmCam)
        self.btnZoomOut2.setGeometry QRect(325, 875, 25, 25))
        self.btnZoomOut2.setText("")
        self.btnZoomOut2.setObjectName("btnZoomOut2")
        self.btnZoomIn2 = QtWidgets.QPushButton(self.frmCam)
        self.btnZoomIn2.setGeometry QRect(295, 875, 25, 25))
        self.btnZoomIn2.setText("")
        self.btnZoomIn2.setObjectName("btnZoomIn2")
        self.btnRtsp0 = QtWidgets.QPushButton(self.frmCam)
        self.btnRtsp0.setGeometry QRect(265, 425, 25, 25))
        self.btnRtsp0.setText("")
        self.btnRtsp0.setObjectName("btnRtsp0")
        self.btnRtsp1 = QtWidgets.QPushButton(self.frmCam)
        self.btnRtsp1.setGeometry QRect(685, 425, 25, 25))
        self.btnRtsp1.setText("")
        self.btnRtsp1.setObjectName("btnRtsp1")
        self.btnRtsp2 = QtWidgets.QPushButton(self.frmCam)
        self.btnRtsp2.setGeometry QRect(265, 875, 25, 25))
        self.btnRtsp2.setText("")
        self.btnRtsp2.setObjectName("btnRtsp2")
        self.btnRtsp3 = QtWidgets.QPushButton(self.frmCam)
        self.btnRtsp3.setGeometry QRect(685, 875, 25, 25))
        self.btnRtsp3.setText("")
        self.btnRtsp3.setObjectName("btnRtsp3")

        self.btnCam0Exp = QtWidgets.QPushButton(self.frmCam)
        self.btnCam0Exp.setGeometry QRect(385, 425, 25, 25))
        self.btnCam0Exp.setText("")
        self.btnCam0Exp.setObjectName("btnCam0Exp")
        self.btnCam1Exp = QtWidgets.QPushButton(self.frmCam)
        self.btnCam1Exp.setGeometry QRect(804, 425, 25, 25))
        self.btnCam1Exp.setText("")
        self.btnCam1Exp.setObjectName("btnCam1Exp")
        self.btnCam2Exp = QtWidgets.QPushButton(self.frmCam)
        self.btnCam2Exp.setGeometry QRect(385, 875, 25, 25))
        self.btnCam2Exp.setText("")
        self.btnCam2Exp.setObjectName("btnCam2Exp")
        self.btnCam3Exp = QtWidgets.QPushButton(self.frmCam)
        self.btnCam3Exp.setGeometry QRect(804, 875, 25, 25))
        self.btnCam3Exp.setText("")
        self.btnCam3Exp.setObjectName("btnCam3Exp")
        self.label_2 = QtWidgets.QLabel(self.frmCam)
        self.label_2.setGeometry QRect(11, 31, 10, 16))
        self.label_2.setText("")
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.frmCam)
        self.label_3.setGeometry QRect(200, 32, 20, 13))
        self.label_3.setText("")
        self.label_3.setObjectName("label_3")
        self.btnDayNightCam0 = QtWidgets.QPushButton(self.frmCam)
        self.btnDayNightCam0.setGeometry QRect(355, 425, 25, 25))
        self.btnDayNightCam0.setText("")
        self.btnDayNightCam0.setObjectName("btnDayNightCam0")
        self.btnDayNightCam2 = QtWidgets.QPushButton(self.frmCam)
        self.btnDayNightCam2.setGeometry QRect(357, 875, 25, 25))
        self.btnDayNightCam2.setText("")
        self.btnDayNightCam2.setObjectName("btnDayNightCam2")
        self.btnDayNightCam1 = QtWidgets.QPushButton(self.frmCam)
        self.btnDayNightCam1.setGeometry QRect(776, 425, 25, 25))
        self.btnDayNightCam1.setText("")
        self.btnDayNightCam1.setObjectName("btnDayNightCam1")
        self.btnDayNightCam3 = QtWidgets.QPushButton(self.frmCam)
        self.btnDayNightCam3.setGeometry QRect(776, 875, 25, 25))
        self.btnDayNightCam3.setText("")
        self.btnDayNightCam3.setObjectName("btnDayNightCam3")
        self.btnCam0Arch = QtWidgets.QPushButton('0', self.frmCam)
        self.btnCam0Arch.setGeometry QRect(30, 425, 25, 25))
        self.btnCam0Arch.setText("")
        self.btnCam0Arch.setObjectName("btnCam0Arch")
        self.name = 'qwerty'
        self.btnCam1Arch = QtWidgets.QPushButton('1', self.frmCam)
        self.btnCam1Arch.setGeometry QRect(450, 425, 25, 25))
        self.btnCam1Arch.setText("")
        self.btnCam1Arch.setObjectName("btnCam1Arch")
        self.name = 'sec'
        self.btnCam3Arch = QtWidgets.QPushButton(self.frmCam)
        self.btnCam3Arch.setGeometry QRect(450, 875, 25, 25))
        self.btnCam3Arch.setText("")
        self.btnCam3Arch.setObjectName("btnCam3Arch")
        self.btnCam2Arch = QtWidgets.QPushButton(self.frmCam)
        self.btnCam2Arch.setGeometry QRect(30, 875, 25, 25))
        self.btnCam2Arch.setText("")
        self.btnCam2Arch.setObjectName("btnCam2Arch")
        self.leCam0Rec = QtWidgets.QLineEdit(self.frmCam)
        self.leCam0Rec.setGeometry QRect(30, 70, 61, 16))
        self.leCam0Rec.setText("")
        self.leCam0Rec.setObjectName("leCam0Rec")
        self.leCam1Rec = QtWidgets.QLineEdit(self.frmCam)
        self.leCam1Rec.setGeometry QRect(450, 70, 61, 16))
        self.leCam1Rec.setText("")
        self.leCam1Rec.setObjectName("leCam1Rec")
        self.leCam3Rec = QtWidgets.QLineEdit(self.frmCam)
        self.leCam3Rec.setGeometry QRect(450, 520, 61, 16))
        self.leCam3Rec.setText("")
        self.leCam3Rec.setObjectName("leCam3Rec")
        self.leCam2Rec = QtWidgets.QLineEdit(self.frmCam)
        self.leCam2Rec.setGeometry QRect(30, 520, 61, 16))
        self.leCam2Rec.setText("")
        self.leCam2Rec.setObjectName("leCam2Rec")
        self.lblMainBoxCams.raise_()
        self.lblCam4.raise_()
        self.lblTitleCams.raise_()
        self.lblCatCams.raise_()
        self.lblIconCams.raise_()
        self.lblCam1.raise_()
        self.lblCam2.raise_()
        self.lblCam0.raise_()
        self.leCam0.raise_()
        self.leCam1.raise_()
        self.leCam2.raise_()
        self.leCam3.raise_()
        self.btnCam0Exp.raise_()
        self.btnCam1Exp.raise_()
        self.btnCam2Exp.raise_()
        self.btnCam3Exp.raise_()
        self.label_2.raise_()
        self.label_3.raise_()
        self.btnDayNightCam0.raise_()
        self.btnDayNightCam2.raise_()
        self.btnDayNightCam1.raise_()
        self.btnDayNightCam3.raise_()
        self.btnCam0Arch.raise_()
        self.btnCam1Arch.raise_()
        self.btnCam3Arch.raise_()
        self.btnCam2Arch.raise_()
        self.leCam0Rec.raise_()
        self.leCam1Rec.raise_()
        self.leCam3Rec.raise_()
        self.leCam2Rec.raise_()

        self.btnCamSet0.raise_()
        self.btnCamSet2.raise_()
        self.btnCamSet1.raise_()
        self.btnCamSet3.raise_()
        self.btnZoomOut0.raise_()
        self.btnZoomIn0.raise_()
        self.btnZoomOut1.raise_()
        self.btnZoomIn1.raise_()
        self.btnZoomOut3.raise_()
        self.btnZoomIn3.raise_()
        self.btnZoomOut2.raise_()
        self.btnZoomIn2.raise_()
        self.btnRtsp0.raise_()
        self.btnRtsp1.raise_()
        self.btnRtsp2.raise_()
        self.btnRtsp3.raise_()

        self.frmTable = QtWidgets.QFrame(self.mainFormPr)
        self.frmTable.setGeometry QRect(854, 132, 866, 441))
        self.frmTable.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmTable.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmTable.setObjectName("frmTable")
        self.lblBoxJournal = QtWidgets.QLabel(self.frmTable)
        self.lblBoxJournal.setGeometry QRect(6, 0, 215, 40))
        font = QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lblBoxJournal.setFont(font)
        self.lblBoxJournal.setText("")
        self.lblBoxJournal.setAlignment Qt.AlignLeading Qt.AlignLeft Qt.AlignTop)
        self.lblBoxJournal.setObjectName("lblBoxJournal")
        self.lblNameJournal = QtWidgets.QLabel(self.frmTable)
        self.lblNameJournal.setGeometry QRect(49, 7, 165, 20))
        font = QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lblNameJournal.setFont(font)
        self.lblNameJournal.setObjectName("lblNameJournal")
        self.lblIconJournal = QtWidgets.QLabel(self.frmTable)
        self.lblIconJournal.setGeometry QRect(22, 5, 23, 23))
        self.lblIconJournal.setText("")
        self.lblIconJournal.setObjectName("lblIconJournal")
        self.lblTbl = QtWidgets.QLabel(self.frmTable)
        self.lblTbl.setEnabled(False)
        self.lblTbl.setGeometry QRect(6, 30, 854, 411))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lblTbl.sizePolicy().hasHeightForWidth())
        self.lblTbl.setSizePolicy(sizePolicy)
        self.lblTbl.setSizeIncrement QSize(0, 10))
        self.lblTbl.setText("")
        self.lblTbl.setObjectName("lblTbl")
        self.leSearchTbl = QtWidgets.QLineEdit(self.frmTable)
        self.leSearchTbl.setGeometry QRect(22, 40, 596, 31))
        font = QFont()
        font.setPointSize(12)
        self.leSearchTbl.setFont(font)
        self.leSearchTbl.setText("")
        self.leSearchTbl.setObjectName("leSearchTbl")
        self.tblMainData = QtWidgets.QTableWidget(self.frmTable)
        self.tblMainData.setGeometry QRect(23, 80, 821, 351))
        self.tblMainData.setFrameShape(QtWidgets.QFrame.NoFrame)
        self.tblMainData.setVerticalScrollBarPolicy Qt.ScrollBarAlwaysOff)
        self.tblMainData.setHorizontalScrollBarPolicy Qt.ScrollBarAlwaysOff)
        self.tblMainData.setShowGrid(True)
        self.tblMainData.setGridStyle Qt.SolidLine)
        self.tblMainData.setRowCount(0)
        self.tblMainData.setColumnCount(13)
        self.tblMainData.setObjectName("tblMainData")
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QColor(84, 122, 181))
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle Qt.SolidPattern)
        item.setForeground(brush)
        self.tblMainData.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QColor(84, 122, 181))
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle Qt.SolidPattern)
        item.setForeground(brush)
        self.tblMainData.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QColor(84, 122, 181))
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle Qt.SolidPattern)
        item.setForeground(brush)
        self.tblMainData.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QColor(84, 122, 181))
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle Qt.SolidPattern)
        item.setForeground(brush)
        self.tblMainData.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QColor(84, 122, 181))
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle Qt.SolidPattern)
        item.setForeground(brush)
        self.tblMainData.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QColor(84, 122, 181))
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle Qt.SolidPattern)
        item.setForeground(brush)
        self.tblMainData.setHorizontalHeaderItem(5, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QColor(84, 122, 181))
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle Qt.SolidPattern)
        item.setForeground(brush)
        self.tblMainData.setHorizontalHeaderItem(6, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QColor(84, 122, 181))
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle Qt.SolidPattern)
        item.setForeground(brush)
        self.tblMainData.setHorizontalHeaderItem(7, item)
        item = QtWidgets.QTableWidgetItem()
        item.setBackground(QColor(84, 122, 181))
        brush = QBrush(QColor(255, 255, 255))
        brush.setStyle Qt.SolidPattern)
        item.setForeground(brush)
        self.tblMainData.setHorizontalHeaderItem(8, item)
        self.tblMainData.horizontalHeader().setVisible(True)
        self.tblMainData.horizontalHeader().setCascadingSectionResizes(False)
        self.tblMainData.horizontalHeader().setDefaultSectionSize(91)
        self.tblMainData.horizontalHeader().setHighlightSections(True)
        self.tblMainData.horizontalHeader().setMinimumSectionSize(31)
        self.tblMainData.horizontalHeader().setSortIndicatorShown(True)
        self.tblMainData.horizontalHeader().setStretchLastSection(False)
        self.tblMainData.verticalHeader().setVisible(False)

        print('serhio123')

        if (globalValues.colorForm == 0):
            self.tblMainData.setStyleSheet(
                "QTableWidget {background-color: rgb(235,235,235);\n"
                                           "border: 1px solid rgb(150,150,150);\n"
                                           "gridline-color: rgb(89,89,89);\n"
                                           "border-radius:5px;\n"
                                           "border-top-left-radius: 0px;\n"
                                           "border-top-right-radius: 0px;\n"
                                           "color:black;}\n"
                                           "\n"
                                           "QLineEdit {background-color: white;}\n"
                                           "\n"
                                           "QHeaderView::section {"
                # "gridline-color: rgb(89,89,89);\n"
                                           "background-color: rgb(142,187,208);\n"
                                           "color: black;};")

        elif (globalValues.colorForm == 1):
            self.tblMainData.setStyleSheet(
                "QTableWidget {background-color: rgb(42,42,42);\n"
                                           "border: 1px solid rgb(63,63,63);\n"
                                           "gridline-color: rgb(89,89,89);\n"
                                           "border-radius:5px;\n"

                                           "color:white;}\n"
                                           "\n"
                                           "QLineEdit {background-color: white;}\n"
                                           "\n"
                                           "QHeaderView::section {\n"
                                           # "gridline-color: rgb(89,89,89);\n"
                                           "background-color: rgb(50,75,115);};")

        self.btnSearch = QtWidgets.QPushButton(self.frmTable)
        self.btnSearch.setGeometry QRect(701, 40, 102, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnSearch.setFont(font)
        icon = QIcon()
        self.btnSearch.setIcon(icon)
        self.btnSearch.setIconSize QSize(25, 25))
        self.btnSearch.setObjectName("btnSearch")
        self.comboSearch = QtWidgets.QComboBox(self.frmTable)
        self.comboSearch.setGeometry QRect(626, 40, 69, 31))
        font = QFont()
        font.setPointSize(12)
        self.comboSearch.setFont(font)
        self.comboSearch.setObjectName("comboSearch")
        self.btnRefresh = QtWidgets.QPushButton(self.frmTable)
        self.btnRefresh.setGeometry QRect(812, 40, 31, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnRefresh.setFont(font)
        self.btnRefresh.setText("")
        icon1 = QIcon()
        self.btnRefresh.setIcon(icon1)
        self.btnRefresh.setIconSize QSize(25, 25))
        self.btnRefresh.setObjectName("btnRefresh")
        self.label_4 = QtWidgets.QLabel(self.frmTable)
        self.label_4.setGeometry QRect(8, 29, 10, 16))
        self.label_4.setText("")
        self.label_4.setObjectName("label_4")
        self.label_5 = QtWidgets.QLabel(self.frmTable)
        self.label_5.setGeometry QRect(207, 32, 20, 8))
        self.label_5.setText("")
        self.label_5.setObjectName("label_5")
        self.verticalScrollBar = QtWidgets.QScrollBar(self.frmTable)
        self.verticalScrollBar.setGeometry QRect(842, 103, 10, 328))
        self.verticalScrollBar.setOrientation Qt.Vertical)
        self.verticalScrollBar.setObjectName("verticalScrollBar")
        self.DateEdit = QtWidgets.QDateEdit(self.frmTable)
        self.DateEdit.setGeometry QRect(22, 40, 146, 30))
        self.DateEdit.setObjectName("DTEitFix")
        self.lblTbl.raise_()
        self.btnDelZakazSergey = QtWidgets.QPushButton(self.frmTable)
        self.btnDelZakazSergey.setGeometry QRect(22, 40, 100, 30))
        self.btnDelZakazSergey.setObjectName("DelZakazSergey")
        font = QFont()
        font.setPointSize(12)
        self.btnDelZakazSergey.setFont(font)
        icon = QIcon()
        self.btnDelZakazSergey.setIcon(icon)
        self.btnDelZakazSergey.setIconSize QSize(25, 25))
        self.lblBoxJournal.raise_()
        self.lblNameJournal.raise_()
        self.lblIconJournal.raise_()
        self.leSearchTbl.raise_()
        self.tblMainData.raise_()
        self.btnSearch.raise_()
        self.comboSearch.raise_()
        self.btnRefresh.raise_()
        self.label_4.raise_()
        self.label_5.raise_()
        self.verticalScrollBar.raise_()
        self.DateEdit.raise_()
        self.btnDelZakazSergey.raise_()
        self.frmMenu = QtWidgets.QFrame(self.mainFormPr)
        self.frmMenu.setGeometry QRect(1720, 132, 201, 941))
        self.frmMenu.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmMenu.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmMenu.setObjectName("frmMenu")
        self.lblBoxSets = QtWidgets.QLabel(self.frmMenu)
        self.lblBoxSets.setEnabled(False)
        self.lblBoxSets.setGeometry QRect(0, 0, 191, 931))
        self.lblBoxSets.setText("")
        self.lblBoxSets.setObjectName("lblBoxSets")
        self.btnAboutSetSystem = QtWidgets.QPushButton(self.frmMenu)
        self.btnAboutSetSystem.setGeometry QRect(15, 541, 161, 33))
        font = QFont()
        font.setPointSize(12)
        self.btnAboutSetSystem.setFont(font)
        self.btnAboutSetSystem.setObjectName("btnAboutSetSystem")
        self.btnStWorkingSystem = QtWidgets.QPushButton(self.frmMenu)
        self.btnStWorkingSystem.setGeometry QRect(15, 481, 161, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnStWorkingSystem.setFont(font)
        self.btnStWorkingSystem.setObjectName("btnStWorkingSystem")
        self.lblIconMainSet = QtWidgets.QLabel(self.frmMenu)
        self.lblIconMainSet.setGeometry QRect(14, 9, 23, 23))
        self.lblIconMainSet.setText("")
        self.lblIconMainSet.setObjectName("lblIconMainSet")
        self.lblTitleMainSet = QtWidgets.QLabel(self.frmMenu)
        self.lblTitleMainSet.setGeometry QRect(13, 34, 165, 21))
        font = QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lblTitleMainSet.setFont(font)
        self.lblTitleMainSet.setObjectName("lblTitleMainSet")
        self.btnStatusDevs = QtWidgets.QPushButton(self.frmMenu)
        self.btnStatusDevs.setGeometry QRect(15, 421, 161, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnStatusDevs.setFont(font)
        self.btnStatusDevs.setObjectName("btnStatusDevs")

        self.btnScaner = QtWidgets.QPushButton(self.frmMenu)
        self.btnScaner.setGeometry QRect(15, 391, 161, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnScaner.setFont(font)
        self.btnScaner.setObjectName("btnScaner")

        self.lblIconSetSystem = QtWidgets.QLabel(self.frmMenu)
        self.lblIconSetSystem.setGeometry QRect(14, 218, 23, 23))
        self.lblIconSetSystem.setText("")
        self.lblIconSetSystem.setObjectName("lblIconSetSystem")
        self.btnJournal = QtWidgets.QPushButton(self.frmMenu)
        self.btnJournal.setGeometry QRect(15, 61, 161, 34))
        font = QFont()
        font.setPointSize(12)
        self.btnJournal.setFont(font)
        self.btnJournal.setObjectName("btnJournal")
        self.btnGard = QtWidgets.QPushButton(self.frmMenu)
        self.btnGard.setGeometry QRect(15, 331, 161, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnGard.setFont(font)
        self.btnGard.setObjectName("btnGard")
        self.lineSetMain = QtWidgets.QFrame(self.frmMenu)
        self.lineSetMain.setGeometry QRect(43, 21, 137, 4))
        self.lineSetMain.setFrameShape(QtWidgets.QFrame.HLine)
        self.lineSetMain.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.lineSetMain.setObjectName("lineSetMain")
        self.btnAboutMain = QtWidgets.QPushButton(self.frmMenu)
        self.btnAboutMain.setGeometry QRect(15, 181, 161, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnAboutMain.setFont(font)
        self.btnAboutMain.setObjectName("btnAboutMain")
        self.btnWeight = QtWidgets.QPushButton(self.frmMenu)
        self.btnWeight.setGeometry QRect(15, 301, 161, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnWeight.setFont(font)
        self.btnWeight.setObjectName("btnWeight")
        self.btnDB = QtWidgets.QPushButton(self.frmMenu)
        self.btnDB.setGeometry QRect(15, 361, 161, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnDB.setFont(font)
        self.btnDB.setObjectName("btnDB")
        self.lineSetsSystem = QtWidgets.QFrame(self.frmMenu)
        self.lineSetsSystem.setGeometry QRect(43, 230, 137, 4))
        self.lineSetsSystem.setFrameShape(QtWidgets.QFrame.HLine)
        self.lineSetsSystem.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.lineSetsSystem.setObjectName("lineSetsSystem")
        self.btnSetCams = QtWidgets.QPushButton(self.frmMenu)
        self.btnSetCams.setGeometry QRect(15, 270, 161, 34))
        font = QFont()
        font.setPointSize(12)
        self.btnSetCams.setFont(font)
        self.btnSetCams.setObjectName("btnSetCams")
        self.lblNameSetSystem = QtWidgets.QLabel(self.frmMenu)
        self.lblNameSetSystem.setGeometry QRect(13, 243, 165, 21))
        font = QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.lblNameSetSystem.setFont(font)
        self.lblNameSetSystem.setObjectName("lblNameSetSystem")
        self.btnDetectionCars = QtWidgets.QPushButton(self.frmMenu)
        self.btnDetectionCars.setGeometry QRect(15, 451, 161, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnDetectionCars.setFont(font)
        self.btnDetectionCars.setObjectName("btnDetectionCars")
        self.btnColorScheme = QtWidgets.QPushButton(self.frmMenu)
        self.btnColorScheme.setGeometry QRect(15, 511, 161, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnColorScheme.setFont(font)
        self.btnColorScheme.setObjectName("btnColorScheme")
        self.btnuserchange = QtWidgets.QPushButton(self.frmMenu)
        self.btnuserchange.setGeometry QRect(150, 897, 30, 30))
        self.btnuserchange.setText("")
        self.btnuserchange.setObjectName("btnuserchange")
        self.lineSetsSystem_2 = QtWidgets.QFrame(self.frmMenu)
        self.lineSetsSystem_2.setGeometry QRect(15, 890, 160, 4))
        self.lineSetsSystem_2.setFrameShape(QtWidgets.QFrame.HLine)
        self.lineSetsSystem_2.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.lineSetsSystem_2.setObjectName("lineSetsSystem_2")
        self.le_UserName = QtWidgets.QLineEdit(self.frmMenu)
        self.le_UserName.setGeometry QRect(64, 900, 81, 20))
        self.le_UserName.setText("")
        self.le_UserName.setAlignment Qt.AlignRight | Qt.AlignTrailing | Qt.AlignVCenter)
        self.le_UserName.setObjectName("lineEdit")
        self.btnJournalTS = QtWidgets.QPushButton(self.frmMenu)
        self.btnJournalTS.setGeometry QRect(15, 92, 161, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnJournalTS.setFont(font)
        self.btnJournalTS.setObjectName("btnJournalTS")
        self.btnBDTS = QtWidgets.QPushButton(self.frmMenu)
        self.btnBDTS.setGeometry QRect(15, 152, 161, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnBDTS.setFont(font)
        self.btnBDTS.setObjectName("btnBDTS")
        self.btnAddZakaz = QtWidgets.QPushButton(self.frmMenu)
        self.btnAddZakaz.setGeometry QRect(15, 122, 161, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnAddZakaz.setFont(font)
        self.btnAddZakaz.setObjectName("btnAddZakaz")
        self.lblBoxSets.raise_()
        self.btnStWorkingSystem.raise_()
        self.lblIconMainSet.raise_()
        self.lblTitleMainSet.raise_()
        self.btnStatusDevs.raise_()
        self.btnScaner.raise_()
        self.lblIconSetSystem.raise_()
        self.btnGard.raise_()
        self.lineSetMain.raise_()
        self.btnAboutMain.raise_()
        self.btnDB.raise_()
        self.lineSetsSystem.raise_()
        self.lblNameSetSystem.raise_()
        self.btnDetectionCars.raise_()
        self.btnuserchange.raise_()
        self.lineSetsSystem_2.raise_()
        self.le_UserName.raise_()
        self.btnSetCams.raise_()
        self.btnAboutSetSystem.raise_()
        self.btnColorScheme.raise_()
        self.btnWeight.raise_()
        self.btnJournal.raise_()
        self.btnJournalTS.raise_()
        self.btnBDTS.raise_()
        self.btnAddZakaz.raise_()
        self.frmHeadImage = QtWidgets.QFrame(self.mainFormPr)
        self.frmHeadImage.setGeometry QRect(0, 39, 1911, 91))
        self.frmHeadImage.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmHeadImage.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmHeadImage.setObjectName("frmHeadImage")
        self.lblBoxUpCompany = QtWidgets.QLabel(self.frmHeadImage)
        self.lblBoxUpCompany.setEnabled(False)
        self.lblBoxUpCompany.setGeometry QRect(9, 0, 1902, 86))
        self.lblBoxUpCompany.setText("")
        self.lblBoxUpCompany.setObjectName("lblBoxUpCompany")
        self.frmHeadExit = QtWidgets.QFrame(self.mainFormPr)
        self.frmHeadExit.setGeometry QRect(0, 0, 1921, 31))
        self.frmHeadExit.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmHeadExit.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmHeadExit.setObjectName("frmHeadExit")
        self.lblBarClose = QtWidgets.QLabel(self.frmHeadExit)
        self.lblBarClose.setGeometry QRect(25, 0, 1896, 30))
        font = QFont()
        font.setFamily("MS Shell Dlg 2")
        font.setPointSize(12)
        font.setBold(False)
        font.setItalic(False)
        font.setUnderline(True)
        font.setWeight(50)
        self.lblBarClose.setFont(font)
        self.lblBarClose.setObjectName("lblBarClose")
        self.lblIconBarClose = QtWidgets.QLabel(self.frmHeadExit)
        self.lblIconBarClose.setGeometry QRect(2, 0, 20, 30))
        self.lblIconBarClose.setText("")
        self.lblIconBarClose.setObjectName("lblIconBarClose")
        self.btnCloseMainForm = QtWidgets.QPushButton(self.frmHeadExit)
        self.btnCloseMainForm.setGeometry QRect(1890, 2, 26, 26))
        self.btnCloseMainForm.setText("")
        self.btnCloseMainForm.setObjectName("btnCloseMainForm")
        self.label = QtWidgets.QLabel(self.frmHeadExit)
        self.label.setGeometry QRect(0, 0, 50, 30))
        self.label.setText("")
        self.label.setObjectName("label")
        self.label.raise_()
        self.lblBarClose.raise_()
        self.lblIconBarClose.raise_()
        self.btnCloseMainForm.raise_()
        self.frmCamAllMon = QtWidgets.QFrame(self.mainFormPr)
        self.frmCamAllMon.setGeometry QRect(0, 129, 1920, 1011))
        self.frmCamAllMon.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmCamAllMon.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmCamAllMon.setObjectName("frmCamAllMon")
        self.leCamAllMon = QtWidgets.QLabel(self.frmCamAllMon)
        self.leCamAllMon.setGeometry QRect(122, 0, 1672, 941))
        self.leCamAllMon.setFrameShadow(QtWidgets.QFrame.Plain)
        self.leCamAllMon.setText("")
        self.leCamAllMon.setObjectName("leCamAllMon")
        self.btnCamCallapse = QtWidgets.QPushButton(self.frmCamAllMon)
        self.btnCamCallapse.setGeometry QRect(1800, 913, 28, 28))
        self.btnCamCallapse.setText("")
        self.btnCamCallapse.setObjectName("btnCamCallapse")
        self.frmCntrlPanel = QtWidgets.QFrame(self.mainFormPr)
        self.frmCntrlPanel.setGeometry QRect(854, 573, 866, 501))
        self.frmCntrlPanel.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmCntrlPanel.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmCntrlPanel.setObjectName("frmCntrlPanel")
        self.lblIconDataWeightTr = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblIconDataWeightTr.setGeometry QRect(40, 35, 23, 23))
        self.lblIconDataWeightTr.setText("")
        self.lblIconDataWeightTr.setObjectName("lblIconDataWeightTr")
        self.lblMainSemaLeftOutUp = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblMainSemaLeftOutUp.setEnabled(False)
        self.lblMainSemaLeftOutUp.setGeometry QRect(43, 78, 31, 31))
        self.lblMainSemaLeftOutUp.setText("")
        self.lblMainSemaLeftOutUp.setObjectName("lblMainSemaLeftOutUp")
        self.label_54 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.label_54.setGeometry QRect(85, 72, 41, 111))
        self.label_54.setText("")
        self.label_54.setObjectName("label_54")
        self.label_56 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.label_56.setGeometry QRect(787, 72, 41, 111))
        self.label_56.setText("")
        self.label_56.setObjectName("label_56")
        self.lblMainSemaRightOutDown = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblMainSemaRightOutDown.setEnabled(False)
        self.lblMainSemaRightOutDown.setGeometry QRect(792, 146, 31, 31))
        self.lblMainSemaRightOutDown.setText("")
        self.lblMainSemaRightOutDown.setObjectName("lblMainSemaRightOutDown")
        self.lblContDataWeight = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblContDataWeight.setGeometry QRect(22, 26, 819, 225))
        self.lblContDataWeight.setText("")
        self.lblContDataWeight.setObjectName("lblContDataWeight")
        self.lcdMainWeight = QtWidgets.QLCDNumber(self.frmCntrlPanel)
        self.lcdMainWeight.setGeometry QRect(135, 72, 596, 111))
        self.lcdMainWeight.setObjectName("lcdMainWeight")
        self.label_53 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.label_53.setGeometry QRect(38, 72, 41, 111))
        self.label_53.setText("")
        self.label_53.setObjectName("label_53")
        self.lblMainSemaLeftInUp = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblMainSemaLeftInUp.setEnabled(False)
        self.lblMainSemaLeftInUp.setGeometry QRect(90, 78, 31, 31))
        self.lblMainSemaLeftInUp.setText("")
        self.lblMainSemaLeftInUp.setObjectName("lblMainSemaLeftInUp")
        self.lblMainSemaLeftOutDown = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblMainSemaLeftOutDown.setEnabled(False)
        self.lblMainSemaLeftOutDown.setGeometry QRect(43, 146, 31, 31))
        self.lblMainSemaLeftOutDown.setText("")
        self.lblMainSemaLeftOutDown.setObjectName("lblMainSemaLeftOutDown")
        self.lblMainSemaRightInDown = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblMainSemaRightInDown.setEnabled(False)
        self.lblMainSemaRightInDown.setGeometry QRect(745, 146, 31, 31))
        self.lblMainSemaRightInDown.setText("")
        self.lblMainSemaRightInDown.setObjectName("lblMainSemaRightInDown")
        self.lblMainSemaRightOutUp = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblMainSemaRightOutUp.setEnabled(False)
        self.lblMainSemaRightOutUp.setGeometry QRect(792, 78, 31, 31))
        self.lblMainSemaRightOutUp.setText("")
        self.lblMainSemaRightOutUp.setObjectName("lblMainSemaRightOutUp")
        self.line_8 = QtWidgets.QFrame(self.frmCntrlPanel)
        self.line_8.setGeometry QRect(370, 50, 131, 16))
        self.line_8.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_8.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_8.setObjectName("line_8")
        self.label_55 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.label_55.setGeometry QRect(740, 72, 41, 111))
        self.label_55.setText("")
        self.label_55.setObjectName("label_55")
        self.lblNameWeight_3 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblNameWeight_3.setGeometry QRect(370, 32, 135, 20))
        font = QFont()
        font.setPointSize(12)
        self.lblNameWeight_3.setFont(font)
        self.lblNameWeight_3.setObjectName("lblNameWeight_3")
        self.lblMainSemaLeftInDown = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblMainSemaLeftInDown.setEnabled(False)
        self.lblMainSemaLeftInDown.setGeometry QRect(90, 146, 31, 31))
        self.lblMainSemaLeftInDown.setText("")
        self.lblMainSemaLeftInDown.setObjectName("lblMainSemaLeftInDown")
        self.lblMainSemaRightInUp = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblMainSemaRightInUp.setEnabled(False)
        self.lblMainSemaRightInUp.setGeometry QRect(745, 78, 31, 31))
        self.lblMainSemaRightInUp.setText("")
        self.lblMainSemaRightInUp.setObjectName("lblMainSemaRightInUp")
        self.lblTbl_2 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblTbl_2.setEnabled(False)
        self.lblTbl_2.setGeometry QRect(6, 10, 854, 479))
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Preferred, QtWidgets.QSizePolicy.Preferred)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.lblTbl_2.sizePolicy().hasHeightForWidth())
        self.lblTbl_2.setSizePolicy(sizePolicy)
        self.lblTbl_2.setSizeIncrement QSize(0, 10))
        self.lblTbl_2.setText("")
        self.lblTbl_2.setObjectName("lblTbl_2")
        self.btnMainSemaREDOUT = QtWidgets.QPushButton(self.frmCntrlPanel)
        self.btnMainSemaREDOUT.setGeometry QRect(180, 202, 161, 28))
        font = QFont()
        font.setPointSize(12)
        self.btnMainSemaREDOUT.setFont(font)
        self.btnMainSemaREDOUT.setObjectName("btnMainSemaREDOUT")
        self.btnMainSemaALLRED = QtWidgets.QPushButton(self.frmCntrlPanel)
        self.btnMainSemaALLRED.setGeometry QRect(360, 202, 161, 28))
        font = QFont()
        font.setPointSize(12)
        self.btnMainSemaALLRED.setFont(font)
        self.btnMainSemaALLRED.setObjectName("btnMainSemaALLRED")
        self.btnMainSemaREDIN = QtWidgets.QPushButton(self.frmCntrlPanel)
        self.btnMainSemaREDIN.setGeometry QRect(540, 202, 161, 28))
        font = QFont()
        font.setPointSize(12)
        self.btnMainSemaREDIN.setFont(font)
        self.btnMainSemaREDIN.setObjectName("btnMainSemaREDIN")
        self.lblContDataWeight_2 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblContDataWeight_2.setGeometry QRect(484, 258, 358, 211))
        self.lblContDataWeight_2.setText("")
        self.lblContDataWeight_2.setObjectName("lblContDataWeight_2")
        self.lblContDataWeight_3 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblContDataWeight_3.setGeometry QRect(23, 258, 451, 211))
        self.lblContDataWeight_3.setText("")
        self.lblContDataWeight_3.setObjectName("lblContDataWeight_3")
        self.btnMainShlgOpen = QtWidgets.QPushButton(self.frmCntrlPanel)
        self.btnMainShlgOpen.setGeometry QRect(520, 307, 141, 28))
        font = QFont()
        font.setPointSize(12)
        self.btnMainShlgOpen.setFont(font)
        self.btnMainShlgOpen.setObjectName("btnMainShlgOpen")
        self.btnMainShlgClose = QtWidgets.QPushButton(self.frmCntrlPanel)
        self.btnMainShlgClose.setGeometry QRect(520, 342, 141, 28))
        font = QFont()
        font.setPointSize(12)
        self.btnMainShlgClose.setFont(font)
        self.btnMainShlgClose.setObjectName("btnMainShlgClose")
        self.line_9 = QtWidgets.QFrame(self.frmCntrlPanel)
        self.line_9.setGeometry QRect(638, 282, 76, 16))
        self.line_9.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_9.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_9.setObjectName("line_9")
        self.lblNameWeight_4 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblNameWeight_4.setGeometry QRect(640, 264, 74, 20))
        font = QFont()
        font.setPointSize(12)
        self.lblNameWeight_4.setFont(font)
        self.lblNameWeight_4.setObjectName("lblNameWeight_4")
        self.label_45 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.label_45.setGeometry QRect(520, 388, 141, 61))
        self.label_45.setText("")
        self.label_45.setObjectName("label_45")
        self.lblMainShlgOpen = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblMainShlgOpen.setEnabled(False)
        self.lblMainShlgOpen.setGeometry QRect(610, 413, 31, 31))
        self.lblMainShlgOpen.setText("")
        self.lblMainShlgOpen.setObjectName("lblMainShlgOpen")
        self.label_20 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.label_20.setGeometry QRect(603, 390, 47, 15))
        font = QFont()
        font.setPointSize(10)
        self.label_20.setFont(font)
        self.label_20.setObjectName("label_20")
        self.label_21 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.label_21.setGeometry QRect(534, 390, 47, 15))
        font = QFont()
        font.setPointSize(10)
        self.label_21.setFont(font)
        self.label_21.setObjectName("label_21")
        self.lblMainShlgClose = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblMainShlgClose.setEnabled(False)
        self.lblMainShlgClose.setGeometry QRect(540, 413, 31, 31))
        self.lblMainShlgClose.setText("")
        self.lblMainShlgClose.setObjectName("lblMainShlgClose")
        self.lblMainShlgStateImage = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblMainShlgStateImage.setGeometry QRect(670, 310, 151, 140))
        self.lblMainShlgStateImage.setText("")
        self.lblMainShlgStateImage.setObjectName("lblMainShlgStateImage")
        self.lblNameWeight_5 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblNameWeight_5.setGeometry QRect(180, 264, 141, 20))
        font = QFont()
        font.setPointSize(12)
        self.lblNameWeight_5.setFont(font)
        self.lblNameWeight_5.setObjectName("lblNameWeight_5")
        self.line_10 = QtWidgets.QFrame(self.frmCntrlPanel)
        self.line_10.setGeometry QRect(178, 282, 141, 16))
        self.line_10.setFrameShape(QtWidgets.QFrame.HLine)
        self.line_10.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line_10.setObjectName("line_10")
        self.lblNameWeight_6 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblNameWeight_6.setGeometry QRect(80, 307, 161, 20))
        font = QFont()
        font.setPointSize(12)
        self.lblNameWeight_6.setFont(font)
        self.lblNameWeight_6.setObjectName("lblNameWeight_6")
        self.lblNameWeight_7 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblNameWeight_7.setGeometry QRect(80, 347, 161, 20))
        font = QFont()
        font.setPointSize(12)
        self.lblNameWeight_7.setFont(font)
        self.lblNameWeight_7.setObjectName("lblNameWeight_7")
        self.lblNameWeight_8 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblNameWeight_8.setGeometry QRect(80, 387, 161, 20))
        font = QFont()
        font.setPointSize(12)
        self.lblNameWeight_8.setFont(font)
        self.lblNameWeight_8.setObjectName("lblNameWeight_8")
        self.lblNameWeight_9 = QtWidgets.QLabel(self.frmCntrlPanel)
        self.lblNameWeight_9.setGeometry QRect(80, 427, 161, 20))
        font = QFont()
        font.setPointSize(12)
        self.lblNameWeight_9.setFont(font)
        self.lblNameWeight_9.setObjectName("lblNameWeight_9")
        self.btnMainGRZCh1 = QtWidgets.QPushButton(self.frmCntrlPanel)
        self.btnMainGRZCh1.setGeometry QRect(400, 300, 33, 33))
        font = QFont()
        font.setPointSize(12)
        self.btnMainGRZCh1.setFont(font)
        self.btnMainGRZCh1.setText("")
        self.btnMainGRZCh1.setObjectName("btnMainGRZCh1")
        self.btnMainGRZCh2 = QtWidgets.QPushButton(self.frmCntrlPanel)
        self.btnMainGRZCh2.setGeometry QRect(400, 340, 33, 33))
        font = QFont()
        font.setPointSize(12)
        self.btnMainGRZCh2.setFont(font)
        self.btnMainGRZCh2.setText("")
        self.btnMainGRZCh2.setObjectName("btnMainGRZCh2")
        self.btnMainGRZCh3 = QtWidgets.QPushButton(self.frmCntrlPanel)
        self.btnMainGRZCh3.setGeometry QRect(400, 380, 33, 33))
        font = QFont()
        font.setPointSize(12)
        self.btnMainGRZCh3.setFont(font)
        self.btnMainGRZCh3.setText("")
        self.btnMainGRZCh3.setObjectName("btnMainGRZCh3")
        self.btnMainGRZCh4 = QtWidgets.QPushButton(self.frmCntrlPanel)
        self.btnMainGRZCh4.setGeometry QRect(400, 420, 33, 33))
        font = QFont()
        font.setPointSize(12)
        self.btnMainGRZCh4.setFont(font)
        self.btnMainGRZCh4.setText("")
        self.btnMainGRZCh4.setObjectName("btnMainGRZCh4")
        self.btnHide = QtWidgets.QPushButton(self.frmCntrlPanel)
        self.btnHide.setGeometry QRect(839, 10, 20, 20))
        self.btnHide.setText("")
        self.btnHide.setObjectName("btnHide")
        # self.leMainGRZGardOut = QtWidgets.QLineEdit(self.frmCntrlPanel)
        # self.leMainGRZGardOut.setGeometry QRect(220, 340, 151, 33))
        # self.leMainGRZGardOut.setObjectName("leMainGRZGardOut")
        # self.leMainGRZWeightOut = QtWidgets.QLineEdit(self.frmCntrlPanel)
        # self.leMainGRZWeightOut.setGeometry QRect(220, 420, 151, 33))
        # self.leMainGRZWeightOut.setObjectName("leMainGRZWeightOut")
        # self.leMainGRZWeightIn = QtWidgets.QLineEdit(self.frmCntrlPanel)
        # self.leMainGRZWeightIn.setGeometry QRect(220, 380, 151, 33))
        # self.leMainGRZWeightIn.setObjectName("leMainGRZWeightIn")
        # self.leMainGRZGardIn = QtWidgets.QLineEdit(self.frmCntrlPanel)
        # self.leMainGRZGardIn.setGeometry QRect(220, 300, 151, 33))
        # self.leMainGRZGardIn.setObjectName("leMainGRZGardIn")
        self.leMainGRZGardOutBack = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardOutBack.setGeometry QRect(220, 340, 162, 33))
        self.leMainGRZGardOutBack.setObjectName("leMainGRZGardOutBack")
        self.leMainGRZWeightOutBack = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightOutBack.setGeometry QRect(220, 420, 162, 33))
        self.leMainGRZWeightOutBack.setObjectName("leMainGRZWeightOutBack")
        self.leMainGRZWeightInBack = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightInBack.setGeometry QRect(220, 380, 162, 33))
        self.leMainGRZWeightInBack.setObjectName("leMainGRZWeightInBack")
        self.leMainGRZGardInBack = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardInBack.setGeometry QRect(220, 300, 162, 33))
        self.leMainGRZGardInBack.setObjectName("leMainGRZGardInBack")
        self.leMainGRZGardIn1 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardIn1.setGeometry QRect(232, 303, 18, 26))
        self.leMainGRZGardIn1.setText("")
        self.leMainGRZGardIn1.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardIn1.setObjectName("leMainGRZGardIn1")
        self.leMainGRZGardIn2 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardIn2.setGeometry QRect(253, 303, 18, 26))
        self.leMainGRZGardIn2.setText("")
        self.leMainGRZGardIn2.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardIn2.setObjectName("leMainGRZGardIn2")
        self.leMainGRZGardIn3 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardIn3.setGeometry QRect(264, 303, 18, 26))
        self.leMainGRZGardIn3.setText("")
        self.leMainGRZGardIn3.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardIn3.setObjectName("leMainGRZGardIn3")
        self.leMainGRZGardIn4 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardIn4.setGeometry QRect(275, 303, 18, 26))
        self.leMainGRZGardIn4.setText("")
        self.leMainGRZGardIn4.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardIn4.setObjectName("leMainGRZGardIn4")
        self.leMainGRZGardIn6 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardIn6.setGeometry QRect(305, 303, 18, 26))
        self.leMainGRZGardIn6.setText("")
        self.leMainGRZGardIn6.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardIn6.setObjectName("leMainGRZGardIn6")
        self.leMainGRZGardIn5 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardIn5.setGeometry QRect(292, 303, 18, 26))
        self.leMainGRZGardIn5.setText("")
        self.leMainGRZGardIn5.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardIn5.setObjectName("leMainGRZGardIn5")
        self.leMainGRZGardIn7 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardIn7.setGeometry QRect(338, 303, 18, 26))
        self.leMainGRZGardIn7.setText("")
        self.leMainGRZGardIn7.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardIn7.setObjectName("leMainGRZGardIn7")
        self.leMainGRZGardIn8 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardIn8.setGeometry QRect(349, 303, 18, 26))
        self.leMainGRZGardIn8.setText("")
        self.leMainGRZGardIn8.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardIn8.setObjectName("leMainGRZGardIn8")
        self.leMainGRZGardIn9 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardIn9.setGeometry QRect(360, 303, 18, 26))
        self.leMainGRZGardIn9.setText("")
        self.leMainGRZGardIn9.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardIn9.setObjectName("leMainGRZGardIn9")
        self.leMainGRZGardOut8 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardOut8.setGeometry QRect(349, 343, 18, 26))
        self.leMainGRZGardOut8.setText("")
        self.leMainGRZGardOut8.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardOut8.setObjectName("leMainGRZGardOut8")
        self.leMainGRZGardOut5 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardOut5.setGeometry QRect(292, 343, 18, 26))
        self.leMainGRZGardOut5.setText("")
        self.leMainGRZGardOut5.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardOut5.setObjectName("leMainGRZGardOut5")
        self.leMainGRZGardOut2 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardOut2.setGeometry QRect(253, 343, 18, 26))
        self.leMainGRZGardOut2.setText("")
        self.leMainGRZGardOut2.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardOut2.setObjectName("leMainGRZGardOut2")
        self.leMainGRZGardOut1 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardOut1.setGeometry QRect(232, 343, 18, 26))
        self.leMainGRZGardOut1.setText("")
        self.leMainGRZGardOut1.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardOut1.setObjectName("leMainGRZGardOut1")
        self.leMainGRZGardOut9 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardOut9.setGeometry QRect(360, 343, 18, 26))
        self.leMainGRZGardOut9.setText("")
        self.leMainGRZGardOut9.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardOut9.setObjectName("leMainGRZGardOut9")
        self.leMainGRZGardOut3 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardOut3.setGeometry QRect(264, 343, 18, 26))
        self.leMainGRZGardOut3.setText("")
        self.leMainGRZGardOut3.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardOut3.setObjectName("leMainGRZGardOut3")
        self.leMainGRZGardOut7 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardOut7.setGeometry QRect(338, 343, 18, 26))
        self.leMainGRZGardOut7.setText("")
        self.leMainGRZGardOut7.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardOut7.setObjectName("leMainGRZGardOut7")
        self.leMainGRZGardOut4 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardOut4.setGeometry QRect(275, 343, 18, 26))
        self.leMainGRZGardOut4.setText("")
        self.leMainGRZGardOut4.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardOut4.setObjectName("leMainGRZGardOut4")
        self.leMainGRZGardOut6 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZGardOut6.setGeometry QRect(305, 343, 18, 26))
        self.leMainGRZGardOut6.setText("")
        self.leMainGRZGardOut6.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZGardOut6.setObjectName("leMainGRZGardOut6")
        self.leMainGRZWeightIn8 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightIn8.setGeometry QRect(349, 383, 18, 26))
        self.leMainGRZWeightIn8.setText("")
        self.leMainGRZWeightIn8.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightIn8.setObjectName("leMainGRZWeightIn8")
        self.leMainGRZWeightIn5 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightIn5.setGeometry QRect(292, 383, 18, 26))
        self.leMainGRZWeightIn5.setText("")
        self.leMainGRZWeightIn5.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightIn5.setObjectName("leMainGRZWeightIn5")
        self.leMainGRZWeightIn2 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightIn2.setGeometry QRect(253, 383, 18, 26))
        self.leMainGRZWeightIn2.setText("")
        self.leMainGRZWeightIn2.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightIn2.setObjectName("leMainGRZWeightIn2")
        self.leMainGRZWeightIn1 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightIn1.setGeometry QRect(232, 383, 18, 26))
        self.leMainGRZWeightIn1.setText("")
        self.leMainGRZWeightIn1.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightIn1.setObjectName("leMainGRZWeightIn1")
        self.leMainGRZWeightIn9 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightIn9.setGeometry QRect(360, 383, 18, 26))
        self.leMainGRZWeightIn9.setText("")
        self.leMainGRZWeightIn9.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightIn9.setObjectName("leMainGRZWeightIn9")
        self.leMainGRZWeightIn3 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightIn3.setGeometry QRect(264, 383, 18, 26))
        self.leMainGRZWeightIn3.setText("")
        self.leMainGRZWeightIn3.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightIn3.setObjectName("leMainGRZWeightIn3")
        self.leMainGRZWeightIn7 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightIn7.setGeometry QRect(338, 383, 18, 26))
        self.leMainGRZWeightIn7.setText("")
        self.leMainGRZWeightIn7.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightIn7.setObjectName("leMainGRZWeightIn7")
        self.leMainGRZWeightIn4 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightIn4.setGeometry QRect(275, 383, 18, 26))
        self.leMainGRZWeightIn4.setText("")
        self.leMainGRZWeightIn4.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightIn4.setObjectName("leMainGRZWeightIn4")
        self.leMainGRZWeightIn6 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightIn6.setGeometry QRect(305, 383, 18, 26))
        self.leMainGRZWeightIn6.setText("")
        self.leMainGRZWeightIn6.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightIn6.setObjectName("leMainGRZWeightIn6")
        self.leMainGRZWeightOut8 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightOut8.setGeometry QRect(349, 423, 18, 26))
        self.leMainGRZWeightOut8.setText("")
        self.leMainGRZWeightOut8.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightOut8.setObjectName("leMainGRZWeightOut8")
        self.leMainGRZWeightOut5 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightOut5.setGeometry QRect(292, 423, 18, 26))
        self.leMainGRZWeightOut5.setText("")
        self.leMainGRZWeightOut5.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightOut5.setObjectName("leMainGRZWeightOut5")
        self.leMainGRZWeightOut2 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightOut2.setGeometry QRect(253, 423, 18, 26))
        self.leMainGRZWeightOut2.setText("")
        self.leMainGRZWeightOut2.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightOut2.setObjectName("leMainGRZWeightOut2")
        self.leMainGRZWeightOut1 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightOut1.setGeometry QRect(232, 423, 18, 26))
        self.leMainGRZWeightOut1.setText("")
        self.leMainGRZWeightOut1.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightOut1.setObjectName("leMainGRZWeightOut1")
        self.leMainGRZWeightOut9 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightOut9.setGeometry QRect(360, 423, 18, 26))
        self.leMainGRZWeightOut9.setText("")
        self.leMainGRZWeightOut9.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightOut9.setObjectName("leMainGRZWeightOut9")
        self.leMainGRZWeightOut3 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightOut3.setGeometry QRect(264, 423, 18, 26))
        self.leMainGRZWeightOut3.setText("")
        self.leMainGRZWeightOut3.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightOut3.setObjectName("leMainGRZWeightOut3")
        self.leMainGRZWeightOut7 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightOut7.setGeometry QRect(338, 423, 18, 26))
        self.leMainGRZWeightOut7.setText("")
        self.leMainGRZWeightOut7.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightOut7.setObjectName("leMainGRZWeightOut7")
        self.leMainGRZWeightOut4 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightOut4.setGeometry QRect(275, 423, 18, 26))
        self.leMainGRZWeightOut4.setText("")
        self.leMainGRZWeightOut4.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightOut4.setObjectName("leMainGRZWeightOut4")
        self.leMainGRZWeightOut6 = QtWidgets.QLineEdit(self.frmCntrlPanel)
        self.leMainGRZWeightOut6.setGeometry QRect(305, 423, 18, 26))
        self.leMainGRZWeightOut6.setText("")
        self.leMainGRZWeightOut6.setAlignment Qt.AlignLeading | Qt.AlignLeft | Qt.AlignVCenter)
        self.leMainGRZWeightOut6.setObjectName("leMainGRZWeightOut6")
        self.lblTbl_2.raise_()
        self.lblContDataWeight.raise_()
        self.lcdMainWeight.raise_()
        self.label_53.raise_()
        self.lblMainSemaLeftOutDown.raise_()
        self.line_8.raise_()
        self.label_55.raise_()
        self.lblNameWeight_3.raise_()
        self.lblMainSemaRightInUp.raise_()
        self.lblMainSemaLeftOutUp.raise_()
        self.lblIconDataWeightTr.raise_()
        self.lblMainSemaRightInDown.raise_()
        self.label_54.raise_()
        self.label_56.raise_()
        self.lblMainSemaLeftInUp.raise_()
        self.lblMainSemaLeftInDown.raise_()
        self.lblMainSemaRightOutDown.raise_()
        self.lblMainSemaRightOutUp.raise_()
        self.btnMainSemaREDOUT.raise_()
        self.btnMainSemaALLRED.raise_()
        self.btnMainSemaREDIN.raise_()
        self.lblContDataWeight_2.raise_()
        self.lblContDataWeight_3.raise_()
        self.btnMainShlgOpen.raise_()
        self.btnMainShlgClose.raise_()
        self.line_9.raise_()
        self.lblNameWeight_4.raise_()
        self.label_45.raise_()
        self.lblMainShlgOpen.raise_()
        self.label_20.raise_()
        self.label_21.raise_()
        self.lblMainShlgClose.raise_()
        self.lblMainShlgStateImage.raise_()
        self.lblNameWeight_5.raise_()
        self.line_10.raise_()
        self.lblNameWeight_6.raise_()
        self.lblNameWeight_7.raise_()
        self.lblNameWeight_8.raise_()
        self.lblNameWeight_9.raise_()
        self.btnMainGRZCh1.raise_()
        self.btnMainGRZCh2.raise_()
        self.btnMainGRZCh3.raise_()
        self.btnMainGRZCh4.raise_()
        self.btnHide.raise_()
        # self.leMainGRZGardOut.raise_()
        # self.leMainGRZWeightOut.raise_()
        # self.leMainGRZWeightIn.raise_()
        # self.leMainGRZGardIn.raise_()
        self.leMainGRZGardOutBack.raise_()
        self.leMainGRZWeightOutBack.raise_()
        self.leMainGRZWeightInBack.raise_()
        self.leMainGRZGardInBack.raise_()
        self.leMainGRZGardIn1.raise_()
        self.leMainGRZGardIn2.raise_()
        self.leMainGRZGardIn3.raise_()
        self.leMainGRZGardIn4.raise_()
        self.leMainGRZGardIn6.raise_()
        self.leMainGRZGardIn5.raise_()
        self.leMainGRZGardIn7.raise_()
        self.leMainGRZGardIn8.raise_()
        self.leMainGRZGardIn9.raise_()
        self.leMainGRZGardOut8.raise_()
        self.leMainGRZGardOut5.raise_()
        self.leMainGRZGardOut2.raise_()
        self.leMainGRZGardOut1.raise_()
        self.leMainGRZGardOut9.raise_()
        self.leMainGRZGardOut3.raise_()
        self.leMainGRZGardOut7.raise_()
        self.leMainGRZGardOut4.raise_()
        self.leMainGRZGardOut6.raise_()
        self.leMainGRZWeightIn8.raise_()
        self.leMainGRZWeightIn5.raise_()
        self.leMainGRZWeightIn2.raise_()
        self.leMainGRZWeightIn1.raise_()
        self.leMainGRZWeightIn9.raise_()
        self.leMainGRZWeightIn3.raise_()
        self.leMainGRZWeightIn7.raise_()
        self.leMainGRZWeightIn4.raise_()
        self.leMainGRZWeightIn6.raise_()
        self.leMainGRZWeightOut8.raise_()
        self.leMainGRZWeightOut5.raise_()
        self.leMainGRZWeightOut2.raise_()
        self.leMainGRZWeightOut1.raise_()
        self.leMainGRZWeightOut9.raise_()
        self.leMainGRZWeightOut3.raise_()
        self.leMainGRZWeightOut7.raise_()
        self.leMainGRZWeightOut4.raise_()
        self.leMainGRZWeightOut6.raise_()
        self.frmArchive = QtWidgets.QFrame(self.mainFormPr)
        self.frmArchive.setGeometry QRect(0, 129, 1920, 1011))
        self.frmArchive.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmArchive.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmArchive.setObjectName("frmArchive")
        self.le_ArchVideoBack = QtWidgets.QLabel(self.frmArchive)
        self.le_ArchVideoBack.setGeometry QRect(9, 0, 1512, 941))
        self.le_ArchVideoBack.setText("")
        self.le_ArchVideoBack.setObjectName("le_ArchVideoBack")
        self.btnBackToMain = QtWidgets.QPushButton(self.frmArchive)
        self.btnBackToMain.setGeometry QRect(1715, 910, 196, 31))
        font = QFont()
        font.setPointSize(10)
        self.btnBackToMain.setFont(font)
        self.btnBackToMain.setObjectName("btnBackToMain")
        self.le_8 = QtWidgets.QLineEdit(self.frmArchive)
        self.le_8.setGeometry QRect(20, 6, 300, 20))
        self.le_8.setObjectName("le_8")
        self.frmArchMPlay = QtWidgets.QFrame(self.frmArchive)
        self.frmArchMPlay.setGeometry QRect(9, 910, 1512, 31))
        self.frmArchMPlay.setFrameShape(QtWidgets.QFrame.StyledPanel)
        self.frmArchMPlay.setFrameShadow(QtWidgets.QFrame.Raised)
        self.frmArchMPlay.setObjectName("frmArchMPlay")
        self.SldrArchMplay = Slider(self.frmArchMPlay)
        self.SldrArchMplay.setGeometry QRect(185, 6, 1305, 16))
        self.SldrArchMplay.setPageStep(1)
        self.SldrArchMplay.setOrientation Qt.Horizontal)
        self.SldrArchMplay.setTickPosition(QtWidgets.QSlider.NoTicks)
        self.SldrArchMplay.setObjectName("SldrArchMplay")
        self.btnStopArchMPlay = QtWidgets.QPushButton(self.frmArchMPlay)
        self.btnStopArchMPlay.setGeometry QRect(0, 0, 31, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnStopArchMPlay.setFont(font)
        self.btnStopArchMPlay.setText("")
        self.btnStopArchMPlay.setIconSize QSize(25, 25))
        self.btnStopArchMPlay.setObjectName("btnStopArchMPlay")
        self.btnPlayArchMPlay = QtWidgets.QPushButton(self.frmArchMPlay)
        self.btnPlayArchMPlay.setGeometry QRect(62, 0, 31, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnPlayArchMPlay.setFont(font)
        self.btnPlayArchMPlay.setText("")
        self.btnPlayArchMPlay.setIconSize QSize(25, 25))
        self.btnPlayArchMPlay.setObjectName("btnPlayArchMPlay")
        self.btnStepforwardArchMPlay = QtWidgets.QPushButton(self.frmArchMPlay)
        self.btnStepforwardArchMPlay.setGeometry QRect(93, 0, 31, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnStepforwardArchMPlay.setFont(font)
        self.btnStepforwardArchMPlay.setText("")
        self.btnStepforwardArchMPlay.setIconSize QSize(25, 25))
        self.btnStepforwardArchMPlay.setObjectName("btnStepforwardArchMPlay")
        self.btnStepbackArchMPlay = QtWidgets.QPushButton(self.frmArchMPlay)
        self.btnStepbackArchMPlay.setGeometry QRect(31, 0, 31, 31))
        font = QFont()
        font.setPointSize(12)
        self.btnStepbackArchMPlay.setFont(font)
        self.btnStepbackArchMPlay.setText("")
        self.btnStepbackArchMPlay.setIconSize QSize(25, 25))
        self.btnStepbackArchMPlay.setObjectName("btnStepbackArchMPlay")
        self.cbArchMPlay = QtWidgets.QComboBox(self.frmArchMPlay)
        self.cbArchMPlay.setGeometry QRect(124, 0, 48, 31))
        self.cbArchMPlay.setObjectName("cbArchMPlay")
        self.cbArchMPlay.addItem("")
        self.cbArchMPlay.addItem("")
        self.cbArchMPlay.addItem("")
        self.cbArchMPlay.addItem("")
        self.label_6 = QtWidgets.QLabel(self.frmArchive)
        self.label_6.setGeometry QRect(1610, 5, 211, 21))
        self.label_6.setObjectName("label_6")
        self.label_7 = QtWidgets.QLabel(self.frmArchive)
        self.label_7.setGeometry QRect(1519, 0, 392, 31))
        self.label_7.setText("")
        self.label_7.setObjectName("label_7")
        self.vScrollArch = QtWidgets.QScrollBar(self.frmArchive)
        self.vScrollArch.setGeometry QRect(1901, 30, 10, 881))
        self.vScrollArch.setOrientation Qt.Vertical)
        self.vScrollArch.setObjectName("vScrollArch")
        self.line = QtWidgets.QFrame(self.frmArchive)
        self.line.setGeometry QRect(1521, 30, 388, 2))
        self.line.setStyleSheet("background-color: rgb(42,42,42);")
        self.line.setFrameShape(QtWidgets.QFrame.HLine)
        self.line.setFrameShadow(QtWidgets.QFrame.Sunken)
        self.line.setObjectName("line")
        self.le_ArchVideoStream = QVideoWidget(self.frmArchive)
        self.le_ArchVideoStream.setGeometry QRect(11, 31, 1509, 849))
        self.le_ArchVideoStream.setObjectName("le_ArchVideoStream")
        self.le_ArchFilename = QtWidgets.QLineEdit(self.frmArchive)
        self.le_ArchFilename.setGeometry QRect(140, 6, 411, 20))
        self.le_ArchFilename.setText("")
        self.le_ArchFilename.setObjectName("le_ArchFilename")
        self.lstArchive = QtWidgets.QTreeWidget(self.frmArchive)
        self.lstArchive.setGeometry QRect(1519, 6, 392, 904))
        self.lstArchive.setObjectName("treeArchive")
        self.label_10 = QtWidgets.QLabel(self.frmArchive)
        self.label_10.setGeometry QRect(1715, 910, 1, 31))
        self.label_10.setText("")
        self.label_10.setObjectName("label_10")
        self.label_9 = QtWidgets.QLabel(self.frmArchive)
        self.label_9.setGeometry QRect(1521, 910, 1, 31))
        self.label_9.setText("")
        self.label_9.setObjectName("label_9")
        self.label_8 = QtWidgets.QLabel(self.frmArchive)
        self.label_8.setGeometry QRect(1522, 910, 193, 31))
        self.label_8.setText("")
        self.label_8.setObjectName("label_8")
        self.btnBackToMain.raise_()
        self.lstArchive.raise_()
        self.label_7.raise_()
        self.label_6.raise_()
        self.line.raise_()
        self.le_ArchVideoBack.raise_()
        self.le_ArchVideoStream.raise_()
        self.vScrollArch.raise_()
        self.frmArchMPlay.raise_()
        self.le_8.raise_()
        self.le_ArchFilename.raise_()
        self.label_10.raise_()
        self.label_9.raise_()
        self.label_8.raise_()
        self.frmArchive.raise_()
        self.frmCamAllMon.raise_()
        self.frmHeadExit.raise_()
        self.frmHeadImage.raise_()
        self.frmMenu.raise_()
        self.frmTable.raise_()
        self.frmCam.raise_()
        self.frmCntrlPanel.raise_()
        self.action123 = QtWidgets.QAction(self)
        self.action123.setCheckable(True)
        self.action123.setObjectName("action123")

        self.btnDayNightCam0.hide()
        self.btnDayNightCam1.hide()
        self.btnDayNightCam2.hide()
        self.btnDayNightCam3.hide()

        # self.lblMainSemaLeftOutUp.setStyleSheet(
        #     "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
        #     "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        #
        # self.lblMainSemaLeftOutDown.setStyleSheet(
        #     "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
        #     "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        #
        # self.lblMainSemaLeftInUp.setStyleSheet(
        #     "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
        #     "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        #
        # self.lblMainSemaLeftInDown.setStyleSheet(
        #     "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
        #     "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        #
        # self.lblMainSemaRightInUp.setStyleSheet(
        #     "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
        #     "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        #
        # self.lblMainSemaRightInDown.setStyleSheet(
        #     "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
        #     "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        #
        # self.lblMainSemaRightOutUp.setStyleSheet(
        #     "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
        #     "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        #
        # self.lblMainSemaRightOutDown.setStyleSheet(
        #     "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
        #     "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        #
        # self.lblMainShlgClose.setStyleSheet(
        #     "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
        #     "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
        #
        # self.lblMainShlgOpen.setStyleSheet(
        #     "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
        #     "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")

        self.comboSearch.addItem("")
        self.comboSearch.addItem("")
        self.comboSearch.addItem("")
        self.comboSearch.addItem("")
        self.retranslateUi()
     QMetaObject.connectSlotsByName(self)
        # self.runThJournal()
        self.firstSets()

    def retranslateUi(self):
        _translate = QCoreApplication.translate
        self.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.lblCatCams.setText(_translate("MainWindow", "Видеонаблюдение"))
        self.lblCam1.setText(_translate("MainWindow", "        КАМЕРА ВЫЕЗД"))
        self.lblCam2.setText(_translate("MainWindow", "   КАМЕРА ВЕСЫ ВЪЕЗД"))
        self.lblCam4.setText(_translate("MainWindow", "  КАМЕРА ВЕСЫ  ВЫЕЗД"))
        self.lblCam0.setText(_translate("MainWindow", "        КАМЕРА ВЪЕЗД"))
        self.btnCam0Exp.setToolTip(_translate("MainWindow", "Растянуть"))
        self.btnCam1Exp.setToolTip(_translate("MainWindow", "Растянуть"))
        self.btnCam2Exp.setToolTip(_translate("MainWindow", "Растянуть"))
        self.btnCam3Exp.setToolTip(_translate("MainWindow", "Растянуть"))
        self.btnDayNightCam0.setToolTip(_translate("MainWindow", "Растянуть"))
        self.btnDayNightCam2.setToolTip(_translate("MainWindow", "Растянуть"))
        self.btnDayNightCam1.setToolTip(_translate("MainWindow", "Растянуть"))
        self.btnDayNightCam3.setToolTip(_translate("MainWindow", "Растянуть"))
        self.btnCam0Arch.setToolTip(_translate("MainWindow", "Растянуть"))
        self.btnCam1Arch.setToolTip(_translate("MainWindow", "Растянуть"))
        self.btnCam3Arch.setToolTip(_translate("MainWindow", "Растянуть"))
        self.btnCam2Arch.setToolTip(_translate("MainWindow", "Растянуть"))
        self.lblNameJournal.setText(_translate("MainWindow", "Системный журнал"))
        self.tblMainData.setSortingEnabled(True)
        item = self.tblMainData.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "Талон №"))
        item = self.tblMainData.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "Дата"))
        item = self.tblMainData.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "Время работ"))
        item = self.tblMainData.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "№ ГРЗ ТС"))
        item = self.tblMainData.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "М.пустой,кг"))
        item = self.tblMainData.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "М.груженый,кг"))
        item = self.tblMainData.horizontalHeaderItem(6)
        item.setText(_translate("MainWindow", "М.грунта,кг"))
        item = self.tblMainData.horizontalHeaderItem(7)
        item.setText(_translate("MainWindow", "Объём грунта"))
        item = self.tblMainData.horizontalHeaderItem(8)
        item.setText(_translate("MainWindow", "Состояние З/Н"))
        self.btnSearch.setText(_translate("MainWindow", "      Поиск"))
        self.btnDelZakazSergey.setText('Удалить')
        self.btnAboutSetSystem.setText(_translate("MainWindow", "Справка"))
        self.btnStWorkingSystem.setText(_translate("MainWindow", "Режим работы ПО"))
        self.lblTitleMainSet.setText(_translate("MainWindow", "Панель управления"))
        self.btnStatusDevs.setText(_translate("MainWindow", "Статус системы"))
        self.btnScaner.setText(_translate("MainWindow", "Сканер"))
        self.btnJournal.setText(_translate("MainWindow", "Журнал событий"))
        self.btnGard.setText(_translate("MainWindow", "Контроль доступа"))
        self.btnAboutMain.setText(_translate("MainWindow", "Справка"))
        self.btnWeight.setText(_translate("MainWindow", "Весовой контроль"))
        self.btnDB.setText(_translate("MainWindow", "Базы данных"))
        self.btnSetCams.setText(_translate("MainWindow", "Видеонаблюдение"))
        self.lblNameSetSystem.setText(_translate("MainWindow", "Настройки системы"))
        self.btnDetectionCars.setText(_translate("MainWindow", "Отслеживание ТС"))
        self.btnColorScheme.setText(_translate("MainWindow", "Цветовая тема"))
        self.btnuserchange.setToolTip(_translate("MainWindow", "Сменить пользователя"))
        self.btnJournalTS.setText(_translate("MainWindow", "Журнал З/Н ТС"))
        self.btnBDTS.setText(_translate("MainWindow", "База ТС"))
        self.btnAddZakaz.setText(_translate("MainWindow", "Добавить З/Н"))
        self.lblBarClose.setText(_translate("MainWindow", "АСМК-ГРУНТ"))
        self.btnCamCallapse.setToolTip(_translate("MainWindow", "Растянуть"))
        self.lblIconDataWeightTr.setToolTip(_translate("MainWindow", "Весы и светофоры"))
        self.lblNameWeight_3.setText(_translate("MainWindow", "Весовой комплекс"))
        self.btnMainSemaREDOUT.setText(_translate("MainWindow", "Разрешить въезд"))
        self.btnMainSemaALLRED.setText(_translate("MainWindow", "Кругом красный"))
        self.btnMainSemaREDIN.setText(_translate("MainWindow", "Разрешить выезд"))
        self.btnMainShlgOpen.setText(_translate("MainWindow", "Открыть"))
        self.btnMainShlgClose.setText(_translate("MainWindow", "Закрыть"))
        self.lblNameWeight_4.setText(_translate("MainWindow", "Шлагбаум"))
        self.label_20.setText(_translate("MainWindow", "Открыт"))
        self.label_21.setText(_translate("MainWindow", "Закрыт"))
        self.lblNameWeight_5.setText(_translate("MainWindow", "Распознавание ГРЗ"))
        self.lblNameWeight_6.setText(_translate("MainWindow", "КПП   ВЪЕЗД"))
        self.lblNameWeight_7.setText(_translate("MainWindow", "КПП   ВЫЕЗД"))
        self.lblNameWeight_8.setText(_translate("MainWindow", "ВЕСЫ ВЪЕЗД"))
        self.lblNameWeight_9.setText(_translate("MainWindow", "ВЕСЫ ВЫЕЗД"))
        # self.btnArchBackToMain.setText(_translate("MainWindow", "Выбрать файл"))
        # self.btnArchOpenFile.setText(_translate("MainWindow", "Закрыть архив"))
        self.le_8.setText(_translate("MainWindow", "Название файла:"))
        self.label_6.setText(_translate("MainWindow", "Список файлов"))
        self.cbArchMPlay.setItemText(0, _translate("MainWindow", "0.5x"))
        self.cbArchMPlay.setItemText(1, _translate("MainWindow", "1x"))
        self.cbArchMPlay.setItemText(2, _translate("MainWindow", "1.5x"))
        self.cbArchMPlay.setItemText(3, _translate("MainWindow", "2x"))
        self.comboSearch.setItemText(0, _translate("MainWindow", "Нет"))
        self.comboSearch.setItemText(1, _translate("MainWindow", "Талон"))
        self.comboSearch.setItemText(2, _translate("MainWindow", "ГРЗ"))
        self.comboSearch.setItemText(3, _translate("MainWindow", "Дата"))
        self.action123.setText(_translate("MainWindow", "123"))
        self.btnBackToMain.setText(_translate("MainWindow", "Назад в гланое меню     "))

    # def addVScroll(self):
    #     self.verticalScrollBar = QtWidgets.QScrollBar(self.frmTable)
    #     self.verticalScrollBar.setGeometry QRect(841, 102, 10, 327))
    #     self.verticalScrollBar.setOrientation Qt.Vertical)
    #     self.verticalScrollBar.setObjectName("verticalScrollBar")
    #     self.verticalScrollBar.setMaximum(self.tblMainData.verticalScrollBar().maximum())
    #
    #     self.tblMainData.setVerticalScrollBarPolicy Qt.ScrollBarAlwaysOff)
    #     self.tblMainData.verticalHeader().setVisible(False)

    def sync_func(self):
        self.tblMainData.verticalScrollBar().setValue(self.verticalScrollBar.value())

    def sync_func_strg(self):
        self.lstArchive.verticalScrollBar().setValue(self.vScrollArch.value())

    def firstSets(self):
            # createImgsWeight(0, ['rtsp://admin:qwe123456@10.2.165.101:554/cam/realmonitor?channel=1&subtype=1'], ['10.2.165.101'], 'E:/', 100001)

            globalValues.lstTimingDayNight = self.lstTimingFromXls()

            print(globalValues.lstTimingDayNight)

            if (globalValues.colorForm == 1):
                self.leCam0.setStyleSheet("background-color: rgb(0,0,0);\n"
                              "border-radius: 5px;\n"
                              "border: 2px solid rgb(42,42,42);\n"
                              "image: url(" + globalValues.pathStyleImgs + "camDefault400x400.png);")
            else:
                self.leCam0.setStyleSheet("background-color: rgb(235,235,235);\n"
                              "border-radius: 5px;\n"
                              "border: 2px solid rgb(150,150,150);\n"
                              "image: url(" + globalValues.pathStyleImgs + "camDefault400x400grey.png);")

            self.le_UserName.setEnabled(False)
            self.lstArchive.clear()
            self.lstArchive.header
            self.le_ArchFilename.hide()
            self.le_8.setEnabled(False)
            self.mediaPlayerVideoStrg = QMediaPlayer(None, QMediaPlayer.VideoSurface)
            self.mediaPlayerVideoStrg.setVideoOutput(self.le_ArchVideoStream)
            # self.setMediaFrm(globalValues.pathFileVideo, self.mediaPlayerVideoStrg)

            self.SldrArchMplay.sliderMoved.connect(self.setPositionStrg)
            self.SldrArchMplay.sliderPressed.connect(self.pressSldrStrg)
            self.mediaPlayerVideoStrg.positionChanged.connect(self.positionChangedStrg)
            self.mediaPlayerVideoStrg.durationChanged.connect(self.durationChangedStrg)
            self.mediaPlayerVideoStrg.error.connect(self.handleError)

            self.btnPlayArchMPlay.clicked.connect(self.playVideoStrg)
            # self.btnPauseArchMPlay.clicked.connect(self.pauseVideoStrg)
            self.btnStopArchMPlay.clicked.connect(self.stopVideoStrg)
            self.btnStepbackArchMPlay.clicked.connect(self.jumpBackStrg)
            self.btnStepforwardArchMPlay.clicked.connect(self.jumpForwardStrg)
            self.cbArchMPlay.currentIndexChanged.connect(self.playBackRateStrg)

            self.btnZoomIn1.clicked.connect(self.measuringVolume)

            self.cbArchMPlay.setCurrentIndex(1)

            # self.createListVideos()

            self.lstArchive.itemSelectionChanged.connect(self.setMediaInFrmClk)

            self.leCam0Rec.setEnabled(False)
            self.leCam1Rec.setEnabled(False)
            self.leCam2Rec.setEnabled(False)
            self.leCam3Rec.setEnabled(False)

            self.lblMainSemaLeftOutUp.setStyleSheet(
                "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
            self.lblMainSemaLeftOutDown.setStyleSheet(
                "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
            self.lblMainSemaLeftInUp.setStyleSheet(
                "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
            self.lblMainSemaLeftInDown.setStyleSheet(
                "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
            self.lblMainSemaRightInUp.setStyleSheet(
                "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
            self.lblMainSemaRightInDown.setStyleSheet(
                "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
            self.lblMainSemaRightOutUp.setStyleSheet(
                "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
            self.lblMainSemaRightOutDown.setStyleSheet(
                "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
            self.lblMainShlgOpen.setStyleSheet(
                "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
            self.lblMainShlgClose.setStyleSheet(
                "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
            self.lblMainShlgStateImage.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                "image: url(" + globalValues.pathStyleImgs + "iconshlMIDDLE.png);")

            self.leCam0Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                 "image: url(" + globalValues.pathStyleImgs + "iconrec13grey.png);")

            self.leCam1Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrec13grey.png);")

            self.leCam2Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrec13grey.png);")

            self.leCam3Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrec13grey.png);")

            # self.setFontLst(6, 'Arial', [self.leMainGRZGardIn1, self.leMainGRZGardIn2, self.leMainGRZGardIn3,
            #                               self.leMainGRZGardIn4, self.leMainGRZGardIn5, self.leMainGRZGardIn6,
            #                               self.leMainGRZGardIn7, self.leMainGRZGardIn8, self.leMainGRZGardIn9,
            #                               self.leMainGRZGardOut1, self.leMainGRZGardOut2, self.leMainGRZGardOut3,
            #                               self.leMainGRZGardOut4, self.leMainGRZGardOut5, self.leMainGRZGardOut6,
            #                               self.leMainGRZGardOut7, self.leMainGRZGardOut8, self.leMainGRZGardOut9,
            #                               self.leMainGRZWeightIn1, self.leMainGRZWeightIn2, self.leMainGRZWeightIn3,
            #                               self.leMainGRZWeightIn4, self.leMainGRZWeightIn5, self.leMainGRZWeightIn6,
            #                               self.leMainGRZWeightIn7, self.leMainGRZWeightIn8, self.leMainGRZWeightIn9,
            #                               self.leMainGRZWeightOut1, self.leMainGRZWeightOut2, self.leMainGRZWeightOut3,
            #                               self.leMainGRZWeightOut4, self.leMainGRZWeightOut5, self.leMainGRZWeightOut6,
            #                               self.leMainGRZWeightOut7, self.leMainGRZWeightOut8, self.leMainGRZWeightOut9])

            self.leMainGRZGardInBack.setEnabled(False)
            self.leMainGRZGardIn1.setEnabled(False)
            self.leMainGRZGardIn2.setEnabled(False)
            self.leMainGRZGardIn3.setEnabled(False)
            self.leMainGRZGardIn4.setEnabled(False)
            self.leMainGRZGardIn5.setEnabled(False)
            self.leMainGRZGardIn6.setEnabled(False)
            self.leMainGRZGardIn7.setEnabled(False)
            self.leMainGRZGardIn8.setEnabled(False)
            self.leMainGRZGardIn9.setEnabled(False)

            self.leMainGRZGardOutBack.setEnabled(False)
            self.leMainGRZGardOut1.setEnabled(False)
            self.leMainGRZGardOut2.setEnabled(False)
            self.leMainGRZGardOut3.setEnabled(False)
            self.leMainGRZGardOut4.setEnabled(False)
            self.leMainGRZGardOut5.setEnabled(False)
            self.leMainGRZGardOut6.setEnabled(False)
            self.leMainGRZGardOut7.setEnabled(False)
            self.leMainGRZGardOut8.setEnabled(False)
            self.leMainGRZGardOut9.setEnabled(False)

            self.leMainGRZWeightInBack.setEnabled(False)
            self.leMainGRZWeightIn1.setEnabled(False)
            self.leMainGRZWeightIn2.setEnabled(False)
            self.leMainGRZWeightIn3.setEnabled(False)
            self.leMainGRZWeightIn4.setEnabled(False)
            self.leMainGRZWeightIn5.setEnabled(False)
            self.leMainGRZWeightIn6.setEnabled(False)
            self.leMainGRZWeightIn7.setEnabled(False)
            self.leMainGRZWeightIn8.setEnabled(False)
            self.leMainGRZWeightIn9.setEnabled(False)

            self.leMainGRZWeightOutBack.setEnabled(False)
            self.leMainGRZWeightOut1.setEnabled(False)
            self.leMainGRZWeightOut2.setEnabled(False)
            self.leMainGRZWeightOut3.setEnabled(False)
            self.leMainGRZWeightOut4.setEnabled(False)
            self.leMainGRZWeightOut5.setEnabled(False)
            self.leMainGRZWeightOut6.setEnabled(False)
            self.leMainGRZWeightOut7.setEnabled(False)
            self.leMainGRZWeightOut8.setEnabled(False)
            self.leMainGRZWeightOut9.setEnabled(False)

            self.lstLight = [[self.comboSearch, "background-color: rgb(227,227,227);\n"
                                                "color: black;\n"
                                                "border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);"],
                             [self.tblMainData,
                              "QTableWidget {background-color: rgb(235,235,235);\n"
                                                "border: 1px solid rgb(150,150,150);\n"
                                                "gridline-color: rgb(42,42,42);\n"
                                                "border-radius:5px;\n"
                                                "border-bottom-right-radius: 0px;\n"
                                                "color:black;}\n"
                                                "QLineEdit {background-color: white;}\n"
                                                "QHeaderView::section {\n"
                              #                   "gridline-color: rgb(89,89,89);\n"
                                                "background-color: rgb(142,187,208);\n"
                                                "color: black;};\n"],
                             [self.mainFormPr, "background-color: rgb(242,242,242);"],
                             [self.frmCam, "background-color: rgb(252,252,252);"],
                             [self.lblTitleCams, "background-color: rgb(242,242,242);\n"
                                                 "color: rgb(0,0,0);\n"
                                                 "border-radius: 5px;\n"
                                                 "border-left: 2px solid rgb(205,205,205);\n"
                                                 "border-top: 2px solid rgb(150,150,150);\n"
                                                 "border-right: 2px solid rgb(150,150,150);\n"
                                                 ""],
                             [self.lblCatCams, "background-color: rgb(242,242,242);\n"
                                               "color: rgb(0,0,0);"],
                             [self.lblIconCams,
                              "background-color:qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                              "image: url(" + globalValues.pathStyleImgs + "iconcam2.png);"],
                             [self.lblMainBoxCams, "background-color: rgb(242,242,242);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 2px solid rgb(205,205,205); "],
                             [self.lblCam1,
                              "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(252,252,252, 255), stop:1 rgba(150,150,150, 255));\n"
                              "color: rgb(0,0,0);\n"
                              "border-radius: 5px;\n"
                              "border: 2px solid rgb(150,150,150);"],
                             [self.lblCam2,
                              "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(252,252,252, 255), stop:1 rgba(150,150,150, 255));\n"
                              "color: rgb(0,0,0);\n"
                              "border-radius: 5px;\n"
                              "border: 2px solid rgb(150,150,150);"],
                             [self.lblCam4,
                              "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(252,252,252, 255), stop:1 rgba(150,150,150, 255));\n"
                              "color: rgb(0,0,0);\n"
                              "border-radius: 5px;\n"
                              "border: 2px solid rgb(150,150,150);"],
                             [self.lblCam0,
                              "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(252,252,252, 255), stop:1 rgba(150,150,150, 255));\n"
                              "color: rgb(0,0,0);\n"
                              "border-radius: 5px;\n"
                              "border: 2px solid rgb(150,150,150);"],
                             [self.leCam0, "background-color: rgb(235,235,235);\n"
                                           "border-radius: 5px;\n"
                                           "border: 2px solid rgb(150,150,150);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "camDefault400x400grey.png);"],
                             [self.leCam1, "background-color: rgb(235,235,235);\n"
                                           "border-radius: 5px;\n"
                                           "border: 2px solid rgb(150,150,150);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "camDefault400x400grey.png);"],
                             [self.leCam2, "background-color: rgb(235,235,235);\n"
                                           "border-radius: 5px;\n"
                                           "border: 2px solid rgb(150,150,150);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "camDefault400x400grey.png);"],
                             [self.leCam3, "background-color: rgb(235,235,235);\n"
                                           "border-radius: 5px;\n"
                                           "border: 2px solid rgb(150,150,150);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "camDefault400x400grey.png);"],
                             [self.btnCam0Exp, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "background-color: rgb(227,227,227);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconmakebigpnggrey1717.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);};"],
                             [self.btnCam1Exp, "QPushButton:!hover {border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "background-color: rgb(227,227,227);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconmakebigpnggrey1717.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);};"],
                             [self.btnCam2Exp, "QPushButton:!hover {border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "background-color: rgb(227,227,227);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconmakebigpnggrey1717.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);};"],
                             [self.btnCam3Exp, "QPushButton:!hover {border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "background-color: rgb(227,227,227);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconmakebigpnggrey1717.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);};"],
                             [self.label_2, "background-color: rgb(242,242,242);"],
                             [self.label_3, "background-color: rgb(242,242,242);"],
                             [self.btnDayNightCam0, "QPushButton:!hover {border-radius: 3px;\n"
                                                    "border: 1px solid rgb(135,135,135);\n"
                                                    "background-color: rgb(227,227,227);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "icondaynight3grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                    "border-radius: 5px;\n"
                                                    "border: 1px solid rgb(135,135,135);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                    "border-radius: 5px;\n"
                                                    "border: 1px solid rgb(135,135,135);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "icondaynight3white.png);};"],
                             [self.btnDayNightCam2, "QPushButton:!hover {border-radius: 3px;\n"
                                                    "border: 1px solid rgb(135,135,135);\n"
                                                    "background-color: rgb(227,227,227);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "icondaynight3grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                    "border-radius: 5px;\n"
                                                    "border: 1px solid rgb(135,135,135);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                    "border-radius: 5px;\n"
                                                    "border: 2px solid rgb(135,135,135);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "icondaynight3white.png);};"],
                             [self.btnDayNightCam1, "QPushButton:!hover {border-radius: 3px;\n"
                                                    "border: 1px solid rgb(135,135,135);\n"
                                                    "background-color: rgb(227,227,227);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "icondaynight3grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                    "border-radius: 5px;\n"
                                                    "border: 1px solid rgb(135,135,135);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                    "border-radius: 5px;\n"
                                                    "border: 1px solid rgb(135,135,135);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "icondaynight3white.png);};"],
                             [self.btnDayNightCam3, "QPushButton:!hover {border-radius: 3px;\n"
                                                    "border: 1px solid rgb(135,135,135);\n"
                                                    "background-color: rgb(227,227,227);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "icondaynight3grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                    "border-radius: 5px;\n"
                                                    "border: 1px solid rgb(135,135,135);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                    "border-radius: 5px;\n"
                                                    "border: 2px solid rgb(135,135,135);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "icondaynight3white.png);};"],
                             [self.btnCam0Arch, "QPushButton:!hover {border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "background-color: rgb(227,227,227);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconarch6grey.png);}\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);};"],
                             [self.btnCam1Arch, "QPushButton:!hover {border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "background-color: rgb(227,227,227);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconarch6grey.png);}\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);};"],
                             [self.btnCam3Arch, "QPushButton:!hover {border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "background-color: rgb(227,227,227);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconarch6grey.png);}\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);};"],
                             [self.btnCam2Arch, "QPushButton:!hover {border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "background-color: rgb(227,227,227);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconarch6grey.png);}\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);};"],
                             [self.frmTable, "background-color: rgb(252,252,252);"],
                             [self.lblBoxJournal, "background-color: rgb(242,242,242);\n"
                                                  "color: rgb(0,0,0);\n"
                                                  "border-radius: 5px;\n"
                                                  "border-left: 2px solid rgb(205,205,205);\n"
                                                  "border-top: 2px solid rgb(150,150,150);\n"
                                                  "border-right: 2px solid rgb(150,150,150);\n"
                                                  ""],
                             [self.lblNameJournal, "background-color: rgb(242,242,242);\n"
                                                   "color: rgb(0,0,0);"],
                             [self.lblIconJournal,
                              "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                              "image: url(" + globalValues.pathStyleImgs + "iconworkpanel1.png);"],
                             [self.lblTbl, "background-color: rgb(242,242,242);\n"
                                           "border-radius: 5px;\n"
                                           "border: 2px solid rgb(205,205,205);"],
                             [self.leSearchTbl, "background-color: rgb(235,235,235);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border-radius: 3px;\n"
                                                "border: 1px solid rgb(150,150,150);"],
                             [self.btnSearch, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                              "color: rgb(0,0,0);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconsearch8.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconsearchwhite.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconsearchwhite.png);};"],
                             [self.btnDelZakazSergey, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                              "color: rgb(0,0,0);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(135,135,135);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(135,135,135);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(135,135,135);}\n"],
                             [self.btnRefresh, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                               "color: rgb(0,0,0);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconrefreshgrey.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconrefreshwhite.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconrefreshwhite.png);};"],
                             [self.label_5, "background-color: rgb(242,242,242);"],
                             [self.label_4, "background-color: rgb(242,242,242);"],
                             [self.verticalScrollBar, "QScrollBar:vertical {\n"
                                                      "border: 1px solid rgb(89,89,89);\n"
                                                      " background: rgb(255,255,255);\n"
                                                      "width:10px;\n"
                                                      "margin: 0px 0px 0px 0px;\n"
                                                      "}\n"
                                                      "QScrollBar::handle:vertical {\n"
                                                      "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0 rgb(142,187,208), stop: 0.5 rgb(142,187,208), stop:1 rgb(142,187,208));\n"
                                                      "min-height: 0px;\n"
                                                      "}\n"
                                                      "QScrollBar::add-line:vertical {\n"
                                                      " background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop: 0 rgb(142,187,208), stop: 0.5 rgb(142,187,208),  stop:1 rgb(142,187,208));\n"
                                                      "height: 0px;\n"
                                                      "subcontrol-position: bottom;\n"
                                                      "subcontrol-origin: margin;\n"
                                                      "}\n"
                                                      "QScrollBar::sub-line:vertical {\n"
                                                      "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0  rgb(142,187,208), stop: 0.5 rgb(142,187,208),  stop:1 rgb(142,187,208));\n"
                                                      "height: 0 px;\n"
                                                      "subcontrol-position: top;\n"
                                                      "subcontrol-origin: margin;\n"
                                                      "}"],
                             [self.DateEdit, "background-color: rgb(227,227,227);\n"
                                             "color: black;\n"
                                             "border-radius: 3px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "font: 11pt \"MS Shell Dlg 2\";\n"
                                             "border: 1px solid rgb(63,63,63)"],
                             [self.frmMenu, "background-color: rgb(252,252,252);"],
                             [self.lblBoxSets, "background-color: rgb(235,235,235);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(150,150,150);"],
                             [self.btnAboutSetSystem, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                      "color: rgb(0,0,0);\n"
                                                      "border-radius:3px;\n"
                                                      "border:1px solid rgb(135,135,135);}\n"
                                                      "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                      "color: rgb(255, 255, 255);\n"
                                                      "border-radius:3px;\n"
                                                      "border:1px solid rgb(63,63,63);}\n"
                                                      "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                      "color: rgb(255, 255, 255);\n"
                                                      "border-radius:3px;\n"
                                                      "border:1px solid rgb(63,63,63);};"],
                             [self.btnStWorkingSystem, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                       "color: rgb(0,0,0);\n"
                                                       "border:1px solid rgb(135,135,135);}\n"
                                                       "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                       "color: rgb(255, 255, 255);\n"
                                                       "border-radius:3px;\n"
                                                       "border:1px solid rgb(63,63,63);}\n"
                                                       "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                       "color: rgb(255, 255, 255);\n"
                                                       "border-radius:3px;\n"
                                                       "border:1px solid rgb(63,63,63);};"],
                             [self.lblIconMainSet,
                              "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                              "image: url(" + globalValues.pathStyleImgs + "iconwork5.png);"],
                             [self.lblTitleMainSet, "background-color: rgb(235,235,235);\n"
                                                    "color: rgb(160,160,160);"],
                             [self.btnStatusDevs, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                  "color: rgb(0,0,0);\n"
                                                  "border:1px solid rgb(135,135,135);}\n"
                                                  "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);}\n"
                                                  "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);};"],
                             [self.btnScaner, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                  "color: rgb(0,0,0);\n"
                                                  "border:1px solid rgb(135,135,135);}\n"
                                                  "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);}\n"
                                                  "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);};"],
                             [self.lblIconSetSystem,
                              "background-color:qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                              "image: url(" + globalValues.pathStyleImgs + "iset1.png);\n"
                              "border-radius: 7px;"],
                             [self.btnJournal, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                               "color: rgb(0,0,0);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(135,135,135);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(63,63,63);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(63,63,63);};"],
                             [self.btnGard, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                            "color: rgb(0,0,0);\n"
                                            "border:1px solid rgb(135,135,135);}\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);}\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);};"],
                             [self.lineSetMain, "background-color: rgb(252,252,252);"],
                             [self.btnAboutMain, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                 "color: rgb(0,0,0);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(135,135,135);}\n"
                                                 "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);};"],
                             [self.btnWeight, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                              "color: rgb(0,0,0);\n"
                                              "border:1px solid rgb(135,135,135);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);};"],
                             [self.btnDB, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                          "color: rgb(0,0,0);\n"
                                          "border:1px solid rgb(135,135,135);}\n"
                                          "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                          "color: rgb(255, 255, 255);\n"
                                          "border-radius:3px;\n"
                                          "border:1px solid rgb(63,63,63);}\n"
                                          "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                          "color: rgb(255, 255, 255);\n"
                                          "border-radius:3px;\n"
                                          "border:1px solid rgb(63,63,63);};"],
                             [self.lineSetsSystem, "background-color: rgb(252,252,252);"],
                             [self.btnSetCams, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                               "color: rgb(0,0,0);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(135,135,135);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(63,63,63);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius:3px;\n"
                                               "border:1px solid rgb(63,63,63);};"],
                             [self.lblNameSetSystem, "background-color: rgb(235,235,235);\n"
                                                     "color: rgb(160,160,160);"],
                             [self.btnDetectionCars, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                     "color: rgb(0,0,0);\n"
                                                     "border:1px solid rgb(135,135,135);}\n"
                                                     "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);}\n"
                                                     "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);};"],
                             [self.btnColorScheme, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                   "color: rgb(0,0,0);\n"
                                                   "border:1px solid rgb(135,135,135);}\n"
                                                   "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                   "color: rgb(255, 255, 255);\n"
                                                   "border-radius:3px;\n"
                                                   "border:1px solid rgb(63,63,63);}\n"
                                                   "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                   "color: rgb(255, 255, 255);\n"
                                                   "border-radius:3px;\n"
                                                   "border:1px solid rgb(63,63,63);};"],
                             [self.btnuserchange, "QPushButton:!hover {background-color: rgb(235,235,235);\n"
                                                  "image: url(" + globalValues.pathStyleImgs + "iconuchange.png);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(235,235,235);}\n"
                                                  "QPushButton:hover {background-color: rgb(75,75,75);\n"
                                                  "image: url(" + globalValues.pathStyleImgs + "iconuserch6.png);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(75,75,75);}\n"
                                                  "QPushButton:hover:pressed {background-color: rgb(75,75,75);\n"
                                                  "image: url(" + globalValues.pathStyleImgs + "iconuserch2.png);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(75,75,75);};"],
                             [self.lineSetsSystem_2, "background-color: rgb(235,235,235);"],
                             [self.le_UserName, "background-color: rgb(235,235,235);\n"
                                                "color: rgb(0,0,0);\n"
                                                "font: 12pt \"Arial\";\n"
                                                "border:1px solid(235,235,235);"],
                             [self.btnJournalTS, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                 "color: rgb(0,0,0);\n"
                                                 "border:1px solid rgb(135,135,135);}\n"
                                                 "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);};"],
                             [self.btnBDTS, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                            "color: rgb(0,0,0);\n"
                                            "border:1px solid rgb(135,135,135);}\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);}\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "color: rgb(255, 255, 255);\n"
                                            "border-radius:3px;\n"
                                            "border:1px solid rgb(63,63,63);};"],
                             [self.btnAddZakaz, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                "color: rgb(0,0,0);\n"
                                                "border:1px solid rgb(135,135,135);}\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(63,63,63);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(63,63,63);};"],
                             [self.frmHeadImage, "background-color: rgb(242,242,242);"],
                             [self.lblBoxUpCompany, "background-image: url(" + globalValues.pathStyleImgs + "iconheadwhite5.png);\n"
                                                    "border-radius: 5px;\n"
                                                    "border: 3px solid rgb(205,205,205);\n"],
                                                    # "image: url(" + globalValues.pathStyleImgs + "sinapswhiteHEAD2.png);"],
                             [self.frmHeadExit, "background-color: rgb(205,205,205);"],
                             [self.lblBarClose, "background-color: rgb(255,255,255);\n"
                                                "color: rgb(0,0,0);\n"
                                                ""],
                             [self.lblIconBarClose, "background-color: rgb(255,255,255);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "sinaps1717.png);"],
                             [self.btnCloseMainForm, "QPushButton:!hover{ background-color: rgb(255,255,255);\n"
                                                     "border-radius: 5px;\n"
                                                     "image: url(" + globalValues.pathStyleImgs + "iconextbl1919.png);}\n"
                                                     "QPushButton:hover { background-color: rgb(84,122,181);\n"
                                                     "border-radius: 3px;\n"
                                                     "image: url(" + globalValues.pathStyleImgs + "iconext1919.png);}\n"
                                                     "QPushButton:hover:pressed { background-color: rgb(50,75,115);\n"
                                                     "border-radius: 3px;\n"
                                                     "image: url(" + globalValues.pathStyleImgs + "iconext1919.png);};\n"
                                                     ""],
                             [self.label, "background-color: rgb(255,255,255);"],
                             [self.frmCntrlPanel, "background-color: rgb(252,252,252);\n"
                                                  ""],
                             [self.lblIconDataWeightTr, "background-color: rgb(235,235,235);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconweightnewgrey2323.png);"],
                             [self.label_54, "background-color: rgb(227,227,227);\n"
                                             "border-radius: 20px;\n"
                                             "border: 2px solid rgb(150,150,150);"],
                             [self.label_56, "background-color: rgb(227,227,227);\n"
                                             "border-radius: 20px;\n"
                                             "border: 2px solid rgb(150,150,150);"],
                             [self.lblContDataWeight, "background-color: rgb(235,235,235);\n"
                                                      "color: white;\n"
                                                      "border-radius: 5px;\n"
                                                      "border: 2px solid rgb(150,150,150);"],
                             [self.lcdMainWeight, "background-color: rgb(227,227,227);\n"
                                                  "color: rgb(0,0,0);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(135,135,135);"],
                             [self.label_53, "background-color: rgb(227,227,227);\n"
                                             "border-radius: 20px;\n"
                                             "border: 2px solid rgb(150,150,150);"],
                             [self.line_8, "background-color: rgb(235,235,235);\n"
                                           "color: white;"],
                             [self.label_55, "background-color: rgb(227,227,227);\n"
                                             "border-radius: 20px;\n"
                                             "border: 2px solid rgb(150,150,150);"],
                             [self.lblNameWeight_3, "background-color: rgb(235,235,235);\n"
                                                    "color: black;"],
                             [self.lblTbl_2, "background-color: rgb(242,242,242);\n"
                                             "border-radius: 5px;\n"
                                             "border: 2px solid rgb(205,205,205);"],
                             [self.btnMainSemaREDOUT, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                      "color: rgb(0,0,0);\n"
                                                      "border-radius:3px;\n"
                                                      "border:1px solid rgb(135,135,135);}\n"
                                                      "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                      "color: rgb(255, 255, 255);\n"
                                                      "border-radius:3px;\n"
                                                      "border:1px solid rgb(63,63,63);}\n"
                                                      "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                      "color: rgb(255, 255, 255);\n"
                                                      "border-radius:3px;\n"
                                                      "border:1px solid rgb(63,63,63);};"],
                             [self.btnMainSemaALLRED, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                      "color: rgb(0,0,0);\n"
                                                      "border-radius:3px;\n"
                                                      "border:1px solid rgb(135,135,135);}\n"
                                                      "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                      "color: rgb(255, 255, 255);\n"
                                                      "border-radius:3px;\n"
                                                      "border:1px solid rgb(63,63,63);}\n"
                                                      "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                      "color: rgb(255, 255, 255);\n"
                                                      "border-radius:3px;\n"
                                                      "border:1px solid rgb(63,63,63);};"],
                             [self.btnMainSemaREDIN, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                     "color: rgb(0,0,0);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(135,135,135);}\n"
                                                     "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);}\n"
                                                     "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);};"],
                             [self.lblContDataWeight_2, "background-color: rgb(235,235,235);\n"
                                                        "color: white;\n"
                                                        "border-radius: 5px;\n"
                                                        "border: 2px solid rgb(150,150,150);"],
                             [self.lblContDataWeight_3, "background-color: rgb(235,235,235);\n"
                                                        "color: white;\n"
                                                        "border-radius: 5px;\n"
                                                        "border: 2px solid rgb(150,150,150);"],
                             [self.btnMainShlgOpen, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                    "color: rgb(0,0,0);\n"
                                                    "border-radius:3px;\n"
                                                    "border:1px solid rgb(135,135,135);}\n"
                                                    "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:3px;\n"
                                                    "border:1px solid rgb(63,63,63);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:3px;\n"
                                                    "border:1px solid rgb(63,63,63);};"],
                             [self.btnMainShlgClose, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                     "color: rgb(0,0,0);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(135,135,135);}\n"
                                                     "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);}\n"
                                                     "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);};"],
                             [self.line_9, "background-color: rgb(235,235,235);\n"
                                           "color: black;"],
                             [self.lblNameWeight_4, "background-color: rgb(235,235,235);\n"
                                                    "color: black;"],
                             [self.label_45, "background-color: rgb(227,227,227);\n"
                                             "border-radius: 5px;\n"
                                             "border:1px solid rgb(135,135,135);"],
                             [self.label_20, "background-color: rgb(227,227,227);\n"
                                             "color: black;"],
                             [self.label_21, "background-color: rgb(227,227,227);\n"
                                             "color: black;"],
                             [self.lblMainShlgStateImage, "background-color: rgb(235,235,235);\n"
                                                          "image: url(" + globalValues.pathStyleImgs + "iconshlMIDDLE.png);"],
                             [self.lblNameWeight_5, "background-color: rgb(235,235,235);\n"
                                                    "color: black;"],
                             [self.line_10, "background-color: rgb(235,235,235);\n"
                                            "color: black;"],
                             [self.lblNameWeight_6, "background-color: rgb(235,235,235);\n"
                                                    "color: black;"],
                             [self.lblNameWeight_7, "background-color: rgb(235,235,235);\n"
                                                    "color: black;"],
                             [self.lblNameWeight_8, "background-color: rgb(235,235,235);\n"
                                                    "color: black;"],
                             [self.lblNameWeight_9, "background-color: rgb(235,235,235);\n"
                                                    "color: black;"],
                             [self.btnMainGRZCh1, "QPushButton:!hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum.png);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);}\n"
                                                  "QPushButton:hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum4.png);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);}\n"
                                                  "QPushButton:hover:pressed {image: url(" + globalValues.pathStyleImgs + "icondelcarnum2.png);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);};"],
                             [self.btnMainGRZCh2, "QPushButton:!hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum.png);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);}\n"
                                                  "QPushButton:hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum4.png);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);}\n"
                                                  "QPushButton:hover:pressed {image: url(" + globalValues.pathStyleImgs + "icondelcarnum2.png);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);};"],
                             [self.btnMainGRZCh3, "QPushButton:!hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum.png);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);}\n"
                                                  "QPushButton:hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum4.png);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);}\n"
                                                  "QPushButton:hover:pressed {image: url(" + globalValues.pathStyleImgs + "icondelcarnum2.png);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);};"],
                             [self.btnMainGRZCh4, "QPushButton:!hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum.png);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);}\n"
                                                  "QPushButton:hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum4.png);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);}\n"
                                                  "QPushButton:hover:pressed {image: url(" + globalValues.pathStyleImgs + "icondelcarnum2.png);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border-radius:3px;\n"
                                                  "border:1px solid rgb(63,63,63);};"],
                             [self.btnHide, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                            "border-top-right-radius: 5px;\n"
                                            "border-bottom-left-radius: 6px;\n"
                                            "border: 2px solid rgb(205,205,205);\n"
                                            "border-bottom: 1px solid rgb(150,150,150);\n"
                                            "border-left: 1px solid rgb(150,150,150);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconhide6.png);}\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "border-top-right-radius: 5px;\n"
                                            "border-bottom-left-radius: 6px;\n"
                                            "border: 2px solid rgb(205,205,205);\n"
                                            "border-bottom: 1px solid rgb(150,150,150);\n"
                                            "border-left: 1px solid rgb(150,150,150);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconhide7.png);}\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "border-top-right-radius: 5px;\n"
                                            "border-bottom-left-radius: 6px;\n"
                                            "border: 2px solid rgb(205,205,205);\n"
                                            "border-bottom: 1px solid rgb(150,150,150);\n"
                                            "border-left: 1px solid rgb(150,150,150);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconhide7.png);};"],
                             [self.leMainGRZGardInBack, "background-color: rgb(75,75,75);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                        "border-radius: 3px;\n"
                                                        "border: 1px solid rgb(0,0,0);"],
                             [self.leMainGRZGardOutBack, "background-color: rgb(75,75,75);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                         "border-radius: 3px;\n"
                                                         "border: 1px solid rgb(0,0,0);"],
                             [self.leMainGRZWeightInBack, "background-color: rgb(75,75,75);\n"
                                                          "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                          "border-radius: 3px;\n"
                                                          "border: 1px solid rgb(0,0,0);"],
                             [self.leMainGRZWeightOutBack, "background-color: rgb(75,75,75);\n"
                                                           "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                           "border-radius: 3px;\n"
                                                           "border: 1px solid rgb(0,0,0);"],
                             [self.leMainGRZGardIn8,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardIn5,
                              "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardIn2,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardIn1,
                              "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardIn9,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardIn3,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardIn7,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardIn4,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardIn6,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardOut3,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardOut1,
                              "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardOut8,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardOut4,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardOut5,
                              "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardOut2,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardOut6,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardOut7,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZGardOut9,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightIn9,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightIn1,
                              "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightIn2,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightIn5,
                              "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightIn3,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightIn4,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightIn7,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightIn6,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightIn8,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightOut1,
                              "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightOut6,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightOut4,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightOut9,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightOut2,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightOut3,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightOut5,
                              "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightOut8,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leMainGRZWeightOut7,
                              "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                              "border: none;\n"
                              "color: black;\n"
                              "font: 14pt \"MS Shell Dlg 2\";"],
                             [self.leCamAllMon, "background-color: rgb(235,235,235);\n"
                                                "border-radius: 10px;\n"
                                                "border: 4px solid rgb(150,150,150);"],
                             [self.btnCamCallapse, "QPushButton:!hover {border-radius: 5px;\n"
                                                   "border: 2px solid rgb(135,135,135);\n"
                                                   "background-color: rgb(227,227,227);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "iconmakesmall2020grey1.png);}\n"
                                                   "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 2px solid rgb(150,150,150);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "iconmakesmall2020.png);}\n"
                                                   "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 2px solid rgb(150,150,150);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "iconmakesmall2020.png);};"],
                             [self.le_ArchVideoBack, "background-color: rgb(235,235,235);\n"
                                                     "border-radius: 5px;\n"
                                                     "border: 2px solid rgb(150,150,150);\n"
                                                     "border-top-right-radius: 0px;\n"
                                                     "border-bottom-right-radius: 0px;\n"
                                                     "border-right: 1px solid rgb(150,150,150);"],
                             [self.frmArchMPlay, "background-color: rgb(227,227,227);\n"
                                                 "border-radius: 5px;\n"
                                                 "border: 2px solid rgb(150,150,150);\n"
                                                 "border-top: 1px solid rgb(150,150,150);\n"
                                                 "border-top-left-radius: 0px;\n"
                                                 "border-top-right-radius: 0px;\n"
                                                 "border-bottom-right-radius: 0px;\n"
                                                 "border-right: 1px solid rgb(150,150,150);"],
                             [self.SldrArchMplay, "QSlider {border-radius: none;\n"
                                                  "border: none;}\n"
                                                  "QSlider::groove:horizontal {\n"
                                                  "    background-color: rgb(252,252,252);\n"
                                                  "    border: 1px solid rgb(64,64,64);\n"
                                                  "    height: 14px;\n"
                                                  "    margin: 0px;\n"
                                                  "    }\n"
                                                  "QSlider::handle:horizontal {\n"
                                                  "    background-color: rgb(64,64,64);\n"
                                                  "    border: 1px solid rgb(64,64,64);\n"
                                                  "    height: 40px;\n"
                                                  "    width: 10px;\n"
                                                  "    margin: -15px 0px;}\n"
                                                  ""],
                             [self.btnStopArchMPlay, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:0px;\n"
                                                     "border:2px solid rgb(150,150,150);\n"
                                                     "border-left:1px solid rgb(150,150,150);\n"
                                                     "border-right:1px solid rgb(150,150,150);\n"
                                                     "border-bottom: 2px solid rgb(150,150,150);\n"
                                                     "border-bottom-left-radius: 5px;\n"
                                                     "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);}\n"
                                                     "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:0px;\n"
                                                     "border:2px solid rgb(150,150,150);\n"
                                                     "border-left:1px solid rgb(150,150,150);\n"
                                                     "border-right:1px solid rgb(150,150,150);\n"
                                                     "border-bottom: 2px solid rgb(150,150,150);\n"
                                                     "border-bottom-left-radius: 5px;\n"
                                                     "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);}\n"
                                                     "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:0px;\n"
                                                     "border:2px solid rgb(150,150,150);\n"
                                                     "border-left:1px solid rgb(150,150,150);\n"
                                                     "border-right:1px solid rgb(150,150,150);\n"
                                                     "border-bottom: 2px solid rgb(150,150,150);\n"
                                                     "border-bottom-left-radius: 5px;\n"
                                                     "image: url(" + globalValues.pathStyleImgs + "iconstop2grey.png);};"],
                             [self.btnPlayArchMPlay, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:0px;\n"
                                                     "border:2px solid rgb(150,150,150);\n"
                                                     "border-right: 1px solid rgb(150,150,150);\n"
                                                     "border-top: 2px solid rgb(150,150,150);\n"
                                                     "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                     "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:0px;\n"
                                                     "border:2px solid rgb(150,150,150);\n"
                                                     "border-right: 1px solid rgb(150,150,150);\n"
                                                     "border-top: 2px solid rgb(150,150,150);\n"
                                                     "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                     "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:0px;\n"
                                                     "border:2px solid rgb(150,150,150);\n"
                                                     "border-right: 1px solid rgb(150,150,150);\n"
                                                     "border-top: 2px solid rgb(150,150,150);\n"
                                                     "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);};"],
                             [self.btnStepforwardArchMPlay, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                            "color: rgb(255, 255, 255);\n"
                                                            "border-radius:0px;\n"
                                                            "border:2px solid rgb(150,150,150);\n"
                                                            "border-left:1px solid rgb(150,150,150);\n"
                                                            "border-right:1px solid rgb(150,150,150);\n"
                                                            "border-bottom: 2px solid rgb(150,150,150);\n"
                                                            "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}\n"
                                                            "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                            "color: rgb(255, 255, 255);\n"
                                                            "border-radius:0px;\n"
                                                            "border:2px solid rgb(150,150,150);\n"
                                                            "border-left:1px solid rgb(150,150,150);\n"
                                                            "border-right:1px solid rgb(150,150,150);\n"
                                                            "border-bottom: 2px solid rgb(150,150,150);\n"
                                                            "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}\n"
                                                            "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                            "color: rgb(255, 255, 255);\n"
                                                            "border-radius:0px;\n"
                                                            "border:2px solid rgb(150,150,150);\n"
                                                            "border-left:1px solid rgb(150,150,150);\n"
                                                            "border-right:1px solid rgb(150,150,150);\n"
                                                            "border-bottom: 2px solid rgb(150,150,150);\n"
                                                            "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5grey.png);}"],
                             [self.btnStepbackArchMPlay, "QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:0px;\n"
                                                         "border:2px solid rgb(150,150,150);\n"
                                                         "border-left:1px solid rgb(150,150,150);\n"
                                                         "border-right:1px solid rgb(150,150,150);\n"
                                                         "border-bottom: 2px solid rgb(150,150,150);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}\n"
                                                         "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:0px;\n"
                                                         "border:2px solid rgb(150,150,150);\n"
                                                         "border-left:1px solid rgb(150,150,150);\n"
                                                         "border-right:1px solid rgb(150,150,150);\n"
                                                         "border-bottom: 2px solid rgb(150,150,150);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}\n"
                                                         "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:0px;\n"
                                                         "border:2px solid rgb(150,150,150);\n"
                                                         "border-left:1px solid rgb(150,150,150);\n"
                                                         "border-right:1px solid rgb(150,150,150);\n"
                                                         "border-bottom: 2px solid rgb(150,150,150);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5grey.png);}"],
                             [self.cbArchMPlay, "QComboBox {\n"
                                                "background: white;\n"
                                                "border: 2px solid rgb(150,150,150);\n"
                                                "border-right: 1px solid rgb(150,150,150);\n"
                                                "border-bottom: 2px solid rgb(150,150,150);\n"
                                                "border-radius: 0px;\n"
                                                "}\n"
                                                "QComboBox:editable {\n"
                                                "background-color: rgb(142,187,208);\n"
                                                "color: black;\n"
                                                "border-radius: 0px;\n"
                                                "}"],
                             [self.le_8, "background-color: rgb(235,235,235);\n"
                                         "border-radius: 10px;\n"
                                         "font: 10pt \"MS Shell Dlg 2\";"],
                             [self.label_6, "background-color: rgb(235,235,235);\n"
                                            "color: black;\n"
                                            "font: 11pt \"MS Shell Dlg 2\";"],
                             [self.line, "background-color: rgb(242,242,242);"],
                             [self.label_7, "background-color: rgb(235,235,235);\n"
                                            "border-radius: 5px;\n"
                                            "border: 2px solid rgb(150,150,150);\n"
                                            "border-left-color: rgb(250,250,250);\n"
                                            "border-bottom-right-radius: 0px;\n"
                                            "border-top-left-radius: 0px;\n"
                                            "border-bottom-left-radius: 0px;\n"
                                            "border-bottom: 1px;"],
                             [self.vScrollArch, "QScrollBar:vertical {\n"
                                                "border: 1px solid rgb(89,89,89);\n"
                                                " background: rgb(255,255,255);\n"
                                                "width:10px;\n"
                                                "margin: 0px 0px 0px 0px;\n"
                                                "}\n"
                                                "QScrollBar::handle:vertical {\n"
                                                "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0 rgb(142,187,208), stop: 0.5 rgb(142,187,208), stop:1 rgb(142,187,208));\n"
                                                "min-height: 0px;\n"
                                                "}\n"
                                                "QScrollBar::add-line:vertical {\n"
                                                " background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop: 0 rgb(142,187,208), stop: 0.5 rgb(142,187,208),  stop:1 rgb(142,187,208));\n"
                                                "height: 0px;\n"
                                                "subcontrol-position: bottom;\n"
                                                "subcontrol-origin: margin;\n"
                                                "}\n"
                                                "QScrollBar::sub-line:vertical {\n"
                                                "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0  rgb(142,187,208), stop: 0.5 rgb(142,187,208),  stop:1 rgb(142,187,208));\n"
                                                "height: 0 px;\n"
                                                "subcontrol-position: top;\n"
                                                "subcontrol-origin: margin;\n"
                                                "}"],
                             [self.btnBackToMain, "QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                                  "border-radius: 0px;\n"
                                                  "border-bottom-right-radius: 3px;\n"
                                                  "color: rgb(0,0,0);\n"
                                                  "border-bottom: 2px solid rgb(150,150,150);\n"
                                                  "border-right: 2px solid rgb(150,150,150);\n"
                                                  "border-top: 1px solid rgb(150,150,150);\n"
                                                  "image: url(" + globalValues.pathStyleImgs + "iconright4grey.png);}\n"
                                                  "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                  "border-radius: 0px;\n"
                                                  "border-bottom-right-radius: 3px;\n"
                                                  "color: rgb(255,255,255);\n"
                                                  "border-bottom: 2px solid rgb(150,150,150);\n"
                                                  "border-right: 2px solid rgb(150,150,150);\n"
                                                  "border-top: 2px solid rgb(150,150,150);\n"
                                                  "image: url(" + globalValues.pathStyleImgs + "iconright4.png);}\n"
                                                  "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                  "border-radius: 0px;\n"
                                                  "border-bottom-right-radius: 3px;\n"
                                                  "color: rgb(255,255,255);\n"
                                                  "border-bottom: 2px solid rgb(150,150,150);\n"
                                                  "border-right: 2px solid rgb(150,150,150);\n"
                                                  "border-top: 2px solid rgb(150,150,150);\n"
                                                  "image: url(" + globalValues.pathStyleImgs + "iconright4.png);}"],
                             [self.le_ArchVideoStream, "background-color: rgb(235,235,235);\n"
                                                       "border: none;"],
                             [self.le_ArchFilename, "background-color: rgb(235,235,235);\n"
                                                    "border: none;\n"
                                                    "font: 10pt \"MS Shell Dlg 2\";"],
                             [self.lstArchive, "QTreeWidget {\n"
                                                "background-color: rgb(242,242,242);\n"
                                                "border-radius: 5px;\n"
                                                "border-top-left-radius : 0px;\n"
                                                "border-bottom-left-radius: 0px;\n"
                                                "border: 2px solid rgb(150,150,150);\n"
                                                "border-left-color: none;\n"
                                                "border-top-right-radius: 0px;\n"
                                                "border-bottom-right-radius: 0px;\n"
                                                "border-bottom: 0px solid rgb(150,150,150);\n"
                                                "border-top: 0px solid rgb(150,150,150);\n"
                                                "color: black;\n"
                                                "font: 11pt \"MS Shell Dlg 2\";\n"
                                                "padding-left: 7px;\n"
                                                "}\n"
                                                "QTreeWidget::item {\n"
                                                "margin: 2px;\n"
                                                "};"],
                             [self.label_8, "background-color: rgb(227,227,227);\n"
                                            "border:1px solid rgb(150,150,150);\n"
                                            "border-left: 0px solid rgb(150,150,150);\n"
                                            "border-bottom: 2px solid rgb(150,150,150);"],
                             [self.label_9, "background-color: rgb(255,255,255);\n"
                                            "border: 1px solid rgb(150,150,150);\n"
                                            "border-top: 1px solid rgb(150,150,150);\n"
                                            "border-bottom: 2px solid rgb(150,150,150);\n"
                                            "border-left: 0px solid rgb(150,150,150);\n"
                                            "border-right: 0px solid rgb(150,150,150);"],
                             [self.label_10, "background-color: rgb(255,255,255);\n"
                                             "border: 1px solid rgb(150,150,150);\n"
                                             "border-top: 1px solid rgb(150,150,150);\n"
                                             "border-bottom: 2px solid rgb(150,150,150);\n"
                                             "border-left: 0px solid rgb(150,150,150);\n"
                                             "border-right: 0px solid rgb(150,150,150);"],
                             [self.btnCamSet0, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "background-color: rgb(227,227,227);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconcamset4.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);};"],
                             [self.btnCamSet2, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "background-color: rgb(227,227,227);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconcamset4.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);};"],
                             [self.btnCamSet1, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "background-color: rgb(227,227,227);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconcamset4.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);};"],
                             [self.btnCamSet3, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "background-color: rgb(227,227,227);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconcamset4.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);};"],
                             [self.btnZoomIn0, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "background-color: rgb(227,227,227);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomin3.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);};"],
                             [self.btnZoomOut0, "QPushButton:!hover {border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "background-color: rgb(227,227,227);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconzoomout3.png);}\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "border-radius: 5px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "border-radius: 5px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);};"],
                             [self.btnZoomOut1, "QPushButton:!hover {border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "background-color: rgb(227,227,227);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconzoomout3.png);}\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "border-radius: 5px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "border-radius: 5px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);};"],
                             [self.btnZoomIn1, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "background-color: rgb(227,227,227);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomin3.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);};"],
                             [self.btnZoomOut2, "QPushButton:!hover {border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "background-color: rgb(227,227,227);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconzoomout3.png);}\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "border-radius: 5px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "border-radius: 5px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);};"],
                             [self.btnZoomIn2, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "background-color: rgb(227,227,227);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomin3.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);};"],
                             [self.btnZoomOut3, "QPushButton:!hover {border-radius: 3px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "background-color: rgb(227,227,227);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconzoomout3.png);}\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "border-radius: 5px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "border-radius: 5px;\n"
                                                "border: 1px solid rgb(135,135,135);\n"
                                                "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);};"],
                             [self.btnZoomIn3, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "background-color: rgb(227,227,227);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomin3.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);};"],
                             [self.btnRtsp0, "QPushButton:!hover {border-radius: 3px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "background-color: rgb(227,227,227);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrtsp4.png);}\n"
                                             "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                             "border-radius: 5px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);}\n"
                                             "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                             "border-radius: 5px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);};"],
                             [self.btnRtsp1, "QPushButton:!hover {border-radius: 3px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "background-color: rgb(227,227,227);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrtsp4.png);}\n"
                                             "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                             "border-radius: 5px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);}\n"
                                             "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                             "border-radius: 5px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);};"],
                             [self.btnRtsp2, "QPushButton:!hover {border-radius: 3px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "background-color: rgb(227,227,227);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrtsp4.png);}\n"
                                             "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                             "border-radius: 5px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);}\n"
                                             "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                             "border-radius: 5px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);};"],
                             [self.btnRtsp3, "QPushButton:!hover {border-radius: 3px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "background-color: rgb(227,227,227);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrtsp4.png);}\n"
                                             "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                             "border-radius: 5px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);}\n"
                                             "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                             "border-radius: 5px;\n"
                                             "border: 1px solid rgb(135,135,135);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);};"]]

            # iconzoomin3white.png

            self.lstDark = [[self.comboSearch, "background-color: rgb(89,89,89);\n"
                                               "color: white;\n"
                                               "border-radius: 3px;\n"
                                               "border: 1px solid rgb(63,63,63)"],
                            [self.tblMainData,
                             "QTableWidget {background-color: rgb(42,42,42);\n"
                                               "border: 1px solid rgb(63,63,63);\n"
                                               "border-radius:5px;\n"
                                               "border-bottom-right-radius: 0px;\n"
                                               "gridline-color: rgb(89,89,89);\n"
                                               "color:white;}\n"
                                               "QLineEdit {background-color: white;\n"
                             "                  color: rgb(0,0,0);}\n"
                                               "QHeaderView::section {\n"
                                               "gridline-color: rgb(89,89,89);\n"
                                               # "background-color: rgb(50,75,115);};\n"
                                               "color: black;};\n"
                             ],
                             # ""],
                            [self.mainFormPr, "background-color: rgb(66,66,66);"],
                            [self.frmCam, "background-color: rgb(66,66,66);"],
                            [self.lblTitleCams, "background-color: rgb(62,62,62);\n"
                                                "color: rgb(255,255,255);\n"
                                                "border-radius: 5px;"],
                            [self.lblCatCams, "background-color: rgb(62,62,62);\n"
                                              "color: rgb(255,255,255);"],
                            [self.lblIconCams,
                             "background-color:qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                             "image: url(" + globalValues.pathStyleImgs + "iconcam2.png);"],
                            [self.lblMainBoxCams, "background-color: rgb(62,62,62);\n"
                                                  "border-radius: 5px;"],
                            [self.lblCam1, "\n"
                                           "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 255), stop:1 rgba(62, 62, 62, 255));\n"
                                           "color: rgb(255,255,255);\n"
                                           "border-radius: 8px;\n"
                                           "border: 2px solid rgb(43,43,44);"],
                            [self.lblCam2, "\n"
                                           "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 255), stop:1 rgba(62, 62, 62, 255));\n"
                                           "color: rgb(255,255,255);\n"
                                           "border-radius: 8px;\n"
                                           "border: 2px solid rgb(43,43,44);"],
                            [self.lblCam4, "\n"
                                           "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 255), stop:1 rgba(62, 62, 62, 255));\n"
                                           "color: rgb(255,255,255);\n"
                                           "border-radius: 8px;\n"
                                           "border: 2px solid rgb(43,43,44);"],
                            [self.lblCam0, "\n"
                                           "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 255), stop:1 rgba(62, 62, 62, 255));\n"
                                           "color: rgb(255,255,255);\n"
                                           "border-radius: 8px;\n"
                                           "border: 2px solid rgb(43,43,44);"],
                            [self.leCam0, "background-color: rgb(0,0,0);\n"
                                          "border-radius: 5px;\n"
                                          "border: 2px solid rgb(42,42,42);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "camDefault400x400.png);"],
                            [self.leCam1, "background-color: rgb(0,0,0);\n"
                                          "border-radius: 5px;\n"
                                          "border: 2px solid rgb(42,42,42);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "camDefault400x400.png);"],
                            [self.leCam2, "background-color: rgb(0,0,0);\n"
                                          "border-radius: 5px;\n"
                                          "border: 2px solid rgb(42,42,42);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "camDefault400x400.png);"],
                            [self.leCam3, "background-color: rgb(0,0,0);\n"
                                          "border-radius: 5px;\n"
                                          "border: 2px solid rgb(42,42,42);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "camDefault400x400.png);"],
                            [self.btnCam0Exp, "QPushButton:!hover {border-radius: 5px;\n"
                                              "border: 2px solid rgb(41,41,41);\n"
                                              "background-color: rgb(75,75,75);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "border-radius: 5px;\n"
                                              "border: 2px solid rgb(41,41,41);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "border-radius: 5px;\n"
                                              "border: 2px solid rgb(41,41,41);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);};"],
                            [self.btnCam1Exp, "QPushButton:!hover {border-radius: 5px;\n"
                                              "border: 2px solid rgb(41,41,41);\n"
                                              "background-color: rgb(75,75,75);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "border-radius: 5px;\n"
                                              "border: 2px solid rgb(41,41,41);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "border-radius: 5px;\n"
                                              "border: 2px solid rgb(41,41,41);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);};"],
                            [self.btnCam2Exp, "QPushButton:!hover {border-radius: 5px;\n"
                                              "border: 2px solid rgb(41,41,41);\n"
                                              "background-color: rgb(75,75,75);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "border-radius: 5px;\n"
                                              "border: 2px solid rgb(41,41,41);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "border-radius: 5px;\n"
                                              "border: 2px solid rgb(41,41,41);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);};"],
                            [self.btnCam3Exp, "QPushButton:!hover {border-radius: 5px;\n"
                                              "border: 2px solid rgb(41,41,41);\n"
                                              "background-color: rgb(75,75,75);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "border-radius: 5px;\n"
                                              "border: 2px solid rgb(41,41,41);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "border-radius: 5px;\n"
                                              "border: 2px solid rgb(41,41,41);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconmakebigpng1717.png);};"],
                            [self.label_2, "background-color: rgb(62,62,62);"],
                            [self.label_3, "background-color: rgb(62,62,62);"],
                            [self.btnDayNightCam0, "QPushButton:!hover {border-radius: 3px;\n"
                                                   "border: 1px solid rgb(0,0,0);\n"
                                                   "background-color: rgb(75,75,75);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);}\n"
                                                   "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 1px solid rgb(135,135,135);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);}\n"
                                                   "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 1px solid rgb(135,135,135);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);};"],
                            [self.btnDayNightCam2, "QPushButton:!hover {border-radius: 3px;\n"
                                                   "border: 1px solid rgb(0,0,0);\n"
                                                   "background-color: rgb(75,75,75);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);}\n"
                                                   "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 1px solid rgb(135,135,135);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);}\n"
                                                   "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 1px solid rgb(135,135,135);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);};"],
                            [self.btnDayNightCam1, "QPushButton:!hover {border-radius: 3px;\n"
                                                   "border: 1px solid rgb(0,0,0);\n"
                                                   "background-color: rgb(75,75,75);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);}\n"
                                                   "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 1px solid rgb(135,135,135);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);}\n"
                                                   "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 1px solid rgb(135,135,135);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);};"],
                            [self.btnDayNightCam3, "QPushButton:!hover {border-radius: 3px;\n"
                                                   "border: 1px solid rgb(0,0,0);\n"
                                                   "background-color: rgb(75,75,75);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);}\n"
                                                   "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 1px solid rgb(135,135,135);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);}\n"
                                                   "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 1px solid rgb(135,135,135);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "icondaynight2white.png);};"],
                            [self.btnCam0Arch, "QPushButton:!hover {border-radius: 5px;\n"
                                               "border: 2px solid rgb(41,41,41);\n"
                                               "background-color: rgb(75,75,75);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(41,41,41);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(41,41,41);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);};"],
                            [self.btnCam1Arch, "QPushButton:!hover {border-radius: 5px;\n"
                                               "border: 2px solid rgb(41,41,41);\n"
                                               "background-color: rgb(75,75,75);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(41,41,41);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(41,41,41);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);};"],
                            [self.btnCam3Arch, "QPushButton:!hover {border-radius: 5px;\n"
                                               "border: 2px solid rgb(41,41,41);\n"
                                               "background-color: rgb(75,75,75);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(41,41,41);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(41,41,41);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);};"],
                            [self.btnCam2Arch, "QPushButton:!hover {border-radius: 5px;\n"
                                               "border: 2px solid rgb(41,41,41);\n"
                                               "background-color: rgb(75,75,75);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(41,41,41);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 2px solid rgb(41,41,41);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconarch6.png);};"],
                            [self.frmTable, "background-color: rgb(66,66,66);\n"
                                            "color: white;"],
                            [self.lblBoxJournal, "background-color: rgb(62,62,62);\n"
                                                 "color: rgb(255,255,255);\n"
                                                 "border-radius: 5px;"],
                            [self.lblNameJournal, "background-color: rgb(62,62,62);\n"
                                                  "color: rgb(255,255,255);"],
                            [self.lblIconJournal,
                             "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                             "image: url(" + globalValues.pathStyleImgs + "iconworkpanel1.png);"],
                            [self.lblTbl, "background-color: rgb(62,62,62);\n"
                                          "border-radius: 5px;"],
                            [self.leSearchTbl, "background-color: rgb(42, 42, 42);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border-radius: 3px;"],
                            [self.btnSearch, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconsearchwhite.png);}\n"
                                             "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconsearchwhite.png);}\n"
                                             "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);\n"
                                             "image: url(" + globalValues.pathStyleImgs + "iconsearchwhite.png);};"],
                            [self.btnDelZakazSergey, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);}\n"
                                             "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);}\n"
                                             "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border-radius:3px;\n"
                                             "border:1px solid rgb(63,63,63);}\n"],
                            [self.btnRefresh, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconrefreshwhite.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconrefreshwhite.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconrefreshwhite.png);};"],
                            [self.label_4, "background-color: rgb(62,62,62);"],
                            [self.label_5, "background-color: rgb(62,62,62);"],
                            [self.verticalScrollBar, "QScrollBar:vertical {\n"
                                                     "border: 1px solid rgb(63,63,63);\n"
                                                     " background: rgb(63,63,63);\n"
                                                     "width:10px;\n"
                                                     "margin: 0px 0px 0px 0px;\n"
                                                     "}\n"
                                                     "QScrollBar::add-page:vertical {background: rgb(89,89,89);}\n"
                                                     "QScrollBar::sub-page:vertical {background: rgb(89,89,89);}\n"
                                                     "QScrollBar::handle:vertical {\n"
                                                     "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0 rgb(50,75,115), stop: 0.5 rgb(50,75,115), stop:1 rgb(50,75,115));\n"
                                                     "min-height: 0px;\n"
                                                     "}\n"
                                                     "QScrollBar::add-line:vertical {\n"
                                                     " background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop: 0 rgb(50,75,115), stop: 0.5 rgb(50,75,115),  stop:1 rgb(50,75,115));\n"
                                                     "height: 0px;\n"
                                                     "subcontrol-position: bottom;\n"
                                                     "subcontrol-origin: margin;\n"
                                                     "}\n"
                                                     "QScrollBar::sub-line:vertical {\n"
                                                     "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0  rgb(50,75,115), stop: 0.5 rgb(50,75,115),  stop:1 rgb(50,75,115));\n"
                                                     "height: 0 px;\n"
                                                     "subcontrol-position: top;\n"
                                                     "subcontrol-origin: margin;\n"
                                                     "}"],
                            [self.DateEdit, "background-color: rgb(89,89,89);\n"
                                            "color: rgb(255,255,255);\n"
                                            "font: 11pt \"MS Shell Dlg 2\";\n"
                                            "border-radius: 3px;\n"
                                            "border: 1px solid rgb(63,63,63)"],
                            [self.frmMenu, "background-color: rgb(66,66,66);"],
                            [self.lblBoxSets, "background-color: rgb(75,75,75);\n"
                                              "border-radius: 5px;"],
                            [self.btnAboutSetSystem, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);}\n"
                                                     "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);}\n"
                                                     "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);};"],
                            [self.btnStWorkingSystem, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                      "color: rgb(255, 255, 255);\n"
                                                      "border:1px solid rgb(63,63,63);}\n"
                                                      "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                      "color: rgb(255, 255, 255);\n"
                                                      "border:1px solid rgb(63,63,63);}\n"
                                                      "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                      "color: rgb(255, 255, 255);\n"
                                                      "border:1px solid rgb(63,63,63);};"],
                            [self.lblIconMainSet,
                             "background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                             "image: url(" + globalValues.pathStyleImgs + "iconwork5.png);"],
                            [self.lblTitleMainSet, "background-color: rgb(75,76,75);\n"
                                                   "color: rgb(47,47,47);"],
                            [self.btnStatusDevs, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border:1px solid rgb(63,63,63);};"],
                            [self.btnScaner, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border:1px solid rgb(63,63,63);};"],
                            [self.lblIconSetSystem,
                             "background-color:qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                             "image: url(" + globalValues.pathStyleImgs + "iset1.png);\n"
                             "border-radius: 7px;"],
                            [self.btnJournal, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);};"],
                            [self.btnGard, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                           "color: rgb(255, 255, 255);\n"
                                           "border:1px solid rgb(63,63,63);}\n"
                                           "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                           "color: rgb(255, 255, 255);\n"
                                           "border:1px solid rgb(63,63,63);}\n"
                                           "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                           "color: rgb(255, 255, 255);\n"
                                           "border:1px solid rgb(63,63,63);};"],
                            [self.lineSetMain, "background-color: rgb(75,75,75);"],
                            [self.btnAboutMain, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(63,63,63);}\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(63,63,63);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border-radius:3px;\n"
                                                "border:1px solid rgb(63,63,63);};"],
                            [self.btnWeight, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border:1px solid rgb(63,63,63);}\n"
                                             "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border:1px solid rgb(63,63,63);}\n"
                                             "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                             "color: rgb(255, 255, 255);\n"
                                             "border:1px solid rgb(63,63,63);};"],
                            [self.btnDB, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                         "color: rgb(255, 255, 255);\n"
                                         "border:1px solid rgb(63,63,63);}\n"
                                         "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                         "color: rgb(255, 255, 255);\n"
                                         "border:1px solid rgb(63,63,63);}\n"
                                         "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                         "color: rgb(255, 255, 255);\n"
                                         "border:1px solid rgb(63,63,63);};"],
                            [self.lineSetsSystem, "background-color: rgb(75,75,75);"],
                            [self.btnSetCams, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "color: rgb(255, 255, 255);\n"
                                              "border-radius:3px;\n"
                                              "border:1px solid rgb(63,63,63);};"],
                            [self.lblNameSetSystem, "background-color: rgb(75,76,75);\n"
                                                    "color: rgb(47,47,47);"],
                            [self.btnDetectionCars, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border:1px solid rgb(63,63,63);}\n"
                                                    "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border:1px solid rgb(63,63,63);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border:1px solid rgb(63,63,63);};"],
                            [self.btnColorScheme, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border:1px solid rgb(63,63,63);}\n"
                                                  "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border:1px solid rgb(63,63,63);}\n"
                                                  "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                  "color: rgb(255, 255, 255);\n"
                                                  "border:1px solid rgb(63,63,63);};"],
                            [self.btnuserchange, "QPushButton:!hover {background-color: rgb(75,75,75);\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconuchange.png);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(75,75,75);}\n"
                                                 "QPushButton:hover {background-color: rgb(75,75,75);\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconuserch6.png);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(75,75,75);}\n"
                                                 "QPushButton:hover:pressed {background-color: rgb(75,75,75);\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconuserch2.png);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(75,75,75);};"],
                            [self.lineSetsSystem_2, "background-color: rgb(75,75,75);"],
                            [self.le_UserName, "background-color: rgb(75,75,75);\n"
                                               "color: rgb(255,255,255);\n"
                                               "font: 12pt \"Arial\";\n"
                                               "border:1px solid(75,75,75);"],
                            [self.btnJournalTS, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border:1px solid rgb(63,63,63);}\n"
                                                "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border:1px solid rgb(63,63,63);}\n"
                                                "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                "color: rgb(255, 255, 255);\n"
                                                "border:1px solid rgb(63,63,63);};"],
                            [self.btnBDTS, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                           "color: rgb(255, 255, 255);\n"
                                           "border:1px solid rgb(63,63,63);}\n"
                                           "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                           "color: rgb(255, 255, 255);\n"
                                           "border:1px solid rgb(63,63,63);}\n"
                                           "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                           "color: rgb(255, 255, 255);\n"
                                           "border:1px solid rgb(63,63,63);};"],
                            [self.btnAddZakaz, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border:1px solid rgb(63,63,63);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border:1px solid rgb(63,63,63);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "color: rgb(255, 255, 255);\n"
                                               "border:1px solid rgb(63,63,63);};"],
                            [self.frmHeadImage, "background-color: rgb(66,66,66);"],
                            [self.lblBoxUpCompany, "background-image: url(" + globalValues.pathStyleImgs + "iconhead.png);\n"
                                                   "border-radius: 5px;\n"
                                                   "border: 3px solid rgb(62,62,62);\n"],
                                                   # "image: url(" + globalValues.pathStyleImgs + "sinapsHEAD2.png);"],
                            [self.lblBarClose, "background-color: rgb(255,255,255);\n"
                                               "color: rgb(0,0,0);\n"
                                               ""],
                            [self.lblIconBarClose, "background-color: rgb(255,255,255);\n"
                                                   "image: url(" + globalValues.pathStyleImgs + "sinaps1717.png);"],
                            [self.btnCloseMainForm, "QPushButton:!hover{ background-color: rgb(255,255,255);\n"
                                                    "border-radius: 5px;\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconextbl1919.png);}\n"
                                                    "QPushButton:hover { background-color: rgb(84,122,181);\n"
                                                    "border-radius: 3px;\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconext1919.png);}\n"
                                                    "QPushButton:hover:pressed { background-color: rgb(50,75,115);\n"
                                                    "border-radius: 3px;\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconext1919.png);};\n"
                                                    ""],
                            [self.label, "background-color: rgb(255,255,255);"],
                            [self.leCamAllMon, "background-color: rgb(0,0,0);\n"
                                               "border-radius: 10px;\n"
                                               "border: 4px solid rgb(42,42,42);"],
                            [self.btnCamCallapse, "QPushButton:!hover {border-radius: 5px;\n"
                                                  "border: 2px solid rgb(41,41,41);\n"
                                                  "background-color: rgb(75,75,75);\n"
                                                  "image: url(" + globalValues.pathStyleImgs + "iconmakesmall2020.png);}\n"
                                                  "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                  "border-radius: 5px;\n"
                                                  "border: 2px solid rgb(41,41,41);\n"
                                                  "image: url(" + globalValues.pathStyleImgs + "iconmakesmall2020.png);}\n"
                                                  "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                  "border-radius: 5px;\n"
                                                  "border: 2px solid rgb(41,41,41);\n"
                                                  "image: url(" + globalValues.pathStyleImgs + "iconmakesmall2020.png);};"],
                            [self.frmCntrlPanel, "background-color: rgb(66,66,66);"],
                            [self.lblIconDataWeightTr, "background-color: rgb(75,75,75);\n"
                                                       "image: url(" + globalValues.pathStyleImgs + "iconweightnew2323.png);"],
                            [self.label_54, "background-color: rgb(89,89,89);\n"
                                            "border-radius: 20px;"],
                            [self.label_56, "background-color: rgb(89,89,89);\n"
                                            "border-radius: 20px;"],
                            [self.lblContDataWeight, "background-color: rgb(75,75,75);\n"
                                                     "border-radius: 5px;"],
                            [self.lcdMainWeight, "background-color: rgb(89,89,89);"],
                            [self.label_53, "background-color: rgb(89,89,89);\n"
                                            "border-radius: 20px;"],
                            [self.line_8, "background-color: rgb(75,75,75);"],
                            [self.label_55, "background-color: rgb(89,89,89);\n"
                                            "border-radius: 20px;"],
                            [self.lblNameWeight_3, "background-color: rgb(75,75,75);\n"
                                                   "color: white;"],
                            [self.lblTbl_2, "background-color: rgb(62,62,62);\n"
                                            "border-radius: 5px;"],
                            [self.btnMainSemaREDOUT, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);}\n"
                                                     "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);}\n"
                                                     "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);};"],
                            [self.btnMainSemaALLRED, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);}\n"
                                                     "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);}\n"
                                                     "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                     "color: rgb(255, 255, 255);\n"
                                                     "border-radius:3px;\n"
                                                     "border:1px solid rgb(63,63,63);};"],
                            [self.btnMainSemaREDIN, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:3px;\n"
                                                    "border:1px solid rgb(63,63,63);}\n"
                                                    "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:3px;\n"
                                                    "border:1px solid rgb(63,63,63);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:3px;\n"
                                                    "border:1px solid rgb(63,63,63);};"],
                            [self.lblContDataWeight_2, "background-color: rgb(75,75,75);\n"
                                                       "border-radius: 5px;"],
                            [self.lblContDataWeight_3, "background-color: rgb(75,75,75);\n"
                                                       "border-radius: 5px;"],
                            [self.btnMainShlgOpen, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                   "color: rgb(255, 255, 255);\n"
                                                   "border-radius:3px;\n"
                                                   "border:1px solid rgb(63,63,63);}\n"
                                                   "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                   "color: rgb(255, 255, 255);\n"
                                                   "border-radius:3px;\n"
                                                   "border:1px solid rgb(63,63,63);}\n"
                                                   "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                   "color: rgb(255, 255, 255);\n"
                                                   "border-radius:3px;\n"
                                                   "border:1px solid rgb(63,63,63);};"],
                            [self.btnMainShlgClose, "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:3px;\n"
                                                    "border:1px solid rgb(63,63,63);}\n"
                                                    "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:3px;\n"
                                                    "border:1px solid rgb(63,63,63);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:3px;\n"
                                                    "border:1px solid rgb(63,63,63);};"],
                            [self.line_9, "background-color: rgb(75,75,75);"],
                            [self.lblNameWeight_4, "background-color: rgb(75,75,75);\n"
                                                   "color: white;"],
                            [self.label_45, "background-color: rgb(89,89,89);\n"
                                            "border-radius: 5px;"],
                            [self.label_20, "background-color: rgb(89,89,89);\n"
                                            "color: white;"],
                            [self.label_21, "background-color: rgb(89,89,89);\n"
                                            "color: white;"],
                            [self.lblMainShlgStateImage,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "image: url(" + globalValues.pathStyleImgs + "iconshlMIDDLE.png);"],
                            [self.lblNameWeight_5, "background-color: rgb(75,75,75);\n"
                                                   "color: white;"],
                            [self.line_10, "background-color: rgb(75,75,75);"],
                            [self.lblNameWeight_6, "background-color: rgb(75,75,75);\n"
                                                   "color: white;"],
                            [self.lblNameWeight_7, "background-color: rgb(75,75,75);\n"
                                                   "color: white;"],
                            [self.lblNameWeight_8, "background-color: rgb(75,75,75);\n"
                                                   "color: white;"],
                            [self.lblNameWeight_9, "background-color: rgb(75,75,75);\n"
                                                   "color: white;"],
                            [self.btnMainGRZCh1, "QPushButton:!hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum.png);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum4.png);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover:pressed {image: url(" + globalValues.pathStyleImgs + "icondelcarnum2.png);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);};"],
                            [self.btnMainGRZCh2, "QPushButton:!hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum.png);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum4.png);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover:pressed {image: url(" + globalValues.pathStyleImgs + "icondelcarnum2.png);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);};"],
                            [self.btnMainGRZCh3, "QPushButton:!hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum.png);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum4.png);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover:pressed {image: url(" + globalValues.pathStyleImgs + "icondelcarnum2.png);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);};"],
                            [self.btnMainGRZCh4, "QPushButton:!hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum.png);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover {image: url(" + globalValues.pathStyleImgs + "icondelcarnum4.png);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);}\n"
                                                 "QPushButton:hover:pressed {image: url(" + globalValues.pathStyleImgs + "icondelcarnum2.png);\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-radius:3px;\n"
                                                 "border:1px solid rgb(63,63,63);};"],
                            [self.btnHide, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                           "border-top-right-radius: 5px;\n"
                                           "border-bottom-left-radius: 6px;\n"
                                           "border: 2px solid rgb(63,63,63);\n"
                                           "border-bottom: 1px solid rgb(75,75,75);\n"
                                           "border-left: 1px solid rgb(75,75,75);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "iconhide7.png);}\n"
                                           "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                           "border-top-right-radius: 5px;\n"
                                           "border-bottom-left-radius: 6px;\n"
                                           "border: 2px solid rgb(63,63,63);\n"
                                           "border-bottom: 1px solid rgb(62,62,62);\n"
                                           "border-left: 1px solid rgb(62,62,62);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "iconhide7.png);}\n"
                                           "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                           "border-top-right-radius: 5px;\n"
                                           "border-bottom-left-radius: 6px;\n"
                                           "border: 2px solid rgb(63,63,63);\n"
                                           "border-bottom: 1px solid rgb(62,62,62);\n"
                                           "border-left: 1px solid rgb(62,62,62);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "iconhide7.png);};"],
                            [self.leMainGRZGardOutBack, "background-color: rgb(75,75,75);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                        "border-radius: 3px;\n"
                                                        "border: 1px solid rgb(0,0,0);"],
                            [self.leMainGRZWeightOutBack, "background-color: rgb(75,75,75);\n"
                                                          "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                          "border-radius: 3px;\n"
                                                          "border: 1px solid rgb(0,0,0);"],
                            [self.leMainGRZWeightInBack, "background-color: rgb(75,75,75);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                         "border-radius: 3px;\n"
                                                         "border: 1px solid rgb(0,0,0);"],
                            [self.leMainGRZGardInBack, "background-color: rgb(75,75,75);\n"
                                                       "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                       "border-radius: 3px;\n"
                                                       "border: 1px solid rgb(0,0,0);"],
                            [self.leMainGRZGardIn1,
                             "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardIn2,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardIn3,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardIn4,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardIn6,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardIn5,
                             "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardIn7,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardIn8,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardIn9,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardOut8,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardOut5,
                             "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardOut2,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardOut1,
                             "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardOut9,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardOut3,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardOut7,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardOut4,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZGardOut6,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightIn8,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightIn5,
                             "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightIn2,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightIn1,
                             "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightIn9,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightIn3,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightIn7,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightIn4,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightIn6,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightOut8,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightOut5,
                             "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightOut2,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightOut1,
                             "background-color:qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightOut9,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightOut3,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightOut7,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightOut4,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.leMainGRZWeightOut6,
                             "background-color: qlineargradient(spread:pad, x1:0, y1:1, x2:0, y2:0, stop:0 rgba(255, 255, 255, 0), stop:1 rgba(255, 255, 255, 0));\n"
                             "border: none;\n"
                             "color: black;\n"
                             "font: 14pt \"MS Shell Dlg 2\";"],
                            [self.le_ArchVideoBack, "background-color: rgb(0,0,0);\n"
                                                    "border-radius: 5px;\n"
                                                    "border: 2px solid rgb(42,42,42);\n"
                                                    "border-top-right-radius: 0px;\n"
                                                    "border-bottom-right-radius: 0px;\n"
                                                    "border-right-color: rgb(0,0,0);\n"
                                                    "border-right: 1px solid rgb(42,42,42)"],
                            [self.btnBackToMain, "QPushButton:!hover {background-color: rgb(66,66,66);\n"
                                                 "border-radius: 0px;\n"
                                                 "border-bottom-right-radius: 3px;\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-bottom: 2px solid rgb(42,42,42);\n"
                                                 "border-right: 2px solid rgb(42,42,42);\n"
                                                 "border-top: 2px solid rgb(42,42,42);\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconright4.png);}\n"
                                                 "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                                 "border-radius: 0px;\n"
                                                 "border-bottom-right-radius: 3px;\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-bottom: 2px solid rgb(42,42,42);\n"
                                                 "border-right: 2px solid rgb(42,42,42);\n"
                                                 "border-top: 2px solid rgb(42,42,42);\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconright4.png);}\n"
                                                 "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                                 "border-radius: 0px;\n"
                                                 "border-bottom-right-radius: 3px;\n"
                                                 "color: rgb(255, 255, 255);\n"
                                                 "border-bottom: 2px solid rgb(42,42,42);\n"
                                                 "border-right: 2px solid rgb(42,42,42);\n"
                                                 "border-top: 2px solid rgb(42,42,42);\n"
                                                 "image: url(" + globalValues.pathStyleImgs + "iconright4.png);}"],
                            [self.le_8, "background-color: rgb(0,0,0);\n"
                                        "border-radius: 10px;\n"
                                        "font: 10pt \"MS Shell Dlg 2\";\n"
                                        "color: rgb(255,255,255);"],
                            [self.frmArchMPlay, "background-color: rgb(89,89,89);\n"
                                                "border-radius: 5px;\n"
                                                "border: 2px solid rgb(42,42,42);\n"
                                                "border-top: 1px solid rgb(42,42,42);\n"
                                                "border-top-left-radius: 0px;\n"
                                                "border-top-right-radius: 0px;\n"
                                                "border-bottom-right-radius: 0px;\n"
                                                "border-right: 1px solid rgb(42,42,42);"],
                            [self.SldrArchMplay, "QSlider {border-radius: none;\n"
                                                 "border: none;}\n"
                                                 "QSlider::groove:horizontal {\n"
                                                 "    background-color: rgb(252,252,252);\n"
                                                 "    border: 1px solid rgb(64,64,64);\n"
                                                 "    height: 14px;\n"
                                                 "    margin: 0px;\n"
                                                 "    }\n"
                                                 "QSlider::handle:horizontal {\n"
                                                 "    background-color: rgb(64,64,64);\n"
                                                 "    border: 1px solid rgb(64,64,64);\n"
                                                 "    height: 40px;\n"
                                                 "    width: 10px;\n"
                                                 "    margin: -15px 0px;}\n"
                                                 ""],
                            [self.btnStopArchMPlay, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "border-bottom-left-radius: 5px;\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "border-bottom-left-radius: 5px;\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "border-bottom-left-radius: 5px;\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconstop2.png);};"],
                            [self.btnPlayArchMPlay, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "border-top: 2px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "border-top: 2px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "border-top: 2px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);};"],
                            [self.btnStepforwardArchMPlay, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                           "color: rgb(255, 255, 255);\n"
                                                           "border-radius:0px;\n"
                                                           "border:2px solid rgb(42,42,42);\n"
                                                           "border-left:1px solid rgb(42,42,42);\n"
                                                           "border-right:1px solid rgb(42,42,42);\n"
                                                           "border-top: 2px solid rgb(42,42,42);\n"
                                                           "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}\n"
                                                           "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                           "color: rgb(255, 255, 255);\n"
                                                           "border-radius:0px;\n"
                                                           "border:2px solid rgb(42,42,42);\n"
                                                           "border-left:1px solid rgb(42,42,42);\n"
                                                           "border-right:1px solid rgb(42,42,42);\n"
                                                           "border-top: 2px solid rgb(42,42,42);\n"
                                                           "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}\n"
                                                           "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                           "color: rgb(255, 255, 255);\n"
                                                           "border-radius:0px;\n"
                                                           "border:2px solid rgb(42,42,42);\n"
                                                           "border-left:1px solid rgb(42,42,42);\n"
                                                           "border-right:1px solid rgb(42,42,42);\n"
                                                           "border-top: 2px solid rgb(42,42,42);\n"
                                                           "image: url(" + globalValues.pathStyleImgs + "iconplaystepforward5.png);}"],
                            [self.btnStepbackArchMPlay, "QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplaystepback5.png);}"],
                            [self.cbArchMPlay, "QComboBox {\n"
                                               "background: white;\n"
                                               "border: 2px solid rgb(42,42,42);\n"
                                               "border-right: 1px solid rgb(42,42,42);\n"
                                               "border-bottom: 2px solid rgb(42,42,42);\n"
                                               "border-radius: 0px;\n"
                                               "}\n"
                                               "QComboBox:editable {\n"
                                               "background-color: rgb(50,75,115);\n"
                                               "color: black;\n"
                                               "border-radius: 0px;\n"
                                               "}"],
                            [self.label_6, "background-color: rgb(42,42,42);\n"
                                           "color: white;\n"
                                           "font: 11pt \"MS Shell Dlg 2\";"],
                            [self.label_7, "background-color: rgb(42,42,42);\n"
                                           "border-radius: 5px;\n"
                                           "border: 2px solid rgb(42,42,42);\n"
                                           "border-left-color: rgb(75,75,75);\n"
                                           "border-bottom-right-radius: 0px;\n"
                                           "border-top-left-radius: 0px;\n"
                                           "border-bottom-left-radius: 0px;\n"
                                           "border-bottom: 1px;"],
                            [self.vScrollArch, "QScrollBar:vertical {\n"
                                               "border: 1px solid rgb(63,63,63);\n"
                                               " background: rgb(63,63,63);\n"
                                               "width:10px;\n"
                                               "margin: 0px 0px 0px 0px;\n"
                                               "}\n"
                                               "QScrollBar::add-page:vertical {background: rgb(89,89,89);}\n"
                                               "QScrollBar::sub-page:vertical {background: rgb(89,89,89);}\n"
                                               "QScrollBar::handle:vertical {\n"
                                               "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0 rgb(50,75,115), stop: 0.5 rgb(50,75,115), stop:1 rgb(50,75,115));\n"
                                               "min-height: 0px;\n"
                                               "}\n"
                                               "QScrollBar::add-line:vertical {\n"
                                               " background: qlineargradient(x1:0, y1:0, x2:1, y2:0, stop: 0 rgb(50,75,115), stop: 0.5 rgb(50,75,115),  stop:1 rgb(50,75,115));\n"
                                               "height: 0px;\n"
                                               "subcontrol-position: bottom;\n"
                                               "subcontrol-origin: margin;\n"
                                               "}\n"
                                               "QScrollBar::sub-line:vertical {\n"
                                               "background: qlineargradient(x1:0, y1:0, x2:1, y2:0,stop: 0  rgb(50,75,115), stop: 0.5 rgb(50,75,115),  stop:1 rgb(50,75,115));\n"
                                               "height: 0 px;\n"
                                               "subcontrol-position: top;\n"
                                               "subcontrol-origin: margin;\n"
                                               "}"],
                            [self.line, "background-color: rgb(42,42,42);"],
                            [self.le_ArchVideoStream, "background-color: rgb(0,0,0);\n"
                                                      "border: none;"],
                            [self.le_ArchFilename, "background-color: rgb(0,0,0);\n"
                                                   "border: none;\n"
                                                   "font: 10pt \"MS Shell Dlg 2\";"],
                            [self.lstArchive, "QTreeWidget {\n"
                                               "background-color: rgb(42,42,42);\n"
                                               "border-radius: 5px;\n"
                                               "border-top-left-radius : 0px;\n"
                                               "border-bottom-left-radius: 0px;\n"
                                               "border: 2px solid rgb(42,42,42);\n"
                                               "border-left: 0px solid rgb(42,42,42);\n"
                                               "border-top-right-radius: 0px;\n"
                                               "border-bottom-right-radius: 0px;\n"
                                               "border-bottom: 4px solid rgb(42,42,42);\n"
                                               "color: white;\n"
                                               "font: 11pt \"MS Shell Dlg 2\";\n"
                                               "padding-left: 7px;\n"
                                               "}\n"
                                               "QTreeWidget::item {\n"
                                               "margin: 2px;\n"
                                               "};"],
                            [self.label_10, "background-color: rgb(85,85,85);\n"
                                            "border: 1px solid rgb(42,42,42);\n"
                                            "border-top: 1px solid rgb(42,42,42);\n"
                                            "border-bottom: 2px solid rgb(42,42,42);\n"
                                            "border-left: 0px solid rgb(42,42,42);\n"
                                            "border-right: 0px solid rgb(42,42,42);"],
                            [self.label_9, "background-color: rgb(120,120,120);\n"
                                           "border: 1px solid rgb(42,42,42);\n"
                                           "border-top: 1px solid rgb(42,42,42);\n"
                                           "border-bottom: 2px solid rgb(42,42,42);\n"
                                           "border-left: 0px solid rgb(42,42,42);\n"
                                           "border-right: 0px solid rgb(42,42,42);"],
                            [self.label_8, "background-color: rgb(89,89,89);\n"
                                           "border:1px solid rgb(42,42,42);\n"
                                           "border-left: 0px solid rgb(42,42,42);\n"
                                           "border-bottom: 2px solid rgb(42,42,42);"],
                            [self.btnCamSet0, "QPushButton:!hover {border-radius: 3px;\n"
                                              "border: 1px solid rgb(0,0,0);\n"
                                              "background-color: rgb(75,75,75);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);};"],
                            [self.btnCamSet2, "QPushButton:!hover {border-radius: 3px;\n"
                                              "border: 1px solid rgb(0,0,0);\n"
                                              "background-color: rgb(75,75,75);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);};"],
                            [self.btnCamSet1, "QPushButton:!hover {border-radius: 3px;\n"
                                              "border: 1px solid rgb(0,0,0);\n"
                                              "background-color: rgb(75,75,75);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);};"],
                            [self.btnCamSet3, "QPushButton:!hover {border-radius: 3px;\n"
                                              "border: 1px solid rgb(0,0,0);\n"
                                              "background-color: rgb(75,75,75);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconcamset4white.png);};"],
                            [self.btnZoomOut0, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(0,0,0);\n"
                                               "background-color: rgb(75,75,75);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);};"],
                            [self.btnZoomIn0, "QPushButton:!hover {border-radius: 3px;\n"
                                              "border: 1px solid rgb(0,0,0);\n"
                                              "background-color: rgb(75,75,75);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);};"],
                            [self.btnZoomOut1, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(0,0,0);\n"
                                               "background-color: rgb(75,75,75);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);};"],
                            [self.btnZoomIn1, "QPushButton:!hover {border-radius: 3px;\n"
                                              "border: 1px solid rgb(0,0,0);\n"
                                              "background-color: rgb(75,75,75);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);};"],
                            [self.btnZoomOut3, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(0,0,0);\n"
                                               "background-color: rgb(75,75,75);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);};"],
                            [self.btnZoomIn3, "QPushButton:!hover {border-radius: 3px;\n"
                                              "border: 1px solid rgb(0,0,0);\n"
                                              "background-color: rgb(75,75,75);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);};"],
                            [self.btnZoomOut2, "QPushButton:!hover {border-radius: 3px;\n"
                                               "border: 1px solid rgb(0,0,0);\n"
                                               "background-color: rgb(75,75,75);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);}\n"
                                               "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);}\n"
                                               "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                               "border-radius: 5px;\n"
                                               "border: 1px solid rgb(135,135,135);\n"
                                               "image: url(" + globalValues.pathStyleImgs + "iconzoomout3white.png);};"],
                            [self.btnZoomIn2, "QPushButton:!hover {border-radius: 3px;\n"
                                              "border: 1px solid rgb(0,0,0);\n"
                                              "background-color: rgb(75,75,75);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);}\n"
                                              "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);}\n"
                                              "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                              "border-radius: 5px;\n"
                                              "border: 1px solid rgb(135,135,135);\n"
                                              "image: url(" + globalValues.pathStyleImgs + "iconzoomin3white.png);};"],
                            [self.btnRtsp0, "QPushButton:!hover {border-radius: 3px;\n"
                                            "border: 1px solid rgb(0,0,0);\n"
                                            "background-color: rgb(75,75,75);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);}\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "border-radius: 5px;\n"
                                            "border: 1px solid rgb(135,135,135);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);}\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "border-radius: 5px;\n"
                                            "border: 1px solid rgb(135,135,135);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);};"],
                            [self.btnRtsp1, "QPushButton:!hover {border-radius: 3px;\n"
                                            "border: 1px solid rgb(0,0,0);\n"
                                            "background-color: rgb(75,75,75);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);}\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "border-radius: 5px;\n"
                                            "border: 1px solid rgb(135,135,135);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);}\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "border-radius: 5px;\n"
                                            "border: 1px solid rgb(135,135,135);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);};"],
                            [self.btnRtsp2, "QPushButton:!hover {border-radius: 3px;\n"
                                            "border: 1px solid rgb(0,0,0);\n"
                                            "background-color: rgb(75,75,75);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);}\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "border-radius: 5px;\n"
                                            "border: 1px solid rgb(135,135,135);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);}\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "border-radius: 5px;\n"
                                            "border: 1px solid rgb(135,135,135);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);};"],
                            [self.btnRtsp3, "QPushButton:!hover {border-radius: 3px;\n"
                                            "border: 1px solid rgb(0,0,0);\n"
                                            "background-color: rgb(75,75,75);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);}\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "border-radius: 5px;\n"
                                            "border: 1px solid rgb(135,135,135);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);}\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "border-radius: 5px;\n"
                                            "border: 1px solid rgb(135,135,135);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconrtsp4white.png);};"]]

            self.lengthLight = len(self.lstLight)
            self.lengthDark = len(self.lstDark)

            self.changeCOLORMainPanelGrunt()

            self.DateEdit.hide()

            font = QFont()
            font.setPointSize(14)
            font.setFamily('Arial')

            self.btnMainSemaREDIN.hide()
            self.btnMainSemaALLRED.hide()
            self.btnMainSemaREDOUT.hide()
            # self.leMainGRZWeightIn.setFont(font)
            # self.leMainGRZWeightOut.setFont(font)
            # self.leMainGRZGardIn.setFont(font)
            # self.leMainGRZGardOut.setFont(font)

            # self.leMainGRZWeightIn.setText('  Р 000 МЕ   777')
            # self.leMainGRZWeightOut.setText('  Р 000 МЕ    77')
            # self.leMainGRZGardIn.setText('  Р 000 МЕ   999')
            # self.leMainGRZGardOut.setText('  Р 000 МЕ   797')

            dateToday = datetime.date.today().strftime('%d.%m.%Y')
            str_date_today = str(dateToday)
            date = QDateTime.fromString(str_date_today, 'd.M.yyyy')
            self.DateEdit.setDateTime(date)

            myQHeaderView = self.tblMainData.horizontalHeader()
            myQHeaderView.setSectionResizeMode(0, QtWidgets.QHeaderView.Fixed)
            myQHeaderView.setSectionResizeMode(1, QtWidgets.QHeaderView.Fixed)
            myQHeaderView.setSectionResizeMode(2, QtWidgets.QHeaderView.Fixed)
            myQHeaderView.setSectionResizeMode(3, QtWidgets.QHeaderView.Fixed)
            myQHeaderView.setSectionResizeMode(4, QtWidgets.QHeaderView.Fixed)
            myQHeaderView.setSectionResizeMode(5, QtWidgets.QHeaderView.Fixed)
            myQHeaderView.setSectionResizeMode(6, QtWidgets.QHeaderView.Fixed)
            myQHeaderView.setSectionResizeMode(7, QtWidgets.QHeaderView.Fixed)
            myQHeaderView.setSectionResizeMode(8, QtWidgets.QHeaderView.Fixed)
            myQHeaderView.setStretchLastSection(False)
            self.tblMainData.setHorizontalHeader(myQHeaderView)

            # self.tblMainData.ho.setResizeMode(QHeaderView.Fixed)
            # self.autoStMainWorking()
            self.frmCamAllMon.hide()
            self.tblMainData.setContextMenuPolicy Qt.CustomContextMenu)
            self.tblMainData.viewport().installEventFilter(self)
            self.tblMainData.customContextMenuRequested.connect(self.generateMenu)
            self.lstArchive.viewport().installEventFilter(self)
            # self.tblMainData.setEditTriggers(QtWidgets.QAbstractItemView.EditingState)
            # self.installEventFilter(self)

            self.verticalScrollBar.valueChanged.connect(self.sync_func)
            self.vScrollArch.valueChanged.connect(self.sync_func_strg)
            self.btnCloseMainForm.clicked.connect(self.exitSystem)
            self.btnSetCams.clicked.connect(self.openPanelDevCams)
            # self.btnZakaz.clicked.connect(self.openPanelZakaz)
            self.btnWeight.clicked.connect(self.openPanelDevWeight)
            self.btnGard.clicked.connect(self.openPanelDevGard)
            self.btnDB.clicked.connect(self.openPanelDevDB)
            self.btnStatusDevs.clicked.connect(self.openPanelDevState)
            self.btnScaner.clicked.connect(self.openPanelDevScaner)
            self.btnDetectionCars.clicked.connect(self.openPanelMapUrl)
            # self.tblMainData.doubleClicked.connect(self.openWebBrowser)
            self.tblMainData.doubleClicked.connect(self.openPanelReadyOrLinkDet)
            self.btnSearch.clicked.connect(self.searchInTbl)
            self.btnRefresh.clicked.connect(self.refreshTblMain)
            self.btnJournal.clicked.connect(self.openPanelJournal)
            self.btnBDTS.clicked.connect(self.openPanelBaseTS)

            # self.btnCam0Exp.clicked.connect(self.openWindowCam0)
            self.btnCam0Exp.clicked.connect(self.openFullCam1)
            self.btnCam1Exp.clicked.connect(self.openFullCam2)
            self.btnCam2Exp.clicked.connect(self.openFullCam3)
            self.btnCam3Exp.clicked.connect(self.openFullCam4)
            self.btnCamCallapse.clicked.connect(self.closeWinFullCam)
            self.btnAboutMain.clicked.connect(self.openPanelAboutMain)
            self.btnAboutSetSystem.clicked.connect(self.openPanelAboutSysSets)
            self.btnStWorkingSystem.clicked.connect(self.openPanelStWorking)
            self.btnuserchange.clicked.connect(self.changeLoginAcc)

            self.btnMainSemaREDOUT.clicked.connect(self.trOpenDriveIn)
            self.btnMainSemaALLRED.clicked.connect(self.trCancelDrive)
            self.btnMainSemaREDIN.clicked.connect(self.trOpenDriveOut)

            self.btnMainGRZCh1.clicked.connect(self.clearGRZGardIn)
            self.btnMainGRZCh2.clicked.connect(self.clearGRZGardOut)
            self.btnMainGRZCh3.clicked.connect(self.clearGRZcarWeightIn)
            self.btnMainGRZCh4.clicked.connect(self.clearGRZcarWeightOut)

            self.btnColorScheme.clicked.connect(self.changeColorScheme)
            self.comboSearch.currentIndexChanged.connect(self.changeStCbSearch)
            self.btnJournalTS.clicked.connect(self.openPanelJournalTS)

            self.btnRtsp0.clicked.connect(self.openPanRtsp)
            self.btnRtsp1.clicked.connect(self.openPanRtsp)
            self.btnRtsp2.clicked.connect(self.openPanRtsp)
            self.btnRtsp3.clicked.connect(self.openPanRtsp)

            # self.lstArchive.itemClicked.connect(self.chgScrollStrg)

            self.tblMainData.setColumnWidth(0, 70)
            self.tblMainData.setColumnWidth(1, 80)
            self.tblMainData.setColumnWidth(2, 90)
            self.tblMainData.setColumnWidth(3, 80)
            self.tblMainData.setColumnWidth(4, 90)
            self.tblMainData.setColumnWidth(5, 90)
            self.tblMainData.setColumnWidth(6, 90)
            self.tblMainData.setColumnWidth(7, 90)
            self.tblMainData.setColumnWidth(8, 138)
            self.tblMainData.setColumnWidth(9, 0)
            self.tblMainData.setColumnWidth(10, 0)
            self.tblMainData.setColumnWidth(11, 0)
            self.tblMainData.setColumnWidth(12, 0)
            self.tblMainData.setFont(QFont("Arial", 9, QFont.Bold))
            self.tblMainData.verticalHeader().hide()
            self.tblMainData.horizontalScrollBar().hide()

            self.leSearchTbl.setPlaceholderText('Поиск')

            # startThCams(1, rtsp_cam_out_gard)
            # startThCams(2, rtsp_cam_in_gard)
            # startThCams(3, rtsp_cam_out_weight)
            # startThCams(4, rtsp_cam_in_weight)

            # self.leCam0.setPixmap(pixCamNotCon)
            # self.leCam1.setPixmap(pixCamNotCon)
            # self.leCam2.setPixmap(pixCamNotCon)
            # self.leCam3.setPixmap(pixCamNotCon)



            # globalValues.my_sql_db = 'cars'
            # globalValues.my_sql_localhost = 'localhost'
            # globalValues.my_sql_name = 'sergey'
            # globalValues.my_sql_password = '34ubitav'
            # globalValues.my_sql_port = 3306

            # connectDBMySQL()

            # self.startThreadsMain()

            # globalValues.my_pg_db = self.leDBNamePg.text()
            # globalValues.my_pg_localhost = self.leDBIpPg.text()
            # globalValues.my_pg_name = self.leDBLoginPg.text()
            # globalValues.my_pg_password = self.leDBPassPg.text()
            # globalValues.my_pg_port = int(self.leDBPortPg.text())
            #
            # self.leDBIpPg.setText('localhost')
            # self.leDBLoginPg.setText('postgres')
            # self.leDBPassPg.setText('sinaps281082')
            # self.leDBNamePg.setText('trassir3')
            # self.leDBPortPg.setText('5432')
            #
            # self.leDBIpMy.setText('localhost')
            # self.leDBLoginMy.setText('sergey')
            # self.leDBPassMy.setText('34ubitav')
            # self.leDBNameMy.setText('cars')
            # self.leDBPortMy.setText('3306')



            # self.tblMainData.set

            # self.startThreadsMain()
            self.frmCntrlPanel.show()

            self.btnHide.clicked.connect(self.hideHandStWork)
            self.btnAddZakaz.clicked.connect(self.openPanelAddZN)

            self.autoStMainWorking()

            self.checkFolderLongPath(globalValues.pathReports)

            self.checkFolderLongPath(globalValues.pathScanImgs)

            self.btnCam0Arch.clicked.connect(partial(self.calluser, '0'))
            self.btnCam1Arch.clicked.connect(partial(self.calluser, '1'))
            self.btnCam2Arch.clicked.connect(partial(self.calluser, '2'))
            self.btnCam3Arch.clicked.connect(partial(self.calluser, '3'))


            self.btnCamSet0.clicked.connect(partial(self.changeSetsCamNightDay, '0'))
            self.btnCamSet1.clicked.connect(partial(self.changeSetsCamNightDay, '1'))
            self.btnCamSet2.clicked.connect(partial(self.changeSetsCamNightDay, '2'))
            self.btnCamSet3.clicked.connect(partial(self.changeSetsCamNightDay, '3'))

            self.btnBackToMain.clicked.connect(self.hideFrmStorage)
            self.btnDelZakazSergey.clicked.connect(self.startThDelSergey)
            self.frmArchive.hide()

            self.lstArchive.verticalScrollBar().hide()

            pathNameObj = globalValues.pathDefaultData + '/Sinaps/name.txt'
            if (os.path.exists(pathNameObj)):
                f = open(pathNameObj, 'r')
                data = f.read()
                dataLst = data.split(' ')
                globalValues.nameObject = dataLst[0] + ' ' + dataLst[1]
                globalValues.namePolygon = dataLst[2] + ' ' + dataLst[3]
                f.close()

            # print('NameObj: ' + globalValues.nameObject)

            f = open(pathFileColor, 'w')
            f.write(str(globalValues.colorForm))
            f.close()

            if globalValues.debug:
                checkAndChangeStr('B126PE69', 303, [self.leMainGRZGardIn1, self.leMainGRZGardIn2, self.leMainGRZGardIn3, self.leMainGRZGardIn4, self.leMainGRZGardIn5, self.leMainGRZGardIn6, self.leMainGRZGardIn7, self.leMainGRZGardIn8, self.leMainGRZGardIn9])
                checkAndChangeStr('TK36777', 383, [self.leMainGRZWeightIn1, self.leMainGRZWeightIn2, self.leMainGRZWeightIn3, self.leMainGRZWeightIn4, self.leMainGRZWeightIn5, self.leMainGRZWeightIn6, self.leMainGRZWeightIn7, self.leMainGRZWeightIn8, self.leMainGRZWeightIn9])
                checkAndChangeStr('K789PL99', 343, [self.leMainGRZGardOut1, self.leMainGRZGardOut2, self.leMainGRZGardOut3, self.leMainGRZGardOut4, self.leMainGRZGardOut5, self.leMainGRZGardOut6, self.leMainGRZGardOut7, self.leMainGRZGardOut8, self.leMainGRZGardOut9])
                checkAndChangeStr('L666KU777', 423, [self.leMainGRZWeightOut1, self.leMainGRZWeightOut2, self.leMainGRZWeightOut3, self.leMainGRZWeightOut4, self.leMainGRZWeightOut5, self.leMainGRZWeightOut6, self.leMainGRZWeightOut7, self.leMainGRZWeightOut8, self.leMainGRZWeightOut9])

            self.btnZoomIn0.clicked.connect(self.createEventToWritingVideoChl_0)
            self.btnZoomIn1.clicked.connect(self.createEventToWritingVideoChl_1)
            self.btnZoomIn2.clicked.connect(self.createEventToWritingVideoChl_2)
            self.btnZoomIn3.clicked.connect(self.createEventToWritingVideoChl_3)

            if (globalValues.debugClose):
                self.btnZoomOut0.clicked.connect(self.debugExitSystem)

            pathFileObj = globalValues.pathDefFldr + '/dataChangeObject.txt'
            if (os.path.exists(pathFileObj) == False):
                f_obj = open(pathFileObj, 'w')
                data_to_wrt = 'o 0'
                f_obj.write(data_to_wrt)
                f_obj.close()

            try:
                pathFileRtsp = globalValues.pathDefaultData + '/Sinaps/dataRtsp.txt'
                print(pathFileRtsp)
                if (os.path.exists(pathFileRtsp)):
                    f = open(pathFileRtsp, 'r')
                    dataRead = f.read()
                    print(dataRead)
                    f.close()
                    if (dataRead != ''):
                        checkFile = True
                        lstData = dataRead.split('\n')
                        for i in range(len(lstData)):
                            globalValues.rtspMainLink[i] = lstData[i]
                else:
                    for i in range(6):
                        globalValues.rtspMainLink[i] = 'rtsp://login:pwd@strIP:554/cam/realmonitor?channel=1&subtype=1'

                print('MainLink: ')
                # print(globalValues.rtspMainLink)

            except Exception as ex:
                globalValues.writeLogData('Чтение данных из файла rtsp главный класс', str(ex))

    def screenImgDisplay(self):

        pathImg = globalValues.curDisk + '/Sinaps/screenShot.png'

        try:
            if(os.path.exists(pathImg)):
                os.remove(pathImg)
            # print('Start')
            # global app
            QScreen.grabWindow(app.primaryScreen(), QApplication.desktop().winId()).save(pathImg, 'png')
            print('End')
            # screen = pyautogui.screenshot(pathImg)
            # subprocess.call(['attrib', '+H', pathImg])
        except Exception as ex:
            print(ex)

            globalValues.writeLogData('Функция создания скриншота рабочего стола', str(ex))

    def measuringVolume(self):
        try:
            globalValues.is_read_depth = True
        except Exception as ex:
            globalValues.writeLogData('Измерить объём', str(ex))

    def showFrmStorage(self, numChl):
        try:
            self.frmCntrlPanel.hide()
            self.frmCam.hide()
            self.frmCamAllMon.hide()
            self.frmMenu.hide()
            self.frmTable.hide()
            self.frmArchive.show()
            name_channel = ''
            if (int(numChl) == 0):
                name_channel = 'КппВъезд'
                pathFldr = globalValues.pathStrg + '/' + name_channel
                self.createListVideos(pathFldr)
            elif (int(numChl) == 1):
                name_channel = 'КппВыезд'
                pathFldr = globalValues.pathStrg + '/' + name_channel
                self.createListVideos(pathFldr)
            elif (int(numChl) == 2):
                name_channel = 'ВесыВъезд'
                pathFldr = globalValues.pathStrg + '/' + name_channel
                self.createListVideos(pathFldr)
            elif (int(numChl) == 3):
                name_channel = 'ВесыВыезд'
                pathFldr = globalValues.pathStrg + '/' + name_channel
                self.createListVideos(pathFldr)

            globalValues.curNameChl = name_channel


            # strName = globalValues.pathFileVideo
            # listName = strName.split('/')
            #
            # self.le_8.setText('Название файла: ' + listName[len(listName) - 1])

            print(globalValues.pathStrg)

        except Exception as ex:
            globalValues.writeLogData('Функция запуска просмотра архива', str(ex))

    def calluser(self, name):
        try:
            self.mediaPlayerVideoStrg.setMedia(QMediaContent(QUrl.fromLocalFile('')))
            self.le_8.setText('Файл не выбран')
            print('checkingIn!!!')
            self.showFrmStorage(name)
            th_strg_scrol = threading.Thread(target=thChangeScrollStrg, args=(True, ))
            th_strg_scrol.start()
        except Exception as ex:
            globalValues.writeLogData('Функция открытия фрэйма архива', str(ex))

    def hideFrmStorage(self):
        try:
            self.frmCntrlPanel.show()
            self.frmCam.show()
            self.frmCamAllMon.hide()
            self.frmMenu.show()
            self.frmTable.show()
            self.frmArchive.hide()
            self.lstArchive.clearSelection()
        except Exception as ex:
            globalValues.writeLogData('Функция запуска просмотра архива', str(ex))

    def setMediaInFrmClk(self):
        try:
            nameCurFile = self.lstArchive.currentItem().text(0)

            isCheck = '.mp4' in nameCurFile

            if (isCheck):
                namePar = self.lstArchive.currentItem().parent().text(0)

                pathFileCur = globalValues.pathStrg + '/' + globalValues.curNameChl + '/' + namePar + '/' + nameCurFile

                globalValues.pathMainVideoStrg = pathFileCur

                self.le_8.setText('Название файла: ' + nameCurFile)

                self.setMediaFrm(pathFileCur, self.mediaPlayerVideoStrg)
            if (globalValues.colorForm == 1):
                self.btnPlayArchMPlay.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-right: 1px solid rgb(150,150,150);\n"
                                                    "border-top: 2px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-right: 1px solid rgb(150,150,150);\n"
                                                    "border-top: 2px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(150,150,150);\n"
                                                    "border-right: 1px solid rgb(150,150,150);\n"
                                                    "border-top: 2px solid rgb(150,150,150);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);};")
            else:
                self.btnPlayArchMPlay.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "border-top: 2px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "border-top: 2px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                    "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                    "color: rgb(255, 255, 255);\n"
                                                    "border-radius:0px;\n"
                                                    "border:2px solid rgb(42,42,42);\n"
                                                    "border-left:1px solid rgb(42,42,42);\n"
                                                    "border-right:1px solid rgb(42,42,42);\n"
                                                    "border-top: 2px solid rgb(42,42,42);\n"
                                                    "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);};")
        except Exception as ex:
            globalValues.writeLogData('Функция смены видеофайла в окне', str(ex))

    def setMediaFrm(self, pathFile, obMedia):
        if (pathFile != ''):
            obMedia.setMedia(QMediaContent(QUrl.fromLocalFile(pathFile)))
            # self.btnPlayKPPOUT.setEnabled(True)
            print('checkingpath: ' + pathFile)

    def hideHandStWork(self):
        try:
            if (self.checkHideHandSt):
                self.frmCntrlPanel.setGeometry(854, 573, 866, 501)
                self.frmTable.setGeometry QRect(854, 132, 866, 441))
                self.lblTbl.setGeometry QRect(6, 30, 854, 411))
                self.tblMainData.setGeometry QRect(23, 80, 820, 350))
                self.lblContDataWeight.show()
                self.lblNameWeight_3.show()
                self.checkHideHandSt = False
                self.verticalScrollBar.setGeometry QRect(841, 102, 10, 327))

                if (globalValues.colorForm == 1):
                    self.btnHide.setStyleSheet("QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                            "border-top-right-radius: 10px;\n"
                                            "border-bottom-left-radius: 6px;\n"
                                            "border: 2px solid rgb(205,205,205);\n"
                                            "border-bottom: 1px solid rgb(150,150,150);\n"
                                            "border-left: 1px solid rgb(150,150,150);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconhide6.png);}\n"
                                            "\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "border-top-right-radius: 10px;\n"
                                            "border-bottom-left-radius: 6px;\n"
                                            "border: 2px solid rgb(205,205,205);\n"
                                            "border-bottom: 1px solid rgb(150,150,150);\n"
                                            "border-left: 1px solid rgb(150,150,150);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconhide7.png);}\n"
                                            "\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "border-top-right-radius: 10px;\n"
                                            "border-bottom-left-radius: 6px;\n"
                                            "border: 2px solid rgb(205,205,205);\n"
                                            "border-bottom: 1px solid rgb(150,150,150);\n"
                                            "border-left: 1px solid rgb(150,150,150);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconhide7.png);};")
                else:
                    self.btnHide.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                           "border-top-right-radius: 10px;\n"
                                           "border-bottom-left-radius: 6px;\n"
                                           "border: 2px solid rgb(63,63,63);\n"
                                           "border-bottom: 1px solid rgb(75,75,75);\n"
                                           "border-left: 1px solid rgb(75,75,75);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "iconhide7.png);}\n"
                                           "\n"
                                           "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                           "border-top-right-radius: 10px;\n"
                                           "border-bottom-left-radius: 6px;\n"
                                           "border: 2px solid rgb(63,63,63);\n"
                                           "border-bottom: 1px solid rgb(0,0,0);\n"
                                           "border-left: 1px solid rgb(0,0,0);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "iconhide7.png);}\n"
                                           "\n"
                                           "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                           "border-top-right-radius: 10px;\n"
                                           "border-bottom-left-radius: 6px;\n"
                                           "border: 2px solid rgb(63,63,63);\n"
                                           "border-bottom: 1px solid rgb(0,0,0);\n"
                                           "border-left: 1px solid rgb(0,0,0);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "iconhide7.png);};")

                th = threading.Thread(target=thChangeScroll, args=(True, ))
                th.start()

            else:
                self.frmCntrlPanel.setGeometry(854, 1044, 866, 30)
                self.frmTable.setGeometry QRect(854, 132, 866, 911))
                self.lblTbl.setGeometry QRect(6, 30, 854, 881))
                self.tblMainData.setGeometry QRect(23, 80, 820, 820))
                self.lblContDataWeight.hide()
                self.lblNameWeight_3.hide()
                self.verticalScrollBar.setGeometry QRect(841, 102, 10, 797))
                if (globalValues.colorForm == 1):
                    self.btnHide.setStyleSheet("QPushButton:!hover {background-color: rgb(227,227,227);\n"
                                            "border-top-right-radius: 10px;\n"
                                            "border-bottom-left-radius: 6px;\n"
                                            "border: 2px solid rgb(205,205,205);\n"
                                            "border-bottom: 1px solid rgb(150,150,150);\n"
                                            "border-left: 1px solid rgb(150,150,150);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconopen6.png);}\n"
                                            "\n"
                                            "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            "border-top-right-radius: 10px;\n"
                                            "border-bottom-left-radius: 6px;\n"
                                            "border: 2px solid rgb(205,205,205);\n"
                                            "border-bottom: 1px solid rgb(150,150,150);\n"
                                            "border-left: 1px solid rgb(150,150,150);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconopen7.png);}\n"
                                            "\n"
                                            "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            "border-top-right-radius: 10px;\n"
                                            "border-bottom-left-radius: 6px;\n"
                                            "border: 2px solid rgb(205,205,205);\n"
                                            "border-bottom: 1px solid rgb(150,150,150);\n"
                                            "border-left: 1px solid rgb(150,150,150);\n"
                                            "image: url(" + globalValues.pathStyleImgs + "iconopen7.png);};")
                else:
                    self.btnHide.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                           "border-top-right-radius: 10px;\n"
                                           "border-bottom-left-radius: 6px;\n"
                                           "border: 2px solid rgb(63,63,63);\n"
                                           "border-bottom: 1px solid rgb(75,75,75);\n"
                                           "border-left: 1px solid rgb(75,75,75);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "iconopen7.png);}\n"
                                           "\n"
                                           "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                           "border-top-right-radius: 10px;\n"
                                           "border-bottom-left-radius: 6px;\n"
                                           "border: 2px solid rgb(63,63,63);\n"
                                           "border-bottom: 1px solid rgb(0,0,0);\n"
                                           "border-left: 1px solid rgb(0,0,0);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "iconopen7.png);}\n"
                                           "\n"
                                           "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                           "border-top-right-radius: 10px;\n"
                                           "border-bottom-left-radius: 6px;\n"
                                           "border: 2px solid rgb(63,63,63);\n"
                                           "border-bottom: 1px solid rgb(0,0,0);\n"
                                           "border-left: 1px solid rgb(0,0,0);\n"
                                           "image: url(" + globalValues.pathStyleImgs + "iconopen7.png);};")
                self.checkHideHandSt = True
                th = threading.Thread(target=thChangeScroll, args=(True, ))
                th.start()
        except Exception as ex:
            globalValues.writeLogData('Функция скрытия панели ручного режима', str(ex))

    def changeStCbSearch(self):
        try:
            dataCb = self.comboSearch.currentIndex()

            if (dataCb == 3):
                self.DateEdit.show()
                self.leSearchTbl.hide()
            else:
                self.DateEdit.hide()
                self.leSearchTbl.setGeometry(22, 40, 596, 30)
                self.leSearchTbl.show()

        except Exception as ex:
            globalValues.writeLogData('Функция смены фильра поиска панели журнала', str(ex))

    def changeColorScheme(self):
        try:
            self.changeCOLORMainPanelGrunt()
            self.uiDev.changeCOLORMainPanelGrunt(0.4)
            self.refreshOneTblMain = True

            globalValues.writeEventToDBJournalMain('Цветовая гамма', 'Выполнена смена цветовой гаммы программы')

            pathFileColor = globalValues.pathDefaultData + '/Sinaps/dataColor.txt'
            f = open(pathFileColor, 'w')
            f.write(str(globalValues.colorForm))
            f.close()

        except Exception as ex:
            globalValues.writeLogData('Функция смены цветовой гаммы ПО', str(ex))

    def changeColor(self, object, str):
            object.setStyleSheet(str)

    def changeCOLORMainPanelGrunt(self):

            delta = int((4/self.lengthDark)*1000)

            if (globalValues.colorForm == 0):

                    dataCur = globalValues.checkCamStrg

                    for i in range(4):
                        if (dataCur[i] == False):
                            if (i == 0):
                                self.leCam0Rec.setStyleSheet(
                                    "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "iconrec13grey.png);\n"
                                    "border-radius: 8px;")
                            elif (i == 1):
                                self.leCam1Rec.setStyleSheet(
                                    "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "iconrec13grey.png);\n"
                                    "border-radius: 8px;")
                            elif (i == 2):
                                self.leCam2Rec.setStyleSheet(
                                    "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "iconrec13grey.png);\n"
                                    "border-radius: 8px;")
                            elif (i == 3):
                                self.leCam3Rec.setStyleSheet(
                                    "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "iconrec13grey.png);\n"
                                    "border-radius: 8px;")

                    for i in range(self.lengthLight):
                        startTime = round(time.time() * 1000)
                        while True:
                                    obj = self.lstLight[i][0]
                                    style = self.lstLight[i][1]
                                    self.changeColor(obj, style)
                                    # print(style)
                                    if (abs(round(time.time()*1000) - startTime) > delta):
                                            break

                    globalValues.colorForm = 1

            elif (globalValues.colorForm == 1):

                    dataCur = globalValues.checkCamStrg

                    for i in range(4):
                        if (dataCur[i] == False):
                            if (i == 0):
                                self.leCam0Rec.setStyleSheet(
                                    "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "iconrec13lightgrey.png);\n"
                                    "border-radius: 8px;")
                            elif (i == 1):
                                self.leCam1Rec.setStyleSheet(
                                    "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "iconrec13lightgrey.png);\n"
                                    "border-radius: 8px;")
                            elif (i == 2):
                                self.leCam2Rec.setStyleSheet(
                                    "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "iconrec13lightgrey.png);\n"
                                    "border-radius: 8px;")
                            elif (i == 3):
                                self.leCam3Rec.setStyleSheet(
                                    "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "iconrec13lightgrey.png);\n"
                                    "border-radius: 8px;")

                    for i in range(self.lengthDark):
                        startTime = round(time.time() * 1000)
                        while True:
                                    obj = self.lstDark[i][0]
                                    style = self.lstDark[i][1]
                                    self.changeColor(obj, style)
                                    # print(style)
                                    if (abs(round(time.time()*1000) - startTime) > delta):
                                            break

                    globalValues.colorForm = 0

    def clearGRZcarWeightIn(self):
        self.leMainGRZWeightIn1.setText('')
        self.leMainGRZWeightIn2.setText('')
        self.leMainGRZWeightIn3.setText('')
        self.leMainGRZWeightIn4.setText('')
        self.leMainGRZWeightIn5.setText('')
        self.leMainGRZWeightIn6.setText('')
        self.leMainGRZWeightIn7.setText('')
        self.leMainGRZWeightIn8.setText('')
        self.leMainGRZWeightIn9.setText('')

    def clearGRZcarWeightOut(self):
        self.leMainGRZWeightOut1.setText('')
        self.leMainGRZWeightOut2.setText('')
        self.leMainGRZWeightOut3.setText('')
        self.leMainGRZWeightOut4.setText('')
        self.leMainGRZWeightOut5.setText('')
        self.leMainGRZWeightOut6.setText('')
        self.leMainGRZWeightOut7.setText('')
        self.leMainGRZWeightOut8.setText('')
        self.leMainGRZWeightOut9.setText('')

    def clearGRZGardIn(self):
        self.leMainGRZGardIn1.setText('')
        self.leMainGRZGardIn2.setText('')
        self.leMainGRZGardIn3.setText('')
        self.leMainGRZGardIn4.setText('')
        self.leMainGRZGardIn5.setText('')
        self.leMainGRZGardIn6.setText('')
        self.leMainGRZGardIn7.setText('')
        self.leMainGRZGardIn8.setText('')
        self.leMainGRZGardIn9.setText('')

    def clearGRZGardOut(self):
        self.leMainGRZGardOut1.setText('')
        self.leMainGRZGardOut2.setText('')
        self.leMainGRZGardOut3.setText('')
        self.leMainGRZGardOut4.setText('')
        self.leMainGRZGardOut5.setText('')
        self.leMainGRZGardOut6.setText('')
        self.leMainGRZGardOut7.setText('')
        self.leMainGRZGardOut8.setText('')
        self.leMainGRZGardOut9.setText('')

    def startThHandWorking(self):
            th_working_hand = threading.Thread(target=self.thStHandWorking)
            th_working_hand.start()

    def thStHandWorking(self):

        start_time = round(time.time())

        value_old_0 = 0
        value_old_1 = 0
        value_old_2 = 0
        value_old_3 = 0
        try:
            dataOld = 0
            num_weight = 0
            num_traf = 0
            while True:
                if (globalValues.stopAll):
                        break

                if (globalValues.stopThStHandWorking):
                        break

                if ( abs(round(time.time() - start_time)) > 1):
                        start_time = round(time.time())

                        if (globalValues.outComWeight == False):
                            num_weight = 0
                            valueCurWeight = globalValues.value
                            if (valueCurWeight == 0 and valueCurWeight != dataOld):
                                self.frmCntrlPanel.updateGeometry()
                            dataOld = valueCurWeight
                            self.lcdMainWeight.display(str(valueCurWeight))
                        else:
                            if num_weight == 0:
                                self.lcdMainWeight.display(str(0))
                                self.frmCntrlPanel.updateGeometry()
                                num_weight += 1


                        if (globalValues.outComTraf == False):

                            num_traf = 0

                            if (globalValues.check_sema_1_out != value_old_0):
                                if (globalValues.check_sema_1_out == 1):
                                    self.lblMainSemaLeftOutUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgRed + ');};')
                                    self.lblMainSemaLeftOutDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGray + ');};')
                                if (globalValues.check_sema_1_out == 2):
                                    self.lblMainSemaLeftOutDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGreen + ');};')
                                    self.lblMainSemaLeftOutUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGray + ');};')

                            if (globalValues.check_sema_1_in != value_old_1):
                                if (globalValues.check_sema_1_in == 1):
                                    self.lblMainSemaLeftInUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgRed + ');};')
                                    self.lblMainSemaLeftInDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGray + ');};')
                                if (globalValues.check_sema_1_in == 2):
                                    self.lblMainSemaLeftInDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGreen + ');};')
                                    self.lblMainSemaLeftInUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGray + ');};')

                            if (globalValues.check_sema_2_in != value_old_2):
                                if (globalValues.check_sema_2_in == 1):
                                    self.lblMainSemaRightInUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgRed + ');};')
                                    self.lblMainSemaRightInDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGray + ');};')
                                if (globalValues.check_sema_2_in == 2):
                                    self.lblMainSemaRightInDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGreen + ');};')
                                    self.lblMainSemaRightInUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGray + ');};')

                            if (globalValues.check_sema_2_out != value_old_3):
                                if (globalValues.check_sema_2_out == 1):
                                    self.lblMainSemaRightOutUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgRed + ');};')
                                    self.lblMainSemaRightOutDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGray + ');};')
                                if (globalValues.check_sema_2_out == 2):
                                    self.lblMainSemaRightOutDown.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGreen + ');};')
                                    self.lblMainSemaRightOutUp.setStyleSheet('QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n'
                            'image: url(' + pathImgGray + ');};')

                            value_old_0 = globalValues.check_sema_1_out
                            value_old_1 = globalValues.check_sema_1_in
                            value_old_2 = globalValues.check_sema_2_in
                            value_old_3 = globalValues.check_sema_2_out
                        else:
                            if num_traf == 0:
                                self.lblMainSemaLeftOutUp.setStyleSheet(
                                    "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
                                self.lblMainSemaLeftOutDown.setStyleSheet(
                                    "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
                                self.lblMainSemaLeftInUp.setStyleSheet(
                                    "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
                                self.lblMainSemaLeftInDown.setStyleSheet(
                                    "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
                                self.lblMainSemaRightInUp.setStyleSheet(
                                    "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
                                self.lblMainSemaRightInDown.setStyleSheet(
                                    "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
                                self.lblMainSemaRightOutUp.setStyleSheet(
                                    "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
                                self.lblMainSemaRightOutDown.setStyleSheet(
                                    "QLabel:!hover {background-color: qlineargradient(spread:pad, x1:1, y1:0, x2:1, y2:1, stop:0 rgba(89, 89, 89, 0), stop:1 rgba(62, 62, 62, 0));\n"
                                    "image: url(" + globalValues.pathStyleImgs + "icongreykrug3131.png);};")
                                num_traf += 1
                sleep(0.2)

                # self.updateGeometry()

            print('FinishingThreadHandSt!!!')
        except Exception as ex:
            globalValues.writeLogData('Поток ручного режима работы системы', str(ex))

    def autoStMainWorking(self):
        try:
            globalValues.startHandSt = False
            globalValues.stopThStHandWorking = False
            self.startThHandWorking()
            # globalValues.stopThStHandWorking = True
            self.currentModeStWorking = 0
            self.btnMainShlgOpen.hide()
            self.btnMainShlgClose.hide()
            self.btnMainGRZCh3.setEnabled(False)
            self.btnMainGRZCh4.setEnabled(False)
            self.btnMainGRZCh1.setEnabled(False)
            self.btnMainGRZCh2.setEnabled(False)
            self.btnMainSemaREDIN.hide()
            self.btnMainSemaREDOUT.hide()
            self.btnMainSemaALLRED.hide()
        except Exception as ex:
            globalValues.writeLogData('Функция запуска автоматического режима работы ПО', str(ex))

    def handStMainWorking(self):

        try:
            globalValues.startHandSt = True
            # globalValues.stopThStHandWorking = False
            # self.startThHandWorking()
            # sleep(0.3)
            print('StartingHand!')
            self.currentModeStWorking = 1
            self.btnMainShlgOpen.show()
            self.btnMainShlgClose.show()
            self.btnMainGRZCh3.setEnabled(True)
            self.btnMainGRZCh4.setEnabled(True)
            self.btnMainGRZCh1.setEnabled(True)
            self.btnMainGRZCh2.setEnabled(True)
            if (self.check_good_con_traffic):
                self.btnMainSemaREDIN.show()
                self.btnMainSemaREDOUT.show()
                self.btnMainSemaALLRED.show()
            self.checkHideHandSt = True
            self.hideHandStWork()

        except Exception as ex:
            globalValues.writeLogData('Функция запуска ручного режима работы ПО', str(ex))

        # self.frmCntrlPanel.show()
        # self.frmTable.setFixedSize(824, 440)
        # self.lblTbl.setFixedSize(810, 410)
        # self.tblMainData.setFixedSize(780, 350)

    def trOpenDriveIn(self):
        globalValues.trafDefault = True

    def trCancelDrive(self):
        globalValues.trafMeasure = True

    def trOpenDriveOut(self):
        globalValues.trafWeightReady = True

    def eventFilter(self, source, event):
        try:

            # print('checkingLord')

            if (event.type() == QEvent.MouseButtonPress and
                    event.buttons() == Qt.RightButton and
                    source is self.tblMainData.viewport()):

                    if (globalValues.curUserName != 'operator'):

                        item = self.tblMainData.itemAt(event.pos())
                        x0 = event.globalPos().x()
                        y0 = event.globalPos().y()

                        print('CheckingInTbl!!!' + str(x0) + '   ' + str(y0))

                        if item is not None:
                                if (self.is_open_pan_context):
                                        self.is_open_pan_context = False
                                        self.is_set_context_show = True
                                        globalValues.checkEditTbl = True

                                        print('Table Item:', item.row(), item.column())

                                        self.item_cur_row = item.row()
                                        self.item_cur_col = item.column()

                                        uiContext = Ui_ContextTable()
                                        self.screenImgDisplay()
                                        numCurCols = int(self.item_cur_col)
                                        if ( numCurCols == 0 or numCurCols == 2):
                                            # uiContext.setFixedSize(160, 90)
                                            uiContext.frmConBtn.setGeometry(x0, y0, 160, 90)
                                        else:
                                            # uiContext.setFixedSize(160, 120)
                                            uiContext.frmConBtn.setGeometry(x0, y0, 160, 120)
                                            # uiContext.btnConDel.setStyleSheet(
                                            #         "QPushButton:!hover {background-color: rgb(89,89,89);\n"
                                            #         "color: rgb(255, 255, 255);\n"
                                            #         "border:1px solid rgb(63,63,63);\n"
                                            #         "border-left-color: black;\n"
                                            #         "border-right-color: black}\n"
                                            #         "\n"
                                            #         "QPushButton:hover {background-color: rgb(84,122,181);\n"
                                            #         "color: rgb(255, 255, 255);\n"
                                            #         "border:1px solid rgb(63,63,63);\n"
                                            #         "border-left-color: black;\n"
                                            #         "border-right-color: black}\n"
                                            #         "\n"
                                            #         "QPushButton:hover:pressed {background-color: rgb(50,75,115);\n"
                                            #         "color: rgb(255, 255, 255);\n"
                                            #         "border:1px solid rgb(63,63,63);\n"
                                            #         "border-left-color: black;\n"
                                            #         "border-right-color: black};")

                                        # uiContext.startClickPython()
                                        uiContext.exec_()

                                        self.is_open_pan_context = True
                                        self.defCon = uiContext.defCon
                                        print('QQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQ' + str(self.defCon))
                                        self.conEventJob()
                                        print('checkingEndContextMenu')


            if (event.type() == QEvent.MouseButtonPress and
                    event.buttons() == Qt.LeftButton and self.is_set_context_show):
                    self.is_set_context_show = False

            if (event.type() == QEvent.MouseButtonPress and
                    event.buttons() == Qt.LeftButton and source is self.lstArchive.viewport()):
                print('qwerty345')
                th_chg_scrl_strg = threading.Thread(target=thChangeScrollStrg, args=(True, ))
                th_chg_scrl_strg.start()

            if (event.type() == QEvent.Wheel and source is self.tblMainData.viewport()):
                # event = QEvent.Wheel
                # QW
                # print(event.delta())
                print('checking!!!')
                delta = event.angleDelta().y()
                v = (-5, 5) [delta > 0]
                self.pos += v
                print(self.pos)
                self.verticalScrollBar.setValue(self.tblMainData.verticalScrollBar().value())
                # print(numDegrees)
            return super(Ui_MainFormGrunt, self).eventFilter(source, event)

        except Exception as ex:
            globalValues.writeLogData('Функция обработки событий в главном окне программы', str(ex))

    # def wheelEvent(self, event):
    #     print('qwertty!!!')
    #     print(event.pixelDelta())

    def generateMenu(self, pos):
            print("pos======", pos)
            # self.menu.exec_(self.tblMainData.mapToGlobal(pos))

            # +++

    def searchInTbl(self):
        try:
            # countRow = self.tblMainData.rowCount()
            # for i in range(countRow):
            #     self.tblMainData.removeRow(i)
            # # self.tblMainData.clear()
            # self.tblMainData.updateGeometry()
            # # time.sleep(5)
            # self.refreshTblMain()
            # time.sleep(0.3)
            j = 0

            globalValues.searchData = True
            globalValues.checkEditTbl = True

            sleep(0.6)

            # countItems = self.tblMainData.rowCount()
            numCurIndex = self.comboSearch.currentIndex()
            globalValues.searchData = self.leSearchTbl.text()

            print(numCurIndex)

            if (numCurIndex == 1):
                globalValues.findingMainOrder = True
            elif(numCurIndex == 2):
                globalValues.findingMainGRZ = True
            elif (numCurIndex == 3):
                globalValues.searchData = self.DateEdit.text()
                globalValues.findingMainDate = True

            globalValues.checkEditTbl = False
            globalValues.refreshTblMain = True

            # str_search = ''
            # if (j == 1):
            #
            #     print(str_search)
            # else:
            #     str_search = self.leSearchTbl.text()
            # list_del_rows = []
            #
            #
            # if ((j == 0 or j == 3 and countItems > 0) or j == 1):
            #     for i in range(countItems - 1, -1, -1):
            #             is_into = str_search in str(self.tblMainData.item(i, j).text())
            #             if (is_into == False):
            #                     self.tblMainData.removeRow(i)
            #     globalValues.stopUpdateMainTbl = True
            #
            # if ((self.leSearchTbl.text() == '' and j != 1) or j == 99):
            #     globalValues.stopUpdateMainTbl = False
            #     self.refreshOneTblMain = True


            strDataInJournal = 'Выполнен поиск по ' + str(self.comboSearch.currentText()) + ' ' + globalValues.searchData
            globalValues.writeEventToDBJournalMain('Системный журнал', strDataInJournal)

        except Exception as ex:
                globalValues.writeLogData('Функция поиска по главной таблице', str(ex))

    def changeLoginAcc(self):
        self.uiLogin = Ui_panel_autologin()
        # self.ui.leLogin.setText('admin')
        # self.ui.lePassword.setText('sinaps281082')
        self.uiLogin.isChangeLogin = True
        self.uiLogin.exec_()
        if (self.uiLogin.check_good_autologin):
            self.le_UserName.setText(globalValues.curUserName)
            self.changeAccount()

    def openPanRtsp(self):
        try:
            panRtsp = Ui_panRtspCams()
            panRtsp.exec_()
        except Exception as ex:
            globalValues.writeLogData('Функция открытия панели настройки ртсп', str(ex))

    def openPanelStWorking(self):
        try:
                uiStWork = Ui_select_st_working()
                uiStWork.comboBox.setCurrentIndex(self.currentModeStWorking)
                uiStWork.exec_()
                if (uiStWork.checkClickOk):

                    strDataMode = uiStWork.stWorkingCb

                    if (strDataMode == 'Автоматический'):
                        self.autoStMainWorking()
                    else:
                        self.handStMainWorking()

                    strDataToTS = 'Выполнена смена на ' + strDataMode
                    globalValues.writeEventToDBJournalMain('Режим работы ПО', strDataToTS)

        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели смены режима работы ', str(ex))

    def openPanelCon(self):
       try:
            uiContext = Ui_ContextTable()
            self.screenImgDisplay()
            self.is_first_con_menu = False
            uiContext.setGeometry(self.xContext, self.yContext, 160, 30)
            uiContext.startClickPython()
            uiContext.exec_()
            self.defCon = uiContext.defCon
            print('QQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQQ' + str(self.defCon))
            self.conEventJob()

       except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели контестного меню ', str(ex))

    def openPanelZakaz(self):
        try:
            self.ui = PanelZakaz()
            dateToday = datetime.date.today().strftime('%d.%m.%Y')
            self.ui.leDate.setText(str(dateToday))
            if (uiMain.tblMainData.rowCount() != 0):
                    lastEl = uiMain.tblMainData.item(0, 0).text()
                    numNewZakaz = int(lastEl) + 1
            else:
                    numNewZakaz = 1
            self.ui.leNumberZakaz.setText(str(numNewZakaz))
            self.ui.strIndexRow = ''
            self.ui.exec_()
            if (self.ui.is_save_data):
                    strDataInJournal = 'Добавлен новый З/Н на выполнение работ'
                    globalValues.writeEventToDBJournalMain('Панель управления', strDataInJournal)
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели заказа ', str(ex))

    def openPanelTalon(self):
        try:
            indexRow = self.tblMainData.currentRow()
            panTalon = Ui_Talon()

            panTalon.le_TakeGruntLocation.setText(globalValues.nameObject)
            panTalon.le_TalonNum.setText(self.tblMainData.item(indexRow, 0).text())
            panTalon.le_TalonDate.setText(self.tblMainData.item(indexRow, 1).text())
            panTalon.le_GRZNumber.setText(self.tblMainData.item(indexRow, 3).text())
            panTalon.le_Weight.setText(self.tblMainData.item(indexRow, 6).text())
            panTalon.le_Volume.setText(self.tblMainData.item(indexRow, 7).text())
            panTalon.le_OffGruntLocation.setText(globalValues.namePolygon)

            try:
                if (self.checkMySql):
                    if globalValues.debug:
                        self.con = pymysql.connect(host='localhost',
                                              port=3306,
                                              user='sergey',
                                              passwd='34ubitav',
                                              db=globalValues.dbMySqlName)
                    else:
                        self.con = pymysql.connect(host=globalValues.my_sql_localhost,
                                              port=globalValues.my_sql_port,
                                              user=globalValues.my_sql_name,
                                              passwd=globalValues.my_sql_password,
                                              db=globalValues.dbMySqlName)

                with self.con:
                    cur = self.con.cursor()

                    cur.execute("SELECT id, time_entry, time_check_out FROM " + globalValues.tblsDB[1])

                    rowsMain = cur.fetchall()

                    for row in reversed(rowsMain):
                        if (str(row[0]) == self.tblMainData.item(indexRow, 9).text()):
                            panTalon.le_TimeCome.setText(str(row[1]))
                            panTalon.le_TimeGone.setText(str(row[2]))
                            break

                    cur.execute("SELECT id, name_company, car_model FROM " + globalValues.tblsDB[2])

                    rowsOrder = cur.fetchall()

                    for row in reversed(rowsOrder):
                        if (str(row[0]) == self.tblMainData.item(indexRow, 10).text()):
                            panTalon.le_Organisation.setText(str(row[1]))
                            panTalon.le_CarModel.setText(str(row[2]))
                            break

                    cur.close()
            except Exception as ex:
                globalValues.writeLogData('Функция считывания данных из БД', str(ex))

            strTalon = panTalon.le_TalonNum.text()

            strDataToTS = 'Выполнено открытие талона #' + strTalon
            globalValues.writeEventToDBJournalMain('Главное окно', strDataToTS)

            panTalon.exec_()
        except Exception as ex:
            globalValues.writeLogData('Функция открытия ', str(ex))

    def openPanelReadyOrLinkDet(self):
        try:
            indexRow = self.tblMainData.currentRow()
            indexCol = self.tblMainData.currentColumn()
            print('IndexCol: ' + str(indexCol))

            if (indexCol == 8):
                try:
                    uiReportZN = Ui_panel_report()

                    if globalValues.debug:
                        self.con = pymysql.connect(host='localhost',
                                              port=3306,
                                              user='sergey',
                                              passwd='34ubitav',
                                              db=globalValues.dbMySqlName)

                    with self.con:
                        cur = self.con.cursor()

                        number_GRZ = self.tblMainData.item(indexRow, 3).text()
                        number_order = self.tblMainData.item(indexRow, 0).text()
                        date_order = self.tblMainData.item(indexRow, 1).text()
                        st_order = self.tblMainData.item(indexRow, 8).text()
                        weight_object = self.tblMainData.item(indexRow, 6).text()
                        date_complete_order = ''
                        volume_object = self.tblMainData.item(indexRow, 7).text()
                        name_company = ''
                        model_ts = ''
                        time_in_object = ''
                        time_out_object = ''
                        time_in_polygon = ''
                        time_in_road = ''
                        volume_polygon = ''
                        weight_polygon = ''

                        id = int(self.tblMainData.item(indexRow, 10).text())
                        print(id)
                        cur.execute("SELECT * FROM " + globalValues.tblsDB[2])
                        rowsOrder = cur.fetchall()
                        for row in rowsOrder:
                            if (row[0] == id):
                                name_company = str(row[4])
                                model_ts = str(row[5])
                                # time_in_object = str(row[10])
                                # time_out_object = str(row[11])
                                break



                        id = int(self.tblMainData.item(indexRow, 12).text())
                        print(id)
                        cur.execute("SELECT * FROM " + globalValues.tblsDB[5])
                        rowsPoly = cur.fetchall()



                        for row in rowsPoly:
                            if (row[0] == id):
                                time_in_polygon = str(row[3])
                                time_in_road = str(row[4])
                                volume_polygon = str(row[2])
                                weight_polygon = str(row[1])
                                break

                        cur.execute("SELECT * FROM " + globalValues.tblsDB[1])
                        idMain = int(self.tblMainData.item(indexRow, 9).text())
                        rowsMain = cur.fetchall()
                        for row in rowsMain:
                            if (row[0] == id):
                                time_in_object = str(row[10])
                                time_out_object = str(row[11])
                                break

                        uiReportZN.le_Organisation.setText(name_company)
                        uiReportZN.le_GRZNumber.setText(number_GRZ)
                        uiReportZN.le_CarModel.setText(model_ts)
                        uiReportZN.le_ZakazNumber.setText(number_order)
                        uiReportZN.le_ZakazDate.setText(date_order)
                        uiReportZN.le_ZakazState.setText(st_order)
                        uiReportZN.le_ZakazDate_2.setText(date_complete_order)
                        # uiReportZN.le_ComeToObject.setText(time_in_object)
                        # uiReportZN.le_GoneFromObject.setText(time_out_object)
                        uiReportZN.le_ComeToPoligon.setText(time_in_polygon)
                        uiReportZN.le_DriveOverTime.setText(time_in_road)
                        uiReportZN.le_WeightObject.setText(weight_object)
                        uiReportZN.le_WeightPoligon.setText(weight_polygon)
                        uiReportZN.le_VolumeObject.setText(volume_object)
                        uiReportZN.le_VolumePoligon.setText(volume_polygon)

                        cur.close()

                    uiReportZN.le_ComeToObject.setText(time_in_object)
                    uiReportZN.le_GoneFromObject.setText(time_out_object)
                    strZN = uiReportZN.le_ZakazNumber.text()

                    pathFiles = glob.glob(globalValues.pathReports + '/*.pdf')
                    print(pathFiles)
                    if (len(pathFiles) != 0):
                        dataSearchOrder = strZN
                        dataSearchFormat = '.pdf'

                        for el in pathFiles:
                            print(el)
                            is_order = dataSearchOrder in el
                            is_fmt = dataSearchFormat in el
                            if (is_order and is_fmt):
                                uiReportZN.btnFilterPrint.setEnabled(True)
                                break

                        if (is_order == False or is_fmt == False):
                            uiReportZN.btnFilterPrint.setEnabled(False)
                    else:
                        uiReportZN.btnFilterPrint.setEnabled(False)

                    pathImgWeight = ''
                    pathImgScan = ''

                    print(number_order)
                    print(globalValues.pathWeightImg)

                    rep = self.checkAndPathImgsForReport(number_order, globalValues.pathWeightImg)
                    print(rep)
                    print('123123123')
                    print(rep[0])
                    print(rep[1])
                    print(rep[2])
                    if rep[0]:
                        pathImgWeight = rep[1]
                        pathImgScan = rep[2]

                    print(os.path.exists(pathImgWeight))

                    if (pathImgWeight != '' and os.path.exists(pathImgWeight)):
                        # print('123')
                        img = QPixmap(pathImgWeight)
                        img = img.scaled(410, 410, Qt.KeepAspectRatio)
                        print(pathImgWeight)
                        # pathImgWeight = 'E:/ACMK/Кадры/100001_weight.jpg'
                        uiReportZN.le_frmKPPOUT.setPixmap(img)
                        uiReportZN.le_frmKPPOUT.raise_()

                    if (pathImgScan != '' and os.path.exists(pathImgScan)):
                        img = QPixmap(pathImgScan)
                        img = img.scaled(410, 410, Qt.KeepAspectRatio)
                        uiReportZN.le_frmKPPIN.setPixmap(img)
                        # uiReportZN.le`

                    strData = str(number_order)
                    strDataToTS = 'Выполнено открытие отчёта #' + strData
                    globalValues.writeEventToDBJournalMain('Главное окно', strDataToTS)

                    uiReportZN.exec_()
                except Exception as ex:
                    globalValues.writeLogData('Функция работы с панелью отчёта о З/Н', str(ex))
            elif (indexCol == 3):
                self.openWebBrowser()
            elif (indexCol == 0):
                self.openPanelTalon()
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели редактирования заказа или слежения ТС ', str(ex))

    def openPanelMapUrl(self):
        try:
            self.uiMap = Ui_DialogMap()
            self.uiMap.exec_()
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели настройки ссылок слежения ТС ', str(ex))

    def openPanelDevCams(self):
        try:
            # self.uiDev.numTab = 1
            self.uiDev.switchDisabled(1)
            self.uiDev.check()
            globalValues.stopThreadCams = False
            globalValues.stopEncryptedDecrypt = False
            self.uiDev.checkStartDB = True
            self.uiDev.startCheckConMySql()
            if self.uiDev.checkLogin():
                self.uiDev.checkArchive()
            time.sleep(0.25)
            self.uiDev.exec_()
            globalValues.stopThreadCams = True
            globalValues.stopEncryptedDecrypt = True

            global listRtsp
            listRtsp = self.uiDev.listRtsp

            # print('RTSPEdiiiiittttt: ' + str(listRtsp))

            self.rtsp_cam_out_gard = str(self.uiDev.listRtsp[0])
            self.rtsp_cam_in_gard = str(self.uiDev.listRtsp[1])
            self.rtsp_cam_in_weight = str(self.uiDev.listRtsp[2])
            self.rtsp_cam_out_weight = str(self.uiDev.listRtsp[3])

            globalValues.stopThreadCams = True
            sleep(0.2)
            self.updateGeometry()
            globalValues.stopThreadCams = False

            if (self.uiDev.is_good_cam_0):
                self.checkStCams[0] = True
                print('checkStart0')
                startThCams(1, self.rtsp_cam_out_gard)
            else:
                self.checkStCams[0] = False
            if (self.uiDev.is_good_cam_1):
                self.checkStCams[1] = True
                print('checkStart1')
                startThCams(2, self.rtsp_cam_in_gard)
            else:
                self.checkStCams[1] = False
            if (self.uiDev.is_good_cam_2):
                self.checkStCams[2] = True
                print('checkStart2')
                startThCams(3, self.rtsp_cam_in_weight)
            else:
                self.checkStCams[2] = False
            if (self.uiDev.is_good_cam_3):
                self.checkStCams[3] = True
                print('checkStart3')
                startThCams(4, self.rtsp_cam_out_weight)
            else:
                self.checkStCams[3] = False

            if (self.uiDev.is_good_cam_4):
                self.checkStCams[4] = True
                print('checkStart3')
                # startThCams(4, self.rtsp_cam_out_weight)
            else:
                self.checkStCams[4] = False

            if (self.uiDev.is_good_cam_5):
                self.checkStCams[5] = True
                print('checkStart3')
                # startThCams(4, self.rtsp_cam_out_weight)
            else:
                self.checkStCams[5] = False

        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели настройки камер ', str(ex))

    def openPanelDevWeight(self):
        try:
            # self.uiDev.numTab = 2
            self.uiDev.switchDisabled(2)
            time.sleep(0.25)
            self.uiDev.exec_()
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели весов ', str(ex))

    def openPanelDevGard(self):
        try:
            # self.uiDev.numTab = 3
            self.uiDev.switchDisabled(3)
            time.sleep(0.25)
            self.uiDev.exec_()
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели контроля доступа ', str(ex))

    def openPanelDevDB(self):
        try:
            # self.uiDev.numTab = 4
            self.uiDev.switchDisabled(4)
            self.uiDev.check()
            globalValues.stopThreadCams = False
            globalValues.stopEncryptedDecrypt = False
            self.uiDev.checkStartDB = True
            self.uiDev.startCheckConMySql()
            time.sleep(0.25)
            self.uiDev.exec_()
            # globalValues.stopThreadCams = True
            # globalValues.stopEncryptedDecrypt = True
            # globalValues.stopThreadCams = True
            # globalValues.stopEncryptedDecrypt = True
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели БД', str(ex))

    def openPanelDevScaner(self):
        try:
            # self.uiDev.numTab = 5
            globalValues.stopVideoScaner = False
            self.uiDev.switchDisabled(5)
            time.sleep(0.25)
            self.uiDev.lblVideoScan.setPixmap(QPixmap(self.pathDefNotCamScan))
            self.uiDev.exec_()
            globalValues.stopVideoScaner = True
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели сканера ', str(ex))

    def openPanelDevState(self):
        try:
            # self.uiDev.numTab = 6
            self.uiDev.switchDisabled(6)
            time.sleep(0.25)
            self.uiDev.exec_()
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели статуса системы ', str(ex))

    # def openPanelDevState(self):
    #     try:
    #         self.uiDev.numTab = 5
    #         self.uiDev.switchDisabled()
    #         self.uiDev.exec_()
    #     except Exception as ex:
    #             globalValues.writeLogData('Функция выполнения панели статуса системы ', str(ex))

    def openPanelLogin(self):
        try:
                self.screenImgDisplay()
                print('StartingLogin!')
                data = ''
                if (os.path.exists(self.pathFileErr)):
                    file_err = open(self.pathFileErr, 'r')
                    data = file_err.read()
                    file_err.close()

                if (data == '0'):
                    globalValues.checkRestartExe = True
                else:
                    file_err = open(self.pathFileErr, 'w')
                    file_err.write('0')
                    file_err.close()

                uiLog = Ui_panel_autologin()
                if globalValues.debugAdmin:
                    uiLog.leLogin.setText('admin')
                    if globalValues.debugLogin:
                        uiLog.lePassword.setText('sinaps281082')
                    else:
                        uiLog.lePassword.setText('sinaps281082')
                if globalValues.debugOperator:
                    uiLog.leLogin.setText('operator')
                    uiLog.lePassword.setText('34ubitav')
                if globalValues.debugSergey:
                    uiLog.leLogin.setText('sergey')
                    if globalValues.debugLogin:
                        uiLog.lePassword.setText('123')
                    else:
                        uiLog.lePassword.setText('qwE123')
                if (self.checkLoginEvent):
                    uiLog.btnInput.setGeometry QRect(903, 650, 120, 31))
                    uiLog.btnInput.setText('Подтвердить')
                if (globalValues.checkRestartExe):
                    uiLog.show()
                    uiLog.btnInput.click()
                else:
                    uiLog.exec_()
                if (self.checkLoginEvent and uiLog.check_good_autologin):
                    print('checkingSergio!')
                    my_thread_del_row_tbl_main = threading.Thread(target=self.startDelElsTbl)
                    my_thread_del_row_tbl_main.start()
                    self.is_open_pan_context = True
                elif (uiLog.check_good_autologin):
                    self.changeAccount()
                    self.openPanelSetDevices()
                else:
                    if (self.checkLoginEvent):
                        self.checkLoginEvent = False
                        # self.uiLog.close()
                    else:
                        # self.uiLog.close()
                        self.exitLogin = True
                        print('ExitSystem!!!')
                        globalValues.stopAll = True
                        # globalValues.stopThreadCams = True
                        # globalValues.stopCamInSet = True
                        # globalValues.stopSysPanel = True
                        # globalValues.stopEncryptedDecrypt = True
                        # globalValues.stopUpdateMainTbl = True
                        # globalValues.stopThreadCams = True
                        # sleep(2.5)
                        self.uiDev.close()
                        sleep(4)
                        self.close()
                        sleep(1)

        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели логина ', str(ex))

    def openPanelSetDevices(self):
            try:
                    th_measuring = threading.Thread(target=self.uiDev.measuringVol)
                    th_measuring.start()
                    self.uiDev.error_weight_signal.connect(self.errorWeightMes)
                    self.uiDev.is_good_con_weight.connect(self.goodConWeight)
                    self.uiDev.error_traffic_signal.connect(self.errorTrafficMes)
                    self.uiDev.is_good_con_traffic.connect(self.goodConTraffic)
                    # self.uiDev.closePanel.connect(self.hidePanelSet)
                    self.uiDev.advices.setCurrentIndex(1)
                    # self.uiDev.checkConDB()
                    # self.uiDev.checkConDBpg()
                    print('dwqdwdwd')
                    # self.uiDev.openPorts()
                    # self.uiDev.check()

                    time.sleep(0.3)
                    self.uiDev.startCheckConMySql()
                    time.sleep(0.75)
                    self.uiDev.startCheckConPgSql()

                    #
                    print('qweqwewqe')
                    self.uiDev.exec_()

                    if (globalValues.curUserName != 'sergey'):
                        self.btnRtsp0.hide()
                        self.btnRtsp1.hide()
                        self.btnRtsp2.hide()
                        self.btnRtsp3.hide()
                        self.btnCamSet0.hide()
                        self.btnCamSet1.hide()
                        self.btnCamSet2.hide()
                        self.btnCamSet3.hide()
                        self.btnZoomIn0.hide()
                        self.btnZoomIn1.hide()
                        self.btnZoomIn2.hide()
                        self.btnZoomIn3.hide()
                        self.btnZoomOut0.hide()
                        self.btnZoomOut1.hide()
                        self.btnZoomOut2.hide()
                        self.btnZoomOut3.hide()
                    else:
                        self.btnRtsp0.show()
                        self.btnRtsp1.show()
                        self.btnRtsp2.show()
                        self.btnRtsp3.show()
                        self.btnCamSet0.show()
                        self.btnCamSet1.show()
                        self.btnCamSet2.show()
                        self.btnCamSet3.show()
                        self.btnZoomIn0.show()
                        self.btnZoomIn1.show()
                        self.btnZoomIn2.show()
                        self.btnZoomIn3.show()
                        self.btnZoomOut0.show()
                        self.btnZoomOut1.show()
                        self.btnZoomOut2.show()
                        self.btnZoomOut3.show()
                    globalValues.stopThreadCams = True
                    globalValues.stopEncryptedDecrypt = True
                    # globalValues.stopThreadSysMenu = True


                    print('RTSPPPPPPPPPPPPPPPPPPPPPP: ' + str(self.uiDev.listRtsp))

                    global listRtsp
                    listRtsp = self.uiDev.listRtsp

                    self.rtsp_cam_out_gard = str(self.uiDev.listRtsp[0])
                    self.rtsp_cam_in_gard = str(self.uiDev.listRtsp[1])
                    self.rtsp_cam_in_weight = str(self.uiDev.listRtsp[2])
                    self.rtsp_cam_out_weight = str(self.uiDev.listRtsp[3])

                    print('Check:   qwe' + str(self.uiDev.is_good_cam_0))
                    print(self.rtsp_cam_out_gard)

                    self.updateGeometry()
                    globalValues.stopThreadCams = False

                    if (self.uiDev.is_good_cam_0):
                        self.checkStCams[0] = True
                        print('checkStart0')
                        startThCams(1, self.rtsp_cam_out_gard)
                    else:
                        self.checkStCams[0] = False
                    if (self.uiDev.is_good_cam_1):
                        self.checkStCams[1] = True
                        print('checkStart1')
                        startThCams(2, self.rtsp_cam_in_gard)
                    else:
                        self.checkStCams[1] = False
                    if (self.uiDev.is_good_cam_2):
                        self.checkStCams[2] = True
                        print('checkStart2')
                        startThCams(3, self.rtsp_cam_in_weight)
                    else:
                        self.checkStCams[2] = False
                    if (self.uiDev.is_good_cam_3):
                        self.checkStCams[3] = True
                        print('checkStart3')
                        startThCams(4, self.rtsp_cam_out_weight)
                    else:
                        self.checkStCams[3] = False

                    if (self.uiDev.is_good_cam_4):
                        self.checkStCams[4] = True
                        print('checkStart4')
                        # startThCams(4, self.rtsp_cam_out_weight)
                    else:
                        self.checkStCams[4] = False

                    if (self.uiDev.is_good_cam_5):
                        self.checkStCams[5] = True
                        print('checkStart5')
                        # startThCams(4, self.rtsp_cam_out_weight)
                    else:
                        self.checkStCams[5] = False

                    if (self.uiDev.goodLoginDBMySQL):
                        self.connectDBMySQL()
                    if (self.uiDev.goodLoginDBPgSQL and self.uiDev.goodLoginDBMySQL):
                        startThPGSQL()

                    sleep(0.2)
                    startThUpdateTable()
                    globalValues.checkUpdate = True

            except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели системных настроек ', str(ex))
            self.le_UserName.setText(globalValues.curUserName)
            self.show()

    def openPanelAboutMain(self):
        try:
                self.uiAboutMain = Ui_about_system_menu()
                self.uiAboutMain.exec_()
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели информации о главном окне ', str(ex))

    def openPanelAboutSysSets(self):
        try:
                self.ui = Ui_AboutSysMenu()
                self.ui.exec_()
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели информации о системном табе ', str(ex))

    def openPanelJournalTS(self):
        try:
            uiJournalTS = Ui_PanelJournalTS()
            uiJournalTS.exec_()
        except Exception as ex:
            globalValues.writeLogData('Функция запуска панели журнала ТС', str(ex))

    def openPanelJournal(self):
        try:
            self.uiJournal = Ui_Journal()

            if (globalValues.curUserName == 'operator'):
                self.uiJournal.btnFilterDel.hide()
                self.uiJournal.btnFilter.setGeometry(840, 100, 181, 31)
            else:
                self.uiJournal.btnFilterDel.show()
                self.uiJournal.btnFilter.setGeometry(740, 100, 181, 31)

            self.uiJournal.exec_()
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели журналирования системы ', str(ex))

    def openPanelBaseTS(self):
        try:
                self.uiPanelBaseTS = Ui_TS_Base()

                # sleep(0.2)
                # uiPanelBaseTS.exec_()

                if (globalValues.curUserName == 'operator'):
                    self.uiPanelBaseTS.btnDelTS.hide()
                    self.uiPanelBaseTS.btnAddTS.hide()
                else:
                    self.uiPanelBaseTS.btnDelTS.show()
                    self.uiPanelBaseTS.btnAddTS.show()

                if (globalValues.curUserName != 'sergey'):
                    self.uiPanelBaseTS.exec_()
                else:
                    self.uiPanelBaseTS.show()
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели база ТС ', str(ex))

    def openPanelAddZN(self):
        try:
                uiPanelAddZN = Ui_AddZakaz()
                # sleep(0.2)
                uiPanelAddZN.exec_()
        except Exception as ex:
                globalValues.writeLogData('Функция выполнения панели база ТС ', str(ex))

    def checkFolderLongPath(self, pathFolder):
        try:
            i = 0
            listPath = []
            for element in pathFolder:
                if (element == '/' and i != 0):
                    listPath.append(pathFolder[0:i])
                i += 1
            # listPath = pathFolder.split('/')
            # listPath.pop(0)
            listPath.append(pathFolder)
            print('ChekingPath:', listPath)
            for elPath in listPath:
                if (os.path.exists(elPath) == False):
                    try:
                        os.mkdir(elPath)
                        # os.system('sudo chmod -R 777' + elPath)
                    except Exception as ex:
                        globalValues.writeLogData('Функция проверки и создания папки хранилища', str(ex))

            if (len(listPath) > 0):
                return True
            else:
                return False
        except Exception as ex:
            globalValues.writeLogData('Функция проверки ссылки на папку хранилища', str(ex))
            return False

    def startThreadsMain(self):

            # th_out_gard = Thread(self)
            # th_out_gard.rtsp_link = 'rtsp://admin:admin@10.2.165.103:554/cam/realmonitor?channel=1&subtype=0'
            # # th_out_gard.rtsp_link = 'rtsp://admin:admin@10.2.165.103:554/out.h264'
            # th_out_gard.start()
            # th_out_gard.changePixmap.connect(self.setImageOutGard)
            #
            # sleep(0.1)
            #
            # th_in_gard = Thread(self)
            # th_in_gard.rtsp_link = 'rtsp://admin:admin@10.2.165.104:554/cam/realmonitor?channel=1&subtype=0'
            # th_in_gard.start()
            # th_in_gard.changePixmap.connect(self.setImageInGard)
            #
            # sleep(0.1)
            # #
            # th_in_weight = Thread(self)
            # th_in_weight.rtsp_link = 'rtsp://admin:qwe123456@10.2.165.101:554/cam/realmonitor?channel=1&subtype=0'
            #
            # th_in_weight.start()
            # th_in_weight.changePixmap.connect(self.setImageInWeight)
            #
            # sleep(0.1)
            #
            # th_out_weight = Thread(self)
            # th_out_weight.rtsp_link = 'rtsp://admin:qwe123456@10.2.165.102:554/cam/realmonitor?channel=1&subtype=0'
            #
            # th_out_weight.start()
            # th_out_weight.changePixmap.connect(self.setImageOutWeight)

            # th_out_gard.rtsp_link = 'rtsp://admin:admin@10.2.165.103:554/cam/realmonitor?channel=1&subtype=0'
            # th_in_gard.rtsp_link = 'rtsp://admin:admin@10.2.165.104:554/cam/realmonitor?channel=1&subtype=0'
            # th_in_weight.rtsp_link = 'rtsp://admin:qwe123456@10.2.165.101:554/cam/realmonitor?channel=1&subtype=0'
            # th_out_weight.rtsp_link = 'rtsp://admin:qwe123456@10.2.165.102:554/cam/realmonitor?channel=1&subtype=0'

            startThUpdateTable()

            # startThPGSQL()

    def setImageOutGard(self, image, check):

            if (self.check_full_cam0):
                    image = image.scaled(1670, 940, Qt.KeepAspectRatio)
                    self.leCamAllMon.setPixmap(QPixmap.fromImage(image))
            else:
                    image = image.scaled(400, 400, Qt.KeepAspectRatio)
                    self.leCam0.setPixmap(QPixmap.fromImage(image))

    def setImageInGard(self, image, check):
            if (self.check_full_cam1):
                    image = image.scaled(1670, 940, Qt.KeepAspectRatio)
                    self.leCamAllMon.setPixmap(QPixmap.fromImage(image))
            else:
                    image = image.scaled(400, 400, Qt.KeepAspectRatio)
                    self.leCam1.setPixmap(QPixmap.fromImage(image))

    def setImageInWeight(self, image, check):
            if (self.check_full_cam2):
                    image = image.scaled(1670, 940, Qt.KeepAspectRatio)
                    self.leCamAllMon.setPixmap(QPixmap.fromImage(image))
            else:
                    image = image.scaled(400, 400, Qt.KeepAspectRatio)
                    self.leCam2.setPixmap(QPixmap.fromImage(image))

    def setImageOutWeight(self, image, check):
            if (self.check_full_cam3):
                    image = image.scaled(1670, 940, Qt.KeepAspectRatio)
                    self.leCamAllMon.setPixmap(QPixmap.fromImage(image))
            else:
                    image = image.scaled(400, 400, Qt.KeepAspectRatio)
                    self.leCam3.setPixmap(QPixmap.fromImage(image))

    def openWindowCam0(self, event):

            if (self.check_full_cam0):
                    self.check_full_cam0 = False
            else:
                    self.check_full_cam0 = True
            self.frmCam.hide()
            self.frmTable.hide()
            self.frmMenu.hide()
            self.frmCntrlPanel.hide()
            self.frmCamAllMon.show()

    def openWindowCam1(self, event):
            if (self.check_full_cam1):
                    self.check_full_cam1 = False
            else:
                    self.check_full_cam1 = True
            self.frmCam.hide()
            self.frmTable.hide()
            self.frmMenu.hide()
            self.frmCntrlPanel.hide()
            self.frmCamAllMon.show()

    def openWindowCam2(self, event):
            if (self.check_full_cam2):
                    self.check_full_cam2 = False
            else:
                    self.check_full_cam2 = True
            self.frmCam.hide()
            self.frmTable.hide()
            self.frmMenu.hide()
            self.frmCntrlPanel.hide()
            self.frmCamAllMon.show()

    def openWindowCam3(self, event):
            if (self.check_full_cam3):
                    self.check_full_cam3 = False
            else:
                    self.check_full_cam3 = True
            self.frmCam.hide()
            self.frmTable.hide()
            self.frmMenu.hide()
            self.frmCntrlPanel.hide()
            self.frmCamAllMon.show()

    def createListVideos(self, pathFldrs):
        try:

            self.lstArchive.clear()

            if (os.path.exists(pathFldrs)):

                listPath = os.listdir(pathFldrs)
                k = 0
                print(listPath)
                for el in listPath:
                    if (os.path.isfile(os.path.join(pathFldrs, el)) == False):
                        itemAddFldr = QtWidgets.QTreeWidgetItem(self.lstArchive)
                        itemAddFldr.setText(0, str(el))
                        # itemInKPP.setFont(0, font)
                        # itemInKPP.setIcon(0, QIcon(self.pathCamNotCon))
                        listPathFls = glob.glob(pathFldrs + '/' + el + '/*.mp4')
                        print('List: ' + str(listPathFls))
                        self.lstArchive.addTopLevelItem(itemAddFldr)

                        curItem = self.lstArchive.topLevelItem(k)

                        k += 1

                        for elFls in listPathFls:
                            # print('Yes!' + str(elFls))
                            if (os.path.isfile(elFls)):
                                # print('Yes! ' + str(elFls))
                                listNames = str(elFls).split('\\')
                                nameFls = listNames[len(listNames) - 1]
                                QtWidgets.QTreeWidgetItem(curItem, [nameFls])
                                # itemFls.setText(0, str(elFls))
                                # itemAddFldr.addChild(itemFls)

            # self.pathFolderVideoStrg = 'E:/Архив/14.05.2020'



        except Exception as ex:
            globalValues.writeLogData('Функция создания списка видеофайлов', str(ex))

    # def closeWinFullCam(self, event):
    #         self.frmCam.show()
    #         self.frmTable.show()
    #         self.frmMenu.show()
    #         self.frmCamAllMon.hide()
    #         self.check_full_cam0 = False
    #         self.check_full_cam1 = False
    #         self.check_full_cam2 = False
    #         self.check_full_cam3 = False

    def exitSystem(self):
            if (self.exitLogin == False):
                uiMes = Ui_mes_box()
                uiMes.lblStrInfo.setText('Вы уверены, что хотите выйти?')
                # uiMes.btnOK.hide()
                # uiMes.label_iconques.setStyleSheet("background-color: rgb(75,75,75);\n"
                #                                    "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
                # uiMes.btnСancel.setText('Продолжить')
                uiMes.exec_()
                if (uiMes.checkCont):
                    file_err = open(self.pathFileErr, 'w')
                    file_err.write('1')
                    file_err.close()
                    globalValues.stopAll = True
                    globalValues.stopThreadCams = True
                    globalValues.stopCamInSet = True
                    globalValues.stopSysPanel = True
                    self.uiDev.close()
                    sleep(0.5)
                    self.close()

            else:
                file_err = open(self.pathFileErr, 'w')
                file_err.write('1')
                file_err.close()
                globalValues.stopAll = True
                globalValues.stopThreadCams = True
                globalValues.stopCamInSet = True
                globalValues.stopSysPanel = True
                # globalValues.stopEncryptedDecrypt = True
                # globalValues.stopUpdateMainTbl = True
                # globalValues.stopThreadCams = True
                self.uiDev.close()
                sleep(0.5)
                self.close()
                # uiMes = Ui_mes_box()
                # uiMes.lblStrInfo.setText('Вы уверены, что хотите выйти?')
                # # uiMes.btnOK.hide()
                # # uiMes.label_iconques.setStyleSheet("background-color: rgb(75,75,75);\n"
                # #                                    "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
                # # uiMes.btnСancel.setText('Продолжить')
                # uiMes.exec_()
                # if (uiMes.checkCont):
                #     globalValues.stopAll = True
                #     globalValues.stopThreadCams = True
                #     globalValues.stopCamInSet = True
                #     globalValues.stopSysPanel = True
                #     self.uiDev.close()
                #     sleep(0.5)
                #     self.close()

    def debugExitSystem(self):
        try:
            uiMes = Ui_mes_box()
            uiMes.lblStrInfo.setText('Вы уверены, что хотите выйти?')
            # uiMes.btnOK.hide()
            # uiMes.label_iconques.setStyleSheet("background-color: rgb(75,75,75);\n"
            #                                    "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
            # uiMes.btnСancel.setText('Продолжить')
            uiMes.exec_()
            if (uiMes.checkCont):
                globalValues.stopAll = True
                globalValues.stopThreadCams = True
                globalValues.stopCamInSet = True
                globalValues.stopSysPanel = True
                self.uiDev.close()
                sleep(0.5)
                self.close()
        except Exception as ex:
            globalValues.writeLogData('Функция закрытия программы в режиме отладки', str(ex))

    def conEventJob(self):

            if (self.defCon == 1):
                try:
                    globalValues.stopUpdateMainTbl = True
                    item = self.tblMainData.item(int(self.item_cur_row), int(self.item_cur_col))
                    # item.
                    item.setFlags Qt.ItemIsSelectable | Qt.ItemIsEnabled | Qt.ItemIsEditable)

                    # if (globalValues.colorForm == 0):

                    item.setBackground(QColor(84, 122, 181))
                        # item.set

                    globalValues.checkEditTbl = True
                    print('EditTable')

                    uiMain.tblMainData.editItem(item)
                    self.is_open_pan_context = True

                    strDataInJournal = 'Выполнено редактирование ячейки (' + str(self.item_cur_row) + ',' + str(self.item_cur_col) + ')'
                    globalValues.writeEventToDBJournalMain('Системный журнал', strDataInJournal)
                except Exception as ex:
                    globalValues.writeLogData('Функция редактирования ячейки таблицы главного окна', str(ex))

            elif (self.defCon == 2):
                    globalValues.stopUpdateMainTbl = True
                    try:
                            print('runEditingData!')
                            my_thread_edit_tbl_main= threading.Thread(target=self.startEditElsTbl)
                            my_thread_edit_tbl_main.start()
                    except Exception as ex:
                            globalValues.writeLogData('Функция запуска потока редактирования ячейки БД MySql', str(ex))
                    globalValues.searchData = False
                    globalValues.checkEditTbl = False
                    globalValues.is_edit_data = True

                    # thUpdateMainTable()

            elif (self.defCon == 3):
                    try:
                            self.checkLoginEvent = True
                            self.openPanelLogin()
                    except Exception as ex:
                            globalValues.writeLogData('Функция удаления З/Н', str(ex))

            elif (self.defCon == 4):
                    try:
                            print('startingDefaulValuesInItem!')
                            my_thread_edit_tblEl_main= threading.Thread(target=self.editDefElTbl)
                            my_thread_edit_tblEl_main.start()
                    except Exception as ex:
                            globalValues.writeLogData('Функция запуска потока задания дефолтного значения ячейки БД MySql', str(ex))
                    globalValues.searchData = False
                    globalValues.checkEditTbl = False

            else:
                    self.is_open_pan_context = True

     pyqtSlot()
    def conEventClickMouse(self):
        self.is_open_pan_context = True
        self.uiContext.hide()

    def startDelElsTbl(self):
            globalValues.searchData = False
            globalValues.checkEditTbl = False
            indexRow = self.item_cur_row
            print('CurIndex: ' + str(indexRow))
            dataNumberOrder = self.tblMainData.item(indexRow, 0).text()

            try:

                    con = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                          user=globalValues.my_sql_name,
                                          passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)

                    cur = con.cursor()

                    # with con:
                    if True:
                            print(self.tblMainData.item(indexRow, 9).text())
                            print(self.tblMainData.item(indexRow, 10).text())
                            print(self.tblMainData.item(indexRow, 11).text())
                            print(self.tblMainData.item(indexRow, 12).text())

                            numberId = int(self.tblMainData.item(indexRow, 9).text())
                            query = ("DELETE FROM " + globalValues.tblsDB[1] + " where id = (%s)")
                            cur.execute(query, numberId)
                            con.commit()

                            numberIdAllData = int(self.tblMainData.item(indexRow, 10).text())
                            query = ("DELETE FROM " + globalValues.tblsDB[2] + " where id = (%s)")
                            cur.execute(query, numberIdAllData)
                            con.commit()

                            numberIdMap = int(self.tblMainData.item(indexRow, 11).text())
                            query = ("DELETE FROM " + globalValues.tblsDB[4] + " where id = (%s)")
                            cur.execute(query, numberIdMap)
                            con.commit()

                            numberIdPoly = int(self.tblMainData.item(indexRow, 12).text())
                            query = ("DELETE FROM " + globalValues.tblsDB[5] + " where id = (%s)")
                            cur.execute(query, numberIdPoly)
                            con.commit()

                    strDataInJournal = 'Выполнено удаление З/Н #' + str(dataNumberOrder)
                    globalValues.writeEventToDBJournalMain('Системный журнал', strDataInJournal)

                    cur.close()
                    con.close()

            except Exception as ex:
                    globalValues.writeLogData('Функция удаления записи из БД MySql', str(ex))

            globalValues.is_edit_data = True
            self.refreshOneTblMain = True

    def startEditElsTbl(self):
            try:
                    indexItem = int(self.item_cur_row)
                    item = self.tblMainData.item(int(self.item_cur_row), int(self.item_cur_col))
                    item.setFlags Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    if (globalValues.colorForm == 1):
                        item.setBackground(QColor(235, 235, 235))
                    else:
                        item.setBackground(QColor(42, 42, 42))

                    con = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                          user=globalValues.my_sql_name,
                                          passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)

                    cur = con.cursor()

                    # with con:
                    if True:
                            dataItem = item.text()
                            global listID
                            intIDMySQL = int(self.tblMainData.item(indexItem, 9).text())

                            global listIDZAKAZData
                            intIDMYZakazAll = int(self.tblMainData.item(indexItem, 10).text())
                            if (self.item_cur_col == 0):
                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET number_order = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()

                                    query = ("UPDATE " + globalValues.tblsDB[2] + " SET number_order = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMYZakazAll))
                                    con.commit()

                            if (self.item_cur_col == 1):
                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET date = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()

                            if (self.item_cur_col == 2):
                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET number_grz = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()

                                    query = ("UPDATE " + globalValues.tblsDB[2] + " SET car_grz = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMYZakazAll))
                                    con.commit()

                            if (self.item_cur_col == 3):
                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET time = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()

                            if (self.item_cur_col == 4):
                                    check = 'кг' in dataItem
                                    if (check == False):
                                        dataItem = dataItem + ' кг'
                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET weight_empty = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()

                            if (self.item_cur_col == 5):
                                    check = 'кг' in dataItem
                                    if (check == False):
                                        dataItem = dataItem + ' кг'
                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET weight_load = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()

                            if (self.item_cur_col == 6):
                                    check = 'кг' in dataItem
                                    if (check == False):
                                        dataItem = dataItem + ' кг'
                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET weight = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()

                            if (self.item_cur_col == 7):
                                    check = 'кг' in dataItem
                                    if (check == False):
                                        dataItem = dataItem + ' м3'
                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET volume = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()

                            if (self.item_cur_col == 8):
                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET state_order = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()

                    strDataInJournal = 'Выполнено сохранение ячейки (' + str(self.item_cur_row) + ',' + str(self.item_cur_col) + ')'
                    globalValues.writeEventToDBJournalMain('Системный журнал', strDataInJournal)

                    cur.close()
                    con.close()

                    self.refreshOneTblMain = True

            except Exception as ex:
                    globalValues.writeLogData('Функция сохранения в БД MySql изменений в главной таблице ', str(ex))

    def editDefElTbl(self):
            try:
                    item = self.tblMainData.item(int(self.item_cur_row), int(self.item_cur_col))
                    item.setFlags Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                    item.setBackground(QColor(42, 42, 42))

                    con = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                          user=globalValues.my_sql_name,
                                          passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)

                    cur = con.cursor()

                    # with con:
                    if True:
                            global listID
                            intIDMySQL = int(self.tblMainData.item(int(self.item_cur_row), 9).text())

                            global listIDZAKAZData
                            intIDMYZakazAll = listIDZAKAZData[int(self.item_cur_row)]

                            if (self.item_cur_col == 4):
                                    dataItem = 'не измерена'

                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET weight_empty = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()


                            if (self.item_cur_col == 5):
                                    dataItem = 'не измерена'

                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET weight_load = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()


                            if (self.item_cur_col == 6):
                                    dataItem = 'не измерена'

                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET weight = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()

                            if (self.item_cur_col == 7):
                                    dataItem = 'не измерен'

                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET volume = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()

                            if (self.item_cur_col == 8):
                                    dataItem = 'выполняется'

                                    query = ("UPDATE " + globalValues.tblsDB[1] + " SET state_order = (%s) WHERE id = (%s)")
                                    cur.execute(query, (dataItem, intIDMySQL))
                                    con.commit()

                    strDataInJournal = 'Выполнено редактирование по умолчанию ячейки  (' + str(self.item_cur_row) + ',' + str(self.item_cur_col) + ')'
                    globalValues.writeEventToDBJournalMain('Системный журнал', strDataInJournal)

                    globalValues.is_edit_data = True
                    self.refreshOneTblMain = True

                    cur.close()
                    con.close()

            except Exception as ex:
                    globalValues.writeLogData('Функция сохранения в БД MySql дефолтных значений  ', str(ex))

    #  pyqtSlot()
    # def print(self):
    #         print('Serhio!!')
    #         Ui_MainFormGrunt.show()
    #         # th = Thread(self)
    #         # th.changePixmap.connect(self.setImage)
    #         # th.start()

     pyqtSlot()
    def errorWeightMes(self):
         try:
            # reply = QMessageBox.critical(self, 'Панель', 'Весы. Ошибка связи с COM-портом.', QMessageBox.Yes)

            # uiMes = Ui_mes_box()
            # uiMes.lblStrInfo.setText('Весы. Ошибка связи с COM-портом!')
            # uiMes.btnOK.hide()
            # if (globalValues.colorForm == 1):
            #     uiMes.label_iconques.setStyleSheet("background-color: rgb(235,235,235);\n"
            #                                        "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
            # else:
            #     uiMes.label_iconques.setStyleSheet("background-color: rgb(75,75,75);\n"
            #                                        "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
            # uiMes.btnCancel.setText('Продолжить')
            # uiMes.exec()

            globalValues.writeEventToDBJournalMain('Весовой контроль', 'Весы. Ошибка связи с COM-портом!')

            self.check_error_weight = True
            self.check_good_con_weight = False
            if (self.check_error_traffic == False):
                    # self.hide()
                    # self.openPanelDevCams()
                    print('errorWeight!!!')
         except Exception as ex:
             globalValues.writeLogData('Функция обработки ошибок порта весов', str(ex))

     pyqtSlot()
    def errorTrafficMes(self):
        try:
            # reply = QMessageBox.critical(self, 'Панель', 'Светофоры. Ошибка связи с COM-портом.', QMessageBox.Yes)
            self.btnMainSemaREDIN.hide()
            self.btnMainSemaALLRED.hide()
            self.btnMainSemaREDOUT.hide()
            # uiMes = Ui_mes_box()
            # uiMes.lblStrInfo.setText('Светофоры. Ошибка связи с COM-портом!')
            # uiMes.btnOK.hide()
            # if (globalValues.colorForm == 1):
            #     uiMes.label_iconques.setStyleSheet("background-color: rgb(235,235,235);\n"
            #                                        "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
            # else:
            #     uiMes.label_iconques.setStyleSheet("background-color: rgb(75,75,75);\n"
            #                                        "image: url(" + globalValues.pathStyleImgs + "icnAtt.png);")
            # uiMes.btnCancel.setText('Продолжить')
            # uiMes.exec()

            globalValues.writeEventToDBJournalMain('Весовой контроль', 'Светофоры. Ошибка связи с COM-портом!')

            self.check_error_traffic = True
            self.check_good_con_traffic = False
            if (self.check_error_weight == False):
                    # self.hide()
                    # self.openPanelDevCams()
                    print('errorTraffic!!!')

        except Exception as ex:
            globalValues.writeLogData('Функция обработки ошибок порта светофоров', str(ex))

     pyqtSlot()
    def goodConWeight(self):
            self.check_good_con_weight = True
            # if (self.check_good_con_traffic):
            #         connectDBMySQL()
                    # self.uiDev.hide()
                    # uiMain.show()
                    # self.startThreadsMain()

     pyqtSlot()
    def goodConTraffic(self):
            self.check_good_con_traffic = True
            # self.btnMainSemaREDIN.show()
            # self.btnMainSemaALLRED.show()
            # self.btnMainSemaREDOUT.show()
            # if (self.check_good_con_weight):
            #         connectDBMySQL()
            #         self.uiDev.hide()
            #         uiMain.show()
            #         self.startThreadsMain()

    def message(self):
            # reply = QMessageBox.question(self, 'Message',
            #                                    "Are you sure to quit?", QMessageBox.Yes, QMessageBox.No)
            msg = QMessageBox()
            msg.setWindowTitle('Mes')
            msg.setText('Are you ready?')
            msg.setIcon(QMessageBox.Information)
            msg.setStandardButtons(QMessageBox.No | QMessageBox.Yes)
            msg.setDetailedText('details')
            msg.buttonClicked.connect(self.popup_buttons)

            x = msg.exec_()

    def popup_buttons(self, i):
            print(i.text())

    # def closeEvent(self, event):
    #         print('End')
            # event.accept()

     pyqtSlot()
    def readAndTrDataWeight(self, data):
            print(str(data) + '111111111111111111111111122222222222222222222222323232323')

    def openWebBrowser(self):
        try:
            if (self.checkMySql):
                    self.con = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                          user=globalValues.my_sql_name,
                                          passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)


            indexRow = self.tblMainData.currentRow()

            with self.con:
                    cur = self.con.cursor()

                    cur.execute("SELECT * FROM " + globalValues.tblsDB[4])

                    rows = cur.fetchall()

                    valIt = 0

                    for row in reversed(rows):
                        if (valIt == indexRow):
                            dataUrl = str(row[1])
                            webbrowser.open(dataUrl, new=2)
                            break
                        valIt += 1
                    cur.close()
        except Exception as ex:
            globalValues.writeLogData('Функция открытия ссылки в браузере для отслеживания ТС', str(ex))

    def refreshTblMain(self):
        try:
                globalValues.checkEditMode = False
                globalValues.searchData = False
                globalValues.checkEditTbl = False
                # startThUpdateTable()
                globalValues.stopUpdateMainTbl = False
                globalValues.refreshTblMain = True
                strDataInJournal = 'Выполнен запуск автообновления'
                globalValues.writeEventToDBJournalMain('Системный журнал', strDataInJournal)
        except Exception as ex:
                globalValues.writeLogData('Функция обновления данных таблицы главного окна', str(ex))

    def checkMySql(self):
        try:
                if (self.con.open):
                        return False
                else:
                        return True
        except Exception as ex:
                globalValues.writeLogData('Функция проверки подключения к бд MySql', str(ex))

    def connectDBMySQL(self):
        i = 0
        while True:

            try:

                    self.con = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                               user=globalValues.my_sql_name,
                                               passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)

                    if (self.con.open):
                        break
            except Exception as ex:
                    globalValues.writeLogData('Функция подключения к БД MySql, главная форма', str(ex))
                    i += 1

            if (i > 2):
                break

            sleep(3)

    def openFullCam1(self):
            try:
                if (self.checkStCams[0]):
                    curRtsp = ''
                    self.numCurChannel = 0
                    curRtsp = str(listRtsp[self.numCurChannel + 6])
                    if (curRtsp != '' and len(curRtsp) > 7):
                            globalValues.stopThreadCams = True
                            sleep(1)
                            globalValues.stopThreadCams = False
                            self.frmMenu.hide()
                            self.frmCam.hide()
                            self.frmTable.hide()
                            self.frmCntrlPanel.hide()
                            self.frmCamAllMon.show()
                            startThCams(99, curRtsp)
                            strDataInJournal = 'Запущен полноэкранный просмотр канала ' + str(self.listChannel[self.numCurChannel])
                            globalValues.writeEventToDBJournalMain('Камеры', strDataInJournal)

            except Exception as ex:
                    strDataInLog = 'Функция открытия камеры ' + str(curRtsp)
                    globalValues.writeLogData(strDataInLog, str(ex))

    def openFullCam2(self):
            try:
                if (self.checkStCams[1]):
                    curRtsp = ''
                    self.numCurChannel = 1
                    curRtsp = str(listRtsp[self.numCurChannel + 6])
                    if (curRtsp != '' and len(curRtsp) > 7):
                            globalValues.stopThreadCams = True
                            sleep(1)
                            globalValues.stopThreadCams = False
                            self.frmMenu.hide()
                            self.frmCam.hide()
                            self.frmTable.hide()
                            self.frmCntrlPanel.hide()
                            self.frmCamAllMon.show()
                            startThCams(99, curRtsp)
                            strDataInJournal = 'Запущен полноэкранный просмотр канала ' + str(self.listChannel[self.numCurChannel])
                            globalValues.writeEventToDBJournalMain('Камеры', strDataInJournal)
            except Exception as ex:
                    strDataInLog = 'Функция открытия камеры ' + str(curRtsp)
                    globalValues.writeLogData(strDataInLog, str(ex))

    def openFullCam3(self):
            try:
                if (self.checkStCams[2]):
                    curRtsp = ''
                    self.numCurChannel = 2
                    curRtsp = str(listRtsp[self.numCurChannel + 6])
                    if (curRtsp != '' and len(curRtsp) > 7):
                            globalValues.stopThreadCams = True
                            sleep(1)
                            globalValues.stopThreadCams = False
                            self.frmMenu.hide()
                            self.frmCam.hide()
                            self.frmTable.hide()
                            self.frmCntrlPanel.hide()
                            self.frmCamAllMon.show()
                            startThCams(99, curRtsp)
                            strDataInJournal = 'Запущен полноэкранный просмотр канала ' + str(self.listChannel[self.numCurChannel])
                            globalValues.writeEventToDBJournalMain('Камеры', strDataInJournal)
            except Exception as ex:
                    strDataInLog = 'Функция открытия камеры ' + str(curRtsp)
                    globalValues.writeLogData(strDataInLog, str(ex))

    def openFullCam4(self):
            try:
                if (self.checkStCams[3]):
                    curRtsp = ''
                    self.numCurChannel = 3
                    curRtsp = str(listRtsp[self.numCurChannel + 6])
                    if (curRtsp != '' and len(curRtsp) > 7):
                            globalValues.stopThreadCams = True
                            sleep(1)
                            globalValues.stopThreadCams = False
                            self.frmMenu.hide()
                            self.frmCam.hide()
                            self.frmTable.hide()
                            self.frmCntrlPanel.hide()
                            self.frmCamAllMon.show()
                            startThCams(99, curRtsp)
                            strDataInJournal = 'Запущен полноэкранный просмотр канала ' + str(self.listChannel[self.numCurChannel])
                            globalValues.writeEventToDBJournalMain('Камеры', strDataInJournal)
            except Exception as ex:
                    strDataInLog = 'Функция открытия камеры ' + str(curRtsp)
                    globalValues.writeLogData(strDataInLog, str(ex))

    def closeWinFullCam(self):
        try:
            # print('checkingCloseFull')
            globalValues.stopThreadCams = True
            sleep(0.35)
            globalValues.stopThreadCams = False
            # sleep(0.2)
            self.frmCamAllMon.hide()
            self.frmCam.show()
            self.frmTable.show()
            self.frmMenu.show()
            self.frmCntrlPanel.show()
            # uiMain.frmCntrlPanel.()
            strDataInJournal = 'Остановлен режим полноэкранного просмотра канала ' + str(self.listChannel[self.numCurChannel])
            globalValues.writeEventToDBJournalMain('Камеры', strDataInJournal)
            rtsp_cam_out_gard = listRtsp[0]
            rtsp_cam_in_gard = listRtsp[1]
            rtsp_cam_out_weight = listRtsp[2]
            rtsp_cam_in_weight = listRtsp[3]
            if (self.checkStCams[0]):
                startThCams(1, rtsp_cam_out_gard)
            if (self.checkStCams[1]):
                startThCams(2, rtsp_cam_in_gard)
            if (self.checkStCams[2]):
                startThCams(3, rtsp_cam_out_weight)
            if (self.checkStCams[3]):
                startThCams(4, rtsp_cam_in_weight)

            # th_in_gard = Thread(self)
            # th_in_gard.rtsp_link = 'rtsp://admin:admin@10.2.165.104:554/cam/realmonitor?channel=1&subtype=0'
            # th_in_gard.start()
            # th_in_gard.changePixmap.connect(self.setImageInGard)
            #
            # sleep(0.1)
            # #
            # th_in_weight = Thread(self)
            # th_in_weight.rtsp_link = 'rtsp://admin:qwe123456@10.2.165.101:554/cam/realmonitor?channel=1&subtype=0'
            #
            # th_in_weight.start()
            # th_in_weight.changePixmap.connect(self.setImageInWeight)
            #
            # sleep(0.1)
            #
            # th_out_weight = Thread(self)
            # th_out_weight.rtsp_link = 'rtsp://admin:qwe123456@10.2.165.102:554/cam/realmonitor?channel=1&subtype=0'
            #
            # th_out_weight.start()
            # th_out_weight.changePixmap.connect(self.setImageOutWeight)
        except Exception as ex:
            globalValues.writeLogData('Функция выхода из полноэкранного просмотра видеопотока', str(ex))
    #WorkingWithVideo

    def playVideoStrg(self):
        # print('StartingPlay!!!')
        checkMp4 = 'mp4' in globalValues.pathMainVideoStrg
        if (os.path.exists(globalValues.pathMainVideoStrg) and checkMp4):
            try:
                self.playBackRateStrg()
                print('StartingPlay!!!')
                if self.mediaPlayerVideoStrg.state() != QMediaPlayer.PlayingState:
                    self.mediaPlayerVideoStrg.play()
                    if (globalValues.colorForm == 1):
                        self.btnPlayArchMPlay.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:0px;\n"
                                                         "border:2px solid rgb(150,150,150);\n"
                                                         "border-right: 1px solid rgb(150,150,150);\n"
                                                         "border-top: 2px solid rgb(150,150,150);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.png);}\n"
                                                         "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:0px;\n"
                                                         "border:2px solid rgb(150,150,150);\n"
                                                         "border-right: 1px solid rgb(150,150,150);\n"
                                                         "border-top: 2px solid rgb(150,150,150);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.png);}\n"
                                                         "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:0px;\n"
                                                         "border:2px solid rgb(150,150,150);\n"
                                                         "border-right: 1px solid rgb(150,150,150);\n"
                                                         "border-top: 2px solid rgb(150,150,150);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconpaus2grey.png);};")
                    else:
                        self.btnPlayArchMPlay.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconpaus2.png);};")
                else:
                    if (globalValues.colorForm == 1):
                        self.btnPlayArchMPlay.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:0px;\n"
                                                         "border:2px solid rgb(150,150,150);\n"
                                                         "border-right: 1px solid rgb(150,150,150);\n"
                                                         "border-top: 2px solid rgb(150,150,150);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                         "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:0px;\n"
                                                         "border:2px solid rgb(150,150,150);\n"
                                                         "border-right: 1px solid rgb(150,150,150);\n"
                                                         "border-top: 2px solid rgb(150,150,150);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                         "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:0px;\n"
                                                         "border:2px solid rgb(150,150,150);\n"
                                                         "border-right: 1px solid rgb(150,150,150);\n"
                                                         "border-top: 2px solid rgb(150,150,150);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);};")
                    else:
                        self.btnPlayArchMPlay.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);};")
                    self.pauseVideoStrg()
            except Exception as ex:
                globalValues.writeLogData('Функция запуска видеофайла главная форма', str(ex))

    def pauseVideoStrg(self):
        checkMp4 = 'mp4' in globalValues.pathMainVideoStrg
        if (os.path.exists(globalValues.pathMainVideoStrg) and checkMp4):
            if self.mediaPlayerVideoStrg.state() != QMediaPlayer.PausedState:
                self.mediaPlayerVideoStrg.pause()

    def stopVideoStrg(self):
        checkMp4 = 'mp4' in globalValues.pathMainVideoStrg
        if (os.path.exists(globalValues.pathMainVideoStrg) and checkMp4):
            try:
                if self.mediaPlayerVideoStrg.state() != QMediaPlayer.StoppedState:
                    self.mediaPlayerVideoStrg.stop()
                    if (globalValues.colorForm == 1):
                        self.btnPlayArchMPlay.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:0px;\n"
                                                         "border:2px solid rgb(150,150,150);\n"
                                                         "border-right: 1px solid rgb(150,150,150);\n"
                                                         "border-top: 2px solid rgb(150,150,150);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                         "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:0px;\n"
                                                         "border:2px solid rgb(150,150,150);\n"
                                                         "border-right: 1px solid rgb(150,150,150);\n"
                                                         "border-top: 2px solid rgb(150,150,150);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                         "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                         "color: rgb(255, 255, 255);\n"
                                                         "border-radius:0px;\n"
                                                         "border:2px solid rgb(150,150,150);\n"
                                                         "border-right: 1px solid rgb(150,150,150);\n"
                                                         "border-top: 2px solid rgb(150,150,150);\n"
                                                         "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);};")
                    else:
                        self.btnPlayArchMPlay.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);};")
            except Exception as ex:
                globalValues.writeLogData('Функция остановки видеофайла главная форма', str(ex))

    def positionChangedStrg(self, position):
        try:
            self.SldrArchMplay.setValue(position)
            if (position >= self.SldrArchMplay.maximum()):
                self.mediaPlayerVideoStrg.setPosition(0)
                self.mediaPlayerVideoStrg.stop()
                if (globalValues.colorForm == 1):
                    self.btnPlayArchMPlay.setStyleSheet("QPushButton:!hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-right: 1px solid rgb(150,150,150);\n"
                                                        "border-top: 2px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(142,187,208);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-right: 1px solid rgb(150,150,150);\n"
                                                        "border-top: 2px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(112,157,178);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(150,150,150);\n"
                                                        "border-right: 1px solid rgb(150,150,150);\n"
                                                        "border-top: 2px solid rgb(150,150,150);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay6grey.png);};")
                else:
                    self.btnPlayArchMPlay.setStyleSheet("QPushButton:!hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover {background-color: rgb(50,75,115);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);}\n"
                                                        "QPushButton:hover:pressed {background-color: rgb(84,122,181);\n"
                                                        "color: rgb(255, 255, 255);\n"
                                                        "border-radius:0px;\n"
                                                        "border:2px solid rgb(42,42,42);\n"
                                                        "border-left:1px solid rgb(42,42,42);\n"
                                                        "border-right:1px solid rgb(42,42,42);\n"
                                                        "border-top: 2px solid rgb(42,42,42);\n"
                                                        "image: url(" + globalValues.pathStyleImgs + "iconplay4.png);};")
        except Exception as ex:
            globalValues.writeLogData('Функция смены позиции видеофайла главная форма', str(ex))

    def durationChangedStrg(self, duration):
        try:
            self.SldrArchMplay.setRange(0, duration)
        except Exception as ex:
            globalValues.writeLogData('Функция смены позиции видеофайла главная форма', str(ex))

    def setPositionStrg(self, position):
        try:
            self.mediaPlayerVideoStrg.setPosition(position)
        except Exception as ex:
            globalValues.writeLogData('Функция смены позиции видеофайла главная форма', str(ex))

    def jumpBackStrg(self):

        try:
            delta = int (self.SldrArchMplay.maximum()/10)
            pos = 0
            if (self.SldrArchMplay.sliderPosition() - delta <= 0):
                pos = 0
            else:
                pos = int(self.SldrArchMplay.sliderPosition() - delta)
            self.SldrArchMplay.setValue(pos)
            self.mediaPlayerVideoStrg.setPosition(pos)

        except Exception as ex:
            globalValues.writeLogData('Функция промотки назад видеофайла архива', str(ex))

    def jumpForwardStrg(self):
        try:
            maxValSldr  = self.SldrArchMplay.maximum()
            delta = int( maxValSldr / 10)
            pos = 0
            if (self.SldrArchMplay.sliderPosition() + delta >= maxValSldr):
                pos = maxValSldr
            else:
                pos = int(self.SldrArchMplay.sliderPosition() + delta)
            self.SldrArchMplay.setValue(pos)
            self.mediaPlayerVideoStrg.setPosition(pos)

        except Exception as ex:
            globalValues.writeLogData('Функция промотки вперёд видеофайла архива', str(ex))

    def playBackRateStrg(self):
        try:
            dataRate = self.cbArchMPlay.currentText()
            dataRate = dataRate.replace('x', '')
            dataRate = float(dataRate)
            # print(dataRate)
            self.mediaPlayerVideoStrg.setPlaybackRate(dataRate)
        except Exception as ex:
            globalValues.writeLogData('Функция смены скорости воспроизведения видеофайла архива', str(ex))

    def pressSldrStrg(self):
        try:
            self.setPositionStrg(self.SldrArchMplay.sliderPosition())
        except Exception as ex:
            globalValues.writeLogData('Функция выставления позиции воспроизведения видеофайла', str(ex))

    def handleError(self):
        try:
            globalValues.writeLogData("Ошибка воспроизведения медиаплеера архива", self.mediaPlayerVideoStrg.errorString())
        except Exception as ex:
            globalValues.writeLogData('Функция обработки ошибки видеофайла главная форма', str(ex))

    def strCurDate(self):
        dateToday = datetime.date.today().strftime('%d.%m.%Y')
        str_date_today = str(dateToday)
        return str_date_today

    def strCurTime(self):
        str_cur_time = str(datetime.datetime.time(datetime.datetime.now()))
        len_cur_time = len(str_cur_time)
        str_cur_time = str_cur_time[0: (len_cur_time - 7)]
        str_cur_time = str_cur_time.replace(':', '.')
        return str_cur_time

    def setFontLst(self, fontSize, fontType, lstLe):
        try:
            font = QFont()
            font.setPointSize(fontSize)
            font.setFamily(fontType)
            for el in lstLe:
                el.setFont(font)
                print('ChangeFontEl')
        except Exception as ex:
            globalValues.writeLogData('Функция настройки шрифтов', str(ex))

    def checkAndPathImgsForReport(self, numOrder, pathImg):
        try:
            lengthPath = len(pathImg)
            pathFile = ''
            isWeight = False
            isScan = False
            pathImgWeight = ''
            pathImgScan = ''

            for name in glob.glob(pathImg + '/*.jpg'):
                print(name)
                pathFile = name
                name = name[lengthPath:len(name)]
                print(name)
                checkOrder = str(numOrder) in name
                checkWeight = 'weight' in name
                checkScan = 'scan' in name
                print(checkScan)
                print(checkOrder)
                pathFile = pathFile.replace('\\', '/')
                if (checkOrder and checkWeight):
                    pathImgWeight = pathFile
                    isWeight = True
                if (checkOrder and checkScan):
                    pathImgScan = pathFile
                    isScan = True
                if (isScan and isWeight):
                    return True, pathImgWeight, pathImgScan

            if (isWeight or isScan):
                return True, pathImgWeight, pathImgScan
            else:
                return False, pathImgWeight, pathImgScan
        except Exception as ex:
            globalValues.writeLogData('Функция проверки картинок для отчёта', str(ex))

    def changeAccount(self):
        try:
            if (globalValues.curUserName == 'sergey'):
                self.btnRtsp0.show()
                self.btnRtsp1.show()
                self.btnRtsp2.show()
                self.btnRtsp3.show()
                self.btnCamSet0.show()
                self.btnCamSet1.show()
                self.btnCamSet2.show()
                self.btnCamSet3.show()
                self.btnZoomIn0.show()
                self.btnZoomIn1.show()
                self.btnZoomIn2.show()
                self.btnZoomIn3.show()
                self.btnZoomOut0.show()
                self.btnZoomOut1.show()
                self.btnZoomOut2.show()
                self.btnZoomOut3.show()
                self.btnSetCams.setEnabled(True)
                self.btnSetCams.setToolTip('')
                self.btnGard.setEnabled(True)
                self.btnGard.setToolTip('')
                self.btnDB.setEnabled(True)
                self.btnDB.setToolTip('')
                self.btnStWorkingSystem.setEnabled(True)
                self.btnStWorkingSystem.setToolTip('')
                self.leSearchTbl.setGeometry QRect(130, 40, 488, 31))
                self.btnDelZakazSergey.hide()
            else:
                self.btnDelZakazSergey.hide()
                self.leSearchTbl.setGeometry QRect(22, 40, 596, 31))
                self.btnRtsp0.hide()
                self.btnRtsp1.hide()
                self.btnRtsp2.hide()
                self.btnRtsp3.hide()
                self.btnCamSet0.hide()
                self.btnCamSet1.hide()
                self.btnCamSet2.hide()
                self.btnCamSet3.hide()
                self.btnZoomIn0.hide()
                self.btnZoomIn1.hide()
                self.btnZoomIn2.hide()
                self.btnZoomIn3.hide()
                self.btnZoomOut0.hide()
                self.btnZoomOut1.hide()
                self.btnZoomOut2.hide()
                self.btnZoomOut3.hide()

                if (globalValues.curUserName == 'operator'):
                    self.btnSetCams.setEnabled(False)
                    self.btnSetCams.setToolTip('Отсутствуют права доступа!')
                    self.btnGard.setEnabled(False)
                    self.btnGard.setToolTip('Отсутствуют права доступа!')
                    self.btnDB.setEnabled(False)
                    self.btnDB.setToolTip('Отсутствуют права доступа!')
                    self.btnStWorkingSystem.setEnabled(False)
                    self.btnStWorkingSystem.setToolTip('Отсутствуют права доступа!')

                elif (globalValues.curUserName == 'admin'):
                    self.btnSetCams.setEnabled(True)
                    self.btnSetCams.setToolTip('')
                    self.btnGard.setEnabled(True)
                    self.btnGard.setToolTip('')
                    self.btnDB.setEnabled(True)
                    self.btnDB.setToolTip('')
                    self.btnStWorkingSystem.setEnabled(True)
                    self.btnStWorkingSystem.setToolTip('')


        except Exception as ex:
            globalValues.writeLogData('Функция смены аккаунта', str(ex))

    def startThDelSergey(self):
        th_del_orders = threading.Thread(target = self.thDelZakazSergey)
        th_del_orders.start()

    def thDelZakazSergey(self):
        try:
            con = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                  user=globalValues.my_sql_name,
                                  passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)

            globalValues.searchData = False
            globalValues.checkEditTbl = False



            countTree = self.tblMainData.rowCount()

            lstPool = []
            lst1 = []


            # for j in range(countTree - 1, -1, -1):
            for j in range(countTree):
                lst1.append(self.tblMainData.item(j, 0).text())
                lst1.append(self.tblMainData.item(j, 8).text())
                lst1.append(self.tblMainData.item(j, 9).text())
                lst1.append(self.tblMainData.item(j, 10).text())
                lst1.append(self.tblMainData.item(j, 11).text())
                lst1.append(self.tblMainData.item(j, 12).text())
                lstPool.append(lst1)
                lst1 = []

            # print('Pool1: ' + str(lstPool))
            # print('Pool2: ' + str(lstPoolSt))

            print(lstPool)

            for i in range(countTree):

                if (globalValues.stopAll):
                    break

                print('Checking: ' + str(i))
                # print(self.tblMainData.item(i, 0).text())
                dataNumberOrder = lstPool[i][0]
                dataStOrder = lstPool[i][1]
                if (dataStOrder == 'выполняется'):


                    indexRow = i


                    print('CurIndex: ' + str(indexRow))

                    try:

                            if (con.open == False):
                                if globalValues.debug:
                                    self.con = pymysql.connect(host='localhost',
                                                               port=3306,
                                                               user='sergey',
                                                               passwd='34ubitav',
                                                               db=globalValues.dbMySqlName)
                                else:
                                    self.con = pymysql.connect(host=globalValues.my_sql_localhost,
                                                               port=globalValues.my_sql_port,
                                                               user=globalValues.my_sql_name,
                                                               passwd=globalValues.my_sql_password,
                                                               db=globalValues.dbMySqlName)

                            cur = con.cursor()

                            # with con:
                            if True:
                                    # print(self.tblMainData.item(indexRow, 9).text())
                                    # print(self.tblMainData.item(indexRow, 10).text())
                                    # print(self.tblMainData.item(indexRow, 11).text())
                                    # print(self.tblMainData.item(indexRow, 12).text())

                                    numberId = int(lstPool[i][2])
                                    query = ("DELETE FROM " + globalValues.tblsDB[1] + " where id = (%s)")
                                    cur.execute(query, numberId)
                                    con.commit()

                                    numberIdAllData = int(lstPool[i][3])
                                    query = ("DELETE FROM " + globalValues.tblsDB[2] + " where id = (%s)")
                                    cur.execute(query, numberIdAllData)
                                    con.commit()

                                    numberIdMap = int(lstPool[i][4])
                                    query = ("DELETE FROM " + globalValues.tblsDB[4] + " where id = (%s)")
                                    cur.execute(query, numberIdMap)
                                    con.commit()

                                    numberIdPoly = int(lstPool[i][5])
                                    query = ("DELETE FROM " + globalValues.tblsDB[5] + " where id = (%s)")
                                    cur.execute(query, numberIdPoly)
                                    con.commit()

                            strDataInJournal = 'Выполнено удаление З/Н #' + str(dataNumberOrder)
                            globalValues.writeEventToDBJournalMain('Системный журнал', strDataInJournal)

                            cur.close()

                    except Exception as ex:
                            globalValues.writeLogData('Функция удаления записи из БД MySql', str(ex))

                    time.sleep(0.25)

            globalValues.is_edit_data = True
            self.refreshOneTblMain = True

        except Exception as ex:
            globalValues.writeLogData('Функция удаления з/н, которые не выполняются', str(ex))

    def createEventToWritingVideoChl_0(self):
        globalValues.numEvent[0] = 1

    def createEventToWritingVideoChl_1(self):
        globalValues.numEvent[1] = 1

    def createEventToWritingVideoChl_2(self):
        globalValues.numEvent[2] = 1

    def createEventToWritingVideoChl_3(self):
        globalValues.numEvent[3] = 1

    def changeSetsCamNightDay(self, name):
        try:
            print('checkingName: ' + name)
            numCam = int(name)
            j = 2*numCam
            if (numCam > 1):
                dataLogin = 'admin'
                # dataPassWd = 'qwE12345'
                dataPassWd = 'sinaps281082'
            else:
                dataLogin = 'admin'
                # dataPassWd = 'admin'
                dataPassWd = 'sinaps281082'
            strIp = globalValues.listIp[numCam]

            print(dataLogin)
            print(dataPassWd)
            print(strIp)

            cam = ONVIFCamera(strIp, 80, dataLogin, dataPassWd, globalValues.pathOnVif)

            # print(cam.devicemgmt.GetDeviceInformation())
            check = True

            if check:
                dt = cam.devicemgmt.GetSystemDateAndTime()

                # media_service = cam.create_media_service()
                #
                # profiles = media_service.GetProfiles()
                #
                # video = media_service.GetVideoSources()
                #
                # image = cam.create_imaging_service()

                media_service = cam.create_media_service()

                # profiles = media_service.GetProfiles()

                # video = media_service.GetVideoSources()


                image = cam.create_imaging_service()

                request = cam.imaging.create_type('SetImagingSettings')
                video_sources = cam.media.GetVideoSources()  # get video source to fetch token

                request.VideoSourceToken = video_sources[0].token  # ptz.ContinuousMove(requestc)

                data = cam.imaging.create_type('GetImagingSettings')
                data.VideoSourceToken = video_sources[0].token


                dataInfo = cam.imaging.GetImagingSettings(data)
                dataInfo = str(dataInfo)
                print("GetImgSet: " + dataInfo)
                strSearch = "IrCutFilter': 'OFF"
                print(strSearch)
                checkIs = strSearch in dataInfo

                print("dataCam:" + str(dataInfo))
                print(checkIs)

                dataStCam = 'OFF'

                if checkIs:
                    dataStCam = 'ON'

                request.ImagingSettings = {
                    # 'Brightness': 50,
                    # 'Contrast': 55,
                    'IrCutFilter': dataStCam
                }
                cam.imaging.SetImagingSettings(request)

                # data = cam.imaging.create_type('GetImagingSettings')
                # data.VideoSourceToken = video_sources[0].token
                # print(cam.imaging.GetImagingSettings(data))
        except Exception as ex:
            globalValues.writeLogData('Функция работы с камерами по onvif', str(ex))

    def changeCamNightDayAuto(self, name, str_cur_time_st):
        try:
            print('qwertyhgfssd')
            print('checkingName: ' + name)
            numCam = int(name)
            j = 2*numCam
            if (numCam > 1 and numCam < 4):
                dataLogin = 'admin'
                # dataPassWd = 'qwE12345'
                dataPassWd = 'sinaps281082'
            else:
                dataLogin = 'admin'
                # dataPassWd = 'admin'
                dataPassWd = 'sinaps281082'
            strIp = globalValues.listIp[numCam]

            print(dataLogin)
            print(dataPassWd)
            print(strIp)

            cam = ONVIFCamera(strIp, 80, dataLogin, dataPassWd, globalValues.pathOnVif)

            print(cam.devicemgmt.GetDeviceInformation())
            check = True

            if check:

                dt = cam.devicemgmt.GetSystemDateAndTime()

                # media_service = cam.create_media_service()

                #
                # profiles = media_service.GetProfiles()

                #
                # video = media_service.GetVideoSources()

                #
                # image = cam.create_imaging_service()

                media_service = cam.create_media_service()

                # profiles = media_service.GetProfiles()

                # video = media_service.GetVideoSources()


                image = cam.create_imaging_service()

                request = cam.imaging.create_type('SetImagingSettings')
                video_sources = cam.media.GetVideoSources()  # get video source to fetch token

                request.VideoSourceToken = video_sources[0].token  # ptz.ContinuousMove(requestc)

                data = cam.imaging.create_type('GetImagingSettings')
                data.VideoSourceToken = video_sources[0].token


                dataInfo = cam.imaging.GetImagingSettings(data)
                dataInfo = str(dataInfo)
                print("GetImgSet: " + dataInfo)
                # strSearch = "IrCutFilter': 'OFF"
                # print(strSearch)
                # checkIs = strSearch in dataInfo
                #
                # print("dataCam:" + str(dataInfo))
                # print(checkIs)
                #
                # dataStCam = 'OFF'
                #
                # if checkIs:
                #     dataStCam = 'ON'

                request.ImagingSettings = {
                    # 'Brightness': 50,
                    # 'Contrast': 55,
                    'IrCutFilter': str_cur_time_st
                }
                cam.imaging.SetImagingSettings(request)

                # data = cam.imaging.create_type('GetImagingSettings')
                # data.VideoSourceToken = video_sources[0].token
                # print(cam.imaging.GetImagingSettings(data))
        except Exception as ex:
            globalValues.writeLogData('Функция работы с камерами по onvif', str(ex))

    def lstTimingFromXls(self):

        try:

            # print('Path xls: ' + str(globalValues.pathFileTiming))

            pathXls = globalValues.pathFileTiming
            # pathXls = str(os.getenv('APPDATA')) + r'\Sinaps\timing.xlsx'

            # print(pathXls)

            book = load_workbook(pathXls)

            sheets = book.sheetnames

            dataListIP = []

            rtsp_link_list = []

            lst_time_data = []

            firstCall = True

            lst_data = []

            for sheet in sheets:
                print(str(sheet))
                if (str(sheet) == 'Лист1'):
                    sheet_main = book[sheet]
                    i = 1
                    while True:
                        # print(sheet_main.cell(i, 1).value)
                        if (sheet_main.cell(i, 1).value != None):
                            date = sheet_main.cell(i, 1).value
                            time_in = str(sheet_main.cell(i, 3).value)
                            time_out = str(sheet_main.cell(i, 4).value)
                            lst_all_data = [date, time_in, time_out]
                            if (sheet_main.cell(i, 1).value == 1):
                                if (firstCall == False):
                                    lst_data.append(lst_time_data)
                                else:
                                    firstCall = False
                                lst_time_data = []

                            print(sheet_main.cell(i, 1).value)
                            # valueIP = sheet_main.cell(i, 1).value
                            lst_time_data.append(lst_all_data)
                        else:
                            lst_data.append(lst_time_data)
                            break
                        i += 1

            # print(lst_data)
            # print(len(lst_data))

            return lst_data
        except Exception as ex:
            globalValues.writeLogData('Функция считывания данных из файла эксель', str(ex))

def strCurTime():
        str_cur_time = str(datetime.datetime.time(datetime.datetime.now()))
        len_cur_time = len(str_cur_time)
        str_cur_time = str_cur_time[0: (len_cur_time - 7)]
        str_cur_time = str_cur_time.replace(':', '.')
        return str_cur_time

def startThUpdateTable():
        try:
                if (globalValues.checkUpdate):
                    print('StartUpdateTbl!')
                    globalValues.checkUpdate = False
                    my_thread_updating_main_tbl = threading.Thread(target=thUpdateMainTable)
                    my_thread_updating_main_tbl.start()

        except Exception as ex:
                globalValues.writeLogData('Функция запуска потока обновления главной таблицы ', str(ex))

# @profile
def thUpdateMainTable():

    # if True:
    try:
        firstCallDay = [True, True, True, True, True, True]
        firstCallNight = [True, True, True, True, True, True]
        startTimeCheck = round(time.time()*100)

        firstInCheck = True

        start_time_clear_cache = round(time.time()*100)
        passwd = globalValues.passwdLogin
        # command = 'sh -c "/usr/bin/echo 3 > /proc/sys/vm/drop_caches"'
        command = 'echo 3 | sudo tee /proc/sys/vm/drop_caches'

        while True:
            checkBreak = False

            countRowsOld = 0

            # if True:
            try:
                if globalValues.debug:
                    con = pymysql.connect(host='localhost',
                                          port=3306,
                                          user='sergey',
                                          passwd='34ubitav',
                                          db=globalValues.dbMySqlName)
                else:
                    con = pymysql.connect(host=globalValues.my_sql_localhost,
                                               port=globalValues.my_sql_port,
                                               user=globalValues.my_sql_name,
                                               passwd=globalValues.my_sql_password,
                                               db=globalValues.dbMySqlName)

                cur = con.cursor()

                start_time = round(time.time()*100)
                dataOld = ''
                start_time_day_night = round(time.time()*100)

                while True:

                    check = False
                    dataCur = globalValues.checkCamStrg

                    if (abs(round(time.time()*100) - start_time_clear_cache) > globalValues.timeClearCache):
                        start_time_clear_cache = round(time.time() * 100)

                        # os.system('sudo -S sh -c "/usr/bin/echo 3 > /proc/sys/vm/drop_caches"')
                        # os.popen("sudo -S %s" % (command), 'w').write(passwd)
                        os.system(command)
                        gc.collect()
                        # print(gc.get_threshold())
                        # while True:
                        # eat()
                        print('ClearingCache!!!')

                    if (str(dataCur) != dataOld):

                        # print('CheckingChangeRecCam: ')
                        # print(dataCur)

                        for i in range(4):
                            if (dataCur[i] == False):
                                if (i == 0):
                                    if (globalValues.colorForm == 0):
                                        uiMain.leCam0Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                "image: url(" + globalValues.pathStyleImgs + "iconrec13lightgrey.png);\n"
                                                                "border-radius: 8px;")
                                    else:
                                        uiMain.leCam0Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                       "image: url(" + globalValues.pathStyleImgs + "iconrec13grey.png);\n"
                                                                       "border-radius: 8px;")
                                elif (i == 1):
                                    if (globalValues.colorForm == 0):
                                        uiMain.leCam1Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                "image: url(" + globalValues.pathStyleImgs + "iconrec13lightgrey.png);\n"
                                                                "border-radius: 8px;")
                                    else:
                                        uiMain.leCam1Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                       "image: url(" + globalValues.pathStyleImgs + "iconrec13grey.png);\n"
                                                                       "border-radius: 8px;")
                                elif (i == 2):
                                    if (globalValues.colorForm == 0):
                                        uiMain.leCam2Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                "image: url(" + globalValues.pathStyleImgs + "iconrec13lightgrey.png);\n"
                                                                "border-radius: 8px;")
                                    else:
                                        uiMain.leCam2Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                       "image: url(" + globalValues.pathStyleImgs + "iconrec13grey.png);\n"
                                                                       "border-radius: 8px;")
                                elif (i == 3):
                                    if (globalValues.colorForm == 0):
                                        uiMain.leCam3Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                "image: url(" + globalValues.pathStyleImgs + "iconrec13lightgrey.png);\n"
                                                                "border-radius: 8px;")
                                    else:
                                        uiMain.leCam3Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                       "image: url(" + globalValues.pathStyleImgs + "iconrec13grey.png);\n"
                                                                       "border-radius: 8px;")
                            else:
                                if (i == 0):
                                    if (globalValues.colorForm == 0):
                                        uiMain.leCam0Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                       "image: url(" + globalValues.pathStyleImgs + "iconrec13.png);\n"
                                                                       "border-radius: 8px;")
                                    else:
                                        uiMain.leCam0Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                "image: url(" + globalValues.pathStyleImgs + "iconrec13.png);\n"
                                                                "border-radius: 8px;")
                                elif (i == 1):
                                    if (globalValues.colorForm == 0):
                                        uiMain.leCam1Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                       "image: url(" + globalValues.pathStyleImgs + "iconrec13.png);\n"
                                                                       "border-radius: 8px;")
                                    else:
                                        uiMain.leCam1Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                "image: url(" + globalValues.pathStyleImgs + "iconrec13.png);\n"
                                                                "border-radius: 8px;")
                                if (i == 2):
                                    if (globalValues.colorForm == 0):
                                        uiMain.leCam2Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                       "image: url(" + globalValues.pathStyleImgs + "iconrec13.png);\n"
                                                                       "border-radius: 8px;")
                                    else:
                                        uiMain.leCam2Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                "image: url(" + globalValues.pathStyleImgs + "iconrec13.png);\n"
                                                                "border-radius: 8px;")
                                if (i == 3):
                                    if (globalValues.colorForm == 0):
                                        uiMain.leCam3Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                       "image: url(" + globalValues.pathStyleImgs + "iconrec13.png);\n"
                                                                       "border-radius: 8px;")
                                    else:
                                        uiMain.leCam3Rec.setStyleSheet("background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(0, 0, 0, 0), stop:1 rgba(255, 255, 255, 0));\n"
                                                                "image: url(" + globalValues.pathStyleImgs + "iconrec13.png);\n"
                                                                "border-radius: 8px;")

                            # nameOb.clear()

                        dataOld = str(globalValues.checkCamStrg)

                    if (globalValues.checkEditTbl == False):

                        if (((abs(round(time.time()*100) - start_time) > 150 or checkBreak) and globalValues.stopUpdateMainTbl == False) or uiMain.refreshOneTblMain or globalValues.refreshTblMain):

                            # print(globalValues.refreshTblMain)

                            start_time = round(time.time()*100)

                            if ((globalValues.searchData == False) or (uiMain.leSearchTbl.text() == '') or uiMain.refreshOneTblMain or globalValues.refreshTblMain):
                                # if True:
                                try:
                                        print('StartingWorkingWithBD!')

                                        if (con.open == False):
                                            con = pymysql.connect(host=globalValues.my_sql_localhost,
                                                                       port=globalValues.my_sql_port,
                                                                       user=globalValues.my_sql_name,
                                                                       passwd=globalValues.my_sql_password,
                                                                       db=globalValues.dbMySqlName)

                                            time.sleep(0.2)

                                            print('Try ConBD!!!')

                                        # print(con.open)
                                        # with con:
                                        with con:

                                                rows = []

                                                # print(con.open)

                                                if (globalValues.checkPoolTS):
                                                    globalValues.checkPoolTS = False
                                                    sqlQ = ("SELECT reg_number, weight_empty FROM " + globalValues.tblsDB[6])
                                                    cur.execute(sqlQ)
                                                    rowsGRZPool = cur.fetchall()
                                                    globalValues.listPoolTS = []
                                                    globalValues.listWeightEmptyTS = []
                                                    for row in rowsGRZPool:
                                                        globalValues.listPoolTS.append(str(row[0]))
                                                        globalValues.listWeightEmptyTS.append(str(row[1]))

                                                    print(rowsGRZPool)

                                                    print('ChekingPool!!!')
                                                    # print(globalValues.listPoolTS)
                                                    # print(globalValues.listWeightEmptyTS)

                                                #
                                                # print(globalValues.listPoolTS)
                                                # print('checkingListSt!!!')

                                                sqlID = 'SELECT id FROM ' +globalValues.tblsDB[1]+ ' order by id desc limit 1'

                                                cur.execute(sqlID)

                                                rows_id = cur.fetchall()

                                                num_end_order = 0

                                                if (len(rows_id) != 0):
                                                    num_end_order = rows_id[0][0]

                                                num_start_order = str(num_end_order - globalValues.valNumShowOeder)

                                                sqlCom = ''

                                                if (globalValues.findingMainOrder):

                                                    globalValues.checkEditTbl = True

                                                    sqlCom = "SELECT * FROM " + globalValues.tblsDB[1] + " WHERE number_order LIKE " +globalValues.searchData+ ""

                                                elif(globalValues.findingMainGRZ):

                                                    globalValues.checkEditTbl = True

                                                    sqlCom = "SELECT * FROM " + globalValues.tblsDB[1] + " WHERE number_grz LIKE '" + globalValues.searchData + "%'"

                                                elif (globalValues.findingMainDate):

                                                    globalValues.checkEditTbl = True

                                                    sqlCom = "SELECT * FROM " + globalValues.tblsDB[1] + " WHERE date LIKE '" + globalValues.searchData + "%'"

                                                elif (num_end_order > globalValues.valNumShowOeder):

                                                    print(num_end_order)

                                                    sqlCom = 'SELECT * FROM ' +globalValues.tblsDB[1]+ ' LIMIT ' + num_start_order + ', ' + str(num_end_order) + ''

                                                else:

                                                    sqlCom = 'SELECT * FROM ' + globalValues.tblsDB[1]

                                                cur.execute(sqlCom)

                                                rows = cur.fetchall()

                                                print('SerhioPerezNew!')

                                                countRows = 0

                                                for row in rows:
                                                        countRows += 1

                                                print(countRows, ' ', countRowsOld)
                                                print(uiMain.refreshOneTblMain)
                                                print(globalValues.refreshTblMain)

                                                if ((countRows != 0 and countRows != countRowsOld) or uiMain.refreshOneTblMain or globalValues.refreshTblMain):

                                                        print('SergggioCol!')

                                                        valIt = countRows - 1
                                                        valWdg = 0
                                                        check_bad_data = False

                                                        # print('Check123')

                                                        if (globalValues.refreshTblMain):
                                                            check_bad_data = True
                                                            globalValues.refreshTblMain = False

                                                        # print('Check123')

                                                        if (countRows == int(uiMain.tblMainData.rowCount()) and uiMain.refreshOneTblMain == False and globalValues.refreshTblMain == False):

                                                                iter = 0
                                                                for row in reversed(rows):
                                                                        if (str(row[1]) != uiMain.tblMainData.item(iter, 0).text()):
                                                                                check_bad_data = True

                                                                        if (str(row[8]) != uiMain.tblMainData.item(iter, 1).text()):
                                                                                check_bad_data = True

                                                                        if (str(row[2]) != uiMain.tblMainData.item(iter, 2).text()):
                                                                                check_bad_data = True

                                                                        if (str(row[9]) != uiMain.tblMainData.item(iter, 3).text()):
                                                                                check_bad_data = True

                                                                        if (str(row[3]) != uiMain.tblMainData.item(iter, 4).text()):
                                                                                check_bad_data = True

                                                                        if (str(row[4]) != uiMain.tblMainData.item(iter, 5).text()):
                                                                                check_bad_data = True

                                                                        if (str(row[5]) != uiMain.tblMainData.item(iter, 6).text()):
                                                                                check_bad_data = True

                                                                        if (str(row[6]) != uiMain.tblMainData.item(iter, 7).text()):
                                                                                check_bad_data = True

                                                                        if (str(row[7]) != uiMain.tblMainData.item(iter, 8).text()):
                                                                                check_bad_data = True

                                                                        if (globalValues.is_edit_data):
                                                                                globalValues.is_edit_data = False
                                                                                check_bad_data = True

                                                                        if (check_bad_data):
                                                                                break
                                                                        iter += 1

                                                        else:
                                                                check_bad_data = True

                                                        print('Check123: ', globalValues.findingMainOrder)

                                                        numOrders = 0

                                                        # check_bad_data = False

                                                        if (check_bad_data):
                                                                uiMain.tblMainData.setRowCount(0)
                                                                global listID
                                                                listID = []

                                                                # release_list(globalValues.listCurTsSt)
                                                                # release_list(globalValues.timelistTs)

                                                                globalValues.listCurTsSt = []
                                                                globalValues.timelistTs = []
                                                                val_st = 0

                                                                for row in reversed(rows):

                                                                        # sleep(0.005)

                                                                        # print(row)

                                                                        # numOrders += 1

                                                                        valColor = globalValues.colorForm

                                                                        listID.append(str(row[0]))

                                                                        uiMain.tblMainData.insertRow(valWdg)

                                                                        # print('checkingRefreshTbl!!!')

                                                                        # print(str(row) + 'qwewqewqewqe')

                                                                        item = QtWidgets.QTableWidgetItem(str(row[1]))
                                                                        item.setTextAlignment(Qt.AlignCenter)
                                                                        if (valColor == 0):
                                                                            item.setForeground(QBrush(Qt.white))
                                                                        else:
                                                                            item.setForeground(QBrush(Qt.black))
                                                                        item.setFlags Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        uiMain.tblMainData.setItem(valWdg, 0, item)

                                                                        item = QtWidgets.QTableWidgetItem(str(row[8]))
                                                                        # print(str(row[8]))
                                                                        item.setTextAlignment(Qt.AlignCenter)
                                                                        if (valColor == 0):
                                                                            item.setForeground(QBrush(Qt.white))
                                                                        else:
                                                                            item.setForeground(QBrush(Qt.black))
                                                                        item.setFlags(
                                                                         Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        # print(item.text())
                                                                        uiMain.tblMainData.setItem(valWdg, 1, item)
                                                                        item = QtWidgets.QTableWidgetItem(str(row[2]))
                                                                        item.setTextAlignment(Qt.AlignCenter)
                                                                        if (valColor == 0):
                                                                            item.setForeground(QBrush(Qt.white))
                                                                        else:
                                                                            item.setForeground(QBrush(Qt.black))
                                                                        item.setFlags(
                                                                             Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        uiMain.tblMainData.setItem(valWdg, 3, item)

                                                                        item = QtWidgets.QTableWidgetItem(str(row[9]))
                                                                        item.setTextAlignment(Qt.AlignCenter)
                                                                        if (valColor == 0):
                                                                            item.setForeground(QBrush(Qt.white))
                                                                        else:
                                                                            item.setForeground(QBrush(Qt.black))
                                                                        item.setFlags(
                                                                         Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        uiMain.tblMainData.setItem(valWdg, 2, item)

                                                                        item = QtWidgets.QTableWidgetItem(str(row[3]))
                                                                        item.setTextAlignment(Qt.AlignCenter)
                                                                        if (valColor == 0):
                                                                            item.setForeground(QBrush(Qt.white))
                                                                        else:
                                                                            item.setForeground(QBrush(Qt.black))
                                                                        item.setFlags(
                                                                             Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        uiMain.tblMainData.setItem(valWdg, 4, item)

                                                                        item = QtWidgets.QTableWidgetItem(str(row[4]))
                                                                        item.setTextAlignment(Qt.AlignCenter)
                                                                        if (valColor == 0):
                                                                            item.setForeground(QBrush(Qt.white))
                                                                        else:
                                                                            item.setForeground(QBrush(Qt.black))
                                                                        item.setFlags(
                                                                             Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        uiMain.tblMainData.setItem(valWdg, 5, item)

                                                                        item = QtWidgets.QTableWidgetItem(str(row[5]))
                                                                        item.setTextAlignment(Qt.AlignCenter)
                                                                        if (valColor == 0):
                                                                            item.setForeground(QBrush(Qt.white))
                                                                        else:
                                                                            item.setForeground(QBrush(Qt.black))
                                                                        item.setFlags(
                                                                             Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        uiMain.tblMainData.setItem(valWdg, 6, item)

                                                                        item = QtWidgets.QTableWidgetItem(str(row[6]))
                                                                        item.setTextAlignment(Qt.AlignCenter)
                                                                        if (valColor == 0):
                                                                            item.setForeground(QBrush(Qt.white))
                                                                        else:
                                                                            item.setForeground(QBrush(Qt.black))
                                                                        item.setFlags(
                                                                             Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        uiMain.tblMainData.setItem(valWdg, 7, item)

                                                                        item = QtWidgets.QTableWidgetItem(str(row[7]))
                                                                        item.setTextAlignment(Qt.AlignCenter)
                                                                        if (valColor == 0):
                                                                            item.setForeground(QBrush(Qt.white))
                                                                        else:
                                                                            item.setForeground(QBrush(Qt.black))
                                                                        item.setFlags(
                                                                             Qt.ItemIsSelectable | Qt.ItemIsEnabled)

                                                                        if (str(row[7]) == 'не выполняется'):
                                                                            if (valColor == 0):
                                                                                item.setBackground(QColor(149, 26, 26))
                                                                                item.setForeground(QColor(255, 255, 255))
                                                                            else:
                                                                                item.setBackground(QColor(197, 104, 104))
                                                                                item.setForeground(QColor(0, 0, 0))
                                                                        elif (str(row[7]) == 'выполняется'):
                                                                            if (valColor == 0):
                                                                                item.setBackground(QColor(149, 26, 26))
                                                                                item.setForeground(QColor(255, 255, 255))
                                                                            else:
                                                                                item.setBackground(QColor(197, 104, 104))
                                                                                item.setForeground(QColor(0, 0, 0))
                                                                            globalValues.listCurTsSt.append([])
                                                                            globalValues.listCurTsSt[valWdg].append(str(row[2]))
                                                                            globalValues.listCurTsSt[valWdg].append(True)
                                                                        elif (str(row[7]) == 'перемещение'):
                                                                            if (valColor == 0):
                                                                                item.setBackground(QColor(6, 81, 16))
                                                                                item.setForeground(QColor(255, 255, 255))
                                                                            else:
                                                                                item.setBackground(QColor(84, 181, 100))
                                                                                item.setForeground(QColor(0, 0, 0))
                                                                            globalValues.listCurTsSt.append([])
                                                                            globalValues.timelistTs.append([])
                                                                            globalValues.listCurTsSt[valWdg].append(str(row[2]))

                                                                            print('checking!!!')
                                                                            print(valWdg)
                                                                            print(globalValues.timelistTs)

                                                                            globalValues.timelistTs[val_st].append(str(row[2]))
                                                                            globalValues.listCurTsSt[valWdg].append(False)
                                                                            globalValues.timelistTs[val_st].append(str(row[11]))
                                                                            globalValues.timelistTs[val_st].append(str(row[8]))
                                                                            val_st += 1

                                                                        uiMain.tblMainData.setItem(valWdg, 8, item)

                                                                        item = QtWidgets.QTableWidgetItem(str(row[0]))
                                                                        item.setFlags(
                                                                         Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        uiMain.tblMainData.setItem(valWdg, 9, item)

                                                                        valIt -= 1
                                                                        valWdg += 1

                                                                        # if (numOrders > globalValues.valNumShowOeder):
                                                                        #     break



                                                                print('checkingStartBD!')

                                                                sqlQ = ''
                                                                if (globalValues.findingMainOrder):
                                                                    sqlQ = "SELECT id FROM " + globalValues.tblsDB[2] + " WHERE number_order LIKE " +globalValues.searchData+ ""
                                                                elif (globalValues.findingMainGRZ):
                                                                    sqlQ = "SELECT id FROM " + globalValues.tblsDB[2] + " WHERE car_grz LIKE '" + globalValues.searchData + "%'"
                                                                elif (globalValues.findingMainDate):
                                                                    sqlQ = "SELECT id FROM " + globalValues.tblsDB[2] + " WHERE date_work LIKE '" + globalValues.searchData + "%'"
                                                                elif (num_end_order > globalValues.valNumShowOeder):
                                                                    sqlQ = 'SELECT id FROM ' + globalValues.tblsDB[2] + ' LIMIT ' + num_start_order + ', ' + str(num_end_order) + ''
                                                                else:
                                                                    sqlQ = ("SELECT id FROM " + globalValues.tblsDB[2])
                                                                cur.execute(sqlQ)
                                                                print('checkingEndBD!')
                                                                rowsOrder = []
                                                                rowsOrder = cur.fetchall()

                                                                i = 0

                                                                global listIDZAKAZData
                                                                # release_list(listIDZAKAZData)
                                                                listIDZAKAZData = []

                                                                for rowOrder in reversed(rowsOrder):
                                                                    item = QtWidgets.QTableWidgetItem(str(rowOrder[0]))
                                                                    item.setFlags Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                    uiMain.tblMainData.setItem(i, 10, item)
                                                                    listIDZAKAZData.append(str(rowOrder[0]))
                                                                    i += 1



                                                                rowsOrder = []
                                                                rowsTrack = []
                                                                if (globalValues.findingMainOrder or globalValues.findingMainDate or globalValues.findingMainGRZ):
                                                                    rowsTrack = listID
                                                                    i = 0
                                                                    for rowTrack in rowsTrack:
                                                                        item = QtWidgets.QTableWidgetItem(
                                                                            str(rowTrack))
                                                                        item.setFlags(
                                                                         Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        uiMain.tblMainData.setItem(i, 11, item)
                                                                        i += 1
                                                                else:
                                                                    if (num_end_order > globalValues.valNumShowOeder):
                                                                        sqlQ = 'SELECT id FROM ' + globalValues.tblsDB[4] + ' LIMIT ' + num_start_order + ', ' + str(num_end_order) + ''
                                                                    else:
                                                                        sqlQ = ("SELECT id FROM " + globalValues.tblsDB[4])
                                                                    cur.execute(sqlQ)

                                                                    rowsTrack = cur.fetchall()

                                                                    i = 0
                                                                    for rowTrack in reversed(rowsTrack):
                                                                        item = QtWidgets.QTableWidgetItem(str(rowTrack[0]))
                                                                        item.setFlags(
                                                                         Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        uiMain.tblMainData.setItem(i, 11, item)
                                                                        i += 1

                                                                rowsTrack = []

                                                                rowsPoly = []
                                                                if (globalValues.findingMainOrder or globalValues.findingMainDate or globalValues.findingMainGRZ):
                                                                    rowsPoly = listID
                                                                    i = 0
                                                                    for rowPoly in rowsPoly:
                                                                        item = QtWidgets.QTableWidgetItem(
                                                                            str(rowPoly))
                                                                        item.setFlags(
                                                                         Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        uiMain.tblMainData.setItem(i, 12, item)
                                                                        i += 1
                                                                else:
                                                                    if (num_end_order > globalValues.valNumShowOeder):
                                                                        sqlQ = 'SELECT id FROM ' + globalValues.tblsDB[5] + ' LIMIT ' + num_start_order + ', ' + str(num_end_order) + ''
                                                                    else:
                                                                        sqlQ = ("SELECT id FROM " + globalValues.tblsDB[5])
                                                                    cur.execute(sqlQ)

                                                                    rowsPoly = cur.fetchall()

                                                                    i = 0
                                                                    for rowPoly in reversed(rowsPoly):
                                                                        item = QtWidgets.QTableWidgetItem(str(rowPoly[0]))
                                                                        item.setFlags Qt.ItemIsSelectable | Qt.ItemIsEnabled)
                                                                        uiMain.tblMainData.setItem(i, 12, item)
                                                                        i += 1

                                                                rowsPoly = []

                                                                rows = []

                                                                # global listIDZAKAZData
                                                                # # release_list(listIDZAKAZData)
                                                                # listIDZAKAZData = []
                                                                #
                                                                # if (num_end_order > globalValues.valNumShowOeder):
                                                                #     sqlQ = 'SELECT id FROM ' + globalValues.tblsDB[2] + ' LIMIT ' + num_start_order + ', ' + str(num_end_order) + ''
                                                                # else:
                                                                # cur.execute("SELECT * FROM " + globalValues.tblsDB[2])
                                                                #
                                                                # rowsZakaz = cur.fetchall()
                                                                #
                                                                # for row in reversed(rowsZakaz):
                                                                #         listIDZAKAZData.append(str(row[0]))

                                                        print('Check123')
                                                        # print(globalValues.timelistTs)

                                                        uiMain.refreshOneTblMain = False

                                                countRowsOld = countRows
                                                globalValues.findingMainOrder = False
                                                globalValues.findingMainGRZ = False
                                                globalValues.findingMainDate = False
                                                # print('qwerty123')
                                                th = threading.Thread(target=thChangeScroll, args=(True,))
                                                th.start()

                                                # release_list(rows)

                                except Exception as ex:
                                    globalValues.writeLogData('Функция разового обновления таблицы главного окна ', str(ex))
                                    checkBreak = True

                            if (globalValues.stopAll):
                                globalValues.checkUpdate = True
                                break

                            if (checkBreak):
                                cur.close()
                                con.close()
                                break

                    else:
                        sleep(0.005)

                    if ((abs(round(time.time()*100) - start_time_day_night) > 18000) or (abs(round(time.time()*100) - start_time_day_night) > 4500 and firstInCheck)):

                        firstInCheck = False

                        if (globalValues.lstTimingDayNight != []):

                            if (uiMain.checkStCams[0]):
                                res_def = time_cur_in_out(globalValues.lstTimingDayNight)

                                val_cur_time = res_def[0]
                                val_time_in = res_def[1]
                                val_time_out = res_def[2]

                                if (val_time_in <= val_cur_time <= val_time_out):

                                    if firstCallDay[0]:
                                        print('WorkMyCodOnvifDay')
                                        uiMain.changeCamNightDayAuto('0', 'ON')
                                        firstCallDay[0] = False
                                        firstCallNight[0] = True
                                else:

                                    if firstCallNight[0]:
                                        print('WorkMyCodOnvifNight')
                                        uiMain.changeCamNightDayAuto('0', 'OFF')
                                        firstCallDay[0] = True
                                        firstCallNight[0] = False

                                # print('GoodCam2')

                            if (uiMain.checkStCams[1]):

                                res_def = time_cur_in_out(globalValues.lstTimingDayNight)

                                val_cur_time = res_def[0]
                                val_time_in = res_def[1]
                                val_time_out = res_def[2]

                                if (val_time_in <= val_cur_time <= val_time_out):

                                    if firstCallDay[1]:
                                        print('WorkMyCodOnvifDay')
                                        uiMain.changeCamNightDayAuto('1', 'ON')
                                        firstCallDay[1] = False
                                        firstCallNight[1] = True
                                else:

                                    if firstCallNight[1]:
                                        print('WorkMyCodOnvifNight')
                                        uiMain.changeCamNightDayAuto('1', 'OFF')
                                        firstCallDay[1] = True
                                        firstCallNight[1] = False

                            if (uiMain.checkStCams[2]):
                                res_def = time_cur_in_out(globalValues.lstTimingDayNight)

                                val_cur_time = res_def[0]
                                val_time_in = res_def[1]
                                val_time_out = res_def[2]

                                if (val_time_in <= val_cur_time <= val_time_out):

                                    if firstCallDay[2]:
                                        print('WorkMyCodOnvifDay')
                                        uiMain.changeCamNightDayAuto('2', 'ON')
                                        firstCallDay[2] = False
                                        firstCallNight[2] = True
                                else:

                                    if firstCallNight[2]:
                                        print('WorkMyCodOnvifNight')
                                        uiMain.changeCamNightDayAuto('2', 'OFF')
                                        firstCallDay[2] = True
                                        firstCallNight[2] = False

                                # print('GoodCam2')

                            if (uiMain.checkStCams[3]):

                                res_def = time_cur_in_out(globalValues.lstTimingDayNight)

                                val_cur_time = res_def[0]
                                val_time_in = res_def[1]
                                val_time_out = res_def[2]

                                if (val_time_in <= val_cur_time <= val_time_out):

                                    if firstCallDay[3]:
                                        print('WorkMyCodOnvifDay')
                                        uiMain.changeCamNightDayAuto('3', 'ON')
                                        firstCallDay[3] = False
                                        firstCallNight[3] = True
                                else:

                                    if firstCallNight[3]:
                                        print('WorkMyCodOnvifNight')
                                        uiMain.changeCamNightDayAuto('3', 'OFF')
                                        firstCallDay[3] = True
                                        firstCallNight[3] = False

                            if (uiMain.checkStCams[4]):

                                res_def = time_cur_in_out(globalValues.lstTimingDayNight)

                                val_cur_time = res_def[0]
                                val_time_in = res_def[1]
                                val_time_out = res_def[2]

                                if (val_time_in <= val_cur_time <= val_time_out):

                                    if firstCallDay[4]:
                                        print('WorkMyCodOnvifDay')
                                        uiMain.changeCamNightDayAuto('4', 'ON')
                                        firstCallDay[4] = False
                                        firstCallNight[4] = True
                                else:

                                    if firstCallNight[4]:
                                        print('WorkMyCodOnvifNight')
                                        uiMain.changeCamNightDayAuto('4', 'OFF')
                                        firstCallDay[4] = True
                                        firstCallNight[4] = False

                                # print('GoodCam2')

                            if (uiMain.checkStCams[5]):

                                res_def = time_cur_in_out(globalValues.lstTimingDayNight)

                                val_cur_time = res_def[0]
                                val_time_in = res_def[1]
                                val_time_out = res_def[2]

                                if (val_time_in <= val_cur_time <= val_time_out):

                                    if firstCallDay[5]:
                                        print('WorkMyCodOnvifDay')
                                        uiMain.changeCamNightDayAuto('5', 'ON')
                                        firstCallDay[5] = False
                                        firstCallNight[5] = True
                                else:

                                    if firstCallNight[5]:
                                        print('WorkMyCodOnvifNight')
                                        uiMain.changeCamNightDayAuto('5', 'OFF')
                                        firstCallDay[5] = True
                                        firstCallNight[5] = False

                            # print('GoodCam3')

                        start_time_day_night = abs(round(time.time()*100))

                        print('WorkingCircleRefreshMainTable!!!')

                    if (globalValues.stopAll):
                        break

                    # gc.collect()

                    sleep(0.1)

            except Exception as ex:
                globalValues.writeLogData('Функция обновления таблицы главного окна ', str(ex))

            if (globalValues.stopAll or globalValues.checkEditTbl):
                break

            sleep(0.1)

            if (abs(round(time.time()*100) - startTimeCheck) > 200):
                uiMain.frmCam.updateGeometry()

    except Exception as ex:
        globalValues.writeLogData('Поток обновления таблицы главного окна', str(ex))

#Working with PgDB and change MySql after
def startThPGSQL():
        print('qwerty234')
        thread_pg_my_sql = threading.Thread(target=thPgMySQL, args=(True, ))
        thread_pg_my_sql.start()

# @profile
def thPgMySQL(checkStart):

     records_old = []

     # channel_hik_right = 'tqHi1zpY_OZvAOM4O' #Весы въезд
     # channel_hik_left = 'OCRYvVkf_OZvAOM4O' #Весы выезд
     # channel_active_left = 'NzRMUFEY_OZvAOM4O' # КПП въезд
     # channel_active_right = 'ztqCdlKM_OZvAOM4O' # КПП выезд

     channel_hik_right = 'enwNXgse_lSkzON1V'
     channel_hik_left = 'dFpO0w65_lSkzON1V'
     channel_active_left = 'enwNXgse_lSkzON1V'
     channel_active_right = 'dFpO0w65_lSkzON1V'

     # channel_hik_right = ''
     # channel_hik_left = ''
     # channel_active_left = 'tqHi1zpY_OZvAOM4O'
     # channel_active_right = ''

     checkQuality = False
     firstCallIn = True
     try:

     # if True:
             while True:

                    conn_pg = psycopg2.connect(dbname=globalValues.my_pg_db, user=globalValues.my_pg_name,
                                                   password=globalValues.my_pg_password, host= globalValues.my_pg_localhost, port= globalValues.my_pg_port)

                    con_mysql = pymysql.connect(host=globalValues.my_sql_localhost, port=globalValues.my_sql_port,
                                              user=globalValues.my_sql_name,
                                              passwd=globalValues.my_sql_password, db=globalValues.dbMySqlName)

                    try:
                    # if True:

                        cursor = conn_pg.cursor()
                        cur = con_mysql.cursor()

                        while True:
                                # print('123123')

                                if (con_mysql.open == False):
                                    con_mysql = pymysql.connect(host=globalValues.my_sql_localhost,
                                                                port=globalValues.my_sql_port,
                                                                user=globalValues.my_sql_name,
                                                                passwd=globalValues.my_sql_password,
                                                                db=globalValues.dbMySqlName)
                                    cur = con_mysql.cursor()

                                    # print('reconnectMySql')

                                if conn_pg is None:
                                    conn_pg = psycopg2.connect(dbname=globalValues.my_pg_db,
                                                               user=globalValues.my_pg_name,
                                                               password=globalValues.my_pg_password,
                                                               host=globalValues.my_pg_localhost,
                                                               port=globalValues.my_pg_port)
                                    cursor = conn_pg.cursor()

                                    # print('reconnectPgSql')

                                with conn_pg:

                                        cursor.execute('SELECT plate, channel, time_enter, quality  FROM auto_log order by id desc limit 1')

                                        # print('werwerwer')

                                        recordsPG = cursor.fetchall()

                                        # print("RecordsPG: " + str(records))

                                        for rowPG in recordsPG:
                                            quality = float(rowPG[3])
                                            if (quality > globalValues.valDetect):
                                                # print('checkingQuality: ')
                                                # print(quality)
                                                checkQuality = True
                                            else:
                                                checkQuality = False
                                            break

                                        checkZN = False

                                        # print(records_old)

                                        if (recordsPG != records_old and checkQuality):
                                                # print('qwerty12')
                                                list_grz_zakaz = []
                                                list_num_talon = []
                                                curListRtsp = []
                                                curListIp = []

                                                # with con_mysql:
                                                with con_mysql:
                                                        cur.execute("SELECT number_grz, state_order, number_order FROM " + globalValues.tblsDB[1])

                                                        rows_my = cur.fetchall()

                                                        time.sleep(0.3)

                                                        iter = 0
                                                        if (globalValues.testSystem == False):
                                                            for row_my in reversed(rows_my):
                                                                    # print(str(row_my[0] + ' --- ' + str(row_my[1])))
                                                                    if (str(row_my[1]) == 'выполняется'):
                                                                            list_grz_zakaz.insert(iter, str(row_my[0]))
                                                                            list_num_talon.insert(iter, str(row_my[2]))
                                                                            iter += 1

                                                            # print('AllListGRZ: ', list_grz_zakaz)
                                                            # print('AllNTalon: ', list_num_talon)

                                                for rowPG in recordsPG:
                                                        strGRZ = str(rowPG[0])
                                                        strGRZ = strGRZ.replace('(', '').replace(')', '').replace(',', '').replace('\'', '').replace(' ', '').replace('?', '').replace('/', '').upper()
                                                        j = 0

                                                        # print('ListOrderInProgress: ' + str(list_grz_zakaz))

                                                        for str_db_grz in list_grz_zakaz:
                                                                iter_s = 0
                                                                check_good_grz = 0
                                                                lenAr = min(len(str_db_grz), len(strGRZ))
                                                                countSymb = 0
                                                                countVal = 0
                                                                for i in range(lenAr):
                                                                        # print('CheckingElements!!!')
                                                                        # print(strGRZ[iter_s])
                                                                        # print(str_db_grz[iter_s])

                                                                        numVals = 1
                                                                        numLetter = 0

                                                                        if (globalValues.debugOrder == False):
                                                                            numVals = 2
                                                                            numLetter = 2

                                                                        if (0 < iter_s < 4):
                                                                            if (strGRZ[iter_s] == str_db_grz[iter_s]):
                                                                                countVal += 1
                                                                        if ((iter_s == 0) or (iter_s == 4) or (iter_s == 5)):
                                                                            if (strGRZ[iter_s] == str_db_grz[iter_s]):
                                                                                countSymb += 1

                                                                        if (countVal >= numVals and countSymb >= numLetter):
                                                                                checkZN = True
                                                                                print('goooooooooooooooooood')
                                                                                print(str(rowPG[1]))
                                                                                globalValues.numTalonVideo = list_num_talon[j]
                                                                                globalValues.numGRZVideo = str_db_grz
                                                                                name_ch = str(rowPG[1])
                                                                                globalValues.curChannelCam = name_ch
                                                                                if (name_ch == channel_active_right):
                                                                                    globalValues.numEvent[1] = 1
                                                                                    # print('qwewqewqe')
                                                                                elif (name_ch == channel_active_left):
                                                                                    # print('qwewqe')
                                                                                    globalValues.numEvent[0] = 1
                                                                                elif (name_ch == channel_hik_right):
                                                                                    # print('wefwef')
                                                                                    globalValues.numEvent[2] = 1
                                                                                elif (name_ch == channel_hik_left):
                                                                                    # print('wefwef')
                                                                                    globalValues.numEvent[3] = 1

                                                                                if (name_ch == channel_hik_left or name_ch == channel_hik_right):
                                                                                        # print(str_db_grz)
                                                                                        # curTalon = list_num_talon[j]
                                                                                        if (name_ch == channel_hik_right):
                                                                                            curListRtsp.append(globalValues.listRtsp[3])
                                                                                            curListIp.append(globalValues.listIp[3])
                                                                                            curListRtsp.append(globalValues.listRtsp[2])
                                                                                            curListIp.append(globalValues.listIp[2])
                                                                                        else:
                                                                                            curListRtsp.append(globalValues.listRtsp[2])
                                                                                            curListIp.append(globalValues.listIp[2])
                                                                                            curListRtsp.append(globalValues.listRtsp[3])
                                                                                            curListIp.append(globalValues.listIp[3])
                                                                                        # th_create_img_weight = threading.Thread(target=createImgsWeight, args = (0, curListRtsp, curListIp, globalValues.pathWeightImg, curTalon, ))
                                                                                        # th_create_img_weight.start()
                                                                                        globalValues.str_grz_weight = str_db_grz
                                                                                break

                                                                        iter_s += 1
                                                                if checkZN:
                                                                    break
                                                                j += 1
                                                        # print(strGRZ)
                                                        if checkZN:
                                                            break

                                                # print('checkPG123')

                                                if (checkZN == False):

                                                    print('checkingNumber!!!')

                                                    checkZNinPool = False
                                                    strGRZCur = ''
                                                    for rowPG in recordsPG:
                                                        strGRZ = str(rowPG[0])
                                                        strGRZ = strGRZ.replace('(', '').replace(')', '').replace(',', '').replace('\'', '').replace(' ', '').replace('?', '').replace('/', '').upper()

                                                        # print(globalValues.listPoolTS)

                                                        for str_db_grz in globalValues.listPoolTS:

                                                            print('checkingNumber123!!!')

                                                            iter_s = 0
                                                            check_good_grz = 0
                                                            lenAr = min(len(str_db_grz), len(strGRZ))
                                                            countSymb = 0
                                                            countVal = 0
                                                            for i in range(lenAr):
                                                                # print(strGRZ)
                                                                #
                                                                # print('CheckingElements!!!')
                                                                # print(strGRZ[iter_s])
                                                                # print(str_db_grz[iter_s])

                                                                numVals = 0
                                                                numLetter = 0

                                                                if (globalValues.debugOrder == False):
                                                                    numVals = 2
                                                                    numLetter = 2

                                                                if (0 < iter_s < 4):
                                                                    if (strGRZ[iter_s] == str_db_grz[iter_s]):
                                                                        countVal += 1
                                                                if ((iter_s == 0) or (iter_s == 4) or (iter_s == 5)):
                                                                    if (strGRZ[iter_s] == str_db_grz[iter_s]):
                                                                        countSymb += 1

                                                                if (countVal >= numVals and countSymb >= numLetter):
                                                                    print('goooddddddd!!!!!')
                                                                    checkZNinPool = True
                                                                    strGRZCur = str_db_grz
                                                                    globalValues.curChannelCam = str(rowPG[1])
                                                                    name_ch = globalValues.curChannelCam
                                                                    if (name_ch == channel_hik_left or name_ch == channel_hik_right):
                                                                        if (name_ch == channel_hik_right):
                                                                            curListRtsp.append(globalValues.listRtsp[3])
                                                                            curListIp.append(globalValues.listIp[3])
                                                                            curListRtsp.append(globalValues.listRtsp[2])
                                                                            curListIp.append(globalValues.listIp[2])
                                                                        else:
                                                                            curListRtsp.append(globalValues.listRtsp[2])
                                                                            curListIp.append(globalValues.listIp[2])
                                                                            curListRtsp.append(globalValues.listRtsp[3])
                                                                            curListIp.append(globalValues.listIp[3])
                                                                        globalValues.str_grz_weight = str_db_grz

                                                                    break
                                                                iter_s += 1
                                                            if checkZNinPool:
                                                                break

                                                        # print(strGRZ)
                                                        if checkZNinPool:
                                                            break

                                                    # print('checkingZNinPool: ', checkZNinPool)

                                                    if checkZNinPool:

                                                        checkGoodTime = True

                                                        if (len(globalValues.timelistTs) != 0):

                                                            delta = globalValues.deltaTime

                                                            checkGoodTime = False

                                                            k = 0

                                                            check = True

                                                            for el in globalValues.timelistTs:

                                                                # print('Checking: ')
                                                                # print(el)

                                                                k += 1

                                                                if (el[0] == strGRZCur):
                                                                    check = False
                                                                    str_time = strCurTime()
                                                                    str_time_start = str(el[1]).replace('.', '')
                                                                    val_min_start = int(str_time_start[2:4])
                                                                    val_start_time = int(str_time_start)
                                                                    str_time_end = str_time.replace('.', '')
                                                                    val_min_end = int(str_time_end[2:4])
                                                                    val_end_time = int(str_time_end)

                                                                    val_date = int(str(el[2])[0:2])
                                                                    dateToday = datetime.date.today().strftime('%d.%m.%Y')
                                                                    date_work = str(dateToday)
                                                                    val_date_now = int(date_work[0:2])

                                                                    cur_delta = 0

                                                                    # print(val_end_time)
                                                                    # print(val_start_time)

                                                                    if (val_end_time >= val_start_time and val_date != val_date_now):
                                                                        cur_delta = 9999
                                                                        checkGoodTime = True
                                                                        break

                                                                    elif (val_end_time >= val_start_time and val_date == val_date_now):
                                                                        if (val_min_end >= val_min_start):
                                                                            cur_delta = val_end_time - val_start_time - 4000
                                                                            cur_delta = val_end_time - val_start_time
                                                                            if (val_end_time - val_start_time) > delta:

                                                                                checkGoodTime = True
                                                                                # print('111')
                                                                                break
                                                                        else:
                                                                            cur_delta = val_end_time - val_start_time - 4000
                                                                            if (val_end_time - val_start_time - 4000) > delta:

                                                                                checkGoodTime = True
                                                                                # print('111222')
                                                                                break

                                                                    elif (val_end_time < val_start_time and val_date != val_date_now):
                                                                        if (val_min_end >= val_min_start):
                                                                            cur_delta = 240000 + val_end_time - val_start_time
                                                                            if (240000 + val_end_time - val_start_time) > delta:

                                                                                checkGoodTime = True
                                                                                print('1113333')
                                                                                break
                                                                        else:
                                                                            cur_delta = 240000 + val_end_time - val_start_time - 4000
                                                                            if (240000 + val_end_time - val_start_time - 4000) > delta:

                                                                                checkGoodTime = True
                                                                                print('111444')
                                                                                break

                                                                    # print(str(cur_delta))

                                                            if check:
                                                                checkGoodTime = True

                                                            # print('CheckingDeltaTime: ' + str(delta))



                                                            # print(str(cur_delta))

                                                        if (checkGoodTime):
                                                            try:
                                                                # with con_mysql:

                                                                # print('startCreateZN!')

                                                                print('CheckIn123!')

                                                                with con_mysql:
                                                                    cur.execute("SELECT * FROM " + globalValues.tblsDB[6])
                                                                    rows = cur.fetchall()
                                                                    print('qwe: ', rows)
                                                                    countRows = 0

                                                                    str_grz = ''
                                                                    car_name = ''
                                                                    date_work = ''
                                                                    name_company = ''
                                                                    car_model = ''

                                                                    dataSearch = strGRZCur.lower()
                                                                    is_search = False
                                                                    print(dataSearch)


                                                                    for row in rows:
                                                                        print(str(row[4]).lower())
                                                                        is_search = dataSearch in str(row[4]).lower()
                                                                        if is_search:
                                                                            str_grz = str(row[4]).upper()
                                                                            car_name = str(row[2])
                                                                            name_company = str(row[1])
                                                                            car_model = str(row[3])
                                                                            # print('checkingSearchData!')
                                                                            break

                                                                    # print(is_search)

                                                                    # is_search = False

                                                                    try:
                                                                        if is_search:
                                                                            # with con_mysql:
                                                                            with con_mysql:

                                                                                # print('StartingCreateOrder!!!')

                                                                                cur.execute("SELECT number_order FROM " + globalValues.tblsDB[1] + ' order by id desc limit 1')
                                                                                rows = cur.fetchall()

                                                                                number_order = 0
                                                                                if (len(rows) != 0):
                                                                                    number_order = str(int(rows[0][0]) + 1)
                                                                                else:
                                                                                    number_order = 100000

                                                                                # print(len(rows))
                                                                                # if (len(rows) != 0):
                                                                                #     for row in reversed(rows):
                                                                                #         try:
                                                                                #             number_order = str(int(row[0]) + 1)
                                                                                #         except Exception as ex:
                                                                                #             globalValues.writeLogData('Присвоение нового номера талона',str(ex))
                                                                                #         break

                                                                                print(number_order)
                                                                                print(str_grz)

                                                                                globalValues.numTalonVideo = number_order
                                                                                globalValues.numGRZVideo = str_grz
                                                                                name_ch = globalValues.curChannelCam
                                                                                if (name_ch == channel_active_right):
                                                                                    globalValues.numEvent[1] = 1
                                                                                elif (name_ch == channel_active_left):
                                                                                    globalValues.numEvent[0] = 1
                                                                                elif (name_ch == channel_hik_right):
                                                                                    globalValues.numEvent[2] = 1
                                                                                elif (name_ch == channel_hik_left):
                                                                                    globalValues.numEvent[3] = 1

                                                                                print(car_name)
                                                                                print(name_company)
                                                                                print(car_model)
                                                                                dateToday = datetime.date.today().strftime('%d.%m.%Y')
                                                                                date_work = str(dateToday)
                                                                                print(date_work)
                                                                                str_time = uiMain.strCurTime()

                                                                                print('sergioMartinez: ', number_order)

                                                                                query = ("INSERT INTO " + globalValues.tblsDB[1] + " (number_order, number_grz, date, time_entry, state_order, name_company) VALUES ( %s, %s, %s, %s, \'выполняется\', %s)")
                                                                                cur.execute(query, (number_order, str_grz, date_work, str_time, name_company))
                                                                                con_mysql.commit()

                                                                                query = ("INSERT INTO " + globalValues.tblsDB[2] + " (car_name, number_order, date_work, name_company, car_model, car_grz) VALUES (%s, %s, %s, %s, %s, %s)")
                                                                                cur.execute(query, (car_name, number_order, date_work, name_company, car_model, str_grz))
                                                                                con_mysql.commit()

                                                                                sqlMy = "INSERT INTO " + globalValues.tblsDB[4] + " (url) VALUES (%s)"
                                                                                cur.execute(sqlMy, ('https://hosting.wialon.com/'))
                                                                                con_mysql.commit()

                                                                                query = ("INSERT INTO " + globalValues.tblsDB[5] + " (weight, car_grz, name_company, date_work) VALUES ( \'не измерена\', %s, %s, %s)")
                                                                                cur.execute(query, (str_grz, name_company, date_work))
                                                                                con_mysql.commit()

                                                                                time.sleep(0.1)

                                                                                globalValues.refreshTblMain = True

                                                                                # print('checkingImg!')
                                                                                # print(curListRtsp)
                                                                                # print(curListIp)
                                                                                # print(globalValues.pathWeightImg)
                                                                                #
                                                                                # print('FinishingCreateOrder!!!')


                                                                                # curTalon = str(number_order)
                                                                                # print(curTalon)

                                                                                # th_create_img_weight = threading.Thread(target=createImgsWeight, args=(0, curListRtsp, curListIp, globalValues.pathWeightImg, curTalon, ))
                                                                                # th_create_img_weight.start()

                                                                    except Exception as ex:
                                                                        globalValues.writeLogData('Создание нового З/Н в автоматическом режиме', str(ex))

                                                            except Exception as ex:
                                                                globalValues.writeLogData('Создание З/Н в авто режиме', str(ex))

                                                if (firstCallIn):
                                                   firstCallIn = False
                                                   # print('FirstCallInPg!!!')
                                                else:
                                                    for rowPG in recordsPG:
                                                        name_ch = str(rowPG[1])

                                                        print(str(rowPG[1]))
                                                        print(str(rowPG[0]))
                                                        print(name_ch)
                                                        print('CheckingBDPgDataGRZ')

                                                        if (name_ch == channel_hik_left):
                                                                    if (str(rowPG[0] != '')):
                                                                        print('checkingIn1')
                                                                        strData = str(rowPG[0]).upper()
                                                                        height = 423
                                                                        checkAndChangeStr(strData, height, [uiMain.leMainGRZWeightOut1, uiMain.leMainGRZWeightOut2, uiMain.leMainGRZWeightOut3, uiMain.leMainGRZWeightOut4, uiMain.leMainGRZWeightOut5, uiMain.leMainGRZWeightOut6, uiMain.leMainGRZWeightOut7, uiMain.leMainGRZWeightOut8, uiMain.leMainGRZWeightOut9])
                                                                        print('checkingIn1')
                                                                        # if (req[0]):
                                                                        #     uiMain.leMainGRZWeightOut.setText(req[1])

                                                        if (name_ch == channel_hik_right):
                                                                    if (str(rowPG[0] != '')):
                                                                        print('checkingIn2')
                                                                        strData = str(rowPG[0]).upper()
                                                                        height = 383
                                                                        checkAndChangeStr(strData, height, [uiMain.leMainGRZWeightIn1, uiMain.leMainGRZWeightIn2, uiMain.leMainGRZWeightIn3, uiMain.leMainGRZWeightIn4, uiMain.leMainGRZWeightIn5, uiMain.leMainGRZWeightIn6, uiMain.leMainGRZWeightIn7, uiMain.leMainGRZWeightIn8, uiMain.leMainGRZWeightIn9])
                                                                        print('checkingIn2')
                                                                        # if (req[0]):
                                                                        #     uiMain.leMainGRZWeightIn.setText(req[1])

                                                        if (name_ch == channel_active_left):
                                                                    if (str(rowPG[0] != '')):
                                                                        print('checkingIn3')
                                                                        strData = str(rowPG[0]).upper()
                                                                        height = 303
                                                                        checkAndChangeStr(strData, height, [uiMain.leMainGRZGardIn1, uiMain.leMainGRZGardIn2, uiMain.leMainGRZGardIn3, uiMain.leMainGRZGardIn4, uiMain.leMainGRZGardIn5, uiMain.leMainGRZGardIn6, uiMain.leMainGRZGardIn7, uiMain.leMainGRZGardIn8, uiMain.leMainGRZGardIn9])
                                                                        print('checkingIn3')
                                                                        # if (req[0]):
                                                                        #     uiMain.leMainGRZGardIn.setText(req[1])

                                                        if (name_ch == channel_active_right):
                                                                    if (str(rowPG[0] != '')):
                                                                        print('checkingIn4')
                                                                        strData = str(rowPG[0]).upper()
                                                                        height = 343
                                                                        checkAndChangeStr(strData, height, [uiMain.leMainGRZGardOut1, uiMain.leMainGRZGardOut2, uiMain.leMainGRZGardOut3, uiMain.leMainGRZGardOut4, uiMain.leMainGRZGardOut5, uiMain.leMainGRZGardOut6, uiMain.leMainGRZGardOut7, uiMain.leMainGRZGardOut8, uiMain.leMainGRZGardOut9])
                                                                        print('checkingIn4')
                                                                        # if (req[0]):
                                                                        #     uiMain.leMainGRZGardOut.setText(req[1])

                                                        time.sleep(0.15)

                                                        # uiMain.updateGeometry()
                                                        # uiMain.frmCntrlPanel.updateGeometry()
                                                        #
                                                        # if (globalValues.colorForm == 1):
                                                        #
                                                        #     uiMain.lblTbl_2.setStyleSheet("background-color: rgb(242,242,242);\n"
                                                        #                     "border-radius: 5px;\n"
                                                        #                     "border: 2px solid rgb(205,205,205);")
                                                        #     uiMain.lblContDataWeight_3.setStyleSheet("background-color: rgb(235,235,235);\n"
                                                        #      "color: white;\n"
                                                        #      "border-radius: 5px;\n"
                                                        #      "border: 2px solid rgb(150,150,150);")
                                                        #     uiMain.leMainGRZGardInBack.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                        #      "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                        #      "border-radius: 3px;\n"
                                                        #      "border: 1px solid rgb(0,0,0);")
                                                        #     uiMain.leMainGRZGardOutBack.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                        #      "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                        #      "border-radius: 3px;\n"
                                                        #      "border: 1px solid rgb(0,0,0);")
                                                        #     uiMain.leMainGRZWeightInBack.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                        #      "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                        #      "border-radius: 3px;\n"
                                                        #      "border: 1px solid rgb(0,0,0);")
                                                        #     uiMain.leMainGRZWeightOutBack.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                        #      "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                        #      "border-radius: 3px;\n"
                                                        #      "border: 1px solid rgb(0,0,0);")
                                                        #
                                                        # else:
                                                        #
                                                        #     uiMain.lblTbl_2.setStyleSheet("background-color: rgb(62,62,62);\n"
                                                        #                     "border-radius: 5px;")
                                                        #     uiMain.lblContDataWeight_3.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                        #      "border-radius: 5px;")
                                                        #     uiMain.leMainGRZGardOutBack.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                        #      "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                        #      "border-radius: 3px;\n"
                                                        #      "border: 1px solid rgb(0,0,0);")
                                                        #     uiMain.leMainGRZWeightOutBack.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                        #      "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                        #      "border-radius: 3px;\n"
                                                        #      "border: 1px solid rgb(0,0,0);")
                                                        #     uiMain.leMainGRZWeightInBack.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                        #      "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                        #      "border-radius: 3px;\n"
                                                        #      "border: 1px solid rgb(0,0,0);")
                                                        #     uiMain.leMainGRZGardInBack.setStyleSheet("background-color: rgb(75,75,75);\n"
                                                        #      "image: url(" + globalValues.pathStyleImgs + "iconcarnum4.png);\n"
                                                        #      "border-radius: 3px;\n"
                                                        #      "border: 1px solid rgb(0,0,0);")

                                        records_old = recordsPG

                                        firstCallIn = False

                                if (globalValues.stopAll):
                                    break

                                gc.collect()

                                time.sleep(0.1)

                        print('checkPointExit1!')

                        cursor.close()
                        cur.close()

                        print('ExitCircleFindingGRZ!!!')

                    except Exception as ex:
                        globalValues.writeLogData('Поток работы с бд PgSql', str(ex))



                    time.sleep(1)

                    if (globalValues.stopAll):
                        break

     except Exception as ex:
             globalValues.writeLogData('Поток работы с БД PgSql', str(ex))

#Working with Cams

def startThCams(numCam, rtsp_link):
    print('StartingCamInMain!!!' + str(rtsp_link))
    try:
            th_cam_main = threading.Thread(target=thSetVideoCam, args=(rtsp_link, numCam,))
            th_cam_main.start()
            # sleep(0.1)

    except Exception as ex:
            strDataInLog = 'Функция запуска камеры основной формы ' + str(rtsp_link)
            globalValues.writeLogData(strDataInLog, str(ex))

def thSetVideoCam(rtsp_link, numCam):

        frame_rate = 20
        prev = 0
        import time
        # os.environ["OPENCV_FFMPEG_CAPTURE_OPTIONS"] = "rtsp_transport;udp"
        strCamInKPP = 'Камера въезд КПП'
        strCamOutKPP = 'Камера выезд КПП'
        strCamInWeight = 'Камера въезд Весы'
        strCAMOutWeight = 'Камера выезд Весы'


        try:

                while True:
                    if (globalValues.stopThreadCams):
                        break


                    # print(numCam)

                    cap = cv2.VideoCapture(rtsp_link, cv2.CAP_FFMPEG)

                    if (cap.isOpened()):

                            if (numCam == 1):
                                    strDataInJournal = 'Выполнено успешное открытие камеры для канала: ' + strCamInKPP
                                    globalValues.writeEventToDBJournalMain('Камеры главное окно', strDataInJournal)
                            elif (numCam == 2):
                                    strDataInJournal = 'Выполнено успешное открытие камеры для канала: ' + strCamOutKPP
                                    globalValues.writeEventToDBJournalMain('Камеры главное окно', strDataInJournal)
                            elif (numCam == 3):
                                    strDataInJournal = 'Выполнено успешное открытие камеры для канала: ' + strCamInWeight
                                    globalValues.writeEventToDBJournalMain('Камеры главное окно', strDataInJournal)
                            elif (numCam == 4):
                                    strDataInJournal = 'Выполнено успешное открытие камеры для канала: ' + strCAMOutWeight
                                    globalValues.writeEventToDBJournalMain('Камеры главное окно', strDataInJournal)

                    cap.set(cv2.CAP_PROP_BUFFERSIZE, 2)
                    # cap.set()
                    width_default = 400
                    koef_ratio = 1980/width_default
                    height_default = int(1020/koef_ratio)
                    width_full = 1670
                    height_full = 940
                    # print(str(koef_ratio))

                    while (globalValues.stopAll == False):

                        if (globalValues.stopThreadCams):
                            break

                        ret, frame = cap.read()
                        # print(numCam)
                        time_elapsed = time.time() - prev

                        if time_elapsed > 1. / frame_rate:
                            prev = time.time()

                            if (ret == False):
                                # print('NotCams!')
                                break

                            # print('Good Cam' + str(numCam))

                            if (numCam == 1):
                                # print('In')
                                frame = cv2.resize(frame, (width_default, height_default))
                                rgbImage = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                                h, w, ch = rgbImage.shape
                                bytesPerLine = ch * w
                                imgQt = QImage(rgbImage.data, w, h, bytesPerLine,
                                               QImage.Format_RGB888)
                                pixCam = QPixmap(imgQt)
                                uiMain.leCam0.setPixmap(pixCam)

                            elif (numCam == 2):
                                # print('In')
                                frame = cv2.resize(frame, (width_default, height_default))
                                rgbImage = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                                h, w, ch = rgbImage.shape
                                bytesPerLine = ch * w
                                imgQt = QImage(rgbImage.data, w, h, bytesPerLine,
                                               QImage.Format_RGB888)
                                pixCam = QPixmap(imgQt)
                                uiMain.leCam1.setPixmap(pixCam)

                            elif (numCam == 3):
                                # print('In')
                                frame = cv2.resize(frame, (width_default, height_default))
                                rgbImage = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                                h, w, ch = rgbImage.shape
                                bytesPerLine = ch * w
                                imgQt = QImage(rgbImage.data, w, h, bytesPerLine,
                                               QImage.Format_RGB888)
                                pixCam = QPixmap(imgQt)
                                uiMain.leCam2.setPixmap(pixCam)

                            elif (numCam == 4):
                                # print('In')
                                frame = cv2.resize(frame, (width_default, height_default))
                                rgbImage = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                                h, w, ch = rgbImage.shape
                                bytesPerLine = ch * w
                                imgQt = QImage(rgbImage.data, w, h, bytesPerLine,
                                               QImage.Format_RGB888)
                                pixCam = QPixmap(imgQt)
                                uiMain.leCam3.setPixmap(pixCam)

                            elif (numCam == 99):
                                frame = cv2.resize(frame, (width_full, height_full))
                                rgbImage = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                                h, w, ch = rgbImage.shape
                                bytesPerLine = ch * w
                                imgQt = QImage(rgbImage.data, w, h, bytesPerLine,
                                               QImage.Format_RGB888)
                                pixCam = QPixmap(imgQt)
                                uiMain.leCamAllMon.setPixmap(pixCam)



                    if (numCam == 1):
                        if (globalValues.colorForm == 1):
                            uiMain.leCam0.setStyleSheet("background-color: rgb(235,235,235);\n"
                                          "border-radius: 5px;\n"
                                          "border: 2px solid rgb(150,150,150);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "camDefault400x400grey.png);")
                        else:
                            uiMain.leCam0.setStyleSheet("background-color: rgb(0,0,0);\n"
                                          "border-radius: 5px;\n"
                                          "border: 2px solid rgb(42,42,42);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "camDefault400x400.png);")
                    elif (numCam == 2):
                        if (globalValues.colorForm == 1):
                            uiMain.leCam1.setStyleSheet("background-color: rgb(235,235,235);\n"
                                          "border-radius: 5px;\n"
                                          "border: 2px solid rgb(150,150,150);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "camDefault400x400grey.png);")
                        else:
                            uiMain.leCam1.setStyleSheet("background-color: rgb(0,0,0);\n"
                                          "border-radius: 5px;\n"
                                          "border: 2px solid rgb(42,42,42);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "camDefault400x400.png);")
                    elif (numCam == 3):
                        if (globalValues.colorForm == 1):
                            uiMain.leCam2.setStyleSheet("background-color: rgb(235,235,235);\n"
                                          "border-radius: 5px;\n"
                                          "border: 2px solid rgb(150,150,150);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "camDefault400x400grey.png);")
                        else:
                            uiMain.leCam2.setStyleSheet("background-color: rgb(0,0,0);\n"
                                          "border-radius: 5px;\n"
                                          "border: 2px solid rgb(42,42,42);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "camDefault400x400.png);")
                    elif (numCam == 4):
                        if (globalValues.colorForm == 1):
                            uiMain.leCam3.setStyleSheet("background-color: rgb(235,235,235);\n"
                                          "border-radius: 5px;\n"
                                          "border: 2px solid rgb(150,150,150);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "camDefault400x400grey.png);")
                        else:
                            uiMain.leCam3.setStyleSheet("background-color: rgb(0,0,0);\n"
                                          "border-radius: 5px;\n"
                                          "border: 2px solid rgb(42,42,42);\n"
                                          "image: url(" + globalValues.pathStyleImgs + "camDefault400x400.png);")
                    elif (numCam == 99):
                        # if (globalValues.colorForm == 1):
                        #     pixCamNotConGreyFull = pixCamNotConGrey.scaled(1670, 940, Qt.KeepAspectRatio)
                        #     uiMain.leCamAllMon.setPixmap(pixCamNotConGreyFull)
                        # else:
                        #     pixCamNotConFull = pixCamNotCon.scaled(1670, 940, Qt.KeepAspectRatio)
                        #     uiMain.leCamAllMon.setPixmap(pixCamNotConFull)
                        print('defaultCamCase!')
                    cap.release()

                    if (globalValues.stopAll):
                        print('ExitCam: ' + str(numCam))
                        break


        except Exception as ex:
                strDataInLog = 'Функция отображения камеры ' + str(rtsp_link)
                globalValues.writeLogData(strDataInLog, str(ex))

        print('ExitCam: ' + str(numCam))

def thChangeScroll(check):
    try:

        num_delta = 200
        global firstCallChangeScroll
        if(firstCallChangeScroll):
            firstCallChangeScroll = False
            num_delta = 700

        # print('Delta: ' + str(num_delta))

        start_time = round(time.time() * 100)
        while True:
            numRows = uiMain.tblMainData.verticalScrollBar().maximum()
            delta = round(abs(time.time() * 100) - start_time)

            # print(num_delta)

            if numRows != 0:
                # print('changeBarScroll!!!')
                uiMain.verticalScrollBar.setMaximum(uiMain.tblMainData.verticalScrollBar().maximum())
                break
            # print('checkingScroll!!!: ' + str(numRows))

            if delta > num_delta:
                if (check):
                    # print('changeScrollTimer')
                    uiMain.verticalScrollBar.setMaximum(uiMain.tblMainData.verticalScrollBar().maximum())
                break

            time.sleep(0.1)

            if (globalValues.stopAll):
                break
    except Exception as ex:
        globalValues.writeLogData('Поток обработки изменения скролла', str(ex))

def thChangeScrollStrg(check):
    try:

        num_delta = 200
        global firstCallChangeScroll
        if(firstCallChangeScroll):
            firstCallChangeScroll = False
            num_delta = 700

        # print('Delta: ' + str(num_delta))

        start_time = round(time.time() * 100)
        while True:
            numRows = uiMain.lstArchive.verticalScrollBar().maximum()
            delta = round(abs(time.time() * 100) - start_time)

            # print(num_delta)

            if numRows != 0:
                # print('changeBarScroll!!!')
                uiMain.vScrollArch.setMaximum(uiMain.lstArchive.verticalScrollBar().maximum())
                break
            # print('checkingScroll!!!: ' + str(numRows))

            if delta > num_delta:
                if (check):
                    # print('changeScrollTimer')
                    uiMain.vScrollArch.setMaximum(uiMain.lstArchive.verticalScrollBar().maximum())
                break

            time.sleep(0.1)

            if (globalValues.stopAll):
                break
    except Exception as ex:
        globalValues.writeLogData('Поток обработки изменения скролла архива', str(ex))

def checkFolderLongPath(pathFolder):
    try:
        i = 0
        listPath = []
        for element in pathFolder:
            if (element == '/' and i != 0):
                listPath.append(pathFolder[0:i])
            i += 1
        # if pathFolder[0] == '/':
        #     pathFolder = pathFolder[1:len(pathFolder)]
        # listPath = pathFolder.split('/')
        # listPath.pop(0)
        listPath.append(pathFolder)
        print('CheckingPath!!', )

        # print(listPath)

        for elPath in listPath:
            if (os.path.exists(elPath) == False):
                try:
                    os.mkdir(elPath)
                except Exception as ex:
                    globalValues.writeLogData('Функция проверки и создания папки хранилища', str(ex))

    except Exception as ex:
        globalValues.writeLogData('Функция проверки ссылки на папку хранилища', str(ex))

def checkAndChangeStr(strData, height, leList):

    try:
        print(strData)
        lengthStr = len(strData)
        if (lengthStr < 5):
            return False, strData
        valCheck = 0
        strDataNew = ''
        for el in strData:
            if (el == '?'):
                valCheck += 1
                el = ' '
            strDataNew += el
            if (valCheck > 3):
                return False, strDataNew
        length = len(strDataNew)
        strGRZCur = strDataNew
        print('checking!!!')
        print(strGRZCur)
        print(length)

        leList[2].show()
        leList[1].setGeometry QRect(253, height, 18, 26))
        leList[4].setGeometry QRect(292, height, 18, 26))
        leList[5].setGeometry QRect(305, height, 18, 26))

        if (length == 8):

            # strGRZCur = '   ' + strGRZCur[0] + ' ' + strGRZCur[1:4] + ' ' + strGRZCur[4:6] + '    ' + strGRZCur[6:length]
            leList[6].setGeometry QRect(344, height, 18, 26))
            leList[7].setGeometry QRect(355, height, 18, 26))
            leList[8].hide()

            for i in range(8):
                leList[i].setText(strGRZCur[i])
                print(strGRZCur[i])

        elif (length == 9):

            # strGRZCur = '   ' + strGRZCur[0] + ' ' + strGRZCur[1:4] + ' ' + strGRZCur[4:6] + '   ' + strGRZCur[6:length]
            leList[6].setGeometry QRect(338, height, 18, 26))
            leList[7].setGeometry QRect(349, height, 18, 26))
            leList[8].show()

            for i in range(length):
                leList[i].setText(strGRZCur[i])

        elif (length == 7):
            leList[6].setGeometry QRect(344, height, 18, 26))
            leList[7].setGeometry QRect(355, height, 18, 26))
            leList[8].hide()
            leList[2].hide()
            leList[1].setGeometry QRect(244, height, 18, 26))
            leList[4].setGeometry QRect(288, height, 18, 26))
            leList[5].setGeometry QRect(300, height, 18, 26))
            for i in range(length):
                if (i >= 2):
                    leList[i+1].setText(strGRZCur[i])
                else:
                    leList[i].setText(strGRZCur[i])



        strGRZCur = strGRZCur.upper()
        return True, strGRZCur

    except Exception as ex:
        globalValues.writeLogData('Функция обработки грз из бд pg', str(ex))

def ping(host):
        """
        Returns True if host (str) responds to a ping request.
        Remember that a host may not respond to a ping (ICMP) request even if the host name is valid.
        """

        # Option for the number of packets as a function of
        param = '-c'

        # Building the command. Ex: "ping -c 1 google.com"
        command = ['ping', param, '1', host]

        return subprocess.call(command) == 0

def checkCamInNetwork(strIps):
    checkGoodIp = False
    # strIp = '192.168.0.88'
    strIpHTTP = 'http://' + strIps
    # print(strIpHTTP)
    start_time = round(time.time())
    try:
        dataPing = ping(strIps)
    except Exception as ex:
        return checkGoodIp
    if (str(dataPing) == 'None'):
        checkGoodIp = False
        return checkGoodIp
        # print(abs(round(time.time()) - start_time))
    else:
        try:
            # print('goodping')
            data = requests.head(strIpHTTP, verify=False, timeout=4)

            checkGoodIp = True
            # print('Good opening: ' + strIps)
            return checkGoodIp
            # print(abs(round(time.time()) - start_time))
        except Exception as ex:
            checkGoodIp = False
            return checkGoodIp
            # print(abs(round(time.time()) - start_time))

def time_cur_in_out(lst_data_timing):
    try:
        dateToday = datetime.date.today().strftime('%d.%m.%Y')
        str_date_today = str(dateToday)

        lstDate = str_date_today.split('.')
        #
        # print(lstDate[0])
        # print(lstDate[1])

        cur_time = datetime.datetime.time(datetime.datetime.now())
        str_cur_time = str(cur_time)

        lstTiming = str_cur_time.split(':')
        hour = lstTiming[0]
        min = lstTiming[1]
        # print(hour)
        # print(min)

        str_cur_time = hour + min

        num_month = int(lstDate[1]) - 1
        num_day = int(lstDate[0]) - 1



        # print(lst_data_new[num_month][num_day])

        str_time_in = str(lst_data_timing[num_month][num_day][1])
        str_time_in = str_time_in[0:len(str_time_in) - 3]
        str_time_in = str_time_in.replace(':', '')
        val_time_in = int(str_time_in)
        str_time_out = str(lst_data_timing[num_month][num_day][2])
        str_time_out = str_time_out[0:len(str_time_out) - 3]
        str_time_out = str_time_out.replace(':', '')
        val_time_out = int(str_time_out)

        val_cur_time = int(str_cur_time)

        print('DataFromFile: ')

        print(val_cur_time)
        print(val_time_in)
        print(val_time_out)

        return val_cur_time, val_time_in, val_time_out


    except Exception as ex:
        globalValues.writeLogData('Функция создания текущего времени, времени восхода, времени заката', str(ex))

def createImgsWeight(numChanel, listRtsp, listIP, pathSaveFldr, numOrder):
    try:
        checkFolderLongPath(pathSaveFldr)
        checkJob = False
        numCheckIP = 0
        numCv2 = 0
        while True:
            str_Ip = listIP[numChanel]
            rtsp_link = listRtsp[numChanel]
            if (checkCamInNetwork(str_Ip)):
                print('sergio1')
                # cap = cv2.VideoCapture(rtsp_link, cv2.CAP_FFMPEG)
                cap = cv2.VideoCapture(rtsp_link)

                cap.set(cv2.CAP_PROP_BUFFERSIZE, 2)

                if (cap.isOpened()):
                    print('CamIsGood')
                    while (True):
                        if (globalValues.stopAll):
                            break
                        ret, frame = cap.read()
                        if ret:
                            pathImg = 'E:/ACMK/imgs/' + str(numOrder) + '_weight.jpg'
                            print(pathImg)
                            cv2.imwrite(pathImg, frame)

                            checkJob = True
                            break
                    # cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
                else:
                    numCv2 += 1
            else:
                if (numChanel == 0):
                    numChanel = 1
                elif (numChanel == 1):
                    numChanel = 0
                str_Ip = listIP[numChanel]
                rtsp_link = listRtsp[numChanel]

                if (checkCamInNetwork(str_Ip)):
                    # cap = cv2.VideoCapture(rtsp_link, cv2.CAP_FFMPEG)
                    cap = cv2.VideoCapture(rtsp_link)

                    if (cap.isOpened()):
                        while (True):
                            if (globalValues.stopAll):
                                break
                            ret, frame = cap.read()
                            if ret:
                                pathImg = pathSaveFldr + '/' + str(numOrder) + '_weight.jpg'
                                cv2.imwrite(pathImg, frame)
                                checkJob = True
                                break
                    else:
                        numCv2 += 1
                else:
                    numCheckIP += 1

            if (checkJob == False and (numCv2 > 1 or numCheckIP > 2)):
                break
            if (checkJob):
                break

    except Exception as ex:
        globalValues.writeLogData('Функция создания скрина события, весы', str(ex))

if __name__ == "__main__":



        locale = QLocale.system().name()
        # translator = QTranslator(app)
        # translator.load('{}qtbase_{}.qm'.format(I18N_QT_PATH, locale))
        # app.installTranslator(translator)

        app_icon = QIcon()
        app_icon.addFile(globalValues.pathStyleImgs + 'sinaps1717.png', QSize(16, 16))
        app.setWindowIcon(app_icon)

        try:
            pathFileStorage = globalValues.pathDefaultData + '/Sinaps'
            pathFileColor = globalValues.pathDefaultData + '/Sinaps/dataColor.txt'
            checkFolderLongPath(pathFileStorage)
            if (os.path.exists(pathFileColor)):
                f = open(pathFileColor, 'r')
                valueRead = int(f.read())
                if (valueRead == 0):
                    globalValues.colorForm = 1
                elif(valueRead == 1):
                    globalValues.colorForm = 0
                f.close()
        except Exception as ex:
            globalValues.writeLogData('Чтение данных из файла цветовой гаммы программы', str(ex))

        uiMain = Ui_MainFormGrunt()
        print('checking!!!')
        uiMain.uiDev.changeCOLORMainPanelGrunt(1.2)

        print('checking!!!123')

        if globalValues.debug:
            startThUpdateTable()
            uiMain.show()
        else:
            uiMain.openPanelLogin()

        try:
            sys.exit(app.exec_())
        except Exception as ex:
            print("Exiting")
            globalValues.writeLogData('ExitSys: ', str(ex))
            print(str(ex))
